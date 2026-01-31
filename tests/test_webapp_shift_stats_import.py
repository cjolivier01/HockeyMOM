import importlib.util
import os


def _load_app_module():
    spec = importlib.util.spec_from_file_location("webapp_app", "tools/webapp/app.py")
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore
    return mod


def should_parse_stats_inputs_player_stats_csv():
    os.environ["HM_WEBAPP_SKIP_DB_INIT"] = "1"
    os.environ["HM_WATCH_ROOT"] = "/tmp/hm-incoming-test"
    mod = _load_app_module()

    csv_text = (
        "Jersey #,Player,GP,Goals,Assists,Shots,SOG,xG,Giveaways,Takeaways,"
        "Controlled Entry For (On-Ice),Controlled Entry Against (On-Ice),"
        "Controlled Exit For (On-Ice),Controlled Exit Against (On-Ice),"
        "Plus Minus,GF Counted,GA Counted,Shifts,TOI Total,Average Shift,Median Shift,Longest Shift,Shortest Shift,"
        "Period 1 TOI,Period 1 Shifts,Period 1 GF,Period 1 GA,TOI Total (Video)\n"
        "8,Adam Ro,1,1,2,5,4,3,0,1,2,3,4,5,1,1,0,10,12:34,0:45,0:42,1:10,0:10,4:10,3,1,0,12:00\n"
    )

    rows = mod.parse_shift_stats_player_stats_csv(csv_text)
    assert len(rows) == 1
    r = rows[0]
    assert r["jersey_number"] == "8"
    assert r["name_norm"] == "adamro"
    assert r["stats"]["goals"] == 1
    assert r["stats"]["assists"] == 2
    assert r["stats"]["shots"] == 5
    assert r["stats"]["sog"] == 4
    assert r["stats"]["expected_goals"] == 3
    assert r["stats"]["shifts"] == 10
    assert r["stats"]["toi_seconds"] == 12 * 60 + 34
    assert r["stats"]["video_toi_seconds"] == 12 * 60
    assert r["stats"]["sb_avg_shift_seconds"] == 45
    assert r["stats"]["sb_median_shift_seconds"] == 42
    assert r["stats"]["sb_longest_shift_seconds"] == 70
    assert r["stats"]["sb_shortest_shift_seconds"] == 10

    p1 = r["period_stats"][1]
    assert p1["toi_seconds"] == 4 * 60 + 10
    assert p1["shifts"] == 3
    assert p1["gf"] == 1
    assert p1["ga"] == 0


def should_parse_stats_inputs_game_stats_csv():
    os.environ["HM_WEBAPP_SKIP_DB_INIT"] = "1"
    os.environ["HM_WATCH_ROOT"] = "/tmp/hm-incoming-test"
    mod = _load_app_module()

    csv_text = "Stat,sharks-12-1-r2\nGoals For,3\nShots For,22\nxG For,5\n"
    out = mod.parse_shift_stats_game_stats_csv(csv_text)
    assert out["_label"] == "sharks-12-1-r2"
    assert out["Goals For"] == "3"
    assert out["Shots For"] == "22"
    assert out["xG For"] == "5"


def should_format_seconds_to_toi_strings():
    os.environ["HM_WEBAPP_SKIP_DB_INIT"] = "1"
    os.environ["HM_WATCH_ROOT"] = "/tmp/hm-incoming-test"
    mod = _load_app_module()

    assert mod.format_seconds_to_mmss_or_hhmmss(12 * 60 + 34) == "12:34"
    assert mod.format_seconds_to_mmss_or_hhmmss(1 * 3600 + 2 * 60 + 3) == "1:02:03"


def should_parse_stats_inputs_otg_ota_columns():
    os.environ["HM_WEBAPP_SKIP_DB_INIT"] = "1"
    os.environ["HM_WATCH_ROOT"] = "/tmp/hm-incoming-test"
    mod = _load_app_module()

    csv_text = "Jersey #,Player,OT Goals,OT Assists\n8,Adam Ro,1,2\n"
    rows = mod.parse_shift_stats_player_stats_csv(csv_text)
    assert len(rows) == 1
    r = rows[0]
    assert r["stats"]["ot_goals"] == 1
    assert r["stats"]["ot_assists"] == 2


def should_parse_stats_inputs_turnovers_forced_keeps_columns_separate():
    os.environ["HM_WEBAPP_SKIP_DB_INIT"] = "1"
    os.environ["HM_WATCH_ROOT"] = "/tmp/hm-incoming-test"
    mod = _load_app_module()

    csv_text = "Jersey #,Player,Turnovers (forced),Giveaways,Takeaways\n8,Adam Ro,3,2,1\n"
    rows = mod.parse_shift_stats_player_stats_csv(csv_text)
    assert len(rows) == 1
    r = rows[0]
    assert r["stats"]["giveaways"] == 2
    assert r["stats"]["turnovers_forced"] == 3
    assert r["stats"]["takeaways"] == 1


def should_not_include_empty_period_columns_in_consolidated_stats():
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    rows = [
        {"player": "1_Ethan", "gp": "1", "goals": "1", "assists": "0", "P4_GF": "1"},
        {"player": "2_Other", "gp": "1", "goals": "0", "assists": "0"},
    ]
    df, cols = mod._build_stats_dataframe(rows, [1, 2, 3, 4], include_shifts_in_stats=False)  # type: ignore[attr-defined]
    assert "P4_GA" not in cols
    assert "P4_GF" in cols


def should_not_write_times_files_when_no_scripts():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        outdir = Path(td)

        mod._write_video_times_and_scripts(  # type: ignore[attr-defined]
            outdir, {"1_Ethan": [("0:00", "0:10")]}, create_scripts=False
        )
        mod._write_scoreboard_times(  # type: ignore[attr-defined]
            outdir, {"1_Ethan": [(1, "0:00", "0:10")]}, create_scripts=False
        )

        stats_dir = outdir / "stats"
        stats_dir.mkdir(parents=True, exist_ok=True)

        ctx = mod.EventLogContext(  # type: ignore[attr-defined]
            event_counts_by_player={},
            event_counts_by_type_team={("Goal", "For"): 1},
            event_instances={},
            event_player_rows=[
                {
                    "event_type": "Goal",
                    "team": "For",
                    "player": "1_Ethan",
                    "jersey": "1",
                    "period": 1,
                    "video_s": 100,
                    "game_s": 200,
                }
            ],
            team_roster={},
            team_excluded={},
        )

        mod._write_event_summaries_and_clips(  # type: ignore[attr-defined]
            outdir,
            stats_dir,
            ctx,
            conv_segments_by_period={},
            create_scripts=False,
        )
        mod._write_player_event_highlights(  # type: ignore[attr-defined]
            outdir,
            ctx,
            conv_segments_by_period={},
            player_keys=["1_Ethan"],
            create_scripts=False,
        )

        assert (stats_dir / "event_summary.csv").exists()
        assert (stats_dir / "event_summary.xlsx").exists()
        assert (stats_dir / "event_players.csv").exists()
        assert (stats_dir / "event_players.xlsx").exists()
        assert not list(outdir.glob("*_times.txt"))
        assert not list(outdir.glob("clip_*.sh"))


def should_not_write_player_shift_clip_files_without_shifts_flag():
    import importlib.util
    import tempfile
    from pathlib import Path

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_shift_files", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        base = Path(td)
        xls_path = base / "shift_sheet.xlsx"
        df = pd.DataFrame(
            [
                ["1st Period", "", "", "", "", ""],
                [
                    "Jersey No",
                    "Name",
                    mod.LABEL_START_SB,  # type: ignore[attr-defined]
                    mod.LABEL_END_SB,  # type: ignore[attr-defined]
                    mod.LABEL_START_V,  # type: ignore[attr-defined]
                    mod.LABEL_END_V,  # type: ignore[attr-defined]
                ],
                ["8", "Adam Ro", "0:10", "0:20", "0:10", "0:20"],
            ]
        )
        df.to_excel(xls_path, index=False, header=False)

        out_base = base / "out"
        outdir, stats_rows, _periods, _per_player_events, _pair_rows = mod.process_sheet(  # type: ignore[attr-defined]
            xls_path=xls_path,
            sheet_name=None,
            outdir=out_base,
            keep_goalies=True,
            goals=[],
            include_shifts_in_stats=False,
            skip_validation=True,
            create_scripts=True,
        )
        assert (outdir / "clip_all.sh").exists()
        assert (outdir / "stats" / "player_stats.csv").exists()
        assert stats_rows
        assert not (outdir / "stats" / "all_events_summary.csv").exists()
        assert not (outdir / "stats" / "all_events_summary.xlsx").exists()

        for row in stats_rows:
            pk = row["player"]
            assert not (outdir / f"{pk}_video_times.txt").exists()
            assert not (outdir / f"{pk}_scoreboard_times.txt").exists()
            assert not (outdir / f"clip_{pk}.sh").exists()

        out_base2 = base / "out2"
        outdir2, stats_rows2, _periods2, _per_player_events2, _pair_rows2 = mod.process_sheet(  # type: ignore[attr-defined]
            xls_path=xls_path,
            sheet_name=None,
            outdir=out_base2,
            keep_goalies=True,
            goals=[],
            include_shifts_in_stats=True,
            skip_validation=True,
            create_scripts=True,
        )
        assert (outdir2 / "clip_all.sh").exists()
        assert stats_rows2
        assert (outdir2 / "stats" / "summary_stats.csv").exists()
        assert (outdir2 / "stats" / "summary_stats.xlsx").exists()
        assert (outdir2 / "stats" / "all_events_summary.csv").exists()
        assert (outdir2 / "stats" / "all_events_summary.xlsx").exists()
        for row in stats_rows2:
            pk = row["player"]
            assert (outdir2 / f"{pk}_video_times.txt").exists()
            assert (outdir2 / f"{pk}_scoreboard_times.txt").exists()
            assert (outdir2 / f"clip_{pk}.sh").exists()


def should_write_pair_on_ice_xlsx_when_shifts_enabled(tmp_path):
    import importlib.util
    from pathlib import Path

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_pair", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    base = Path(tmp_path)
    xls_path = base / "shift_sheet.xlsx"
    df = pd.DataFrame(
        [
            ["1st Period", "", "", "", "", ""],
            [
                "Jersey No",
                "Name",
                mod.LABEL_START_SB,  # type: ignore[attr-defined]
                mod.LABEL_END_SB,  # type: ignore[attr-defined]
                mod.LABEL_START_V,  # type: ignore[attr-defined]
                mod.LABEL_END_V,  # type: ignore[attr-defined]
            ],
            ["8", "Adam Ro", "0:10", "0:40", "0:10", "0:40"],
            ["9", "Bee Two", "0:20", "0:30", "0:20", "0:30"],
        ]
    )
    df.to_excel(xls_path, index=False, header=False)

    outdir, _rows, _periods, _events, _pair_rows = mod.process_sheet(  # type: ignore[attr-defined]
        xls_path=xls_path,
        sheet_name=None,
        outdir=base / "out",
        keep_goalies=True,
        goals=[],
        include_shifts_in_stats=True,
        skip_validation=True,
        create_scripts=False,
    )
    assert (outdir / "stats" / "pair_on_ice.csv").exists()
    assert (outdir / "stats" / "pair_on_ice.xlsx").exists()


def should_write_pair_on_ice_without_shifts_but_not_publish_toi(tmp_path):
    import importlib.util
    from pathlib import Path

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_pair_no_shifts", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    base = Path(tmp_path)
    xls_path = base / "shift_sheet.xlsx"
    df = pd.DataFrame(
        [
            ["1st Period", "", "", "", "", ""],
            [
                "Jersey No",
                "Name",
                mod.LABEL_START_SB,  # type: ignore[attr-defined]
                mod.LABEL_END_SB,  # type: ignore[attr-defined]
                mod.LABEL_START_V,  # type: ignore[attr-defined]
                mod.LABEL_END_V,  # type: ignore[attr-defined]
            ],
            ["8", "Adam Ro", "0:10", "0:40", "0:10", "0:40"],
            ["9", "Bee Two", "0:20", "0:30", "0:20", "0:30"],
        ]
    )
    df.to_excel(xls_path, index=False, header=False)

    outdir, _rows, _periods, _events, _pair_rows = mod.process_sheet(  # type: ignore[attr-defined]
        xls_path=xls_path,
        sheet_name=None,
        outdir=base / "out",
        keep_goalies=True,
        goals=[],
        include_shifts_in_stats=False,
        skip_validation=True,
        create_scripts=False,
    )
    csv_text = (outdir / "stats" / "pair_on_ice.csv").read_text(encoding="utf-8")
    assert "Overlap %" in csv_text.splitlines()[0]
    assert "Player TOI" not in csv_text.splitlines()[0]
    assert "Overlap," not in csv_text.splitlines()[0]

    # Per-player stats should include the pair section, but without absolute time values.
    stats_text = (outdir / "stats" / "8_Adam_Ro_stats.txt").read_text(encoding="utf-8")
    assert "On-ice with teammates (by TOI%)" in stats_text
    assert "00:00:" not in stats_text
    assert "+0" not in stats_text


def should_not_write_team_assist_clip_scripts_and_sanitize_event_filenames():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_team_clips", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        outdir = Path(td)
        stats_dir = outdir / "stats"
        stats_dir.mkdir(parents=True, exist_ok=True)

        ctx = mod.EventLogContext(  # type: ignore[attr-defined]
            event_counts_by_player={},
            event_counts_by_type_team={
                ("Assist", "Blue"): 1,
                ("TurnoverForced", "Blue"): 1,
                ("CreatedTurnover", "Blue"): 1,
                ("Giveaway", "Blue"): 1,
                ("Giveaway", "White"): 1,
            },
            event_instances={
                ("Assist", "Blue"): [{"period": 1, "video_s": 100, "game_s": 200}],
                ("TurnoverForced", "Blue"): [{"period": 1, "video_s": 110, "game_s": 210}],
                ("CreatedTurnover", "Blue"): [{"period": 1, "video_s": 115, "game_s": 215}],
                ("Giveaway", "Blue"): [{"period": 1, "video_s": 120, "game_s": 220}],
                ("Giveaway", "White"): [{"period": 1, "video_s": 130, "game_s": 230}],
            },
            event_player_rows=[],
            team_roster={},
            team_excluded={},
        )

        mod._write_event_summaries_and_clips(  # type: ignore[attr-defined]
            outdir,
            stats_dir,
            ctx,
            conv_segments_by_period={},
            create_scripts=True,
            focus_team="Blue",
        )

        # Assist should not generate team-level event clip scripts or timestamp files.
        assert not list(outdir.glob("*Assist*"))

        # Event script/timestamp names should be sanitized (no parentheses).
        all_paths = list(outdir.rglob("*"))
        assert all_paths, "expected some files to be generated"
        assert not any("(" in p.name or ")" in p.name for p in all_paths)

        # Turnovers (forced) should use the sanitized filename form.
        assert (outdir / "clip_events_Turnovers_forced_For.sh").exists()
        script_text = (outdir / "clip_events_Turnovers_forced_For.sh").read_text(encoding="utf-8")
        assert "--blink-event-text" in script_text
        assert "--blink-event-label" in script_text
        assert '--blink-event-label "FORCED TURNOVER"' in script_text

        # Timestamp lines may include event-moment time(s) after start/end.
        ts_line = (
            (outdir / "events_Turnovers_forced_For_video_times.txt")
            .read_text(encoding="utf-8")
            .strip()
            .splitlines()[0]
        )
        assert len(ts_line.split()) >= 3

        # Giveaways should have For/Against team-level clip scripts.
        assert (outdir / "clip_events_Giveaway_For.sh").exists()
        assert (outdir / "clip_events_Giveaway_Against.sh").exists()

        # Created Turnovers should blink as singular.
        assert (outdir / "clip_events_Created_Turnovers_For.sh").exists()
        created_script = (outdir / "clip_events_Created_Turnovers_For.sh").read_text(
            encoding="utf-8"
        )
        assert '--blink-event-label "CREATED TURNOVER"' in created_script


def should_process_t2s_only_game_without_spreadsheets():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_t2s", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    class DummyT2SApi:
        def get_game_details(self, game_id: int):  # noqa: ARG002
            return {
                "stats": {
                    "homePlayers": [{"number": "1", "position": "F", "name": "Ethan Olivier"}],
                    "awayPlayers": [{"number": "99", "position": "F", "name": "Opponent Player"}],
                    "homeScoring": [
                        {
                            "period": "OT",
                            "time": "0:45",
                            "extra": "",
                            "goal": {"text": "#1 Ethan Olivier"},
                            "assist1": "",
                            "assist2": "",
                        }
                    ],
                    "awayScoring": [],
                }
            }

    # Install dummy T2S API to avoid network/DB.
    mod._t2s_api = DummyT2SApi()  # type: ignore[attr-defined]
    mod._t2s_api_loaded = True  # type: ignore[attr-defined]

    with tempfile.TemporaryDirectory() as td:
        base = Path(td)
        outdir = base / "out"
        hockey_db_dir = base / "db"
        final_outdir, stats_rows, periods, per_player_events = mod.process_t2s_only_game(  # type: ignore[attr-defined]
            t2s_id=51602,
            side="home",
            outdir=outdir,
            label="t2s-51602",
            hockey_db_dir=hockey_db_dir,
            include_shifts_in_stats=False,
        )

        assert (final_outdir / "stats" / "player_stats.csv").exists()
        assert (final_outdir / "stats" / "game_stats.csv").exists()

        rows_by_player = {r["player"]: r for r in stats_rows}
        assert rows_by_player["1_Ethan_Olivier"]["goals"] == "1"
        assert rows_by_player["1_Ethan_Olivier"]["assists"] == "0"
        assert rows_by_player["1_Ethan_Olivier"]["ot_goals"] == "1"
        assert rows_by_player["1_Ethan_Olivier"]["ot_assists"] == "0"
        assert 4 in periods  # OT -> period 4

        assert len(per_player_events["1_Ethan_Olivier"]["goals"]) == 1


def should_parse_t2s_only_token_with_side_and_label():
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_parse", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    assert mod._parse_t2s_only_token("t2s=53445:HOME:stockton-r2") == (  # type: ignore[attr-defined]
        53445,
        "home",
        "stockton-r2",
    )
    assert mod._parse_t2s_spec("53445:AWAY:stockton-r2") == (  # type: ignore[attr-defined]
        53445,
        "away",
        "stockton-r2",
    )


def should_fail_goals_from_t2s_when_api_missing():
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_t2s_missing", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    mod._t2s_api = None  # type: ignore[attr-defined]
    mod._t2s_api_loaded = True  # type: ignore[attr-defined]
    mod._t2s_api_import_error = "ModuleNotFoundError: dummy"  # type: ignore[attr-defined]

    try:
        mod.goals_from_t2s(51602, side="home")  # type: ignore[attr-defined]
        assert False, "expected goals_from_t2s to raise when T2S API is unavailable"
    except RuntimeError as e:
        msg = str(e)
        assert "failed to import" in msg
        assert "ModuleNotFoundError: dummy" in msg


def should_fail_goals_from_t2s_when_api_call_fails():
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_t2s_call_fail", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    class DummyT2SApi:
        def get_game_details(self, game_id, **kwargs):
            raise RuntimeError("boom")

    mod._t2s_api = DummyT2SApi()  # type: ignore[attr-defined]
    mod._t2s_api_loaded = True  # type: ignore[attr-defined]
    mod._t2s_api_import_error = None  # type: ignore[attr-defined]

    try:
        mod.goals_from_t2s(51602, side="home")  # type: ignore[attr-defined]
        assert False, "expected goals_from_t2s to raise when T2S API call fails"
    except RuntimeError as e:
        assert "boom" in str(e)


def should_parse_long_sheet_turnover_and_giveaway_distinct_events():
    import importlib.util

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_long", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    df = pd.DataFrame(
        [
            ["1st Period", "Video Time", "Scoreboard", "Team", "Shots", "Shots on Goal", "Assist"],
            ["Turnover", "0:10", "0:20", "Blue", "#5", "Caused by #7", "Takeaway by #8"],
            ["Unforced TO", "0:11", "0:21", "Blue", "#9", "Takeaway by #11", ""],
            ["Giveaway", "0:12", "0:22", "Blue", "#10", "Takeaway by #12", ""],
        ]
    )

    events, goal_rows, jerseys_by_team = mod._parse_long_left_event_table(df)  # type: ignore[attr-defined]
    assert goal_rows == []
    got = [(e.event_type, e.team, list(e.jerseys)) for e in events]
    assert got == [
        ("TurnoverForced", "Blue", [5]),
        ("CreatedTurnover", "White", [7]),
        ("Takeaway", "White", [8]),
        ("Giveaway", "Blue", [9]),
        ("Takeaway", "White", [11]),
        ("Giveaway", "Blue", [10]),
        ("Takeaway", "White", [12]),
    ]
    assert jerseys_by_team == {"Blue": {5, 9, 10}, "White": {7, 8, 11, 12}}


def should_error_on_unforced_turnover_with_multiple_opponent_jerseys():
    import importlib.util

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_long_err", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    df = pd.DataFrame(
        [
            ["1st Period", "Video Time", "Scoreboard", "Team", "Shots", "Shots on Goal", "Assist"],
            ["Unforced TO", "0:10", "0:20", "Blue", "#5", "Takeaway by #7 #8", ""],
        ]
    )

    try:
        mod._parse_long_left_event_table(df)  # type: ignore[attr-defined]
        assert False, "expected unforced turnover row to raise on multiple opponent jerseys"
    except ValueError as e:
        assert "unforced turnover row" in str(e).lower()


def should_parse_long_sheet_sharks_12_1_r3_turnover_roles():
    import importlib.util
    from collections import Counter
    from pathlib import Path

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_long_fixture", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    xls_path = Path(__file__).resolve().parent / "testdata" / "sharks-12-1-r3-long-54153.xlsx"
    df = pd.read_excel(xls_path, sheet_name=0, header=None)

    events, goal_rows, jerseys_by_team = mod._parse_long_left_event_table(df)  # type: ignore[attr-defined]
    assert goal_rows, "expected fixture to contain goal rows"
    assert jerseys_by_team.get("Blue") and jerseys_by_team.get("White")

    counts = Counter(e.event_type for e in events)
    assert counts["TurnoverForced"] == 51
    assert counts["Giveaway"] == 10
    assert counts["CreatedTurnover"] == 20
    assert counts["Takeaway"] == 53

    def _has(event_type: str, team: str, period: int, game_s: int, jersey: int) -> bool:
        return any(
            e.event_type == event_type
            and e.team == team
            and e.period == period
            and e.game_s == game_s
            and jersey in (e.jerseys or ())
            for e in events
        )

    # Forced turnover with both Caused By and Takeaway players.
    assert _has("TurnoverForced", "White", 1, 11 * 60 + 55, 13)
    assert _has("CreatedTurnover", "Blue", 1, 11 * 60 + 55, 29)
    assert _has("Takeaway", "Blue", 1, 11 * 60 + 55, 81)

    # Unforced turnover: giveaway + takeaway.
    assert _has("Giveaway", "White", 1, 12 * 60 + 13, 14)
    assert _has("Takeaway", "Blue", 1, 12 * 60 + 13, 57)

    # Regression: typo "Unforced OT" must not be mis-parsed as an OT period header.
    assert _has("TurnoverForced", "White", 3, 29, 59)
    assert _has("CreatedTurnover", "Blue", 3, 29, 16)
    assert _has("Takeaway", "Blue", 3, 29, 16)


def should_write_game_stats_consolidated_preserves_result_order():
    import importlib.util
    import tempfile
    from pathlib import Path

    import pandas as pd

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_game_stats_order", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        base = Path(td)
        outdir_a = base / "A" / "stats"
        outdir_b = base / "B" / "stats"
        outdir_a.mkdir(parents=True, exist_ok=True)
        outdir_b.mkdir(parents=True, exist_ok=True)

        pd.DataFrame({"Stat": ["Goals For"], "game-a": [1]}).to_csv(
            outdir_a / "game_stats.csv", index=False
        )
        pd.DataFrame({"Stat": ["Goals For"], "game-b": [2]}).to_csv(
            outdir_b / "game_stats.csv", index=False
        )

        results = [
            {"label": "B", "outdir": outdir_b.parent},
            {"label": "A", "outdir": outdir_a.parent},
        ]
        assert mod._write_game_stats_consolidated_files(base, results) is True  # type: ignore[attr-defined]

        df = pd.read_csv(base / "game_stats_consolidated.csv")
        assert list(df.columns) == ["Stat", "game-b", "game-a"]


def should_order_player_stats_consolidated_sheets_reverse_file_list():
    import importlib.util
    import sys
    import tempfile
    from pathlib import Path

    import pandas as pd
    from openpyxl import load_workbook

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_sheet_order", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    # In --file-list runs, newer games tend to be appended at the bottom. The consolidated
    # workbook should list per-game sheets in reverse file-list order (newest first).
    with tempfile.TemporaryDirectory() as td:
        base = Path(td)
        outdir = base / "out"
        file_list = base / "games.txt"

        def _write_minimal_sheet(path: Path) -> None:
            df = pd.DataFrame(
                [
                    ["1st Period", "", "", "", "", ""],
                    [
                        "Jersey No",
                        "Name",
                        mod.LABEL_START_SB,  # type: ignore[attr-defined]
                        mod.LABEL_END_SB,  # type: ignore[attr-defined]
                        mod.LABEL_START_V,  # type: ignore[attr-defined]
                        mod.LABEL_END_V,  # type: ignore[attr-defined]
                    ],
                    ["8", "Adam Ro", "0:10", "0:20", "0:10", "0:20"],
                ]
            )
            df.to_excel(path, index=False, header=False)

        old_game = base / "old.xlsx"
        new_game = base / "new.xlsx"
        _write_minimal_sheet(old_game)
        _write_minimal_sheet(new_game)

        # HOME/AWAY is now required for spreadsheet-backed games when using --file-list.
        file_list.write_text(f"{old_game}:HOME\n{new_game}:HOME\n", encoding="utf-8")

        argv_prev = sys.argv[:]
        try:
            sys.argv = [
                "parse_stats_inputs.py",
                "--file-list",
                str(file_list),
                "--outdir",
                str(outdir),
                "--no-scripts",
                "--skip-validation",
            ]
            mod.main()  # type: ignore[attr-defined]
        finally:
            sys.argv = argv_prev

        wb = load_workbook(outdir / "player_stats_consolidated.xlsx")
        assert wb.sheetnames[0] == "Cumulative"
        assert wb.sheetnames[1:] == ["new", "old"]


def should_aggregate_all_turnover_types_in_consolidated_player_stats():
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_agg", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    game1_rows = [
        {
            "player": "1_Ethan",
            "goals": "0",
            "assists": "0",
            "turnovers_forced": "1",
            "created_turnovers": "2",
            "giveaways": "3",
            "takeaways": "4",
        },
        {
            "player": "2_Other",
            "goals": "0",
            "assists": "0",
            "turnovers_forced": "0",
            "created_turnovers": "0",
            "giveaways": "1",
            "takeaways": "0",
        },
    ]
    game2_rows = [
        {
            "player": "1_Ethan",
            "goals": "0",
            "assists": "0",
            "turnovers_forced": "2",
            "created_turnovers": "",
            "giveaways": "0",
            "takeaways": "1",
        },
        {
            "player": "2_Other",
            "goals": "0",
            "assists": "0",
            "turnovers_forced": "1",
            "created_turnovers": "",
            "giveaways": "0",
            "takeaways": "0",
        },
    ]

    aggregated_rows, _periods, per_game_denoms = mod._aggregate_stats_rows(  # type: ignore[attr-defined]
        [(game1_rows, [1]), (game2_rows, [1])]
    )
    assert per_game_denoms["turnovers_forced_per_game"] == 2
    assert per_game_denoms["giveaways_per_game"] == 2
    assert per_game_denoms["takeaways_per_game"] == 2
    assert per_game_denoms["created_turnovers_per_game"] == 1

    by_player = {r["player"]: r for r in aggregated_rows}
    ethan = by_player["1_Ethan"]
    assert ethan["turnovers_forced"] == "3"
    assert ethan["turnovers_forced_per_game"] == "1.5"
    assert ethan["giveaways"] == "3"
    assert ethan["giveaways_per_game"] == "1.5"
    assert ethan["takeaways"] == "5"
    assert ethan["takeaways_per_game"] == "2.5"
    assert ethan["created_turnovers"] == "2"
    assert ethan["created_turnovers_per_game"] == "2.0"


def should_select_tracking_output_video_prefers_highest_number():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_video", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        game_dir = Path(td) / "chicago-4"
        stats_dir = game_dir / "stats"
        stats_dir.mkdir(parents=True, exist_ok=True)
        sheet_path = stats_dir / "game_stats.xlsx"
        sheet_path.write_text("dummy", encoding="utf-8")

        (game_dir / "tracking_output-with-audio.mp4").write_text("", encoding="utf-8")
        (game_dir / "tracking_output-with-audio-2.mp4").write_text("", encoding="utf-8")
        best = game_dir / "tracking_output-with-audio-10.mp4"
        best.write_text("", encoding="utf-8")

        picked = mod._find_tracking_output_video_for_sheet_path(sheet_path)  # type: ignore[attr-defined]
        assert picked == best


def should_write_combined_highlight_times_with_deduped_xg_goal():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_highlights", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        outdir = Path(td) / "out"
        outdir.mkdir(parents=True, exist_ok=True)

        # Identity scoreboard->video mapping for period 1.
        conv = {1: [(0, 500, 0, 500)]}

        goal = mod.GoalEvent("GF", 1, "1:40")  # 100s
        assist = mod.GoalEvent("GF", 1, "3:20")  # 200s
        per_player_events = {
            "1_Ethan": {"goals": [goal], "assists": [assist], "gf_on_ice": [], "ga_on_ice": []}
        }

        # ExpectedGoal at the goal time should be skipped (goals count as xG internally).
        ctx = mod.EventLogContext(
            event_counts_by_player={},
            event_counts_by_type_team={},
            event_instances={},
            event_player_rows=[
                {
                    "event_type": "ExpectedGoal",
                    "team": "For",
                    "player": "1_Ethan",
                    "jersey": 1,
                    "period": 1,
                    "video_s": None,
                    "game_s": 100,
                },
                {
                    "event_type": "ExpectedGoal",
                    "team": "For",
                    "player": "1_Ethan",
                    "jersey": 1,
                    "period": 1,
                    "video_s": None,
                    "game_s": 300,
                },
                {
                    "event_type": "Takeaway",
                    "team": "For",
                    "player": "1_Ethan",
                    "jersey": 1,
                    "period": 1,
                    "video_s": 400,
                    "game_s": None,
                },
            ],
            team_roster={},
            team_excluded={},
        )

        mod._write_player_combined_highlights(  # type: ignore[attr-defined]
            outdir,
            event_log_context=ctx,
            conv_segments_by_period=conv,
            per_player_goal_events=per_player_events,
            player_keys=["1_Ethan"],
            create_scripts=True,
        )

        ts_file = outdir / "events_Highlights_1_Ethan_video_times.txt"
        assert ts_file.exists()
        assert ts_file.read_text(encoding="utf-8").strip().splitlines() == [
            "00:01:20 00:01:50",  # Goal (20s pre, 10s post)
            "00:03:00 00:03:30",  # Assist (uses goal window)
            "00:04:50 00:05:05",  # xG (10s pre, 5s post)
            "00:06:30 00:06:45",  # Takeaway (10s pre, 5s post)
        ]


def should_write_season_highlight_script_with_embedded_video_paths():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_season_scripts", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        base = Path(td)
        base_outdir = base / "out"
        base_outdir.mkdir(parents=True, exist_ok=True)

        # Two games with stats dirs and tracking videos.
        g1 = base / "game1"
        g2 = base / "game2"
        (g1 / "stats").mkdir(parents=True, exist_ok=True)
        (g2 / "stats").mkdir(parents=True, exist_ok=True)
        sheet1 = g1 / "stats" / "game_stats.xlsx"
        sheet2 = g2 / "stats" / "game_stats.xlsx"
        sheet1.write_text("dummy", encoding="utf-8")
        sheet2.write_text("dummy", encoding="utf-8")

        # Game1: prefer highest numbered video.
        (g1 / "tracking_output-with-audio.mp4").write_text("", encoding="utf-8")
        (g1 / "tracking_output-with-audio-3.mp4").write_text("", encoding="utf-8")
        best1 = g1 / "tracking_output-with-audio-12.mp4"
        best1.write_text("", encoding="utf-8")
        # Game2: only plain.
        best2 = g2 / "tracking_output-with-audio.mp4"
        best2.write_text("", encoding="utf-8")

        # Per-game output dirs with per-player highlight timestamps.
        out1 = base_outdir / "game1" / "per_player"
        out2 = base_outdir / "game2" / "per_player"
        out1.mkdir(parents=True, exist_ok=True)
        out2.mkdir(parents=True, exist_ok=True)
        (out1 / "events_Highlights_1_Ethan_video_times.txt").write_text(
            "00:00:10 00:00:20\n", encoding="utf-8"
        )
        (out2 / "events_Highlights_1_Ethan_video_times.txt").write_text(
            "00:00:30 00:00:40\n", encoding="utf-8"
        )

        results = [
            {"label": "game1", "outdir": out1, "sheet_path": sheet1, "video_path": None},
            {"label": "game2", "outdir": out2, "sheet_path": sheet2, "video_path": None},
        ]

        mod._write_season_highlight_scripts(  # type: ignore[attr-defined]
            base_outdir, results, create_scripts=True, use_yaml_order=True
        )

        script = base_outdir / "season_highlights" / "clip_season_highlights_1_Ethan.sh"
        assert script.exists()
        content = script.read_text(encoding="utf-8")
        assert str(best1) in content
        assert str(best2) in content


def should_write_season_highlight_script_uses_absolute_timestamp_paths_for_relative_outdir():
    import importlib.util
    import os
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_season_scripts_rel", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        old_cwd = os.getcwd()
        os.chdir(td)
        try:
            base_outdir = Path("player_shifts")
            base_outdir.mkdir(parents=True, exist_ok=True)

            game_label = "chicago-1"
            game_dir = Path(game_label)
            (game_dir / "stats").mkdir(parents=True, exist_ok=True)
            sheet = game_dir / "stats" / "game_stats.xlsx"
            sheet.write_text("dummy", encoding="utf-8")
            video = game_dir / "tracking_output-with-audio.mp4"
            video.write_text("", encoding="utf-8")

            out = base_outdir / game_label / "per_player"
            out.mkdir(parents=True, exist_ok=True)
            ts = out / "events_Highlights_1_Ethan_video_times.txt"
            ts.write_text("00:00:10 00:00:20\n", encoding="utf-8")

            results = [
                {"label": game_label, "outdir": out, "sheet_path": sheet, "video_path": None}
            ]
            mod._write_season_highlight_scripts(base_outdir, results, create_scripts=True)  # type: ignore[attr-defined]

            script = base_outdir / "season_highlights" / "clip_season_highlights_1_Ethan.sh"
            assert script.exists()
            content = script.read_text(encoding="utf-8")

            assert f'TS_FILE="{ts.resolve()}"' in content
            assert f'VIDEO="{video.resolve()}"' in content
        finally:
            os.chdir(old_cwd)


def should_write_season_highlight_script_uses_ffmpeg_concat_for_season_join():
    import importlib.util
    import tempfile
    from pathlib import Path

    spec = importlib.util.spec_from_file_location(
        "parse_stats_inputs_mod_season_scripts_ffmpeg", "scripts/parse_stats_inputs.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)  # type: ignore

    with tempfile.TemporaryDirectory() as td:
        base = Path(td)
        base_outdir = base / "out"
        base_outdir.mkdir(parents=True, exist_ok=True)

        game = base / "game1"
        (game / "stats").mkdir(parents=True, exist_ok=True)
        sheet = game / "stats" / "game_stats.xlsx"
        sheet.write_text("dummy", encoding="utf-8")
        video = game / "tracking_output-with-audio.mp4"
        video.write_text("", encoding="utf-8")

        out = base_outdir / "game1" / "per_player"
        out.mkdir(parents=True, exist_ok=True)
        (out / "events_Highlights_1_Ethan_video_times.txt").write_text(
            "00:00:10 00:00:20\n", encoding="utf-8"
        )

        results = [{"label": "game1", "outdir": out, "sheet_path": sheet, "video_path": None}]
        mod._write_season_highlight_scripts(base_outdir, results, create_scripts=True)  # type: ignore[attr-defined]

        script = base_outdir / "season_highlights" / "clip_season_highlights_1_Ethan.sh"
        assert script.exists()
        content = script.read_text(encoding="utf-8")

        assert 'echo "file \'$f\'" >> "$LIST_FILE"' in content
        assert 'ffmpeg -f concat -safe 0 -i "$LIST_FILE" -c copy "$OUT_FILE"' in content
        assert "--video-file-list" not in content
