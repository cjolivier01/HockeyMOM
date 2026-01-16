from __future__ import annotations

import io
import json

import pytest


@pytest.fixture()
def client(monkeypatch, webapp_db):
    _django_orm, _m = webapp_db
    monkeypatch.setenv("HM_WEBAPP_SKIP_DB_INIT", "1")
    from django.test import Client

    return Client()


def should_export_table_as_xlsx_and_redact_shift_toi_columns(client):
    payload = {
        "title": "Player Stats",
        "sheet_name": "Player Stats",
        "filename": "player_stats.xlsx",
        "headers": ["Jersey #", "Player", "TOI", "Shifts", "Goals"],
        "col_keys": [None, None, "toi_seconds", "shifts", None],
        "rows": [["1", "Alice", "10:00", "20", "2"]],
        "freeze_cols": 2,
    }
    resp = client.post(
        "/api/xlsx/table",
        data=json.dumps(payload),
        content_type="application/json",
    )
    assert resp.status_code == 200
    assert (
        resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    assert str(ws.cell(row=1, column=1).value or "").strip() == "Player Stats"

    headers = [
        str(ws.cell(row=2, column=i).value or "").replace("\n", " ").strip() for i in range(1, 6)
    ]
    headers = [h for h in headers if h]
    assert headers == ["Jersey #", "Player", "Goals"]

    assert str(ws.freeze_panes) == "C3"


def should_reject_non_post_methods(client):
    resp = client.get("/api/xlsx/table")
    assert resp.status_code == 405
