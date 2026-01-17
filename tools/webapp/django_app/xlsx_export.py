from __future__ import annotations

import io
import re
from typing import Optional


SHIFT_STAT_KEYS: set[str] = {
    "toi_seconds",
    "toi_seconds_per_game",
    "shifts",
    "shifts_per_game",
    "video_toi_seconds",
    "sb_avg_shift_seconds",
    "sb_median_shift_seconds",
    "sb_longest_shift_seconds",
    "sb_shortest_shift_seconds",
}


def _wrap_header_after_words(header: str, *, words_per_line: int = 2) -> str:
    if not header:
        return ""
    parts = str(header).strip().split()
    if len(parts) <= words_per_line:
        return str(header).strip()
    lines = [" ".join(parts[i : i + words_per_line]) for i in range(0, len(parts), words_per_line)]
    return "\n".join(lines)


def _sanitize_sheet_name(name: str) -> str:
    s = str(name or "").strip() or "Sheet"
    # Excel sheet name restrictions: max 31 chars, no []:*?/\
    s = re.sub(r"[\[\]\:\*\?\/\\]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        s = "Sheet"
    return s[:31]


def _sanitize_filename(name: str) -> str:
    s = str(name or "").strip() or "table"
    s = re.sub(r"[^\w\-. ]+", "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace(" ", "_")
    if not s:
        s = "table"
    if not s.lower().endswith(".xlsx"):
        s = f"{s}.xlsx"
    return s


def _header_len(header: str) -> int:
    if header is None:
        return 0
    parts = str(header).splitlines()
    return max((len(p) for p in parts), default=len(str(header)))


def _autosize_columns(ws, *, headers: list[str], rows: list[list[str]]) -> None:
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    for idx, h in enumerate(headers, start=1):
        max_len = _header_len(h)
        for r in rows:
            if idx - 1 >= len(r):
                continue
            v = r[idx - 1]
            vlen = len(str(v or ""))
            if vlen > max_len:
                max_len = vlen
        # Leave extra room for Excel's filter dropdown arrow so it doesn't overlap header text.
        width = min(max(max_len + 4, 10), 80)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _should_redact_column(*, header: str, key: Optional[str]) -> bool:
    if key:
        if str(key).strip() in SHIFT_STAT_KEYS:
            return True

    h = str(header or "").strip().casefold()
    # Never export TOI/shift columns (privacy / coach-usage concerns).
    if "toi" in h:
        return True
    if "ice time" in h or "ice-time" in h:
        return True
    if "shift" in h:
        return True
    return False


def redact_shift_columns(
    *, headers: list[str], rows: list[list[str]], col_keys: Optional[list[Optional[str]]] = None
) -> tuple[list[str], list[list[str]], Optional[list[Optional[str]]]]:
    if not headers:
        return headers, rows, col_keys
    keys = list(col_keys) if col_keys else [None] * len(headers)
    keep_idx: list[int] = []
    out_headers: list[str] = []
    out_keys: list[Optional[str]] = []
    for i, h in enumerate(headers):
        k = keys[i] if i < len(keys) else None
        if _should_redact_column(header=str(h or ""), key=(str(k).strip() if k else None)):
            continue
        keep_idx.append(i)
        out_headers.append(str(h or ""))
        out_keys.append(k)

    out_rows: list[list[str]] = []
    for r in rows or []:
        rr: list[str] = []
        for i in keep_idx:
            rr.append("" if i >= len(r) else str(r[i] or ""))
        out_rows.append(rr)
    return out_headers, out_rows, (out_keys if col_keys else None)


def build_table_xlsx_bytes(
    *,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    freeze_cols: int = 0,
    sheet_name: Optional[str] = None,
    words_per_line: int = 2,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    safe_title = str(title or "").strip() or "Table"
    safe_sheet_name = _sanitize_sheet_name(sheet_name or safe_title)

    headers_wrapped = [
        _wrap_header_after_words(str(h), words_per_line=words_per_line) for h in (headers or [])
    ]
    rows_str = [[str(v or "") for v in (r or [])] for r in (rows or [])]

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name

    ncols = len(headers_wrapped)
    if ncols <= 0:
        # Still produce a workbook with a single title cell.
        ws["A1"].value = safe_title
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Theme colors match scripts/parse_stats_inputs.py.
    teal_fill = PatternFill(fill_type="solid", start_color="FF009688", end_color="FF009688")
    header_font = Font(color="FF000000", bold=True)
    title_font = Font(color="FF000000", bold=True, size=14)
    band_a = PatternFill(fill_type="solid", start_color="FFE6E6E6", end_color="FFE6E6E6")
    band_b = PatternFill(fill_type="solid", start_color="FFF2F2F2", end_color="FFF2F2F2")

    white_side = Side(style="thin", color="FFFFFFFF")
    white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

    title_row = 1
    header_row = 2
    data_start_row = 3

    # Title row (merged across all columns)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)
    ws.row_dimensions[title_row].height = 24.0
    for c in range(1, ncols + 1):
        cell = ws.cell(row=title_row, column=c)
        cell.fill = teal_fill
        if c == 1:
            cell.value = safe_title
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    max_lines = 1
    for c, h in enumerate(headers_wrapped, start=1):
        cell = ws.cell(row=header_row, column=c)
        cell.value = h
        cell.fill = teal_fill
        cell.font = header_font
        max_lines = max(max_lines, len(str(h or "").splitlines()))
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[header_row].height = max(15.0, 15.0 * max_lines)

    # Data rows
    for i, r in enumerate(rows_str):
        rr = data_start_row + i
        fill = band_a if (i % 2 == 0) else band_b
        for c in range(1, ncols + 1):
            cell = ws.cell(row=rr, column=c)
            cell.value = r[c - 1] if c - 1 < len(r) else ""
            cell.fill = fill

    # Borders (title + header + data)
    last_row = data_start_row + max(len(rows_str), 1) - 1 if rows_str else header_row
    for r in range(title_row, last_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = white_border

    # Auto-filter on header row.
    try:
        last_col = get_column_letter(ncols)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
    except Exception:
        # Best-effort: some openpyxl versions/edge cases can fail to apply filters; avoid failing export.
        pass

    # Right-align time/duration-like columns.
    try:
        right_cols: list[int] = []
        for idx, cell in enumerate(ws[header_row], start=1):
            name = str(cell.value or "").replace("\n", " ").strip()
            if any(tok in name for tok in ["Time", "Overlap", "Video", "Duration"]):
                right_cols.append(idx)
        for r in range(data_start_row, data_start_row + len(rows_str)):
            for c in right_cols:
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(
                    horizontal="right", vertical=cell.alignment.vertical or "center"
                )
    except Exception:
        # Best-effort: alignment failures should not prevent exporting a usable spreadsheet.
        pass

    # Freeze title+header rows and optionally N left columns.
    try:
        freeze_cols_i = int(freeze_cols or 0)
    except Exception:
        freeze_cols_i = 0
    if freeze_cols_i < 0:
        freeze_cols_i = 0
    try:
        start_col = get_column_letter(min(max(freeze_cols_i + 1, 1), ncols))
        ws.freeze_panes = f"{start_col}{data_start_row}"
    except Exception:
        ws.freeze_panes = "A3"

    _autosize_columns(ws, headers=headers_wrapped, rows=rows_str)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_table_xlsx(
    *,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    col_keys: Optional[list[Optional[str]]] = None,
    freeze_cols: int = 0,
    sheet_name: Optional[str] = None,
    filename: Optional[str] = None,
) -> tuple[str, bytes]:
    headers2, rows2, keys2 = redact_shift_columns(headers=headers, rows=rows, col_keys=col_keys)
    out_bytes = build_table_xlsx_bytes(
        title=title,
        headers=headers2,
        rows=rows2,
        freeze_cols=freeze_cols,
        sheet_name=sheet_name,
        words_per_line=2,
    )
    return _sanitize_filename(filename or title), out_bytes
