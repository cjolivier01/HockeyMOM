#!/usr/bin/env python3
"""
Extract shifts and stats from the 'dh-tv-12-1.xls' style sheet.

Outputs per-player:
  - {Jersey}_{Name}_stats.txt                -> scoreboard-time-based stats (plus/minus, goals for/against counted, etc.)

When `--shifts` is set (and `--no-scripts` is not), also writes per-player shift clip helpers:
  - {Jersey}_{Name}_video_times.txt          -> "videoStart videoEnd"
  - {Jersey}_{Name}_scoreboard_times.txt     -> "period scoreboardStart scoreboardEnd"

Goals can be specified via:
  --goal GF:2/13:45 --goal GA:1/05:12 ...
  --goal GF:OT/0:45  # OT goals use period 'OT' (treated as period 4)
or as a file with lines like:
  GF:1/13:47
  GA:2/09:15
  # comments and blank lines allowed

Alternatively, provide a TimeToScore spec to auto-fill goals (or run a game with no spreadsheets):
  --t2s 51602 --home                 # Your team is home (home scoring = GF)
  --t2s 51602 --away                 # Your team is away (away scoring = GF)
  --t2s 51602:HOME:stockton-r2       # Also sets the game label for T2S-only games

Install deps (for .xls):
  pip install pandas xlrd

Example:
  python scripts/parse_shift_spreadsheet.py \
      --input dh-tv-12-1.xls \
      --outdir player_focus \
      --t2s 51602 --away \
      --keep-goalies
"""

import argparse
import datetime
import os
import re
import statistics
import sys
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple

import pandas as pd

# Ensure repo root is on sys.path so optional `hmlib.*` imports work when running
# this script directly from `scripts/`.
_repo_root = Path(__file__).resolve().parents[1]
if str(_repo_root) not in sys.path:
    sys.path.insert(0, str(_repo_root))

# Optional import of TimeToScore API (available in this repo). We lazily import it
# to avoid pulling in heavy `hmlib` dependencies when users run this script without T2S.
_t2s_api: Any = None
_t2s_api_loaded = False
_t2s_api_import_error: Optional[str] = None


def _get_t2s_api() -> Any:
    global _t2s_api, _t2s_api_loaded, _t2s_api_import_error
    if _t2s_api_loaded:
        return _t2s_api
    _t2s_api_loaded = True
    try:  # pragma: no cover - optional at runtime
        from hmlib.time2score import api as t2s_api  # type: ignore

        _t2s_api = t2s_api
        _t2s_api_import_error = None
    except Exception as e:  # noqa: BLE001
        _t2s_api = None
        _t2s_api_import_error = f"{type(e).__name__}: {e}"
    return _t2s_api

# Header labels as they appear in the sheet
LABEL_START_SB = "Shift Start (Scoreboard Time)"
LABEL_END_SB = "Shift End (Scoreboard Time)"
LABEL_START_V = "Shift Start (Video Time)"
LABEL_END_V = "Shift End (Video Time)"

# Clip window durations (seconds)
EVENT_CLIP_PRE_S = 10
EVENT_CLIP_POST_S = 5
GOAL_CLIP_PRE_S = 20
GOAL_CLIP_POST_S = 10


# ----------------------------- utilities -----------------------------


def sanitize_name(s: str) -> str:
    s = s.strip().replace(" ", "_")
    return re.sub(r"[^A-Za-z0-9_\-]", "", s)


def is_period_label(x: object) -> bool:
    return parse_period_label(x) is not None


def parse_period_label(x: object) -> Optional[int]:
    """
    Extract the period number from a label.
    Handles variants like 'Period 1', '1st Period', '1st Period (Blue team)', and 'OT'.
    """
    try:
        s = str(x).strip()
    except Exception:
        return None
    if not s:
        return None
    # Overtime. Some sheets label this like "OT", "Overtime", or "OT 3 on 3".
    # Treat OT labels as period 4, but only when the cell itself is an OT header
    # (avoid matching event labels like "Unforced OT").
    if re.match(r"(?i)^(?:ot|overtime)(?:\b|\d)", s):
        return 4
    m = re.search(r"(?i)(\d+)(?:st|nd|rd|th)?\s*period", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    m = re.search(r"(?i)period\s*(\d+)", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def _int_floor_seconds_component(sec_str: str) -> int:
    """Return integer seconds, flooring if fractional (e.g., '12.7' -> 12)."""
    try:
        return int(float(sec_str))
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Invalid seconds component '{sec_str}': {e}")


def parse_flex_time_to_seconds(s: str) -> int:
    """
    Accepts H:MM:SS(.fff) or M:SS(.fff) or MM:SS(.fff).
    Returns total seconds (int), flooring fractional seconds.
    """
    s = s.strip()
    parts = s.split(":")
    if len(parts) == 3:
        h, m, sec = parts
        return int(h) * 3600 + int(m) * 60 + _int_floor_seconds_component(sec)
    elif len(parts) == 2:
        m, sec = parts
        return int(m) * 60 + _int_floor_seconds_component(sec)
    elif len(parts) == 1:
        # Support 'SS(.fff)' format (e.g., '58.2' -> 0:58)
        return _int_floor_seconds_component(parts[0])
    else:
        raise ValueError(f"Invalid time format '{s}'. Expected M:SS or H:MM:SS.")


def seconds_to_mmss_or_hhmmss(t: int) -> str:
    """Pretty printer for seconds: HH:MM:SS if >= 3600 else M:SS with minutes not zero-padded."""
    if t < 0:
        t = 0
    h = t // 3600
    r = t % 3600
    m = r // 60
    s = r % 60
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    else:
        return f"{m}:{s:02d}"


def seconds_to_hhmmss(t: int) -> str:
    """Always format seconds as HH:MM:SS with zero-padded hours."""
    if t < 0:
        t = 0
    h = t // 3600
    r = t % 3600
    m = r // 60
    s = r % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _clip_pre_post_s_for_event_type(event_type: str) -> Tuple[int, int]:
    et = str(event_type or "").strip().lower()
    if et in {"goal", "assist"}:
        return GOAL_CLIP_PRE_S, GOAL_CLIP_POST_S
    return EVENT_CLIP_PRE_S, EVENT_CLIP_POST_S


def forward_fill_header_labels(header_row: pd.Series) -> Dict[str, List[int]]:
    """
    Given a header row with merged-like cells (label then NaNs across its span),
    forward-fill labels across columns to group column indices by label.
    """
    labels_by_col: List[Optional[str]] = []
    current = None
    for c in range(len(header_row)):
        val = header_row.iloc[c]
        if pd.notna(val) and str(val).strip():
            current = str(val).strip()
        labels_by_col.append(current)

    groups: Dict[str, List[int]] = {}
    for idx, lab in enumerate(labels_by_col):
        if not lab:
            continue
        groups.setdefault(lab, []).append(idx)
    return groups


def _normalize_header_label(label: str) -> str:
    """Normalize header label for comparison (case/spacing/punctuation insensitive)."""
    return re.sub(r"[^a-z0-9]+", "", label.lower())


def _resolve_header_columns(groups: Dict[str, List[int]], *candidates: str) -> List[int]:
    """Return columns for the first candidate label that matches after normalization."""
    norm_map = {_normalize_header_label(k): v for k, v in groups.items() if k}
    for cand in candidates:
        key = _normalize_header_label(cand)
        cols = norm_map.get(key)
        if cols:
            return cols
    # Fallback: allow substring match to tolerate slight label variants
    for cand in candidates:
        key = _normalize_header_label(cand)
        for norm_label, cols in norm_map.items():
            if cols and key and key in norm_label:
                return cols
    return []


def extract_pairs_from_row(
    row: pd.Series, start_cols: List[int], end_cols: List[int]
) -> List[Tuple[str, str]]:
    """
    From start/end column groups, collect non-empty strings and pair positionally.
    Start/End order in the sheet can be higher->lower or lower->higher; pairing is positional only.
    If a value is a time (datetime.time or Timestamp with time), keep only hour:minute.
    """

    def format_cell(val) -> str:
        if pd.isna(val):
            return ""
        # Already string → return trimmed
        if isinstance(val, str):
            s = val.strip()
            # Treat textual NaN/None markers as empty cells
            if s.lower() in {"nan", "none"}:
                return ""
            return s
        # datetime.time (Excel time)
        if isinstance(val, datetime.time):
            return val.strftime("%H:%M")
        # pandas Timestamp (could include date/time)
        if isinstance(val, pd.Timestamp):
            if val.time() != datetime.time(0, 0):  # has time portion
                return val.strftime("%H:%M")
            else:
                return val.strftime("%Y-%m-%d")  # just a date
        # Fallback
        return str(val).strip()

    starts = [format_cell(row[c]) for c in start_cols if format_cell(row[c])]
    ends = [format_cell(row[c]) for c in end_cols if format_cell(row[c])]
    n = min(len(starts), len(ends))
    return [(starts[i], ends[i]) for i in range(n)]


# Historically we tried to normalize certain end-of-period times (e.g., 15:00 or 20:00)
# to 0:00 when sheets encoded the period end as its nominal start time. This proved
# brittle and could distort genuine in-period times (e.g., 12:00). We now leave
# scoreboard end times exactly as entered in the sheet.
def _normalize_sb_end_time(t: str) -> str:
    return t


@contextmanager
def _working_directory(path: Path) -> Iterator[None]:
    """Temporarily switch working directory, creating it if needed."""
    prev = Path.cwd()
    path.mkdir(parents=True, exist_ok=True)
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(prev)


def _duration_to_seconds(val: str) -> int:
    try:
        return parse_flex_time_to_seconds(str(val)) if val not in ("", None) else 0
    except Exception:
        return 0


def _seconds_to_compact_hms(t: int) -> str:
    """
    Compact duration formatting:
      - H:MM:SS if >= 3600
      - M:SS if >= 60
      - SS if < 60
    """
    if t < 0:
        t = 0
    if t < 60:
        return str(int(t))
    h = t // 3600
    r = t % 3600
    m = r // 60
    s = r % 60
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def _format_duration(total_seconds: int) -> str:
    return _seconds_to_compact_hms(total_seconds)


def _autosize_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Best-effort column auto-width for Excel sheets."""
    try:
        col_widths = []
        for col in df.columns:
            header_len = max([len(x) for x in str(col).splitlines()] or [len(str(col))])
            max_len = max([header_len] + [len(str(x)) for x in df[col].astype(str).fillna("")])
            col_widths.append(min(max_len + 2, 80))
        if writer.engine == "openpyxl":
            from openpyxl.utils import get_column_letter

            ws = writer.sheets.get(sheet_name)
            if ws:
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
        elif writer.engine == "xlsxwriter":
            ws = writer.sheets.get(sheet_name)
            if ws:
                for i, width in enumerate(col_widths):
                    ws.set_column(i, i, width)
    except Exception:
        pass


def _wrap_header_after_words(header: str, *, words_per_line: int = 3) -> str:
    if not header:
        return header
    parts = str(header).strip().split()
    if len(parts) <= words_per_line:
        return str(header)
    lines = [" ".join(parts[i : i + words_per_line]) for i in range(0, len(parts), words_per_line)]
    return "\n".join(lines)


def _write_styled_xlsx_table(
    xlsx_path: Path,
    df: pd.DataFrame,
    *,
    sheet_name: str,
    title: str,
    words_per_line: int = 2,
    number_formats: Optional[Dict[str, str]] = None,
    text_columns: Optional[List[str]] = None,
    align_right_columns: Optional[List[str]] = None,
) -> None:
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=words_per_line) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            _apply_excel_table_style(writer, sheet_name, title=title, df=df_excel)
            try:
                ws = writer.sheets.get(sheet_name)
                if ws is not None:
                    header_row = 2
                    data_start_row = 3
                    nrows = int(getattr(df_excel, "shape", (0, 0))[0] or 0)
                    if nrows > 0:
                        header_cells = list(ws[header_row])
                        col_idx_by_name: Dict[str, int] = {}
                        for idx, cell in enumerate(header_cells, start=1):
                            if cell.value is None:
                                continue
                            col_idx_by_name[str(cell.value).replace("\n", " ").strip()] = idx

                        def _apply_number_format(col_name: str, fmt: str) -> None:
                            col_idx = col_idx_by_name.get(str(col_name).strip())
                            if not col_idx:
                                return
                            for r in range(data_start_row, data_start_row + nrows):
                                ws.cell(row=r, column=col_idx).number_format = fmt

                        def _apply_alignment_right(col_name: str) -> None:
                            col_idx = col_idx_by_name.get(str(col_name).strip())
                            if not col_idx:
                                return
                            from openpyxl.styles import Alignment

                            for r in range(data_start_row, data_start_row + nrows):
                                cell = ws.cell(row=r, column=col_idx)
                                cell.alignment = Alignment(
                                    horizontal="right", vertical=cell.alignment.vertical or "center"
                                )

                        # Apply explicit number formats (e.g., 1-decimal percent columns).
                        for col_name, fmt in (number_formats or {}).items():
                            _apply_number_format(col_name, fmt)

                        # Force key text columns to "Text" format (best-effort; Excel spellcheck
                        # disable is not supported via openpyxl).
                        auto_text_cols = {"player", "Player", "teammate", "Teammate"}
                        for col_name in set((text_columns or [])) | auto_text_cols:
                            _apply_number_format(col_name, "@")

                        # Right-align duration/time columns (best-effort).
                        auto_right_cols: set[str] = set(align_right_columns or [])
                        for name in col_idx_by_name.keys():
                            n = name.strip()
                            if any(tok in n for tok in ["TOI", "Time", "Overlap", "Average Shift", "Median Shift", "Longest Shift", "Shortest Shift"]):
                                auto_right_cols.add(n)
                        for col_name in auto_right_cols:
                            _apply_alignment_right(col_name)
            except Exception:
                pass
            _autosize_columns(writer, sheet_name, df_excel)
    except Exception:
        pass


def _apply_excel_header_wrap(writer: pd.ExcelWriter, sheet_name: str, *, header_row: int = 1) -> None:
    """
    Enable wrap-text for a header row when using openpyxl.
    """
    try:
        if writer.engine != "openpyxl":
            return
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            return
        from openpyxl.styles import Alignment

        max_lines = 1
        for cell in ws[header_row]:
            if cell.value:
                max_lines = max(max_lines, len(str(cell.value).splitlines()))
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[header_row].height = max(15.0, 15.0 * max_lines)
    except Exception:
        return


def _apply_excel_table_style(
    writer: pd.ExcelWriter,
    sheet_name: str,
    *,
    title: str,
    df: pd.DataFrame,
) -> None:
    """
    Apply a simple "teal header + banded gray rows" theme to a sheet written
    with pandas. Assumes the DataFrame was written with `startrow=1` so that:
      - Row 1 is available for the title
      - Row 2 is the header row
      - Row 3.. are data rows
    """
    try:
        if writer.engine != "openpyxl":
            return
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            return

        ncols = int(getattr(df, "shape", (0, 0))[1] or 0)
        nrows = int(getattr(df, "shape", (0, 0))[0] or 0)
        if ncols <= 0:
            return

        from openpyxl.styles import Alignment, Font, PatternFill

        teal_fill = PatternFill(fill_type="solid", start_color="FF009688", end_color="FF009688")
        header_font = Font(color="FF000000", bold=True)
        title_font = Font(color="FF000000", bold=True, size=14)
        band_a = PatternFill(fill_type="solid", start_color="FFE6E6E6", end_color="FFE6E6E6")
        band_b = PatternFill(fill_type="solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
        from openpyxl.styles import Border, Side

        white_side = Side(style="thin", color="FFFFFFFF")
        white_border = Border(
            left=white_side,
            right=white_side,
            top=white_side,
            bottom=white_side,
        )

        title_row = 1
        header_row = 2
        data_start_row = 3

        # Title row (merged across all columns)
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.value = title
        title_cell.fill = teal_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 24.0

        # Ensure the merged region is fully teal in Excel viewers that don't
        # propagate styles from the top-left cell.
        for c in range(1, ncols + 1):
            cell = ws.cell(row=title_row, column=c)
            cell.fill = teal_fill

        # Header styling
        for c in range(1, ncols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = teal_fill
            cell.font = header_font

        _apply_excel_header_wrap(writer, sheet_name, header_row=header_row)

        # Banded rows (data)
        for i in range(nrows):
            r = data_start_row + i
            fill = band_a if (i % 2 == 0) else band_b
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = fill

        # White grid borders (title + header + data)
        last_row = data_start_row + max(nrows, 1) - 1 if nrows > 0 else header_row
        for r in range(title_row, last_row + 1):
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).border = white_border

        # Right-align duration/time-like columns (strings such as '54', '1:02', '2:10:03').
        try:
            header_cells = list(ws[header_row])
            right_cols: List[int] = []
            for idx, cell in enumerate(header_cells, start=1):
                if cell.value is None:
                    continue
                name = str(cell.value).replace("\n", " ").strip()
                if any(
                    tok in name
                    for tok in [
                        "TOI",
                        "Time",
                        "Overlap",
                        "Average Shift",
                        "Median Shift",
                        "Longest Shift",
                        "Shortest Shift",
                        "Video",
                        "Duration",
                    ]
                ):
                    right_cols.append(idx)
            if right_cols and nrows > 0:
                for i in range(nrows):
                    r = data_start_row + i
                    for c in right_cols:
                        cell = ws.cell(row=r, column=c)
                        cell.alignment = Alignment(
                            horizontal="right", vertical=cell.alignment.vertical or "center"
                        )
        except Exception:
            pass

        # Freeze the first column (Player/Stat) and the title+header rows so player
        # names stay visible while scrolling in Excel/LibreOffice.
        ws.freeze_panes = "B3"
    except Exception:
        return


def _collect_sheet_jerseys(
    xls_path: Path, sheet_name: Optional[str], keep_goalies: bool
) -> set[str]:
    df = pd.read_excel(xls_path, sheet_name=(0 if sheet_name is None else sheet_name), header=None)
    (
        used_event_log,
        _video_pairs_by_player,
        sb_pairs_by_player,
        _conv_segments_by_period,
        event_log_context,
    ) = _parse_event_log_layout(df)
    if not used_event_log:
        (
            _video_pairs_by_player,
            sb_pairs_by_player,
            _conv_segments_by_period,
            _validation_errors,
        ) = _parse_per_player_layout(df, keep_goalies=keep_goalies, skip_validation=True)

    jerseys: set[str] = set()
    for pk in sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(pk)
        if norm:
            jerseys.add(norm)

    if not jerseys and event_log_context is not None:
        for lst in (event_log_context.team_roster or {}).values():
            for num in lst or []:
                norm = _normalize_jersey_number(num)
                if norm:
                    jerseys.add(norm)

    return jerseys


# ----------------------------- parsing sheet -----------------------------


def find_period_blocks(df: pd.DataFrame) -> List[Tuple[int, int, int]]:
    """
    Returns a list of (period_number, start_row_idx, end_row_idx_exclusive) for each 'Period X' block.
    """
    # Use position-based indexing for robustness: some Excel files may not
    # create a column label "0" even with header=None, so df[0] can KeyError.
    col0 = df.iloc[:, 0]
    starts: List[int] = []
    periods: List[int] = []
    for i, v in col0.items():
        pnum = parse_period_label(v)
        if pnum is None:
            continue
        starts.append(i)
        periods.append(pnum)
    starts.append(len(df))
    blocks = []
    for i in range(len(periods)):
        blocks.append((periods[i], starts[i], starts[i + 1]))
    return blocks


def find_header_row(df: pd.DataFrame, start: int, end: int) -> Optional[int]:
    """
    Within [start, end), locate the header row (typically the row with 'Jersey No' in col 0).
    Often pattern is: 'Period', blank, header.
    """
    for r in range(start, min(end, start + 12)):
        if str(df.iloc[r, 0]).strip().lower() in ["jersey no", "jersey number"]:
            return r
    # Fallback (Period + blank + header)
    return start + 2 if start + 2 < end else None


# ----------------------------- goals parsing -----------------------------


@dataclass
class GoalEvent:
    kind: str  # "GF" or "GA"
    period: int
    t_str: str
    scorer: Optional[str] = None
    assists: List[str] = field(default_factory=list)
    t_sec: int = field(init=False)
    is_game_tying: bool = False
    is_game_winning: bool = False

    def __post_init__(self) -> None:
        self.t_str = self.t_str.strip()
        self.assists = [a for a in self.assists if a]
        self.t_sec = parse_flex_time_to_seconds(self.t_str)

    def __str__(self) -> str:  # preserve prior textual representation
        return f"{self.kind}:{self.period}/{self.t_str}"

    __repr__ = __str__


def parse_goal_token(token: str) -> GoalEvent:
    """
    Token: GF:2/13:45, GA:1/05:12, or GF:OT/0:45 (case-insensitive on GF/GA and OT).
    """
    token = token.strip()
    m = re.fullmatch(r"(?i)(GF|GA)\s*:\s*([^/]+)\s*/\s*([0-9:]+)", token)
    if not m:
        raise ValueError(f"Bad goal token '{token}'. Expected GF:period/time or GA:period/time")
    kind = m.group(1).upper()
    period = parse_period_token(m.group(2))
    if period is None:
        raise ValueError(
            f"Bad goal token '{token}': invalid period '{m.group(2)}' (use 1/2/3/4 or OT)."
        )
    t_str = m.group(3)
    return GoalEvent(kind, period, t_str)


def parse_period_token(x: object) -> Optional[int]:
    """
    Parse a period token from various sources:
      - ints ("1", 1)
      - labels ("Period 1", "1st Period")
      - overtime ("OT", "Overtime")

    Convention: OT == period 4.
    """
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, int):
        return x
    try:
        s = str(x).strip()
    except Exception:
        return None
    if not s:
        return None
    # Labels like "1st Period" / "OT"
    p = parse_period_label(s)
    if p is not None:
        return p
    # Plain number like "1"
    if re.fullmatch(r"\d+", s):
        try:
            return int(s)
        except Exception:
            return None
    # Fallback: any number embedded in the token
    m = re.search(r"(\d+)", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def _annotate_goal_roles(goals: List[GoalEvent]) -> None:
    """
    Annotate GoalEvent entries in-place with game-tying / game-winning flags.

    - Game-tying goal: GF that changes the score from trailing to tied.
    - Game-winning goal: when our team wins (GF > GA at end), the first goal
      scored by our team after the last time the game was tied.
    """
    if not goals:
        return

    # Work on indices into the original list to avoid reordering it.
    order = sorted(range(len(goals)), key=lambda idx: (goals[idx].period, goals[idx].t_sec))
    before_scores: List[Tuple[int, int]] = []
    after_scores: List[Tuple[int, int]] = []
    gf = 0
    ga = 0
    last_tie_pos = -1

    for pos, idx in enumerate(order):
        ev = goals[idx]
        before_scores.append((gf, ga))
        if ev.kind == "GF":
            gf += 1
        else:
            ga += 1
        after_scores.append((gf, ga))
        if gf == ga:
            last_tie_pos = pos

    final_gf, final_ga = gf, ga

    # Reset flags then mark game-tying goals.
    for pos, idx in enumerate(order):
        ev = goals[idx]
        ev.is_game_tying = False
        ev.is_game_winning = False
        bf_for, bf_against = before_scores[pos]
        af_for, af_against = after_scores[pos]
        if ev.kind == "GF" and bf_for < bf_against and af_for == af_against:
            ev.is_game_tying = True

    # Game-winning goal (if we finished ahead).
    if final_gf > final_ga and order:
        win_pos = last_tie_pos + 1
        if 0 <= win_pos < len(order):
            ev = goals[order[win_pos]]
            if ev.kind == "GF":
                ev.is_game_winning = True


def _normalize_jersey_number(token: Any) -> Optional[str]:
    if token is None:
        return None
    try:
        text = str(token).strip()
    except Exception:
        return None
    if not text:
        return None
    m = re.search(r"(\d+)", text)
    if not m:
        return None
    num = m.group(1).lstrip("0")
    return num or "0"


def _extract_jersey_number(cell: Any) -> Optional[str]:
    if isinstance(cell, dict):
        cell = cell.get("text") or cell.get("link") or ""
    return _normalize_jersey_number(cell)


def _scoring_numbers_from_row(row: Any) -> Tuple[Optional[str], List[str]]:
    if not isinstance(row, dict):
        return None, []
    scorer = _extract_jersey_number(row.get("goal"))
    assists: List[str] = []
    for key in ("assist1", "assist2"):
        num = _extract_jersey_number(row.get(key))
        if num:
            assists.append(num)
    return scorer, assists


def _infer_t2s_from_filename(path: Path) -> Optional[int]:
    stem = path.stem
    m = re.search(r"^(.*)-(\d+)$", stem)
    if not m:
        return None
    try:
        game_id = int(m.group(2))
    except Exception:
        return None
    # Treat only sufficiently large numeric suffixes as TimeToScore ids.
    # Smaller suffixes (e.g., 'chicago-1') are considered part of the game name.
    return game_id if game_id >= 10000 else None


def _base_label_from_path(path: Any) -> str:
    p = Path(path)
    stem = p.stem
    # If the filename encodes a T2S id as a trailing numeric suffix (>= 10000),
    # drop that suffix from the label. Otherwise keep the stem as-is.
    m = re.search(r"^(.*)-(\d+)$", stem)
    if m:
        try:
            suffix_num = int(m.group(2))
        except Exception:
            suffix_num = None
        if suffix_num is not None and suffix_num >= 10000:
            base = m.group(1) or stem
            # Treat companion long sheets (e.g., 'game-long-54111.xlsx') as belonging
            # to the same game label as their non-long counterpart.
            if base.endswith("-long"):
                base = base[: -len("-long")]
            return base or stem
    label = stem or p.name
    if label.endswith("-long"):
        label = label[: -len("-long")]
    return label


def _is_long_sheet_path(path: Path) -> bool:
    try:
        stem = Path(path).stem
    except Exception:
        return False
    return re.search(r"(?i)(?:^|-)long(?:-|$)", stem) is not None


def _select_tracking_output_video(video_dir: Path) -> Optional[Path]:
    """
    Choose a `tracking_output-with-audio*.mp4` file in `video_dir`.

    If any numbered `tracking_output-with-audio-<N>.mp4` files exist, return the one
    with the largest N. Otherwise fall back to `tracking_output-with-audio.mp4`.
    """
    try:
        video_dir = Path(video_dir)
    except Exception:
        return None

    best_num: Optional[int] = None
    best_path: Optional[Path] = None
    try:
        for p in video_dir.glob("tracking_output-with-audio-*.mp4"):
            m = re.match(r"^tracking_output-with-audio-(\d+)\.mp4$", p.name)
            if not m:
                continue
            try:
                n = int(m.group(1))
            except Exception:
                continue
            if best_num is None or n > best_num:
                best_num = n
                best_path = p
    except Exception:
        best_path = None

    if best_path is not None and best_path.exists():
        return best_path

    plain = video_dir / "tracking_output-with-audio.mp4"
    return plain if plain.exists() else None


def _find_tracking_output_video_for_sheet_path(sheet_path: Path) -> Optional[Path]:
    """
    Locate the tracking video for a sheet assumed to live under `<game_dir>/stats/`.
    """
    try:
        sheet_path = Path(sheet_path)
    except Exception:
        return None
    stats_dir = sheet_path.parent
    video_dir = stats_dir.parent if stats_dir.name.lower() == "stats" else stats_dir
    return _select_tracking_output_video(video_dir)


@dataclass(frozen=True)
class InputEntry:
    path: Optional[Path]
    side: Optional[str]
    t2s_id: Optional[int] = None
    label: Optional[str] = None


def _parse_t2s_spec(token: Any) -> Optional[Tuple[int, Optional[str], Optional[str]]]:
    """
    Parse a TimeToScore spec token, formatted like:
      - 51602
      - 51602:HOME
      - 51602:HOME:stockton-r2
      - t2s=51602:HOME:stockton-r2

    Returns (t2s_id, side, label).
    """
    raw = str(token or "").strip()
    if not raw:
        return None

    m = re.match(r"(?i)^\s*(?:t2s\s*=\s*)?(\d+)(.*)$", raw)
    if not m:
        return None
    try:
        t2s_id = int(m.group(1))
    except Exception:
        return None

    rest = (m.group(2) or "").strip()
    if rest.startswith(":"):
        rest = rest[1:]
    rest_parts = [p.strip() for p in rest.split(":")] if rest else []

    side: Optional[str] = None
    label: Optional[str] = None
    if rest_parts:
        first = rest_parts[0].upper()
        if first in {"HOME", "AWAY"}:
            side = first.lower()
            label = ":".join(rest_parts[1:]).strip() if len(rest_parts) > 1 else None
        else:
            label = ":".join(rest_parts).strip()
    if label == "":
        label = None
    return t2s_id, side, label


def _parse_t2s_only_token(token: str) -> Optional[Tuple[int, Optional[str], Optional[str]]]:
    """
    Parse a special `--file-list` token representing a TimeToScore-only game
    (no spreadsheets present), formatted like:
      - t2s=51602
      - t2s=51602:HOME
      - t2s=51602:HOME:stockton-r2
    """
    raw = str(token or "").strip()
    if not re.match(r"(?i)^\s*t2s\s*=", raw):
        return None
    return _parse_t2s_spec(raw)


def _parse_input_token(token: str, base_dir: Optional[Path] = None) -> Tuple[Path, Optional[str]]:
    raw = token.strip()
    side: Optional[str] = None
    if ":" in raw:
        raw_path, suffix = raw.rsplit(":", 1)
        if suffix.upper() in {"HOME", "AWAY"}:
            side = suffix.lower()
            raw = raw_path
    p = Path(raw).expanduser()
    if base_dir and not p.is_absolute():
        p = (base_dir / p).resolve()
    return p, side


def _is_spreadsheet_input_path(path: Path) -> bool:
    try:
        return path.is_file() and path.suffix.lower() in {".xls", ".xlsx"}
    except Exception:
        return False


def _should_ignore_spreadsheet_input(path: Path) -> bool:
    name = path.name
    if not name:
        return True
    if name.startswith(".") or name.startswith("~$"):
        return True
    stem = path.stem.lower()
    if stem == "goals":
        return True
    if stem.startswith("player_stats"):
        return True
    return False


def _discover_spreadsheet_inputs_in_dir(dir_path: Path) -> List[Path]:
    try:
        paths = [p for p in dir_path.iterdir() if _is_spreadsheet_input_path(p)]
    except Exception:
        return []
    out: List[Path] = []
    for p in paths:
        if _should_ignore_spreadsheet_input(p):
            continue
        out.append(p)
    out.sort(key=lambda p: p.name.lower())
    return out


def _expand_dir_input_to_game_sheets(dir_path: Path) -> List[Path]:
    """
    Expand a directory passed to --input into the game sheet(s) inside:
      - exactly one non-'-long' shift sheet (primary)
      - zero or more companion '*-long*' sheets

    Also supports passing a game directory by checking `<dir>/stats/` as a fallback.
    """
    candidates = _discover_spreadsheet_inputs_in_dir(dir_path)
    stats_dir = dir_path / "stats"
    if not candidates and stats_dir.is_dir():
        candidates = _discover_spreadsheet_inputs_in_dir(stats_dir)
        dir_path = stats_dir

    if not candidates:
        raise ValueError(f"No input .xls/.xlsx sheets found in {dir_path}")

    # Directory inputs are expected to correspond to a single game label.
    by_label: Dict[str, List[Path]] = {}
    for p in candidates:
        by_label.setdefault(_base_label_from_path(p), []).append(p)
    if len(by_label) != 1:
        labels = ", ".join(sorted(by_label.keys()))
        raise ValueError(
            f"Directory {dir_path} contains multiple game labels ({labels}); "
            "pass a specific file path or use --file-list."
        )

    only_label = next(iter(by_label.keys()))
    paths = by_label[only_label]
    primaries = [p for p in paths if not _is_long_sheet_path(p)]
    long_paths = [p for p in paths if _is_long_sheet_path(p)]
    if not primaries and long_paths:
        raise ValueError(
            f"Directory {dir_path} has only '*-long*' sheets for {only_label}; "
            "add the primary shift sheet or pass it explicitly."
        )
    if len(primaries) != 1:
        raise ValueError(
            f"Directory {dir_path} has {len(primaries)} primary sheets for {only_label}; "
            f"expected exactly 1: {[p.name for p in primaries]}"
        )
    return [primaries[0]] + long_paths


def load_goals(goals_inline: Iterable[str], goals_file: Optional[Path]) -> List[GoalEvent]:
    events: List[GoalEvent] = []
    # Inline
    for tok in goals_inline or []:
        if not tok:
            continue
        events.append(parse_goal_token(tok))
    # File
    if goals_file:
        with goals_file.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                # Strip UTF-8 BOM if present (common when files are created on Windows).
                line = line.lstrip("\ufeff")
                if not line or line.startswith("#"):
                    continue
                events.append(parse_goal_token(line))
    return events


def _format_goal_time_cell(val: Any) -> Optional[str]:
    """
    Convert a goal time cell (often an Excel time) into an M:SS-style string
    compatible with parse_flex_time_to_seconds.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime.time):
        # Excel stores times as H:M; reuse the same formatting as shift sheets
        return val.strftime("%H:%M")
    if isinstance(val, pd.Timestamp):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if not s:
        return None
    parts = s.split(":")
    # Convert H:MM:SS (Excel-style) into M:SS with minutes = H*60 + MM
    if len(parts) == 3:
        try:
            h = int(parts[0])
            m = int(parts[1])
            sec = int(parts[2])
            total_min = h * 60 + m
            return f"{total_min}:{sec:02d}"
        except Exception:
            pass
    return s


def _goals_from_goals_xlsx(goals_xlsx: Path) -> List[GoalEvent]:
    """
    Parse goals from a goals.xlsx sheet that has side-by-side GF / GA tables.

    Layout example (no headers row index assumed):

        row 0:  'GF'  ...  'GA' ...
        row 1:  (blank)
        row 2:  'Period' 'Time' 'Goal' 'Assist 1' 'Assist 2' ... 'Period' 'Time' 'Goal' 'Assist 1' 'Assist 2'
        row 3+: data rows
    """
    if not goals_xlsx.exists():
        return []

    df = pd.read_excel(goals_xlsx, header=None)
    nrows, ncols = df.shape

    def _find_label(label: str) -> Optional[Tuple[int, int]]:
        target = _normalize_header_label(label)
        for r in range(nrows):
            for c in range(ncols):
                val = df.iat[r, c]
                if pd.isna(val):
                    continue
                if _normalize_header_label(str(val)) == target:
                    return r, c
        return None

    def _parse_table(label_rc: Tuple[int, int], kind: str) -> List[GoalEvent]:
        base_r, base_c = label_rc
        # Find header row: look a few rows below the label for a 'Period' cell
        header_r: Optional[int] = None
        for r in range(base_r + 1, min(nrows, base_r + 8)):
            for c in range(base_c, ncols):
                val = df.iat[r, c]
                if pd.isna(val):
                    continue
                if _normalize_header_label(str(val)) == "period":
                    header_r = r
                    break
            if header_r is not None:
                break
        if header_r is None:
            return []

        # Map header labels to columns starting from the table's base column
        label_to_col: Dict[str, int] = {}
        for c in range(base_c, ncols):
            val = df.iat[header_r, c]
            if pd.isna(val):
                continue
            key = _normalize_header_label(str(val))
            if key:
                label_to_col.setdefault(key, c)

        period_col = label_to_col.get("period")
        time_col = label_to_col.get("time")
        goal_col = label_to_col.get("goal")
        assist1_col = label_to_col.get("assist1") or label_to_col.get("assist") or label_to_col.get(
            "assist_1"
        )
        assist2_col = label_to_col.get("assist2") or label_to_col.get("assist_2")

        if period_col is None or time_col is None or goal_col is None:
            return []

        events: List[GoalEvent] = []
        for r in range(header_r + 1, nrows):
            period_val = df.iat[r, period_col]
            time_val = df.iat[r, time_col]
            goal_val = df.iat[r, goal_col]
            if all(pd.isna(x) for x in (period_val, time_val, goal_val)):
                continue
            period = parse_period_token(period_val)
            if period is None:
                continue
            t_str = _format_goal_time_cell(time_val)
            if not t_str:
                continue
            scorer = _extract_jersey_number(goal_val)
            assists: List[str] = []
            for col in (assist1_col, assist2_col):
                if col is None:
                    continue
                val = df.iat[r, col]
                num = _extract_jersey_number(val)
                if num:
                    assists.append(num)
            events.append(GoalEvent(kind, period, t_str, scorer=scorer, assists=assists))
        return events

    events: List[GoalEvent] = []
    gf_rc = _find_label("GF")
    ga_rc = _find_label("GA")
    if gf_rc:
        events.extend(_parse_table(gf_rc, "GF"))
    if ga_rc:
        events.extend(_parse_table(ga_rc, "GA"))
    events.sort(key=lambda e: (e.period, e.t_sec))
    return events


def goals_from_t2s(game_id: int, *, side: str) -> List[GoalEvent]:
    """
    Retrieve goals from TimeToScore for a game id and map them to GF/GA based on
    the selected side (home/away).
    """
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        details = f": {_t2s_api_import_error}" if _t2s_api_import_error else ""
        raise RuntimeError(f"TimeToScore API not available (failed to import hmlib.time2score.api){details}")

    info = t2s_api.get_game_details(game_id)
    if not isinstance(info, dict):
        raise RuntimeError(f"TimeToScore API returned invalid game details type for game {game_id}: {type(info)}")
    stats = info.get("stats")
    if not isinstance(stats, dict) or not stats:
        keys = ", ".join(sorted(list(info.keys()))) if isinstance(info, dict) else ""
        raise RuntimeError(
            f"TimeToScore returned no usable stats for game {game_id} (keys=[{keys}]); cannot compute goals."
        )

    home_sc = stats.get("homeScoring") or []
    away_sc = stats.get("awayScoring") or []

    events: List[GoalEvent] = []

    def _mk_event(kind: str, row: Any) -> Optional[GoalEvent]:
        if not isinstance(row, dict):
            return None
        period_val = row.get("period")
        time_val = row.get("time")
        if period_val is None or time_val is None:
            return None
        per = parse_period_token(period_val)
        if per is None:
            return None
        t_str = str(time_val).strip()
        # Normalize fractional seconds by flooring
        try:
            sec = parse_flex_time_to_seconds(t_str)
        except Exception:
            return None
        mm = sec // 60
        ss = sec % 60
        scorer_num, assist_nums = _scoring_numbers_from_row(row)
        return GoalEvent(kind, per, f"{mm}:{ss:02d}", scorer=scorer_num, assists=assist_nums)

    # Home goals are GF if side == home else GA
    for row in home_sc:
        ev = _mk_event("GF" if side == "home" else "GA", row)
        if ev:
            events.append(ev)
    # Away goals are GF if side == away else GA
    for row in away_sc:
        ev = _mk_event("GF" if side == "away" else "GA", row)
        if ev:
            events.append(ev)

    # Sort by period then time for determinism
    events.sort(key=lambda e: (e.period, e.t_sec))
    return events


# ----------------------------- core processing -----------------------------


def compute_interval_seconds(a: str, b: str) -> Tuple[int, int]:
    """
    Return (lo, hi) in seconds for a scoreboard interval defined by two times (order-agnostic).
    Works even if the scoreboard counts down (e.g., 15:00 -> 09:21) or up (e.g., 11:27 -> 29:54).
    """
    sa, sb = parse_flex_time_to_seconds(a), parse_flex_time_to_seconds(b)
    return (sa, sb) if sa <= sb else (sb, sa)


def interval_contains(t: int, lo: int, hi: int) -> bool:
    return lo <= t <= hi


def fmt_pairs_for_file(pairs: List[Tuple[str, str]]) -> str:
    return "\n".join(f"{a} {b}" for a, b in pairs)


def _fmt_per_shift(numer: int, shifts: int) -> str:
    if shifts <= 0:
        return ""
    return f"{(numer / shifts):.2f}"


def _fmt_plus_minus(x: int) -> str:
    if x == 0:
        return "0"
    return f"{x:+d}"


def _merge_intervals(intervals: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    if not intervals:
        return []
    merged: List[Tuple[int, int]] = []
    for lo, hi in sorted(intervals, key=lambda x: (x[0], x[1])):
        if not merged:
            merged.append((lo, hi))
            continue
        prev_lo, prev_hi = merged[-1]
        if lo <= prev_hi:
            merged[-1] = (prev_lo, max(prev_hi, hi))
        else:
            merged.append((lo, hi))
    return merged


def _intersection_seconds(a: List[Tuple[int, int]], b: List[Tuple[int, int]]) -> int:
    """
    Sum the overlap duration between two lists of non-overlapping, sorted intervals.
    Intervals are treated as continuous with duration (hi - lo).
    """
    i = 0
    j = 0
    total = 0
    while i < len(a) and j < len(b):
        alo, ahi = a[i]
        blo, bhi = b[j]
        lo = max(alo, blo)
        hi = min(ahi, bhi)
        if hi > lo:
            total += hi - lo
        if ahi <= bhi:
            i += 1
        else:
            j += 1
    return total


def _compute_pair_on_ice_rows(
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    goals_by_period: Dict[int, List["GoalEvent"]],
) -> List[Dict[str, Any]]:
    """
    Produce directed (player, teammate) rows for pair on-ice overlap and goal +/- together.

    - overlap is computed from scoreboard shift intervals (unioned by period)
    - goal counts include goals during overlap, but skip goals at a player's shift start
      (matches the plus/minus behavior in `_compute_player_stats`)
    """
    all_players = sorted([p for p in sb_pairs_by_player.keys() if p])
    if not all_players:
        return []

    intervals_by_player_period: Dict[str, Dict[int, List[Tuple[int, int]]]] = {}
    raw_intervals_by_player_period: Dict[str, Dict[int, List[Tuple[int, int, Optional[int]]]]] = {}
    toi_sec_by_player: Dict[str, int] = {}

    for player in all_players:
        per_period: Dict[int, List[Tuple[int, int]]] = {}
        raw_by_period: Dict[int, List[Tuple[int, int, Optional[int]]]] = {}
        for period, a, b in sb_pairs_by_player.get(player, []):
            lo, hi = compute_interval_seconds(a, b)
            per_period.setdefault(period, []).append((lo, hi))
            try:
                start_sec: Optional[int] = parse_flex_time_to_seconds(a)
            except Exception:
                start_sec = None
            raw_by_period.setdefault(period, []).append((lo, hi, start_sec))
        merged_by_period: Dict[int, List[Tuple[int, int]]] = {}
        toi_total = 0
        for period, intervals in per_period.items():
            merged = _merge_intervals(intervals)
            merged_by_period[period] = merged
            toi_total += sum(hi - lo for lo, hi in merged)
        intervals_by_player_period[player] = merged_by_period
        raw_intervals_by_player_period[player] = raw_by_period
        toi_sec_by_player[player] = toi_total

    # Exclude players with no recorded TOI from pair tables (likely absent or a goalie).
    players = [p for p in all_players if int(toi_sec_by_player.get(p, 0) or 0) > 0]
    if not players:
        return []

    jersey_to_player: Dict[str, str] = {}
    for pk in players:
        jersey_part = str(pk).split("_", 1)[0]
        norm = _normalize_jersey_number(jersey_part)
        if norm:
            jersey_to_player.setdefault(norm, pk)

    def _on_ice_for_goal(player: str, period: int, t_sec: int) -> bool:
        # Match `_compute_player_stats` behavior:
        #  - count a goal if it falls within any shift interval for the player
        #  - BUT skip a goal that occurs exactly at the *start* of that specific shift
        #    (important when a goal occurs at the boundary between consecutive shifts).
        for lo, hi, start_sec in raw_intervals_by_player_period.get(player, {}).get(period, []):
            if not interval_contains(t_sec, lo, hi):
                continue
            if start_sec is not None and t_sec == start_sec:
                continue
            return True
        return False

    def _on_ice_any(player: str, period: int, t_sec: int) -> bool:
        for lo, hi in intervals_by_player_period.get(player, {}).get(period, []):
            if interval_contains(t_sec, lo, hi):
                return True
        return False

    overlap_by_pair: Dict[Tuple[str, str], int] = {}
    for i in range(len(players)):
        for j in range(i + 1, len(players)):
            a = players[i]
            b = players[j]
            overlap = 0
            a_periods = intervals_by_player_period.get(a, {})
            b_periods = intervals_by_player_period.get(b, {})
            for period in set(a_periods.keys()) & set(b_periods.keys()):
                overlap += _intersection_seconds(a_periods.get(period, []), b_periods.get(period, []))
            overlap_by_pair[(a, b)] = overlap

    goals_by_pair: Dict[Tuple[str, str], Tuple[int, int]] = {}
    player_goals_on_ice_by_pair: Dict[Tuple[str, str], int] = {}
    player_assists_on_ice_by_pair: Dict[Tuple[str, str], int] = {}
    collab_goals_by_pair: Dict[Tuple[str, str], int] = {}
    collab_assists_by_pair: Dict[Tuple[str, str], int] = {}
    player_total_pm: Dict[str, int] = {p: 0 for p in players}
    for period, goals in (goals_by_period or {}).items():
        if not goals:
            continue
        for ev in goals:
            on_ice = [p for p in players if _on_ice_for_goal(p, period, int(ev.t_sec))]
            on_ice_any = [p for p in players if _on_ice_any(p, period, int(ev.t_sec))]
            for p in on_ice:
                if ev.kind == "GF":
                    player_total_pm[p] = int(player_total_pm.get(p, 0) or 0) + 1
                else:
                    player_total_pm[p] = int(player_total_pm.get(p, 0) or 0) - 1

            # Attribute goal/assist counts only for our team's goals (GF) when scorer/assists are known.
            if ev.kind == "GF":
                scorer_key: Optional[str] = None
                if ev.scorer:
                    scorer_key = jersey_to_player.get(_normalize_jersey_number(ev.scorer) or "")
                assist_keys: List[str] = []
                for a in (ev.assists or []):
                    ak = jersey_to_player.get(_normalize_jersey_number(a) or "")
                    if ak:
                        assist_keys.append(ak)

                if scorer_key and scorer_key in on_ice_any:
                    for teammate in on_ice_any:
                        if teammate == scorer_key:
                            continue
                        player_goals_on_ice_by_pair[(scorer_key, teammate)] = int(
                            player_goals_on_ice_by_pair.get((scorer_key, teammate), 0) or 0
                        ) + 1

                for ak in assist_keys:
                    if ak not in on_ice_any:
                        continue
                    for teammate in on_ice_any:
                        if teammate == ak:
                            continue
                        player_assists_on_ice_by_pair[(ak, teammate)] = int(
                            player_assists_on_ice_by_pair.get((ak, teammate), 0) or 0
                        ) + 1

                # Direct collaboration: scorer<->assister pairs on the same goal.
                if scorer_key:
                    for ak in assist_keys:
                        if ak == scorer_key:
                            continue
                        collab_goals_by_pair[(scorer_key, ak)] = int(
                            collab_goals_by_pair.get((scorer_key, ak), 0) or 0
                        ) + 1
                        collab_assists_by_pair[(ak, scorer_key)] = int(
                            collab_assists_by_pair.get((ak, scorer_key), 0) or 0
                        ) + 1

            if len(on_ice) < 2:
                continue
            for i in range(len(on_ice)):
                for j in range(i + 1, len(on_ice)):
                    a, b = sorted((on_ice[i], on_ice[j]))
                    gf, ga = goals_by_pair.get((a, b), (0, 0))
                    if ev.kind == "GF":
                        gf += 1
                    else:
                        ga += 1
                    goals_by_pair[(a, b)] = (gf, ga)

    rows: List[Dict[str, Any]] = []
    for player in players:
        player_toi = int(toi_sec_by_player.get(player, 0) or 0)
        for teammate in players:
            if teammate == player:
                continue
            a, b = sorted((player, teammate))
            overlap = int(overlap_by_pair.get((a, b), 0) or 0)
            gf, ga = goals_by_pair.get((a, b), (0, 0))
            pct = (100.0 * overlap / player_toi) if player_toi > 0 else 0.0
            rows.append(
                {
                    "player": player,
                    "teammate": teammate,
                    "shift_games": 1,
                    "player_toi_seconds": player_toi,
                    "overlap_seconds": overlap,
                    "overlap_pct": pct,
                    "gf_together": int(gf),
                    "ga_together": int(ga),
                    "player_goals_on_ice_together": int(player_goals_on_ice_by_pair.get((player, teammate), 0) or 0),
                    "player_assists_on_ice_together": int(
                        player_assists_on_ice_by_pair.get((player, teammate), 0) or 0
                    ),
                    "goals_collab_with_teammate": int(collab_goals_by_pair.get((player, teammate), 0) or 0),
                    "assists_collab_with_teammate": int(collab_assists_by_pair.get((player, teammate), 0) or 0),
                    "player_total_plus_minus": int(player_total_pm.get(player, 0) or 0),
                    "teammate_total_plus_minus": int(player_total_pm.get(teammate, 0) or 0),
                    "plus_minus_together": int(gf) - int(ga),
                }
            )
    return rows


def summarize_shift_lengths_sec(pairs: List[Tuple[str, str]]) -> Dict[str, str]:
    """
    Given scoreboard time string pairs, compute durations in seconds and summarize.
    """
    lengths = []
    for a, b in pairs:
        lo, hi = compute_interval_seconds(a, b)
        lengths.append(hi - lo)
    if not lengths:
        return {
            "num_shifts": "0",
            "toi_total": "0",
            "toi_avg": "0",
            "toi_median": "0",
            "toi_longest": "0",
            "toi_shortest": "0",
        }
    return {
        "num_shifts": str(len(lengths)),
        "toi_total": _format_duration(sum(lengths)),
        "toi_avg": _format_duration(int(sum(lengths) / len(lengths))),
        "toi_median": _format_duration(int(statistics.median(lengths))),
        "toi_longest": _format_duration(max(lengths)),
        "toi_shortest": _format_duration(min(lengths)),
    }


def per_period_toi(pairs_by_period: Dict[int, List[Tuple[str, str]]]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for period, pairs in pairs_by_period.items():
        total = 0
        for a, b in pairs:
            lo, hi = compute_interval_seconds(a, b)
            total += hi - lo
        out[period] = _format_duration(total)
    return out


@dataclass
class EventLogContext:
    event_counts_by_player: Dict[str, Dict[str, int]]
    event_counts_by_type_team: Dict[Tuple[str, str], int]
    event_instances: Dict[Tuple[str, str], List[Dict[str, Any]]]
    event_player_rows: List[Dict[str, Any]]
    team_roster: Dict[str, List[int]]
    team_excluded: Dict[str, List[int]]


@dataclass(frozen=True)
class LongEvent:
    event_type: str
    team: str
    period: int
    video_s: Optional[int]
    game_s: Optional[int]
    jerseys: Tuple[int, ...] = ()


def _parse_long_mmss_time_to_seconds(cell: Any) -> Optional[int]:
    """
    Parse times as they appear in the '-long' sheets.

    These sheets commonly store MM:SS values as Excel time-of-day cells, so
    pandas yields datetime.time like 23:56:00 (meaning 23:56, not 23 hours).

    Accepts:
      - datetime.time / pd.Timestamp: interpret hour as minutes, minute as seconds
      - strings like '24:50' or '00:56:00': interpret as MM:SS (ignore 3rd component)
    """
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    if isinstance(cell, datetime.time):
        return int(cell.hour) * 60 + int(cell.minute)
    if isinstance(cell, pd.Timestamp):
        return int(cell.hour) * 60 + int(cell.minute)
    try:
        s = str(cell).strip()
    except Exception:
        return None
    if not s:
        return None
    # Skip header labels like 'Video Time'
    if s.lower() in {"video time", "scoreboard", "team"}:
        return None
    parts = s.split(":")
    try:
        if len(parts) == 3:
            a_s, b_s, c_s = parts
            a = int(a_s)
            b = int(b_s)
            c = int(c_s)
            # Some sheets emit long video times as H:MM:SS (e.g., 1:09:19, 1:00:00).
            # Others emit MM:SS as a time-of-day-like string HH:MM:SS (e.g., 23:56:00 meaning 23:56).
            #
            # Heuristic:
            #   - If seconds are non-zero, it's almost certainly H:MM:SS.
            #   - If seconds are zero, treat small unpadded leading fields (e.g., "1:00:00") as H:MM:SS,
            #     otherwise treat as MM:SS encoded as HH:MM:SS.
            if c != 0:
                return a * 3600 + b * 60 + c
            if 0 < a < 4 and len(a_s) == 1:
                return a * 3600 + b * 60 + c
            return a * 60 + b
        if len(parts) == 2:
            mm, ss = parts
            return int(mm) * 60 + int(ss)
        if len(parts) == 1:
            return int(float(parts[0]))
    except Exception:
        return None
    return None


def _parse_long_team(cell: Any) -> Optional[str]:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    try:
        s = str(cell).strip()
    except Exception:
        return None
    if not s:
        return None
    sl = s.lower()
    if "blue" in sl or sl.startswith("blu"):
        return "Blue"
    if "white" in sl or sl.startswith("whi"):
        return "White"
    return None


def _extract_jerseys_from_cell(cell: Any) -> List[int]:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return []
    if isinstance(cell, (int, float)) and not pd.isna(cell):
        try:
            n = int(cell)
        except Exception:
            return []
        return [n] if 1 <= n <= 98 else []
    try:
        s = str(cell).strip()
    except Exception:
        return []
    if not s:
        return []
    # Don't treat times as jersey numbers.
    if re.match(r"^\d{1,3}:\d{2}(:\d{2})?$", s):
        return []
    nums: List[int] = []
    for m in re.finditer(r"#?(\d{1,2})(?!\d)", s):
        try:
            n = int(m.group(1))
        except Exception:
            continue
        if 1 <= n <= 98:
            nums.append(n)
    # Dedupe while preserving order.
    seen: set[int] = set()
    out: List[int] = []
    for n in nums:
        if n in seen:
            continue
        seen.add(n)
        out.append(n)
    return out


def _parse_long_left_event_table(
    df: pd.DataFrame,
) -> Tuple[List[LongEvent], List[Dict[str, Any]], Dict[str, set[int]]]:
    """
    Parse the leftmost per-period event table found in '*-long*.xlsx' sheets.

    Returns:
      - list of LongEvent entries (for event summaries/clips)
      - list of goal rows with scorer/assists for optional goal inference
      - observed jerseys by team color (Blue/White) for team inference
    """
    if df.empty or df.shape[1] < 6:
        return [], [], {}

    # Identify period header rows in column 0 (e.g., '1st Period', '2nd Period', '3rd period').
    col0 = df.iloc[:, 0]
    period_rows: List[Tuple[int, int]] = []
    for r, v in col0.items():
        p = parse_period_label(v)
        if p is not None:
            period_rows.append((int(r), int(p)))

    if not period_rows:
        return [], [], {}

    # Append sentinel end row.
    period_rows_sorted = sorted(period_rows, key=lambda x: x[0])
    period_rows_sorted.append((int(df.shape[0]), -1))

    events: List[LongEvent] = []
    goal_rows: List[Dict[str, Any]] = []
    jerseys_by_team: Dict[str, set[int]] = {"Blue": set(), "White": set()}

    def _col_for(header_row: int, *candidates: str) -> Optional[int]:
        # Map normalized header names to column indices.
        label_to_col: Dict[str, int] = {}
        for c in range(df.shape[1]):
            v = df.iat[header_row, c]
            if pd.isna(v):
                continue
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            label_to_col[_normalize_header_label(s)] = c
        for cand in candidates:
            key = _normalize_header_label(cand)
            if key in label_to_col:
                return label_to_col[key]
        # Fallback: substring match
        for cand in candidates:
            key = _normalize_header_label(cand)
            for hdr, c in label_to_col.items():
                if key and key in hdr:
                    return c
        return None

    for idx in range(len(period_rows_sorted) - 1):
        header_r, period = period_rows_sorted[idx]
        end_r = period_rows_sorted[idx + 1][0]

        # The header row itself contains column names (Video Time, Scoreboard, Team, etc.).
        video_col = _col_for(header_r, "Video Time")
        sb_col = _col_for(header_r, "Scoreboard", "Game Time")
        team_col = _col_for(header_r, "Team")
        shots_col = _col_for(header_r, "Shots", "Shot")
        sog_col = _col_for(header_r, "Shots on Goal", "Shot on Goal", "SOG")
        assist_col = _col_for(header_r, "Assist", "Assists")

        for r in range(header_r + 1, end_r):
            team = _parse_long_team(df.iat[r, team_col] if team_col is not None else None)
            if not team:
                continue
            vsec = _parse_long_mmss_time_to_seconds(df.iat[r, video_col]) if video_col is not None else None
            gsec = _parse_long_mmss_time_to_seconds(df.iat[r, sb_col]) if sb_col is not None else None

            label = df.iat[r, 0]
            label_s = str(label).strip() if isinstance(label, str) else ""
            label_l = label_s.lower()

            shooter = (
                _extract_jerseys_from_cell(df.iat[r, shots_col]) if shots_col is not None else []
            )
            assists = (
                _extract_jerseys_from_cell(df.iat[r, assist_col]) if assist_col is not None else []
            )
            marker = df.iat[r, sog_col] if sog_col is not None else None
            marker_s = str(marker).strip() if isinstance(marker, str) else ""
            marker_l = marker_s.lower()

            # Long sheets encode turnovers in a few different row labels over time:
            #  - "Turnover" / "Turnover (forced)"  -> forced turnover (lost puck)
            #  - "Unforced TO" (and occasional typos like "Unforced OT") -> unforced turnover (giveaway)
            #  - legacy: "Giveaway"
            is_unforced_to = label_l.startswith("unforced")
            is_turnover = ("turnover" in label_l) and (not is_unforced_to)
            is_giveaway = ("giveaway" in label_l) and (not is_turnover) and (not is_unforced_to)
            is_expected_goal = "expected goal" in label_l
            is_controlled_entry = ("controlled" in label_l) and ("entr" in label_l)
            is_controlled_exit = ("controlled" in label_l) and ("exit" in label_l)
            is_rush = "rush" in label_l
            is_goal = (label_l == "goal") or (marker_l == "goal")
            is_sog = marker_l in {"sog", "goal"}

            if is_turnover or is_giveaway or is_unforced_to:
                # Turnover rows use:
                #  - "Shots" for the player who lost possession
                #  - then 1-2 jersey numbers describing the opponent roles:
                #      * forced turnover: optional "Caused by" and always "Takeaway"
                #      * unforced turnover: always just "Takeaway"
                giving = shooter
                detail_a = marker if sog_col is not None else None
                detail_b = df.iat[r, assist_col] if assist_col is not None else None
                detail_a_nums = _extract_jerseys_from_cell(detail_a)
                detail_b_nums = _extract_jerseys_from_cell(detail_b)

                other_team = "White" if team == "Blue" else "Blue"

                giving_event_type = "TurnoverForced" if is_turnover else "Giveaway"
                giveaway = giving[:1]

                caused_by: List[int] = []
                takeaway: List[int] = []

                if is_turnover:
                    # Forced turnover:
                    #  - If "Assist" has a jersey, treat it as the Takeaway player and
                    #    interpret "Shots on Goal" as the Caused By player (if present).
                    #  - Otherwise, use ordering across the detail cells:
                    #      1 jersey  -> Takeaway
                    #      2 jerseys -> Caused By, Takeaway
                    if detail_b_nums:
                        if len(detail_b_nums) > 1:
                            raise ValueError(
                                f"forced turnover row has multiple takeaway jerseys: row={r} value={detail_b!r}"
                            )
                        takeaway = detail_b_nums[:1]
                        if detail_a_nums:
                            if len(detail_a_nums) > 1:
                                raise ValueError(
                                    f"forced turnover row has multiple caused-by jerseys: row={r} value={detail_a!r}"
                                )
                            caused_by = detail_a_nums[:1]
                    else:
                        combined: List[int] = []
                        for n in detail_a_nums + detail_b_nums:
                            if n not in combined:
                                combined.append(n)
                        if len(combined) > 2:
                            raise ValueError(
                                f"forced turnover row has too many opponent jerseys: row={r} values={[detail_a, detail_b]!r}"
                            )
                        if len(combined) == 2:
                            caused_by = combined[:1]
                            takeaway = combined[1:2]
                        elif len(combined) == 1:
                            takeaway = combined[:1]
                            # If the only detail is explicitly a "Caused by" entry,
                            # count it as both CreatedTurnover + Takeaway (fallback for incomplete rows).
                            try:
                                a_txt = str(detail_a).lower() if detail_a is not None else ""
                            except Exception:
                                a_txt = ""
                            if "caused" in a_txt:
                                caused_by = combined[:1]
                                takeaway = combined[:1]
                else:
                    # Giveaway / Unforced TO: first jersey is the giveaway; second (opponent) is the takeaway.
                    if len(giving) > 1:
                        raise ValueError(
                            f"unforced turnover row has multiple giveaway jerseys: row={r} value={df.iat[r, shots_col] if shots_col is not None else None!r}"
                        )
                    combined: List[int] = []
                    for n in detail_a_nums + detail_b_nums:
                        if n not in combined:
                            combined.append(n)
                    if len(combined) > 1:
                        raise ValueError(
                            f"unforced turnover row has multiple opponent jerseys: row={r} values={[detail_a, detail_b]!r}"
                        )
                    takeaway = combined[:1]

                if giveaway:
                    events.append(
                        LongEvent(
                            event_type=giving_event_type,
                            team=team,
                            period=period,
                            video_s=vsec,
                            game_s=gsec,
                            jerseys=tuple(giveaway),
                        )
                    )
                if caused_by:
                    events.append(
                        LongEvent(
                            event_type="CreatedTurnover",
                            team=other_team,
                            period=period,
                            video_s=vsec,
                            game_s=gsec,
                            jerseys=tuple(caused_by),
                        )
                    )
                if takeaway:
                    events.append(
                        LongEvent(
                            event_type="Takeaway",
                            team=other_team,
                            period=period,
                            video_s=vsec,
                            game_s=gsec,
                            jerseys=tuple(takeaway),
                        )
                    )
                for j in giveaway:
                    jerseys_by_team.setdefault(team, set()).add(int(j))
                for j in caused_by + takeaway:
                    jerseys_by_team.setdefault(other_team, set()).add(int(j))
                continue

            if shooter:
                events.append(
                    LongEvent(
                        event_type="Shot",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=tuple(shooter),
                    )
                )
            if shooter and is_sog:
                events.append(
                    LongEvent(
                        event_type="SOG",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=tuple(shooter),
                    )
                )
            if shooter and is_goal:
                events.append(
                    LongEvent(
                        event_type="Goal",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=tuple(shooter),
                    )
                )
                scorer = shooter[0] if shooter else None
                goal_rows.append(
                    {
                        "team": team,
                        "period": period,
                        "game_s": gsec,
                        "scorer": scorer,
                        "assists": assists,
                    }
                )
            if assists:
                events.append(
                    LongEvent(
                        event_type="Assist",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=tuple(assists),
                    )
                )
            # Expected goals (xG): count both explicit "Expected Goal" rows and all goals.
            if is_expected_goal or is_goal:
                events.append(
                    LongEvent(
                        event_type="ExpectedGoal",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=tuple(shooter),
                    )
                )
            if is_controlled_entry:
                events.append(
                    LongEvent(
                        event_type="ControlledEntry",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                    )
                )
            if is_controlled_exit:
                events.append(
                    LongEvent(
                        event_type="ControlledExit",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                    )
                )
            if is_rush:
                events.append(
                    LongEvent(
                        event_type="Rush",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                    )
                )

            for j in shooter + assists:
                jerseys_by_team.setdefault(team, set()).add(int(j))

    # Remove empty defaults if nothing was seen.
    jerseys_by_team = {k: v for k, v in jerseys_by_team.items() if v}
    return events, goal_rows, jerseys_by_team


def _infer_focus_team_from_long_sheet(
    our_jerseys: set[str],
    jerseys_by_team: Dict[str, set[int]],
) -> Optional[str]:
    if not our_jerseys or not jerseys_by_team:
        return None
    blue = {str(x) for x in jerseys_by_team.get("Blue", set())}
    white = {str(x) for x in jerseys_by_team.get("White", set())}
    blue_ov = len(our_jerseys & blue)
    white_ov = len(our_jerseys & white)
    if blue_ov == 0 and white_ov == 0:
        return None
    if blue_ov == white_ov:
        return None
    return "Blue" if blue_ov > white_ov else "White"


def _event_log_context_from_long_events(
    long_events: List[LongEvent],
    *,
    jersey_to_players: Dict[str, List[str]],
    focus_team: Optional[str],
    jerseys_by_team: Dict[str, set[int]],
) -> EventLogContext:
    event_counts_by_player: Dict[str, Dict[str, int]] = {}
    event_counts_by_type_team: Dict[Tuple[str, str], int] = {}
    event_instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    event_player_rows: List[Dict[str, Any]] = []

    MAX_TEAM_PLAYERS = 20
    team_roster: Dict[str, List[int]] = {}
    team_excluded: Dict[str, List[int]] = {}

    def _register_and_flag(team: str, jerseys: Iterable[int]) -> None:
        seen_local: set[int] = set()
        ordered: List[int] = []
        for j in jerseys:
            jj = int(j)
            if jj in seen_local:
                continue
            seen_local.add(jj)
            ordered.append(jj)
        roster = team_roster.setdefault(team, [])
        excluded = team_excluded.setdefault(team, [])
        for j in ordered:
            if j in roster:
                continue
            if len(roster) < MAX_TEAM_PLAYERS:
                roster.append(j)
            else:
                if j not in excluded:
                    excluded.append(j)

    # Seed rosters from the observed jerseys for better inference/debugging.
    for team, nums in (jerseys_by_team or {}).items():
        if not team or not nums:
            continue
        _register_and_flag(team, sorted(nums))

    for ev in long_events:
        team = ev.team
        etype = ev.event_type
        key = (etype, team)
        event_counts_by_type_team[key] = event_counts_by_type_team.get(key, 0) + 1
        event_instances.setdefault(key, []).append(
            {"period": ev.period, "video_s": ev.video_s, "game_s": ev.game_s}
        )

        if not ev.jerseys:
            continue

        _register_and_flag(team, ev.jerseys)
        for jersey in ev.jerseys:
            jersey_norm = _normalize_jersey_number(jersey)
            player_keys: List[str] = []
            if focus_team is not None and team == focus_team and jersey_norm:
                player_keys = jersey_to_players.get(jersey_norm, [])
            if not player_keys:
                player_keys = [f"{team}_{int(jersey)}"]
            for pk in player_keys:
                d = event_counts_by_player.setdefault(pk, {})
                d[etype] = d.get(etype, 0) + 1
                event_player_rows.append(
                    {
                        "event_type": etype,
                        "team": team,
                        "player": pk,
                        "jersey": int(jersey),
                        "period": ev.period,
                        "video_s": ev.video_s,
                        "game_s": ev.game_s,
                    }
                )

    return EventLogContext(
        event_counts_by_player=event_counts_by_player,
        event_counts_by_type_team=event_counts_by_type_team,
        event_instances=event_instances,
        event_player_rows=event_player_rows,
        team_roster=team_roster,
        team_excluded=team_excluded,
    )


def _merge_event_log_contexts(
    a: Optional[EventLogContext],
    b: Optional[EventLogContext],
) -> Optional[EventLogContext]:
    if a is None:
        return b
    if b is None:
        return a

    merged_counts_by_player: Dict[str, Dict[str, int]] = {}
    for src in (a.event_counts_by_player or {}, b.event_counts_by_player or {}):
        for pk, kinds in src.items():
            dest = merged_counts_by_player.setdefault(pk, {})
            for kind, cnt in (kinds or {}).items():
                try:
                    inc = int(cnt)
                except Exception:
                    inc = 0
                dest[kind] = dest.get(kind, 0) + inc

    merged_counts_by_type_team: Dict[Tuple[str, str], int] = {}
    for src in (a.event_counts_by_type_team or {}, b.event_counts_by_type_team or {}):
        for k, cnt in src.items():
            try:
                inc = int(cnt)
            except Exception:
                inc = 0
            merged_counts_by_type_team[k] = merged_counts_by_type_team.get(k, 0) + inc

    merged_instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for src in (a.event_instances or {}, b.event_instances or {}):
        for k, rows in src.items():
            merged_instances.setdefault(k, []).extend(list(rows or []))

    merged_player_rows: List[Dict[str, Any]] = []
    merged_player_rows.extend(list(a.event_player_rows or []))
    merged_player_rows.extend(list(b.event_player_rows or []))

    def _merge_rosters(
        r1: Dict[str, List[int]], r2: Dict[str, List[int]]
    ) -> Dict[str, List[int]]:
        out: Dict[str, List[int]] = {}
        for src in (r1 or {}, r2 or {}):
            for team, nums in src.items():
                if not team:
                    continue
                cur = out.setdefault(team, [])
                for n in nums or []:
                    try:
                        nn = int(n)
                    except Exception:
                        continue
                    if nn not in cur:
                        cur.append(nn)
        return out

    merged_roster = _merge_rosters(a.team_roster or {}, b.team_roster or {})
    merged_excluded = _merge_rosters(a.team_excluded or {}, b.team_excluded or {})

    return EventLogContext(
        event_counts_by_player=merged_counts_by_player,
        event_counts_by_type_team=merged_counts_by_type_team,
        event_instances=merged_instances,
        event_player_rows=merged_player_rows,
        team_roster=merged_roster,
        team_excluded=merged_excluded,
    )


def _detect_event_log_headers(
    df: pd.DataFrame,
) -> Tuple[Optional[Tuple[int, int]], Optional[Tuple[int, int]]]:
    def _find_cell(value: str) -> Optional[Tuple[int, int]]:
        needle = value.strip().lower()
        for rr in range(df.shape[0]):
            for cc in range(df.shape[1]):
                try:
                    v = df.iat[rr, cc]
                except Exception:
                    continue
                if pd.isna(v):
                    continue
                if str(v).strip().lower() == needle:
                    return (rr, cc)
        return None

    blue_hdr = _find_cell("Shifts (Blue)") or _find_cell("Shifts (blue)")
    white_hdr = _find_cell("Shifts (white)")
    return blue_hdr, white_hdr


def _parse_event_log_layout(df: pd.DataFrame) -> Tuple[
    bool,
    Dict[str, List[Tuple[str, str]]],
    Dict[str, List[Tuple[int, str, str]]],
    Dict[int, List[Tuple[int, int, int, int]]],
    Optional[EventLogContext],
]:
    blue_hdr, white_hdr = _detect_event_log_headers(df)
    if not (blue_hdr or white_hdr):
        return False, {}, {}, {}, None

    # Accumulators
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]] = {}
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]] = {}
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]] = {}

    # Roster cap handling
    MAX_TEAM_PLAYERS = 20
    team_roster: Dict[str, List[int]] = {}
    team_excluded: Dict[str, List[int]] = {}

    def _register_and_flag(team: str, jerseys: List[int]) -> List[int]:
        if not team:
            return []
        seen_local = set()
        ordered: List[int] = []
        for j in jerseys:
            if j in seen_local:
                continue
            seen_local.add(j)
            ordered.append(j)
        roster = team_roster.setdefault(team, [])
        excluded = team_excluded.setdefault(team, [])
        for j in ordered:
            if j in roster:
                continue
            if len(roster) < MAX_TEAM_PLAYERS:
                roster.append(j)
            else:
                if j not in excluded:
                    excluded.append(j)
        return ordered

    def _parse_event_time(cell: Any) -> Optional[int]:
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            return None
        s = str(cell).strip()
        parts = s.split(":")
        try:
            if len(parts) == 3:
                h, m, _s = parts
                return int(h) * 60 + int(m)
            if len(parts) == 2:
                m, sec = parts
                return int(m) * 60 + int(sec)
            if len(parts) == 1:
                return int(float(parts[0]))
        except Exception:
            return None
        return None

    def _period_num_from_label(lbl: Any) -> Optional[int]:
        return parse_period_token(lbl)

    def _parse_event_block(header_rc: Tuple[int, int], team_prefix: str) -> None:
        base_r, base_c = header_rc
        # Locate columns for video/game time near header
        video_col = None
        game_col = None
        period_col = None
        for r in range(base_r + 1, min(df.shape[0], base_r + 4)):
            for c in range(max(0, base_c - 4), min(df.shape[1], base_c + 6)):
                val = df.iat[r, c]
                if pd.notna(val) and isinstance(val, str):
                    s = val.strip().lower()
                    if s == "video time":
                        video_col = c
                    elif s == "game time":
                        game_col = c
                    elif "period" in s:
                        period_col = c
            if video_col is not None and game_col is not None:
                break
        if video_col is None:
            video_col = max(0, base_c - 2)
        if game_col is None:
            game_col = max(0, base_c - 1)
        if period_col is None:
            period_col = base_c + 2

        players_start = base_c
        players_width = 12

        # Build events
        current_period_label: Optional[str] = None
        events: List[Dict[str, Any]] = []
        for r in range(base_r + 1, df.shape[0]):
            vcell = df.iat[r, video_col] if video_col < df.shape[1] else None
            if isinstance(vcell, str) and vcell.strip().lower() == "video time":
                plbl = df.iat[r, period_col] if period_col < df.shape[1] else None
                current_period_label = str(plbl).strip() if pd.notna(plbl) else current_period_label
                continue
            gcell = df.iat[r, game_col] if game_col < df.shape[1] else None
            vsec = _parse_event_time(vcell)
            gsec = _parse_event_time(gcell)
            players: List[int] = []
            for k in range(players_width):
                c = players_start + k
                if c >= df.shape[1]:
                    break
                val = df.iat[r, c]
                if pd.isna(val):
                    continue
                if isinstance(val, (int, float)):
                    n = int(val)
                    if 1 <= n <= 98:
                        players.append(n)
                    continue
                if hasattr(val, "hour") and hasattr(val, "minute"):
                    continue
                s = str(val).strip()
                if not s:
                    continue
                if s.upper() in {"PP", "SH"}:
                    continue
                if re.match(r"^\d{1,2}:\d{2}(?::\d{2})?$", s):
                    continue
                for m in re.finditer(r"#?(\d{1,2})(?!\d)", s):
                    try:
                        n = int(m.group(1))
                    except Exception:
                        continue
                    if 1 <= n <= 98:
                        players.append(n)
            if players:
                players = sorted(set(players))
            if vsec is None and gsec is None and not players:
                continue
            players = _register_and_flag(team_prefix, players)
            events.append(
                {
                    "period": _period_num_from_label(current_period_label),
                    "v": vsec,
                    "g": gsec,
                    "players": players,
                }
            )

        # Walk events to generate per-player shifts
        open_shift: Dict[int, Dict[str, Any]] = {}
        last_period: Optional[int] = None
        for ev in events:
            cur_p = ev.get("period")
            # Close open shifts at period change (scoreboard end -> 0:00)
            if last_period is not None and cur_p is not None and cur_p != last_period:
                for pid, sh in list(open_shift.items()):
                    sv, sg = sh.get("sv"), sh.get("sg")
                    evv, egg = ev.get("v"), 0
                    key = f"{team_prefix}_{pid}"
                    if sv is not None and evv is not None:
                        video_pairs_by_player.setdefault(key, []).append(
                            (seconds_to_hhmmss(int(sv)), seconds_to_hhmmss(int(evv)))
                        )
                    if sg is not None:
                        sb_pairs_by_player.setdefault(key, []).append(
                            (
                                last_period,
                                seconds_to_mmss_or_hhmmss(int(sg)),
                                seconds_to_mmss_or_hhmmss(int(egg)),
                            )
                        )
                        if sv is not None and evv is not None:
                            conv_segments_by_period.setdefault(last_period, []).append(
                                (int(sg), int(egg), int(sv), int(evv))
                            )
                    del open_shift[pid]
            if cur_p is not None:
                last_period = cur_p

            on_ice = set(ev.get("players") or [])
            # Close shifts for players no longer on ice
            for pid, sh in list(open_shift.items()):
                if pid not in on_ice:
                    sv, sg = sh.get("sv"), sh.get("sg")
                    evv, egg = ev.get("v"), ev.get("g")
                    key = f"{team_prefix}_{pid}"
                    if sv is not None and evv is not None:
                        video_pairs_by_player.setdefault(key, []).append(
                            (seconds_to_hhmmss(int(sv)), seconds_to_hhmmss(int(evv)))
                        )
                    if sg is not None and cur_p is not None:
                        end_g = egg if egg is not None else 0
                        sb_pairs_by_player.setdefault(key, []).append(
                            (
                                cur_p,
                                seconds_to_mmss_or_hhmmss(int(sg)),
                                seconds_to_mmss_or_hhmmss(int(end_g)),
                            )
                        )
                        if sv is not None and evv is not None:
                            conv_segments_by_period.setdefault(cur_p, []).append(
                                (int(sg), int(end_g), int(sv), int(evv))
                            )
                    del open_shift[pid]

            # Open shifts for players now on
            for pid in on_ice:
                if pid not in open_shift:
                    open_shift[pid] = {
                        "sv": ev.get("v"),
                        "sg": ev.get("g"),
                        "period": ev.get("period"),
                    }

        # Close any remaining open shifts at last event, scoreboard -> 0:00
        if events and open_shift:
            last_ev = events[-1]
            for pid, sh in list(open_shift.items()):
                key = f"{team_prefix}_{pid}"
                sv, sg = sh.get("sv"), sh.get("sg")
                evv, egg = last_ev.get("v"), 0
                per = sh.get("period") or last_ev.get("period")
                if sv is not None and evv is not None:
                    video_pairs_by_player.setdefault(key, []).append(
                        (seconds_to_hhmmss(int(sv)), seconds_to_hhmmss(int(evv)))
                    )
                if sg is not None and per is not None:
                    sb_pairs_by_player.setdefault(key, []).append(
                        (
                            per,
                            seconds_to_mmss_or_hhmmss(int(sg)),
                            seconds_to_mmss_or_hhmmss(int(egg)),
                        )
                    )
                    if sv is not None and evv is not None:
                        conv_segments_by_period.setdefault(int(per), []).append(
                            (int(sg), int(egg), int(sv), int(evv))
                        )

    if blue_hdr:
        _parse_event_block(blue_hdr, "Blue")
    if white_hdr:
        _parse_event_block(white_hdr, "White")

    # ---- Parse left-side event columns ----
    event_counts_by_player: Dict[str, Dict[str, int]] = {}
    event_counts_by_type_team: Dict[Tuple[str, str], int] = {}
    event_instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    event_player_rows: List[Dict[str, Any]] = []

    # Find left header row and relevant columns up to the Blue/White shifts header column
    limit_col = blue_hdr[1] if blue_hdr else df.shape[1]
    left_header_row: Optional[int] = None
    for r in range(min(8, df.shape[0])):
        for c in range(min(limit_col, df.shape[1])):
            v = df.iat[r, c]
            if pd.notna(v) and isinstance(v, str) and v.strip().lower() == "video time":
                left_header_row = r
                break
        if left_header_row is not None:
            break
    if left_header_row is not None:
        # Map labels -> columns from the header row
        label_to_col: Dict[str, int] = {}
        for c in range(min(limit_col, df.shape[1])):
            v = df.iat[left_header_row, c]
            if pd.notna(v) and isinstance(v, str) and v.strip():
                label_to_col[v.strip().lower()] = c

        vt_col = label_to_col.get("video time")
        gt_col = label_to_col.get("scoreboard")
        shots_col = label_to_col.get("shots")
        goals_col = label_to_col.get("goals")
        assists_col = label_to_col.get("assist")
        # tolerate wording variations and capitalization
        entries_col = None
        exits_col = None
        for k, c in label_to_col.items():
            kl = k.lower()
            if "controlled" in kl and "blue" in kl and "entr" in kl:
                entries_col = c
            if "controlled" in kl and "exit" in kl:
                exits_col = c

        # Guess team column
        team_col: Optional[int] = None
        best_count = 0
        for c in range(min(limit_col, df.shape[1])):
            cnt = 0
            for r in range(left_header_row + 1, min(df.shape[0], left_header_row + 80)):
                v = df.iat[r, c]
                if isinstance(v, str) and v.strip() in ("Blue", "White"):
                    cnt += 1
            if cnt > best_count:
                best_count = cnt
                team_col = c

        def _parse_team_from_text(s: Optional[str]) -> Optional[str]:
            if not s:
                return None
            t = s.lower()
            if "blue" in t:
                return "Blue"
            if "white" in t:
                return "White"
            return None

        def _extract_nums(s: Optional[str]) -> List[int]:
            if not s:
                return []
            s = s.strip()
            if re.match(r"^\d{1,2}:\d{2}(?::\d{2})?$", s):
                return []
            nums: List[int] = []
            for m in re.finditer(r"#?(\d{1,2})(?!\d)", s):
                try:
                    n = int(m.group(1))
                except Exception:
                    continue
                if 1 <= n <= 98:
                    nums.append(n)
            seen = set()
            out: List[int] = []
            for n in nums:
                if n in seen:
                    continue
                seen.add(n)
                out.append(n)
            return out

        def _record_event(
            kind: str,
            team: Optional[str],
            jersey_list: List[int],
            period_label: Optional[str],
            vsec: Optional[int],
            gsec: Optional[int],
        ) -> None:
            if not team:
                return
            event_counts_by_type_team[(kind, team)] = (
                event_counts_by_type_team.get((kind, team), 0) + 1
            )
            filtered = _register_and_flag(team, jersey_list)
            period_num = None
            if period_label is not None:
                m = re.search(r"(\d+)", str(period_label))
                if m:
                    period_num = int(m.group(1))
            for j in filtered:
                pk = f"{team}_{int(j)}"
                d = event_counts_by_player.setdefault(pk, {})
                d[kind] = d.get(kind, 0) + 1
                event_player_rows.append(
                    {
                        "event_type": kind,
                        "team": team,
                        "player": pk,
                        "jersey": int(j),
                        "period": period_num,
                        "video_s": vsec,
                        "game_s": gsec,
                    }
                )
            event_instances.setdefault((kind, team), []).append(
                {"period": period_num, "video_s": vsec, "game_s": gsec}
            )

        # Walk data rows to collect events
        current_period: Optional[str] = None
        for r in range(left_header_row + 1, df.shape[0]):
            row_vals = [df.iat[r, c] for c in range(min(limit_col, df.shape[1]))]
            for v in row_vals:
                if isinstance(v, str) and "period" in v.lower():
                    current_period = v.strip()
                    break

            team_val = None
            if team_col is not None:
                tv = df.iat[r, team_col]
                if isinstance(tv, str) and tv.strip() in ("Blue", "White"):
                    team_val = tv.strip()
            row_vsec = _parse_event_time(df.iat[r, vt_col]) if vt_col is not None else None
            row_gsec = _parse_event_time(df.iat[r, gt_col]) if gt_col is not None else None

            if shots_col is not None:
                sv = df.iat[r, shots_col]
                if isinstance(sv, str) and sv.strip():
                    t = team_val or _parse_team_from_text(sv)
                    jerseys = _extract_nums(sv)
                    _record_event("Shot", t, jerseys, current_period, row_vsec, row_gsec)
            if goals_col is not None:
                gv = df.iat[r, goals_col]
                if isinstance(gv, str) and gv.strip():
                    t = team_val or _parse_team_from_text(gv)
                    jerseys = _extract_nums(gv)
                    _record_event("Goal", t, jerseys, current_period, row_vsec, row_gsec)
                    # Goals also count as expected goals (xG).
                    _record_event("ExpectedGoal", t, jerseys, current_period, row_vsec, row_gsec)
            if assists_col is not None:
                av = df.iat[r, assists_col]
                if isinstance(av, str) and av.strip():
                    t = team_val or _parse_team_from_text(av)
                    jerseys = _extract_nums(av)
                    _record_event("Assist", t, jerseys, current_period, row_vsec, row_gsec)
            if entries_col is not None:
                ev = df.iat[r, entries_col]
                if isinstance(ev, str) and ev.strip():
                    t = team_val or _parse_team_from_text(ev)
                    jerseys = _extract_nums(ev)
                    _record_event("ControlledEntry", t, jerseys, current_period, row_vsec, row_gsec)
            if exits_col is not None:
                xv = df.iat[r, exits_col]
                if isinstance(xv, str) and xv.strip():
                    t = team_val or _parse_team_from_text(xv)
                    jerseys = _extract_nums(xv)
                    _record_event("ControlledExit", t, jerseys, current_period, row_vsec, row_gsec)
            label = df.iat[r, 0]
            if isinstance(label, str) and label.strip().lower() == "expected goal":
                text_cell = None
                if shots_col is not None:
                    text_cell = df.iat[r, shots_col]
                if not (isinstance(text_cell, str) and text_cell.strip()) and goals_col is not None:
                    text_cell = df.iat[r, goals_col]
                t = None
                jerseys: List[int] = []
                if isinstance(text_cell, str) and text_cell.strip():
                    t = team_val or _parse_team_from_text(text_cell)
                    jerseys = _extract_nums(text_cell)
                _record_event("ExpectedGoal", t, jerseys, current_period, row_vsec, row_gsec)

    event_log_context = EventLogContext(
        event_counts_by_player=event_counts_by_player,
        event_counts_by_type_team=event_counts_by_type_team,
        event_instances=event_instances,
        event_player_rows=event_player_rows,
        team_roster=team_roster,
        team_excluded=team_excluded,
    )

    return (
        True,
        video_pairs_by_player,
        sb_pairs_by_player,
        conv_segments_by_period,
        event_log_context,
    )


def _parse_per_player_layout(df: pd.DataFrame, keep_goalies: bool, skip_validation: bool) -> Tuple[
    Dict[str, List[Tuple[str, str]]],
    Dict[str, List[Tuple[int, str, str]]],
    Dict[int, List[Tuple[int, int, int, int]]],
    int,
]:
    blocks = find_period_blocks(df)
    if not blocks:
        raise ValueError("No 'Period N' sections found in column A.")

    video_pairs_by_player: Dict[str, List[Tuple[str, str]]] = {}
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]] = {}
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]] = {}

    MAX_SHIFT_SECONDS = 30 * 60  # 30 minutes

    def _report_validation(
        kind: str, period: int, player_key: str, a: str, b: str, reason: str
    ) -> None:
        print(
            f"[validation] {kind} | Player={player_key} | Period={period} | start='{a}' end='{b}' -> {reason}",
            file=sys.stderr,
        )

    validation_errors = 0

    # Parse each period block
    for period_num, blk_start, blk_end in blocks:
        header_row_idx = find_header_row(df, blk_start, blk_end)
        if header_row_idx is None:
            raise ValueError(f"Could not locate header row for Period {period_num}")
        data_start = header_row_idx + 1

        header = df.iloc[header_row_idx]
        groups = forward_fill_header_labels(header)

        start_sb_cols = _resolve_header_columns(
            groups,
            LABEL_START_SB,
            "Shift start (Scoreboard time)",
            "Shift Start (Scoreboard time)",
        )
        end_sb_cols = _resolve_header_columns(groups, LABEL_END_SB, "Shift end (Scoreboard time)")
        start_v_cols = _resolve_header_columns(groups, LABEL_START_V)
        end_v_cols = _resolve_header_columns(groups, LABEL_END_V)
        for lab, cols in [
            (LABEL_START_SB, start_sb_cols),
            (LABEL_END_SB, end_sb_cols),
            (LABEL_START_V, start_v_cols),
            (LABEL_END_V, end_v_cols),
        ]:
            if not cols:
                raise ValueError(f"Missing header columns for '{lab}' in Period {period_num}")

        # Iterate rows (players) in the block
        for r in range(data_start, blk_end):
            jersey = str(df.iloc[r, 0]).strip()
            name = str(df.iloc[r, 1]).strip()

            if is_period_label(df.iloc[r, 0]) or (not jersey and not name):
                break
            jersey_lower = jersey.lower()
            # Skip header-like rows that may appear again (e.g., overtime sections)
            if jersey_lower in {"jersey no", "jersey number"}:
                continue
            if not jersey or jersey_lower == "nan":
                continue
            # Skip goalies like "(G) 37"
            if not keep_goalies and "(" in jersey and ")" in jersey:
                continue

            player_key = f"{sanitize_name(jersey)}_{sanitize_name(name)}"

            video_pairs = extract_pairs_from_row(df.iloc[r], start_v_cols, end_v_cols)
            sb_pairs = extract_pairs_from_row(df.iloc[r], start_sb_cols, end_sb_cols)
            if sb_pairs:
                sb_pairs = [(a, _normalize_sb_end_time(b)) for a, b in sb_pairs]

            if not skip_validation:
                for va, vb in video_pairs:
                    try:
                        vsa = parse_flex_time_to_seconds(va)
                        vsb = parse_flex_time_to_seconds(vb)
                    except Exception as e:
                        _report_validation(
                            "VIDEO", period_num, player_key, va, vb, f"unparseable time: {e}"
                        )
                        validation_errors += 1
                        continue
                    if vsa >= vsb:
                        _report_validation(
                            "VIDEO",
                            period_num,
                            player_key,
                            va,
                            vb,
                            "start must be before end (strictly increasing)",
                        )
                        validation_errors += 1
                    dur = vsb - vsa if vsb >= vsa else 0
                    if dur > MAX_SHIFT_SECONDS:
                        _report_validation(
                            "VIDEO",
                            period_num,
                            player_key,
                            va,
                            vb,
                            f"duration {seconds_to_mmss_or_hhmmss(dur)} exceeds limit 30:00",
                        )
                        validation_errors += 1

                for sa, sb in sb_pairs:
                    try:
                        ssa = parse_flex_time_to_seconds(sa)
                        ssb = parse_flex_time_to_seconds(sb)
                    except Exception as e:
                        _report_validation(
                            "SCOREBOARD", period_num, player_key, sa, sb, f"unparseable time: {e}"
                        )
                        validation_errors += 1
                        continue
                    if ssa == ssb:
                        _report_validation(
                            "SCOREBOARD",
                            period_num,
                            player_key,
                            sa,
                            sb,
                            "start equals end (zero-length shift)",
                        )
                        validation_errors += 1
                    dur = abs(ssb - ssa)
                    if dur > MAX_SHIFT_SECONDS:
                        _report_validation(
                            "SCOREBOARD",
                            period_num,
                            player_key,
                            sa,
                            sb,
                            f"duration {seconds_to_mmss_or_hhmmss(dur)} exceeds limit 30:00",
                        )
                        validation_errors += 1

            if video_pairs:
                video_pairs_by_player.setdefault(player_key, []).extend(video_pairs)
            if sb_pairs:
                sb_pairs_by_player.setdefault(player_key, []).extend(
                    (period_num, a, b) for a, b in sb_pairs
                )

            nseg = min(len(video_pairs), len(sb_pairs))
            for idx in range(nseg):
                sva, svb = video_pairs[idx]
                sba, sbb = sb_pairs[idx]
                try:
                    v1 = parse_flex_time_to_seconds(sva)
                    v2 = parse_flex_time_to_seconds(svb)
                    s1 = parse_flex_time_to_seconds(sba)
                    s2 = parse_flex_time_to_seconds(sbb)
                except Exception:
                    continue
                conv_segments_by_period.setdefault(period_num, []).append((s1, s2, v1, v2))

    return video_pairs_by_player, sb_pairs_by_player, conv_segments_by_period, validation_errors


def _write_video_times_and_scripts(
    outdir: Path,
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]],
    create_scripts: bool,
) -> None:
    if not create_scripts:
        return
    for player_key, v_pairs in video_pairs_by_player.items():
        norm_pairs = []
        for a, b in v_pairs:
            try:
                sa = parse_flex_time_to_seconds(a)
                sb = parse_flex_time_to_seconds(b)
            except Exception:
                continue
            norm_pairs.append((seconds_to_hhmmss(sa), seconds_to_hhmmss(sb)))
        p = outdir / f"{player_key}_video_times.txt"
        p.write_text(
            "\n".join(f"{a} {b}" for a, b in norm_pairs) + ("\n" if norm_pairs else ""),
            encoding="utf-8",
        )

        script_path = outdir / f"clip_{player_key}.sh"
        player_label = player_key.replace("_", " ")
        script_body = """#!/usr/bin/env bash
	set -euo pipefail
	if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
THIS_DIR=\"$(cd \"$(dirname \"${{BASH_SOURCE[0]}}\")\" && pwd)\"
TS_FILE=\"$THIS_DIR/{player_key}_video_times.txt\"
# Parse optional flags
QUICK=0
HQ=0
shift 2 || true
for ARG in \"$@\"; do
  if [ \"$ARG\" = \"--quick\" ] || [ \"$ARG\" = \"-q\" ]; then
    QUICK=1
  elif [ \"$ARG\" = \"--hq\" ]; then
    HQ=1
  fi
done

EXTRA_FLAGS=()
if [ \"$QUICK\" -gt 0 ]; then
  EXTRA_FLAGS+=(\"--quick\" \"1\")
fi
if [ \"$HQ\" -gt 0 ]; then
  export VIDEO_CLIPPER_HQ=1
fi

python -m hmlib.cli.video_clipper -j {nr_jobs} --input \"$INPUT\" --timestamps \"$TS_FILE\" --temp-dir \"$THIS_DIR/temp_clips/{player_key}\" \"{player_label} vs $OPP\" \"${{EXTRA_FLAGS[@]}}\"
""".format(
            nr_jobs=4, player_key=player_key, player_label=player_label
        )
        script_path.write_text(script_body, encoding="utf-8")
        try:
            import os

            os.chmod(script_path, 0o755)
        except Exception:
            pass


def _write_scoreboard_times(
    outdir: Path,
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    *,
    create_scripts: bool,
) -> None:
    if not create_scripts:
        return
    for player_key, sb_list in sb_pairs_by_player.items():
        p = outdir / f"{player_key}_scoreboard_times.txt"
        lines = [f"{period} {a} {b}" for (period, a, b) in sb_list]
        p.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def _write_global_summary_csv(
    stats_dir: Path, sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]]
) -> None:
    summary_rows = []
    for player_key, sb_list in sb_pairs_by_player.items():
        all_pairs = [(a, b) for (_, a, b) in sb_list]
        shift_summary = summarize_shift_lengths_sec(all_pairs)
        row = {
            "player": player_key,
            "num_shifts": int(shift_summary["num_shifts"]),
            "toi_total_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_total"])
                if ":" in shift_summary["toi_total"]
                else 0
            ),
            "toi_avg_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_avg"])
                if ":" in shift_summary["toi_avg"]
                else 0
            ),
            "toi_median_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_median"])
                if ":" in shift_summary["toi_median"]
                else 0
            ),
            "toi_longest_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_longest"])
                if ":" in shift_summary["toi_longest"]
                else 0
            ),
            "toi_shortest_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_shortest"])
                if ":" in shift_summary["toi_shortest"]
                else 0
            ),
        }
        summary_rows.append(row)
    if summary_rows:
        df = pd.DataFrame(summary_rows).sort_values(by="player")
        df.to_csv(stats_dir / "summary_stats.csv", index=False)
        _write_styled_xlsx_table(
            stats_dir / "summary_stats.xlsx",
            df,
            sheet_name="summary_stats",
            title="Shift Summary",
        )


def _write_pair_on_ice_csv(stats_dir: Path, rows: List[Dict[str, Any]], *, include_toi: bool) -> None:
    if not rows:
        return
    def _blank_if_zero(x: Any) -> Any:
        try:
            v = int(x or 0)
        except Exception:
            v = 0
        return "" if v == 0 else v

    out_rows: List[Dict[str, Any]] = []
    for r in rows:
        row_out: Dict[str, Any] = {
            "Player": r.get("player", ""),
            "Teammate": r.get("teammate", ""),
            "Games with Data": int(r.get("shift_games", 0) or 0),
            "Overlap %": float(r.get("overlap_pct", 0.0) or 0.0),
            "GF Together": int(r.get("gf_together", 0) or 0),
            "GA Together": int(r.get("ga_together", 0) or 0),
            "Player Goals (On Ice Together)": _blank_if_zero(r.get("player_goals_on_ice_together", 0)),
            "Player Assists (On Ice Together)": _blank_if_zero(r.get("player_assists_on_ice_together", 0)),
            "Goals Collaborated": _blank_if_zero(r.get("goals_collab_with_teammate", 0)),
            "Assists Collaborated": _blank_if_zero(r.get("assists_collab_with_teammate", 0)),
            "+/- Together": int(r.get("plus_minus_together", 0) or 0),
            "Player Total +/-": int(r.get("player_total_plus_minus", 0) or 0),
            "Teammate Total +/-": int(r.get("teammate_total_plus_minus", 0) or 0),
        }
        # Only publish absolute TOI/overlap time when explicitly enabled.
        if include_toi:
            row_out["Player TOI"] = _seconds_to_compact_hms(int(r.get("player_toi_seconds", 0) or 0))
            row_out["Overlap"] = _seconds_to_compact_hms(int(r.get("overlap_seconds", 0) or 0))
        out_rows.append(row_out)
    df = pd.DataFrame(out_rows)
    try:
        df.sort_values(by=["Player", "Overlap %"], ascending=[True, False], inplace=True)
    except Exception:
        pass
    df.to_csv(stats_dir / "pair_on_ice.csv", index=False)
    _write_styled_xlsx_table(
        stats_dir / "pair_on_ice.xlsx",
        df,
        sheet_name="pair_on_ice",
        title="Pair On-Ice",
        number_formats={"Overlap %": "0.0"},
        text_columns=["Player", "Teammate"],
        align_right_columns=["Player TOI", "Overlap"],
    )


def _write_all_events_summary(
    stats_dir: Path,
    *,
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    goals: List[GoalEvent],
    goals_by_period: Dict[int, List[GoalEvent]],
    event_log_context: Optional["EventLogContext"],
    focus_team: Optional[str],
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
) -> None:
    """
    Write a single per-game table of all known events, including:
      - long-sheet / event-log events (player-attributed and team-only)
      - goal/assist events from the goals list (when available)
      - on-ice player lists for each event (our team only, based on shift times)

    This is intended to be sufficient to recompute most per-game and per-player event-based stats.
    Note: without opponent shift times, opponent on-ice lists cannot be reconstructed here.
    """

    def _team_label(team: Any, *, event_type: Optional[str] = None) -> str:
        team_str = str(team) if team is not None else ""
        if focus_team in {"Blue", "White"} and team_str in {"Blue", "White"}:
            et = str(event_type or "")
            invert = et in {"TurnoverForced", "CreatedTurnover", "Giveaway", "Takeaway"}
            if invert:
                return "Against" if team_str == focus_team else "For"
            return "For" if team_str == focus_team else "Against"
        return team_str or ""

    # Precompute merged shift intervals and shift-start times for on-ice membership.
    intervals_by_player_period: Dict[str, Dict[int, List[Tuple[int, int]]]] = {}
    start_times_by_player_period: Dict[str, Dict[int, set[int]]] = {}
    for player, sb_list in (sb_pairs_by_player or {}).items():
        per_period: Dict[int, List[Tuple[int, int]]] = {}
        start_times: Dict[int, set[int]] = {}
        for period, a, b in sb_list or []:
            lo, hi = compute_interval_seconds(a, b)
            per_period.setdefault(int(period), []).append((lo, hi))
            try:
                start_times.setdefault(int(period), set()).add(parse_flex_time_to_seconds(a))
            except Exception:
                pass
        merged: Dict[int, List[Tuple[int, int]]] = {p: _merge_intervals(iv) for p, iv in per_period.items()}
        intervals_by_player_period[player] = merged
        start_times_by_player_period[player] = start_times

    def _on_ice_players(period: int, game_s: int) -> List[str]:
        out: List[str] = []
        for pk, per_map in intervals_by_player_period.items():
            for lo, hi in per_map.get(period, []):
                if interval_contains(game_s, lo, hi):
                    out.append(pk)
                    break
        return sorted(out)

    def _on_ice_players_pm(period: int, game_s: int) -> List[str]:
        # Plus/minus skips goals exactly at shift start for each player.
        base = _on_ice_players(period, game_s)
        out: List[str] = []
        for pk in base:
            if game_s in start_times_by_player_period.get(pk, {}).get(period, set()):
                continue
            out.append(pk)
        return out

    def _map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    # Group player-attributed long-sheet rows by event identity.
    player_rows_by_event: Dict[Tuple[str, str, int, Optional[int], Optional[int]], Dict[str, Any]] = {}
    if event_log_context is not None:
        for r in event_log_context.event_player_rows or []:
            try:
                et = str(r.get("event_type") or "")
                tm = str(r.get("team") or "")
                per = int(r.get("period") or 0)
                vs = r.get("video_s")
                gs = r.get("game_s")
                key = (et, tm, per, int(vs) if isinstance(vs, (int, float)) else None, int(gs) if isinstance(gs, (int, float)) else None)
                dest = player_rows_by_event.setdefault(key, {"players": set(), "jerseys": set()})
                pk = str(r.get("player") or "").strip()
                if pk:
                    dest["players"].add(pk)
                j = r.get("jersey")
                if j is not None:
                    try:
                        dest["jerseys"].add(int(j))
                    except Exception:
                        pass
            except Exception:
                continue

    rows: List[Dict[str, Any]] = []
    event_id = 0

    # Long-sheet / event-log events.
    if event_log_context is not None:
        for (etype, team), inst_list in sorted((event_log_context.event_instances or {}).items()):
            if goals and str(etype) == "Goal":
                # Prefer canonical Goal/Assist rows from the goals list when available.
                continue
            for inst in inst_list or []:
                period = int(inst.get("period") or 0)
                video_s = inst.get("video_s")
                game_s = inst.get("game_s")
                vs_i = int(video_s) if isinstance(video_s, (int, float)) else None
                gs_i = int(game_s) if isinstance(game_s, (int, float)) else None
                key = (str(etype), str(team), period, vs_i, gs_i)
                pj = player_rows_by_event.get(key, {"players": set(), "jerseys": set()})
                attrib_players = sorted([str(x) for x in pj.get("players", set()) if x])
                attrib_jerseys = sorted([int(x) for x in pj.get("jerseys", set()) if isinstance(x, int)])
                on_ice = _on_ice_players(period, gs_i) if (gs_i is not None and period > 0) else []
                on_ice_pm = _on_ice_players_pm(period, gs_i) if (gs_i is not None and period > 0) else []

                event_id += 1
                rows.append(
                    {
                        "Event ID": event_id,
                        "Source": "long",
                        "Event Type": _display_event_type(str(etype)),
                        "Event Type Raw": str(etype),
                        "Team Raw": str(team),
                        "Team Rel": _team_label(team, event_type=str(etype)),
                        "Period": period if period > 0 else "",
                        "Game Time": seconds_to_mmss_or_hhmmss(gs_i) if gs_i is not None else "",
                        "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                        "Game Seconds": gs_i if gs_i is not None else "",
                        "Video Seconds": vs_i if vs_i is not None else "",
                        "Attributed Players": ",".join(attrib_players),
                        "Attributed Jerseys": ",".join(str(j) for j in attrib_jerseys),
                        "On-Ice Players": ",".join(on_ice),
                        "On-Ice Players (PM)": ",".join(on_ice_pm),
                    }
                )

    # Goal/assist events from the goals list.
    for ev in sorted(goals or [], key=lambda e: (int(getattr(e, "period", 0) or 0), int(getattr(e, "t_sec", 0) or 0))):
        period = int(getattr(ev, "period", 0) or 0)
        gs_i = int(getattr(ev, "t_sec", 0) or 0)
        vs_i = _map_sb_to_video(period, gs_i)
        team_rel = "For" if getattr(ev, "kind", None) == "GF" else ("Against" if getattr(ev, "kind", None) == "GA" else "")
        on_ice = _on_ice_players(period, gs_i) if period > 0 else []
        on_ice_pm = _on_ice_players_pm(period, gs_i) if period > 0 else []

        # Goal scorer row (if known).
        scorer = getattr(ev, "scorer", None)
        if scorer:
            event_id += 1
            rows.append(
                {
                    "Event ID": event_id,
                    "Source": "goals",
                    "Event Type": "Goal",
                    "Event Type Raw": "Goal",
                    "Team Raw": "",
                    "Team Rel": team_rel,
                    "Period": period if period > 0 else "",
                    "Game Time": seconds_to_mmss_or_hhmmss(gs_i),
                    "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                    "Game Seconds": gs_i,
                    "Video Seconds": vs_i if vs_i is not None else "",
                    "Attributed Players": "",
                    "Attributed Jerseys": str(scorer),
                    "On-Ice Players": ",".join(on_ice),
                    "On-Ice Players (PM)": ",".join(on_ice_pm),
                }
            )
        # Assist rows (if any).
        for ast in getattr(ev, "assists", []) or []:
            if not ast:
                continue
            event_id += 1
            rows.append(
                {
                    "Event ID": event_id,
                    "Source": "goals",
                    "Event Type": "Assist",
                    "Event Type Raw": "Assist",
                    "Team Raw": "",
                    "Team Rel": team_rel,
                    "Period": period if period > 0 else "",
                    "Game Time": seconds_to_mmss_or_hhmmss(gs_i),
                    "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                    "Game Seconds": gs_i,
                    "Video Seconds": vs_i if vs_i is not None else "",
                    "Attributed Players": "",
                    "Attributed Jerseys": str(ast),
                    "On-Ice Players": ",".join(on_ice),
                    "On-Ice Players (PM)": ",".join(on_ice_pm),
                }
            )

    # Stable sort by time, then id.
    def _sort_key(r: Dict[str, Any]) -> Tuple[int, int, int]:
        try:
            per = int(r.get("Period") or 0)
        except Exception:
            per = 0
        try:
            gs = int(r.get("Game Seconds") or 0)
        except Exception:
            gs = 0
        try:
            eid = int(r.get("Event ID") or 0)
        except Exception:
            eid = 0
        return (per, gs, eid)

    rows.sort(key=_sort_key)

    cols = [
        "Event ID",
        "Source",
        "Event Type",
        "Event Type Raw",
        "Team Raw",
        "Team Rel",
        "Period",
        "Game Time",
        "Video Time",
        "Game Seconds",
        "Video Seconds",
        "Attributed Players",
        "Attributed Jerseys",
        "On-Ice Players",
        "On-Ice Players (PM)",
    ]
    df = pd.DataFrame([{c: r.get(c, "") for c in cols} for r in rows], columns=cols)
    df.to_csv(stats_dir / "all_events_summary.csv", index=False)
    _write_styled_xlsx_table(
        stats_dir / "all_events_summary.xlsx",
        df,
        sheet_name="all_events",
        title="All Events",
        text_columns=[
            "Event Type",
            "Event Type Raw",
            "Team Raw",
            "Team Rel",
            "Attributed Players",
            "Attributed Jerseys",
            "On-Ice Players",
            "On-Ice Players (PM)",
        ],
    )


def _compute_player_stats(
    player_key: str,
    sb_list: List[Tuple[int, str, str]],
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]],
    goals_by_period: Dict[int, List[GoalEvent]],
) -> Tuple[Dict[str, str], Dict[str, int], Dict[str, int], Dict[int, str]]:
    sb_by_period: Dict[int, List[Tuple[str, str]]] = {}
    for period, a, b in sb_list:
        sb_by_period.setdefault(period, []).append((a, b))

    all_pairs = [(a, b) for (_, a, b) in sb_list]
    shift_summary = summarize_shift_lengths_sec(all_pairs)
    per_period_toi_map = per_period_toi(sb_by_period)

    plus_minus = 0
    counted_gf: List[str] = []
    counted_ga: List[str] = []
    counted_gf_by_period: Dict[int, int] = {}
    counted_ga_by_period: Dict[int, int] = {}
    for period, pairs in sb_by_period.items():
        if period not in goals_by_period:
            continue
        for ev in goals_by_period[period]:
            matched = False
            for a, b in pairs:
                a_sec = parse_flex_time_to_seconds(a)
                b_sec = parse_flex_time_to_seconds(b)
                lo, hi = (a_sec, b_sec) if a_sec <= b_sec else (b_sec, a_sec)
                if not (lo <= ev.t_sec <= hi):
                    continue
                if ev.kind == "GA" and ev.t_sec == a_sec:
                    continue
                elif ev.kind == "GF" and ev.t_sec == a_sec:
                    continue
                matched = True
                break
            if matched:
                if ev.kind == "GF":
                    plus_minus += 1
                    counted_gf.append(f"P{period}:{ev.t_str}")
                    counted_gf_by_period[period] = counted_gf_by_period.get(period, 0) + 1
                else:
                    plus_minus -= 1
                    counted_ga.append(f"P{period}:{ev.t_str}")
                    counted_ga_by_period[period] = counted_ga_by_period.get(period, 0) + 1

    row_map: Dict[str, str] = {
        "player": player_key,
        "shifts": shift_summary["num_shifts"],
        "plus_minus": str(plus_minus),
        "sb_toi_total": shift_summary["toi_total"],
        "sb_avg": shift_summary["toi_avg"],
        "sb_median": shift_summary["toi_median"],
        "sb_longest": shift_summary["toi_longest"],
        "sb_shortest": shift_summary["toi_shortest"],
    }
    row_map["gf_counted"] = str(len(counted_gf))
    row_map["ga_counted"] = str(len(counted_ga))

    v_pairs = video_pairs_by_player.get(player_key, [])
    if v_pairs:
        v_sum = 0
        for a, b in v_pairs:
            lo, hi = compute_interval_seconds(a, b)
            v_sum += hi - lo
        row_map["video_toi_total"] = _format_duration(v_sum)
    else:
        row_map["video_toi_total"] = ""

    # per-period values
    per_counts: Dict[str, int] = {}
    per_counts_gf: Dict[str, int] = {}
    for period, toi in per_period_toi_map.items():
        row_map[f"P{period}_toi"] = toi
    for period, pairs in sb_by_period.items():
        per_counts[f"P{period}_shifts"] = len(pairs)
    for period, cnt in counted_gf_by_period.items():
        per_counts_gf[f"P{period}_GF"] = cnt
    for period, cnt in counted_ga_by_period.items():
        per_counts_gf[f"P{period}_GA"] = per_counts_gf.get(
            f"P{period}_GA", 0
        )  # placeholder to ensure keys
    # Return row_map and per-period counts; plus per_period_toi_map for columns
    return row_map, per_counts, {**{k: 0 for k in []}}, per_period_toi_map


def _build_stats_dataframe(
    stats_table_rows: List[Dict[str, str]],
    all_periods_seen: List[int],
    sort_for_cumulative: bool = False,
    *,
    include_shifts_in_stats: bool,
) -> Tuple[pd.DataFrame, List[str]]:
    def _has_any_value(key: str) -> bool:
        for r in stats_table_rows:
            if not r:
                continue
            v = r.get(key, "")
            if v is None:
                continue
            if str(v).strip() != "":
                return True
        return False

    periods = sorted(all_periods_seen)
    # Keep key summary metrics grouped early for easier reading in consolidated sheets.
    summary_cols = ["player", "gp", "goals", "assists", "points", "ppg"]
    summary_cols += [
        "plus_minus",
        "plus_minus_per_game",
        "gf_counted",
        "gf_per_game",
        "ga_counted",
        "ga_per_game",
    ]
    if include_shifts_in_stats:
        summary_cols += ["gf_per_shift", "ga_per_shift"]

    summary_cols += [
        "shots",
        "shots_per_game",
        "sog",
        "expected_goals",
        "sog_per_game",
        "expected_goals_per_game",
        "expected_goals_per_sog",
        "turnovers_forced",
        "turnovers_forced_per_game",
        "created_turnovers",
        "created_turnovers_per_game",
        "giveaways",
        "giveaways_per_game",
        "takeaways",
        "takeaways_per_game",
        "controlled_entry_for",
        "controlled_entry_for_per_game",
        "controlled_entry_against",
        "controlled_entry_against_per_game",
        "controlled_exit_for",
        "controlled_exit_for_per_game",
        "controlled_exit_against",
        "controlled_exit_against_per_game",
        "gt_goals",
        "gw_goals",
        "ot_goals",
        "ot_assists",
    ]
    if include_shifts_in_stats:
        summary_cols += [
            "shifts",
            "shifts_per_game",
            "goals_per_shift",
            "assists_per_shift",
            "points_per_shift",
            "shots_per_shift",
            "sog_per_shift",
            "expected_goals_per_shift",
            "turnovers_forced_per_shift",
            "created_turnovers_per_shift",
            "giveaways_per_shift",
            "takeaways_per_shift",
            "controlled_entry_for_per_shift",
            "controlled_entry_against_per_shift",
            "controlled_exit_for_per_shift",
            "controlled_exit_against_per_shift",
        ]

    sb_cols = (
        ["sb_toi_total", "sb_toi_per_game", "sb_avg", "sb_median", "sb_longest", "sb_shortest"]
        if include_shifts_in_stats
        else []
    )
    video_cols = ["video_toi_total"] if include_shifts_in_stats else []
    period_toi_cols = (
        [f"P{p}_toi" for p in periods if _has_any_value(f"P{p}_toi")] if include_shifts_in_stats else []
    )
    period_shift_cols = (
        [f"P{p}_shifts" for p in periods if _has_any_value(f"P{p}_shifts")]
        if include_shifts_in_stats
        else []
    )
    # Only include per-period goal columns when they have at least one entry.
    period_gf_cols = [f"P{p}_GF" for p in periods if _has_any_value(f"P{p}_GF")]
    period_ga_cols = [f"P{p}_GA" for p in periods if _has_any_value(f"P{p}_GA")]
    # Place video TOI as the last column in the table so that
    # scoreboard-based stats and per-period splits appear first.
    cols = (
        summary_cols
        + sb_cols
        + period_toi_cols
        + period_shift_cols
        + period_gf_cols
        + period_ga_cols
        + video_cols
    )

    rows_sorted: List[Dict[str, str]] = list(stats_table_rows)
    if sort_for_cumulative:
        # Stable multi-key sort:
        #  - final primary key: points (descending)
        #  - tie-breakers (in order of increasing precedence due to stable sort):
        #      player name, then assists, then goals.
        def _intval(row: Dict[str, str], key: str) -> int:
            try:
                return int(str(row.get(key, 0) or 0))
            except Exception:
                return 0

        rows_sorted.sort(key=lambda r: r.get("player", ""))
        rows_sorted.sort(key=lambda r: _intval(r, "assists"))
        rows_sorted.sort(key=lambda r: _intval(r, "goals"))
        rows_sorted.sort(key=lambda r: _intval(r, "points"), reverse=True)
    else:
        # Per-game sheets: simple alphabetical order by player.
        rows_sorted.sort(key=lambda r: r.get("player", ""))

    rows_for_print: List[List[str]] = [
        [r.get(c, "") for c in cols] for r in rows_sorted
    ]
    df = pd.DataFrame(rows_for_print, columns=cols)
    return df, cols


def _display_col_name(key: str) -> str:
    """
    Human-friendly column names: remove internal prefixes/suffixes,
    replace underscores with spaces, and capitalize words.
    """
    # Explicit overrides for common fields
    overrides = {
        "player": "Player",
        "gp": "GP",
        "goals": "Goals",
        "goals_per_shift": "Goals per Shift",
        "assists": "Assists",
        "assists_per_shift": "Assists per Shift",
        "points": "Points",
        "ppg": "PPG",
        "points_per_shift": "Points per Shift",
        "shots": "Shots",
        "shots_per_game": "Shots per Game",
        "shots_per_shift": "Shots per Shift",
        "sog": "SOG",
        "sog_per_game": "SOG per Game",
        "sog_per_shift": "SOG per Shift",
        "expected_goals": "xG",
        "expected_goals_per_game": "xG per Game",
        "expected_goals_per_shift": "xG per Shift",
        "expected_goals_per_sog": "xG per SOG",
        "turnovers_forced": "Turnovers (forced)",
        "turnovers_forced_per_game": "Turnovers (forced) per Game",
        "turnovers_forced_per_shift": "Turnovers (forced) per Shift",
        "created_turnovers": "Created Turnovers",
        "created_turnovers_per_game": "Created Turnovers per Game",
        "created_turnovers_per_shift": "Created Turnovers per Shift",
        "giveaways": "Giveaways",
        "giveaways_per_game": "Giveaways per Game",
        "giveaways_per_shift": "Giveaways per Shift",
        "takeaways": "Takeaways",
        "takeaways_per_game": "Takeaways per Game",
        "takeaways_per_shift": "Takeaways per Shift",
        "controlled_entry_for": "Controlled Entry For (On-Ice)",
        "controlled_entry_for_per_game": "Controlled Entry For (On-Ice) per Game",
        "controlled_entry_for_per_shift": "Controlled Entry For (On-Ice) per Shift",
        "controlled_entry_against": "Controlled Entry Against (On-Ice)",
        "controlled_entry_against_per_game": "Controlled Entry Against (On-Ice) per Game",
        "controlled_entry_against_per_shift": "Controlled Entry Against (On-Ice) per Shift",
        "controlled_exit_for": "Controlled Exit For (On-Ice)",
        "controlled_exit_for_per_game": "Controlled Exit For (On-Ice) per Game",
        "controlled_exit_for_per_shift": "Controlled Exit For (On-Ice) per Shift",
        "controlled_exit_against": "Controlled Exit Against (On-Ice)",
        "controlled_exit_against_per_game": "Controlled Exit Against (On-Ice) per Game",
        "controlled_exit_against_per_shift": "Controlled Exit Against (On-Ice) per Shift",
        "gt_goals": "GT Goals",
        "gw_goals": "GW Goals",
        "ot_goals": "OT Goals",
        "ot_assists": "OT Assists",
        "shifts": "Shifts",
        "shifts_per_game": "Shifts per Game",
        "plus_minus": "Goal +/-",
        "plus_minus_per_game": "Goal +/- per Game",
        "gf_counted": "GF Counted",
        "gf_per_game": "GF per Game",
        "gf_per_shift": "GF per Shift",
        "ga_counted": "GA Counted",
        "ga_per_game": "GA per Game",
        "ga_per_shift": "GA per Shift",
        "sb_toi_total": "TOI Total",
        "sb_toi_per_game": "TOI per Game",
        "sb_avg": "Average Shift",
        "sb_median": "Median Shift",
        "sb_longest": "Longest Shift",
        "sb_shortest": "Shortest Shift",
        "video_toi_total": "TOI Total (Video)",
    }
    if key in overrides:
        return overrides[key]

    # Period-specific columns
    m = re.fullmatch(r"P(\d+)_toi", key)
    if m:
        return f"Period {m.group(1)} TOI"
    m = re.fullmatch(r"P(\d+)_shifts", key)
    if m:
        return f"Period {m.group(1)} Shifts"
    m = re.fullmatch(r"P(\d+)_GF", key)
    if m:
        return f"Period {m.group(1)} GF"
    m = re.fullmatch(r"P(\d+)_GA", key)
    if m:
        return f"Period {m.group(1)} GA"

    # Generic fallback: split on underscores and capitalize words,
    # preserving common hockey/stat acronyms.
    parts = key.split("_")
    out_parts = []
    for part in parts:
        up = part.upper()
        if up in {"TOI", "GF", "GA", "GT", "GW"}:
            out_parts.append(up)
        else:
            out_parts.append(part.capitalize())
    return " ".join(out_parts)


def _display_player_name(raw: str) -> str:
    """
    Human-friendly player label from an internal key like '59_Ryan_S_Donahue'.
    Format: two-character jersey (right-aligned) + space + name with spaces.
    Example: '59_Ryan_S_Donahue' -> '59 Ryan S Donahue'
             '8_Adam_Ro'        -> ' 8 Adam Ro'
    """
    if not raw:
        return ""
    text = str(raw)
    parts = text.split("_", 1)
    if len(parts) == 2:
        jersey_part, name_part = parts
        # Extract numeric jersey if present; otherwise use the raw jersey_part.
        m = re.search(r"(\\d+)", jersey_part)
        num = m.group(1) if m else jersey_part
        jersey_fmt = f"{num:>2}"
        name = name_part.replace("_", " ").strip()
        return f"{jersey_fmt} {name}"
    # Fallback: just replace underscores with spaces
    return text.replace("_", " ")


def _display_event_type(event_type: str) -> str:
    et = str(event_type or "").strip()
    if et == "ExpectedGoal":
        return "xG"
    if et == "TurnoverForced":
        return "Turnovers (forced)"
    if et == "CreatedTurnover":
        return "Created Turnovers"
    return et


def _blink_event_label(event_type: str) -> str:
    """
    Label used for the flashing (blink) overlay in event clip scripts.

    Keep this short, uppercase, and singular (even when the displayed stat is plural).
    """
    et = str(event_type or "").strip()
    if not et:
        return ""

    # Prefer explicit mappings for cases where the displayed stat is plural or includes annotations.
    mapping: Dict[str, str] = {
        "ExpectedGoal": "xG",
        "SOG": "SOG",
        "Goal": "GOAL",
        "Assist": "ASSIST",
        "TurnoverForced": "FORCED TURNOVER",
        "CreatedTurnover": "CREATED TURNOVER",
        "Giveaway": "GIVEAWAY",
        "Takeaway": "TAKEAWAY",
    }
    if et in mapping:
        return mapping[et]

    label = re.sub(r"[()]", "", _display_event_type(et)).strip()
    if label and label != "xG":
        label = label.upper()
    return label


def _write_player_stats_text_and_csv(
    stats_dir: Path,
    stats_table_rows: List[Dict[str, str]],
    all_periods_seen: List[int],
    *,
    include_shifts_in_stats: bool,
) -> None:
    df, cols = _build_stats_dataframe(
        stats_table_rows,
        all_periods_seen,
        sort_for_cumulative=False,
        include_shifts_in_stats=include_shifts_in_stats,
    )
    # Pretty-print player names for display tables
    if "player" in df.columns:
        df["player"] = df["player"].apply(_display_player_name)
    rows_for_print = df.values.tolist()

    # Human-friendly display column names
    disp_cols = [_display_col_name(c) for c in cols]

    widths = [len(c) for c in disp_cols]
    for row in rows_for_print:
        for i, cell in enumerate(row):
            if len(str(cell)) > widths[i]:
                widths[i] = len(str(cell))

    def fmt_row(values: List[str]) -> str:
        return "  ".join(str(v).ljust(widths[i]) for i, v in enumerate(values))

    lines = [fmt_row(disp_cols)]
    lines.append(fmt_row(["-" * w for w in widths]))
    for row in rows_for_print:
        lines.append(fmt_row(row))
    (stats_dir / "player_stats.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")

    import csv  # local import

    # DataFrame with display headers for CSV/XLSX output
    df_display = df.copy()
    df_display.columns = disp_cols

    csv_rows = [dict(zip(disp_cols, row)) for row in rows_for_print]
    try:
        df_display.to_csv(stats_dir / "player_stats.csv", index=False)
    except Exception:
        with (stats_dir / "player_stats.csv").open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=disp_cols)
            w.writeheader()
            for r in csv_rows:
                w.writerow(r)
    try:
        with pd.ExcelWriter(stats_dir / "player_stats.xlsx", engine="openpyxl") as writer:
            df_excel = df_display.copy()
            df_excel.columns = [_wrap_header_after_words(c, words_per_line=2) for c in disp_cols]
            df_excel.to_excel(writer, sheet_name="player_stats", index=False, startrow=1)
            _apply_excel_table_style(writer, "player_stats", title="Player Stats", df=df_excel)
            _autosize_columns(writer, "player_stats", df_excel)
    except Exception:
        pass


def _write_game_stats_files_t2s_only(
    stats_dir: Path,
    *,
    label: str,
    t2s_id: int,
    goals: List[GoalEvent],
    periods: Optional[List[int]] = None,
) -> None:
    """
    Write per-game stats for a TimeToScore-only game (no spreadsheets).

    This mirrors `_write_game_stats_files(...)` but doesn't require an input XLSX path.
    """
    gf = sum(1 for g in goals if getattr(g, "kind", None) == "GF")
    ga = sum(1 for g in goals if getattr(g, "kind", None) == "GA")
    goal_diff = gf - ga

    row: Dict[str, Any] = {
        "T2S ID": str(int(t2s_id)),
        "Score (For-Against)": f"{gf}-{ga}",
        "Goals For": gf,
        "Goals Against": ga,
        "Goal Diff": goal_diff,
    }

    # Per-period goal totals (from goal list).
    period_set: set[int] = set()
    for x in periods or []:
        if isinstance(x, int):
            period_set.add(int(x))
    for g in goals or []:
        try:
            period_set.add(int(getattr(g, "period", 0) or 0))
        except Exception:
            continue
    for p in sorted({pp for pp in period_set if isinstance(pp, int) and pp > 0}):
        row[f"Period {p} Goals For"] = sum(1 for g in goals if g.kind == "GF" and g.period == p)
        row[f"Period {p} Goals Against"] = sum(1 for g in goals if g.kind == "GA" and g.period == p)

    # No event-log/long-sheet stats available for T2S-only games; keep keys stable with blanks.
    row.update(
        {
            "Shots For": "",
            "Shots Against": "",
            "SOG For": "",
            "SOG Against": "",
            "xG For": "",
            "xG Against": "",
            "xG per SOG (For)": "",
            "xG per SOG (Against)": "",
            "Turnovers (forced) For": "",
            "Turnovers (forced) Against": "",
            "Created Turnovers For": "",
            "Created Turnovers Against": "",
            "Giveaways For": "",
            "Giveaways Against": "",
            "Takeaways For": "",
            "Takeaways Against": "",
            "Controlled Entry For": "",
            "Controlled Entry Against": "",
            "Controlled Exit For": "",
            "Controlled Exit Against": "",
            "Rush For": "",
            "Rush Against": "",
        }
    )

    # Transpose: stats are rows, and the game label is the column header.
    df = pd.DataFrame({"Stat": list(row.keys()), label: list(row.values())})
    df.to_csv(stats_dir / "game_stats.csv", index=False)

    try:
        with pd.ExcelWriter(stats_dir / "game_stats.xlsx", engine="openpyxl") as writer:
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="game_stats", index=False, startrow=1)
            _apply_excel_table_style(writer, "game_stats", title="Game Stats", df=df_excel)
            _autosize_columns(writer, "game_stats", df_excel)
    except Exception:
        pass


def _write_game_stats_files(
    stats_dir: Path,
    *,
    xls_path: Path,
    periods: List[int],
    goals: List[GoalEvent],
    event_log_context: Optional[EventLogContext],
    focus_team: Optional[str],
) -> None:
    """
    Write per-game stats as a compact 2-column table (Stat x Value) as CSV + XLSX.

    This intentionally excludes any shift/TOI information.
    """
    label = _base_label_from_path(xls_path)
    t2s_id = _infer_t2s_from_filename(xls_path)

    gf = sum(1 for g in goals if getattr(g, "kind", None) == "GF")
    ga = sum(1 for g in goals if getattr(g, "kind", None) == "GA")
    goal_diff = gf - ga

    row: Dict[str, Any] = {
        "T2S ID": str(t2s_id) if t2s_id is not None else "",
        "Score (For-Against)": f"{gf}-{ga}",
        "Goals For": gf,
        "Goals Against": ga,
        "Goal Diff": goal_diff,
    }

    # Per-period goal totals (from goal list). Include OT if present even when the shift
    # sheet doesn't have an OT section.
    period_set: set[int] = set()
    for x in periods or []:
        if isinstance(x, int):
            period_set.add(int(x))
    for g in goals or []:
        try:
            period_set.add(int(getattr(g, "period", 0) or 0))
        except Exception:
            continue
    for p in sorted({pp for pp in period_set if isinstance(pp, int) and pp > 0}):
        row[f"Period {p} Goals For"] = sum(1 for g in goals if g.kind == "GF" and g.period == p)
        row[f"Period {p} Goals Against"] = sum(1 for g in goals if g.kind == "GA" and g.period == p)

    # Event-based team stats (from event log context; usually from '*-long*' sheets)
    counts = (event_log_context.event_counts_by_type_team if event_log_context else None) or {}
    has_event_counts = bool(counts)

    def _for_against(event_type: str) -> Tuple[Any, Any]:
        if not has_event_counts or focus_team not in {"Blue", "White"}:
            return "", ""
        opp = "White" if focus_team == "Blue" else "Blue"
        k_for = (event_type, focus_team)
        k_against = (event_type, opp)
        # If this stat type wasn't collected for this game, leave blank (not 0).
        if k_for not in counts and k_against not in counts:
            return "", ""
        return int(counts.get(k_for, 0) or 0), int(counts.get(k_against, 0) or 0)

    shots_for, shots_against = _for_against("Shot")
    sog_for, sog_against = _for_against("SOG")
    xg_for, xg_against = _for_against("ExpectedGoal")
    turnovers_forced_for, turnovers_forced_against = _for_against("TurnoverForced")
    created_turnovers_for, created_turnovers_against = _for_against("CreatedTurnover")
    giveaways_for, giveaways_against = _for_against("Giveaway")
    takeaways_for, takeaways_against = _for_against("Takeaway")
    ce_for, ce_against = _for_against("ControlledEntry")
    cx_for, cx_against = _for_against("ControlledExit")
    rush_for, rush_against = _for_against("Rush")

    # Keep columns stable; leave blank when mapping isn't available.
    row.update(
        {
            "Shots For": shots_for,
            "Shots Against": shots_against,
            "SOG For": sog_for,
            "SOG Against": sog_against,
            "xG For": xg_for,
            "xG Against": xg_against,
            "xG per SOG (For)": (
                f"{(int(xg_for) / int(sog_for)):.2f}"
                if isinstance(xg_for, (int, float))
                and isinstance(sog_for, (int, float))
                and int(sog_for) > 0
                else ""
            ),
            "xG per SOG (Against)": (
                f"{(int(xg_against) / int(sog_against)):.2f}"
                if isinstance(xg_against, (int, float))
                and isinstance(sog_against, (int, float))
                and int(sog_against) > 0
                else ""
            ),
            "Turnovers (forced) For": turnovers_forced_for,
            "Turnovers (forced) Against": turnovers_forced_against,
            "Created Turnovers For": created_turnovers_for,
            "Created Turnovers Against": created_turnovers_against,
            "Giveaways For": giveaways_for,
            "Giveaways Against": giveaways_against,
            "Takeaways For": takeaways_for,
            "Takeaways Against": takeaways_against,
            "Controlled Entry For": ce_for,
            "Controlled Entry Against": ce_against,
            "Controlled Exit For": cx_for,
            "Controlled Exit Against": cx_against,
            "Rush For": rush_for,
            "Rush Against": rush_against,
        }
    )

    # Transpose: stats are rows, and the game label is the column header.
    df = pd.DataFrame({"Stat": list(row.keys()), label: list(row.values())})
    df.to_csv(stats_dir / "game_stats.csv", index=False)

    try:
        with pd.ExcelWriter(stats_dir / "game_stats.xlsx", engine="openpyxl") as writer:
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="game_stats", index=False, startrow=1)
            _apply_excel_table_style(writer, "game_stats", title="Game Stats", df=df_excel)
            _autosize_columns(writer, "game_stats", df_excel)
    except Exception:
        pass


def _load_game_stats_csv_as_series(path: Path) -> Tuple[str, pd.Series]:
    """
    Load stats/game_stats.csv (2-column "Stat", "<game_label>") into a Series indexed by Stat.
    Returns (value_column_label, series).
    """
    df = pd.read_csv(path)
    if "Stat" not in df.columns:
        raise ValueError("missing Stat column")
    value_cols = [c for c in df.columns if c != "Stat"]
    if len(value_cols) != 1:
        raise ValueError("expected exactly one value column")
    value_col = str(value_cols[0])
    series = df.set_index("Stat")[value_col]
    # Preserve row order from the CSV.
    return value_col, series


def _write_game_stats_consolidated_files(
    base_outdir: Path,
    ordered_results: List[Dict[str, Any]],
) -> bool:
    """
    Join per-game stats/game_stats.csv tables into a single consolidated CSV/XLSX:
      - rows: Stat
      - columns: each game's label (value column header from game_stats.csv)
    """
    series_list: List[pd.Series] = []
    for r in ordered_results:
        outdir_val = r.get("outdir")
        if not outdir_val:
            continue
        outdir = Path(outdir_val)
        stats_csv = outdir / "stats" / "game_stats.csv"
        if not stats_csv.exists():
            continue
        try:
            value_col, series = _load_game_stats_csv_as_series(stats_csv)
        except Exception as e:  # noqa: BLE001
            print(f"[game_stats_consolidated] Failed to read {stats_csv}: {e}", file=sys.stderr)
            continue
        # Prefer the label from the CSV (matches per-game output); fall back to the group's label.
        label = str(value_col or r.get("label") or outdir.name)
        series_list.append(series.rename(label))

    if not series_list:
        return False

    df = pd.concat(series_list, axis=1, sort=False)
    df.index.name = "Stat"
    df_out = df.reset_index()

    csv_path = base_outdir / "game_stats_consolidated.csv"
    xlsx_path = base_outdir / "game_stats_consolidated.xlsx"
    try:
        df_out.to_csv(csv_path, index=False)
    except Exception as e:  # noqa: BLE001
        print(f"[game_stats_consolidated] Failed to write {csv_path}: {e}", file=sys.stderr)
        return False

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_excel = df_out.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="game_stats", index=False, startrow=1)
            _apply_excel_table_style(
                writer, "game_stats", title="Game Stats (All Games)", df=df_excel
            )
            _autosize_columns(writer, "game_stats", df_excel)
    except Exception as e:  # noqa: BLE001
        print(f"[game_stats_consolidated] Failed to write {xlsx_path}: {e}", file=sys.stderr)
        return False

    return True


def _write_pair_on_ice_consolidated_files(
    base_outdir: Path,
    ordered_results: List[Dict[str, Any]],
    *,
    include_toi: bool,
) -> bool:
    """
    Join per-game pair-on-ice rows into a single consolidated CSV/XLSX.

    This avoids reading the per-game CSVs so we can:
      - keep full-precision overlap math internally
      - optionally omit absolute TOI publishing when `--shifts` is not set
    """
    agg: Dict[Tuple[str, str], Dict[str, int]] = {}
    # Compute per-player total +/- from the per-game player stats rows so that:
    #  - totals match the consolidated player stats sheets
    #  - games with shift data for only one skater still contribute to totals
    total_pm_by_player: Dict[str, int] = {}
    for r in ordered_results:
        for prow in (r.get("stats") or []):
            try:
                player = str(prow.get("player") or "").strip()
                pm_raw = prow.get("plus_minus", "")
                if not player or str(pm_raw).strip() == "":
                    continue
                total_pm_by_player[player] = int(total_pm_by_player.get(player, 0) or 0) + int(str(pm_raw))
            except Exception:
                continue

        for raw in (r.get("pair_on_ice") or []):
            try:
                player = str(raw.get("player") or "").strip()
                teammate = str(raw.get("teammate") or "").strip()
                if not player or not teammate:
                    continue

                key = (player, teammate)
                dest = agg.setdefault(
                    key,
                    {
                        "shift_games": 0,
                        "player_toi_seconds": 0,
                        "overlap_seconds": 0,
                        "gf_together": 0,
                        "ga_together": 0,
                        "player_goals_on_ice_together": 0,
                        "player_assists_on_ice_together": 0,
                        "goals_collab_with_teammate": 0,
                        "assists_collab_with_teammate": 0,
                    },
                )
                dest["shift_games"] += int(raw.get("shift_games", 1) or 1)
                dest["player_toi_seconds"] += int(raw.get("player_toi_seconds", 0) or 0)
                dest["overlap_seconds"] += int(raw.get("overlap_seconds", 0) or 0)
                dest["gf_together"] += int(raw.get("gf_together", 0) or 0)
                dest["ga_together"] += int(raw.get("ga_together", 0) or 0)
                dest["player_goals_on_ice_together"] += int(raw.get("player_goals_on_ice_together", 0) or 0)
                dest["player_assists_on_ice_together"] += int(raw.get("player_assists_on_ice_together", 0) or 0)
                dest["goals_collab_with_teammate"] += int(raw.get("goals_collab_with_teammate", 0) or 0)
                dest["assists_collab_with_teammate"] += int(raw.get("assists_collab_with_teammate", 0) or 0)
            except Exception:
                continue

    rows: List[Dict[str, Any]] = []
    if not agg:
        # Still write an empty consolidated file set so users always see it
        # alongside other consolidated outputs.
        empty_cols = [
            "Player",
            "Teammate",
            "Games with Data",
            "Overlap %",
            "GF Together",
            "GA Together",
            "Player Goals (On Ice Together)",
            "Player Assists (On Ice Together)",
            "Goals Collaborated",
            "Assists Collaborated",
            "+/- Together",
            "Player Total +/-",
            "Teammate Total +/-",
        ]
        if include_toi:
            empty_cols.insert(3, "Player TOI")
            empty_cols.insert(4, "Overlap")
        df_out = pd.DataFrame([], columns=empty_cols)
        csv_out = base_outdir / "pair_on_ice_consolidated.csv"
        xlsx_out = base_outdir / "pair_on_ice_consolidated.xlsx"
        try:
            df_out.to_csv(csv_out, index=False)
        except Exception:
            return False
        _write_styled_xlsx_table(
            xlsx_out,
            df_out,
            sheet_name="pair_on_ice",
            title="Pair On-Ice (All Games)",
            number_formats={"Overlap %": "0.0"},
            text_columns=["Player", "Teammate"],
        )
        return True

    def _blank_if_zero(x: Any) -> Any:
        try:
            v = int(x or 0)
        except Exception:
            v = 0
        return "" if v == 0 else v

    for (player, teammate), d in agg.items():
        toi = int(d.get("player_toi_seconds", 0) or 0)
        overlap = int(d.get("overlap_seconds", 0) or 0)
        gf = int(d.get("gf_together", 0) or 0)
        ga = int(d.get("ga_together", 0) or 0)
        pct = 100.0 * overlap / toi if toi > 0 else 0.0
        row_out: Dict[str, Any] = {
            "Player": player,
            "Teammate": teammate,
            "Games with Data": int(d.get("shift_games", 0) or 0),
            "Overlap %": pct,
            "GF Together": gf,
            "GA Together": ga,
            "Player Goals (On Ice Together)": _blank_if_zero(d.get("player_goals_on_ice_together", 0)),
            "Player Assists (On Ice Together)": _blank_if_zero(d.get("player_assists_on_ice_together", 0)),
            "Goals Collaborated": _blank_if_zero(d.get("goals_collab_with_teammate", 0)),
            "Assists Collaborated": _blank_if_zero(d.get("assists_collab_with_teammate", 0)),
            "+/- Together": gf - ga,
            "Player Total +/-": int(total_pm_by_player.get(player, 0) or 0),
            "Teammate Total +/-": int(total_pm_by_player.get(teammate, 0) or 0),
        }
        if include_toi:
            row_out["Player TOI"] = _seconds_to_compact_hms(toi)
            row_out["Overlap"] = _seconds_to_compact_hms(overlap)
        rows.append(row_out)

    df_out = pd.DataFrame(rows)
    try:
        df_out.sort_values(by=["Player", "Overlap %"], ascending=[True, False], inplace=True)
    except Exception:
        pass

    csv_out = base_outdir / "pair_on_ice_consolidated.csv"
    xlsx_out = base_outdir / "pair_on_ice_consolidated.xlsx"
    try:
        df_out.to_csv(csv_out, index=False)
    except Exception:
        return False

    _write_styled_xlsx_table(
        xlsx_out,
        df_out,
        sheet_name="pair_on_ice",
        title="Pair On-Ice (All Games)",
        number_formats={"Overlap %": "0.0"},
        text_columns=["Player", "Teammate"],
        align_right_columns=["Player TOI", "Overlap"],
    )
    return True


def _aggregate_stats_rows(
    stats_sets: List[Tuple[List[Dict[str, str]], List[int]]],
) -> Tuple[List[Dict[str, str]], List[int], Dict[str, int]]:
    def _game_has_any_value(rows: List[Dict[str, str]], key: str) -> bool:
        for r in rows:
            if not r:
                continue
            v = r.get(key, "")
            if v is None:
                continue
            if str(v).strip() != "":
                return True
        return False

    # Count how many games have player-attributed stats available for each
    # long-sheet-derived stat key. This is used as the denominator for per-game
    # averages so that games without a `*-long*` sheet (or without a newer stat
    # type) don't dilute the per-game numbers.
    long_stat_keys = [
        "shots",
        "sog",
        "expected_goals",
        "turnovers_forced",
        "created_turnovers",
        "giveaways",
        "takeaways",
        "controlled_entry_for",
        "controlled_entry_against",
        "controlled_exit_for",
        "controlled_exit_against",
    ]
    games_with_long_stat: Dict[str, int] = {k: 0 for k in long_stat_keys}
    for rows, _periods in stats_sets:
        for k in long_stat_keys:
            if _game_has_any_value(rows, k):
                games_with_long_stat[k] += 1

    per_game_denoms: Dict[str, int] = {f"{k}_per_game": v for k, v in games_with_long_stat.items()}

    agg: Dict[str, Dict[str, Any]] = {}
    all_periods: set[int] = set()

    def _ensure(player: str) -> Dict[str, Any]:
        if player not in agg:
            agg[player] = {
                "player": player,
                "goals": 0,
                "assists": 0,
                "goals_shift": 0,
                "assists_shift": 0,
                "ot_goals": 0,
                "ot_assists": 0,
                "shots": 0,
                "sog": 0,
                "expected_goals": 0,
                "turnovers_forced": 0,
                "created_turnovers": 0,
                "giveaways": 0,
                "takeaways": 0,
                "controlled_entry_for": 0,
                "controlled_entry_against": 0,
                "controlled_exit_for": 0,
                "controlled_exit_against": 0,
                "gp": 0,
                "shifts": 0,
                "plus_minus": 0,
                "gf_counted": 0,
                "ga_counted": 0,
                "gf_counted_shift": 0,
                "ga_counted_shift": 0,
                "sb_toi_total_sec": 0,
                "video_toi_total_sec": 0,
                "sb_longest_sec": 0,
                "sb_shortest_sec": None,
                # Per-shift denominators (sum of shifts across games where the stat is available).
                "shots_shifts_denom": 0,
                "sog_shifts_denom": 0,
                "expected_goals_shifts_denom": 0,
                "turnovers_forced_shifts_denom": 0,
                "created_turnovers_shifts_denom": 0,
                "giveaways_shifts_denom": 0,
                "takeaways_shifts_denom": 0,
                "controlled_entry_for_shifts_denom": 0,
                "controlled_entry_against_shifts_denom": 0,
                "controlled_exit_for_shifts_denom": 0,
                "controlled_exit_against_shifts_denom": 0,
            }
        return agg[player]

    for rows, periods in stats_sets:
        for p in periods:
            all_periods.add(p)
        for row in rows:
            player = row.get("player", "")
            if not player:
                continue
            dest = _ensure(player)
            dest["goals"] += int(str(row.get("goals", 0) or 0))
            dest["assists"] += int(str(row.get("assists", 0) or 0))
            dest["ot_goals"] += int(str(row.get("ot_goals", 0) or 0))
            dest["ot_assists"] += int(str(row.get("ot_assists", 0) or 0))
            dest["shots"] += int(str(row.get("shots", 0) or 0))
            dest["sog"] += int(str(row.get("sog", 0) or 0))
            dest["expected_goals"] += int(str(row.get("expected_goals", 0) or 0))
            dest["turnovers_forced"] += int(str(row.get("turnovers_forced", 0) or 0))
            dest["created_turnovers"] += int(str(row.get("created_turnovers", 0) or 0))
            dest["giveaways"] += int(str(row.get("giveaways", 0) or 0))
            dest["takeaways"] += int(str(row.get("takeaways", 0) or 0))
            dest["controlled_entry_for"] += int(str(row.get("controlled_entry_for", 0) or 0))
            dest["controlled_entry_against"] += int(
                str(row.get("controlled_entry_against", 0) or 0)
            )
            dest["controlled_exit_for"] += int(str(row.get("controlled_exit_for", 0) or 0))
            dest["controlled_exit_against"] += int(
                str(row.get("controlled_exit_against", 0) or 0)
            )
            # Each per-game stats row corresponds to one game played (GP),
            # including cases where the player only appears on the T2S roster.
            dest["gp"] += 1
            shifts_raw = str(row.get("shifts", "") or "").strip()
            shifts_i = int(str(shifts_raw or 0))
            has_shift_sheet = shifts_raw != ""
            dest["shifts"] += shifts_i
            dest["plus_minus"] += int(str(row.get("plus_minus", 0) or 0))
            dest["gf_counted"] += int(str(row.get("gf_counted", 0) or 0))
            dest["ga_counted"] += int(str(row.get("ga_counted", 0) or 0))

            # Per-shift numerators should only use games where shift times exist.
            if has_shift_sheet:
                dest["goals_shift"] += int(str(row.get("goals", 0) or 0))
                dest["assists_shift"] += int(str(row.get("assists", 0) or 0))
                dest["gf_counted_shift"] += int(str(row.get("gf_counted", 0) or 0))
                dest["ga_counted_shift"] += int(str(row.get("ga_counted", 0) or 0))

                # Per-shift denominators are stat-specific: only include shifts from games
                # where that stat is actually present (e.g., long sheets existed).
                def _has_stat(key: str) -> bool:
                    return str(row.get(key, "") or "").strip() != ""

                if shifts_i > 0:
                    for k in [
                        "shots",
                        "sog",
                        "expected_goals",
                        "turnovers_forced",
                        "created_turnovers",
                        "giveaways",
                        "takeaways",
                        "controlled_entry_for",
                        "controlled_entry_against",
                        "controlled_exit_for",
                        "controlled_exit_against",
                    ]:
                        if _has_stat(k):
                            denom_key = f"{k}_shifts_denom"
                            dest[denom_key] = int(dest.get(denom_key, 0) or 0) + shifts_i

            dest["sb_toi_total_sec"] += _duration_to_seconds(row.get("sb_toi_total", ""))
            dest["video_toi_total_sec"] += _duration_to_seconds(row.get("video_toi_total", ""))
            longest = _duration_to_seconds(row.get("sb_longest", ""))
            if longest > dest["sb_longest_sec"]:
                dest["sb_longest_sec"] = longest
            shortest = _duration_to_seconds(row.get("sb_shortest", ""))
            if shortest > 0:
                if dest["sb_shortest_sec"] is None or shortest < dest["sb_shortest_sec"]:
                    dest["sb_shortest_sec"] = shortest
            # per-period counts and toi
            for key, val in row.items():
                if not isinstance(key, str):
                    continue
                if re.fullmatch(r"P\d+_shifts", key):
                    dest[key] = dest.get(key, 0) + int(str(val or 0))
                elif re.fullmatch(r"P\d+_GF", key):
                    dest[key] = dest.get(key, 0) + int(str(val or 0))
                elif re.fullmatch(r"P\d+_GA", key):
                    dest[key] = dest.get(key, 0) + int(str(val or 0))
                elif re.fullmatch(r"P\d+_toi", key):
                    dest[key] = dest.get(key, 0) + _duration_to_seconds(val or "")

    aggregated_rows: List[Dict[str, str]] = []
    for player, data in sorted(agg.items(), key=lambda x: x[0]):
        gp = data.get("gp", 0) or 0
        shifts = data["shifts"] or 0
        total_sec = data["sb_toi_total_sec"]
        avg_sec = int(total_sec / shifts) if shifts else 0
        total_goals = data["goals"]
        total_assists = data["assists"]
        goals_shift = data.get("goals_shift", 0) or 0
        assists_shift = data.get("assists_shift", 0) or 0
        points_shift = goals_shift + assists_shift
        total_ot_goals = data.get("ot_goals", 0) or 0
        total_ot_assists = data.get("ot_assists", 0) or 0
        total_points = total_goals + total_assists
        total_shots = data.get("shots", 0) or 0
        total_sog = data.get("sog", 0) or 0
        total_expected_goals = data.get("expected_goals", 0) or 0
        total_turnovers_forced = data.get("turnovers_forced", 0) or 0
        total_created_turnovers = data.get("created_turnovers", 0) or 0
        total_giveaways = data.get("giveaways", 0) or 0
        total_takeaways = data.get("takeaways", 0) or 0
        total_ce_for = data.get("controlled_entry_for", 0) or 0
        total_ce_against = data.get("controlled_entry_against", 0) or 0
        total_cx_for = data.get("controlled_exit_for", 0) or 0
        total_cx_against = data.get("controlled_exit_against", 0) or 0
        shots_games = per_game_denoms.get("shots_per_game", 0) or 0
        sog_games = per_game_denoms.get("sog_per_game", 0) or 0
        xg_games = per_game_denoms.get("expected_goals_per_game", 0) or 0
        turnovers_forced_games = per_game_denoms.get("turnovers_forced_per_game", 0) or 0
        created_turnovers_games = per_game_denoms.get("created_turnovers_per_game", 0) or 0
        giveaway_games = per_game_denoms.get("giveaways_per_game", 0) or 0
        takeaway_games = per_game_denoms.get("takeaways_per_game", 0) or 0
        ce_for_games = per_game_denoms.get("controlled_entry_for_per_game", 0) or 0
        ce_against_games = per_game_denoms.get("controlled_entry_against_per_game", 0) or 0
        cx_for_games = per_game_denoms.get("controlled_exit_for_per_game", 0) or 0
        cx_against_games = per_game_denoms.get("controlled_exit_against_per_game", 0) or 0
        row: Dict[str, str] = {
            "player": player,
            "gp": str(gp),
            "goals": str(total_goals),
            "assists": str(total_assists),
            "ot_goals": str(total_ot_goals),
            "ot_assists": str(total_ot_assists),
            "points": str(total_points),
            "ppg": f"{(total_points / gp):.1f}" if gp > 0 else "0.0",
            "shots": str(total_shots) if shots_games > 0 else "",
            "shots_per_game": f"{(total_shots / shots_games):.1f}" if shots_games > 0 else "",
            "sog": str(total_sog) if sog_games > 0 else "",
            "sog_per_game": f"{(total_sog / sog_games):.1f}" if sog_games > 0 else "",
            "expected_goals": str(total_expected_goals) if xg_games > 0 else "",
            "expected_goals_per_game": (
                f"{(total_expected_goals / xg_games):.1f}" if xg_games > 0 else ""
            ),
            "expected_goals_per_sog": (
                f"{(total_expected_goals / total_sog):.2f}"
                if xg_games > 0 and sog_games > 0 and total_sog > 0
                else ""
            ),
            "turnovers_forced": str(total_turnovers_forced) if turnovers_forced_games > 0 else "",
            "turnovers_forced_per_game": (
                f"{(total_turnovers_forced / turnovers_forced_games):.1f}"
                if turnovers_forced_games > 0
                else ""
            ),
            "created_turnovers": str(total_created_turnovers) if created_turnovers_games > 0 else "",
            "created_turnovers_per_game": (
                f"{(total_created_turnovers / created_turnovers_games):.1f}"
                if created_turnovers_games > 0
                else ""
            ),
            "giveaways": str(total_giveaways) if giveaway_games > 0 else "",
            "giveaways_per_game": (
                f"{(total_giveaways / giveaway_games):.1f}" if giveaway_games > 0 else ""
            ),
            "takeaways": str(total_takeaways) if takeaway_games > 0 else "",
            "takeaways_per_game": (
                f"{(total_takeaways / takeaway_games):.1f}" if takeaway_games > 0 else ""
            ),
            "controlled_entry_for": str(total_ce_for) if ce_for_games > 0 else "",
            "controlled_entry_for_per_game": (
                f"{(total_ce_for / ce_for_games):.1f}" if ce_for_games > 0 else ""
            ),
            "controlled_entry_against": str(total_ce_against) if ce_against_games > 0 else "",
            "controlled_entry_against_per_game": (
                f"{(total_ce_against / ce_against_games):.1f}" if ce_against_games > 0 else ""
            ),
            "controlled_exit_for": str(total_cx_for) if cx_for_games > 0 else "",
            "controlled_exit_for_per_game": (
                f"{(total_cx_for / cx_for_games):.1f}" if cx_for_games > 0 else ""
            ),
            "controlled_exit_against": str(total_cx_against) if cx_against_games > 0 else "",
            "controlled_exit_against_per_game": (
                f"{(total_cx_against / cx_against_games):.1f}" if cx_against_games > 0 else ""
            ),
            "shifts": str(shifts),
            "shifts_per_game": f"{(shifts / gp):.1f}" if gp > 0 else "",
            "goals_per_shift": _fmt_per_shift(goals_shift, shifts),
            "assists_per_shift": _fmt_per_shift(assists_shift, shifts),
            "points_per_shift": _fmt_per_shift(points_shift, shifts),
            "plus_minus": str(data["plus_minus"]),
            "plus_minus_per_game": (
                f"{(data['plus_minus'] / gp):.1f}" if gp > 0 else ""
            ),
            "gf_counted": str(data["gf_counted"]),
            "gf_per_game": f"{(data['gf_counted'] / gp):.1f}" if gp > 0 else "",
            "ga_counted": str(data["ga_counted"]),
            "ga_per_game": f"{(data['ga_counted'] / gp):.1f}" if gp > 0 else "",
            "gf_per_shift": _fmt_per_shift(int(data.get("gf_counted_shift", 0) or 0), shifts),
            "ga_per_shift": _fmt_per_shift(int(data.get("ga_counted_shift", 0) or 0), shifts),
            "sb_toi_total": _format_duration(total_sec),
            "sb_toi_per_game": (
                _format_duration(int(total_sec / gp)) if gp > 0 and total_sec > 0 else ""
            ),
            "sb_avg": _format_duration(avg_sec) if shifts else "",
            "sb_median": "",
            "sb_longest": _format_duration(data["sb_longest_sec"]),
            "sb_shortest": (
                _format_duration(data["sb_shortest_sec"] or 0) if data["sb_shortest_sec"] else ""
            ),
            "video_toi_total": _format_duration(data["video_toi_total_sec"]),
        }

        # Stat-specific per-shift rates (only over games where the stat exists).
        row["shots_per_shift"] = _fmt_per_shift(total_shots, int(data.get("shots_shifts_denom", 0) or 0))
        row["sog_per_shift"] = _fmt_per_shift(total_sog, int(data.get("sog_shifts_denom", 0) or 0))
        row["expected_goals_per_shift"] = _fmt_per_shift(
            total_expected_goals, int(data.get("expected_goals_shifts_denom", 0) or 0)
        )
        row["turnovers_forced_per_shift"] = _fmt_per_shift(
            total_turnovers_forced, int(data.get("turnovers_forced_shifts_denom", 0) or 0)
        )
        row["created_turnovers_per_shift"] = _fmt_per_shift(
            total_created_turnovers, int(data.get("created_turnovers_shifts_denom", 0) or 0)
        )
        row["giveaways_per_shift"] = _fmt_per_shift(total_giveaways, int(data.get("giveaways_shifts_denom", 0) or 0))
        row["takeaways_per_shift"] = _fmt_per_shift(total_takeaways, int(data.get("takeaways_shifts_denom", 0) or 0))
        row["controlled_entry_for_per_shift"] = _fmt_per_shift(
            total_ce_for, int(data.get("controlled_entry_for_shifts_denom", 0) or 0)
        )
        row["controlled_entry_against_per_shift"] = _fmt_per_shift(
            total_ce_against, int(data.get("controlled_entry_against_shifts_denom", 0) or 0)
        )
        row["controlled_exit_for_per_shift"] = _fmt_per_shift(
            total_cx_for, int(data.get("controlled_exit_for_shifts_denom", 0) or 0)
        )
        row["controlled_exit_against_per_shift"] = _fmt_per_shift(
            total_cx_against, int(data.get("controlled_exit_against_shifts_denom", 0) or 0)
        )
        for p in sorted(all_periods):
            toi_key = f"P{p}_toi"
            shift_key = f"P{p}_shifts"
            gf_key = f"P{p}_GF"
            ga_key = f"P{p}_GA"
            if toi_key in data:
                row[toi_key] = _format_duration(int(data[toi_key]))
            if shift_key in data:
                row[shift_key] = str(data[shift_key])
            if gf_key in data:
                row[gf_key] = str(data[gf_key])
            if ga_key in data:
                row[ga_key] = str(data[ga_key])
        aggregated_rows.append(row)

    return aggregated_rows, sorted(all_periods), per_game_denoms


def _augment_aggregate_with_goal_details(
    aggregated_rows: List[Dict[str, str]],
    per_player_events: Dict[str, Dict[str, List[Tuple[str, GoalEvent]]]],
) -> None:
    """
    Enrich aggregated rows with game-tying / game-winning goal counts
    derived from per-player scoring events across all games.
    """
    if not aggregated_rows:
        return

    totals: Dict[str, Dict[str, int]] = {}
    for player, events in per_player_events.items():
        gt = 0
        gw = 0
        goals = events.get("goals", [])
        for _label, ev in goals:
            if getattr(ev, "is_game_tying", False):
                gt += 1
            if getattr(ev, "is_game_winning", False):
                gw += 1
        totals[player] = {"gt_goals": gt, "gw_goals": gw}

    for row in aggregated_rows:
        player = row.get("player", "")
        t = totals.get(player, {})
        row["gt_goals"] = str(t.get("gt_goals", 0))
        row["gw_goals"] = str(t.get("gw_goals", 0))


def _write_consolidated_workbook(
    out_path: Path,
    sheets: List[Tuple[str, pd.DataFrame]],
    *,
    per_game_denoms: Optional[Dict[str, int]] = None,
) -> None:
    def _disp_col(key: str, *, is_cumulative: bool) -> str:
        base = _display_col_name(key)
        if is_cumulative and per_game_denoms and key in per_game_denoms:
            return f"{base} ({per_game_denoms[key]})"
        return base

    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for name, df in sheets:
                safe_name = re.sub(r"[:\\\\/?*\\[\\]]", "_", name or "Sheet")[:31]
                df_display = df.copy()
                # Pretty player names if present
                if "player" in df_display.columns:
                    df_display["player"] = df_display["player"].apply(_display_player_name)
                is_cumulative = str(name or "").strip().lower() == "cumulative"
                disp_cols = [
                    _wrap_header_after_words(_disp_col(c, is_cumulative=is_cumulative), words_per_line=2)
                    for c in df_display.columns
                ]
                df_display.columns = disp_cols
                df_display.to_excel(writer, sheet_name=safe_name, index=False, startrow=1)
                _apply_excel_table_style(writer, safe_name, title=(name or safe_name), df=df_display)
                _autosize_columns(writer, safe_name, df_display)
    except Exception:
        pass


def _write_cumulative_player_detail_files(
    base_outdir: Path,
    aggregated_rows: List[Dict[str, str]],
    per_player_events: Dict[str, Dict[str, List[Tuple[str, GoalEvent]]]],
    per_game_stats_by_label: Dict[str, List[Dict[str, str]]],
    per_game_pair_on_ice_by_label: Dict[str, List[Dict[str, Any]]],
    *,
    include_shifts_in_stats: bool,
) -> None:
    """
    Write one cumulative per-player stats file summarizing all games,
    including lists of goals, assists, and goals-against with game/period/time
    and game-tying / game-winning annotations for goals.
    """
    if not aggregated_rows:
        return

    outdir = base_outdir / "cumulative_per_player"
    outdir.mkdir(parents=True, exist_ok=True)

    rows_by_player: Dict[str, Dict[str, str]] = {
        r.get("player", ""): r for r in aggregated_rows if r.get("player")
    }

    # Determine which game each player's longest / shortest shift came from (optional).
    longest_by_player: Dict[str, Tuple[int, str, str]] = {}
    shortest_by_player: Dict[str, Tuple[int, str, str]] = {}
    if include_shifts_in_stats:
        for game_label, rows in per_game_stats_by_label.items():
            for row in rows:
                player = row.get("player", "")
                if not player:
                    continue
                sb_long = _duration_to_seconds(row.get("sb_longest", ""))
                if sb_long > 0:
                    cur = longest_by_player.get(player)
                    if cur is None or sb_long > cur[0]:
                        longest_by_player[player] = (sb_long, row.get("sb_longest", ""), game_label)
                sb_short = _duration_to_seconds(row.get("sb_shortest", ""))
                if sb_short > 0:
                    cur_s = shortest_by_player.get(player)
                    if cur_s is None or sb_short < cur_s[0]:
                        shortest_by_player[player] = (sb_short, row.get("sb_shortest", ""), game_label)

    pair_overlap: Dict[Tuple[str, str], Dict[str, int]] = {}
    if per_game_pair_on_ice_by_label:
        for _game_label, rows in per_game_pair_on_ice_by_label.items():
            for raw in rows or []:
                try:
                    player = str(raw.get("player") or "").strip()
                    teammate = str(raw.get("teammate") or "").strip()
                    if not player or not teammate:
                        continue
                    key = (player, teammate)
                    dest = pair_overlap.setdefault(
                        key,
                        {
                            "overlap_seconds": 0,
                            "player_toi_seconds": 0,
                            "gf_together": 0,
                            "ga_together": 0,
                            "shift_games": 0,
                        },
                    )
                    dest["overlap_seconds"] += int(raw.get("overlap_seconds", 0) or 0)
                    dest["player_toi_seconds"] += int(raw.get("player_toi_seconds", 0) or 0)
                    dest["gf_together"] += int(raw.get("gf_together", 0) or 0)
                    dest["ga_together"] += int(raw.get("ga_together", 0) or 0)
                    dest["shift_games"] += int(raw.get("shift_games", 1) or 1)
                except Exception:
                    continue

    def _fmt_tags(ev: GoalEvent) -> str:
        tags: List[str] = []
        if getattr(ev, "is_game_tying", False):
            tags.append("GT")
        if getattr(ev, "is_game_winning", False):
            tags.append("GW")
        return f" [{' '.join(tags)}]" if tags else ""

    for player, row in sorted(rows_by_player.items()):
        events = per_player_events.get(player, {})
        goals_list = events.get("goals", [])
        assists_list = events.get("assists", [])
        gf_on_ice_list = events.get("gf_on_ice", [])
        ga_list = events.get("ga_on_ice", [])

        lines: List[str] = []
        lines.append(f"Player: {_display_player_name(player)}")
        lines.append("")
        lines.append("Overall stats (all games):")
        gp_str = row.get("gp", "0")
        points_str = row.get("points", "0")
        ppg_str = row.get("ppg", "0.0")
        lines.append(f"  Games Played (GP): {gp_str}")
        lines.append(f"  Points (G+A): {points_str} (PPG: {ppg_str})")
        lines.append(
            f"  Goals: {row.get('goals', '0')} "
            f"(GT: {row.get('gt_goals', '0')}, GW: {row.get('gw_goals', '0')}, OT: {row.get('ot_goals', '0')})"
        )
        lines.append(f"  Assists: {row.get('assists', '0')} (OT: {row.get('ot_assists', '0')})")
        lines.append(f"  Goal +/-: {row.get('plus_minus', '0')}")
        if include_shifts_in_stats and row.get("shifts_per_game"):
            lines.append(f"  Shifts per game: {row.get('shifts_per_game')}")
        if row.get("plus_minus_per_game"):
            lines.append(f"  Goal +/- per game: {row.get('plus_minus_per_game')}")
        lines.append(
            f"  Goals For while on ice: {row.get('gf_counted', '0')}, "
            f"Goals Against while on ice: {row.get('ga_counted', '0')}"
        )
        if row.get("gf_per_game") or row.get("ga_per_game"):
            lines.append(
                f"  GF per game: {row.get('gf_per_game', '') or '0.0'}, "
                f"GA per game: {row.get('ga_per_game', '') or '0.0'}"
            )
        if include_shifts_in_stats:
            if row.get("sb_toi_total"):
                lines.append(f"  TOI total (scoreboard): {row.get('sb_toi_total')}")
                if row.get("sb_toi_per_game"):
                    lines.append(
                        f"  TOI per game (scoreboard): {row.get('sb_toi_per_game')}"
                    )
            if row.get("video_toi_total"):
                lines.append(f"  TOI total (video): {row.get('video_toi_total')}")
            # Longest/shortest shift games
            long_info = longest_by_player.get(player)
            if long_info is not None:
                _, dur, game_label = long_info
                lines.append(f"  Longest shift (scoreboard): {dur} ({game_label})")
            short_info = shortest_by_player.get(player)
            if short_info is not None:
                _, dur_s, game_s = short_info
                lines.append(f"  Shortest shift (scoreboard): {dur_s} ({game_s})")

            # Pair on-ice overlap across all games that had shift data.
            teammates = sorted({t for (p, t) in pair_overlap.keys() if p == player and t})
            if teammates:
                rows_disp: List[Tuple[str, float, int, int, int, int, int]] = []
                for teammate in teammates:
                    d = pair_overlap.get(
                        (player, teammate),
                        {
                            "overlap_seconds": 0,
                            "player_toi_seconds": 0,
                            "gf_together": 0,
                            "ga_together": 0,
                            "shift_games": 0,
                        },
                    )
                    overlap_s = int(d.get("overlap_seconds", 0) or 0)
                    denom_toi = int(d.get("player_toi_seconds", 0) or 0)
                    gf = int(d.get("gf_together", 0) or 0)
                    ga = int(d.get("ga_together", 0) or 0)
                    shift_games = int(d.get("shift_games", 0) or 0)
                    pct = 100.0 * overlap_s / denom_toi if denom_toi > 0 else 0.0
                    rows_disp.append(
                        (
                            _display_player_name(teammate),
                            float(pct),
                            int(shift_games),
                            int(gf) - int(ga),
                            int(gf),
                            int(ga),
                            int(overlap_s),
                        )
                    )
                rows_disp.sort(key=lambda x: x[1], reverse=True)
                name_w = max([len(r[0]) for r in rows_disp] + [len("Teammate")])
                lines.append("")
                lines.append("On-ice with teammates (by TOI%, shift games only):")
                if include_shifts_in_stats:
                    lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA  Overlap")
                    for teammate_disp, pct, shift_games, pm, gf, ga, overlap_s in rows_disp:
                        lines.append(
                            f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}  {_format_duration(overlap_s):>7}"
                        )
                else:
                    # Without --shifts, do not publish any absolute time values.
                    lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA")
                    for teammate_disp, pct, shift_games, pm, gf, ga, _overlap_s in rows_disp:
                        lines.append(
                            f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}"
                        )

        # Goals
        lines.append("")
        lines.append("Goals detail:")
        if not goals_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(
                goals_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)
            ):
                lines.append(
                    f"  {game_label}: Period {ev.period}, {ev.t_str}{_fmt_tags(ev)}"
                )

        # Assists
        lines.append("")
        lines.append("Assists detail:")
        if not assists_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(
                assists_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)
            ):
                lines.append(
                    f"  {game_label}: Period {ev.period}, {ev.t_str}{_fmt_tags(ev)}"
                )

        # Goals for / against while on ice
        lines.append("")
        lines.append("Goals for while on ice:")
        if not gf_on_ice_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(
                gf_on_ice_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)
            ):
                lines.append(
                    f"  {game_label}: Period {ev.period}, {ev.t_str}{_fmt_tags(ev)}"
                )

        lines.append("")
        lines.append("Goals against while on ice:")
        if not ga_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(
                ga_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)
            ):
                lines.append(f"  {game_label}: Period {ev.period}, {ev.t_str}")

        (outdir / f"{player}_cumulative_stats.txt").write_text(
            "\n".join(lines) + "\n", encoding="utf-8"
        )


def _write_season_highlight_scripts(
    base_outdir: Path,
    results: List[Dict[str, Any]],
    *,
    create_scripts: bool,
) -> None:
    """
    For multi-game runs, write no-arg per-player season highlight scripts that:
      - clip per-game highlights from `events_Highlights_<player>_video_times.txt`
      - use each game's `tracking_output-with-audio*.mp4` automatically
      - join per-game highlight clips in game order
    """
    if not create_scripts:
        return
    if not results:
        return

    season_dir = base_outdir / "season_highlights"
    season_dir.mkdir(parents=True, exist_ok=True)

    prefix = "events_Highlights_"
    suffix = "_video_times.txt"

    def _has_timestamps(path: Path) -> bool:
        try:
            with path.open("r", encoding="utf-8") as f:
                for line in f:
                    s = line.strip()
                    if not s or s.startswith("#"):
                        continue
                    return True
        except Exception:
            return False
        return False

    players: set[str] = set()
    for r in results:
        outdir = r.get("outdir")
        if not outdir:
            continue
        try:
            for p in Path(outdir).glob(f"{prefix}*{suffix}"):
                name = p.name
                if not name.startswith(prefix) or not name.endswith(suffix):
                    continue
                pk = name[len(prefix) : -len(suffix)]
                if pk:
                    players.add(pk)
        except Exception:
            continue

    if not players:
        return

    written_scripts: List[str] = []
    for player_key in sorted(players):
        # Determine which games this player has highlight events for, in game order.
        game_entries: List[Tuple[str, Path, Path, str]] = []
        missing_videos: List[Tuple[str, str]] = []
        for r in results:
            sheet_path = r.get("sheet_path")
            if sheet_path is None:
                # T2S-only games don't have a reliable scoreboard->video mapping.
                continue
            game_label = str(r.get("label") or "")
            outdir = Path(r.get("outdir"))
            ts_file = outdir / f"{prefix}{player_key}{suffix}"
            if not ts_file.exists() or not _has_timestamps(ts_file):
                continue

            video_path = r.get("video_path")
            video = Path(video_path) if video_path is not None else _find_tracking_output_video_for_sheet_path(Path(sheet_path))
            if video is None or not video.exists():
                missing_videos.append((game_label, str(video) if video is not None else "<missing>"))
                continue

            game_entries.append((game_label, video, ts_file, sanitize_name(game_label)))

        if missing_videos:
            miss_str = ", ".join(f"{g} -> {vp}" for g, vp in missing_videos)
            print(
                f"[season-highlights] WARNING: Missing video(s) for {_display_player_name(player_key)}: {miss_str}",
                file=sys.stderr,
            )

        if not game_entries:
            continue

        player_out_dir = season_dir / player_key
        script_path = season_dir / f"clip_season_highlights_{player_key}.sh"

        script_lines: List[str] = [
            "#!/usr/bin/env bash",
            "set -euo pipefail",
            "",
            "# Optional flags:",
            "#   --quick / -q   lower quality, faster",
            "#   --hq           lossless intermediates (requires NVENC)",
            "",
            "QUICK=0",
            "HQ=0",
            "for ARG in \"$@\"; do",
            "  if [ \"$ARG\" = \"--quick\" ] || [ \"$ARG\" = \"-q\" ]; then",
            "    QUICK=1",
            "  elif [ \"$ARG\" = \"--hq\" ]; then",
            "    HQ=1",
            "  fi",
            "done",
            "",
            "EXTRA_FLAGS=()",
            "if [ \"$QUICK\" -gt 0 ]; then",
            "  EXTRA_FLAGS+=(\"--quick\" \"1\")",
            "fi",
            "if [ \"$HQ\" -gt 0 ]; then",
            "  export VIDEO_CLIPPER_HQ=1",
            "fi",
            "",
            "THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"",
            f"OUT_DIR=\"$THIS_DIR/{player_key}\"",
            "mkdir -p \"$OUT_DIR\"",
            "",
            "GAME_CLIPS=()",
            "",
        ]

        for game_label, video, ts_file, game_safe in game_entries:
            try:
                video_abs = str(Path(video).resolve())
            except Exception:
                video_abs = str(video)
            try:
                ts_abs = str(Path(ts_file).resolve())
            except Exception:
                ts_abs = str(ts_file)
            label_safe = sanitize_name(f"{player_key}__{game_label}__highlights")
            temp_dir = f"$THIS_DIR/temp_clips/{player_key}/{game_safe}"
            script_lines.extend(
                [
                    f"echo \"[{_display_player_name(player_key)}] {game_label}\"",
                    f"VIDEO=\"{video_abs}\"",
                    f"TS_FILE=\"{ts_abs}\"",
                    f"TEMP_DIR=\"{temp_dir}\"",
                    "mkdir -p \"$TEMP_DIR\"",
                    "(",
                    "  cd \"$OUT_DIR\"",
                    "  python -m hmlib.cli.video_clipper -j 4 --input \"$VIDEO\" --timestamps \"$TS_FILE\" --temp-dir \"$TEMP_DIR\" "
                    f"\"{label_safe}\" \"${{EXTRA_FLAGS[@]}}\"",
                    ")",
                    f"GAME_CLIPS+=(\"$OUT_DIR/clips-{label_safe}.mp4\")",
                    "",
                ]
            )

        list_file = "$OUT_DIR/season_clips.txt"
        season_label_safe = sanitize_name(f"{player_key}__season_highlights")

        script_lines.extend(
            [
                f"LIST_FILE=\"{list_file}\"",
                ": > \"$LIST_FILE\"",
                "for f in \"${GAME_CLIPS[@]}\"; do",
                "  echo \"file '$f'\" >> \"$LIST_FILE\"",
                "done",
                "",
                f"OUT_FILE=\"$OUT_DIR/clips-{season_label_safe}.mp4\"",
                "ffmpeg -f concat -safe 0 -i \"$LIST_FILE\" -c copy \"$OUT_FILE\"",
                "",
                "echo \"Wrote: $OUT_FILE\"",
                "",
            ]
        )

        script_path.write_text("\n".join(script_lines) + "\n", encoding="utf-8")
        try:
            os.chmod(script_path, 0o755)
        except Exception:
            pass

        written_scripts.append(script_path.name)

    if written_scripts:
        runner = season_dir / "clip_season_highlights_all.sh"
        runner_body = """#!/usr/bin/env bash
set -euo pipefail
THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"
for s in \"$THIS_DIR\"/clip_season_highlights_*.sh; do
  [ -x \"$s\" ] || continue
  if [ \"$s\" = \"$THIS_DIR/clip_season_highlights_all.sh\" ]; then
    continue
  fi
  echo \"Running $s...\"
  \"$s\" \"$@\"
done
"""
        runner.write_text(runner_body, encoding="utf-8")
        try:
            os.chmod(runner, 0o755)
        except Exception:
            pass


def _infer_side_from_rosters(
    t2s_id: int,
    jersey_numbers: set[str],
    hockey_db_dir: Path,
    *,
    debug: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        import_err = _t2s_api_import_error
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": False,
                    "t2s_api_import_error": import_err,
                    "jerseys_in_sheet_count": len(jersey_numbers or set()),
                    "jerseys_in_sheet": sorted(list(jersey_numbers or set()), key=lambda x: int(x)),
                    "failure": "t2s_api_not_available",
                }
            )
        details = f": {import_err}" if import_err else ""
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: TimeToScore API not available "
            f"(failed to import hmlib.time2score.api){details}.",
            file=sys.stderr,
        )
        return None
    try:
        with _working_directory(hockey_db_dir):
            info = t2s_api.get_game_details(int(t2s_id))
    except Exception as e:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "jerseys_in_sheet_count": len(jersey_numbers or set()),
                    "jerseys_in_sheet": sorted(list(jersey_numbers or set()), key=lambda x: int(x)),
                    "failure": "t2s_fetch_failed",
                    "exception": str(e),
                }
            )
        print(f"[t2s] Failed to load game {t2s_id} for side inference: {e}", file=sys.stderr)
        return None
    stats = (info or {}).get("stats") or {}
    home_players = stats.get("homePlayers") or []
    away_players = stats.get("awayPlayers") or []

    def _nums(rows: Any) -> set[str]:
        out: set[str] = set()
        for r in rows or []:
            num = _normalize_jersey_number((r or {}).get("number"))
            if num:
                out.add(num)
        return out

    home_set = _nums(home_players)
    away_set = _nums(away_players)

    if not jersey_numbers:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "home_roster_count": len(home_set),
                    "away_roster_count": len(away_set),
                    "failure": "no_jersey_numbers_in_sheet",
                }
            )
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: no jersey numbers found in sheet.",
            file=sys.stderr,
        )
        return None
    if not home_set and not away_set:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "jerseys_in_sheet_count": len(jersey_numbers),
                    "jerseys_in_sheet": sorted(list(jersey_numbers), key=lambda x: int(x)),
                    "home_roster_count": 0,
                    "away_roster_count": 0,
                    "failure": "no_roster_numbers_in_t2s",
                }
            )
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: no roster numbers in TimeToScore stats.",
            file=sys.stderr,
        )
        return None

    home_overlap = len(jersey_numbers & home_set)
    away_overlap = len(jersey_numbers & away_set)

    if home_overlap == away_overlap:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "jerseys_in_sheet_count": len(jersey_numbers),
                    "jerseys_in_sheet": sorted(list(jersey_numbers), key=lambda x: int(x)),
                    "home_roster_count": len(home_set),
                    "away_roster_count": len(away_set),
                    "home_overlap": home_overlap,
                    "away_overlap": away_overlap,
                    "home_overlap_jerseys": sorted(
                        list(jersey_numbers & home_set), key=lambda x: int(x)
                    ),
                    "away_overlap_jerseys": sorted(
                        list(jersey_numbers & away_set), key=lambda x: int(x)
                    ),
                    "failure": "overlap_tie",
                }
            )
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: overlap tie (home={home_overlap}, away={away_overlap}).",
            file=sys.stderr,
        )
        return None
    side = "home" if home_overlap > away_overlap else "away"
    if debug is not None:
        debug.clear()
        debug.update(
            {
                "t2s_id": int(t2s_id),
                "t2s_api_available": True,
                "jerseys_in_sheet_count": len(jersey_numbers),
                "jerseys_in_sheet": sorted(list(jersey_numbers), key=lambda x: int(x)),
                "home_roster_count": len(home_set),
                "away_roster_count": len(away_set),
                "home_overlap": home_overlap,
                "away_overlap": away_overlap,
                "chosen_side": side,
            }
        )
    return side


def _get_t2s_team_roster(
    t2s_id: int,
    side: str,
    hockey_db_dir: Path,
) -> Dict[str, str]:
    """
    Return a mapping of normalized jersey number -> player name for the given
    TimeToScore game id and team side ('home' or 'away').

    This is used to credit Games Played (GP) for players who appear on the
    official game roster even if they have no recorded shifts in the sheet.
    """
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        details = f": {_t2s_api_import_error}" if _t2s_api_import_error else ""
        raise RuntimeError(f"TimeToScore API not available (failed to import hmlib.time2score.api){details}")
    try:
        with _working_directory(hockey_db_dir):
            info = t2s_api.get_game_details(int(t2s_id))
    except Exception as e:
        raise RuntimeError(f"TimeToScore API failed to load game {t2s_id} for roster: {e}") from e

    stats = (info or {}).get("stats")
    if not isinstance(stats, dict) or not stats:
        raise RuntimeError(f"TimeToScore returned no usable stats for game {t2s_id}; cannot load roster.")
    players_key = "homePlayers" if side == "home" else "awayPlayers"
    rows = stats.get(players_key) or []

    roster: Dict[str, str] = {}

    for r in rows:
        try:
            raw_num = str((r or {}).get("number")).strip()
        except Exception:
            continue
        if not raw_num:
            continue
        num_norm = _normalize_jersey_number(raw_num)
        if not num_norm:
            continue
        name = str((r or {}).get("name") or "").strip()
        if not name:
            continue
        # One name per jersey; later entries with the same jersey overwrite.
        roster[num_norm] = name

    return roster


def _write_event_summaries_and_clips(
    outdir: Path,
    stats_dir: Path,
    event_log_context: EventLogContext,
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    create_scripts: bool,
    *,
    focus_team: Optional[str] = None,
) -> None:
    raw_evt_by_team = event_log_context.event_counts_by_type_team or {}
    raw_instances = event_log_context.event_instances or {}
    invert_event_type: Set[str] = set()
    if focus_team in {"Blue", "White"}:
        invert_event_type.add("Giveaway")
        turnover_teams = {
            str(team)
            for (etype, team) in raw_evt_by_team.keys()
            if str(etype) == "TurnoverForced"
        }
        if not turnover_teams:
            turnover_teams = {
                str(team)
                for (etype, team) in raw_instances.keys()
                if str(etype) == "TurnoverForced"
            }
        turnover_teams = {t for t in turnover_teams if t in {"Blue", "White"}}
        if len(turnover_teams) > 1:
            invert_event_type.add("TurnoverForced")

    def _no_parens_label(s: str) -> str:
        return re.sub(r"[()]", "", str(s or "")).strip()

    def _team_label(team: Any, *, event_type: Optional[str] = None) -> str:
        team_str = str(team) if team is not None else ""
        if focus_team in {"Blue", "White"} and team_str in {"Blue", "White"}:
            # Some event types are recorded for the team that loses possession, but
            # "For/Against" is intended to be relative to the focus team.
            et = str(event_type or "")
            invert = et in invert_event_type
            if invert:
                return "Against" if team_str == focus_team else "For"
            return "For" if team_str == focus_team else "Against"
        return team_str or "Unknown"

    evt_by_team: Dict[Tuple[str, str], int] = {}
    for (et, tm), cnt in sorted(raw_evt_by_team.items()):
        label = _team_label(tm, event_type=str(et))
        try:
            inc = int(cnt)
        except Exception:
            inc = 0
        evt_by_team[(et, label)] = evt_by_team.get((et, label), 0) + inc

    rows_evt = [
        {"event_type": _display_event_type(et), "team": tm, "count": cnt}
        for (et, tm), cnt in sorted(evt_by_team.items())
    ]
    if rows_evt:
        df_evt = pd.DataFrame(rows_evt)
        df_evt.to_csv(stats_dir / "event_summary.csv", index=False)
        _write_styled_xlsx_table(
            stats_dir / "event_summary.xlsx",
            df_evt,
            sheet_name="event_summary",
            title="Event Summary",
        )

    player_event_rows = event_log_context.event_player_rows or []
    if player_event_rows:

        def _fmt_v(x):
            return (
                seconds_to_hhmmss(int(x))
                if isinstance(x, int)
                else (seconds_to_hhmmss(int(x)) if isinstance(x, float) else "")
            )

        def _fmt_g(x):
            return (
                seconds_to_mmss_or_hhmmss(int(x))
                if isinstance(x, int)
                else (seconds_to_mmss_or_hhmmss(int(x)) if isinstance(x, float) else "")
            )

        rows = []
        for r in player_event_rows:
            rows.append(
                {
                    "event_type": _display_event_type(str(r.get("event_type") or "")),
                    "player": r.get("player"),
                    "jersey": r.get("jersey"),
                    "period": r.get("period"),
                    "video_time": _fmt_v(r.get("video_s")),
                    "game_time": _fmt_g(r.get("game_s")),
                    "team": _team_label(r.get("team"), event_type=str(r.get("event_type") or "")),
                }
            )
        df_players = pd.DataFrame(rows)
        df_players.to_csv(stats_dir / "event_players.csv", index=False)
        _write_styled_xlsx_table(
            stats_dir / "event_players.xlsx",
            df_players,
            sheet_name="event_players",
            title="Event Players",
        )

    # When --no-scripts is set, we still keep event summary CSVs but skip all
    # clip-related timestamp files and helper scripts.
    if not create_scripts:
        return

    instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for (etype, team), lst in raw_instances.items():
        label = _team_label(team, event_type=str(etype))
        instances.setdefault((etype, label), []).extend(list(lst or []))

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    max_sb_by_period: Dict[int, int] = {}
    for period, segs in conv_segments_by_period.items():
        mx = 0
        for s1, s2, _, _ in segs:
            mx = max(mx, s1, s2)
        if mx > 0:
            max_sb_by_period[period] = mx
    for _, lst in instances.items():
        for it in lst:
            p = it.get("period")
            gs = it.get("game_s")
            if isinstance(p, int) and isinstance(gs, (int, float)):
                v = int(gs)
                max_sb_by_period[p] = max(max_sb_by_period.get(p, 0), v)

    def merge_windows(win: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        if not win:
            return []
        win = sorted(win)
        out = [list(win[0])]
        for a, b in win[1:]:
            la, lb = out[-1]
            if a <= lb + 10:
                out[-1][1] = max(lb, b)
            else:
                out.append([a, b])
        return [(a, b) for a, b in out]

    def merge_windows_with_centers(
        win: List[Tuple[int, int, int]],
    ) -> List[Tuple[int, int, List[int]]]:
        """
        Merge overlapping/nearby (<=10s gap) video windows, preserving the set of
        event-center timestamps inside each merged window.
        """
        if not win:
            return []
        win_sorted = sorted(win, key=lambda x: (x[0], x[1], x[2]))
        out: List[Tuple[int, int, List[int]]] = [(win_sorted[0][0], win_sorted[0][1], [win_sorted[0][2]])]
        for a, b, c in win_sorted[1:]:
            la, lb, centers = out[-1]
            if a <= lb + 10:
                out[-1] = (la, max(lb, b), centers + [c])
            else:
                out.append((a, b, [c]))

        merged: List[Tuple[int, int, List[int]]] = []
        for a, b, centers in out:
            seen: set[int] = set()
            centers_sorted: List[int] = []
            for x in sorted(int(v) for v in centers):
                if x in seen:
                    continue
                seen.add(x)
                centers_sorted.append(x)
            merged.append((int(a), int(b), centers_sorted))
        return merged

    clip_scripts: List[str] = []
    for (etype, team), lst in sorted(instances.items()):
        # Skip team-level assist clip scripts; assists are handled in per-player highlights.
        if str(etype) == "Assist":
            continue

        v_windows: List[Tuple[int, int, int]] = []
        sb_windows_by_period: Dict[int, List[Tuple[int, int]]] = {}
        pre_s, post_s = _clip_pre_post_s_for_event_type(str(etype))
        etype_disp = _display_event_type(str(etype))
        etype_safe = sanitize_name(etype_disp) or "Event"
        team_safe = sanitize_name(team) or "Team"
        for it in lst:
            p = it.get("period")
            v = it.get("video_s")
            g = it.get("game_s")
            vsec = None
            if isinstance(v, (int, float)):
                vsec = int(v)
            elif isinstance(g, (int, float)) and isinstance(p, int):
                vsec = map_sb_to_video(int(p), int(g))
            if vsec is not None:
                start = max(0, vsec - int(pre_s))
                end = vsec + int(post_s)
                v_windows.append((start, end, int(vsec)))
            if isinstance(g, (int, float)) and isinstance(p, int):
                gsec = int(g)
                sb_max = max_sb_by_period.get(int(p), None)
                sb_start = gsec + int(pre_s)
                if sb_max is not None:
                    sb_start = min(sb_max, sb_start)
                sb_end = max(0, gsec - int(post_s))
                lo, hi = (sb_end, sb_start) if sb_end <= sb_start else (sb_start, sb_end)
                sb_windows_by_period.setdefault(int(p), []).append((lo, hi))

        v_windows_merged = merge_windows_with_centers(v_windows)
        if v_windows_merged:
            vfile = outdir / f"events_{etype_safe}_{team_safe}_video_times.txt"
            v_lines = []
            for a, b, centers in v_windows_merged:
                tokens = [seconds_to_hhmmss(a), seconds_to_hhmmss(b)]
                tokens.extend(seconds_to_hhmmss(int(c)) for c in (centers or []))
                v_lines.append(" ".join(tokens))
            vfile.write_text("\n".join(v_lines) + "\n", encoding="utf-8")
            if create_scripts:
                script = outdir / f"clip_events_{etype_safe}_{team_safe}.sh"
                label = f"{_no_parens_label(etype_disp)} {_no_parens_label(team)}"
                blink_label = _blink_event_label(str(etype))
                body = f"""#!/usr/bin/env bash
		set -euo pipefail
		if [ $# -lt 2 ]; then
		  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
	THIS_DIR=\"$(cd \"$(dirname \"${{BASH_SOURCE[0]}}\")\" && pwd)\"
	TS_FILE=\"$THIS_DIR/{vfile.name}\"
	shift 2 || true
	python -m hmlib.cli.video_clipper -j 4 --blink-event-text --blink-event-label \"{blink_label}\" --input \"$INPUT\" --timestamps \"$TS_FILE\" --temp-dir \"$THIS_DIR/temp_clips/{etype_safe}_{team_safe}\" \"{label} vs $OPP\" \"$@\"
	"""
                script.write_text(body, encoding="utf-8")
                try:
                    import os as _os

                    _os.chmod(script, 0o755)
                except Exception:
                    pass
                clip_scripts.append(script.name)

        if sb_windows_by_period:
            sfile = outdir / f"events_{etype_safe}_{team_safe}_scoreboard_times.txt"
            s_lines = []
            for p, wins in sorted(sb_windows_by_period.items()):
                wins = merge_windows(wins)
                for lo, hi in wins:
                    s_lines.append(
                        f"{p} {seconds_to_mmss_or_hhmmss(hi)} {seconds_to_mmss_or_hhmmss(lo)}"
                    )
            if s_lines:
                sfile.write_text("\n".join(s_lines) + "\n", encoding="utf-8")

    if clip_scripts and create_scripts:
        all_script = outdir / "clip_events_all.sh"
        all_body = """#!/usr/bin/env bash
set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
shift 2 || true
THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"
for s in {scripts}; do
  echo \"Running $s...\"
  \"$THIS_DIR/$s\" \"$INPUT\" \"$OPP\" \"$@\"
done
""".replace(
            "{scripts}", " ".join(sorted(clip_scripts))
        )
        all_script.write_text(all_body, encoding="utf-8")
        try:
            import os as _os

            _os.chmod(all_script, 0o755)
        except Exception:
            pass

    team_excluded = event_log_context.team_excluded or {}
    if any(v for v in team_excluded.values()):
        import sys as _sys

        for team, excl in team_excluded.items():
            if not excl:
                continue
            excl_str = ", ".join(str(x) for x in excl[:20])
            print(
                f"[warning] Team {team} exceeded 20 unique jerseys; additional jerseys seen: {excl_str} (data kept)",
                file=_sys.stderr,
            )


def _write_player_event_highlights(
    outdir: Path,
    event_log_context: EventLogContext,
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    player_keys: Iterable[str],
    create_scripts: bool,
    *,
    highlight_types: Tuple[str, ...] = ("Goal", "SOG"),
) -> None:
    """
    Generate per-player highlight timestamp files + helper scripts for selected event types.

    Intended for '-long' sheets where events include video time. Falls back to
    mapping scoreboard->video using conv_segments when needed.
    """
    if not create_scripts:
        return
    if not event_log_context or not (event_log_context.event_player_rows or []):
        return

    player_set = set(player_keys or [])
    if not player_set:
        return

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    def merge_windows(win: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        if not win:
            return []
        win_sorted = sorted(win)
        out: List[Tuple[int, int]] = [win_sorted[0]]
        for a, b in win_sorted[1:]:
            la, lb = out[-1]
            if a <= lb:
                out[-1] = (la, max(lb, b))
            else:
                out.append((a, b))
        return out

    def merge_windows_with_centers(
        win: List[Tuple[int, int, int]],
    ) -> List[Tuple[int, int, List[int]]]:
        if not win:
            return []
        win_sorted = sorted(win, key=lambda x: (x[0], x[1], x[2]))
        out: List[Tuple[int, int, List[int]]] = [(win_sorted[0][0], win_sorted[0][1], [win_sorted[0][2]])]
        for a, b, c in win_sorted[1:]:
            la, lb, centers = out[-1]
            if a <= lb:
                out[-1] = (la, max(lb, b), centers + [c])
            else:
                out.append((a, b, [c]))

        merged: List[Tuple[int, int, List[int]]] = []
        for a, b, centers in out:
            seen: set[int] = set()
            centers_sorted: List[int] = []
            for x in sorted(int(v) for v in centers):
                if x in seen:
                    continue
                seen.add(x)
                centers_sorted.append(x)
            merged.append((int(a), int(b), centers_sorted))
        return merged

    by_player_type: Dict[Tuple[str, str], List[Tuple[int, Optional[int], Optional[int]]]] = {}
    for row in event_log_context.event_player_rows or []:
        et = row.get("event_type")
        pk = row.get("player")
        if not isinstance(et, str) or et not in highlight_types:
            continue
        if not isinstance(pk, str) or pk not in player_set:
            continue
        p = row.get("period")
        if not isinstance(p, int):
            continue
        v = row.get("video_s")
        g = row.get("game_s")
        vsec = int(v) if isinstance(v, (int, float)) else None
        gsec = int(g) if isinstance(g, (int, float)) else None
        by_player_type.setdefault((pk, et), []).append((p, vsec, gsec))

    for (pk, etype), rows in sorted(by_player_type.items(), key=lambda x: (x[0][0], x[0][1])):
        v_windows: List[Tuple[int, int, int]] = []
        pre_s, post_s = _clip_pre_post_s_for_event_type(str(etype))
        for period, vsec, gsec in rows:
            vv = vsec
            if vv is None and gsec is not None:
                vv = map_sb_to_video(period, gsec)
            if vv is None:
                continue
            v_windows.append((max(0, int(vv) - int(pre_s)), int(vv) + int(post_s), int(vv)))

        v_windows_merged = merge_windows_with_centers(v_windows)
        if not v_windows_merged:
            continue

        vfile = outdir / f"events_{etype}_{pk}_video_times.txt"
        v_lines = []
        for a, b, centers in v_windows_merged:
            tokens = [seconds_to_hhmmss(a), seconds_to_hhmmss(b)]
            tokens.extend(seconds_to_hhmmss(int(c)) for c in (centers or []))
            v_lines.append(" ".join(tokens))
        vfile.write_text("\n".join(v_lines) + "\n", encoding="utf-8")

        if not create_scripts:
            continue

        label = f"{_display_event_type(etype)} - {_display_player_name(pk)}"
        blink_label = _blink_event_label(str(etype))
        script_name = f"clip_{etype.lower()}_{pk}.sh"
        script = outdir / script_name
        body = f"""#!/usr/bin/env bash
	set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
THIS_DIR=\"$(cd \"$(dirname \"${{BASH_SOURCE[0]}}\")\" && pwd)\"
TS_FILE=\"$THIS_DIR/{vfile.name}\"
shift 2 || true
python -m hmlib.cli.video_clipper -j 4 --blink-event-text --blink-event-label \"{blink_label}\" --input \"$INPUT\" --timestamps \"$TS_FILE\" --temp-dir \"$THIS_DIR/temp_clips/{etype}_{pk}\" \"{label} vs $OPP\" \"$@\"
"""
        script.write_text(body, encoding="utf-8")
        try:
            import os as _os

            _os.chmod(script, 0o755)
        except Exception:
            pass


def _write_player_combined_highlights(
    outdir: Path,
    *,
    event_log_context: Optional[EventLogContext],
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]],
    player_keys: Iterable[str],
    create_scripts: bool,
    highlight_types: Tuple[str, ...] = ("Goal", "Assist", "ExpectedGoal", "Takeaway"),
) -> None:
    """
    Generate a per-player "Highlights" timestamp file that mixes multiple event types
    in chronological order (within the game).

    This is used for season (multi-game) highlight reels so a player's events are not
    grouped by type (e.g., all goals then all assists).
    """
    if not create_scripts:
        return
    player_set = set(player_keys or [])
    if not player_set:
        return

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    def merge_windows(win: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        if not win:
            return []
        win_sorted = sorted(win)
        out: List[Tuple[int, int]] = [win_sorted[0]]
        for a, b in win_sorted[1:]:
            la, lb = out[-1]
            if a <= lb:
                out[-1] = (la, max(lb, b))
            else:
                out.append((a, b))
        return out

    highlight_types_set = set(highlight_types or ())

    # First, seed goal timestamps (period, game_s) from the scoring source so we can
    # dedupe xG rows that are just goals (goals count as xG internally).
    goal_keys_by_player: Dict[str, set[Tuple[int, int]]] = {}
    for pk, info in (per_player_goal_events or {}).items():
        if pk not in player_set:
            continue
        for ev in (info or {}).get("goals", []) or []:
            try:
                goal_keys_by_player.setdefault(pk, set()).add((int(ev.period), int(ev.t_sec)))
            except Exception:
                continue

    # Collect raw event centers (video seconds) per player.
    # We dedupe by (event_type, period, game_s) when possible.
    events_by_player: Dict[str, List[Tuple[int, str, int, Optional[int]]]] = {}
    seen_by_player: Dict[str, set[Tuple[str, int, int]]] = {}

    def _register_event(
        *,
        player: str,
        event_type: str,
        period: int,
        video_s: Optional[int],
        game_s: Optional[int],
    ) -> None:
        if player not in player_set:
            return
        et = str(event_type or "")
        if et not in highlight_types_set:
            return
        try:
            p = int(period)
        except Exception:
            return

        # Dedupe ExpectedGoal rows that correspond to goals for the same player/time.
        if et == "ExpectedGoal":
            if game_s is not None and (p, int(game_s)) in goal_keys_by_player.get(player, set()):
                return

        vsec = int(video_s) if isinstance(video_s, int) else (int(video_s) if isinstance(video_s, float) else None)
        gsec = int(game_s) if isinstance(game_s, int) else (int(game_s) if isinstance(game_s, float) else None)

        if vsec is None and gsec is not None:
            vsec = map_sb_to_video(p, int(gsec))
        if vsec is None:
            return

        # Use game time for dedupe when we have it; otherwise fall back to video time.
        key_time = gsec if gsec is not None else int(vsec)
        seen = seen_by_player.setdefault(player, set())
        if (et, p, int(key_time)) in seen:
            return
        seen.add((et, p, int(key_time)))
        events_by_player.setdefault(player, []).append((int(vsec), et, p, gsec))

        if et == "Goal" and gsec is not None:
            goal_keys_by_player.setdefault(player, set()).add((p, int(gsec)))

    # Prefer long-sheet / event-log times when available.
    if event_log_context is not None:
        for row in event_log_context.event_player_rows or []:
            try:
                pk = row.get("player")
                etype = row.get("event_type")
                period = row.get("period")
                if not isinstance(pk, str) or not isinstance(etype, str) or not isinstance(period, int):
                    continue
                _register_event(
                    player=pk,
                    event_type=etype,
                    period=period,
                    video_s=row.get("video_s"),
                    game_s=row.get("game_s"),
                )
            except Exception:
                continue

    # Fallback for goals/assists when a game has no long sheet.
    for pk, info in (per_player_goal_events or {}).items():
        if pk not in player_set:
            continue
        for ev in (info or {}).get("goals", []) or []:
            _register_event(
                player=pk,
                event_type="Goal",
                period=int(getattr(ev, "period", 0) or 0),
                video_s=None,
                game_s=int(getattr(ev, "t_sec", 0) or 0),
            )
        for ev in (info or {}).get("assists", []) or []:
            _register_event(
                player=pk,
                event_type="Assist",
                period=int(getattr(ev, "period", 0) or 0),
                video_s=None,
                game_s=int(getattr(ev, "t_sec", 0) or 0),
            )

    for pk, evs in sorted(events_by_player.items(), key=lambda x: x[0]):
        if not evs:
            continue
        # Chronological within this game.
        evs_sorted = sorted(evs, key=lambda x: x[0])
        windows: List[Tuple[int, int]] = []
        for vsec, etype, _period, _gsec in evs_sorted:
            pre_s, post_s = _clip_pre_post_s_for_event_type(str(etype))
            windows.append((max(0, int(vsec) - int(pre_s)), int(vsec) + int(post_s)))

        windows = merge_windows(windows)
        if not windows:
            continue

        vfile = outdir / f"events_Highlights_{pk}_video_times.txt"
        v_lines = [f"{seconds_to_hhmmss(a)} {seconds_to_hhmmss(b)}" for a, b in windows]
        vfile.write_text("\n".join(v_lines) + "\n", encoding="utf-8")


def _write_goal_window_files(
    outdir: Path,
    goals: List[GoalEvent],
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
) -> None:
    if not goals:
        return

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    max_sb_by_period: Dict[int, int] = {}
    for period, segs in conv_segments_by_period.items():
        mx = 0
        for s1, s2, _, _ in segs:
            mx = max(mx, s1, s2)
        max_sb_by_period[period] = mx

    gf_lines: List[str] = []
    ga_lines: List[str] = []
    for ev in goals:
        sb_max = max_sb_by_period.get(ev.period, None)
        start_sb = ev.t_sec + GOAL_CLIP_PRE_S
        end_sb = ev.t_sec - GOAL_CLIP_POST_S
        if sb_max is not None:
            start_sb = max(0, min(sb_max, start_sb))
            end_sb = max(0, min(sb_max, end_sb))
        else:
            start_sb = max(0, start_sb)
            end_sb = max(0, end_sb)

        v_center = map_sb_to_video(ev.period, ev.t_sec)
        if v_center is not None:
            v_start = max(0, v_center - GOAL_CLIP_PRE_S)
            v_end = v_center + GOAL_CLIP_POST_S
            start_str = seconds_to_hhmmss(v_start)
            end_str = seconds_to_hhmmss(v_end)
        else:
            v_start = map_sb_to_video(ev.period, start_sb)
            v_end = map_sb_to_video(ev.period, end_sb)
            if v_start is not None and v_end is not None:
                lo, hi = (int(v_start), int(v_end)) if v_start <= v_end else (int(v_end), int(v_start))
                start_str = seconds_to_hhmmss(max(0, lo))
                end_str = seconds_to_hhmmss(max(0, hi))
            else:
                start_str = seconds_to_hhmmss(max(0, start_sb))
                end_str = seconds_to_hhmmss(max(0, end_sb))

        line = f"{start_str} {end_str}"
        if ev.kind == "GF":
            gf_lines.append(line)
        else:
            ga_lines.append(line)

    (outdir / "goals_for.txt").write_text(
        "\n".join(gf_lines) + ("\n" if gf_lines else ""), encoding="utf-8"
    )
    (outdir / "goals_against.txt").write_text(
        "\n".join(ga_lines) + ("\n" if ga_lines else ""), encoding="utf-8"
    )


def _write_clip_all_runner(outdir: Path, create_scripts: bool) -> None:
    if not create_scripts:
        return
    clip_all_path = outdir / "clip_all.sh"
    clip_all_body = """#!/usr/bin/env bash
set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
shift 2 || true
THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"
for s in \"$THIS_DIR\"/clip_*.sh; do
  [ -x \"$s\" ] || continue
  base=\"$(basename \"$s\")\"
  if [ \"$base\" = \"clip_all.sh\" ] || [ \"$base\" = \"clip_events_all.sh\" ]; then
    continue
  fi
  echo \"Running $s...\"
  \"$s\" \"$INPUT\" \"$OPP\" \"$@\"
done
"""
    clip_all_path.write_text(clip_all_body, encoding="utf-8")
    try:
        import os as _os

        _os.chmod(clip_all_path, 0o755)
    except Exception:
        pass


def process_sheet(
    xls_path: Path,
    sheet_name: Optional[str],
    outdir: Path,
    keep_goalies: bool,
    goals: List[GoalEvent],
    roster_map: Optional[Dict[str, str]] = None,
    long_xls_paths: Optional[List[Path]] = None,
    focus_team_override: Optional[str] = None,
    include_shifts_in_stats: bool = False,
    skip_validation: bool = False,
    create_scripts: bool = True,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
    List[Dict[str, Any]],
]:
    target_sheet = 0 if sheet_name is None else sheet_name
    df = pd.read_excel(xls_path, sheet_name=target_sheet, header=None)
    outdir.mkdir(parents=True, exist_ok=True)

    # Try event-log layout first
    (
        used_event_log,
        video_pairs_by_player,
        sb_pairs_by_player,
        conv_segments_by_period,
        event_log_context,
    ) = _parse_event_log_layout(df)

    validation_errors = 0
    if not used_event_log:
        (
            video_pairs_by_player,
            sb_pairs_by_player,
            conv_segments_by_period,
            validation_errors,
        ) = _parse_per_player_layout(df, keep_goalies=keep_goalies, skip_validation=skip_validation)

    # Output subdir depends on format
    format_dir = "event_log" if used_event_log else "per_player"
    outdir = outdir / format_dir
    outdir.mkdir(parents=True, exist_ok=True)
    stats_dir = outdir / "stats"
    stats_dir.mkdir(parents=True, exist_ok=True)

    # Per-player time files and clip scripts
    if include_shifts_in_stats:
        _write_video_times_and_scripts(outdir, video_pairs_by_player, create_scripts=create_scripts)
        _write_scoreboard_times(outdir, sb_pairs_by_player, create_scripts=create_scripts)

    stats_table_rows: List[Dict[str, str]] = []
    all_periods_seen: set[int] = set()

    jersey_to_players: Dict[str, List[str]] = {}
    for pk in sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(pk)
        if norm:
            jersey_to_players.setdefault(norm, []).append(pk)

    # Optionally add TimeToScore roster-only players (no shifts in this sheet)
    roster_only_players: List[str] = []
    if roster_map:
        # Jersey numbers already present in this sheet (normalized).
        seen_normals: set[str] = set()
        for pk in sb_pairs_by_player.keys():
            jersey_part = pk.split("_", 1)[0]
            norm = _normalize_jersey_number(jersey_part)
            if norm:
                seen_normals.add(norm)

        for jersey_norm, name in roster_map.items():
            if not jersey_norm:
                continue
            if jersey_norm in seen_normals:
                # Already have at least one player with this jersey from shifts.
                continue
            jersey_label = jersey_norm
            player_key = f"{sanitize_name(jersey_label)}_{sanitize_name(name)}"
            if player_key in sb_pairs_by_player:
                continue
            # No recorded shifts; keep an explicit empty list.
            sb_pairs_by_player[player_key] = []
            roster_only_players.append(player_key)
            # Map normalized jersey to this player for scoring lookups.
            jersey_to_players.setdefault(jersey_norm, []).append(player_key)

    # Optional '-long' event sheet: merge event context and optionally infer goals.
    focus_team: Optional[str] = focus_team_override
    merged_event_context: Optional[EventLogContext] = event_log_context
    inferred_long_goals: List[GoalEvent] = []

    long_goal_rows_all: List[Dict[str, Any]] = []
    jerseys_by_team_all: Dict[str, set[int]] = {}

    if long_xls_paths:
        for long_path in long_xls_paths:
            if long_path is None:
                continue
            lp = Path(long_path).expanduser()
            if not lp.exists():
                continue
            try:
                long_df = pd.read_excel(lp, sheet_name=0, header=None)
            except Exception as e:  # noqa: BLE001
                print(f"[long] Failed to read {lp}: {e}", file=sys.stderr)
                continue

            long_events, long_goal_rows, jerseys_by_team = _parse_long_left_event_table(long_df)
            if not long_events and not long_goal_rows:
                continue

            for team, nums in (jerseys_by_team or {}).items():
                jerseys_by_team_all.setdefault(team, set()).update(nums or set())
            long_goal_rows_all.extend(long_goal_rows or [])

            if focus_team is None:
                our_jerseys: set[str] = set(jersey_to_players.keys())
                if roster_map:
                    our_jerseys |= set(roster_map.keys())
                focus_team = _infer_focus_team_from_long_sheet(our_jerseys, jerseys_by_team_all)

            long_ctx = _event_log_context_from_long_events(
                long_events,
                jersey_to_players=jersey_to_players,
                focus_team=focus_team,
                jerseys_by_team=jerseys_by_team,
            )
            merged_event_context = _merge_event_log_contexts(merged_event_context, long_ctx)

        if focus_team is None and (long_goal_rows_all or jerseys_by_team_all):
            print(
                "[long] Could not infer whether your team is Blue or White; "
                "use --dark/--light to enable team-relative stats and goal inference.",
                file=sys.stderr,
            )

        if (not goals) and focus_team is not None and long_goal_rows_all:
            for row in long_goal_rows_all:
                team = row.get("team")
                period = row.get("period")
                gsec = row.get("game_s")
                scorer = row.get("scorer")
                assists = row.get("assists") or []
                if team not in {"Blue", "White"}:
                    continue
                if not isinstance(period, int):
                    continue
                if not isinstance(gsec, (int, float)):
                    continue
                kind = "GF" if team == focus_team else "GA"
                t_str = seconds_to_mmss_or_hhmmss(int(gsec))
                scorer_str = str(int(scorer)) if scorer is not None else None
                assists_str: List[str] = []
                for a in assists:
                    if a is None:
                        continue
                    try:
                        assists_str.append(str(int(a)))
                    except Exception:
                        continue
                inferred_long_goals.append(
                    GoalEvent(kind, int(period), t_str, scorer=scorer_str, assists=assists_str)
                )
            inferred_long_goals.sort(key=lambda e: (e.period, e.t_sec))

    # Use inferred long-sheet goals only as a fallback when no goals were provided.
    if not goals and inferred_long_goals:
        goals = inferred_long_goals

    # Stats and plus/minus
    goals_by_period: Dict[int, List[GoalEvent]] = {}
    for ev in goals:
        goals_by_period.setdefault(ev.period, []).append(ev)

    # Annotate game-tying / game-winning roles for goals in this game.
    _annotate_goal_roles(goals)

    pair_on_ice_rows: List[Dict[str, Any]] = []
    pair_on_ice_by_player: Dict[str, List[Dict[str, Any]]] = {}
    if sb_pairs_by_player:
        try:
            pair_on_ice_rows = _compute_pair_on_ice_rows(sb_pairs_by_player, goals_by_period)
            for r in pair_on_ice_rows:
                pk = str(r.get("player") or "")
                if pk:
                    pair_on_ice_by_player.setdefault(pk, []).append(r)
            for pk, lst in pair_on_ice_by_player.items():
                lst.sort(key=lambda x: float(x.get("overlap_pct", 0.0) or 0.0), reverse=True)
            _write_pair_on_ice_csv(stats_dir, pair_on_ice_rows, include_toi=include_shifts_in_stats)
        except Exception:
            pair_on_ice_rows = []
            pair_on_ice_by_player = {}

    # Per-player event details for this game (for multi-game summaries).
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]] = {
        pk: {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []}
        for pk in sb_pairs_by_player.keys()
    }

    goal_assist_counts: Dict[str, Dict[str, int]] = {
        pk: {"goals": 0, "assists": 0} for pk in sb_pairs_by_player.keys()
    }

    def _match_player_keys(num_token: Any) -> List[str]:
        matches: List[str] = []
        candidates: set[str] = set()
        if num_token is not None:
            try:
                txt = str(num_token).strip()
                if txt:
                    candidates.add(txt)
            except Exception:
                pass
        norm = _normalize_jersey_number(num_token)
        if norm:
            candidates.add(norm)
        for cand in candidates:
            matches.extend(jersey_to_players.get(cand, []))
        return list(dict.fromkeys(matches))  # dedupe while preserving order

    for ev in goals:
        if ev.kind != "GF":
            continue
        if ev.scorer:
            for pk in _match_player_keys(ev.scorer):
                goal_assist_counts[pk]["goals"] += 1
                per_player_goal_events[pk]["goals"].append(ev)
        for ast in ev.assists:
            for pk in _match_player_keys(ast):
                goal_assist_counts[pk]["assists"] += 1
                per_player_goal_events[pk]["assists"].append(ev)

    event_log_context = merged_event_context

    # Determine which long-sheet/player-attributed event types exist in this game.
    # This is used to avoid treating missing stats as 0 when a game lacks a `*-long*` sheet
    # or when a newer stat type wasn't collected for older games.
    player_event_types_present: set[str] = set()
    if event_log_context is not None:
        for counts in (event_log_context.event_counts_by_player or {}).values():
            for et in (counts or {}).keys():
                if et:
                    player_event_types_present.add(str(et))

    has_player_shots = "Shot" in player_event_types_present
    has_player_sog = "SOG" in player_event_types_present
    has_player_expected_goals = "ExpectedGoal" in player_event_types_present
    has_player_turnovers_forced = "TurnoverForced" in player_event_types_present
    has_player_created_turnovers = "CreatedTurnover" in player_event_types_present
    has_player_giveaways = "Giveaway" in player_event_types_present
    has_player_takeaways = "Takeaway" in player_event_types_present

    # Team-level event availability for on-ice for/against metrics.
    has_controlled_entry_events = False
    has_controlled_exit_events = False
    if focus_team in {"Blue", "White"} and event_log_context is not None:
        for (etype, _team) in (event_log_context.event_instances or {}).keys():
            if etype == "ControlledEntry":
                has_controlled_entry_events = True
            elif etype == "ControlledExit":
                has_controlled_exit_events = True

    if include_shifts_in_stats:
        _write_all_events_summary(
            stats_dir,
            sb_pairs_by_player=sb_pairs_by_player,
            goals=goals,
            goals_by_period=goals_by_period,
            event_log_context=event_log_context,
            focus_team=focus_team,
            conv_segments_by_period=conv_segments_by_period,
        )

    # Pre-group team-level events by period for on-ice for/against counts.
    on_ice_event_types = {"ControlledEntry", "ControlledExit"}
    team_events_by_period: Dict[int, List[Tuple[str, str, int]]] = {}
    if focus_team is not None and event_log_context is not None:
        for (etype, team), inst_list in (event_log_context.event_instances or {}).items():
            if etype not in on_ice_event_types:
                continue
            for it in inst_list or []:
                p = it.get("period")
                gs = it.get("game_s")
                if not isinstance(p, int) or not isinstance(gs, (int, float)):
                    continue
                team_events_by_period.setdefault(int(p), []).append((etype, str(team), int(gs)))

    for player_key, sb_list in sb_pairs_by_player.items():
        sb_by_period: Dict[int, List[Tuple[str, str]]] = {}
        for period, a, b in sb_list:
            sb_by_period.setdefault(period, []).append((a, b))
        for period in sb_by_period.keys():
            all_periods_seen.add(period)
        all_pairs = [(a, b) for (_, a, b) in sb_list]
        if include_shifts_in_stats:
            shift_summary = summarize_shift_lengths_sec(all_pairs)
            per_period_toi_map = per_period_toi(sb_by_period)
        else:
            shift_summary = {}
            per_period_toi_map = {}

        plus_minus = 0
        counted_gf: List[str] = []
        counted_ga: List[str] = []
        counted_gf_by_period: Dict[int, int] = {}
        counted_ga_by_period: Dict[int, int] = {}
        for period, pairs in sb_by_period.items():
            if period not in goals_by_period:
                continue
            for ev in goals_by_period[period]:
                matched = False
                for a, b in pairs:
                    a_sec = parse_flex_time_to_seconds(a)
                    b_sec = parse_flex_time_to_seconds(b)
                    lo, hi = (a_sec, b_sec) if a_sec <= b_sec else (b_sec, a_sec)
                    if not (lo <= ev.t_sec <= hi):
                        continue
                    if ev.kind == "GA" and ev.t_sec == a_sec:
                        continue
                    elif ev.kind == "GF" and ev.t_sec == a_sec:
                        continue
                    matched = True
                    break
                if matched:
                    if ev.kind == "GF":
                        plus_minus += 1
                        counted_gf.append(f"P{period}:{ev.t_str}")
                        counted_gf_by_period[period] = counted_gf_by_period.get(period, 0) + 1
                        per_player_goal_events[player_key]["gf_on_ice"].append(ev)
                    else:
                        plus_minus -= 1
                        counted_ga.append(f"P{period}:{ev.t_str}")
                        counted_ga_by_period[period] = counted_ga_by_period.get(period, 0) + 1
                        per_player_goal_events[player_key]["ga_on_ice"].append(ev)

        scoring_counts = goal_assist_counts.get(player_key, {"goals": 0, "assists": 0})
        points_val = scoring_counts.get("goals", 0) + scoring_counts.get("assists", 0)
        # Per-game points-per-game (PPG) for this single game.
        ppg_val = float(points_val)
        # Overtime scoring (OT period is period 4; OT2 -> 5, etc.).
        ot_goals_cnt = 0
        ot_assists_cnt = 0
        try:
            ev_map = per_player_goal_events.get(player_key, {}) or {}
            ot_goals_cnt = sum(
                1 for ev in (ev_map.get("goals") or []) if int(getattr(ev, "period", 0) or 0) >= 4
            )
            ot_assists_cnt = sum(
                1
                for ev in (ev_map.get("assists") or [])
                if int(getattr(ev, "period", 0) or 0) >= 4
            )
        except Exception:
            ot_goals_cnt = 0
            ot_assists_cnt = 0

        # On-ice for/against metrics for team-level events (e.g., controlled exits)
        on_ice: Dict[str, int] = {
            "controlled_entry_for": 0,
            "controlled_entry_against": 0,
            "controlled_exit_for": 0,
            "controlled_exit_against": 0,
        }
        if focus_team is not None and team_events_by_period and sb_by_period:
            for period, pairs in sb_by_period.items():
                period_events = team_events_by_period.get(period, [])
                if not period_events or not pairs:
                    continue
                intervals = [compute_interval_seconds(a, b) for a, b in pairs]
                for etype, team, gsec in period_events:
                    in_any = False
                    for lo, hi in intervals:
                        if lo <= gsec <= hi:
                            in_any = True
                            break
                    if not in_any:
                        continue
                    is_for = team == focus_team
                    if etype == "ControlledEntry":
                        key = "controlled_entry_for" if is_for else "controlled_entry_against"
                        on_ice[key] += 1
                    elif etype == "ControlledExit":
                        key = "controlled_exit_for" if is_for else "controlled_exit_against"
                        on_ice[key] += 1

        stats_lines = []
        stats_lines.append(f"Player: {_display_player_name(player_key)}")
        stats_lines.append("Games Played (GP): 1")
        stats_lines.append(f"Goals: {scoring_counts.get('goals', 0)}")
        stats_lines.append(f"Assists: {scoring_counts.get('assists', 0)}")
        stats_lines.append(f"Points (G+A): {points_val}")
        stats_lines.append(f"PPG (points per game): {ppg_val:.1f}")
        if include_shifts_in_stats:
            shifts_cnt = 0
            try:
                shifts_cnt = int(str(shift_summary.get("num_shifts", "0") or 0))
            except Exception:
                shifts_cnt = 0
            stats_lines.append(f"Shifts (scoreboard): {shift_summary.get('num_shifts', '0')}")
            stats_lines.append(f"TOI total (scoreboard): {shift_summary.get('toi_total', '0:00')}")
            stats_lines.append(f"Avg shift: {shift_summary.get('toi_avg', '0:00')}")
            stats_lines.append(f"Median shift: {shift_summary.get('toi_median', '0:00')}")
            stats_lines.append(f"Longest shift: {shift_summary.get('toi_longest', '0:00')}")
            stats_lines.append(f"Shortest shift: {shift_summary.get('toi_shortest', '0:00')}")
            if shifts_cnt > 0:
                stats_lines.append(f"Goals per shift: {_fmt_per_shift(scoring_counts.get('goals', 0), shifts_cnt)}")
                stats_lines.append(
                    f"Assists per shift: {_fmt_per_shift(scoring_counts.get('assists', 0), shifts_cnt)}"
                )
                stats_lines.append(f"Points per shift: {_fmt_per_shift(points_val, shifts_cnt)}")
            if per_period_toi_map:
                stats_lines.append("Per-period TOI (scoreboard):")
                for period in sorted(per_period_toi_map.keys()):
                    stats_lines.append(f"  Period {period}: {per_period_toi_map[period]}")
        stats_lines.append(f"Goal +/-: {plus_minus}")
        if counted_gf:
            stats_lines.append("  GF counted at: " + ", ".join(sorted(counted_gf)))
        if counted_ga:
            stats_lines.append("  GA counted at: " + ", ".join(sorted(counted_ga)))

        if player_key in pair_on_ice_by_player:
            stats_lines.append("")
            stats_lines.append("On-ice with teammates (by TOI%):")
            disp_rows: List[Tuple[str, float, int, int, int, int, int]] = []
            for r in pair_on_ice_by_player.get(player_key, []):
                teammate = str(r.get("teammate") or "")
                if not teammate:
                    continue
                disp_rows.append(
                    (
                        _display_player_name(teammate),
                        float(r.get("overlap_pct", 0.0) or 0.0),
                        int(r.get("shift_games", 0) or 0),
                        int(r.get("plus_minus_together", 0) or 0),
                        int(r.get("gf_together", 0) or 0),
                        int(r.get("ga_together", 0) or 0),
                        int(r.get("overlap_seconds", 0) or 0),
                    )
                )
            name_w = max([len(r[0]) for r in disp_rows] + [len("Teammate")])
            if include_shifts_in_stats:
                stats_lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA  Overlap")
                for teammate_disp, pct, shift_games, pm, gf, ga, overlap_s in disp_rows:
                    stats_lines.append(
                        f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}  {_format_duration(overlap_s):>7}"
                    )
            else:
                # Without --shifts, do not publish any absolute time values.
                stats_lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA")
                for teammate_disp, pct, shift_games, pm, gf, ga, _overlap_s in disp_rows:
                    stats_lines.append(
                        f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}"
                    )

        if event_log_context is not None:
            per_player_events = event_log_context.event_counts_by_player
            ev_counts = per_player_events.get(player_key, {})
            if ev_counts:
                stats_lines.append("")
                stats_lines.append("Event Counts:")
                order = [
                    "Shot",
                    "SOG",
                    "Goal",
                    "Assist",
                    "ExpectedGoal",
                    "TurnoverForced",
                    "CreatedTurnover",
                    "Giveaway",
                    "Takeaway",
                    "ControlledEntry",
                    "ControlledExit",
                ]
                sog_cnt = int(ev_counts.get("SOG", 0) or 0)
                xg_cnt = int(ev_counts.get("ExpectedGoal", 0) or 0)
                for kind in order:
                    if kind == "ExpectedGoal":
                        # If the player has at least one SOG, always show xG even when it's 0
                        # (so parents can see the full SOG/xG line without missing fields).
                        if sog_cnt > 0 or xg_cnt > 0:
                            stats_lines.append(f"  {_display_event_type(kind)}: {xg_cnt}")
                        continue
                    if kind in ev_counts and ev_counts[kind] > 0:
                        stats_lines.append(f"  {_display_event_type(kind)}: {ev_counts[kind]}")
                for kind, cnt in sorted(ev_counts.items()):
                    if kind in order:
                        continue
                    stats_lines.append(f"  {_display_event_type(kind)}: {cnt}")

        if focus_team is not None and any(v > 0 for v in on_ice.values()):
            stats_lines.append("")
            stats_lines.append("On-ice team events (for/against):")
            stats_lines.append(
                f"  ControlledEntry: {on_ice['controlled_entry_for']} for, {on_ice['controlled_entry_against']} against"
            )
            stats_lines.append(
                f"  ControlledExit: {on_ice['controlled_exit_for']} for, {on_ice['controlled_exit_against']} against"
            )

        if include_shifts_in_stats:
            for period, pairs in sorted(sb_by_period.items()):
                stats_lines.append(f"Shifts in Period {period}: {len(pairs)}")

        (stats_dir / f"{player_key}_stats.txt").write_text(
            "\n".join(stats_lines) + "\n", encoding="utf-8"
        )

        row_map: Dict[str, str] = {
            "player": player_key,
            "goals": str(scoring_counts.get("goals", 0)),
            "assists": str(scoring_counts.get("assists", 0)),
            "ot_goals": str(ot_goals_cnt),
            "ot_assists": str(ot_assists_cnt),
            "points": str(points_val),
            "gp": "1",
            "ppg": f"{ppg_val:.1f}",
            "plus_minus": str(plus_minus),
            "plus_minus_per_game": str(plus_minus),
        }
        if include_shifts_in_stats:
            shifts_cnt_row = 0
            try:
                shifts_cnt_row = int(str(shift_summary.get("num_shifts", "0") or 0))
            except Exception:
                shifts_cnt_row = 0
            row_map["shifts"] = str(shifts_cnt_row)
            row_map["shifts_per_game"] = str(shifts_cnt_row)
            row_map["sb_toi_total"] = str(shift_summary.get("toi_total", "0:00"))
            row_map["sb_toi_per_game"] = str(shift_summary.get("toi_total", "0:00"))
            row_map["sb_avg"] = str(shift_summary.get("toi_avg", "0:00"))
            row_map["sb_median"] = str(shift_summary.get("toi_median", "0:00"))
            row_map["sb_longest"] = str(shift_summary.get("toi_longest", "0:00"))
            row_map["sb_shortest"] = str(shift_summary.get("toi_shortest", "0:00"))
            row_map["goals_per_shift"] = _fmt_per_shift(scoring_counts.get("goals", 0), shifts_cnt_row)
            row_map["assists_per_shift"] = _fmt_per_shift(scoring_counts.get("assists", 0), shifts_cnt_row)
            row_map["points_per_shift"] = _fmt_per_shift(points_val, shifts_cnt_row)
        # Event counts (from event logs / long sheets), per game.
        if event_log_context is not None:
            ev_counts = (event_log_context.event_counts_by_player or {}).get(player_key, {})
        else:
            ev_counts = {}
        shots_cnt = int(ev_counts.get("Shot", 0) or 0)
        sog_cnt = int(ev_counts.get("SOG", 0) or 0)
        expected_goals_cnt = int(ev_counts.get("ExpectedGoal", 0) or 0)
        turnovers_forced_cnt = int(ev_counts.get("TurnoverForced", 0) or 0)
        created_turnovers_cnt = int(ev_counts.get("CreatedTurnover", 0) or 0)
        giveaways_cnt = int(ev_counts.get("Giveaway", 0) or 0)
        takeaways_cnt = int(ev_counts.get("Takeaway", 0) or 0)

        if has_player_shots:
            row_map["shots"] = str(shots_cnt)
            row_map["shots_per_game"] = str(shots_cnt)
            if include_shifts_in_stats:
                row_map["shots_per_shift"] = _fmt_per_shift(shots_cnt, shifts_cnt_row)
        else:
            row_map["shots"] = ""
            row_map["shots_per_game"] = ""
            if include_shifts_in_stats:
                row_map["shots_per_shift"] = ""

        if has_player_sog:
            row_map["sog"] = str(sog_cnt)
            row_map["sog_per_game"] = str(sog_cnt)
            if include_shifts_in_stats:
                row_map["sog_per_shift"] = _fmt_per_shift(sog_cnt, shifts_cnt_row)
        else:
            row_map["sog"] = ""
            row_map["sog_per_game"] = ""
            if include_shifts_in_stats:
                row_map["sog_per_shift"] = ""

        if has_player_expected_goals:
            row_map["expected_goals"] = str(expected_goals_cnt)
            row_map["expected_goals_per_game"] = str(expected_goals_cnt)
            if include_shifts_in_stats:
                row_map["expected_goals_per_shift"] = _fmt_per_shift(expected_goals_cnt, shifts_cnt_row)
        else:
            row_map["expected_goals"] = ""
            row_map["expected_goals_per_game"] = ""
            if include_shifts_in_stats:
                row_map["expected_goals_per_shift"] = ""

        if has_player_expected_goals and has_player_sog:
            row_map["expected_goals_per_sog"] = (
                f"{(expected_goals_cnt / sog_cnt):.2f}" if sog_cnt > 0 else ""
            )
        else:
            row_map["expected_goals_per_sog"] = ""

        if has_player_turnovers_forced:
            row_map["turnovers_forced"] = str(turnovers_forced_cnt)
            row_map["turnovers_forced_per_game"] = str(turnovers_forced_cnt)
            if include_shifts_in_stats:
                row_map["turnovers_forced_per_shift"] = _fmt_per_shift(turnovers_forced_cnt, shifts_cnt_row)
        else:
            row_map["turnovers_forced"] = ""
            row_map["turnovers_forced_per_game"] = ""
            if include_shifts_in_stats:
                row_map["turnovers_forced_per_shift"] = ""

        if has_player_created_turnovers:
            row_map["created_turnovers"] = str(created_turnovers_cnt)
            row_map["created_turnovers_per_game"] = str(created_turnovers_cnt)
            if include_shifts_in_stats:
                row_map["created_turnovers_per_shift"] = _fmt_per_shift(created_turnovers_cnt, shifts_cnt_row)
        else:
            row_map["created_turnovers"] = ""
            row_map["created_turnovers_per_game"] = ""
            if include_shifts_in_stats:
                row_map["created_turnovers_per_shift"] = ""

        if has_player_giveaways:
            row_map["giveaways"] = str(giveaways_cnt)
            row_map["giveaways_per_game"] = str(giveaways_cnt)
            if include_shifts_in_stats:
                row_map["giveaways_per_shift"] = _fmt_per_shift(giveaways_cnt, shifts_cnt_row)
        else:
            row_map["giveaways"] = ""
            row_map["giveaways_per_game"] = ""
            if include_shifts_in_stats:
                row_map["giveaways_per_shift"] = ""

        if has_player_takeaways:
            row_map["takeaways"] = str(takeaways_cnt)
            row_map["takeaways_per_game"] = str(takeaways_cnt)
            if include_shifts_in_stats:
                row_map["takeaways_per_shift"] = _fmt_per_shift(takeaways_cnt, shifts_cnt_row)
        else:
            row_map["takeaways"] = ""
            row_map["takeaways_per_game"] = ""
            if include_shifts_in_stats:
                row_map["takeaways_per_shift"] = ""

        if has_controlled_entry_events:
            row_map["controlled_entry_for"] = str(on_ice["controlled_entry_for"])
            row_map["controlled_entry_for_per_game"] = str(on_ice["controlled_entry_for"])
            row_map["controlled_entry_against"] = str(on_ice["controlled_entry_against"])
            row_map["controlled_entry_against_per_game"] = str(on_ice["controlled_entry_against"])
            if include_shifts_in_stats:
                row_map["controlled_entry_for_per_shift"] = _fmt_per_shift(
                    on_ice["controlled_entry_for"], shifts_cnt_row
                )
                row_map["controlled_entry_against_per_shift"] = _fmt_per_shift(
                    on_ice["controlled_entry_against"], shifts_cnt_row
                )
        else:
            row_map["controlled_entry_for"] = ""
            row_map["controlled_entry_for_per_game"] = ""
            row_map["controlled_entry_against"] = ""
            row_map["controlled_entry_against_per_game"] = ""
            if include_shifts_in_stats:
                row_map["controlled_entry_for_per_shift"] = ""
                row_map["controlled_entry_against_per_shift"] = ""

        if has_controlled_exit_events:
            row_map["controlled_exit_for"] = str(on_ice["controlled_exit_for"])
            row_map["controlled_exit_for_per_game"] = str(on_ice["controlled_exit_for"])
            row_map["controlled_exit_against"] = str(on_ice["controlled_exit_against"])
            row_map["controlled_exit_against_per_game"] = str(on_ice["controlled_exit_against"])
            if include_shifts_in_stats:
                row_map["controlled_exit_for_per_shift"] = _fmt_per_shift(on_ice["controlled_exit_for"], shifts_cnt_row)
                row_map["controlled_exit_against_per_shift"] = _fmt_per_shift(
                    on_ice["controlled_exit_against"], shifts_cnt_row
                )
        else:
            row_map["controlled_exit_for"] = ""
            row_map["controlled_exit_for_per_game"] = ""
            row_map["controlled_exit_against"] = ""
            row_map["controlled_exit_against_per_game"] = ""
            if include_shifts_in_stats:
                row_map["controlled_exit_for_per_shift"] = ""
                row_map["controlled_exit_against_per_shift"] = ""

        row_map["gf_counted"] = str(len(counted_gf))
        row_map["gf_per_game"] = str(len(counted_gf))
        row_map["ga_counted"] = str(len(counted_ga))
        row_map["ga_per_game"] = str(len(counted_ga))
        if include_shifts_in_stats:
            row_map["gf_per_shift"] = _fmt_per_shift(len(counted_gf), shifts_cnt_row)
            row_map["ga_per_shift"] = _fmt_per_shift(len(counted_ga), shifts_cnt_row)
        if include_shifts_in_stats:
            v_pairs = video_pairs_by_player.get(player_key, [])
            if v_pairs:
                v_sum = 0
                for a, b in v_pairs:
                    lo, hi = compute_interval_seconds(a, b)
                    v_sum += hi - lo
                row_map["video_toi_total"] = _format_duration(v_sum)
            else:
                row_map["video_toi_total"] = ""
            for period, toi in per_period_toi_map.items():
                row_map[f"P{period}_toi"] = toi
                all_periods_seen.add(period)
            for period, pairs in sb_by_period.items():
                row_map[f"P{period}_shifts"] = str(len(pairs))
        for period, cnt in counted_gf_by_period.items():
            row_map[f"P{period}_GF"] = str(cnt)
            all_periods_seen.add(period)
        for period, cnt in counted_ga_by_period.items():
            row_map[f"P{period}_GA"] = str(cnt)
            all_periods_seen.add(period)

        stats_table_rows.append(row_map)

    # Global CSV (contains TOI); only write when explicitly enabled.
    if include_shifts_in_stats:
        _write_global_summary_csv(stats_dir, sb_pairs_by_player)

    period_list = sorted(all_periods_seen)
    # Consolidated player stats
    if stats_table_rows:
        _write_player_stats_text_and_csv(
            stats_dir,
            stats_table_rows,
            period_list,
            include_shifts_in_stats=include_shifts_in_stats,
        )

    # Per-game team stats (no TOI)
    _write_game_stats_files(
        stats_dir,
        xls_path=xls_path,
        periods=period_list,
        goals=goals,
        event_log_context=event_log_context,
        focus_team=focus_team,
    )

    # Goals windows
    _write_goal_window_files(outdir, goals, conv_segments_by_period)

    # Event summaries
    if event_log_context is not None:
        _write_event_summaries_and_clips(
            outdir,
            stats_dir,
            event_log_context,
            conv_segments_by_period,
            create_scripts=create_scripts,
            focus_team=focus_team,
        )
        _write_player_event_highlights(
            outdir,
            event_log_context,
            conv_segments_by_period,
            sb_pairs_by_player.keys(),
            create_scripts=create_scripts,
        )
        _write_player_combined_highlights(
            outdir,
            event_log_context=event_log_context,
            conv_segments_by_period=conv_segments_by_period,
            per_player_goal_events=per_player_goal_events,
            player_keys=sb_pairs_by_player.keys(),
            create_scripts=create_scripts,
        )
    else:
        # Still write combined per-player highlights for goals/assists when available.
        _write_player_combined_highlights(
            outdir,
            event_log_context=None,
            conv_segments_by_period=conv_segments_by_period,
            per_player_goal_events=per_player_goal_events,
            player_keys=sb_pairs_by_player.keys(),
            create_scripts=create_scripts,
        )

    # Aggregate clip runner (optional scripts)
    _write_clip_all_runner(outdir, create_scripts=create_scripts)

    # Validation summary
    if (not used_event_log) and (not skip_validation) and validation_errors > 0:
        print(
            f"[validation] Completed with {validation_errors} issue(s). See messages above.",
            file=sys.stderr,
        )

    return outdir, stats_table_rows, sorted(all_periods_seen), per_player_goal_events, pair_on_ice_rows


def process_t2s_only_game(
    *,
    t2s_id: int,
    side: str,
    outdir: Path,
    label: str,
    hockey_db_dir: Path,
    include_shifts_in_stats: bool,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
]:
    """
    Process a game using only TimeToScore data (no shift spreadsheets).

    Writes the same `per_player/stats/*` outputs as `process_sheet`, but leaves
    shift/TOI and on-ice (+/-) fields blank since they cannot be derived without
    shift timing data.
    """
    outdir.mkdir(parents=True, exist_ok=True)
    outdir = outdir / "per_player"
    outdir.mkdir(parents=True, exist_ok=True)
    stats_dir = outdir / "stats"
    stats_dir.mkdir(parents=True, exist_ok=True)

    # Load roster + scoring from TimeToScore.
    with _working_directory(hockey_db_dir):
        goals = goals_from_t2s(int(t2s_id), side=side)
    roster_map = _get_t2s_team_roster(int(t2s_id), side, hockey_db_dir)

    # Annotate game-tying / game-winning roles.
    _annotate_goal_roles(goals)

    periods_seen = sorted(
        {
            int(getattr(ev, "period", 0) or 0)
            for ev in goals
            if isinstance(getattr(ev, "period", None), int) and int(getattr(ev, "period", 0) or 0) > 0
        }
    )

    # Build internal player keys from T2S roster (normalized jersey -> name).
    player_keys: List[str] = []
    jersey_to_player: Dict[str, str] = {}
    for jersey_norm, name in sorted((roster_map or {}).items(), key=lambda x: x[0]):
        if not jersey_norm or not name:
            continue
        player_key = f"{sanitize_name(jersey_norm)}_{sanitize_name(name)}"
        player_keys.append(player_key)
        jersey_to_player[jersey_norm] = player_key

    def _sort_player_key(pk: str) -> Tuple[int, str]:
        try:
            jersey_part = pk.split("_", 1)[0]
            norm = _normalize_jersey_number(jersey_part) or ""
            return int(norm), pk
        except Exception:
            return 9999, pk

    player_keys.sort(key=_sort_player_key)

    # Per-player event details for this game (for multi-game summaries).
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]] = {
        pk: {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []} for pk in player_keys
    }
    goal_assist_counts: Dict[str, Dict[str, int]] = {pk: {"goals": 0, "assists": 0} for pk in player_keys}

    def _match_player_keys(num_token: Any) -> List[str]:
        norm = _normalize_jersey_number(num_token)
        if not norm:
            return []
        pk = jersey_to_player.get(norm)
        return [pk] if pk else []

    for ev in goals:
        if ev.kind != "GF":
            continue
        if ev.scorer:
            for pk in _match_player_keys(ev.scorer):
                goal_assist_counts[pk]["goals"] += 1
                per_player_goal_events[pk]["goals"].append(ev)
        for ast in ev.assists:
            for pk in _match_player_keys(ast):
                goal_assist_counts[pk]["assists"] += 1
                per_player_goal_events[pk]["assists"].append(ev)

    stats_table_rows: List[Dict[str, str]] = []
    for player_key in player_keys:
        scoring_counts = goal_assist_counts.get(player_key, {"goals": 0, "assists": 0})
        goals_cnt = int(scoring_counts.get("goals", 0) or 0)
        assists_cnt = int(scoring_counts.get("assists", 0) or 0)
        points_val = goals_cnt + assists_cnt
        ppg_val = float(points_val)

        ev_map = per_player_goal_events.get(player_key, {}) or {}
        ot_goals_cnt = sum(
            1 for ev in (ev_map.get("goals") or []) if int(getattr(ev, "period", 0) or 0) >= 4
        )
        ot_assists_cnt = sum(
            1 for ev in (ev_map.get("assists") or []) if int(getattr(ev, "period", 0) or 0) >= 4
        )

        # Individual per-player stats file (best-effort, without shift-derived fields).
        try:
            lines: List[str] = []
            lines.append(f"Player: {_display_player_name(player_key)}")
            lines.append("")
            lines.append("TimeToScore-only game (no shift spreadsheet)")
            lines.append("")
            lines.append(f"Goals: {goals_cnt}")
            lines.append(f"Assists: {assists_cnt}")
            lines.append(f"OT Goals: {ot_goals_cnt}")
            lines.append(f"OT Assists: {ot_assists_cnt}")
            lines.append(f"Points (G+A): {points_val}")
            lines.append(f"PPG (points per game): {ppg_val:.1f}")
            goals_list = ev_map.get("goals") or []
            assists_list = ev_map.get("assists") or []
            if goals_list:
                lines.append("")
                lines.append("Goals:")
                for ev in sorted(goals_list, key=lambda e: (e.period, e.t_sec)):
                    tags: List[str] = []
                    if getattr(ev, "is_game_tying", False):
                        tags.append("GT")
                    if getattr(ev, "is_game_winning", False):
                        tags.append("GW")
                    tag_str = f" [{' '.join(tags)}]" if tags else ""
                    lines.append(f"  Period {ev.period}, {ev.t_str}{tag_str}")
            if assists_list:
                lines.append("")
                lines.append("Assists:")
                for ev in sorted(assists_list, key=lambda e: (e.period, e.t_sec)):
                    tags2: List[str] = []
                    if getattr(ev, "is_game_tying", False):
                        tags2.append("GT")
                    if getattr(ev, "is_game_winning", False):
                        tags2.append("GW")
                    tag_str2 = f" [{' '.join(tags2)}]" if tags2 else ""
                    lines.append(f"  Period {ev.period}, {ev.t_str}{tag_str2}")
            (stats_dir / f"{player_key}_stats.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
        except Exception:
            pass

        row_map: Dict[str, str] = {
            "player": player_key,
            "goals": str(goals_cnt),
            "assists": str(assists_cnt),
            "ot_goals": str(ot_goals_cnt),
            "ot_assists": str(ot_assists_cnt),
            "points": str(points_val),
            "gp": "1",
            "ppg": f"{ppg_val:.1f}",
            # Long-sheet-derived stats are not available via T2S; leave blank.
            "shots": "",
            "shots_per_game": "",
            "sog": "",
            "sog_per_game": "",
            "expected_goals": "",
            "expected_goals_per_game": "",
            "expected_goals_per_sog": "",
            "giveaways": "",
            "giveaways_per_game": "",
            "takeaways": "",
            "takeaways_per_game": "",
            "controlled_entry_for": "",
            "controlled_entry_for_per_game": "",
            "controlled_entry_against": "",
            "controlled_entry_against_per_game": "",
            "controlled_exit_for": "",
            "controlled_exit_for_per_game": "",
            "controlled_exit_against": "",
            "controlled_exit_against_per_game": "",
            # Shift-derived metrics are not available; leave blank.
            "plus_minus": "",
            "plus_minus_per_game": "",
            "gf_counted": "",
            "gf_per_game": "",
            "ga_counted": "",
            "ga_per_game": "",
        }
        if include_shifts_in_stats:
            row_map.update(
                {
                    "shifts": "",
                    "shifts_per_game": "",
                    "sb_toi_total": "",
                    "sb_toi_per_game": "",
                    "sb_avg": "",
                    "sb_median": "",
                    "sb_longest": "",
                    "sb_shortest": "",
                    "video_toi_total": "",
                }
            )
        stats_table_rows.append(row_map)

    if stats_table_rows:
        _write_player_stats_text_and_csv(
            stats_dir,
            stats_table_rows,
            periods_seen,
            include_shifts_in_stats=include_shifts_in_stats,
        )

    _write_game_stats_files_t2s_only(
        stats_dir,
        label=label,
        t2s_id=int(t2s_id),
        goals=goals,
        periods=periods_seen,
    )

    return outdir, stats_table_rows, periods_seen, per_player_goal_events


# ----------------------------- CLI -----------------------------


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Extract per-player shifts & stats from an Excel sheet like 'dh-tv-12-1.xls'."
    )
    p.add_argument(
        "--input",
        "-i",
        dest="inputs",
        action="append",
        type=str,
        default=[],
        help="Path to input .xls/.xlsx file, or a directory containing one primary sheet plus optional '*-long*' companion sheets. "
        "Can repeat for multiple games. "
        "Append ':HOME' or ':AWAY' to override side for that file.",
    )
    p.add_argument(
        "--file-list",
        type=Path,
        default=None,
        help="Path to a text file containing one .xls/.xlsx path or directory per line (comments/# allowed). "
        "Directories are expanded to the primary sheet plus optional '*-long*' companion sheets. "
        "You can append ':HOME' or ':AWAY' per line. "
        "Lines may also be 't2s=<game_id>[:HOME|AWAY][:game_label]' to process a TimeToScore-only game with no spreadsheets.",
    )
    p.add_argument(
        "--sheet", "-s", type=str, default=None, help="Worksheet name (default: first sheet)."
    )
    p.add_argument(
        "--outdir", "-o", type=Path, default=Path("player_focus"), help="Output directory."
    )
    p.add_argument(
        "--keep-goalies",
        action="store_true",
        help="By default, rows like '(G) 37' are skipped. Use this flag to include them.",
    )
    # Goals
    p.add_argument(
        "--goal",
        "-g",
        action="append",
        default=[],
        help="Goal event token. Example: 'GF:2/13:45' or 'GA:1/05:12'. Can repeat.",
    )
    p.add_argument(
        "--goals-file",
        type=Path,
        default=None,
        help="Path to a text file with one goal per line (GF:period/time or GA:period/time). '#' lines ignored.",
    )
    # TimeToScore integration
    p.add_argument(
        "--t2s",
        type=str,
        default=None,
        help=(
            "TimeToScore spec: '<game_id>[:HOME|AWAY][:game_label]'. "
            "If set and no --goal/--goals-file provided, fetch goals from T2S. "
            "If provided without --input/--file-list, runs a TimeToScore-only game (no spreadsheets)."
        ),
    )
    p.add_argument(
        "--hockey-db-dir",
        type=Path,
        default=Path.home() / ".cache" / "hockeymom",
        help="Directory for the hockey_league.db used when fetching TimeToScore goals (default: ~/.cache/hockeymom).",
    )
    side_group = p.add_mutually_exclusive_group()
    side_group.add_argument(
        "--home",
        action="store_true",
        help="Your team is the home team (with --t2s).",
    )
    side_group.add_argument(
        "--away",
        action="store_true",
        help="Your team is the away team (with --t2s).",
    )
    team_color_group = p.add_mutually_exclusive_group()
    team_color_group.add_argument(
        "--light",
        action="store_true",
        help="For '*-long*' sheets: treat the White team as your team when mapping events to players.",
    )
    team_color_group.add_argument(
        "--dark",
        action="store_true",
        help="For '*-long*' sheets: treat the Blue team as your team when mapping events to players.",
    )
    p.add_argument(
        "--shifts",
        action="store_true",
        help=(
            "Include shift/TOI metrics in parent-facing stats outputs "
            "(stats/*.txt, stats/player_stats.* and consolidated workbook). "
            "Also enables writing per-player shift clip scripts and `*_video_times.txt` / `*_scoreboard_times.txt` files. "
            "By default these are omitted."
        ),
    )
    p.add_argument(
        "--no-scripts",
        action="store_true",
        help="Do not generate clip helper scripts or any '*_times.txt' timestamp files used for clipping.",
    )
    p.add_argument(
        "--skip-validation",
        action="store_true",
        help="Skip validation checks on start/end ordering and excessive durations.",
    )
    return p


def main() -> None:
    args = build_arg_parser().parse_args()
    hockey_db_dir = args.hockey_db_dir.expanduser()
    create_scripts = not args.no_scripts
    include_shifts_in_stats = bool(getattr(args, "shifts", False))
    focus_team_override: Optional[str] = None
    if getattr(args, "light", False):
        focus_team_override = "White"
    elif getattr(args, "dark", False):
        focus_team_override = "Blue"

    t2s_arg_id: Optional[int] = None
    t2s_arg_side: Optional[str] = None
    t2s_arg_label: Optional[str] = None
    if args.t2s is not None:
        parsed = _parse_t2s_spec(args.t2s)
        if parsed is None:
            print(
                f"Error: invalid --t2s spec '{args.t2s}' (expected '<id>[:HOME|AWAY][:label]').",
                file=sys.stderr,
            )
            sys.exit(2)
        t2s_arg_id, t2s_arg_side, t2s_arg_label = parsed

    input_entries: List[InputEntry] = []
    if args.file_list:
        try:
            base_dir = args.file_list.expanduser().resolve().parent
            with args.file_list.open("r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    # Strip UTF-8 BOM if present (common when files are created on Windows).
                    line = line.lstrip("\ufeff")
                    if not line or line.startswith("#"):
                        continue
                    t2s_only = _parse_t2s_only_token(line)
                    if t2s_only is not None:
                        t2s_id, side, label = t2s_only
                        input_entries.append(
                            InputEntry(path=None, side=side, t2s_id=t2s_id, label=label)
                        )
                        continue
                    p, side = _parse_input_token(line, base_dir=base_dir)
                    input_entries.append(InputEntry(path=p, side=side))
        except Exception as e:
            print(f"Error reading --file-list: {e}", file=sys.stderr)
            sys.exit(2)
    for tok in args.inputs or []:
        t2s_only = _parse_t2s_only_token(tok)
        if t2s_only is not None:
            t2s_id, side, label = t2s_only
            input_entries.append(InputEntry(path=None, side=side, t2s_id=t2s_id, label=label))
            continue
        p, side = _parse_input_token(tok)
        input_entries.append(InputEntry(path=p, side=side))

    if not input_entries:
        # Allow a TimeToScore-only run by specifying just `--t2s`.
        if t2s_arg_id is not None:
            input_entries.append(
                InputEntry(path=None, side=t2s_arg_side, t2s_id=int(t2s_arg_id), label=t2s_arg_label)
            )
        else:
            print("Error: at least one --input/--file-list entry or --t2s is required.", file=sys.stderr)
            sys.exit(2)

    # Support passing a directory to --input for single-game runs: discover the
    # primary shift sheet plus optional '*-long*' companion sheet(s) from that
    # directory (or `<dir>/stats`).
    expanded_entries: List[InputEntry] = []
    for entry in input_entries:
        if entry.path is None:
            expanded_entries.append(entry)
            continue
        pp = Path(entry.path).expanduser()
        if pp.is_dir():
            try:
                discovered = _expand_dir_input_to_game_sheets(pp)
            except Exception as e:  # noqa: BLE001
                print(f"Error expanding directory input {pp}: {e}", file=sys.stderr)
                sys.exit(2)
            for fp in discovered:
                expanded_entries.append(InputEntry(path=fp, side=entry.side))
        else:
            expanded_entries.append(InputEntry(path=pp, side=entry.side))
    input_entries = expanded_entries

    base_outdir = args.outdir.expanduser()
    # Group '-long' companion sheets with their non-long counterpart so a game is processed once.
    groups_by_label: Dict[str, Dict[str, Any]] = {}
    display_label_to_key: Dict[str, str] = {}
    for order_idx, entry in enumerate(input_entries):
        side = entry.side
        if entry.t2s_id is not None and entry.path is None:
            group_key = f"t2s-{int(entry.t2s_id)}"
            display_label = str(entry.label or group_key)
            existing_key = display_label_to_key.get(display_label)
            if existing_key is not None and existing_key != group_key:
                print(
                    f"Error: duplicate game label '{display_label}' (conflicts between {existing_key} and {group_key}).",
                    file=sys.stderr,
                )
                sys.exit(2)
            display_label_to_key.setdefault(display_label, group_key)
            g = groups_by_label.setdefault(
                group_key,
                {
                    "label": display_label,
                    "primary": None,
                    "long_paths": [],
                    "side": None,
                    "order": order_idx,
                    "t2s_id_only": int(entry.t2s_id),
                },
            )
        else:
            if entry.path is None:
                continue
            p = Path(entry.path)
            label = _base_label_from_path(p)
            existing_key = display_label_to_key.get(label)
            if existing_key is not None and existing_key != label:
                print(
                    f"Error: duplicate game label '{label}' (conflicts between {existing_key} and {label}).",
                    file=sys.stderr,
                )
                sys.exit(2)
            display_label_to_key.setdefault(label, label)
            g = groups_by_label.setdefault(
                label,
                {"label": label, "primary": None, "long_paths": [], "side": None, "order": order_idx},
            )
        if g.get("side") is None:
            g["side"] = side
        elif side and g.get("side") != side:
            print(
                f"Error: conflicting side overrides for {g.get('label')}: {g.get('side')} vs {side}",
                file=sys.stderr,
            )
            sys.exit(2)

        if entry.path is None:
            continue
        if _is_long_sheet_path(p):
            g["long_paths"].append(p)
        else:
            if g.get("primary") is None:
                g["primary"] = p
            else:
                # Prefer the first primary; keep extras as additional long inputs.
                g["long_paths"].append(p)

    groups = sorted(groups_by_label.values(), key=lambda x: int(x.get("order", 0)))
    multiple_inputs = len(groups) > 1
    results: List[Dict[str, Any]] = []

    if t2s_arg_id is not None and len(groups) > 1:
        print(
            "Error: --t2s can only be used with a single game. "
            "For multi-game runs, use filename suffixes (e.g. 'game-51602.xlsx') "
            "or `t2s=<id>[:HOME|AWAY][:label]` lines in --file-list.",
            file=sys.stderr,
        )
        sys.exit(2)

    def _resolve_goals_for_file(
        in_path: Path, t2s_id: Optional[int], side: Optional[str]
    ) -> List[GoalEvent]:
        g = load_goals(args.goal, args.goals_file)
        if g:
            return g
        # If no t2s id was provided/inferred, fall back to goals.xlsx next to the input.
        if t2s_id is None:
            goals_xlsx = in_path.parent / "goals.xlsx"
            try:
                gx = _goals_from_goals_xlsx(goals_xlsx)
            except Exception as e:  # noqa: BLE001
                print(f"[goals.xlsx] Failed to parse {goals_xlsx}: {e}", file=sys.stderr)
                gx = []
            if gx:
                for gg in reversed(sorted([str(x) for x in gx])):
                    print(f"[goals.xlsx:{in_path.name}] {gg}")
                return gx
            return g
        # With a t2s id, require a side and use TimeToScore data.
        if side is None:
            print(
                "Error: T2S game id "
                f"{t2s_id} provided but side could not be determined while processing "
                f"'{_base_label_from_path(in_path)}' ({in_path}). "
                "Provide --home/--away (single game) or ':HOME' / ':AWAY' in --file-list.",
                file=sys.stderr,
            )
            sys.exit(2)
        try:
            with _working_directory(hockey_db_dir):
                g = goals_from_t2s(int(t2s_id), side=side)
        except Exception as e:  # noqa: BLE001
            print(
                f"Error: failed to fetch goals from TimeToScore for game {t2s_id} while processing "
                f"'{_base_label_from_path(in_path)}' ({in_path}).",
                file=sys.stderr,
            )
            if args.file_list:
                print(f"  File list: {args.file_list}", file=sys.stderr)
            print(f"  Side: {side}", file=sys.stderr)
            print(f"  Cause: {e}", file=sys.stderr)
            sys.exit(2)

        for gg in reversed(sorted([str(x) for x in g])):
            print(f"[t2s:{t2s_id}] {gg}")
        return g

    for idx, g in enumerate(groups):
        t2s_only_id = g.get("t2s_id_only")
        if t2s_only_id is not None:
            label = str(g.get("label") or f"t2s-{int(t2s_only_id)}")
            outdir = base_outdir if not multiple_inputs else base_outdir / label
            side_override: Optional[str] = g.get("side") or (
                "home" if args.home else ("away" if args.away else None)
            )
            if side_override is None:
                print(
                    f"Error: T2S-only game {t2s_only_id} requires a side (provide --home/--away or ':HOME'/'AWAY' on the t2s=... line).",
                    file=sys.stderr,
                )
                sys.exit(2)
            try:
                final_outdir, stats_rows, periods, per_player_events = process_t2s_only_game(
                    t2s_id=int(t2s_only_id),
                    side=str(side_override),
                    outdir=outdir,
                    label=label,
                    hockey_db_dir=hockey_db_dir,
                    include_shifts_in_stats=include_shifts_in_stats,
                )
            except Exception as e:  # noqa: BLE001
                print(
                    f"Error: failed to process TimeToScore-only game {t2s_only_id} ('{label}').",
                    file=sys.stderr,
                )
                if args.file_list:
                    print(f"  File list: {args.file_list}", file=sys.stderr)
                print(f"  Side: {side_override}", file=sys.stderr)
                print(f"  Cause: {e}", file=sys.stderr)
                sys.exit(2)
            results.append(
                {
                    "label": label,
                    "t2s_id": int(t2s_only_id),
                    "order": idx,
                    "outdir": final_outdir,
                    "stats": stats_rows,
                    "periods": periods,
                    "events": per_player_events,
                    "pair_on_ice": [],
                    "sheet_path": None,
                    "video_path": None,
                }
            )
            try:
                print(f"✅ Done. Wrote per-player files to: {final_outdir.resolve()}")
            except Exception:
                print("✅ Done.")
            continue

        in_path = g.get("primary")
        if in_path is None:
            long_paths = g.get("long_paths") or []
            if long_paths:
                in_path = long_paths[0]
            else:
                continue
        in_path = Path(in_path)
        path_side = g.get("side")
        long_paths: List[Path] = [Path(p) for p in (g.get("long_paths") or []) if Path(p) != in_path]

        # Prefer an explicit --t2s value; otherwise, infer a T2S id from the
        # filename only when the trailing numeric suffix is large enough
        # (>= 10000). Smaller suffixes (e.g., 'chicago-1') remain part of the
        # game name and do not trigger T2S usage.
        t2s_id = t2s_arg_id if t2s_arg_id is not None else _infer_t2s_from_filename(in_path)
        label = str(g.get("label") or _base_label_from_path(in_path))
        outdir = base_outdir if not multiple_inputs else base_outdir / label
        manual_goals = load_goals(args.goal, args.goals_file)

        jersey_numbers = set()
        if t2s_id is not None and not manual_goals:
            try:
                jersey_numbers = _collect_sheet_jerseys(in_path, args.sheet, args.keep_goalies)
            except Exception as e:
                print(f"Error parsing sheet for jersey numbers ({in_path}): {e}", file=sys.stderr)

        side_override: Optional[str] = path_side or t2s_arg_side or (
            "home" if args.home else ("away" if args.away else None)
        )
        inferred_side: Optional[str] = None
        side_infer_debug: Dict[str, Any] = {}
        if manual_goals:
            side_to_use = side_override
        else:
            if side_override:
                side_to_use = side_override
            elif t2s_id is not None:
                inferred_side = _infer_side_from_rosters(
                    int(t2s_id), jersey_numbers, hockey_db_dir, debug=side_infer_debug
                )
                side_to_use = inferred_side
            else:
                side_to_use = None

        # If a TimeToScore id is in use and we don't have a side override, we
        # must determine whether our team is Home or Away to map GF/GA.
        if t2s_id is not None and (not manual_goals) and side_to_use is None:
            t2s_source = (
                f"--t2s {args.t2s}"
                if t2s_arg_id is not None
                else f"inferred from filename suffix: {in_path.name}"
            )
            cli_side = "--home" if args.home else ("--away" if args.away else "<none>")
            def _safe_int(token: Any) -> int:
                try:
                    return int(str(token))
                except Exception:
                    return 0

            jerseys_sorted = (
                sorted(list(jersey_numbers or set()), key=_safe_int) if jersey_numbers else []
            )
            sample = ", ".join(jerseys_sorted[:20])
            extra = f", ... (+{len(jerseys_sorted) - 20} more)" if len(jerseys_sorted) > 20 else ""

            print(
                f"Error: cannot determine HOME/AWAY side for TimeToScore game {t2s_id} while processing '{label}'.",
                file=sys.stderr,
            )
            print(f"  Input sheet: {in_path}", file=sys.stderr)
            if args.file_list:
                print(f"  File list: {args.file_list}", file=sys.stderr)
            print(f"  T2S id source: {t2s_source}", file=sys.stderr)
            print("  Side overrides checked:", file=sys.stderr)
            print(f"    - from --file-list ':HOME' / ':AWAY': {path_side or '<none>'}", file=sys.stderr)
            print(f"    - from --t2s spec side: {t2s_arg_side or '<none>'}", file=sys.stderr)
            print(f"    - from CLI flags: {cli_side}", file=sys.stderr)
            print(
                f"  Jerseys found in sheet: {len(jerseys_sorted)}"
                + (f" ({sample}{extra})" if jerseys_sorted else ""),
                file=sys.stderr,
            )
            if side_infer_debug:
                failure = side_infer_debug.get("failure")
                if failure:
                    print(f"  Side inference result: {failure}", file=sys.stderr)
                if side_infer_debug.get("t2s_api_available") is False:
                    import_err = side_infer_debug.get("t2s_api_import_error")
                    details = f": {import_err}" if import_err else ""
                    print(
                        f"  TimeToScore API: not available (failed to import hmlib.time2score.api){details}.",
                        file=sys.stderr,
                    )
                exc = side_infer_debug.get("exception")
                if exc:
                    print(f"  TimeToScore fetch error: {exc}", file=sys.stderr)
                hr = side_infer_debug.get("home_roster_count")
                ar = side_infer_debug.get("away_roster_count")
                if isinstance(hr, int) or isinstance(ar, int):
                    print(
                        f"  TimeToScore roster sizes: home={hr or 0}, away={ar or 0}",
                        file=sys.stderr,
                    )
                ho = side_infer_debug.get("home_overlap")
                ao = side_infer_debug.get("away_overlap")
                if isinstance(ho, int) or isinstance(ao, int):
                    print(
                        f"  Overlap with sheet jerseys: home={ho or 0}, away={ao or 0}",
                        file=sys.stderr,
                    )
                hoj = side_infer_debug.get("home_overlap_jerseys")
                aoj = side_infer_debug.get("away_overlap_jerseys")
                if isinstance(hoj, list) or isinstance(aoj, list):
                    try:
                        h_txt = ", ".join(str(x) for x in (hoj or [])[:20])
                        a_txt = ", ".join(str(x) for x in (aoj or [])[:20])
                    except Exception:
                        h_txt = ""
                        a_txt = ""
                    if h_txt or a_txt:
                        print(
                            f"  Overlap jersey numbers: home=[{h_txt}], away=[{a_txt}]",
                            file=sys.stderr,
                        )
            print(
                "Fix: add ':HOME' or ':AWAY' to this game in --file-list (or run a single game with --home/--away).",
                file=sys.stderr,
            )
            sys.exit(2)

        goals = _resolve_goals_for_file(in_path, t2s_id, side_to_use)

        # Build a TimeToScore roster map (normalized jersey -> name) for GP
        # accounting, so that players listed on the official roster are
        # credited with a game played even if they have no recorded shifts.
        roster_map: Optional[Dict[str, str]] = None
        if t2s_id is not None and side_to_use is not None:
            try:
                roster_map = _get_t2s_team_roster(int(t2s_id), side_to_use, hockey_db_dir)
            except Exception as e:  # noqa: BLE001
                print(
                    f"Error: failed to fetch roster from TimeToScore for game {t2s_id} while processing '{label}'.",
                    file=sys.stderr,
                )
                print(f"  Input sheet: {in_path}", file=sys.stderr)
                if args.file_list:
                    print(f"  File list: {args.file_list}", file=sys.stderr)
                print(f"  Side: {side_to_use}", file=sys.stderr)
                print(f"  Cause: {e}", file=sys.stderr)
                sys.exit(2)

        final_outdir, stats_rows, periods, per_player_events, pair_on_ice_rows = process_sheet(
            xls_path=in_path,
            sheet_name=args.sheet,
            outdir=outdir,
            keep_goalies=args.keep_goalies,
            goals=goals,
            roster_map=roster_map,
            long_xls_paths=long_paths,
            focus_team_override=focus_team_override,
            include_shifts_in_stats=include_shifts_in_stats,
            skip_validation=args.skip_validation,
            create_scripts=create_scripts,
        )
        video_path = _find_tracking_output_video_for_sheet_path(in_path) if create_scripts else None
        results.append(
            {
                "label": label,
                "t2s_id": t2s_id,
                "order": idx,
                "outdir": final_outdir,
                "stats": stats_rows,
                "periods": periods,
                "events": per_player_events,
                "pair_on_ice": pair_on_ice_rows,
                "sheet_path": in_path,
                "video_path": video_path,
            }
        )
        try:
            print(f"✅ Done. Wrote per-player files to: {final_outdir.resolve()}")
        except Exception:
            print("✅ Done.")

    if multiple_inputs:
        agg_rows, agg_periods, per_game_denoms = _aggregate_stats_rows(
            [(r["stats"], r["periods"]) for r in results]
        )

        # Build per-player event lists across all games (with game labels).
        per_player_events: Dict[str, Dict[str, List[Tuple[str, GoalEvent]]]] = {}
        per_game_stats_by_label: Dict[str, List[Dict[str, str]]] = {}
        for r in results:
            game_label = r.get("label", "")
            ev_map: Dict[str, Dict[str, List[GoalEvent]]] = r.get("events", {}) or {}
            per_game_stats_by_label[game_label] = r.get("stats", []) or []
            for player_key, info in ev_map.items():
                dest = per_player_events.setdefault(
                    player_key, {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []}
                )
                for ev in info.get("goals", []):
                    dest["goals"].append((game_label, ev))
                for ev in info.get("assists", []):
                    dest["assists"].append((game_label, ev))
                for ev in info.get("gf_on_ice", []):
                    dest["gf_on_ice"].append((game_label, ev))
                for ev in info.get("ga_on_ice", []):
                    dest["ga_on_ice"].append((game_label, ev))

        # Add GT/GW goal counts into aggregated rows.
        _augment_aggregate_with_goal_details(agg_rows, per_player_events)

        # Cumulative sheet: points-based ordering.
        agg_df, _ = _build_stats_dataframe(
            agg_rows,
            agg_periods,
            sort_for_cumulative=True,
            include_shifts_in_stats=include_shifts_in_stats,
        )
        sheets: List[Tuple[str, pd.DataFrame]] = [("Cumulative", agg_df)]
        has_t2s = any(r.get("t2s_id") is not None for r in results)
        ordered_results = (
            sorted(
                results,
                key=lambda r: (
                    0 if r.get("t2s_id") is not None else 1,
                    r.get("t2s_id") if r.get("t2s_id") is not None else float("inf"),
                    r.get("order", 0),
                ),
            )
            if has_t2s
            else results
        )

        # Workbook tab order: keep Cumulative first, then list games in reverse
        # `--file-list` order so the most recent games (often listed last) appear first.
        per_game_sheet_results = list(reversed(results)) if args.file_list else ordered_results
        for r in per_game_sheet_results:
            # Per-game sheets: keep simple alphabetical ordering by player.
            df, _ = _build_stats_dataframe(
                r["stats"],
                r["periods"],
                sort_for_cumulative=False,
                include_shifts_in_stats=include_shifts_in_stats,
            )
            sheets.append((r["label"], df))
        consolidated_path = base_outdir / "player_stats_consolidated.xlsx"
        _write_consolidated_workbook(consolidated_path, sheets, per_game_denoms=per_game_denoms)
        try:
            print(f"📊 Consolidated workbook: {consolidated_path.resolve()}")
        except Exception:
            print("📊 Consolidated workbook written.")

        # Also write a one-sheet CSV summary for the cumulative sheet.
        try:
            df_csv = agg_df.copy()
            if "player" in df_csv.columns:
                df_csv["player"] = df_csv["player"].apply(_display_player_name)
            df_csv.columns = [
                f"{_display_col_name(c)} ({per_game_denoms[c]})"
                if c in per_game_denoms
                else _display_col_name(c)
                for c in df_csv.columns
            ]
            player_csv_path = base_outdir / "player_stats_consolidated.csv"
            df_csv.to_csv(player_csv_path, index=False)
        except Exception:
            pass

        # Team-level (game) stats consolidated across all games.
        if _write_game_stats_consolidated_files(base_outdir, results):
            try:
                print(
                    f"📊 Game stats consolidated: {(base_outdir / 'game_stats_consolidated.xlsx').resolve()}"
                )
            except Exception:
                print("📊 Game stats consolidated workbook written.")

        # Pair on-ice stats consolidated across all games (only includes games that had shift data).
        if _write_pair_on_ice_consolidated_files(
            base_outdir, results, include_toi=include_shifts_in_stats
        ):
            try:
                print(
                    f"📊 Pair on-ice consolidated: {(base_outdir / 'pair_on_ice_consolidated.xlsx').resolve()}"
                )
            except Exception:
                print("📊 Pair on-ice consolidated workbook written.")

        # Per-player cumulative detail files across all games.
        _write_cumulative_player_detail_files(
            base_outdir,
            agg_rows,
            per_player_events,
            per_game_stats_by_label,
            {str(r.get("label") or ""): (r.get("pair_on_ice") or []) for r in results if r.get("label")},
            include_shifts_in_stats=include_shifts_in_stats,
        )

        # Season (multi-game) highlight reel scripts (skipped with --no-scripts).
        _write_season_highlight_scripts(base_outdir, results, create_scripts=create_scripts)


if __name__ == "__main__":
    main()
