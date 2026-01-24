#!/usr/bin/env python3
"""
Extract shifts and stats from the 'dh-tv-12-1.xls' style sheet.

Outputs per-player:
  - {Jersey}_{Name}_stats.txt                -> scoreboard-time-based stats (plus/minus, goals for/against counted, etc.)

When `--shifts` is set (and `--no-scripts` is not), also writes per-player shift clip helpers:
  - {Jersey}_{Name}_video_times.txt          -> "videoStart videoEnd"
  - {Jersey}_{Name}_scoreboard_times.txt     -> "period scoreboardStart scoreboardEnd"

Goals can be specified via:
  --goal GF:2/13:45 --goal GA:1/05:12 ...
  --goal GF:OT/0:45  # OT goals use period 'OT' (treated as period 4)
or as a file with lines like:
  GF:1/13:47
  GA:2/09:15
  # comments and blank lines allowed

Alternatively, provide a TimeToScore spec to auto-fill goals (or run a game with no spreadsheets):
  --t2s 51602 --home                 # Your team is home (home scoring = GF)
  --t2s 51602 --away                 # Your team is away (away scoring = GF)
  --t2s 51602:HOME:stockton-r2       # Also sets the game label for T2S-only games

Install deps (for .xls):
  pip install pandas xlrd

Example:
  python scripts/parse_stats_inputs.py \
      --input dh-tv-12-1.xls \
      --outdir player_focus \
      --t2s 51602 --away \
      --keep-goalies
"""

import argparse
import base64
import datetime
import html as _html
import os
import re
import statistics
import sys
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple
from urllib.parse import urlencode
from urllib.request import urlopen

import pandas as pd
import yaml

# Ensure repo root is on sys.path so optional `hmlib.*` imports work when running
# this script directly from `scripts/`.
_repo_root = Path(__file__).resolve().parents[1]
if str(_repo_root) not in sys.path:
    sys.path.insert(0, str(_repo_root))

# Optional import of TimeToScore API (available in this repo). We lazily import it
# to avoid pulling in heavy `hmlib` dependencies when users run this script without T2S.
_t2s_api: Any = None
_t2s_api_loaded = False
_t2s_api_import_error: Optional[str] = None
_t2s_scrape_logged: set[tuple[int, str]] = set()


def _get_t2s_api() -> Any:
    global _t2s_api, _t2s_api_loaded, _t2s_api_import_error
    if _t2s_api_loaded:
        return _t2s_api
    _t2s_api_loaded = True
    try:  # pragma: no cover - optional at runtime
        from hmlib.time2score import api as t2s_api  # type: ignore

        _t2s_api = t2s_api
        _t2s_api_import_error = None
    except Exception as e:  # noqa: BLE001
        _t2s_api = None
        _t2s_api_import_error = f"{type(e).__name__}: {e}"
    return _t2s_api


def _log_t2s_scrape(game_id: int, purpose: str) -> None:
    key = (int(game_id), str(purpose or "").strip().lower())
    if key in _t2s_scrape_logged:
        return
    _t2s_scrape_logged.add(key)
    print(f"[t2s:{int(game_id)}] Scrape-only: {purpose}", file=sys.stderr)


# Header labels as they appear in the sheet
LABEL_START_SB = "Shift Start (Scoreboard Time)"
LABEL_END_SB = "Shift End (Scoreboard Time)"
LABEL_START_V = "Shift Start (Video Time)"
LABEL_END_V = "Shift End (Video Time)"

# Clip window durations (seconds)
EVENT_CLIP_PRE_S = 10
EVENT_CLIP_POST_S = 5
GOAL_CLIP_PRE_S = 20
GOAL_CLIP_POST_S = 10


# ----------------------------- utilities -----------------------------


def sanitize_name(s: str) -> str:
    s = s.strip().replace(" ", "_")
    return re.sub(r"[^A-Za-z0-9_\-]", "", s)


def _stats_base_dir_from_env() -> Optional[Path]:
    raw = str(os.environ.get("HOCKEYMOM_STATS_BASE_DIR") or "").strip()
    if not raw:
        return None
    try:
        p = Path(raw).expanduser()
        # If a file is provided accidentally, treat its parent as the base dir.
        if p.exists() and p.is_file():
            p = p.parent
        if not p.exists() or not p.is_dir():
            return None
        return p.resolve()
    except Exception:
        return None


def is_period_label(x: object) -> bool:
    return parse_period_label(x) is not None


def parse_period_label(x: object) -> Optional[int]:
    """
    Extract the period number from a label.
    Handles variants like 'Period 1', '1st Period', '1st Period (Blue team)', and 'OT'.
    """
    try:
        s = str(x).strip()
    except Exception:
        return None
    if not s:
        return None
    # Overtime. Some sheets label this like "OT", "Overtime", or "OT 3 on 3".
    # Treat OT labels as period 4, but only when the cell itself is an OT header
    # (avoid matching event labels like "Unforced OT").
    if re.match(r"(?i)^(?:ot|overtime)(?:\b|\d)", s):
        return 4
    m = re.search(r"(?i)(\d+)(?:st|nd|rd|th)?\s*period", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    m = re.search(r"(?i)period\s*(\d+)", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def _int_floor_seconds_component(sec_str: str) -> int:
    """Return integer seconds, flooring if fractional (e.g., '12.7' -> 12)."""
    try:
        return int(float(sec_str))
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Invalid seconds component '{sec_str}': {e}")


def parse_flex_time_to_seconds(s: str) -> int:
    """
    Accepts H:MM:SS(.fff) or M:SS(.fff) or MM:SS(.fff).
    Returns total seconds (int), flooring fractional seconds.
    """
    s = str(s).strip()
    # Some spreadsheets occasionally contain malformed time strings like "10::52".
    # Normalize repeated/trailing/leading separators so the parser is robust.
    s = re.sub(r"\s+", "", s)
    s = re.sub(r":{2,}", ":", s).strip(":")
    parts = s.split(":")
    if len(parts) == 3:
        h, m, sec = parts
        return int(h) * 3600 + int(m) * 60 + _int_floor_seconds_component(sec)
    elif len(parts) == 2:
        m, sec = parts
        return int(m) * 60 + _int_floor_seconds_component(sec)
    elif len(parts) == 1:
        # Support 'SS(.fff)' format (e.g., '58.2' -> 0:58)
        return _int_floor_seconds_component(parts[0])
    else:
        raise ValueError(f"Invalid time format '{s}'. Expected M:SS or H:MM:SS.")


def seconds_to_mmss_or_hhmmss(t: int) -> str:
    """Pretty printer for seconds: HH:MM:SS if >= 3600 else M:SS with minutes not zero-padded."""
    if t < 0:
        t = 0
    h = t // 3600
    r = t % 3600
    m = r // 60
    s = r % 60
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    else:
        return f"{m}:{s:02d}"


def seconds_to_hhmmss(t: int) -> str:
    """Always format seconds as HH:MM:SS with zero-padded hours."""
    if t < 0:
        t = 0
    h = t // 3600
    r = t % 3600
    m = r // 60
    s = r % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _clip_pre_post_s_for_event_type(event_type: str) -> Tuple[int, int]:
    et = str(event_type or "").strip().lower()
    if et in {"goal", "assist"}:
        return GOAL_CLIP_PRE_S, GOAL_CLIP_POST_S
    return EVENT_CLIP_PRE_S, EVENT_CLIP_POST_S


def _clean_html_fragment(s: str) -> str:
    if not s:
        return ""
    # Lightly strip tags (TimeToScore scoresheet pages are simple tables) and normalize whitespace.
    txt = re.sub(r"(?is)<[^>]+>", " ", str(s))
    txt = _html.unescape(txt)
    txt = txt.replace("\xa0", " ")
    return " ".join(txt.split()).strip()


T2S_CAHA_SCORESHEET_URL = "https://stats.caha.timetoscore.com/oss-scoresheet"


def _fetch_t2s_scoresheet_html(game_id: int, *, timeout_s: float = 20.0) -> str:
    url = f"{T2S_CAHA_SCORESHEET_URL}?{urlencode({'game_id': int(game_id)})}"
    with urlopen(url, timeout=float(timeout_s)) as resp:  # noqa: S310 (trusted URL)
        data = resp.read()
    return data.decode("utf-8", "ignore")


def _t2s_team_names_from_scoresheet_html(html: str) -> Tuple[Optional[str], Optional[str]]:
    # NOTE: the scoresheet HTML has malformed tags in spots (e.g., <td> ... </th>),
    # so use a tolerant regex rather than a strict HTML parser.
    m_vis = re.search(r"(?is)<tr>\s*<th>\s*Visitor\s*</th>\s*<td>\s*([^<]+)", html)
    m_home = re.search(r"(?is)<tr>\s*<th>\s*Home\s*</th>\s*<td>\s*([^<]+)", html)
    visitor = _clean_html_fragment(m_vis.group(1)) if m_vis else None
    home = _clean_html_fragment(m_home.group(1)) if m_home else None
    return (visitor or None, home or None)


def _t2s_team_names_for_game(
    t2s_id: int,
    *,
    allow_remote: bool,
    allow_full_sync: bool,
    hockey_db_dir: Path,
) -> tuple[Optional[str], Optional[str]]:
    home_name: Optional[str] = None
    away_name: Optional[str] = None
    t2s_api = _get_t2s_api()
    if t2s_api is not None:
        try:
            if allow_remote and not allow_full_sync:
                _log_t2s_scrape(int(t2s_id), "game details (team names)")
            with _working_directory(hockey_db_dir):
                info = t2s_api.get_game_details(
                    int(t2s_id),
                    sync_if_missing=(bool(allow_full_sync) if allow_remote else False),
                    fetch_stats_if_missing=bool(allow_remote),
                )
            home = ((info or {}).get("home") or {}).get("team") or {}
            away = ((info or {}).get("away") or {}).get("team") or {}
            home_name = str(home.get("name") or "").strip() or None
            away_name = str(away.get("name") or "").strip() or None
        except Exception as e:  # noqa: BLE001
            print(f"[t2s:{t2s_id}] Failed to resolve team names via API: {e}", file=sys.stderr)
            home_name = None
            away_name = None
    if (not home_name or not away_name) and allow_remote:
        try:
            if not allow_full_sync:
                _log_t2s_scrape(int(t2s_id), "scoresheet HTML (team names)")
            html = _fetch_t2s_scoresheet_html(int(t2s_id))
            visitor, home = _t2s_team_names_from_scoresheet_html(html)
            away_name = str(visitor or "").strip() or away_name
            home_name = str(home or "").strip() or home_name
        except Exception as e:  # noqa: BLE001
            print(
                f"[t2s:{t2s_id}] Failed to fetch/parse scoresheet HTML for team names: {e}",
                file=sys.stderr,
            )
    return home_name, away_name


def _t2s_team_logo_urls_for_game(
    t2s_id: int,
    *,
    allow_remote: bool,
    allow_full_sync: bool,
    hockey_db_dir: Path,
) -> tuple[Optional[str], Optional[str]]:
    if not allow_remote:
        return None, None
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        return None, None
    try:
        if not allow_full_sync:
            _log_t2s_scrape(int(t2s_id), "team logos")
        with _working_directory(hockey_db_dir):
            try:
                info = t2s_api.get_game_details(
                    int(t2s_id),
                    sync_if_missing=(bool(allow_full_sync) if allow_remote else False),
                    fetch_stats_if_missing=bool(allow_remote),
                )
            except TypeError:
                info = t2s_api.get_game_details(int(t2s_id))
    except Exception:
        return None, None
    game = (info or {}).get("game") or {}
    home = ((info or {}).get("home") or {}).get("team") or {}
    away = ((info or {}).get("away") or {}).get("team") or {}
    try:
        season_id = int(game.get("season_id") or 0)
    except Exception:
        season_id = 0
    league_id = None
    try:
        league_id = int(game.get("league") or 0)
    except Exception:
        league_id = None
    try:
        from hmlib.time2score import direct as t2s_direct  # type: ignore
    except Exception:
        return None, None
    home_url = None
    away_url = None
    try:
        home_id = int(home.get("team_id") or 0)
        if home_id:
            home_url = t2s_direct.scrape_team_logo_url(
                "caha", season_id=season_id, team_id=home_id, league_id=league_id
            )
    except Exception:
        home_url = None
    try:
        away_id = int(away.get("team_id") or 0)
        if away_id:
            away_url = t2s_direct.scrape_team_logo_url(
                "caha", season_id=season_id, team_id=away_id, league_id=league_id
            )
    except Exception:
        away_url = None
    return home_url, away_url


def _t2s_default_period_length_seconds_from_scoresheet_html(html: str) -> int:
    # Example: "<td colspan=2>Period Lengths:</td><td align=center>15</td>..."
    m = re.search(r"(?is)Period\s*Lengths\s*:\s*</td>\s*<td[^>]*>\s*(\d+)\s*<", html)
    if m:
        try:
            return int(m.group(1)) * 60
        except Exception:
            pass
    return 15 * 60


def _t2s_penalties_from_scoresheet_html(
    html: str,
    *,
    our_side: Optional[str],
) -> List[Dict[str, Any]]:
    visitor_name, home_name = _t2s_team_names_from_scoresheet_html(html)
    # The scoresheet contains two Penalties tables; the first corresponds to Visitor/away,
    # and the second corresponds to Home.
    tables = re.findall(
        r"(?is)<table[^>]*>\s*<tr>\s*<th[^>]*colspan\s*=\s*8[^>]*>\s*Penalties\s*</th>.*?</table>",
        html,
    )

    out: List[Dict[str, Any]] = []

    def _parse_table(table_html: str, *, side: str, team_name: Optional[str]) -> None:
        for row_html in re.findall(r"(?is)<tr[^>]*>.*?</tr>", table_html):
            # Skip header rows.
            if re.search(r"(?is)<th[^>]*>\s*Per\s*</th>", row_html):
                continue
            cells = re.findall(r"(?is)<t[dh][^>]*>(.*?)</t[dh]>", row_html)
            if len(cells) < 8:
                continue
            # Per, #, Infraction, Min, Off Ice, Start, End, On Ice
            per_raw = _clean_html_fragment(cells[0])
            per = parse_period_token(per_raw)
            if per is None:
                continue
            num_raw = _clean_html_fragment(cells[1])
            infraction = _clean_html_fragment(cells[2])
            minutes_raw = _clean_html_fragment(cells[3])
            off_ice = _clean_html_fragment(cells[4])
            start = _clean_html_fragment(cells[5]) or off_ice
            end = _clean_html_fragment(cells[6])
            on_ice = _clean_html_fragment(cells[7])

            try:
                game_s = parse_flex_time_to_seconds(start)
            except Exception:
                continue

            jersey = _normalize_jersey_number(num_raw)
            details_parts: List[str] = []
            if jersey:
                details_parts.append(f"#{jersey}")
            if infraction:
                details_parts.append(infraction)
            if minutes_raw:
                details_parts.append(f"{minutes_raw}m")
            t_parts: List[str] = []
            if start:
                t_parts.append(f"start {start}")
            if end:
                t_parts.append(f"end {end}")
            if t_parts:
                details_parts.append(f"({', '.join(t_parts)})")
            if on_ice and on_ice != end:
                details_parts.append(f"on ice {on_ice}")

            out.append(
                {
                    "event_type": "Penalty",
                    "source": "t2s",
                    "team_raw": team_name or side,
                    "team_side": side,
                    "for_against": "Against",
                    "period": int(per),
                    "game_s": int(game_s),
                    "details": " ".join([p for p in details_parts if p]).strip(),
                    "attributed_jerseys": [jersey] if jersey else [],
                }
            )

    if len(tables) >= 1:
        _parse_table(tables[0], side="away", team_name=visitor_name)
    if len(tables) >= 2:
        _parse_table(tables[1], side="home", team_name=home_name)
    return out


def _t2s_goalie_changes_from_scoresheet_html(
    html: str,
    *,
    our_side: Optional[str],
) -> List[Dict[str, Any]]:
    visitor_name, home_name = _t2s_team_names_from_scoresheet_html(html)
    default_period_s = _t2s_default_period_length_seconds_from_scoresheet_html(html)

    def _section(title: str) -> List[str]:
        # Capture between <th colspan=2>Title</th> and the next section header / end of table.
        if title == "Home Goalie Changes":
            m = re.search(
                r"(?is)<th[^>]*colspan\s*=\s*2[^>]*>\s*Home Goalie Changes\s*</th>(.*?)(?:<th[^>]*colspan\s*=\s*2[^>]*>\s*Visitor Changes\s*</th>)",
                html,
            )
        else:
            m = re.search(
                r"(?is)<th[^>]*colspan\s*=\s*2[^>]*>\s*Visitor Changes\s*</th>(.*?)(?:</table>)",
                html,
            )
        if not m:
            return []
        seg = m.group(1)
        # Each row is a single-cell <td colspan=2>...</td>
        vals = re.findall(r"(?is)<td[^>]*colspan\s*=\s*2[^>]*>(.*?)</td>", seg)
        out_vals: List[str] = []
        for v in vals:
            vv = _clean_html_fragment(v)
            if vv:
                out_vals.append(vv)
        return out_vals

    out: List[Dict[str, Any]] = []

    def _add_lines(lines: List[str], *, side: str, team_name: Optional[str]) -> None:
        for line in lines:
            # Examples:
            #  - "Joshua T Brown Starting"
            #  - "Empty Net 3-14:47"
            #  - "Joshua Rocha 2-13:42"
            details = str(line).strip()
            period: Optional[int] = None
            game_s: Optional[int] = None
            m = re.search(r"(?i)\b(\d+)\s*-\s*(\d{1,2}:\d{2})\b", details)
            if m:
                period = parse_period_token(m.group(1))
                try:
                    game_s = parse_flex_time_to_seconds(m.group(2))
                except Exception:
                    game_s = None
                details = details[: m.start()].strip()
            elif re.search(r"(?i)\bstarting\b", details):
                period = 1
                game_s = int(default_period_s)

            if period is None:
                continue

            out.append(
                {
                    "event_type": "GoalieChange",
                    "source": "t2s",
                    "team_raw": team_name or side,
                    "team_side": side,
                    "for_against": "",
                    "period": int(period),
                    "game_s": int(game_s) if game_s is not None else None,
                    "details": details,
                    "attributed_jerseys": [],
                }
            )

    _add_lines(_section("Home Goalie Changes"), side="home", team_name=home_name)
    _add_lines(_section("Visitor Changes"), side="away", team_name=visitor_name)
    return out


def t2s_events_from_scoresheet(
    game_id: int,
    *,
    our_side: Optional[str],
    allow_remote: bool = True,
    allow_full_sync: bool = True,
) -> List[Dict[str, Any]]:
    """
    Best-effort TimeToScore (CAHA) event scraper used to enrich `all_events_summary.csv`.

    Includes:
      - penalties (start times)
      - goalie changes (including empty net + starting goalie)
    """
    if not allow_remote:
        print(
            f"[t2s:{game_id}] Skipping scoresheet fetch (cache-only; pass without --t2s-cache-only to enable).",
            file=sys.stderr,
        )
        return []
    if not allow_full_sync:
        _log_t2s_scrape(int(game_id), "scoresheet HTML (penalties/goalies)")
    try:
        html = _fetch_t2s_scoresheet_html(int(game_id))
    except Exception as e:  # noqa: BLE001
        print(f"[t2s:{game_id}] Failed to fetch scoresheet HTML: {e}", file=sys.stderr)
        return []
    out: List[Dict[str, Any]] = []
    try:
        out.extend(_t2s_penalties_from_scoresheet_html(html, our_side=our_side))
    except Exception as e:  # noqa: BLE001
        print(f"[t2s:{game_id}] Failed to parse penalties: {e}", file=sys.stderr)
    try:
        out.extend(_t2s_goalie_changes_from_scoresheet_html(html, our_side=our_side))
    except Exception as e:  # noqa: BLE001
        print(f"[t2s:{game_id}] Failed to parse goalie changes: {e}", file=sys.stderr)
    return out


def forward_fill_header_labels(header_row: pd.Series) -> Dict[str, List[int]]:
    """
    Given a header row with merged-like cells (label then NaNs across its span),
    forward-fill labels across columns to group column indices by label.
    """
    labels_by_col: List[Optional[str]] = []
    current = None
    for c in range(len(header_row)):
        val = header_row.iloc[c]
        if pd.notna(val) and str(val).strip():
            current = str(val).strip()
        labels_by_col.append(current)

    groups: Dict[str, List[int]] = {}
    for idx, lab in enumerate(labels_by_col):
        if not lab:
            continue
        groups.setdefault(lab, []).append(idx)
    return groups


def _sniff_image_content_type(data: bytes) -> Optional[str]:
    if not data:
        return None
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if data.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if data.startswith(b"GIF87a") or data.startswith(b"GIF89a"):
        return "image/gif"
    if len(data) >= 12 and data[0:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    head = data[:2048].lstrip()
    if head.startswith(b"<?xml") or head.lower().startswith(b"<svg") or b"<svg" in head.lower():
        return "image/svg+xml"
    return None


def _content_type_from_ext(path: Path) -> Optional[str]:
    ext = path.suffix.lower()
    if ext == ".png":
        return "image/png"
    if ext in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if ext == ".gif":
        return "image/gif"
    if ext == ".webp":
        return "image/webp"
    if ext == ".svg":
        return "image/svg+xml"
    return None


def _normalize_logo_b64(s: str) -> str:
    ss = str(s or "").strip()
    m = re.match(r"(?is)^data:[^;]+;base64,(.+)$", ss)
    if m:
        ss = m.group(1).strip()
    return re.sub(r"\s+", "", ss)


def _load_logo_fields_from_meta(
    meta: Dict[str, str],
    *,
    base_dir: Optional[Path] = None,
    warn_label: str = "",
) -> Dict[str, str]:
    """
    Build webapp shift_package logo fields from per-game metadata.

    Supports either:
      - home_logo=/path/to/image.png
      - home_team_icon=/path/to/image.png  (alias)
      - home_logo_base64=<base64>
      - home_logo_content_type=image/png  (optional; otherwise guessed)
    and same for away_*.

    Missing files are warnings (non-fatal).
    """

    def _warn(msg: str) -> None:
        if not msg:
            return
        prefix = f"[warning]{' [' + warn_label + ']' if warn_label else ''}"
        print(f"{prefix} {msg}", file=sys.stderr)

    def _resolve_path(p: str) -> Path:
        pp = Path(str(p)).expanduser()
        if base_dir is not None and not pp.is_absolute():
            pp = (Path(base_dir) / pp).resolve()
        return pp

    def _one_side(side: str) -> Dict[str, str]:
        out: Dict[str, str] = {}
        side_l = str(side).strip().lower()
        if side_l not in {"home", "away"}:
            return out

        b64_key = f"{side_l}_logo_base64"
        path_key = f"{side_l}_logo"
        icon_key = f"{side_l}_team_icon"
        ct_key = f"{side_l}_logo_content_type"

        b64_raw = str(meta.get(b64_key) or "").strip()
        path_raw = str(meta.get(path_key) or meta.get(icon_key) or "").strip()
        ct_raw = str(meta.get(ct_key) or "").strip()

        if b64_raw:
            b64_norm = _normalize_logo_b64(b64_raw)
            out[f"{side_l}_logo_b64"] = b64_norm
            if ct_raw:
                out[f"{side_l}_logo_content_type"] = ct_raw
                return out
            try:
                data = base64.b64decode(b64_norm.encode("ascii"), validate=False)
            except Exception:
                data = b""
            out[f"{side_l}_logo_content_type"] = _sniff_image_content_type(data) or "image/png"
            return out

        if not path_raw:
            return out

        logo_path = _resolve_path(path_raw)
        if not logo_path.exists():
            _warn(f"{side_l}_logo path does not exist: {logo_path}")
            return out
        if not logo_path.is_file():
            _warn(f"{side_l}_logo path is not a file: {logo_path}")
            return out

        try:
            data = logo_path.read_bytes()
        except Exception as e:  # noqa: BLE001
            _warn(f"failed to read {side_l}_logo file {logo_path}: {e}")
            return out

        if not data:
            _warn(f"{side_l}_logo file is empty: {logo_path}")
            return out
        if len(data) > 5 * 1024 * 1024:
            _warn(f"{side_l}_logo file too large (>5MB), skipping: {logo_path}")
            return out

        out[f"{side_l}_logo_b64"] = base64.b64encode(data).decode("ascii")
        out[f"{side_l}_logo_content_type"] = (
            ct_raw
            or _content_type_from_ext(logo_path)
            or _sniff_image_content_type(data)
            or "image/png"
        )
        return out

    out_all: Dict[str, str] = {}
    out_all.update(_one_side("home"))
    out_all.update(_one_side("away"))
    return out_all


def _starts_at_from_meta(meta: Dict[str, str], *, warn_label: str = "") -> Optional[str]:
    """
    Resolve an external game's starts_at timestamp from metadata.

    Supported keys:
      - starts_at=<datetime string> (passed through after light normalization)
      - date=<YYYY-MM-DD> (converted to 'YYYY-MM-DD 00:00:00' unless `time` is provided)
      - time=<HH:MM[:SS]> (used with `date` to form 'YYYY-MM-DD HH:MM:SS')

    This is only used for external games (the webapp does not update starts_at on existing games).
    """

    def _warn(msg: str) -> None:
        if not msg:
            return
        prefix = f"[warning]{' [' + warn_label + ']' if warn_label else ''}"
        print(f"{prefix} {msg}", file=sys.stderr)

    def _fmt_dt(dt_obj: datetime.datetime) -> str:
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")

    def _parse_date_only(s: str) -> Optional[str]:
        ss = s.strip()
        if not ss:
            return None
        try:
            d = datetime.date.fromisoformat(ss)
            return f"{d.isoformat()} 00:00:00"
        except Exception:
            pass
        m = re.match(r"^\s*(\d{1,2})/(\d{1,2})/(\d{2,4})\s*$", ss)
        if m:
            mm = int(m.group(1))
            dd = int(m.group(2))
            yy = int(m.group(3))
            if yy < 100:
                yy = 2000 + yy
            try:
                d2 = datetime.date(yy, mm, dd)
                return f"{d2.isoformat()} 00:00:00"
            except Exception:
                return None
        return None

    def _parse_time_only(s: str) -> Optional[datetime.time]:
        ss = s.strip()
        if not ss:
            return None
        try:
            # Accept HH:MM[:SS]
            return datetime.time.fromisoformat(ss)
        except Exception:
            return None

    starts_at_raw = str(meta.get("starts_at") or "").strip()
    if starts_at_raw:
        # Normalize ISO-ish strings to a SQL-friendly format when possible.
        ss = starts_at_raw.replace("T", " ").strip()
        try:
            # Accept full ISO datetime; allow seconds to be omitted.
            dt_obj = datetime.datetime.fromisoformat(ss)
            return _fmt_dt(dt_obj)
        except Exception:
            # If user gives a bare date in starts_at, treat it as date-only.
            parsed = _parse_date_only(ss)
            if parsed:
                return parsed
            # Otherwise pass through (webapp may still accept it).
            return starts_at_raw

    date_raw = str(meta.get("date") or "").strip()
    time_raw = str(meta.get("time") or "").strip()
    if date_raw:
        parsed2 = _parse_date_only(date_raw)
        if parsed2:
            if time_raw:
                t = _parse_time_only(time_raw)
                if t is None:
                    _warn(
                        f"could not parse time={time_raw!r} (expected HH:MM or HH:MM:SS); ignoring"
                    )
                    return parsed2
                try:
                    d = datetime.date.fromisoformat(parsed2.split(" ", 1)[0])
                    dt_obj = datetime.datetime.combine(d, t)
                    return _fmt_dt(dt_obj)
                except Exception:
                    return parsed2
            return parsed2
        _warn(f"could not parse date={date_raw!r} (expected YYYY-MM-DD or M/D/YYYY); ignoring")
        return None
    if time_raw:
        _warn("time=... provided without date=...; ignoring")
    return None


def _starts_at_from_t2s_game_id(
    t2s_game_id: int,
    *,
    hockey_db_dir: Path,
    warn_label: str = "",
    allow_remote: bool = True,
    allow_full_sync: bool = True,
) -> Optional[str]:
    def _warn(msg: str) -> None:
        if not msg:
            return
        prefix = f"[warning]{' [' + warn_label + ']' if warn_label else ''}"
        print(f"{prefix} {msg}", file=sys.stderr)

    try:
        from hmlib.time2score.api import get_game_details
    except Exception as e:  # noqa: BLE001
        _warn(f"failed to import TimeToScore API helpers: {e}")
        return None

    try:
        with _working_directory(hockey_db_dir):
            if allow_remote and not allow_full_sync:
                _log_t2s_scrape(int(t2s_game_id), "game details (start time)")
            info = get_game_details(
                int(t2s_game_id),
                season=None,
                sync_if_missing=(bool(allow_full_sync) if allow_remote else False),
                fetch_stats_if_missing=False,
            )
    except Exception as e:  # noqa: BLE001
        _warn(f"failed to fetch TimeToScore game details for game_id={t2s_game_id}: {e}")
        return None

    st = ((info or {}).get("game") or {}).get("start_time")
    if st is None:
        return None
    if isinstance(st, datetime.datetime):
        return st.strftime("%Y-%m-%d %H:%M:%S")
    ss = str(st).strip()
    if not ss:
        return None
    ss = ss.replace("T", " ").strip()
    try:
        dt_obj = datetime.datetime.fromisoformat(ss)
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        # If we only get a date, normalize to midnight.
        try:
            d = datetime.date.fromisoformat(ss)
            return f"{d.isoformat()} 00:00:00"
        except Exception:
            return None


def _normalize_header_label(label: str) -> str:
    """Normalize header label for comparison (case/spacing/punctuation insensitive)."""
    return re.sub(r"[^a-z0-9]+", "", label.lower())


def _resolve_header_columns(groups: Dict[str, List[int]], *candidates: str) -> List[int]:
    """Return columns for the first candidate label that matches after normalization."""
    norm_map = {_normalize_header_label(k): v for k, v in groups.items() if k}
    for cand in candidates:
        key = _normalize_header_label(cand)
        cols = norm_map.get(key)
        if cols:
            return cols
    # Fallback: allow substring match to tolerate slight label variants
    for cand in candidates:
        key = _normalize_header_label(cand)
        for norm_label, cols in norm_map.items():
            if cols and key and key in norm_label:
                return cols
    return []


def extract_pairs_from_row(
    row: pd.Series, start_cols: List[int], end_cols: List[int]
) -> List[Tuple[str, str]]:
    """
    From start/end column groups, collect non-empty strings and pair positionally.
    Start/End order in the sheet can be higher->lower or lower->higher; pairing is positional only.
    If a value is a time (datetime.time or Timestamp with time), keep only hour:minute.
    """

    def format_cell(val) -> str:
        if pd.isna(val):
            return ""
        # Already string → return trimmed
        if isinstance(val, str):
            s = val.strip()
            # Treat textual NaN/None markers as empty cells
            if s.lower() in {"nan", "none"}:
                return ""
            return s
        # datetime.time (Excel time)
        if isinstance(val, datetime.time):
            return val.strftime("%H:%M")
        # pandas Timestamp (could include date/time)
        if isinstance(val, pd.Timestamp):
            if val.time() != datetime.time(0, 0):  # has time portion
                return val.strftime("%H:%M")
            else:
                return val.strftime("%Y-%m-%d")  # just a date
        # Fallback
        return str(val).strip()

    starts = [format_cell(row[c]) for c in start_cols if format_cell(row[c])]
    ends = [format_cell(row[c]) for c in end_cols if format_cell(row[c])]
    n = min(len(starts), len(ends))
    return [(starts[i], ends[i]) for i in range(n)]


# Historically we tried to normalize certain end-of-period times (e.g., 15:00 or 20:00)
# to 0:00 when sheets encoded the period end as its nominal start time. This proved
# brittle and could distort genuine in-period times (e.g., 12:00). We now leave
# scoreboard end times exactly as entered in the sheet.
def _normalize_sb_end_time(t: str) -> str:
    return t


@contextmanager
def _working_directory(path: Path) -> Iterator[None]:
    """Temporarily switch working directory, creating it if needed."""
    prev = Path.cwd()
    path.mkdir(parents=True, exist_ok=True)
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(prev)


def _duration_to_seconds(val: str) -> int:
    try:
        return parse_flex_time_to_seconds(str(val)) if val not in ("", None) else 0
    except Exception:
        return 0


def _seconds_to_compact_hms(t: int) -> str:
    """
    Compact duration formatting:
      - H:MM:SS if >= 3600
      - M:SS if >= 60
      - SS if < 60
    """
    if t < 0:
        t = 0
    if t < 60:
        return str(int(t))
    h = t // 3600
    r = t % 3600
    m = r // 60
    s = r % 60
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def _format_duration(total_seconds: int) -> str:
    return _seconds_to_compact_hms(total_seconds)


def _autosize_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Best-effort column auto-width for Excel sheets."""
    try:
        col_widths = []
        for col in df.columns:
            header_len = max([len(x) for x in str(col).splitlines()] or [len(str(col))])
            max_len = max([header_len] + [len(str(x)) for x in df[col].astype(str).fillna("")])
            # Leave extra room for Excel's filter dropdown arrow so it doesn't overlap the header text.
            col_widths.append(min(max(max_len + 4, 10), 80))
        if writer.engine == "openpyxl":
            from openpyxl.utils import get_column_letter

            ws = writer.sheets.get(sheet_name)
            if ws:
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
        elif writer.engine == "xlsxwriter":
            ws = writer.sheets.get(sheet_name)
            if ws:
                for i, width in enumerate(col_widths):
                    ws.set_column(i, i, width)
    except Exception:
        pass


def _wrap_header_after_words(header: str, *, words_per_line: int = 3) -> str:
    if not header:
        return header
    parts = str(header).strip().split()
    if len(parts) <= words_per_line:
        return str(header)
    lines = [" ".join(parts[i : i + words_per_line]) for i in range(0, len(parts), words_per_line)]
    return "\n".join(lines)


def _write_styled_xlsx_table(
    xlsx_path: Path,
    df: pd.DataFrame,
    *,
    sheet_name: str,
    title: str,
    words_per_line: int = 2,
    number_formats: Optional[Dict[str, str]] = None,
    text_columns: Optional[List[str]] = None,
    align_right_columns: Optional[List[str]] = None,
    merge_columns: Optional[List[str]] = None,
) -> None:
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=words_per_line)
                for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            _apply_excel_table_style(writer, sheet_name, title=title, df=df_excel)
            try:
                ws = writer.sheets.get(sheet_name)
                if ws is not None:
                    header_row = 2
                    data_start_row = 3
                    nrows = int(getattr(df_excel, "shape", (0, 0))[0] or 0)
                    if nrows > 0:
                        header_cells = list(ws[header_row])
                        col_idx_by_name: Dict[str, int] = {}
                        for idx, cell in enumerate(header_cells, start=1):
                            if cell.value is None:
                                continue
                            col_idx_by_name[str(cell.value).replace("\n", " ").strip()] = idx
                        col_idx_by_norm: Dict[str, int] = {
                            re.sub(r"\s+", " ", str(k).strip()).casefold(): v
                            for k, v in col_idx_by_name.items()
                        }

                        def _apply_number_format(col_name: str, fmt: str) -> None:
                            col_idx = col_idx_by_name.get(str(col_name).strip())
                            if not col_idx:
                                return
                            for r in range(data_start_row, data_start_row + nrows):
                                ws.cell(row=r, column=col_idx).number_format = fmt

                        def _apply_alignment_right(col_name: str) -> None:
                            col_idx = col_idx_by_name.get(str(col_name).strip())
                            if not col_idx:
                                return
                            from openpyxl.styles import Alignment

                            for r in range(data_start_row, data_start_row + nrows):
                                cell = ws.cell(row=r, column=col_idx)
                                cell.alignment = Alignment(
                                    horizontal="right", vertical=cell.alignment.vertical or "center"
                                )

                        # Apply explicit number formats (e.g., 1-decimal percent columns).
                        for col_name, fmt in (number_formats or {}).items():
                            _apply_number_format(col_name, fmt)

                        # Force key text columns to "Text" format (best-effort; Excel spellcheck
                        # disable is not supported via openpyxl).
                        auto_text_cols = {"player", "Player", "teammate", "Teammate"}
                        for col_name in set((text_columns or [])) | auto_text_cols:
                            _apply_number_format(col_name, "@")

                        # Right-align duration/time columns (best-effort).
                        auto_right_cols: set[str] = set(align_right_columns or [])
                        for name in col_idx_by_name.keys():
                            n = name.strip()
                            if any(
                                tok in n
                                for tok in [
                                    "TOI",
                                    "Time",
                                    "Overlap",
                                    "Average Shift",
                                    "Median Shift",
                                    "Longest Shift",
                                    "Shortest Shift",
                                ]
                            ):
                                auto_right_cols.add(n)
                        for col_name in auto_right_cols:
                            _apply_alignment_right(col_name)

                        if merge_columns:
                            from openpyxl.styles import Alignment

                            def _merge_column(col_name: str) -> None:
                                key = re.sub(r"\s+", " ", str(col_name).strip()).casefold()
                                col_idx = col_idx_by_norm.get(key)
                                if not col_idx:
                                    return
                                start = data_start_row
                                prev_val = None
                                for r in range(data_start_row, data_start_row + nrows):
                                    val = ws.cell(row=r, column=col_idx).value
                                    if val is None or str(val).strip() == "":
                                        cur_val = None
                                    else:
                                        cur_val = val
                                    if r == data_start_row:
                                        prev_val = cur_val
                                        continue
                                    if cur_val == prev_val:
                                        continue
                                    if prev_val is not None and r - start > 1:
                                        ws.merge_cells(
                                            start_row=start,
                                            start_column=col_idx,
                                            end_row=r - 1,
                                            end_column=col_idx,
                                        )
                                        ws.cell(row=start, column=col_idx).alignment = Alignment(
                                            horizontal="left",
                                            vertical="center",
                                        )
                                    start = r
                                    prev_val = cur_val
                                if prev_val is not None and (data_start_row + nrows - start) > 1:
                                    ws.merge_cells(
                                        start_row=start,
                                        start_column=col_idx,
                                        end_row=data_start_row + nrows - 1,
                                        end_column=col_idx,
                                    )
                                    ws.cell(row=start, column=col_idx).alignment = Alignment(
                                        horizontal="left",
                                        vertical="center",
                                    )

                            for col_name in merge_columns:
                                _merge_column(col_name)
            except Exception as e:  # noqa: BLE001
                print(
                    f"[warning] Failed to apply Excel merge formatting to sheet '{sheet_name}': {e}",
                    file=sys.stderr,
                )
            _autosize_columns(writer, sheet_name, df_excel)
    except Exception as e:  # noqa: BLE001
        print(f"[warning] Failed to apply Excel formatting: {e}", file=sys.stderr)


def _apply_excel_header_wrap(
    writer: pd.ExcelWriter, sheet_name: str, *, header_row: int = 1
) -> None:
    """
    Enable wrap-text for a header row when using openpyxl.
    """
    try:
        if writer.engine != "openpyxl":
            return
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            return
        from openpyxl.styles import Alignment

        max_lines = 1
        for cell in ws[header_row]:
            if cell.value:
                max_lines = max(max_lines, len(str(cell.value).splitlines()))
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[header_row].height = max(15.0, 15.0 * max_lines)
    except Exception:
        return


def _apply_excel_table_style(
    writer: pd.ExcelWriter,
    sheet_name: str,
    *,
    title: str,
    df: pd.DataFrame,
) -> None:
    """
    Apply a simple "teal header + banded gray rows" theme to a sheet written
    with pandas. Assumes the DataFrame was written with `startrow=1` so that:
      - Row 1 is available for the title
      - Row 2 is the header row
      - Row 3.. are data rows
    """
    try:
        if writer.engine != "openpyxl":
            return
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            return

        ncols = int(getattr(df, "shape", (0, 0))[1] or 0)
        nrows = int(getattr(df, "shape", (0, 0))[0] or 0)
        if ncols <= 0:
            return

        from openpyxl.styles import Alignment, Font, PatternFill

        teal_fill = PatternFill(fill_type="solid", start_color="FF009688", end_color="FF009688")
        header_font = Font(color="FF000000", bold=True)
        title_font = Font(color="FF000000", bold=True, size=14)
        band_a = PatternFill(fill_type="solid", start_color="FFE6E6E6", end_color="FFE6E6E6")
        band_b = PatternFill(fill_type="solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
        from openpyxl.styles import Border, Side

        white_side = Side(style="thin", color="FFFFFFFF")
        white_border = Border(
            left=white_side,
            right=white_side,
            top=white_side,
            bottom=white_side,
        )

        title_row = 1
        header_row = 2
        data_start_row = 3

        # Title row (merged across all columns)
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.value = title
        title_cell.fill = teal_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 24.0

        # Ensure the merged region is fully teal in Excel viewers that don't
        # propagate styles from the top-left cell.
        for c in range(1, ncols + 1):
            cell = ws.cell(row=title_row, column=c)
            cell.fill = teal_fill

        # Header styling
        for c in range(1, ncols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = teal_fill
            cell.font = header_font

        _apply_excel_header_wrap(writer, sheet_name, header_row=header_row)

        # Banded rows (data)
        for i in range(nrows):
            r = data_start_row + i
            fill = band_a if (i % 2 == 0) else band_b
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = fill

        # White grid borders (title + header + data)
        last_row = data_start_row + max(nrows, 1) - 1 if nrows > 0 else header_row
        for r in range(title_row, last_row + 1):
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).border = white_border

        # Enable Excel auto-filters (dropdowns) on the header row.
        try:
            from openpyxl.utils import get_column_letter

            last_col = get_column_letter(ncols)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
        except Exception:  # noqa: BLE001
            # Auto-filter support is optional; ignore failures (e.g., missing/old openpyxl).
            pass

        # Right-align duration/time-like columns (strings such as '54', '1:02', '2:10:03').
        try:
            header_cells = list(ws[header_row])
            right_cols: List[int] = []
            for idx, cell in enumerate(header_cells, start=1):
                if cell.value is None:
                    continue
                name = str(cell.value).replace("\n", " ").strip()
                if any(
                    tok in name
                    for tok in [
                        "TOI",
                        "Time",
                        "Overlap",
                        "Average Shift",
                        "Median Shift",
                        "Longest Shift",
                        "Shortest Shift",
                        "Video",
                        "Duration",
                    ]
                ):
                    right_cols.append(idx)
            if right_cols and nrows > 0:
                for i in range(nrows):
                    r = data_start_row + i
                    for c in right_cols:
                        cell = ws.cell(row=r, column=c)
                        cell.alignment = Alignment(
                            horizontal="right", vertical=cell.alignment.vertical or "center"
                        )
        except Exception:
            pass

        # Freeze key identity columns and the title+header rows so the important
        # left-side columns stay visible while scrolling in Excel/LibreOffice.
        #
        # Default: freeze column A + rows 1-2 (pane starts at B3).
        # Player stats tables now have: [Jersey #, Player, ...] so freeze both.
        try:
            h1 = ws.cell(row=header_row, column=1).value
            h2 = ws.cell(row=header_row, column=2).value
            h3 = ws.cell(row=header_row, column=3).value
            h4 = ws.cell(row=header_row, column=4).value
            h1n = str(h1 or "").replace("\n", " ").strip()
            h2n = str(h2 or "").replace("\n", " ").strip()
            h3n = str(h3 or "").replace("\n", " ").strip()
            h4n = str(h4 or "").replace("\n", " ").strip()

            # Pair-on-ice: freeze both player identity columns and teammate identity columns.
            if (
                h1n == "Player Jersey #"
                and h2n == "Player"
                and h3n == "Teammate Jersey #"
                and h4n == "Teammate"
            ):
                ws.freeze_panes = "E3"
            # Player stats: freeze jersey + player name.
            elif h1n in {"Jersey #", "Jersey", "Jersey No", "Jersey Number"} and h2n == "Player":
                ws.freeze_panes = "C3"
            else:
                ws.freeze_panes = "B3"
        except Exception:
            ws.freeze_panes = "B3"
    except Exception:
        return


def _collect_sheet_jerseys(
    xls_path: Path, sheet_name: Optional[str], keep_goalies: bool
) -> set[str]:
    df = pd.read_excel(xls_path, sheet_name=(0 if sheet_name is None else sheet_name), header=None)
    (
        used_event_log,
        _video_pairs_by_player,
        sb_pairs_by_player,
        _conv_segments_by_period,
        event_log_context,
    ) = _parse_event_log_layout(df)
    if not used_event_log:
        try:
            (
                _video_pairs_by_player,
                sb_pairs_by_player,
                _conv_segments_by_period,
                _validation_errors,
            ) = _parse_per_player_layout(df, keep_goalies=keep_goalies, skip_validation=True)
        except Exception:
            sb_pairs_by_player = {}

    jerseys: set[str] = set()
    for pk in sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(pk)
        if norm:
            jerseys.add(norm)

    if not jerseys and event_log_context is not None:
        for lst in (event_log_context.team_roster or {}).values():
            for num in lst or []:
                norm = _normalize_jersey_number(num)
                if norm:
                    jerseys.add(norm)

    # Long-only fallback: parse embedded long shift tables and/or long-sheet event table rosters.
    if not jerseys:
        try:
            # Embedded shift tables (team blocks with Jersey/Name columns).
            parsed = _parse_long_shift_tables(df)
            for _team, info in (parsed or {}).items():
                sb_any = (info or {}).get("sb_pairs_by_player") or {}
                for pk in sb_any.keys():
                    parts = _parse_player_key(pk)
                    norm = _normalize_jersey_number(parts.jersey)
                    if norm:
                        jerseys.add(norm)
        except Exception:
            pass
    if not jerseys:
        try:
            _events, _goal_rows, jerseys_by_team, _mapping = (
                _parse_long_left_event_table_with_mapping(df)
            )
            for nums in (jerseys_by_team or {}).values():
                for n in nums or set():
                    norm = _normalize_jersey_number(n)
                    if norm:
                        jerseys.add(norm)
        except Exception:
            pass

    return jerseys


# ----------------------------- parsing sheet -----------------------------


def find_period_blocks(df: pd.DataFrame) -> List[Tuple[int, int, int]]:
    """
    Returns a list of (period_number, start_row_idx, end_row_idx_exclusive) for each 'Period X' block.
    """
    # Use position-based indexing for robustness: some Excel files may not
    # create a column label "0" even with header=None, so df[0] can KeyError.
    col0 = df.iloc[:, 0]
    starts: List[int] = []
    periods: List[int] = []
    for i, v in col0.items():
        pnum = parse_period_label(v)
        if pnum is None:
            continue
        starts.append(i)
        periods.append(pnum)
    starts.append(len(df))
    blocks = []
    for i in range(len(periods)):
        blocks.append((periods[i], starts[i], starts[i + 1]))
    return blocks


def find_header_row(df: pd.DataFrame, start: int, end: int) -> Optional[int]:
    """
    Within [start, end), locate the header row (typically the row with 'Jersey No' in col 0).
    Often pattern is: 'Period', blank, header.
    """
    for r in range(start, min(end, start + 12)):
        row = df.iloc[r]
        for c in range(min(len(row), 20)):  # quick scan of first 20 cols
            val = row.iloc[c]
            if not isinstance(val, str):
                continue
            norm = _normalize_header_label(val)
            if norm in {"jerseyno", "jerseynumber"} or ("jersey" in norm and "number" in norm):
                return r
    # Fallback (Period + blank + header)
    return start + 2 if start + 2 < end else None


# ----------------------------- goals parsing -----------------------------


@dataclass
class GoalEvent:
    kind: str  # "GF" or "GA"
    period: int
    t_str: str
    video_t_str: Optional[str] = None
    scorer: Optional[str] = None
    assists: List[str] = field(default_factory=list)
    t_sec: int = field(init=False)
    video_t_sec: Optional[int] = field(init=False, default=None)
    is_game_tying: bool = False
    is_game_winning: bool = False

    def __post_init__(self) -> None:
        self.t_str = self.t_str.strip()
        if self.video_t_str is not None:
            self.video_t_str = str(self.video_t_str).strip()
            if self.video_t_str.lower() in {"nan", "nat", "none"}:
                self.video_t_str = None
        self.assists = [a for a in self.assists if a]
        self.t_sec = parse_flex_time_to_seconds(self.t_str)
        if self.video_t_str:
            self.video_t_sec = parse_flex_time_to_seconds(self.video_t_str)

    def __str__(self) -> str:  # preserve prior textual representation
        return f"{self.kind}:{self.period}/{self.t_str}"

    __repr__ = __str__


def parse_goal_token(token: str) -> GoalEvent:
    """
    Token: GF:2/13:45, GA:1/05:12, or GF:OT/0:45 (case-insensitive on GF/GA and OT).
    """
    token = token.strip()
    m = re.fullmatch(r"(?i)(GF|GA)\s*:\s*([^/]+)\s*/\s*([0-9:]+)", token)
    if not m:
        raise ValueError(f"Bad goal token '{token}'. Expected GF:period/time or GA:period/time")
    kind = m.group(1).upper()
    period = parse_period_token(m.group(2))
    if period is None:
        raise ValueError(
            f"Bad goal token '{token}': invalid period '{m.group(2)}' (use 1/2/3/4 or OT)."
        )
    t_str = m.group(3)
    return GoalEvent(kind, period, t_str)


def parse_period_token(x: object) -> Optional[int]:
    """
    Parse a period token from various sources:
      - ints ("1", 1)
      - labels ("Period 1", "1st Period")
      - overtime ("OT", "Overtime")

    Convention: OT == period 4.
    """
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, int):
        return x
    try:
        s = str(x).strip()
    except Exception:
        return None
    if not s:
        return None
    # Labels like "1st Period" / "OT"
    p = parse_period_label(s)
    if p is not None:
        return p
    # Plain number like "1"
    if re.fullmatch(r"\d+", s):
        try:
            return int(s)
        except Exception:
            return None
    # Fallback: any number embedded in the token
    m = re.search(r"(\d+)", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def _annotate_goal_roles(goals: List[GoalEvent]) -> None:
    """
    Annotate GoalEvent entries in-place with game-tying / game-winning flags.

    - Game-tying goal: GF that changes the score from trailing to tied.
    - Game-winning goal: when our team wins (GF > GA at end), the first goal
      scored by our team after the last time the game was tied.
    """
    if not goals:
        return

    # Work on indices into the original list to avoid reordering it.
    order = sorted(range(len(goals)), key=lambda idx: (goals[idx].period, goals[idx].t_sec))
    before_scores: List[Tuple[int, int]] = []
    after_scores: List[Tuple[int, int]] = []
    gf = 0
    ga = 0
    last_tie_pos = -1

    for pos, idx in enumerate(order):
        ev = goals[idx]
        before_scores.append((gf, ga))
        if ev.kind == "GF":
            gf += 1
        else:
            ga += 1
        after_scores.append((gf, ga))
        if gf == ga:
            last_tie_pos = pos

    final_gf, final_ga = gf, ga

    # Reset flags then mark game-tying goals.
    for pos, idx in enumerate(order):
        ev = goals[idx]
        ev.is_game_tying = False
        ev.is_game_winning = False
        bf_for, bf_against = before_scores[pos]
        af_for, af_against = after_scores[pos]
        if ev.kind == "GF" and bf_for < bf_against and af_for == af_against:
            ev.is_game_tying = True

    # Game-winning goal (if we finished ahead).
    if final_gf > final_ga and order:
        win_pos = last_tie_pos + 1
        if 0 <= win_pos < len(order):
            ev = goals[order[win_pos]]
            if ev.kind == "GF":
                ev.is_game_winning = True


def _normalize_jersey_number(token: Any) -> Optional[str]:
    if token is None:
        return None
    try:
        text = str(token).strip()
    except Exception:
        return None
    if not text:
        return None
    m = re.search(r"(\d+)", text)
    if not m:
        return None
    num = m.group(1).lstrip("0")
    return num or "0"


@dataclass(frozen=True)
class PlayerKeyParts:
    raw: str
    jersey: Optional[str]
    name: str


def _parse_player_key(raw: Any) -> PlayerKeyParts:
    """
    Parse internal player keys like:
      - '1_Ethan_L_Olivier'    -> jersey='1',  name='Ethan L Olivier'
      - '59_Ryan_S_Donahue'    -> jersey='59', name='Ryan S Donahue'
      - 'Blue_1'               -> jersey='1',  name='Blue'

    This is used for user-facing table formatting and for robust jersey extraction
    (avoid ad-hoc `split('_', 1)` parsing).
    """
    try:
        text = str(raw or "").strip()
    except Exception:
        text = ""
    if not text:
        return PlayerKeyParts(raw="", jersey=None, name="")

    parts = [p for p in text.split("_") if p != ""]
    if not parts:
        return PlayerKeyParts(raw=text, jersey=None, name=text.replace("_", " ").strip())

    jersey: Optional[str] = None
    name_parts: List[str] = []

    # Primary format: <jersey>_<first>_<...>_<last>
    jersey_norm = _normalize_jersey_number(parts[0])
    if jersey_norm and len(parts) > 1:
        jersey = jersey_norm
        name_parts = parts[1:]
    else:
        # Secondary: <team>_<jersey>_<optional name...>
        if len(parts) >= 2 and parts[0] in {"Blue", "White"}:
            jersey_norm2 = _normalize_jersey_number(parts[1])
            if jersey_norm2:
                jersey = jersey_norm2
                name_parts = parts[2:] if len(parts) > 2 else [parts[0]]
            else:
                name_parts = parts
        else:
            name_parts = parts

    name = " ".join([p for p in name_parts if p]).replace("_", " ").strip()
    if not name:
        name = text.replace("_", " ").strip()
    return PlayerKeyParts(raw=text, jersey=jersey, name=name)


def _player_sort_key(raw: Any) -> Tuple[str, int, str]:
    p = _parse_player_key(raw)
    try:
        j = int(p.jersey) if p.jersey is not None else 10**9
    except Exception:
        j = 10**9
    return (p.name.lower(), j, p.raw)


def _format_player_name_only(raw: Any) -> str:
    p = _parse_player_key(raw)
    return p.name


def _format_player_name_with_jersey(raw: Any) -> str:
    p = _parse_player_key(raw)
    if p.jersey:
        return f"{p.name} ({p.jersey})"
    return p.name


def _extract_jersey_number(cell: Any) -> Optional[str]:
    if isinstance(cell, dict):
        cell = cell.get("text") or cell.get("link") or ""
    return _normalize_jersey_number(cell)


def _scoring_numbers_from_row(row: Any) -> Tuple[Optional[str], List[str]]:
    if not isinstance(row, dict):
        return None, []
    scorer = _extract_jersey_number(row.get("goal"))
    assists: List[str] = []
    for key in ("assist1", "assist2"):
        num = _extract_jersey_number(row.get(key))
        if num:
            assists.append(num)
    return scorer, assists


def _infer_t2s_from_filename(path: Path) -> Optional[int]:
    stem = path.stem
    m = re.search(r"^(.*)-(\d+)$", stem)
    if not m:
        return None
    try:
        game_id = int(m.group(2))
    except Exception:
        return None
    # Treat only sufficiently large numeric suffixes as TimeToScore ids.
    # Smaller suffixes (e.g., 'chicago-1') are considered part of the game name.
    return game_id if game_id >= 10000 else None


def _base_label_from_path(path: Any) -> str:
    p = Path(path)
    stem = p.stem
    # If the filename encodes a T2S id as a trailing numeric suffix (>= 10000),
    # drop that suffix from the label. Otherwise keep the stem as-is.
    m = re.search(r"^(.*)-(\d+)$", stem)
    if m:
        try:
            suffix_num = int(m.group(2))
        except Exception:
            suffix_num = None
        if suffix_num is not None and suffix_num >= 10000:
            base = m.group(1) or stem
            # Treat companion long sheets (e.g., 'game-long-54111.xlsx') as belonging
            # to the same game label as their non-long counterpart.
            if base.endswith("-long"):
                base = base[: -len("-long")]
            return base or stem
    label = stem or p.name
    if label.endswith("-long"):
        label = label[: -len("-long")]
    return label


def _is_long_sheet_path(path: Path) -> bool:
    try:
        stem = Path(path).stem
    except Exception:
        return False
    return re.search(r"(?i)(?:^|-)long(?:-|$)", stem) is not None


def _select_tracking_output_video(video_dir: Path) -> Optional[Path]:
    """
    Choose a `tracking_output-with-audio*.mp4` file in `video_dir`.

    If any numbered `tracking_output-with-audio-<N>.mp4` files exist, return the one
    with the largest N. Otherwise fall back to `tracking_output-with-audio.mp4`.
    """
    try:
        video_dir = Path(video_dir)
    except Exception:
        return None

    best_num: Optional[int] = None
    best_path: Optional[Path] = None
    try:
        for p in video_dir.glob("tracking_output-with-audio-*.mp4"):
            m = re.match(r"^tracking_output-with-audio-(\d+)\.mp4$", p.name)
            if not m:
                continue
            try:
                n = int(m.group(1))
            except Exception:
                continue
            if best_num is None or n > best_num:
                best_num = n
                best_path = p
    except Exception:
        best_path = None

    if best_path is not None and best_path.exists():
        return best_path

    plain = video_dir / "tracking_output-with-audio.mp4"
    return plain if plain.exists() else None


def _find_tracking_output_video_for_sheet_path(sheet_path: Path) -> Optional[Path]:
    """
    Locate the tracking video for a sheet assumed to live under `<game_dir>/stats/`.
    """
    try:
        sheet_path = Path(sheet_path)
    except Exception:
        return None
    stats_dir = sheet_path.parent
    video_dir = stats_dir.parent if stats_dir.name.lower() == "stats" else stats_dir
    return _select_tracking_output_video(video_dir)


@dataclass(frozen=True)
class InputEntry:
    path: Optional[Path]
    side: Optional[str]
    t2s_id: Optional[int] = None
    label: Optional[str] = None
    meta: dict[str, str] = field(default_factory=dict)
    event_corrections: Any = None


def _parse_t2s_spec(token: Any) -> Optional[Tuple[int, Optional[str], Optional[str]]]:
    """
    Parse a TimeToScore spec token, formatted like:
      - 51602
      - 51602:HOME
      - 51602:HOME:stockton-r2
      - t2s=51602:HOME:stockton-r2

    Returns (t2s_id, side, label).
    """
    raw = str(token or "").strip()
    if not raw:
        return None

    m = re.match(r"(?i)^\s*(?:t2s\s*=\s*)?(\d+)(.*)$", raw)
    if not m:
        return None
    try:
        t2s_id = int(m.group(1))
    except Exception:
        return None

    rest = (m.group(2) or "").strip()
    if rest.startswith(":"):
        rest = rest[1:]
    rest_parts = [p.strip() for p in rest.split(":")] if rest else []

    side: Optional[str] = None
    label: Optional[str] = None
    if rest_parts:
        first = rest_parts[0].upper()
        if first in {"HOME", "AWAY"}:
            side = first.lower()
            label = ":".join(rest_parts[1:]).strip() if len(rest_parts) > 1 else None
        else:
            label = ":".join(rest_parts).strip()
    if label == "":
        label = None
    return t2s_id, side, label


def _parse_t2s_only_token(token: str) -> Optional[Tuple[int, Optional[str], Optional[str]]]:
    """
    Parse a special `--file-list` token representing a TimeToScore-only game
    (no spreadsheets present), formatted like:
      - t2s=51602
      - t2s=51602:HOME
      - t2s=51602:HOME:stockton-r2
    """
    raw = str(token or "").strip()
    if not re.match(r"(?i)^\s*t2s\s*=", raw):
        return None
    return _parse_t2s_spec(raw)


def _parse_input_token(token: str, base_dir: Optional[Path] = None) -> Tuple[Path, Optional[str]]:
    p, side, _meta = _parse_input_token_with_meta(token, base_dir=base_dir)
    return p, side


def _parse_input_token_with_meta(
    token: str, base_dir: Optional[Path] = None
) -> Tuple[Path, Optional[str], dict[str, str]]:
    """
    Parse a non-`t2s=...` input token with optional inline side and metadata:
      - /path/to/stats
      - /path/to/stats:HOME
      - /path/to/stats:AWAY
      - /path/to/stats:AWAY:home_team=Reston Renegades 12AA

    Inline `key=value` segments are primarily a convenience/back-compat for
    users who accidentally use ':' instead of '|key=value' in --file-list.
    """
    raw = token.strip()
    side: Optional[str] = None
    meta: dict[str, str] = {}
    if ":" in raw:
        parts = raw.split(":")
        # Parse from right-to-left:
        #   ...:<HOME|AWAY>[:key=value[:key=value...]]
        while len(parts) > 1 and "=" in parts[-1]:
            k, v = parts[-1].split("=", 1)
            kk = str(k or "").strip().lower()
            vv = str(v or "").strip()
            if kk and vv:
                meta[kk] = vv
            parts.pop()
        if len(parts) > 1 and parts[-1].upper() in {"HOME", "AWAY"}:
            side = parts[-1].lower()
            parts.pop()
        raw = ":".join(parts)
    p = Path(raw).expanduser()
    if base_dir and not p.is_absolute():
        p = (base_dir / p).resolve()
    return p, side, meta


def _load_input_entries_from_yaml_file_list(
    file_list_path: Path, *, base_dir: Path, use_t2s: bool
) -> List[InputEntry]:
    """
    YAML variant of `--file-list`.

    Supported shapes:
    - A dict with `games:` containing a list of entries (recommended)
    - A list of entries
      - Each entry may be a mapping (recommended), or
      - a legacy string line (back-compat; supports optional `|key=value` metadata).

    Entry formats:
    - string (legacy): "/path/to/stats:HOME | owner_email=... | league=..."
    - mapping:
        path: "/path/to/stats"              # required for spreadsheets
        side: HOME|AWAY                     # optional
        label: "my-game-label"              # optional
        # Metadata can be provided either:
        #   - as a nested `meta:` / `metadata:` mapping, or
        #   - as direct subkeys on the mapping (any key not in the reserved set).
        meta:
          home_team: "San Jose Jr Sharks 12AA-2"
          away_team: "Stockton Colts 12AA"
          date: "2025-12-07"
          time: "16:15"
          game_video: "https://youtu.be/..."
      or:
        t2s: 51602                          # TimeToScore-only (no spreadsheets)
        side: HOME|AWAY                     # optional
        label: "stockton-r2"                # optional
      or:
        label: "tv-12-1-r1"                 # required for `sheets:` entries
        shared_long_path: "/path/to/stats"  # optional; used to attach '*-long*' sheets to all sides
        metadata:
          game_video: "https://youtu.be/..."
        sheets:
          - side: AWAY
            path: "/path/to/away/stats"
          - side: HOME
            path: "/path/to/home/stats"
    """

    raw = file_list_path.read_text(encoding="utf-8", errors="ignore").lstrip("\ufeff")
    data = yaml.safe_load(raw) if raw.strip() else None
    if data is None:
        return []

    entries: Any = data
    if isinstance(data, dict):
        entries = data.get("games") or data.get("entries") or []

    if not isinstance(entries, list):
        raise ValueError("YAML file-list must be a list or a dict containing a 'games' list")

    warned_pipe_meta = False

    def _merge_meta(a: dict[str, str], b: dict[str, str]) -> dict[str, str]:
        out = dict(a or {})
        for k, v in (b or {}).items():
            kk = str(k or "").strip().lower()
            vv = str(v or "").strip()
            if not kk or not vv:
                continue
            if kk in out and str(out[kk]) != vv:
                raise ValueError(f"conflicting metadata for key '{kk}': {out[kk]!r} vs {vv!r}")
            out[kk] = vv
        return out

    def _parse_side(raw_side: Any) -> Optional[str]:
        s = str(raw_side or "").strip().upper()
        if s in {"HOME", "AWAY"}:
            return s.lower()
        return None

    out_entries: List[InputEntry] = []
    for idx, item in enumerate(entries):
        if isinstance(item, str):
            line = item.strip().lstrip("\ufeff")
            if not line or line.startswith("#"):
                continue
            if not warned_pipe_meta and "|" in line:
                warned_pipe_meta = True
                print(
                    "[warn] YAML file-list uses legacy `|key=value` metadata separators; prefer mapping entries with subkeys.",
                    file=sys.stderr,
                )
            parts = [p.strip() for p in str(line).split("|") if p.strip()]
            token = parts[0] if parts else ""
            meta: dict[str, str] = {}
            for seg in parts[1:]:
                if "=" not in seg:
                    continue
                k, v = seg.split("=", 1)
                kk = str(k or "").strip().lower()
                vv = str(v or "").strip()
                if kk and vv:
                    meta[kk] = vv

            t2s_only = _parse_t2s_only_token(token)
            if t2s_only is not None:
                if not use_t2s:
                    print(f"[no-time2score] Skipping file-list entry: {token}", file=sys.stderr)
                    continue
                t2s_id, side, label = t2s_only
                out_entries.append(
                    InputEntry(path=None, side=side, t2s_id=t2s_id, label=label, meta=meta)
                )
                continue

            p, side, inline_meta = _parse_input_token_with_meta(token, base_dir=base_dir)
            out_entries.append(InputEntry(path=p, side=side, meta=_merge_meta(meta, inline_meta)))
            continue

        if not isinstance(item, dict):
            raise ValueError(f"YAML games[{idx}] must be a string or mapping")

        # Allow top-level key/value pairs as metadata (in addition to an explicit `meta:` map).
        reserved = {
            "token",
            "path",
            "file",
            "dir",
            "t2s",
            "timetoscore_game_id",
            "side",
            "label",
            "meta",
            "metadata",
            "sheets",
            "shared_long_path",
            "event_corrections",
            "event_correction",
        }
        meta: dict[str, str] = {}
        if isinstance(item.get("meta"), dict):
            meta = _merge_meta(
                meta, {str(k): str(v) for k, v in item["meta"].items() if v is not None}
            )
        if isinstance(item.get("metadata"), dict):
            meta = _merge_meta(
                meta,
                {str(k): str(v) for k, v in item["metadata"].items() if v is not None},
            )
        for k, v in item.items():
            kk = str(k or "").strip()
            if not kk or kk in reserved:
                continue
            if v is None:
                continue
            meta[str(kk).strip().lower()] = str(v).strip()

        event_corrections = item.get("event_corrections")
        if event_corrections is None:
            event_corrections = item.get("event_correction")
        if event_corrections is not None and not isinstance(event_corrections, (list, dict)):
            raise ValueError(f"invalid event_corrections at games[{idx}] (must be list or mapping)")

        if item.get("token"):
            token = str(item.get("token") or "").strip()
            label_key = str(item.get("label") or "").strip() or None
            if not warned_pipe_meta and "|" in token:
                warned_pipe_meta = True
                print(
                    "[warn] YAML file-list uses legacy `token: '... |key=value'` metadata separators; prefer mapping entries with subkeys.",
                    file=sys.stderr,
                )
            parts = [p.strip() for p in token.split("|") if p.strip()]
            token0 = parts[0] if parts else ""
            meta_inline: dict[str, str] = {}
            for seg in parts[1:]:
                if "=" not in seg:
                    continue
                k, v = seg.split("=", 1)
                kk = str(k or "").strip().lower()
                vv = str(v or "").strip()
                if kk and vv:
                    meta_inline[kk] = vv
            meta = _merge_meta(meta, meta_inline)

            t2s_only = _parse_t2s_only_token(token0)
            if t2s_only is not None:
                if not use_t2s:
                    print(f"[no-time2score] Skipping file-list entry: {token0}", file=sys.stderr)
                    continue
                t2s_id, side, label = t2s_only
                out_entries.append(
                    InputEntry(
                        path=None,
                        side=side,
                        t2s_id=t2s_id,
                        label=label,
                        meta=meta,
                        event_corrections=event_corrections,
                    )
                )
                continue

            p, side, inline_meta = _parse_input_token_with_meta(token0, base_dir=base_dir)
            out_entries.append(
                InputEntry(
                    path=p,
                    side=side,
                    label=label_key,
                    meta=_merge_meta(meta, inline_meta),
                    event_corrections=event_corrections,
                )
            )
            continue

        if item.get("t2s") is not None or item.get("timetoscore_game_id") is not None:
            if not use_t2s:
                print(f"[no-time2score] Skipping YAML t2s entry at games[{idx}]", file=sys.stderr)
                continue
            t2s_raw = (
                item.get("t2s") if item.get("t2s") is not None else item.get("timetoscore_game_id")
            )
            try:
                t2s_id = int(t2s_raw)
            except Exception:
                raise ValueError(f"invalid t2s id at games[{idx}]: {t2s_raw!r}") from None
            side = _parse_side(item.get("side"))
            label = str(item.get("label") or "").strip() or None
            out_entries.append(
                InputEntry(
                    path=None,
                    side=side,
                    t2s_id=int(t2s_id),
                    label=label,
                    meta=meta,
                    event_corrections=event_corrections,
                )
            )
            continue

        if item.get("sheets") is not None:
            label_key = str(item.get("label") or "").strip() or None
            if not label_key:
                raise ValueError(f"missing 'label' for games[{idx}] (required when using 'sheets')")
            sheets_raw = item.get("sheets")
            if not isinstance(sheets_raw, list):
                raise ValueError(f"games[{idx}] 'sheets' must be a list")

            shared_long_token = str(item.get("shared_long_path") or "").strip() or None
            if shared_long_token:
                shared_long_path = Path(shared_long_token).expanduser()
                if not shared_long_path.is_absolute():
                    shared_long_path = (base_dir / shared_long_path).resolve()
                if shared_long_path.is_dir():
                    for cand in _discover_spreadsheet_inputs_in_dir(shared_long_path):
                        if not _is_long_sheet_path(cand):
                            continue
                        if _base_label_from_path(cand) != str(label_key):
                            continue
                        out_entries.append(
                            InputEntry(
                                path=cand,
                                side=None,
                                label=label_key,
                                meta=dict(meta),
                                event_corrections=event_corrections,
                            )
                        )
                elif shared_long_path.is_file():
                    if not _is_long_sheet_path(shared_long_path):
                        raise ValueError(
                            f"games[{idx}] shared_long_path must point to a '*-long*' sheet: {shared_long_path}"
                        )
                    out_entries.append(
                        InputEntry(
                            path=shared_long_path,
                            side=None,
                            label=label_key,
                            meta=dict(meta),
                            event_corrections=event_corrections,
                        )
                    )
                else:
                    raise ValueError(
                        f"games[{idx}] shared_long_path does not exist: {shared_long_path}"
                    )

            for sidx, s in enumerate(sheets_raw):
                if isinstance(s, str):
                    token = str(s).strip()
                    if not token:
                        continue
                    p, side, inline_meta = _parse_input_token_with_meta(token, base_dir=base_dir)
                    out_entries.append(
                        InputEntry(
                            path=p,
                            side=side,
                            label=label_key,
                            meta=_merge_meta(meta, inline_meta),
                            event_corrections=event_corrections,
                        )
                    )
                    continue

                if not isinstance(s, dict):
                    raise ValueError(f"games[{idx}].sheets[{sidx}] must be a string or mapping")
                sheet_meta = dict(meta)
                if isinstance(s.get("meta"), dict):
                    sheet_meta = _merge_meta(
                        sheet_meta, {str(k): str(v) for k, v in s["meta"].items() if v is not None}
                    )
                if isinstance(s.get("metadata"), dict):
                    sheet_meta = _merge_meta(
                        sheet_meta,
                        {str(k): str(v) for k, v in s["metadata"].items() if v is not None},
                    )

                path_token = str(s.get("path") or s.get("file") or s.get("dir") or "").strip()
                if not path_token:
                    raise ValueError(f"missing 'path'/'file' for games[{idx}].sheets[{sidx}]")
                side_hint = _parse_side(s.get("side"))
                if side_hint and not re.search(r"(?i):(?:home|away)(?::|$)", path_token):
                    path_token = f"{path_token}:{side_hint.upper()}"
                p, side, inline_meta = _parse_input_token_with_meta(path_token, base_dir=base_dir)
                out_entries.append(
                    InputEntry(
                        path=p,
                        side=side or side_hint,
                        label=label_key,
                        meta=_merge_meta(sheet_meta, inline_meta),
                        event_corrections=event_corrections,
                    )
                )
            continue

        path_token = str(item.get("path") or item.get("file") or item.get("dir") or "").strip()
        if not path_token:
            raise ValueError(f"missing 'path'/'file' for games[{idx}]")
        side_hint = _parse_side(item.get("side"))
        label_key = str(item.get("label") or "").strip() or None
        if side_hint and not re.search(r"(?i):(?:home|away)(?::|$)", path_token):
            path_token = f"{path_token}:{side_hint.upper()}"
        p, side, inline_meta = _parse_input_token_with_meta(path_token, base_dir=base_dir)
        out_entries.append(
            InputEntry(
                path=p,
                side=side or side_hint,
                label=label_key,
                meta=_merge_meta(meta, inline_meta),
                event_corrections=event_corrections,
            )
        )

    return out_entries


def _is_spreadsheet_input_path(path: Path) -> bool:
    try:
        return path.is_file() and path.suffix.lower() in {".xls", ".xlsx"}
    except Exception:
        return False


def _should_ignore_spreadsheet_input(path: Path) -> bool:
    name = path.name
    if not name:
        return True
    if name.startswith(".") or name.startswith("~$"):
        return True
    stem = path.stem.lower()
    if stem == "goals":
        return True
    if stem.startswith("player_stats"):
        return True
    return False


def _discover_spreadsheet_inputs_in_dir(dir_path: Path) -> List[Path]:
    try:
        paths = [p for p in dir_path.iterdir() if _is_spreadsheet_input_path(p)]
    except Exception:
        return []
    out: List[Path] = []
    for p in paths:
        if _should_ignore_spreadsheet_input(p):
            continue
        out.append(p)
    out.sort(key=lambda p: p.name.lower())
    return out


def _expand_dir_input_to_game_sheets(
    dir_path: Path, *, ignore_primary: bool = False, ignore_long: bool = False
) -> List[Path]:
    """
    Expand a directory passed to --input into the game sheet(s) inside:
      - exactly one non-'-long' shift sheet (primary)
      - zero or more companion '*-long*' sheets

    Also supports passing a game directory by checking `<dir>/stats/` as a fallback.
    """
    candidates = _discover_spreadsheet_inputs_in_dir(dir_path)
    stats_dir = dir_path / "stats"
    if not candidates and stats_dir.is_dir():
        candidates = _discover_spreadsheet_inputs_in_dir(stats_dir)
        dir_path = stats_dir

    if not candidates:
        # Goals-only games: allow a directory that contains only goals.xlsx.
        goals_xlsx = dir_path / "goals.xlsx"
        if not goals_xlsx.exists() and stats_dir.is_dir():
            goals_xlsx = stats_dir / "goals.xlsx"
            if goals_xlsx.exists():
                dir_path = stats_dir
        if goals_xlsx.exists():
            return [goals_xlsx]
        raise ValueError(f"No input .xls/.xlsx sheets found in {dir_path}")

    # Directory inputs are expected to correspond to a single game label.
    by_label: Dict[str, List[Path]] = {}
    for p in candidates:
        by_label.setdefault(_base_label_from_path(p), []).append(p)
    if len(by_label) != 1:
        labels = ", ".join(sorted(by_label.keys()))
        raise ValueError(
            f"Directory {dir_path} contains multiple game labels ({labels}); "
            "pass a specific file path or use --file-list."
        )

    only_label = next(iter(by_label.keys()))
    paths = by_label[only_label]
    primaries = [p for p in paths if not _is_long_sheet_path(p)]
    long_paths = [p for p in paths if _is_long_sheet_path(p)]
    if ignore_primary and ignore_long:
        raise ValueError("cannot combine --ignore-primary and --ignore-long")

    # Sheet selection rules:
    # - Default: require exactly one primary, plus any companion '*-long*' sheets.
    # - --ignore-primary: prefer long sheets when present; fall back to primary-only.
    # - --ignore-long: prefer primary sheet when present; fall back to long-only.
    if ignore_primary:
        if long_paths:
            return long_paths
        if len(primaries) != 1:
            raise ValueError(
                f"Directory {dir_path} has {len(primaries)} primary sheets for {only_label}; "
                f"expected exactly 1: {[p.name for p in primaries]}"
            )
        return [primaries[0]]
    if ignore_long:
        if primaries:
            if len(primaries) != 1:
                raise ValueError(
                    f"Directory {dir_path} has {len(primaries)} primary sheets for {only_label}; "
                    f"expected exactly 1: {[p.name for p in primaries]}"
                )
            return [primaries[0]]
        return long_paths

    # Long-only games: allow processing from the embedded long-sheet shift tables.
    if not primaries and long_paths:
        return long_paths
    if len(primaries) != 1:
        raise ValueError(
            f"Directory {dir_path} has {len(primaries)} primary sheets for {only_label}; "
            f"expected exactly 1: {[p.name for p in primaries]}"
        )
    return [primaries[0]] + long_paths


def _label_from_goals_xlsx_path(p: Path) -> str:
    """
    Choose a game label for a goals-only goals.xlsx path.

    Common layouts:
      - /path/to/<game>/stats/goals.xlsx -> label '<game>'
      - /path/to/<game>/goals.xlsx       -> label '<game>'
    """
    try:
        if p.parent.name.lower() == "stats":
            return p.parent.parent.name
        return p.parent.name
    except Exception:
        return p.stem or "goals"


def load_goals(goals_inline: Iterable[str], goals_file: Optional[Path]) -> List[GoalEvent]:
    events: List[GoalEvent] = []
    # Inline
    for tok in goals_inline or []:
        if not tok:
            continue
        events.append(parse_goal_token(tok))
    # File
    if goals_file:
        with goals_file.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                # Strip UTF-8 BOM if present (common when files are created on Windows).
                line = line.lstrip("\ufeff")
                if not line or line.startswith("#"):
                    continue
                events.append(parse_goal_token(line))
    return events


def _format_goal_time_cell(val: Any) -> Optional[str]:
    """
    Convert a goal time cell (often an Excel time) into an M:SS-style string
    compatible with parse_flex_time_to_seconds.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime.time):
        # Excel stores times as H:M; reuse the same formatting as shift sheets
        return val.strftime("%H:%M")
    if isinstance(val, pd.Timestamp):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if not s:
        return None
    parts = s.split(":")
    # Convert H:MM:SS (Excel-style) into M:SS with minutes = H*60 + MM
    if len(parts) == 3:
        try:
            h = int(parts[0])
            m = int(parts[1])
            sec = int(parts[2])
            total_min = h * 60 + m
            return f"{total_min}:{sec:02d}"
        except Exception:
            pass
    return s


def _format_goal_video_time_cell(val: Any) -> Optional[str]:
    """
    Convert a goal video time cell (often an Excel time) into a string compatible with
    parse_flex_time_to_seconds.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, pd.Timestamp):
        return val.strftime("%H:%M:%S")
    s = str(val).strip()
    return s or None


def _goals_from_goals_xlsx(
    goals_xlsx: Path,
    *,
    our_jerseys: Optional[set[str]] = None,
    our_team_name: Optional[str] = None,
) -> List[GoalEvent]:
    """
    Parse goals from a goals.xlsx sheet that has side-by-side GF / GA tables.

    Layout example (no headers row index assumed):

        row 0:  'GF'  ...  'GA' ...
        row 1:  (blank)
        row 2:  'Period' 'Time' 'Goal' 'Assist 1' 'Assist 2' ... 'Period' 'Time' 'Goal' 'Assist 1' 'Assist 2'
        row 3+: data rows
    """
    if not goals_xlsx.exists():
        return []

    df = pd.read_excel(goals_xlsx, header=None)
    nrows, ncols = df.shape

    def _find_label(label: str) -> Optional[Tuple[int, int]]:
        target = _normalize_header_label(label)
        for r in range(nrows):
            for c in range(ncols):
                val = df.iat[r, c]
                if pd.isna(val):
                    continue
                if _normalize_header_label(str(val)) == target:
                    return r, c
        return None

    def _parse_table(label_rc: Tuple[int, int], kind: str) -> List[GoalEvent]:
        base_r, base_c = label_rc
        # Find header row: look a few rows below the label for a 'Period' cell
        header_r: Optional[int] = None
        for r in range(base_r + 1, min(nrows, base_r + 8)):
            for c in range(base_c, ncols):
                val = df.iat[r, c]
                if pd.isna(val):
                    continue
                if _normalize_header_label(str(val)) == "period":
                    header_r = r
                    break
            if header_r is not None:
                break
        if header_r is None:
            return []

        # Map header labels to columns starting from the table's base column
        label_to_col: Dict[str, int] = {}
        for c in range(base_c, ncols):
            val = df.iat[header_r, c]
            if pd.isna(val):
                continue
            key = _normalize_header_label(str(val))
            if key:
                label_to_col.setdefault(key, c)

        period_col = label_to_col.get("period")
        time_col = label_to_col.get("time")
        video_time_col = label_to_col.get("videotime")
        goal_col = label_to_col.get("goal")
        assist1_col = (
            label_to_col.get("assist1")
            or label_to_col.get("assist")
            or label_to_col.get("assist_1")
        )
        assist2_col = label_to_col.get("assist2") or label_to_col.get("assist_2")

        if period_col is None or time_col is None or goal_col is None:
            return []

        events: List[GoalEvent] = []
        for r in range(header_r + 1, nrows):
            period_val = df.iat[r, period_col]
            time_val = df.iat[r, time_col]
            goal_val = df.iat[r, goal_col]
            if all(pd.isna(x) for x in (period_val, time_val, goal_val)):
                continue
            period = parse_period_token(period_val)
            if period is None:
                continue
            t_str = _format_goal_time_cell(time_val)
            if not t_str:
                continue
            v_str = None
            if video_time_col is not None:
                v_str = _format_goal_video_time_cell(df.iat[r, video_time_col])
            scorer = _extract_jersey_number(goal_val)
            assists: List[str] = []
            for col in (assist1_col, assist2_col):
                if col is None:
                    continue
                val = df.iat[r, col]
                num = _extract_jersey_number(val)
                if num:
                    assists.append(num)
            events.append(
                GoalEvent(kind, period, t_str, video_t_str=v_str, scorer=scorer, assists=assists)
            )
        return events

    events: List[GoalEvent] = []
    gf_rc = _find_label("GF")
    ga_rc = _find_label("GA")
    if gf_rc:
        events.extend(_parse_table(gf_rc, "GF"))
    if ga_rc:
        events.extend(_parse_table(ga_rc, "GA"))
    if events:
        events.sort(key=lambda e: (e.period, e.t_sec))
        return events

    def _scan_team_tables() -> List[Tuple[Optional[str], List[GoalEvent]]]:
        """
        Parse a goals.xlsx variant where each table is labeled by team name above it,
        rather than "GF"/"GA", and may include a "Video Time" column.
        """

        def _infer_team_label(header_r: int, start_c: int, end_c: int) -> Optional[str]:
            header_norm = {"period", "time", "goal", "assist", "assist1", "assist2", "videotime"}
            for rr in range(header_r - 1, max(-1, header_r - 6), -1):
                for cc in range(start_c, end_c + 1):
                    v = df.iat[rr, cc]
                    if pd.isna(v):
                        continue
                    s = str(v).strip()
                    if not s:
                        continue
                    if _normalize_header_label(s) in header_norm:
                        continue
                    return s
            return None

        results: List[Tuple[Optional[str], List[GoalEvent]]] = []
        for header_r in range(nrows):
            period_cols = [
                c
                for c in range(ncols)
                if (not pd.isna(df.iat[header_r, c]))
                and _normalize_header_label(str(df.iat[header_r, c])) == "period"
            ]
            if not period_cols:
                continue
            period_cols = sorted(period_cols)
            for idx, start_c in enumerate(period_cols):
                next_c = period_cols[idx + 1] if idx + 1 < len(period_cols) else ncols
                label_to_col: Dict[str, int] = {}
                for c in range(start_c, next_c):
                    val = df.iat[header_r, c]
                    if pd.isna(val):
                        continue
                    key = _normalize_header_label(str(val))
                    if key:
                        label_to_col.setdefault(key, c)

                period_col = label_to_col.get("period")
                time_col = (
                    label_to_col.get("scoreboard")
                    or label_to_col.get("scoreboardtime")
                    or label_to_col.get("time")
                )
                video_time_col = label_to_col.get("videotime")
                goal_col = label_to_col.get("goal") or label_to_col.get("scorer")
                assist1_col = (
                    label_to_col.get("assist1")
                    or label_to_col.get("assist")
                    or label_to_col.get("assist_1")
                )
                assist2_col = label_to_col.get("assist2") or label_to_col.get("assist_2")

                if period_col is None or time_col is None or goal_col is None:
                    continue

                team_label = _infer_team_label(header_r, start_c, max(start_c, next_c - 1))
                table_events: List[GoalEvent] = []
                for r in range(header_r + 1, nrows):
                    period_val = df.iat[r, period_col]
                    time_val = df.iat[r, time_col]
                    goal_val = df.iat[r, goal_col]
                    if all(pd.isna(x) for x in (period_val, time_val, goal_val)):
                        continue
                    if not pd.isna(period_val) and _normalize_header_label(str(period_val)) in {
                        "period",
                        "roster",
                    }:
                        break
                    period = parse_period_token(period_val)
                    if period is None:
                        continue
                    t_str = _format_goal_time_cell(time_val)
                    if not t_str:
                        continue
                    v_str = None
                    if video_time_col is not None:
                        v_str = _format_goal_video_time_cell(df.iat[r, video_time_col])
                    scorer = _extract_jersey_number(goal_val)
                    assists: List[str] = []
                    for col in (assist1_col, assist2_col):
                        if col is None:
                            continue
                        num = _extract_jersey_number(df.iat[r, col])
                        if num:
                            assists.append(num)
                    table_events.append(
                        GoalEvent(
                            "GF",
                            period,
                            t_str,
                            video_t_str=v_str,
                            scorer=scorer,
                            assists=assists,
                        )
                    )
                if table_events:
                    results.append((team_label, table_events))
        return results

    tables = _scan_team_tables()
    if not tables:
        return []

    our_jerseys_norm = {str(x).strip() for x in (our_jerseys or set()) if str(x).strip()}
    our_team_norm = _normalize_header_label(str(our_team_name or ""))

    def _table_overlap(events: List[GoalEvent]) -> int:
        if not our_jerseys_norm or not events:
            return 0
        jerseys: set[str] = set()
        for ev in events:
            if ev.scorer:
                jerseys.add(_normalize_jersey_number(ev.scorer) or "")
            for a in ev.assists or []:
                if a:
                    jerseys.add(_normalize_jersey_number(a) or "")
        jerseys.discard("")
        return len(jerseys & our_jerseys_norm)

    overlaps = [_table_overlap(evts) for _lbl, evts in tables]
    gf_idx: Optional[int] = None
    if our_team_norm:
        matches = []
        for idx, (lbl, _evts) in enumerate(tables):
            lbl_norm = _normalize_header_label(str(lbl or ""))
            if not lbl_norm:
                continue
            if our_team_norm == lbl_norm or our_team_norm in lbl_norm or lbl_norm in our_team_norm:
                matches.append(idx)
        if len(matches) == 1:
            gf_idx = matches[0]
    if gf_idx is None and overlaps and any(o > 0 for o in overlaps):
        best = max(overlaps)
        if overlaps.count(best) == 1:
            gf_idx = overlaps.index(best)
    if gf_idx is None:
        gf_idx = 0

    out_events: List[GoalEvent] = []
    for idx, (_lbl, evts) in enumerate(tables):
        kind = "GF" if idx == gf_idx or len(tables) == 1 else "GA"
        for ev in evts:
            ev.kind = kind
            out_events.append(ev)

    out_events.sort(key=lambda e: (e.period, e.t_sec))
    return out_events


def _rosters_from_goals_xlsx(
    goals_xlsx: Path,
) -> List[Tuple[Optional[str], List[Dict[str, Any]]]]:
    """
    Best-effort roster extraction from goals.xlsx (team-table variant).

    This supports goals.xlsx layouts where the two teams are presented side-by-side, and each side
    optionally contains a "Roster" section below the goal table with columns like:
      - "#", "Name", "Pos"

    Returns a list of (team_label_or_none, roster_records) entries where roster_records contains
    dicts with keys: {jersey_number, name, position?}.
    """
    if not goals_xlsx.exists():
        return []

    df = pd.read_excel(goals_xlsx, header=None)
    nrows, ncols = df.shape

    def _norm_roster_header(v: Any) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        try:
            s = str(v).strip().lower()
        except Exception:
            return ""
        if not s:
            return ""
        # Preserve '#' for roster headers.
        if s == "#":
            return "#"
        s = s.replace("\xa0", " ")
        s = " ".join(s.split())
        return re.sub(r"[^a-z0-9#]+", "", s)

    def _infer_team_label(anchor_r: int, start_c: int, end_c: int) -> Optional[str]:
        skip_norm = {
            "period",
            "time",
            "scoreboard",
            "scoreboardtime",
            "videotime",
            "goal",
            "scorer",
            "assist",
            "assist1",
            "assist2",
            "roster",
            "name",
            "player",
            "playername",
            "pos",
            "position",
            "#",
            "number",
            "num",
        }
        for rr in range(anchor_r - 1, max(-1, anchor_r - 12), -1):
            for cc in range(start_c, end_c + 1):
                v = df.iat[rr, cc]
                if pd.isna(v):
                    continue
                s = str(v).strip()
                if not s:
                    continue
                if s.lower().startswith(("http://", "https://")):
                    continue
                if _normalize_header_label(s) in skip_norm or _norm_roster_header(s) in skip_norm:
                    continue
                return s
        return None

    # Identify team blocks by the first row that looks like a goals-table header row.
    blocks: List[Tuple[int, int, int]] = []
    for r in range(nrows):
        norm_row = []
        for c in range(ncols):
            v = df.iat[r, c]
            if pd.isna(v):
                norm_row.append("")
            else:
                norm_row.append(_normalize_header_label(str(v)))

        period_cols = [c for c, key in enumerate(norm_row) if key == "period"]
        if not period_cols:
            continue
        period_cols = sorted(period_cols)
        for idx, start_c in enumerate(period_cols):
            next_c = period_cols[idx + 1] if idx + 1 < len(period_cols) else ncols
            has_time = any(
                norm_row[c] in {"time", "scoreboard", "scoreboardtime"}
                for c in range(start_c, next_c)
            )
            if not has_time:
                continue
            blocks.append((r, start_c, next_c - 1))
        if blocks:
            break

    if not blocks:
        # Fallback: scan the full sheet for a roster table.
        blocks = [(0, 0, max(0, ncols - 1))]

    out: List[Tuple[Optional[str], List[Dict[str, Any]]]] = []
    for header_r, start_c, end_c in blocks:
        team_label = _infer_team_label(header_r, start_c, end_c)

        roster_label_r: Optional[int] = None
        for rr in range(header_r + 1, nrows):
            found = False
            for cc in range(start_c, end_c + 1):
                v = df.iat[rr, cc]
                if pd.isna(v):
                    continue
                if _normalize_header_label(str(v)) == "roster":
                    roster_label_r = rr
                    found = True
                    break
            if found:
                break
        if roster_label_r is None:
            continue

        roster_header_r: Optional[int] = None
        jersey_col: Optional[int] = None
        name_col: Optional[int] = None
        pos_col: Optional[int] = None
        for rr in range(roster_label_r + 1, min(nrows, roster_label_r + 8)):
            normed = [_norm_roster_header(df.iat[rr, c]) for c in range(start_c, end_c + 1)]
            if not any(normed):
                continue
            jersey_cols = [
                cidx
                for cidx, s in enumerate(normed)
                if s and (("jersey" in s) or s in {"number", "num", "#"})
            ]
            name_cols = [
                cidx
                for cidx, s in enumerate(normed)
                if s and (("name" in s) or s in {"player", "playername"})
            ]
            pos_cols = [
                cidx
                for cidx, s in enumerate(normed)
                if s and (s in {"pos", "position"} or "position" in s)
            ]
            if not jersey_cols or not name_cols:
                continue

            best = None
            for jc in jersey_cols:
                for nc in name_cols:
                    dist = abs(jc - nc)
                    if best is None or dist < best[0]:
                        best = (dist, jc, nc)
            if best is None:
                continue
            _, jc, nc = best
            roster_header_r = rr
            jersey_col = start_c + jc
            name_col = start_c + nc
            pos_col = (start_c + pos_cols[0]) if pos_cols else None
            break

        if roster_header_r is None or jersey_col is None or name_col is None:
            continue

        roster: List[Dict[str, Any]] = []
        seen_jerseys: set[str] = set()
        blank_streak = 0
        for rr in range(roster_header_r + 1, nrows):
            jv = df.iat[rr, jersey_col] if jersey_col < ncols else None
            nv = df.iat[rr, name_col] if name_col < ncols else None

            j_blank = jv is None or (isinstance(jv, float) and pd.isna(jv)) or not str(jv).strip()
            n_blank = nv is None or (isinstance(nv, float) and pd.isna(nv)) or not str(nv).strip()
            if j_blank and n_blank:
                blank_streak += 1
                if blank_streak >= 3:
                    break
                continue
            blank_streak = 0

            # Stop if we hit another "Roster" marker.
            if isinstance(jv, str) and _normalize_header_label(jv) == "roster":
                break

            jersey_norm = _normalize_jersey_number(jv)
            name = str(nv).strip() if not n_blank else ""
            if not jersey_norm or not name:
                continue
            if jersey_norm in seen_jerseys:
                continue
            seen_jerseys.add(jersey_norm)

            rec: Dict[str, Any] = {"jersey_number": jersey_norm, "name": name}
            if pos_col is not None and pos_col < ncols:
                pv = df.iat[rr, pos_col]
                if pv is not None and not (isinstance(pv, float) and pd.isna(pv)):
                    ps = str(pv).strip()
                    if ps:
                        rec["position"] = ps
            roster.append(rec)

        if not roster:
            continue
        if team_label is None:
            team_label = _infer_team_label(roster_header_r, start_c, end_c)
        out.append((team_label, roster))

    return out


def goals_from_t2s(
    game_id: int,
    *,
    side: str,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
) -> List[GoalEvent]:
    """
    Retrieve goals from TimeToScore for a game id and map them to GF/GA based on
    the selected side (home/away).
    """
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        details = f": {_t2s_api_import_error}" if _t2s_api_import_error else ""
        raise RuntimeError(
            f"TimeToScore API not available (failed to import hmlib.time2score.api){details}"
        )

    if allow_remote and not allow_full_sync:
        _log_t2s_scrape(int(game_id), "game details (scoring)")
    try:
        info = t2s_api.get_game_details(
            game_id,
            sync_if_missing=bool(allow_full_sync) if allow_remote else False,
            fetch_stats_if_missing=bool(allow_remote),
        )
    except TypeError:
        # Tests may stub a minimal TimeToScore API that doesn't accept the optional kwargs.
        info = t2s_api.get_game_details(game_id)
    if not isinstance(info, dict):
        raise RuntimeError(
            f"TimeToScore API returned invalid game details type for game {game_id}: {type(info)}"
        )
    stats = info.get("stats")
    if not isinstance(stats, dict) or not stats:
        keys = ", ".join(sorted(list(info.keys()))) if isinstance(info, dict) else ""
        suffix = ""
        if not allow_remote:
            suffix = " (cache-only; run import_time2score or omit --t2s-cache-only)"
        elif not allow_full_sync:
            suffix = " (scrape-only; run import_time2score or omit --t2s-scrape-only)"
        raise RuntimeError(
            f"TimeToScore returned no usable stats for game {game_id} (keys=[{keys}]); cannot compute goals.{suffix}"
        )

    home_sc = stats.get("homeScoring") or []
    away_sc = stats.get("awayScoring") or []

    events: List[GoalEvent] = []

    def _mk_event(kind: str, row: Any) -> Optional[GoalEvent]:
        if not isinstance(row, dict):
            return None
        period_val = row.get("period")
        time_val = row.get("time")
        if period_val is None or time_val is None:
            return None
        per = parse_period_token(period_val)
        if per is None:
            return None
        t_str = str(time_val).strip()
        # Normalize fractional seconds by flooring
        try:
            sec = parse_flex_time_to_seconds(t_str)
        except Exception:
            return None
        mm = sec // 60
        ss = sec % 60
        scorer_num, assist_nums = _scoring_numbers_from_row(row)
        return GoalEvent(kind, per, f"{mm}:{ss:02d}", scorer=scorer_num, assists=assist_nums)

    # Home goals are GF if side == home else GA
    for row in home_sc:
        ev = _mk_event("GF" if side == "home" else "GA", row)
        if ev:
            events.append(ev)
    # Away goals are GF if side == away else GA
    for row in away_sc:
        ev = _mk_event("GF" if side == "away" else "GA", row)
        if ev:
            events.append(ev)

    # Sort by period then time for determinism
    events.sort(key=lambda e: (e.period, e.t_sec))
    return events


# ----------------------------- core processing -----------------------------


def compute_interval_seconds(a: str, b: str) -> Tuple[int, int]:
    """
    Return (lo, hi) in seconds for a scoreboard interval defined by two times (order-agnostic).
    Works even if the scoreboard counts down (e.g., 15:00 -> 09:21) or up (e.g., 11:27 -> 29:54).
    """
    sa, sb = parse_flex_time_to_seconds(a), parse_flex_time_to_seconds(b)
    return (sa, sb) if sa <= sb else (sb, sa)


def interval_contains(t: int, lo: int, hi: int) -> bool:
    return lo <= t <= hi


def fmt_pairs_for_file(pairs: List[Tuple[str, str]]) -> str:
    return "\n".join(f"{a} {b}" for a, b in pairs)


def _fmt_per_shift(numer: int, shifts: int) -> str:
    if shifts <= 0:
        return ""
    return f"{(numer / shifts):.2f}"


def _fmt_plus_minus(x: int) -> str:
    if x == 0:
        return "0"
    return f"{x:+d}"


def _merge_intervals(intervals: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    if not intervals:
        return []
    merged: List[Tuple[int, int]] = []
    for lo, hi in sorted(intervals, key=lambda x: (x[0], x[1])):
        if not merged:
            merged.append((lo, hi))
            continue
        prev_lo, prev_hi = merged[-1]
        if lo <= prev_hi:
            merged[-1] = (prev_lo, max(prev_hi, hi))
        else:
            merged.append((lo, hi))
    return merged


def _intersection_seconds(a: List[Tuple[int, int]], b: List[Tuple[int, int]]) -> int:
    """
    Sum the overlap duration between two lists of non-overlapping, sorted intervals.
    Intervals are treated as continuous with duration (hi - lo).
    """
    i = 0
    j = 0
    total = 0
    while i < len(a) and j < len(b):
        alo, ahi = a[i]
        blo, bhi = b[j]
        lo = max(alo, blo)
        hi = min(ahi, bhi)
        if hi > lo:
            total += hi - lo
        if ahi <= bhi:
            i += 1
        else:
            j += 1
    return total


def _compute_pair_on_ice_rows(
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    goals_by_period: Dict[int, List["GoalEvent"]],
) -> List[Dict[str, Any]]:
    """
    Produce directed (player, teammate) rows for pair on-ice overlap and goal +/- together.

    - overlap is computed from scoreboard shift intervals (unioned by period)
    - goal counts include goals during overlap, but skip goals at a player's shift start
      (matches the plus/minus behavior in `_compute_player_stats`)
    """
    all_players = sorted([p for p in sb_pairs_by_player.keys() if p])
    if not all_players:
        return []

    intervals_by_player_period: Dict[str, Dict[int, List[Tuple[int, int]]]] = {}
    raw_intervals_by_player_period: Dict[str, Dict[int, List[Tuple[int, int, Optional[int]]]]] = {}
    toi_sec_by_player: Dict[str, int] = {}

    for player in all_players:
        per_period: Dict[int, List[Tuple[int, int]]] = {}
        raw_by_period: Dict[int, List[Tuple[int, int, Optional[int]]]] = {}
        for period, a, b in sb_pairs_by_player.get(player, []):
            lo, hi = compute_interval_seconds(a, b)
            per_period.setdefault(period, []).append((lo, hi))
            try:
                start_sec: Optional[int] = parse_flex_time_to_seconds(a)
            except Exception:
                start_sec = None
            raw_by_period.setdefault(period, []).append((lo, hi, start_sec))
        merged_by_period: Dict[int, List[Tuple[int, int]]] = {}
        toi_total = 0
        for period, intervals in per_period.items():
            merged = _merge_intervals(intervals)
            merged_by_period[period] = merged
            toi_total += sum(hi - lo for lo, hi in merged)
        intervals_by_player_period[player] = merged_by_period
        raw_intervals_by_player_period[player] = raw_by_period
        toi_sec_by_player[player] = toi_total

    # Exclude players with no recorded TOI from pair tables (likely absent or a goalie).
    players = [p for p in all_players if int(toi_sec_by_player.get(p, 0) or 0) > 0]
    if not players:
        return []

    jersey_to_player: Dict[str, str] = {}
    for pk in players:
        norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
        if norm:
            jersey_to_player.setdefault(norm, pk)

    def _on_ice_for_goal(player: str, period: int, t_sec: int) -> bool:
        # Match `_compute_player_stats` behavior:
        #  - count a goal if it falls within any shift interval for the player
        #  - BUT skip a goal that occurs exactly at the *start* of that specific shift
        #    (important when a goal occurs at the boundary between consecutive shifts).
        for lo, hi, start_sec in raw_intervals_by_player_period.get(player, {}).get(period, []):
            if not interval_contains(t_sec, lo, hi):
                continue
            if start_sec is not None and t_sec == start_sec:
                continue
            return True
        return False

    def _on_ice_any(player: str, period: int, t_sec: int) -> bool:
        for lo, hi in intervals_by_player_period.get(player, {}).get(period, []):
            if interval_contains(t_sec, lo, hi):
                return True
        return False

    overlap_by_pair: Dict[Tuple[str, str], int] = {}
    for i in range(len(players)):
        for j in range(i + 1, len(players)):
            a = players[i]
            b = players[j]
            overlap = 0
            a_periods = intervals_by_player_period.get(a, {})
            b_periods = intervals_by_player_period.get(b, {})
            for period in set(a_periods.keys()) & set(b_periods.keys()):
                overlap += _intersection_seconds(
                    a_periods.get(period, []), b_periods.get(period, [])
                )
            overlap_by_pair[(a, b)] = overlap

    goals_by_pair: Dict[Tuple[str, str], Tuple[int, int]] = {}
    player_goals_on_ice_by_pair: Dict[Tuple[str, str], int] = {}
    player_assists_on_ice_by_pair: Dict[Tuple[str, str], int] = {}
    collab_goals_by_pair: Dict[Tuple[str, str], int] = {}
    collab_assists_by_pair: Dict[Tuple[str, str], int] = {}
    player_total_pm: Dict[str, int] = {p: 0 for p in players}
    for period, goals in (goals_by_period or {}).items():
        if not goals:
            continue
        for ev in goals:
            on_ice = [p for p in players if _on_ice_for_goal(p, period, int(ev.t_sec))]
            on_ice_any = [p for p in players if _on_ice_any(p, period, int(ev.t_sec))]
            for p in on_ice:
                if ev.kind == "GF":
                    player_total_pm[p] = int(player_total_pm.get(p, 0) or 0) + 1
                else:
                    player_total_pm[p] = int(player_total_pm.get(p, 0) or 0) - 1

            # Attribute goal/assist counts only for our team's goals (GF) when scorer/assists are known.
            if ev.kind == "GF":
                scorer_key: Optional[str] = None
                if ev.scorer:
                    scorer_key = jersey_to_player.get(_normalize_jersey_number(ev.scorer) or "")
                assist_keys: List[str] = []
                for a in ev.assists or []:
                    ak = jersey_to_player.get(_normalize_jersey_number(a) or "")
                    if ak:
                        assist_keys.append(ak)

                if scorer_key and scorer_key in on_ice_any:
                    for teammate in on_ice_any:
                        if teammate == scorer_key:
                            continue
                        player_goals_on_ice_by_pair[(scorer_key, teammate)] = (
                            int(player_goals_on_ice_by_pair.get((scorer_key, teammate), 0) or 0) + 1
                        )

                for ak in assist_keys:
                    if ak not in on_ice_any:
                        continue
                    for teammate in on_ice_any:
                        if teammate == ak:
                            continue
                        player_assists_on_ice_by_pair[(ak, teammate)] = (
                            int(player_assists_on_ice_by_pair.get((ak, teammate), 0) or 0) + 1
                        )

                # Direct collaboration: scorer<->assister pairs on the same goal.
                if scorer_key:
                    for ak in assist_keys:
                        if ak == scorer_key:
                            continue
                        collab_goals_by_pair[(scorer_key, ak)] = (
                            int(collab_goals_by_pair.get((scorer_key, ak), 0) or 0) + 1
                        )
                        collab_assists_by_pair[(ak, scorer_key)] = (
                            int(collab_assists_by_pair.get((ak, scorer_key), 0) or 0) + 1
                        )

            if len(on_ice) < 2:
                continue
            for i in range(len(on_ice)):
                for j in range(i + 1, len(on_ice)):
                    a, b = sorted((on_ice[i], on_ice[j]))
                    gf, ga = goals_by_pair.get((a, b), (0, 0))
                    if ev.kind == "GF":
                        gf += 1
                    else:
                        ga += 1
                    goals_by_pair[(a, b)] = (gf, ga)

    rows: List[Dict[str, Any]] = []
    for player in players:
        player_toi = int(toi_sec_by_player.get(player, 0) or 0)
        for teammate in players:
            if teammate == player:
                continue
            a, b = sorted((player, teammate))
            overlap = int(overlap_by_pair.get((a, b), 0) or 0)
            gf, ga = goals_by_pair.get((a, b), (0, 0))
            pct = (100.0 * overlap / player_toi) if player_toi > 0 else 0.0
            rows.append(
                {
                    "player": player,
                    "teammate": teammate,
                    "shift_games": 1,
                    "player_toi_seconds": player_toi,
                    "overlap_seconds": overlap,
                    "overlap_pct": pct,
                    "gf_together": int(gf),
                    "ga_together": int(ga),
                    "player_goals_on_ice_together": int(
                        player_goals_on_ice_by_pair.get((player, teammate), 0) or 0
                    ),
                    "player_assists_on_ice_together": int(
                        player_assists_on_ice_by_pair.get((player, teammate), 0) or 0
                    ),
                    "goals_collab_with_teammate": int(
                        collab_goals_by_pair.get((player, teammate), 0) or 0
                    ),
                    "assists_collab_with_teammate": int(
                        collab_assists_by_pair.get((player, teammate), 0) or 0
                    ),
                    "player_total_plus_minus": int(player_total_pm.get(player, 0) or 0),
                    "teammate_total_plus_minus": int(player_total_pm.get(teammate, 0) or 0),
                    "plus_minus_together": int(gf) - int(ga),
                }
            )
    return rows


def summarize_shift_lengths_sec(pairs: List[Tuple[str, str]]) -> Dict[str, str]:
    """
    Given scoreboard time string pairs, compute durations in seconds and summarize.
    """
    lengths = []
    for a, b in pairs:
        lo, hi = compute_interval_seconds(a, b)
        lengths.append(hi - lo)
    if not lengths:
        return {
            "num_shifts": "0",
            "toi_total": "0",
            "toi_avg": "0",
            "toi_median": "0",
            "toi_longest": "0",
            "toi_shortest": "0",
        }
    return {
        "num_shifts": str(len(lengths)),
        "toi_total": _format_duration(sum(lengths)),
        "toi_avg": _format_duration(int(sum(lengths) / len(lengths))),
        "toi_median": _format_duration(int(statistics.median(lengths))),
        "toi_longest": _format_duration(max(lengths)),
        "toi_shortest": _format_duration(min(lengths)),
    }


def per_period_toi(pairs_by_period: Dict[int, List[Tuple[str, str]]]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for period, pairs in pairs_by_period.items():
        total = 0
        for a, b in pairs:
            lo, hi = compute_interval_seconds(a, b)
            total += hi - lo
        out[period] = _format_duration(total)
    return out


@dataclass
class EventLogContext:
    event_counts_by_player: Dict[str, Dict[str, int]]
    event_counts_by_type_team: Dict[Tuple[str, str], int]
    event_instances: Dict[Tuple[str, str], List[Dict[str, Any]]]
    event_player_rows: List[Dict[str, Any]]
    team_roster: Dict[str, List[int]]
    team_excluded: Dict[str, List[int]]


@dataclass(frozen=True)
class LongEvent:
    event_type: str
    team: str
    period: int
    video_s: Optional[int]
    game_s: Optional[int]
    jerseys: Tuple[int, ...] = ()


@dataclass
class SpreadsheetEventMappingSummary:
    """
    Summary of how spreadsheet-specific event rows were mapped to output event types.
    """

    raw_row_counts: Dict[Tuple[str, str], int]
    mapped_counts: Dict[Tuple[str, str], Dict[str, int]]
    unmapped_row_counts: Dict[Tuple[str, str], int]

    @classmethod
    def empty(cls) -> "SpreadsheetEventMappingSummary":
        return cls(raw_row_counts={}, mapped_counts={}, unmapped_row_counts={})

    def merge(self, other: Optional["SpreadsheetEventMappingSummary"]) -> None:
        if other is None:
            return
        for k, v in (other.raw_row_counts or {}).items():
            self.raw_row_counts[k] = int(self.raw_row_counts.get(k, 0) or 0) + int(v or 0)
        for k, v in (other.unmapped_row_counts or {}).items():
            self.unmapped_row_counts[k] = int(self.unmapped_row_counts.get(k, 0) or 0) + int(v or 0)
        for k, d in (other.mapped_counts or {}).items():
            dest = self.mapped_counts.setdefault(k, {})
            for et, cnt in (d or {}).items():
                dest[et] = int(dest.get(et, 0) or 0) + int(cnt or 0)


def _print_spreadsheet_event_mapping_summary(
    *, label: str, summary: SpreadsheetEventMappingSummary
) -> None:
    if not summary or not (summary.raw_row_counts or {}):
        return
    print(f"[events] Spreadsheet event mapping summary: {label}")
    items = list((summary.raw_row_counts or {}).items())
    items.sort(key=lambda kv: (-int(kv[1] or 0), kv[0][0], kv[0][1]))
    for (raw_label, raw_marker), nrows in items:
        mapped = (summary.mapped_counts or {}).get((raw_label, raw_marker), {}) or {}
        unmapped = int((summary.unmapped_row_counts or {}).get((raw_label, raw_marker), 0) or 0)
        parts: List[str] = []
        if raw_label:
            parts.append(f"label={raw_label!r}")
        if raw_marker:
            parts.append(f"marker={raw_marker!r}")
        parts_s = " ".join(parts) if parts else "label=<blank>"
        mapped_s = ", ".join(
            f"{et}={int(cnt or 0)}"
            for et, cnt in sorted(mapped.items(), key=lambda x: (-int(x[1] or 0), x[0]))
        )
        if mapped_s:
            mapped_s = f" -> {mapped_s}"
        if unmapped:
            print(f"  - {parts_s}: rows={int(nrows or 0)} UNMAPPED={unmapped}{mapped_s}")
        else:
            print(f"  - {parts_s}: rows={int(nrows or 0)}{mapped_s}")


def _parse_long_mmss_time_to_seconds(cell: Any) -> Optional[int]:
    """
    Parse times as they appear in the '-long' sheets.

    These sheets commonly store MM:SS values as Excel time-of-day cells, so
    pandas yields datetime.time like 23:56:00 (meaning 23:56, not 23 hours).

    Accepts:
      - datetime.time / pd.Timestamp: interpret hour as minutes, minute as seconds
      - strings like '24:50' or '00:56:00': interpret as MM:SS (ignore 3rd component)
    """
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    if isinstance(cell, datetime.time):
        return int(cell.hour) * 60 + int(cell.minute)
    if isinstance(cell, pd.Timestamp):
        return int(cell.hour) * 60 + int(cell.minute)
    try:
        s = str(cell).strip()
    except Exception:
        return None
    if not s:
        return None
    # Skip header labels like 'Video Time'
    if s.lower() in {"video time", "scoreboard", "team"}:
        return None
    parts = s.split(":")
    try:
        if len(parts) == 3:
            a_s, b_s, c_s = parts
            a = int(a_s)
            b = int(b_s)
            c = int(c_s)
            # Some sheets emit long video times as H:MM:SS (e.g., 1:09:19, 1:00:00).
            # Others emit MM:SS as a time-of-day-like string HH:MM:SS (e.g., 23:56:00 meaning 23:56).
            #
            # Heuristic:
            #   - If seconds are non-zero, it's almost certainly H:MM:SS.
            #   - If seconds are zero, treat small unpadded leading fields (e.g., "1:00:00") as H:MM:SS,
            #     otherwise treat as MM:SS encoded as HH:MM:SS.
            if c != 0:
                return a * 3600 + b * 60 + c
            if 0 < a < 4 and len(a_s) == 1:
                return a * 3600 + b * 60 + c
            return a * 60 + b
        if len(parts) == 2:
            mm, ss = parts
            return int(mm) * 60 + int(ss)
        if len(parts) == 1:
            return int(float(parts[0]))
    except Exception:
        return None
    return None


def _parse_long_team(cell: Any) -> Optional[str]:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    try:
        s = str(cell).strip()
    except Exception:
        return None
    if not s:
        return None
    sl = s.lower()
    if "blue" in sl or sl.startswith("blu"):
        return "Blue"
    if "white" in sl or sl.startswith("whi"):
        return "White"
    return None


def _extract_jerseys_from_cell(cell: Any) -> List[int]:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return []
    if isinstance(cell, (int, float)) and not pd.isna(cell):
        try:
            n = int(cell)
        except Exception:
            return []
        return [n] if 1 <= n <= 98 else []
    try:
        s = str(cell).strip()
    except Exception:
        return []
    if not s:
        return []
    # Don't treat times as jersey numbers.
    if re.match(r"^\d{1,3}:\d{2}(:\d{2})?$", s):
        return []
    nums: List[int] = []
    for m in re.finditer(r"#?(\d{1,2})(?!\d)", s):
        try:
            n = int(m.group(1))
        except Exception:
            continue
        if 1 <= n <= 98:
            nums.append(n)
    # Dedupe while preserving order.
    seen: set[int] = set()
    out: List[int] = []
    for n in nums:
        if n in seen:
            continue
        seen.add(n)
        out.append(n)
    return out


def _parse_long_left_event_table(
    df: pd.DataFrame,
) -> Tuple[List[LongEvent], List[Dict[str, Any]], Dict[str, set[int]]]:
    events, goal_rows, jerseys_by_team, _mapping = _parse_long_left_event_table_with_mapping(df)
    return events, goal_rows, jerseys_by_team


def _parse_long_left_event_table_with_mapping(
    df: pd.DataFrame,
) -> Tuple[
    List[LongEvent], List[Dict[str, Any]], Dict[str, set[int]], SpreadsheetEventMappingSummary
]:
    """
    Parse the leftmost per-period event table found in '*-long*.xlsx' sheets.

    Returns:
      - list of LongEvent entries (for event summaries/clips)
      - list of goal rows with scorer/assists for optional goal inference
      - observed jerseys by team color (Blue/White) for team inference
    """
    if df.empty or df.shape[1] < 6:
        return [], [], {}, SpreadsheetEventMappingSummary.empty()

    # Identify period header rows in column 0 (e.g., '1st Period', '2nd Period', '3rd period').
    col0 = df.iloc[:, 0]
    period_rows: List[Tuple[int, int]] = []
    for r, v in col0.items():
        p = parse_period_label(v)
        if p is not None:
            period_rows.append((int(r), int(p)))

    if not period_rows:
        return [], [], {}, SpreadsheetEventMappingSummary.empty()

    # Append sentinel end row.
    period_rows_sorted = sorted(period_rows, key=lambda x: x[0])
    period_rows_sorted.append((int(df.shape[0]), -1))

    events: List[LongEvent] = []
    goal_rows: List[Dict[str, Any]] = []
    jerseys_by_team: Dict[str, set[int]] = {"Blue": set(), "White": set()}
    mapping_summary = SpreadsheetEventMappingSummary.empty()

    def _col_for(header_row: int, *candidates: str) -> Optional[int]:
        # Map normalized header names to column indices.
        label_to_col: Dict[str, int] = {}
        for c in range(df.shape[1]):
            v = df.iat[header_row, c]
            if pd.isna(v):
                continue
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            label_to_col[_normalize_header_label(s)] = c
        for cand in candidates:
            key = _normalize_header_label(cand)
            if key in label_to_col:
                return label_to_col[key]
        # Fallback: substring match
        for cand in candidates:
            key = _normalize_header_label(cand)
            for hdr, c in label_to_col.items():
                if key and key in hdr:
                    return c
        return None

    for idx in range(len(period_rows_sorted) - 1):
        header_r, period = period_rows_sorted[idx]
        end_r = period_rows_sorted[idx + 1][0]

        # The header row itself contains column names (Video Time, Scoreboard, Team, etc.).
        video_col = _col_for(header_r, "Video Time")
        sb_col = _col_for(header_r, "Scoreboard", "Game Time")
        team_col = _col_for(header_r, "Team")
        shots_col = _col_for(header_r, "Shots", "Shot")
        sog_col = _col_for(header_r, "Shots on Goal", "Shot on Goal", "SOG")
        xg_col = _col_for(header_r, "Expected Goal", "Expected Goals", "xG", "XG")
        assist_col = _col_for(header_r, "Assist", "Assists")
        if xg_col is None and sog_col is not None:
            try:
                cand = int(sog_col) + 1
                if 0 <= cand < int(df.shape[1]):
                    hdr = df.iat[header_r, cand]
                    hdr_s = str(hdr).strip().lower() if isinstance(hdr, str) else ""
                    if ("expected" in hdr_s) or (hdr_s in {"xg", "xg (expected goal)"}):
                        xg_col = cand
            except Exception:  # noqa: BLE001
                # Best-effort detection of a companion xG column to the right of SOG; ignore failures.
                pass

        for r in range(header_r + 1, end_r):
            team = _parse_long_team(df.iat[r, team_col] if team_col is not None else None)
            if not team:
                continue
            vsec = (
                _parse_long_mmss_time_to_seconds(df.iat[r, video_col])
                if video_col is not None
                else None
            )
            gsec = (
                _parse_long_mmss_time_to_seconds(df.iat[r, sb_col]) if sb_col is not None else None
            )

            label = df.iat[r, 0]
            label_s = str(label).strip() if isinstance(label, str) else ""
            label_l = label_s.lower()

            shooter = (
                _extract_jerseys_from_cell(df.iat[r, shots_col]) if shots_col is not None else []
            )
            assists = (
                _extract_jerseys_from_cell(df.iat[r, assist_col]) if assist_col is not None else []
            )
            marker = df.iat[r, sog_col] if sog_col is not None else None
            marker_s = str(marker).strip() if isinstance(marker, str) else ""
            marker_l = marker_s.lower()
            xg_cell = df.iat[r, xg_col] if xg_col is not None else None
            try:
                xg_s = str(xg_cell).strip()
            except Exception:
                xg_s = ""
            xg_l = xg_s.lower()

            raw_label_key = str(label_s or "").strip()
            raw_marker_key = str(marker_s or "").strip()
            map_key = (raw_label_key, raw_marker_key)
            mapping_summary.raw_row_counts[map_key] = (
                int(mapping_summary.raw_row_counts.get(map_key, 0) or 0) + 1
            )
            mapped_types_for_row: List[str] = []

            def _add_event(
                *,
                event_type: str,
                team: str,
                period: int,
                video_s: Optional[int],
                game_s: Optional[int],
                jerseys: Optional[Iterable[int]] = None,
            ) -> None:
                events.append(
                    LongEvent(
                        event_type=str(event_type),
                        team=str(team),
                        period=int(period),
                        video_s=video_s,
                        game_s=game_s,
                        jerseys=tuple(int(x) for x in (list(jerseys or []) if jerseys else [])),
                    )
                )
                mapped_types_for_row.append(str(event_type))

            def _record_mapping_for_row() -> None:
                if not mapped_types_for_row and (
                    raw_label_key or raw_marker_key or shooter or assists
                ):
                    mapping_summary.unmapped_row_counts[map_key] = (
                        int(mapping_summary.unmapped_row_counts.get(map_key, 0) or 0) + 1
                    )
                    return
                dest = mapping_summary.mapped_counts.setdefault(map_key, {})
                for et in mapped_types_for_row:
                    dest[et] = int(dest.get(et, 0) or 0) + 1

            # Long sheets encode turnovers in a few different row labels over time:
            #  - "Turnover" / "Turnover (forced)"  -> forced turnover (lost puck)
            #  - "Unforced TO" (and occasional typos like "Unforced OT") -> unforced turnover (giveaway)
            #  - legacy: "Giveaway"
            # Note: "unfoced"/"unforcd" below are intentional, capturing common typos in sheet labels.
            is_unforced_to = label_l.startswith(("unforced", "unfoced", "unforcd"))
            is_turnover = ("turnover" in label_l) and (not is_unforced_to)
            is_giveaway = ("giveaway" in label_l) and (not is_turnover) and (not is_unforced_to)
            # Some long sheets mark xG in a separate column (often immediately right of the SOG column).
            # Treat any explicit 'expected goal'/'xg' marker there as xG.
            xg_marked = False
            if xg_l:
                if ("expected" in xg_l and "goal" in xg_l) or ("xg" in xg_l):
                    xg_marked = True
                elif xg_l in {"1", "true", "yes", "y"}:
                    xg_marked = True
            else:
                if isinstance(xg_cell, (int, float)) and not pd.isna(xg_cell):
                    try:
                        xg_marked = float(xg_cell) != 0.0
                    except Exception:
                        xg_marked = False
            is_expected_goal = (
                ("expected goal" in label_l) or ("expected goal" in marker_l) or xg_marked
            )
            is_controlled_entry = bool(re.search(r"con?trolled", label_l)) and ("entr" in label_l)
            is_controlled_exit = ("controlled" in label_l) and ("exit" in label_l)
            is_rush = "rush" in label_l
            is_goal = (label_l == "goal") or (marker_l == "goal")
            is_sog = (label_l == "sog") or (marker_l in {"sog", "goal"})
            is_completed_pass = ("completed pass" in label_l) or ("complete pass" in label_l)
            is_penalty = "penalty" in label_l
            is_goalie_change = ("goalie" in label_l) and ("change" in label_l)

            if is_completed_pass:
                if shooter:
                    _add_event(
                        event_type="CompletedPass",
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=shooter,
                    )
                    for j in shooter:
                        jerseys_by_team.setdefault(team, set()).add(int(j))
                _record_mapping_for_row()
                continue

            if is_penalty:
                _add_event(
                    event_type="Penalty",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                )
                _record_mapping_for_row()
                continue

            if is_goalie_change:
                _add_event(
                    event_type="GoalieChange",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                )
                _record_mapping_for_row()
                continue

            if is_turnover or is_giveaway or is_unforced_to:
                # Turnover rows use:
                #  - "Shots" for the player who lost possession
                #  - then 1-2 jersey numbers describing the opponent roles:
                #      * forced turnover: optional "Caused by" and always "Takeaway"
                #      * unforced turnover: always just "Takeaway"
                giving = shooter
                detail_a = marker if sog_col is not None else None
                detail_b = df.iat[r, assist_col] if assist_col is not None else None
                detail_a_nums = _extract_jerseys_from_cell(detail_a)
                detail_b_nums = _extract_jerseys_from_cell(detail_b)

                other_team = "White" if team == "Blue" else "Blue"

                giving_event_type = "TurnoverForced" if is_turnover else "Giveaway"
                giveaway = giving[:1]

                caused_by: List[int] = []
                takeaway: List[int] = []

                if is_turnover:
                    # Forced turnover:
                    #  - If "Assist" has a jersey, treat it as the Takeaway player and
                    #    interpret "Shots on Goal" as the Caused By player (if present).
                    #  - Otherwise, use ordering across the detail cells:
                    #      1 jersey  -> Takeaway
                    #      2 jerseys -> Caused By, Takeaway
                    if detail_b_nums:
                        if len(detail_b_nums) > 1:
                            raise ValueError(
                                f"forced turnover row has multiple takeaway jerseys: row={r} value={detail_b!r}"
                            )
                        takeaway = detail_b_nums[:1]
                        if detail_a_nums:
                            if len(detail_a_nums) > 1:
                                raise ValueError(
                                    f"forced turnover row has multiple caused-by jerseys: row={r} value={detail_a!r}"
                                )
                            caused_by = detail_a_nums[:1]
                    else:
                        combined: List[int] = []
                        for n in detail_a_nums + detail_b_nums:
                            if n not in combined:
                                combined.append(n)
                        if len(combined) > 2:
                            raise ValueError(
                                f"forced turnover row has too many opponent jerseys: row={r} values={[detail_a, detail_b]!r}"
                            )
                        if len(combined) == 2:
                            caused_by = combined[:1]
                            takeaway = combined[1:2]
                        elif len(combined) == 1:
                            takeaway = combined[:1]
                            # If the only detail is explicitly a "Caused by" entry,
                            # count it as both CreatedTurnover + Takeaway (fallback for incomplete rows).
                            try:
                                a_txt = str(detail_a).lower() if detail_a is not None else ""
                            except Exception:
                                a_txt = ""
                            if "caused" in a_txt:
                                caused_by = combined[:1]
                                takeaway = combined[:1]
                else:
                    # Giveaway / Unforced TO: first jersey is the giveaway; second (opponent) is the takeaway.
                    if len(giving) > 1:
                        raise ValueError(
                            f"unforced turnover row has multiple giveaway jerseys: row={r} value={df.iat[r, shots_col] if shots_col is not None else None!r}"
                        )
                    combined: List[int] = []
                    for n in detail_a_nums + detail_b_nums:
                        if n not in combined:
                            combined.append(n)
                    if len(combined) > 1:
                        raise ValueError(
                            f"unforced turnover row has multiple opponent jerseys: row={r} values={[detail_a, detail_b]!r}"
                        )
                    takeaway = combined[:1]

                if giveaway:
                    _add_event(
                        event_type=giving_event_type,
                        team=team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=giveaway,
                    )
                if caused_by:
                    _add_event(
                        event_type="CreatedTurnover",
                        team=other_team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=caused_by,
                    )
                if takeaway:
                    _add_event(
                        event_type="Takeaway",
                        team=other_team,
                        period=period,
                        video_s=vsec,
                        game_s=gsec,
                        jerseys=takeaway,
                    )
                for j in giveaway:
                    jerseys_by_team.setdefault(team, set()).add(int(j))
                for j in caused_by + takeaway:
                    jerseys_by_team.setdefault(other_team, set()).add(int(j))
                _record_mapping_for_row()
                continue

            if shooter:
                _add_event(
                    event_type="Shot",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                    jerseys=shooter,
                )
            if shooter and is_sog:
                _add_event(
                    event_type="SOG",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                    jerseys=shooter,
                )
            if shooter and is_goal:
                _add_event(
                    event_type="Goal",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                    jerseys=shooter,
                )
                scorer = shooter[0] if shooter else None
                goal_rows.append(
                    {
                        "team": team,
                        "period": period,
                        "game_s": gsec,
                        "scorer": scorer,
                        "assists": assists,
                    }
                )
            if assists:
                _add_event(
                    event_type="Assist",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                    jerseys=assists,
                )
            # Expected goals (xG): count both explicit "Expected Goal" rows and all goals.
            if is_expected_goal or is_goal:
                _add_event(
                    event_type="ExpectedGoal",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                    jerseys=shooter,
                )
            if is_controlled_entry:
                _add_event(
                    event_type="ControlledEntry",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                )
            if is_controlled_exit:
                _add_event(
                    event_type="ControlledExit",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                )
            if is_rush:
                _add_event(
                    event_type="Rush",
                    team=team,
                    period=period,
                    video_s=vsec,
                    game_s=gsec,
                )

            for j in shooter + assists:
                jerseys_by_team.setdefault(team, set()).add(int(j))
            _record_mapping_for_row()

    # Remove empty defaults if nothing was seen.
    jerseys_by_team = {k: v for k, v in jerseys_by_team.items() if v}
    return events, goal_rows, jerseys_by_team, mapping_summary


def _parse_long_shift_tables(
    df: pd.DataFrame,
) -> Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]]:
    """
    Parse the embedded per-team shift tables found in '*-long*.xlsx' sheets.

    These long sheets often contain (for each team):
      - a title row like "<Team Name> (1st Period)"
      - a row like "2nd Period" / "3rd Period" between blocks
      - a shift table with columns similar to the primary shift spreadsheet:
          - Jersey Number, Player Name
          - Shift start/end (Scoreboard time)
          - Shift Start/End (Video Time)

    Returns:
      {
        "<Team Name>": {
          "sb_pairs_by_player": { "<jersey>_<name>": [(period, start, end), ...], ... },
          "video_pairs_by_player": { "<jersey>_<name>": [(video_start, video_end), ...], ... },
        },
        ...
      }
    """
    if df is None or df.empty:
        return {}

    def _find_text_cells(row: pd.Series) -> List[Tuple[int, str]]:
        out: List[Tuple[int, str]] = []
        for c in range(len(row)):
            v = row.iloc[c]
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            out.append((c, s))
        return out

    def _parse_team_header(s: str) -> Optional[Tuple[str, int]]:
        # Example: "San Jose Jr Sharks 12AA-2 (1st Period)"
        m = re.match(r"(?is)^\s*(.+?)\s*\(([^)]+)\)\s*$", str(s or ""))
        if not m:
            return None
        left = _clean_html_fragment(m.group(1))
        right = _clean_html_fragment(m.group(2))

        # Common format: "<Team Name> (1st Period)"
        p = parse_period_label(right)
        if left and p is not None:
            return left, int(p)

        # Older long sheets sometimes label shift tables by team color only:
        #   "1st Period (White team)" / "1st Period (Blue team)"
        p2 = parse_period_label(left)
        if p2 is not None and right:
            rl = right.casefold()
            if "white" in rl:
                return "White", int(p2)
            if "blue" in rl:
                return "Blue", int(p2)
        return None

    def _parse_period_only(s: str) -> Optional[int]:
        # Example: "2nd Period"
        return parse_period_label(s)

    def _is_shift_table_header(row: pd.Series) -> bool:
        have_jersey = False
        have_name = False
        have_sb_start = False
        have_sb_end = False
        for _c, txt in _find_text_cells(row):
            norm = _normalize_header_label(txt)
            if norm in {"jerseynumber", "jerseyno", "jerseynumber"} or (
                "jersey" in norm and "number" in norm
            ):
                have_jersey = True
            if norm in {"playername", "player"} or ("player" in norm and "name" in norm):
                have_name = True
            if "shiftstart" in norm and "scoreboard" in norm:
                have_sb_start = True
            if "shiftend" in norm and "scoreboard" in norm:
                have_sb_end = True
        return have_jersey and have_name and have_sb_start and have_sb_end

    out: Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]] = {}
    current_team: Optional[str] = None
    current_period: Optional[int] = None

    r = 0
    while r < df.shape[0]:
        row = df.iloc[r]

        # Update team/period context from title lines.
        for _c, txt in _find_text_cells(row):
            header = _parse_team_header(txt)
            if header is not None:
                current_team, current_period = header
                break
        # Update period if we see a period-only marker within a team block.
        if current_team is not None:
            for _c, txt in _find_text_cells(row):
                p = _parse_period_only(txt)
                if p is not None:
                    current_period = int(p)
                    break

        if not _is_shift_table_header(row):
            r += 1
            continue

        if current_team is None or current_period is None:
            r += 1
            continue

        # Identify core columns (jersey/name) and shift time groups.
        jersey_col: Optional[int] = None
        name_col: Optional[int] = None
        for c, txt in _find_text_cells(row):
            norm = _normalize_header_label(txt)
            if jersey_col is None and ("jersey" in norm and "number" in norm):
                jersey_col = c
            if name_col is None and (norm == "playername" or ("player" in norm and "name" in norm)):
                name_col = c

        groups = forward_fill_header_labels(row)
        start_sb_cols = _resolve_header_columns(
            groups,
            LABEL_START_SB,
            "Shift Start (Scoreboard time)",
            "Shift start (Scoreboard time)",
            "Shift start (Scoreboard time)",
        )
        end_sb_cols = _resolve_header_columns(
            groups,
            LABEL_END_SB,
            "Shift End (Scoreboard Time)",
            "Shift end (Scoreboard time)",
            "Shift End (Scoreboard time)",
        )
        start_v_cols = _resolve_header_columns(groups, LABEL_START_V, "Shift Start (Video Time)")
        end_v_cols = _resolve_header_columns(groups, LABEL_END_V, "Shift End (Video Time)")

        if jersey_col is None or name_col is None or not start_sb_cols or not end_sb_cols:
            r += 1
            continue

        team_entry = out.setdefault(
            str(current_team),
            {"sb_pairs_by_player": {}, "video_pairs_by_player": {}},
        )
        sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]] = team_entry["sb_pairs_by_player"]  # type: ignore[assignment]
        video_pairs_by_player: Dict[str, List[Tuple[str, str]]] = team_entry["video_pairs_by_player"]  # type: ignore[assignment]

        # Parse until we hit a new header / period marker.
        blank_streak = 0
        rr = r + 1
        while rr < df.shape[0]:
            row2 = df.iloc[rr]
            # Stop at another header row or at a new team title / period marker.
            if _is_shift_table_header(row2):
                break
            stop = False
            for _c, txt in _find_text_cells(row2):
                if _parse_team_header(txt) is not None:
                    stop = True
                    break
                if _parse_period_only(txt) is not None:
                    stop = True
                    break
            if stop:
                break

            jersey_norm = _normalize_jersey_number(row2.iloc[jersey_col])
            nm_raw = row2.iloc[name_col]
            name = str(nm_raw or "").strip()
            if not jersey_norm or not name or name.lower() in {"nan", "none"}:
                blank_streak += 1
                if blank_streak >= 5:
                    break
                rr += 1
                continue
            blank_streak = 0

            player_key = f"{sanitize_name(jersey_norm)}_{sanitize_name(name)}"
            sb_pairs = extract_pairs_from_row(row2, start_sb_cols, end_sb_cols)
            for a, b in sb_pairs:
                if not a or not b:
                    continue
                sb_pairs_by_player.setdefault(player_key, []).append((int(current_period), a, b))
            if start_v_cols and end_v_cols:
                v_pairs = extract_pairs_from_row(row2, start_v_cols, end_v_cols)
                if v_pairs:
                    video_pairs_by_player.setdefault(player_key, []).extend(v_pairs)
            rr += 1

        r = rr

    return out


def _compare_primary_shifts_to_long_shifts(
    *,
    primary_sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    long_shift_tables_by_team: Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]],
    threshold_seconds: int = 5,
    warn_label: str = "",
    long_sheet_paths: Optional[List[Path]] = None,
) -> Dict[str, Any]:
    """
    Compare the primary (non-long) shift sheet against long-sheet shift tables for the
    matching team (chosen by jersey overlap) and print discrepancies > threshold_seconds.
    """
    if not primary_sb_pairs_by_player or not long_shift_tables_by_team:
        return {}

    def _warn(msg: str) -> None:
        if not msg:
            return
        prefix = f"[long-shifts]{' [' + warn_label + ']' if warn_label else ''}"
        print(f"{prefix} {msg}", file=sys.stderr)

    our_jerseys: set[str] = set()
    for pk in primary_sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
        if norm:
            our_jerseys.add(norm)
    if not our_jerseys:
        return {}

    # Choose the long-sheet team whose jersey set best overlaps our primary sheet.
    best_team: Optional[str] = None
    best_overlap = -1
    best_total = 0
    for team_name, info in (long_shift_tables_by_team or {}).items():
        sb_map_any = (info or {}).get("sb_pairs_by_player") or {}
        jerseys_team: set[str] = set()
        for pk in sb_map_any.keys():
            norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
            if norm:
                jerseys_team.add(norm)
        ov = len(our_jerseys & jerseys_team)
        if ov > best_overlap:
            best_overlap = ov
            best_team = str(team_name)
            best_total = len(jerseys_team)

    if best_team is None or best_overlap <= 0:
        _warn("could not match a long-sheet team to the primary shift sheet (no jersey overlap).")
        return {}

    sb_long_any = (long_shift_tables_by_team.get(best_team) or {}).get("sb_pairs_by_player") or {}

    # Index long-sheet players by jersey for robust matching (names may differ).
    long_player_by_jersey: Dict[str, str] = {}
    for pk in sb_long_any.keys():
        norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
        if not norm:
            continue
        long_player_by_jersey.setdefault(norm, pk)

    # Normalize both sources to per-player per-period lists of (start_sec, end_sec).
    def _to_period_intervals(
        sb_list: List[Tuple[int, str, str]],
    ) -> Dict[int, List[Tuple[int, int]]]:
        out: Dict[int, List[Tuple[int, int]]] = {}
        for per, a, b in sb_list or []:
            try:
                p = int(per)
            except Exception:
                continue
            try:
                sa = parse_flex_time_to_seconds(str(a))
                sb = parse_flex_time_to_seconds(str(b))
            except Exception:
                continue
            out.setdefault(p, []).append((sa, sb))
        # Sort by shift start descending (scoreboard counts down).
        for p, lst in out.items():
            lst.sort(key=lambda x: (-max(x[0], x[1]), -min(x[0], x[1])))
        return out

    def _match_intervals(
        prim: List[Tuple[int, int]], long: List[Tuple[int, int]]
    ) -> List[Tuple[Tuple[int, int], Optional[Tuple[int, int]]]]:
        if not prim:
            return []
        used: set[int] = set()
        matches: List[Tuple[Tuple[int, int], Optional[Tuple[int, int]]]] = []
        for p_int in prim:
            best_j = None
            best_cost = None
            for j, l_int in enumerate(long or []):
                if j in used:
                    continue
                cost = abs(p_int[0] - l_int[0]) + abs(p_int[1] - l_int[1])
                if best_cost is None or cost < best_cost:
                    best_cost = cost
                    best_j = j
            if best_j is None:
                matches.append((p_int, None))
            else:
                used.add(best_j)
                matches.append((p_int, long[best_j]))
        return matches

    total_checked = 0
    total_mismatched = 0
    total_missing_in_long = 0
    total_extra_in_long = 0
    total_primary_shifts = 0
    total_long_shifts = 0
    nonsensical_primary = 0
    nonsensical_long = 0
    total_primary_toi_s = 0
    total_long_toi_s = 0
    total_len_diff_over_threshold = 0
    per_player_rows: List[Dict[str, Any]] = []
    mismatch_rows: List[Dict[str, Any]] = []

    def _dur(a: int, b: int) -> int:
        try:
            return abs(int(b) - int(a))
        except Exception:
            return 0

    MAX_SHIFT_SECONDS = 30 * 60

    for pk_primary, sb_list_primary in sorted(
        primary_sb_pairs_by_player.items(), key=lambda x: x[0]
    ):
        jersey = _normalize_jersey_number(_parse_player_key(pk_primary).jersey)
        if not jersey:
            continue
        if not sb_list_primary:
            continue
        pk_long = long_player_by_jersey.get(jersey)
        if not pk_long:
            continue

        prim_by_period = _to_period_intervals(sb_list_primary)
        long_by_period = _to_period_intervals(sb_long_any.get(pk_long, []) or [])

        player_checked = 0
        player_mismatched = 0
        player_missing = 0
        player_extra = 0
        player_max_ds = 0
        player_max_de = 0
        player_max_dlen = 0
        player_sum_dlen = 0
        player_sum_ds = 0
        player_sum_de = 0
        player_pairs = 0
        player_primary_toi_s = 0
        player_long_toi_s = 0
        player_len_diff_over_threshold = 0

        for period, prim_ints in sorted(prim_by_period.items(), key=lambda x: x[0]):
            long_ints = long_by_period.get(period, [])
            total_primary_shifts += len(prim_ints)
            total_long_shifts += len(long_ints)
            for a, b in prim_ints:
                if _dur(a, b) <= 0 or _dur(a, b) > MAX_SHIFT_SECONDS:
                    nonsensical_primary += 1
                player_primary_toi_s += _dur(a, b)
                total_primary_toi_s += _dur(a, b)
            for a, b in long_ints:
                if _dur(a, b) <= 0 or _dur(a, b) > MAX_SHIFT_SECONDS:
                    nonsensical_long += 1
                player_long_toi_s += _dur(a, b)
                total_long_toi_s += _dur(a, b)

            pairs = _match_intervals(prim_ints, long_ints)
            # Extra shifts in long are those not used by the matching.
            used = sum(1 for _p, long_int in pairs if long_int is not None)
            if len(long_ints) > used:
                extra = len(long_ints) - used
                total_extra_in_long += extra
                player_extra += extra
            for idx, (p_int, l_int) in enumerate(pairs):
                total_checked += 1
                player_checked += 1
                if l_int is None:
                    total_mismatched += 1
                    total_missing_in_long += 1
                    player_missing += 1
                    player_mismatched += 1
                    _warn(
                        f"missing long shift: team={best_team!r} player={pk_primary} period={period} primary={seconds_to_mmss_or_hhmmss(p_int[0])}-{seconds_to_mmss_or_hhmmss(p_int[1])}"
                    )
                    mismatch_rows.append(
                        {
                            "team": best_team,
                            "player": pk_primary,
                            "period": int(period),
                            "shift": int(idx + 1),
                            "kind": "missing_in_long",
                            "primary_start": int(p_int[0]),
                            "primary_end": int(p_int[1]),
                            "long_start": None,
                            "long_end": None,
                            "delta_start_s": None,
                            "delta_end_s": None,
                            "delta_len_s": None,
                        }
                    )
                    continue
                ds = abs(int(p_int[0]) - int(l_int[0]))
                de = abs(int(p_int[1]) - int(l_int[1]))
                dlen = abs(_dur(p_int[0], p_int[1]) - _dur(l_int[0], l_int[1]))
                player_pairs += 1
                player_sum_ds += int(ds)
                player_sum_de += int(de)
                player_sum_dlen += int(dlen)
                player_max_ds = max(player_max_ds, int(ds))
                player_max_de = max(player_max_de, int(de))
                player_max_dlen = max(player_max_dlen, int(dlen))
                if int(dlen) > int(threshold_seconds):
                    total_len_diff_over_threshold += 1
                    player_len_diff_over_threshold += 1
                if ds > int(threshold_seconds) or de > int(threshold_seconds):
                    total_mismatched += 1
                    player_mismatched += 1
                    _warn(
                        f"shift mismatch (> {threshold_seconds}s): team={best_team!r} player={pk_primary} period={period} shift#{idx+1} "
                        f"primary={seconds_to_mmss_or_hhmmss(p_int[0])}-{seconds_to_mmss_or_hhmmss(p_int[1])} "
                        f"long={seconds_to_mmss_or_hhmmss(l_int[0])}-{seconds_to_mmss_or_hhmmss(l_int[1])} "
                        f"Δstart={ds}s Δend={de}s"
                    )
                    mismatch_rows.append(
                        {
                            "team": best_team,
                            "player": pk_primary,
                            "period": int(period),
                            "shift": int(idx + 1),
                            "kind": "mismatch",
                            "primary_start": int(p_int[0]),
                            "primary_end": int(p_int[1]),
                            "long_start": int(l_int[0]),
                            "long_end": int(l_int[1]),
                            "delta_start_s": int(ds),
                            "delta_end_s": int(de),
                            "delta_len_s": int(dlen),
                        }
                    )

        player_total_primary = sum(len(v or []) for v in prim_by_period.values())
        player_total_long = sum(len(v or []) for v in long_by_period.values())
        per_player_rows.append(
            {
                "player": pk_primary,
                "jersey": jersey,
                "primary_shifts": player_total_primary,
                "long_shifts": player_total_long,
                "primary_toi_s": int(player_primary_toi_s),
                "long_toi_s": int(player_long_toi_s),
                "checked_shifts": player_checked,
                "mismatched": player_mismatched,
                "missing_in_long": player_missing,
                "extra_in_long": player_extra,
                "len_diff_over_threshold": player_len_diff_over_threshold,
                "max_delta_start_s": player_max_ds,
                "max_delta_end_s": player_max_de,
                "max_delta_len_s": player_max_dlen,
                "mean_delta_len_s": (player_sum_dlen / player_pairs) if player_pairs > 0 else 0.0,
                "mean_delta_start_s": (player_sum_ds / player_pairs) if player_pairs > 0 else 0.0,
                "mean_delta_end_s": (player_sum_de / player_pairs) if player_pairs > 0 else 0.0,
            }
        )

    _warn(
        f"compared primary vs long shifts for team={best_team!r} (overlap={best_overlap}/{len(our_jerseys)} jerseys; long_team_roster={best_total}); "
        f"checked={total_checked} shifts, mismatched={total_mismatched}."
    )

    # Also count "missing in primary" shifts for matched players (long shifts not present in primary).
    # This is distinct from "extra_in_long" which is based on per-period matching.
    # Here we count any player+period long shifts that have no sufficiently-close primary match at all.
    # (Used only for summary/diagnostics.)

    summary: Dict[str, Any] = {
        "primary_label": str(warn_label or ""),
        "long_sheets": [str(p) for p in (long_sheet_paths or [])],
        "matched_team": str(best_team),
        "threshold_seconds": int(threshold_seconds),
        "primary_jerseys": len(our_jerseys),
        "team_overlap": int(best_overlap),
        "long_team_roster": int(best_total),
        "total_primary_shifts": int(total_primary_shifts),
        "total_long_shifts": int(total_long_shifts),
        "total_compared": int(total_checked),
        "mismatched": int(total_mismatched),
        "missing_in_long": int(total_missing_in_long),
        "extra_in_long": int(total_extra_in_long),
        "nonsensical_primary_shifts": int(nonsensical_primary),
        "nonsensical_long_shifts": int(nonsensical_long),
        "total_primary_toi_s": int(total_primary_toi_s),
        "total_long_toi_s": int(total_long_toi_s),
        "len_diff_over_threshold": int(total_len_diff_over_threshold),
        "per_player": per_player_rows,
        "mismatch_rows": mismatch_rows,
    }
    return summary


def _print_shift_discrepancy_rich_summary(summary: Dict[str, Any]) -> None:
    if not summary:
        return
    try:
        from rich.console import Console  # type: ignore
        from rich.table import Table  # type: ignore
        from rich.text import Text  # type: ignore
    except Exception:
        # Fallback: plain summary line.
        msg = (
            f"[shift-summary] team={summary.get('matched_team')!r} "
            f"mismatched={summary.get('mismatched')} missing_in_long={summary.get('missing_in_long')} "
            f"extra_in_long={summary.get('extra_in_long')}"
        )
        print(msg, file=sys.stderr)
        return

    console = Console(file=sys.stderr)
    title = (
        f"Shift Discrepancy Summary (team={summary.get('matched_team')!r}, "
        f"threshold>{int(summary.get('threshold_seconds') or 0)}s)"
    )
    table = Table(title=title, show_lines=False)
    table.add_column("Jersey", justify="right", no_wrap=True)
    table.add_column("Player", overflow="fold")
    table.add_column("P", justify="right")
    table.add_column("L", justify="right")
    table.add_column("Missing", justify="right")
    table.add_column("Extra", justify="right")
    table.add_column("Mismatch", justify="right")
    table.add_column("Max ΔStart", justify="right")
    table.add_column("Max ΔEnd", justify="right")
    table.add_column("Max ΔLen", justify="right")
    table.add_column("Mean ΔLen", justify="right")

    rows = list(summary.get("per_player") or [])
    rows.sort(
        key=lambda r: (
            int(r.get("mismatched") or 0),
            int(r.get("missing_in_long") or 0),
            int(r.get("extra_in_long") or 0),
        ),
        reverse=True,
    )
    for r in rows[:40]:
        mismatch = int(r.get("mismatched") or 0)
        style = "red" if mismatch > 0 else None
        pk = str(r.get("player") or "")
        jersey = str(r.get("jersey") or "")
        player_disp = _format_player_name_only(pk)
        table.add_row(
            jersey,
            Text(player_disp, style=style) if style else player_disp,
            str(int(r.get("primary_shifts") or 0)),
            str(int(r.get("long_shifts") or 0)),
            str(int(r.get("missing_in_long") or 0)),
            str(int(r.get("extra_in_long") or 0)),
            str(mismatch),
            f"{int(r.get('max_delta_start_s') or 0)}s",
            f"{int(r.get('max_delta_end_s') or 0)}s",
            f"{int(r.get('max_delta_len_s') or 0)}s",
            f"{float(r.get('mean_delta_len_s') or 0.0):.1f}s",
        )

    console.print(table)

    totals = (
        f"Totals: compared={summary.get('total_compared')} mismatched={summary.get('mismatched')} "
        f"missing_in_long={summary.get('missing_in_long')} extra_in_long={summary.get('extra_in_long')} "
        f"primary_shifts={summary.get('total_primary_shifts')} long_shifts={summary.get('total_long_shifts')} "
        f"nonsensical_primary={summary.get('nonsensical_primary_shifts')} nonsensical_long={summary.get('nonsensical_long_shifts')}"
    )
    console.print(totals)


def _write_shift_discrepancy_xlsx(stats_dir: Path, summary: Dict[str, Any]) -> None:
    if not summary:
        return
    try:
        stats_dir.mkdir(parents=True, exist_ok=True)
        out_path = stats_dir / "shift_discrepancies.xlsx"
        per_player = list(summary.get("per_player") or [])
        mismatch_rows = list(summary.get("mismatch_rows") or [])

        def _fmt_secs(val: Any) -> str:
            try:
                return seconds_to_mmss_or_hhmmss(int(val))
            except Exception:
                return ""

        def _safe_sheet_name(raw: str, used: set[str]) -> str:
            name = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(raw or "").strip())
            name = " ".join(name.split())
            if not name:
                name = "Player"
            name = name[:31]
            base = name
            idx = 2
            while name in used:
                suffix = f" {idx}"
                name = (base[: max(0, 31 - len(suffix))] + suffix).strip()
                idx += 1
            used.add(name)
            return name

        overview_rows: list[dict[str, Any]] = [
            {"Metric": "Matched Team", "Value": summary.get("matched_team")},
            {"Metric": "Threshold Seconds", "Value": summary.get("threshold_seconds")},
            {"Metric": "Primary Jerseys", "Value": summary.get("primary_jerseys")},
            {"Metric": "Team Overlap", "Value": summary.get("team_overlap")},
            {"Metric": "Long Team Roster", "Value": summary.get("long_team_roster")},
            {"Metric": "Total Primary Shifts", "Value": summary.get("total_primary_shifts")},
            {"Metric": "Total Long Shifts", "Value": summary.get("total_long_shifts")},
            {"Metric": "Total Compared", "Value": summary.get("total_compared")},
            {"Metric": "Mismatched", "Value": summary.get("mismatched")},
            {"Metric": "Missing In Long", "Value": summary.get("missing_in_long")},
            {"Metric": "Extra In Long", "Value": summary.get("extra_in_long")},
            {
                "Metric": "Length Diffs > Threshold",
                "Value": summary.get("len_diff_over_threshold"),
            },
            {
                "Metric": "Nonsensical Primary Shifts",
                "Value": summary.get("nonsensical_primary_shifts"),
            },
            {
                "Metric": "Nonsensical Long Shifts",
                "Value": summary.get("nonsensical_long_shifts"),
            },
            {"Metric": "Total Primary TOI (s)", "Value": summary.get("total_primary_toi_s")},
            {
                "Metric": "Total Primary TOI",
                "Value": _fmt_secs(summary.get("total_primary_toi_s")),
            },
            {"Metric": "Total Long TOI (s)", "Value": summary.get("total_long_toi_s")},
            {
                "Metric": "Total Long TOI",
                "Value": _fmt_secs(summary.get("total_long_toi_s")),
            },
            {"Metric": "Long Sheets", "Value": ", ".join(summary.get("long_sheets") or [])},
        ]
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # Summary sheet: overall discrepancy metrics.
            df_summary = pd.DataFrame(overview_rows)
            df_summary.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_summary.columns
            ]
            df_summary.to_excel(writer, sheet_name="summary", index=False, startrow=1)
            _apply_excel_table_style(writer, "summary", title="Shift Discrepancies", df=df_summary)
            _autosize_columns(writer, "summary", df_summary)

            used_names: set[str] = {"summary"}
            for row in per_player:
                pk = str(row.get("player") or "")
                jersey = str(row.get("jersey") or "")
                player_disp = _format_player_name_only(pk)
                sheet_name = _safe_sheet_name(f"{jersey} {player_disp}", used_names)

                player_mismatches = [m for m in mismatch_rows if str(m.get("player")) == pk]
                mismatch_table: list[dict[str, Any]] = []
                for m in player_mismatches:
                    mismatch_table.append(
                        {
                            "Period": m.get("period"),
                            "Shift": m.get("shift"),
                            "Kind": m.get("kind"),
                            "Primary Start": _fmt_secs(m.get("primary_start")),
                            "Primary End": _fmt_secs(m.get("primary_end")),
                            "Long Start": _fmt_secs(m.get("long_start")),
                            "Long End": _fmt_secs(m.get("long_end")),
                            "ΔStart (s)": m.get("delta_start_s"),
                            "ΔEnd (s)": m.get("delta_end_s"),
                            "ΔLen (s)": m.get("delta_len_s"),
                        }
                    )

                if not mismatch_table:
                    mismatch_table = [
                        {
                            "Period": "",
                            "Shift": "",
                            "Kind": "no_discrepancies",
                            "Primary Start": "",
                            "Primary End": "",
                            "Long Start": "",
                            "Long End": "",
                            "ΔStart (s)": "",
                            "ΔEnd (s)": "",
                            "ΔLen (s)": "",
                        }
                    ]

                df_player = pd.DataFrame(mismatch_table)
                df_player.columns = [
                    _wrap_header_after_words(str(c), words_per_line=2) for c in df_player.columns
                ]
                df_player.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

                title = (
                    f"{jersey} {player_disp} "
                    f"(Psh={int(row.get('primary_shifts') or 0)}, "
                    f"Lsh={int(row.get('long_shifts') or 0)}, "
                    f"mis={int(row.get('mismatched') or 0)}, "
                    f"miss={int(row.get('missing_in_long') or 0)}, "
                    f"extra={int(row.get('extra_in_long') or 0)})"
                ).strip()
                _apply_excel_table_style(writer, sheet_name, title=title, df=df_player)
                _autosize_columns(writer, sheet_name, df_player)
    except Exception:
        return


def _print_game_inputs_rich_summary(results: List[Dict[str, Any]]) -> None:
    """
    Print a compact, per-game summary of which sources were used to generate stats/events.

    This is intended to make it obvious when, e.g., a long sheet is present but its embedded
    shift tables couldn't be parsed (so opponent shifts/stats are missing).
    """
    if not results or len(results) <= 1:
        return

    try:
        from rich.console import Console  # type: ignore
        from rich.table import Table  # type: ignore
        from rich.text import Text  # type: ignore
    except Exception:
        return

    def _count_players(player_stats_csv: Path) -> Optional[int]:
        try:
            if not player_stats_csv.exists():
                return None
            n = 0
            with player_stats_csv.open("r", encoding="utf-8", errors="ignore") as f:
                for _ in f:
                    n += 1
            # subtract header
            return max(0, n - 1)
        except Exception:
            return None

    def _events_sources(events_csv: Path) -> str:
        try:
            if not events_csv.exists():
                return ""
            import csv

            with events_csv.open("r", encoding="utf-8", errors="ignore") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return ""
                src_key = None
                for cand in ("Source", "source"):
                    if cand in reader.fieldnames:
                        src_key = cand
                        break
                if src_key is None:
                    return ""
                seen: List[str] = []
                for idx, row in enumerate(reader):
                    if idx > 3000:
                        break
                    v = str((row or {}).get(src_key) or "").strip()
                    if not v:
                        continue
                    if v not in seen:
                        seen.append(v)
                # One source per line keeps the cell readable even in narrow terminals.
                return "\n".join(seen)
        except Exception:
            return ""

    def _shift_source_for_side(
        side: str,
        *,
        primary_side: Optional[str],
        primary_path: Optional[Path],
        stats_dir: Path,
    ) -> str:
        n_players = _count_players(stats_dir / "player_stats.csv")
        if n_players is None or n_players <= 0:
            return ""
        if (
            primary_side == side
            and primary_path is not None
            and (not _is_long_sheet_path(primary_path))
        ):
            return "Primary"
        return "Long"

    console = Console(file=sys.stderr)
    table = Table(title="Stats Inputs Summary", show_lines=False)
    table.add_column("Game", no_wrap=True)
    table.add_column("T2S", justify="right", no_wrap=True)
    table.add_column("Primary Side", no_wrap=True)
    table.add_column("Primary", no_wrap=True)
    table.add_column("Long", justify="right", no_wrap=True)
    table.add_column("Home Shifts", no_wrap=True)
    table.add_column("Away Shifts", no_wrap=True)
    table.add_column("Home Players", justify="right", no_wrap=True)
    table.add_column("Away Players", justify="right", no_wrap=True)
    # Wrap header at words (not per letter) and keep sources readable.
    table.add_column("Event\nSources")

    for r in sorted(results, key=lambda x: int(x.get("order") or 0)):
        label = str(r.get("label") or "")
        t2s_id = r.get("t2s_id")
        t2s_disp = str(int(t2s_id)) if isinstance(t2s_id, int) else ""
        primary_side = str(r.get("side") or "").strip().lower() or None
        primary_path = None
        try:
            pp = r.get("primary_path")
            primary_path = Path(pp) if pp else None
        except Exception:
            primary_path = None
        long_paths = list(r.get("long_paths") or [])
        long_count = len(long_paths)

        outdir = Path(r.get("outdir"))
        fmt_dir = outdir.name
        root = outdir.parent.parent
        home_stats = root / "Home" / fmt_dir / "stats"
        away_stats = root / "Away" / fmt_dir / "stats"
        home_players = _count_players(home_stats / "player_stats.csv")
        away_players = _count_players(away_stats / "player_stats.csv")
        home_shift_src = _shift_source_for_side(
            "home", primary_side=primary_side, primary_path=primary_path, stats_dir=home_stats
        )
        away_shift_src = _shift_source_for_side(
            "away", primary_side=primary_side, primary_path=primary_path, stats_dir=away_stats
        )

        events_sources = _events_sources(outdir / "stats" / "all_events_summary.csv")

        def _fmt_players(n: Optional[int]) -> str:
            if n is None:
                return ""
            return str(int(n))

        # Highlight missing opponent shifts when long sheets are present.
        warn_style = None
        if long_count > 0:
            if primary_side == "home" and not away_shift_src:
                warn_style = "red"
            if primary_side == "away" and not home_shift_src:
                warn_style = "red"

        game_cell: Any = label
        if warn_style:
            game_cell = Text(label, style=warn_style)

        table.add_row(
            game_cell,
            t2s_disp,
            primary_side or "",
            (
                "yes"
                if (
                    primary_path is not None
                    and primary_path.exists()
                    and (not _is_long_sheet_path(primary_path))
                )
                else "no"
            ),
            str(long_count) if long_count else "",
            home_shift_src,
            away_shift_src,
            _fmt_players(home_players),
            _fmt_players(away_players),
            events_sources,
        )

    console.print(table)


def _goal_records_from_long_goal_rows(
    goal_rows: List[Dict[str, Any]],
    *,
    focus_team: Optional[str],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in goal_rows or []:
        try:
            team_raw = str(r.get("team") or "").strip()
            period = int(r.get("period") or 0)
        except Exception:
            continue
        gsec = r.get("game_s")
        if not isinstance(gsec, (int, float)):
            continue
        gsec_i = int(gsec)
        scorer = r.get("scorer")
        assists = r.get("assists") or []

        scorer_norm = _normalize_jersey_number(scorer) if scorer is not None else None
        assists_norm: List[str] = []
        if isinstance(assists, (list, tuple)):
            for a in assists:
                aa = _normalize_jersey_number(a)
                if aa:
                    assists_norm.append(aa)
        else:
            aa = _normalize_jersey_number(assists)
            if aa:
                assists_norm.append(aa)

        kind: Optional[str] = None
        if focus_team in {"Blue", "White"} and team_raw in {"Blue", "White"}:
            kind = "GF" if team_raw == focus_team else "GA"

        out.append(
            {
                "source": "long",
                "team_raw": team_raw,
                "period": period,
                "game_s": gsec_i,
                "t_str": seconds_to_mmss_or_hhmmss(gsec_i),
                "kind": kind,
                "scorer": scorer_norm,
                "assists": assists_norm,
            }
        )
    out.sort(key=lambda x: (int(x.get("period") or 0), int(x.get("game_s") or 0)))
    return out


def _goal_records_from_t2s(
    goals: List[GoalEvent],
    *,
    side: str,
) -> List[Dict[str, Any]]:
    """
    Convert GoalEvent rows (relative GF/GA for our team) into goal records tagged with team_side.
    """
    side_l = str(side or "").strip().lower()
    if side_l not in {"home", "away"}:
        side_l = "home"
    opp = "away" if side_l == "home" else "home"

    out: List[Dict[str, Any]] = []
    for g in goals or []:
        period = int(getattr(g, "period", 0) or 0)
        gsec = int(getattr(g, "t_sec", 0) or 0)
        kind = str(getattr(g, "kind", "") or "").strip().upper()
        if kind not in {"GF", "GA"}:
            continue
        team_side = side_l if kind == "GF" else opp
        scorer = _normalize_jersey_number(getattr(g, "scorer", None))
        assists_norm: List[str] = []
        for a in (getattr(g, "assists", None) or []) or []:
            aa = _normalize_jersey_number(a)
            if aa:
                assists_norm.append(aa)
        out.append(
            {
                "source": "t2s",
                "team_side": team_side,
                "period": period,
                "game_s": gsec,
                "t_str": seconds_to_mmss_or_hhmmss(gsec),
                "kind": kind,
                "scorer": scorer,
                "assists": assists_norm,
            }
        )
    out.sort(key=lambda x: (int(x.get("period") or 0), int(x.get("game_s") or 0)))
    return out


def _compare_t2s_vs_long_goals(
    *,
    label: str,
    t2s_id: int,
    side: str,
    t2s_goals: List[GoalEvent],
    long_goal_rows: List[Dict[str, Any]],
    focus_team: Optional[str],
    home_team_name: Optional[str] = None,
    away_team_name: Optional[str] = None,
    match_tolerance_s: int = 2,
) -> List[Dict[str, Any]]:
    """
    Compare TimeToScore (official) goals vs long-sheet recorded goals/assists by time+period.
    Produces one row per discrepancy (missing/extra goal, scorer mismatch, assist mismatch, time mismatch).
    """
    tol = int(match_tolerance_s or 0)
    t2s_recs = _goal_records_from_t2s(t2s_goals, side=side)
    long_recs = _goal_records_from_long_goal_rows(long_goal_rows, focus_team=focus_team)

    unmatched_long = list(long_recs)
    out_rows: List[Dict[str, Any]] = []

    def _fmt_assists(lst: Any) -> str:
        if not lst:
            return ""
        if isinstance(lst, (list, tuple)):
            return ",".join([str(x) for x in lst if str(x).strip()])
        return str(lst)

    def _best_long_match(period: int, gsec: int, kind: Optional[str]) -> Optional[Dict[str, Any]]:
        best = None
        best_dt = None
        for cand in unmatched_long:
            if int(cand.get("period") or 0) != int(period):
                continue
            gs2 = int(cand.get("game_s") or 0)
            dt = abs(gs2 - int(gsec))
            if tol >= 0 and dt > tol:
                continue
            # If long has a GF/GA kind (requires focus_team), prefer matching by kind.
            cand_kind = cand.get("kind")
            if kind in {"GF", "GA"} and cand_kind in {"GF", "GA"} and cand_kind != kind:
                continue
            if best is None or best_dt is None or dt < best_dt:
                best = cand
                best_dt = dt
        return best

    for tg in t2s_recs:
        period = int(tg.get("period") or 0)
        gsec = int(tg.get("game_s") or 0)
        kind = str(tg.get("kind") or "")
        tm_side = str(tg.get("team_side") or "")
        best = _best_long_match(period, gsec, kind)
        if best is None:
            out_rows.append(
                {
                    "label": label,
                    "t2s_id": int(t2s_id),
                    "home_team": str(home_team_name or ""),
                    "away_team": str(away_team_name or ""),
                    "period": period,
                    "game_s": gsec,
                    "time": tg.get("t_str") or "",
                    "team": tm_side,
                    "issue": "T2S-only goal (no long match)",
                    "t2s_scorer": tg.get("scorer") or "",
                    "long_scorer": "",
                    "t2s_assists": _fmt_assists(tg.get("assists") or []),
                    "long_assists": "",
                    "dt_s": "",
                }
            )
            continue

        unmatched_long.remove(best)
        dt_val = abs(int(best.get("game_s") or 0) - gsec)

        # If time differs but still matched within tolerance, show it.
        if dt_val != 0:
            out_rows.append(
                {
                    "label": label,
                    "t2s_id": int(t2s_id),
                    "home_team": str(home_team_name or ""),
                    "away_team": str(away_team_name or ""),
                    "period": period,
                    "game_s": gsec,
                    "time": tg.get("t_str") or "",
                    "team": tm_side,
                    "issue": "Time mismatch",
                    "t2s_scorer": tg.get("scorer") or "",
                    "long_scorer": best.get("scorer") or "",
                    "t2s_assists": _fmt_assists(tg.get("assists") or []),
                    "long_assists": _fmt_assists(best.get("assists") or []),
                    "dt_s": str(dt_val),
                }
            )

        t2s_scorer = str(tg.get("scorer") or "").strip()
        long_scorer = str(best.get("scorer") or "").strip()
        if t2s_scorer and long_scorer and t2s_scorer != long_scorer:
            out_rows.append(
                {
                    "label": label,
                    "t2s_id": int(t2s_id),
                    "home_team": str(home_team_name or ""),
                    "away_team": str(away_team_name or ""),
                    "period": period,
                    "game_s": gsec,
                    "time": tg.get("t_str") or "",
                    "team": tm_side,
                    "issue": "Scorer mismatch",
                    "t2s_scorer": t2s_scorer,
                    "long_scorer": long_scorer,
                    "t2s_assists": _fmt_assists(tg.get("assists") or []),
                    "long_assists": _fmt_assists(best.get("assists") or []),
                    "dt_s": str(dt_val) if dt_val else "",
                }
            )

        t2s_ast = [str(x).strip() for x in (tg.get("assists") or []) if str(x).strip()]
        long_ast = [str(x).strip() for x in (best.get("assists") or []) if str(x).strip()]
        if set(t2s_ast) != set(long_ast):
            missing = sorted([x for x in t2s_ast if x and x not in set(long_ast)])
            extra = sorted([x for x in long_ast if x and x not in set(t2s_ast)])
            out_rows.append(
                {
                    "label": label,
                    "t2s_id": int(t2s_id),
                    "home_team": str(home_team_name or ""),
                    "away_team": str(away_team_name or ""),
                    "period": period,
                    "game_s": gsec,
                    "time": tg.get("t_str") or "",
                    "team": tm_side,
                    "issue": (
                        "Assist mismatch"
                        + (f" (-{','.join(missing)})" if missing else "")
                        + (f" (+{','.join(extra)})" if extra else "")
                    ),
                    "t2s_scorer": t2s_scorer,
                    "long_scorer": long_scorer,
                    "t2s_assists": _fmt_assists(t2s_ast),
                    "long_assists": _fmt_assists(long_ast),
                    "dt_s": str(dt_val) if dt_val else "",
                }
            )

    for lg in unmatched_long:
        out_rows.append(
            {
                "label": label,
                "t2s_id": int(t2s_id),
                "home_team": str(home_team_name or ""),
                "away_team": str(away_team_name or ""),
                "period": int(lg.get("period") or 0),
                "game_s": int(lg.get("game_s") or 0),
                "time": lg.get("t_str") or "",
                "team": str(lg.get("kind") or lg.get("team_raw") or ""),
                "issue": "Long-only goal (no T2S match)",
                "t2s_scorer": "",
                "long_scorer": lg.get("scorer") or "",
                "t2s_assists": "",
                "long_assists": _fmt_assists(lg.get("assists") or []),
                "dt_s": "",
            }
        )

    out_rows.sort(
        key=lambda r: (
            str(r.get("label") or ""),
            int(r.get("period") or 0),
            int(r.get("game_s") or 0),
        )
    )
    return out_rows


def _print_goal_discrepancy_rich_table(rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return
    try:
        from rich.console import Console  # type: ignore
        from rich.table import Table  # type: ignore
        from rich.text import Text  # type: ignore
    except Exception:
        # Fallback: one-line summaries.
        by_game: Dict[str, int] = {}
        for r in rows:
            by_game[str(r.get("label") or "")] = by_game.get(str(r.get("label") or ""), 0) + 1
        items = ", ".join([f"{k}={v}" for k, v in sorted(by_game.items())])
        print(f"[t2s-vs-long] goal discrepancies: {items}", file=sys.stderr)
        return

    console = Console(file=sys.stderr)
    table = Table(title="TimeToScore vs Long-Sheet Goal/Assist Discrepancies", show_lines=False)
    table.add_column("Game", no_wrap=True)
    table.add_column("T2S", justify="right", no_wrap=True)
    table.add_column("P", justify="right", no_wrap=True)
    table.add_column("Time", justify="right", no_wrap=True)
    table.add_column("Team", no_wrap=True)
    table.add_column("Issue")
    table.add_column("T2S Scorer", justify="right", no_wrap=True)
    table.add_column("Long Scorer", justify="right", no_wrap=True)
    table.add_column("T2S Ast", no_wrap=True)
    table.add_column("Long Ast", no_wrap=True)
    table.add_column("Δt(s)", justify="right", no_wrap=True)

    def _issue_style(issue: str) -> str:
        s = str(issue or "").strip().lower()
        if s.startswith("t2s-only goal"):
            return "bold red"
        if s.startswith("long-only goal"):
            return "bold yellow"
        if s.startswith("scorer mismatch"):
            return "bold magenta"
        if s.startswith("assist mismatch"):
            return "bold cyan"
        if s.startswith("time mismatch"):
            return "dim"
        return ""

    for r in rows:
        issue = str(r.get("issue") or "")
        issue_style = _issue_style(issue)
        team = str(r.get("team") or "")
        team_style = (
            "green" if team.lower() == "home" else ("blue" if team.lower() == "away" else "")
        )

        t2s_scorer = str(r.get("t2s_scorer") or "")
        long_scorer = str(r.get("long_scorer") or "")
        scorer_mismatch = (
            bool(t2s_scorer.strip()) and bool(long_scorer.strip()) and t2s_scorer != long_scorer
        )

        t2s_ast_str = str(r.get("t2s_assists") or "")
        long_ast_str = str(r.get("long_assists") or "")

        def _assist_tokens(s: str) -> List[str]:
            parts = [p.strip() for p in str(s or "").split(",")]
            return [p for p in parts if p]

        t2s_ast_tokens = _assist_tokens(t2s_ast_str)
        long_ast_tokens = _assist_tokens(long_ast_str)
        t2s_ast_set = set(t2s_ast_tokens)
        long_ast_set = set(long_ast_tokens)
        assist_mismatch = t2s_ast_set != long_ast_set

        def _render_assists(tokens: List[str], other_set: set[str]) -> Text:
            txt = Text()
            for i, tok in enumerate(tokens):
                if i:
                    txt.append(",")
                if tok not in other_set:
                    txt.append(tok, style="bold cyan")
                else:
                    txt.append(tok)
            return txt

        table.add_row(
            Text(str(r.get("label") or ""), style="bold"),
            str(r.get("t2s_id") or ""),
            str(r.get("period") or ""),
            str(r.get("time") or ""),
            Text(team, style=team_style) if team_style else team,
            Text(issue, style=issue_style) if issue_style else issue,
            Text(t2s_scorer, style="bold red") if scorer_mismatch else t2s_scorer,
            Text(long_scorer, style="bold red") if scorer_mismatch else long_scorer,
            _render_assists(t2s_ast_tokens, long_ast_set) if assist_mismatch else t2s_ast_str,
            _render_assists(long_ast_tokens, t2s_ast_set) if assist_mismatch else long_ast_str,
            Text(str(r.get("dt_s") or ""), style="dim") if str(r.get("dt_s") or "").strip() else "",
        )
    console.print(table)

    # Under the table, print the team name mapping (from TimeToScore) per game.
    try:
        games: Dict[Tuple[str, str], Tuple[str, str]] = {}
        for r in rows:
            label = str(r.get("label") or "")
            t2s_id = str(r.get("t2s_id") or "")
            home_team = str(r.get("home_team") or "")
            away_team = str(r.get("away_team") or "")
            if not label and not t2s_id:
                continue
            if (label, t2s_id) not in games:
                games[(label, t2s_id)] = (home_team, away_team)
        if games:
            console.print("")
            for (label, t2s_id), (home_team, away_team) in sorted(
                games.items(), key=lambda x: (x[0][0], x[0][1])
            ):
                prefix = f"{label} (t2s={t2s_id})".strip()
                if prefix:
                    console.print(Text(prefix, style="bold"))
                if home_team:
                    console.print(Text(f"  Home: {home_team}", style="green"))
                if away_team:
                    console.print(Text(f"  Away: {away_team}", style="blue"))
    except Exception:
        pass


def _write_team_stats_from_long_shift_team(
    *,
    game_out_root: Path,
    format_dir: str,
    team_side: str,
    team_name: str,
    long_shift_tables_by_team: Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]],
    goals: List[GoalEvent],
    event_log_context: Optional["EventLogContext"],
    focus_team: Optional[str],
    include_shifts_in_stats: bool,
    write_shift_rows_csv: bool = False,
    xls_path: Path,
    t2s_rosters_by_side: Optional[Dict[str, Dict[str, str]]] = None,
    create_scripts: bool = False,
    skip_if_exists: bool = False,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
    List[Dict[str, Any]],
]:
    """
    Write stats for a specific long-sheet embedded shift table team into Home/Away/<format_dir>/stats.
    Returns (outdir, stats_rows, periods, per_player_goal_events, pair_on_ice_rows).
    """
    side_l = str(team_side or "").strip().lower()
    if side_l not in {"home", "away"}:
        raise ValueError(f"invalid team_side: {team_side!r}")
    team_subdir = "Away" if side_l == "away" else "Home"
    outdir = Path(game_out_root) / team_subdir / str(format_dir or "per_player")
    stats_dir = outdir / "stats"
    stats_dir.mkdir(parents=True, exist_ok=True)

    if skip_if_exists:
        try:
            existing = stats_dir / "player_stats.csv"
            if existing.exists() and existing.is_file() and existing.stat().st_size > 0:
                return outdir, [], [], {}, []
        except Exception:
            pass

    info = (long_shift_tables_by_team or {}).get(str(team_name)) or {}
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]] = dict(
        (info.get("sb_pairs_by_player") or {})
    )
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]] = dict(
        (info.get("video_pairs_by_player") or {})
    )
    if not sb_pairs_by_player:
        return outdir, [], [], {}, []

    # Optionally add TimeToScore roster-only players for GP accounting.
    if t2s_rosters_by_side and side_l in {"home", "away"}:
        roster = t2s_rosters_by_side.get(side_l) or {}
        seen_normals: set[str] = set()
        for pk in sb_pairs_by_player.keys():
            norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
            if norm:
                seen_normals.add(norm)
        for jersey_norm, name in (roster or {}).items():
            if not jersey_norm or not name:
                continue
            if jersey_norm in seen_normals:
                continue
            player_key = f"{sanitize_name(jersey_norm)}_{sanitize_name(name)}"
            if player_key in sb_pairs_by_player:
                continue
            sb_pairs_by_player[player_key] = []

    if include_shifts_in_stats:
        try:
            _write_video_times_and_scripts(
                outdir, video_pairs_by_player, create_scripts=create_scripts
            )
            _write_scoreboard_times(outdir, sb_pairs_by_player, create_scripts=create_scripts)
        except Exception:
            pass

    # Build jersey->player mapping for scoring attribution.
    jersey_to_players: Dict[str, List[str]] = {}
    for pk in sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
        if norm:
            jersey_to_players.setdefault(norm, []).append(pk)

    goals2 = list(goals or [])
    goals2.sort(key=lambda e: (e.period, e.t_sec))
    _annotate_goal_roles(goals2)

    # Determine which long-sheet/player-attributed event types exist in this game so we can
    # leave blank when the stat type wasn't collected (not 0).
    player_event_types_present: set[str] = set()
    if event_log_context is not None:
        for counts in (event_log_context.event_counts_by_player or {}).values():
            for et in (counts or {}).keys():
                if et:
                    player_event_types_present.add(str(et))
    has_player_shots = "Shot" in player_event_types_present
    has_player_sog = "SOG" in player_event_types_present
    has_player_expected_goals = "ExpectedGoal" in player_event_types_present
    has_player_turnovers_forced = "TurnoverForced" in player_event_types_present
    has_player_created_turnovers = "CreatedTurnover" in player_event_types_present
    has_player_giveaways = "Giveaway" in player_event_types_present
    has_player_takeaways = "Takeaway" in player_event_types_present

    has_controlled_entry_events = False
    has_controlled_exit_events = False
    if focus_team in {"Blue", "White"} and event_log_context is not None:
        for etype, _team in (event_log_context.event_instances or {}).keys():
            if etype == "ControlledEntry":
                has_controlled_entry_events = True
            elif etype == "ControlledExit":
                has_controlled_exit_events = True

    # Pre-group team-level events by period for on-ice for/against counts.
    on_ice_event_types = {"ControlledEntry", "ControlledExit"}
    team_events_by_period: Dict[int, List[Tuple[str, str, int]]] = {}
    if focus_team is not None and event_log_context is not None:
        for (etype, team), inst_list in (event_log_context.event_instances or {}).items():
            if etype not in on_ice_event_types:
                continue
            for it in inst_list or []:
                p = it.get("period")
                gs = it.get("game_s")
                if not isinstance(p, int) or not isinstance(gs, (int, float)):
                    continue
                team_events_by_period.setdefault(int(p), []).append(
                    (str(etype), str(team), int(gs))
                )

    goals_by_period: Dict[int, List[GoalEvent]] = {}
    for ev in goals2 or []:
        goals_by_period.setdefault(int(ev.period), []).append(ev)

    pair_on_ice_rows: List[Dict[str, Any]] = []
    if sb_pairs_by_player:
        try:
            pair_on_ice_rows = _compute_pair_on_ice_rows(sb_pairs_by_player, goals_by_period)
            _write_pair_on_ice_csv(stats_dir, pair_on_ice_rows, include_toi=include_shifts_in_stats)
        except Exception:
            pair_on_ice_rows = []

    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]] = {
        pk: {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []}
        for pk in sb_pairs_by_player.keys()
    }
    goal_assist_counts: Dict[str, Dict[str, int]] = {
        pk: {"goals": 0, "assists": 0} for pk in sb_pairs_by_player.keys()
    }

    def _match_player_keys(num_token: Any) -> List[str]:
        matches: List[str] = []
        candidates2: set[str] = set()
        if num_token is not None:
            try:
                txt = str(num_token).strip()
                if txt:
                    candidates2.add(txt)
            except Exception:
                pass
        norm2 = _normalize_jersey_number(num_token)
        if norm2:
            candidates2.add(norm2)
        for cand in candidates2:
            matches.extend(jersey_to_players.get(cand, []))
        return list(dict.fromkeys(matches))

    for ev in goals2:
        if ev.kind != "GF":
            continue
        if ev.scorer:
            for pk in _match_player_keys(ev.scorer):
                goal_assist_counts[pk]["goals"] += 1
                per_player_goal_events[pk]["goals"].append(ev)
        for ast in ev.assists:
            for pk in _match_player_keys(ast):
                goal_assist_counts[pk]["assists"] += 1
                per_player_goal_events[pk]["assists"].append(ev)

    stats_table_rows: List[Dict[str, str]] = []
    all_periods_seen: set[int] = set()

    for player_key, sb_list in sb_pairs_by_player.items():
        sb_by_period: Dict[int, List[Tuple[str, str]]] = {}
        for period, a, b in sb_list or []:
            sb_by_period.setdefault(int(period), []).append((str(a), str(b)))
        for period in sb_by_period.keys():
            all_periods_seen.add(period)
        all_pairs = [(a, b) for (_, a, b) in sb_list]
        if include_shifts_in_stats:
            shift_summary = summarize_shift_lengths_sec(all_pairs)
            per_period_toi_map = per_period_toi(sb_by_period)
        else:
            shift_summary = {}
            per_period_toi_map = {}

        plus_minus = 0
        counted_gf: List[str] = []
        counted_ga: List[str] = []
        counted_gf_by_period: Dict[int, int] = {}
        counted_ga_by_period: Dict[int, int] = {}
        for period, pairs in sb_by_period.items():
            if period not in goals_by_period:
                continue
            for ev in goals_by_period[period]:
                matched = False
                for a, b in pairs:
                    a_sec = parse_flex_time_to_seconds(a)
                    b_sec = parse_flex_time_to_seconds(b)
                    lo, hi = (a_sec, b_sec) if a_sec <= b_sec else (b_sec, a_sec)
                    if not (lo <= ev.t_sec <= hi):
                        continue
                    if ev.t_sec == a_sec:
                        continue
                    matched = True
                    break
                if matched:
                    if ev.kind == "GF":
                        plus_minus += 1
                        counted_gf.append(f"P{period}:{ev.t_str}")
                        counted_gf_by_period[period] = counted_gf_by_period.get(period, 0) + 1
                        per_player_goal_events[player_key]["gf_on_ice"].append(ev)
                    else:
                        plus_minus -= 1
                        counted_ga.append(f"P{period}:{ev.t_str}")
                        counted_ga_by_period[period] = counted_ga_by_period.get(period, 0) + 1
                        per_player_goal_events[player_key]["ga_on_ice"].append(ev)

        scoring_counts = goal_assist_counts.get(player_key, {"goals": 0, "assists": 0})
        goals_cnt = int(scoring_counts.get("goals", 0) or 0)
        assists_cnt = int(scoring_counts.get("assists", 0) or 0)
        points_val = goals_cnt + assists_cnt

        # Overtime scoring (OT period is period 4; OT2 -> 5, etc.).
        ot_goals_cnt = 0
        ot_assists_cnt = 0
        try:
            ev_map = per_player_goal_events.get(player_key, {}) or {}
            ot_goals_cnt = sum(
                1 for ev in (ev_map.get("goals") or []) if int(getattr(ev, "period", 0) or 0) >= 4
            )
            ot_assists_cnt = sum(
                1 for ev in (ev_map.get("assists") or []) if int(getattr(ev, "period", 0) or 0) >= 4
            )
        except Exception:
            ot_goals_cnt = 0
            ot_assists_cnt = 0

        # Game-tying / game-winning goals for this game.
        gt_goals_cnt = 0
        gw_goals_cnt = 0
        try:
            ev_map = per_player_goal_events.get(player_key, {}) or {}
            gt_goals_cnt, gw_goals_cnt = _count_goal_role_flags(list(ev_map.get("goals") or []))
        except Exception:
            gt_goals_cnt = 0
            gw_goals_cnt = 0

        # On-ice for/against metrics for team-level events (e.g., controlled exits).
        on_ice: Dict[str, int] = {
            "controlled_entry_for": 0,
            "controlled_entry_against": 0,
            "controlled_exit_for": 0,
            "controlled_exit_against": 0,
        }
        if focus_team is not None and team_events_by_period and sb_by_period:
            for period, pairs in sb_by_period.items():
                period_events = team_events_by_period.get(period, [])
                if not period_events or not pairs:
                    continue
                intervals = [compute_interval_seconds(a, b) for a, b in pairs]
                for etype, team, gsec in period_events:
                    in_any = False
                    for lo, hi in intervals:
                        if lo <= gsec <= hi:
                            in_any = True
                            break
                    if not in_any:
                        continue
                    is_for = team == focus_team
                    if etype == "ControlledEntry":
                        key = "controlled_entry_for" if is_for else "controlled_entry_against"
                        on_ice[key] += 1
                    elif etype == "ControlledExit":
                        key = "controlled_exit_for" if is_for else "controlled_exit_against"
                        on_ice[key] += 1

        row_map: Dict[str, str] = {
            "player": player_key,
            "goals": str(goals_cnt),
            "assists": str(assists_cnt),
            "gt_goals": str(gt_goals_cnt),
            "gw_goals": str(gw_goals_cnt),
            "ot_goals": str(ot_goals_cnt),
            "ot_assists": str(ot_assists_cnt),
            "points": str(points_val),
            "gp": "1",
            "plus_minus": str(plus_minus),
        }
        if include_shifts_in_stats:
            shifts_cnt_row = 0
            try:
                shifts_cnt_row = int(str(shift_summary.get("num_shifts", "0") or 0))
            except Exception:
                shifts_cnt_row = 0
            row_map["shifts"] = str(shifts_cnt_row)
            row_map["sb_toi_total"] = str(shift_summary.get("toi_total", "0:00"))
            row_map["sb_avg"] = str(shift_summary.get("toi_avg", "0:00"))
            row_map["sb_median"] = str(shift_summary.get("toi_median", "0:00"))
            row_map["sb_longest"] = str(shift_summary.get("toi_longest", "0:00"))
            row_map["sb_shortest"] = str(shift_summary.get("toi_shortest", "0:00"))

        # Event counts (from event logs / long sheets), per game.
        if event_log_context is not None:
            ev_counts = (event_log_context.event_counts_by_player or {}).get(player_key, {})
        else:
            ev_counts = {}
        has_completed_passes_local = "CompletedPass" in player_event_types_present
        shots_cnt = int(ev_counts.get("Shot", 0) or 0)
        sog_cnt = int(ev_counts.get("SOG", 0) or 0)
        # When SOG isn't explicitly collected, fall back to Shots.
        if not has_player_sog and has_player_shots:
            sog_cnt = shots_cnt
            if sog_cnt > 0 and "SOG" not in ev_counts:
                ev_counts = dict(ev_counts)
                ev_counts["SOG"] = sog_cnt
        expected_goals_cnt = int(ev_counts.get("ExpectedGoal", 0) or 0)
        turnovers_forced_cnt = int(ev_counts.get("TurnoverForced", 0) or 0)
        created_turnovers_cnt = int(ev_counts.get("CreatedTurnover", 0) or 0)
        giveaways_cnt = int(ev_counts.get("Giveaway", 0) or 0)
        takeaways_cnt = int(ev_counts.get("Takeaway", 0) or 0)
        completed_passes_cnt = int(ev_counts.get("CompletedPass", 0) or 0)

        # Per-player stats file (for parity with primary sheet outputs).
        try:
            stats_lines: List[str] = []
            stats_lines.append(f"Player: {_display_player_name(player_key)}")
            stats_lines.append("")
            stats_lines.append("Goals:")
            stats_lines.append(f"  Goals: {goals_cnt}")
            stats_lines.append(f"  Assists: {assists_cnt}")
            stats_lines.append(f"  Points (G+A): {points_val}")
            stats_lines.append(f"  OT Goals: {ot_goals_cnt}")
            stats_lines.append(f"  OT Assists: {ot_assists_cnt}")
            stats_lines.append(f"  GT Goals: {gt_goals_cnt}")
            stats_lines.append(f"  GW Goals: {gw_goals_cnt}")
            stats_lines.append("")
            stats_lines.append(f"+/- (on-ice goals only): {plus_minus:+d}")
            stats_lines.append(f"GF counted: {len(counted_gf)}")
            stats_lines.append(f"GA counted: {len(counted_ga)}")

            if include_shifts_in_stats and shift_summary:
                stats_lines.append("")
                stats_lines.append("Shifts (from long sheet):")
                stats_lines.append(f"  Shifts: {shift_summary.get('num_shifts', 0)}")
                stats_lines.append(f"  TOI (scoreboard): {shift_summary.get('toi_total', '0:00')}")
                stats_lines.append(f"  Avg: {shift_summary.get('toi_avg', '0:00')}")
                stats_lines.append(f"  Median: {shift_summary.get('toi_median', '0:00')}")
                stats_lines.append(f"  Longest: {shift_summary.get('toi_longest', '0:00')}")
                stats_lines.append(f"  Shortest: {shift_summary.get('toi_shortest', '0:00')}")

            if per_period_toi_map:
                stats_lines.append("")
                stats_lines.append("Per-period:")
                for period, toi in sorted(per_period_toi_map.items(), key=lambda x: int(x[0])):
                    num_shifts = len(sb_by_period.get(int(period), []) or [])
                    stats_lines.append(f"  Period {int(period)}: TOI={toi}, shifts={num_shifts}")

            ev_map = per_player_goal_events.get(player_key, {}) or {}
            goals_list = ev_map.get("goals") or []
            assists_list = ev_map.get("assists") or []
            if goals_list:
                stats_lines.append("")
                stats_lines.append("Goals timeline:")
                for ev in sorted(
                    goals_list,
                    key=lambda e: (
                        int(getattr(e, "period", 0) or 0),
                        int(getattr(e, "t_sec", 0) or 0),
                    ),
                ):
                    tags: List[str] = []
                    if getattr(ev, "is_game_tying", False):
                        tags.append("GT")
                    if getattr(ev, "is_game_winning", False):
                        tags.append("GW")
                    tag_str = f" [{' '.join(tags)}]" if tags else ""
                    stats_lines.append(f"  Period {int(ev.period)}, {ev.t_str}{tag_str}")
            if assists_list:
                stats_lines.append("")
                stats_lines.append("Assists timeline:")
                for ev in sorted(
                    assists_list,
                    key=lambda e: (
                        int(getattr(e, "period", 0) or 0),
                        int(getattr(e, "t_sec", 0) or 0),
                    ),
                ):
                    tags2: List[str] = []
                    if getattr(ev, "is_game_tying", False):
                        tags2.append("GT")
                    if getattr(ev, "is_game_winning", False):
                        tags2.append("GW")
                    tag_str2 = f" [{' '.join(tags2)}]" if tags2 else ""
                    stats_lines.append(f"  Period {int(ev.period)}, {ev.t_str}{tag_str2}")

            if event_log_context is not None and ev_counts:
                stats_lines.append("")
                stats_lines.append("Event Counts:")
                order = [
                    "Shot",
                    "SOG",
                    "Goal",
                    "Assist",
                    "ExpectedGoal",
                    "TurnoverForced",
                    "CreatedTurnover",
                    "Giveaway",
                    "Takeaway",
                    "CompletedPass",
                    "ControlledEntry",
                    "ControlledExit",
                ]
                for kind in order:
                    if kind in ev_counts and int(ev_counts.get(kind, 0) or 0) > 0:
                        stats_lines.append(
                            f"  {_display_event_type(kind)}: {int(ev_counts.get(kind, 0) or 0)}"
                        )
                for kind, cnt in sorted(ev_counts.items()):
                    if kind in order:
                        continue
                    if int(cnt or 0) > 0:
                        stats_lines.append(f"  {_display_event_type(str(kind))}: {int(cnt or 0)}")

            if focus_team is not None and any(v > 0 for v in on_ice.values()):
                stats_lines.append("")
                stats_lines.append("On-ice team events (for/against):")
                stats_lines.append(
                    f"  ControlledEntry: {on_ice['controlled_entry_for']} for, {on_ice['controlled_entry_against']} against"
                )
                stats_lines.append(
                    f"  ControlledExit: {on_ice['controlled_exit_for']} for, {on_ice['controlled_exit_against']} against"
                )

            if include_shifts_in_stats:
                for period, pairs in sorted(sb_by_period.items()):
                    stats_lines.append(f"Shifts in Period {period}: {len(pairs)}")

            (stats_dir / f"{player_key}_stats.txt").write_text(
                "\n".join(stats_lines) + "\n", encoding="utf-8"
            )
        except Exception:
            pass

        row_map["shots"] = str(shots_cnt) if has_player_shots else ""
        row_map["sog"] = (
            str(sog_cnt) if (has_player_sog or (has_player_shots and sog_cnt >= 0)) else ""
        )
        row_map["expected_goals"] = str(expected_goals_cnt) if has_player_expected_goals else ""
        if has_player_expected_goals and (has_player_sog or has_player_shots):
            row_map["expected_goals_per_sog"] = (
                f"{(expected_goals_cnt / sog_cnt):.2f}" if sog_cnt > 0 else ""
            )
        else:
            row_map["expected_goals_per_sog"] = ""
        row_map["turnovers_forced"] = (
            str(turnovers_forced_cnt) if has_player_turnovers_forced else ""
        )
        row_map["created_turnovers"] = (
            str(created_turnovers_cnt) if has_player_created_turnovers else ""
        )
        row_map["giveaways"] = str(giveaways_cnt) if has_player_giveaways else ""
        row_map["takeaways"] = str(takeaways_cnt) if has_player_takeaways else ""
        row_map["completed_passes"] = (
            str(completed_passes_cnt) if has_completed_passes_local else ""
        )

        if has_controlled_entry_events:
            row_map["controlled_entry_for"] = str(on_ice["controlled_entry_for"])
            row_map["controlled_entry_against"] = str(on_ice["controlled_entry_against"])
        else:
            row_map["controlled_entry_for"] = ""
            row_map["controlled_entry_against"] = ""
        if has_controlled_exit_events:
            row_map["controlled_exit_for"] = str(on_ice["controlled_exit_for"])
            row_map["controlled_exit_against"] = str(on_ice["controlled_exit_against"])
        else:
            row_map["controlled_exit_for"] = ""
            row_map["controlled_exit_against"] = ""

        row_map["gf_counted"] = str(len(counted_gf))
        row_map["ga_counted"] = str(len(counted_ga))
        if include_shifts_in_stats:
            v_pairs = video_pairs_by_player.get(player_key, [])
            if v_pairs:
                v_sum = 0
                for a, b in v_pairs:
                    lo, hi = compute_interval_seconds(a, b)
                    v_sum += hi - lo
                row_map["video_toi_total"] = _format_duration(v_sum)
            else:
                row_map["video_toi_total"] = ""
            for period, toi in per_period_toi_map.items():
                row_map[f"P{period}_toi"] = toi
                all_periods_seen.add(int(period))
            for period, pairs in sb_by_period.items():
                row_map[f"P{period}_shifts"] = str(len(pairs))
        for period, cnt in counted_gf_by_period.items():
            row_map[f"P{period}_GF"] = str(cnt)
            all_periods_seen.add(int(period))
        for period, cnt in counted_ga_by_period.items():
            row_map[f"P{period}_GA"] = str(cnt)
            all_periods_seen.add(int(period))

        stats_table_rows.append(row_map)

    if include_shifts_in_stats or write_shift_rows_csv:
        try:
            _write_shift_rows_csv(
                stats_dir,
                sb_pairs_by_player=sb_pairs_by_player,
                video_pairs_by_player=video_pairs_by_player,
                source="shift_spreadsheet",
            )
        except Exception as exc:  # noqa: BLE001
            print(f"Warning: failed to write shift_rows.csv: {exc!r}", file=sys.stderr)
    if include_shifts_in_stats:
        _write_global_summary_csv(stats_dir, sb_pairs_by_player)

    period_list = sorted(all_periods_seen)
    if stats_table_rows:
        _write_player_stats_text_and_csv(
            stats_dir,
            stats_table_rows,
            period_list,
            include_shifts_in_stats=include_shifts_in_stats,
        )

    _write_game_stats_files(
        stats_dir,
        xls_path=xls_path,
        periods=period_list,
        goals=goals2,
        event_log_context=event_log_context,
        focus_team=focus_team,
    )

    # Best-effort scoreboard->video conversion segments derived from long-sheet shift times.
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]] = {}
    try:
        for pk, sb_list in (sb_pairs_by_player or {}).items():
            v_list = (video_pairs_by_player or {}).get(pk) or []
            nseg = min(len(sb_list or []), len(v_list or []))
            for idx in range(nseg):
                try:
                    per, sba, sbb = sb_list[idx]
                    sva, svb = v_list[idx]
                    p_i = int(per)
                    s1 = parse_flex_time_to_seconds(str(sba))
                    s2 = parse_flex_time_to_seconds(str(sbb))
                    v1 = parse_flex_time_to_seconds(str(sva))
                    v2 = parse_flex_time_to_seconds(str(svb))
                    conv_segments_by_period.setdefault(p_i, []).append((s1, s2, v1, v2))
                except Exception:
                    continue
    except Exception:
        conv_segments_by_period = {}

    # Goals windows + event summaries (mirror primary sheet output structure).
    try:
        _write_goal_window_files(outdir, goals2, conv_segments_by_period)
    except Exception:
        pass

    if event_log_context is not None:
        try:
            _write_event_summaries_and_clips(
                outdir,
                stats_dir,
                event_log_context,
                conv_segments_by_period,
                create_scripts=create_scripts,
                focus_team=focus_team,
            )
            _write_player_event_highlights(
                outdir,
                event_log_context,
                conv_segments_by_period,
                sb_pairs_by_player.keys(),
                create_scripts=create_scripts,
            )
            _write_player_combined_highlights(
                outdir,
                event_log_context=event_log_context,
                conv_segments_by_period=conv_segments_by_period,
                per_player_goal_events=per_player_goal_events,
                player_keys=sb_pairs_by_player.keys(),
                create_scripts=create_scripts,
            )
        except Exception:
            pass
    else:
        try:
            _write_player_combined_highlights(
                outdir,
                event_log_context=None,
                conv_segments_by_period=conv_segments_by_period,
                per_player_goal_events=per_player_goal_events,
                player_keys=sb_pairs_by_player.keys(),
                create_scripts=create_scripts,
            )
        except Exception:
            pass

    try:
        _write_clip_all_runner(outdir, create_scripts=create_scripts)
    except Exception:
        pass

    return outdir, stats_table_rows, period_list, per_player_goal_events, pair_on_ice_rows


def _write_opponent_team_stats_from_long_shifts(
    *,
    game_out_root: Path,
    format_dir: str,
    our_side: str,
    long_shift_tables_by_team: Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]],
    shift_cmp_summary: Dict[str, Any],
    goals: List[GoalEvent],
    event_log_context: Optional["EventLogContext"],
    focus_team: Optional[str],
    include_shifts_in_stats: bool,
    write_shift_rows_csv: bool = False,
    xls_path: Path,
    t2s_rosters_by_side: Optional[Dict[str, Dict[str, str]]] = None,
    create_scripts: bool = False,
    skip_if_exists: bool = False,
) -> Optional[Path]:
    """
    Best-effort: when a '*-long*' sheet provides embedded shift tables for both teams, write
    stats outputs for the opponent team into the opposite Home/Away subtree.

    This is primarily used to populate the opponent's player stats in the webapp import.
    """
    matched_team = str((shift_cmp_summary or {}).get("matched_team") or "").strip()
    if not matched_team:
        return None

    candidates = [t for t in (long_shift_tables_by_team or {}).keys() if str(t) != matched_team]
    if not candidates:
        return None

    def _total_shifts(team_name: str) -> int:
        info = (long_shift_tables_by_team or {}).get(team_name) or {}
        sb_map = (info.get("sb_pairs_by_player") or {}) if isinstance(info, dict) else {}
        return sum(len(v or []) for v in (sb_map or {}).values())

    opponent_team = max(candidates, key=_total_shifts)

    our_side_l = str(our_side or "").strip().lower()
    if our_side_l not in {"home", "away"}:
        return None
    opp_side_l = "away" if our_side_l == "home" else "home"

    # Flip scoring perspective (their GF is our GA, and vice-versa).
    opp_goals: List[GoalEvent] = []
    for g in goals or []:
        kind = str(getattr(g, "kind", "") or "").strip().upper()
        if kind == "GF":
            kind2 = "GA"
        elif kind == "GA":
            kind2 = "GF"
        else:
            kind2 = kind
        opp_goals.append(
            GoalEvent(
                kind2,
                int(getattr(g, "period", 0) or 0),
                str(getattr(g, "t_str", "") or ""),
                scorer=g.scorer,
                assists=list(g.assists or []),
            )
        )

    opp_focus_team: Optional[str] = None
    if focus_team in {"Blue", "White"}:
        opp_focus_team = "White" if focus_team == "Blue" else "Blue"

    outdir2, _rows, _periods, _events, _pair_rows = _write_team_stats_from_long_shift_team(
        game_out_root=game_out_root,
        format_dir=format_dir,
        team_side=str(opp_side_l),
        team_name=str(opponent_team),
        long_shift_tables_by_team=long_shift_tables_by_team,
        goals=opp_goals,
        event_log_context=event_log_context,
        focus_team=opp_focus_team,
        include_shifts_in_stats=include_shifts_in_stats,
        write_shift_rows_csv=bool(write_shift_rows_csv),
        xls_path=xls_path,
        t2s_rosters_by_side=t2s_rosters_by_side,
        create_scripts=create_scripts,
        skip_if_exists=skip_if_exists,
    )
    return outdir2


def _infer_focus_team_from_long_sheet(
    our_jerseys: set[str],
    jerseys_by_team: Dict[str, set[int]],
) -> Optional[str]:
    if not our_jerseys or not jerseys_by_team:
        return None
    blue = {str(x) for x in jerseys_by_team.get("Blue", set())}
    white = {str(x) for x in jerseys_by_team.get("White", set())}
    blue_ov = len(our_jerseys & blue)
    white_ov = len(our_jerseys & white)
    if blue_ov == 0 and white_ov == 0:
        return None
    if blue_ov == white_ov:
        return None
    return "Blue" if blue_ov > white_ov else "White"


def _infer_focus_team_from_color_rosters(
    our_jerseys: set[str],
    jerseys_by_team: Dict[str, Iterable[int]],
) -> Optional[str]:
    """
    Infer whether "us" is Blue or White by roster overlap.

    This is used when the primary sheet is an event-log layout (no names) and we have
    TimeToScore roster numbers for our team, or when long-sheet inference isn't available.
    """
    if not our_jerseys or not jerseys_by_team:
        return None
    blue = {str(int(x)) for x in (jerseys_by_team.get("Blue") or []) if isinstance(x, (int, float))}
    white = {
        str(int(x)) for x in (jerseys_by_team.get("White") or []) if isinstance(x, (int, float))
    }
    blue_ov = len(our_jerseys & blue)
    white_ov = len(our_jerseys & white)
    if blue_ov == 0 and white_ov == 0:
        return None
    if blue_ov == white_ov:
        return None
    return "Blue" if blue_ov > white_ov else "White"


def _extract_roster_tables_from_df(
    df: pd.DataFrame,
) -> List[Tuple[Optional[str], Dict[str, str]]]:
    """
    Best-effort extraction of roster tables from a sheet.

    We look for header rows that contain both a jersey/number column and a name/player column,
    and then parse subsequent rows as (jersey -> name).

    Returns a list of (team_color_or_none, roster_map) entries. The team is inferred only when
    nearby cells mention "Blue" or "White"; otherwise it is None and callers can assign based
    on jersey overlap with other signals.
    """
    if df is None or df.empty:
        return []

    def _norm_cell(v: Any) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        try:
            s = str(v).strip().lower()
        except Exception:
            return ""
        s = s.replace("\xa0", " ")
        s = " ".join(s.split())
        # Preserve '#' so roster headers like "#" can be detected.
        return re.sub(r"[^a-z0-9#]+", "", s)

    def _scan_team_near(r: int) -> Optional[str]:
        # Look in the header row and a couple of rows above for Blue/White labels.
        found: set[str] = set()
        for rr in range(max(0, r - 2), min(df.shape[0], r + 1)):
            for cc in range(df.shape[1]):
                v = df.iat[rr, cc]
                if pd.isna(v) or not isinstance(v, str):
                    continue
                sl = v.strip().lower()
                if "blue" in sl or sl.startswith("blu"):
                    found.add("Blue")
                if "white" in sl or sl.startswith("whi"):
                    found.add("White")
        if len(found) == 1:
            return next(iter(found))
        return None

    out: List[Tuple[Optional[str], Dict[str, str]]] = []
    # Scan for header rows.
    for r in range(df.shape[0]):
        normed = [_norm_cell(df.iat[r, c]) for c in range(df.shape[1])]
        if not any(normed):
            continue

        jersey_cols = [
            c
            for c, s in enumerate(normed)
            if s and (("jersey" in s) or s in {"number", "num", "#"})
        ]
        name_cols = [
            c
            for c, s in enumerate(normed)
            if s and (("name" in s) or s in {"player", "playername"})
        ]
        if not jersey_cols or not name_cols:
            continue

        # Prefer adjacent columns.
        best = None
        for jc in jersey_cols:
            for nc in name_cols:
                dist = abs(jc - nc)
                if best is None or dist < best[0]:
                    best = (dist, jc, nc)
        if best is None:
            continue
        _, jersey_col, name_col = best

        team_guess = _scan_team_near(r)
        roster: Dict[str, str] = {}
        blank_jersey_streak = 0
        for rr in range(r + 1, df.shape[0]):
            # Stop if we hit another header-like row.
            if rr != r + 1:
                row_normed = [_norm_cell(df.iat[rr, c]) for c in range(df.shape[1])]
                if ("jersey" in "".join(row_normed)) and (
                    "name" in "".join(row_normed) or "player" in "".join(row_normed)
                ):
                    break
            # Stop at period headers (common in these sheets).
            try:
                if parse_period_label(df.iat[rr, 0]) is not None:
                    break
            except Exception:
                # Non-period label or unexpected cell format; ignore and continue scanning.
                pass

            jersey_norm = _normalize_jersey_number(df.iat[rr, jersey_col])
            if not jersey_norm:
                blank_jersey_streak += 1
                if blank_jersey_streak >= 3:
                    break
                continue
            blank_jersey_streak = 0
            try:
                nm_raw = df.iat[rr, name_col]
            except Exception:
                nm_raw = None
            name = str(nm_raw or "").strip()
            if not name or name.lower() in {"nan", "none"}:
                continue
            roster[jersey_norm] = name

        # Require a small minimum to avoid false positives.
        if len(roster) >= 3:
            out.append((team_guess, roster))

    return out


def _resolve_team_color_player_key(
    pk: str,
    roster_name_by_team: Dict[str, Dict[str, str]],
) -> str:
    """
    Resolve placeholder keys like 'Blue_12' to '12_First_Last' when a roster name is known.
    """
    try:
        text = str(pk or "").strip()
    except Exception:
        return pk
    m = re.fullmatch(r"(Blue|White)_(\d+)", text)
    if not m:
        return pk
    team = m.group(1)
    jersey_norm = _normalize_jersey_number(m.group(2))
    if not jersey_norm:
        return pk
    name = (roster_name_by_team.get(team) or {}).get(jersey_norm)
    if not name:
        return pk
    return f"{sanitize_name(jersey_norm)}_{sanitize_name(name)}"


def _rename_dict_keys_merge_lists(
    d: Dict[str, List[Any]],
    key_map: Dict[str, str],
) -> Dict[str, List[Any]]:
    out: Dict[str, List[Any]] = {}
    for k, v in (d or {}).items():
        nk = key_map.get(k, k)
        out.setdefault(nk, []).extend(list(v or []))
    return out


def _rename_event_log_context_players(
    ctx: Optional["EventLogContext"],
    key_map: Dict[str, str],
) -> Optional["EventLogContext"]:
    if ctx is None or not key_map:
        return ctx

    counts_by_player: Dict[str, Dict[str, int]] = {}
    for pk, kinds in (ctx.event_counts_by_player or {}).items():
        nk = key_map.get(pk, pk)
        dest = counts_by_player.setdefault(nk, {})
        for et, cnt in (kinds or {}).items():
            try:
                inc = int(cnt)
            except Exception:
                inc = 0
            dest[et] = dest.get(et, 0) + inc

    player_rows: List[Dict[str, Any]] = []
    for r in ctx.event_player_rows or []:
        rr = dict(r or {})
        pk = str(rr.get("player") or "").strip()
        if pk:
            rr["player"] = key_map.get(pk, pk)
        player_rows.append(rr)

    return EventLogContext(
        event_counts_by_player=counts_by_player,
        event_counts_by_type_team=dict(ctx.event_counts_by_type_team or {}),
        event_instances=dict(ctx.event_instances or {}),
        event_player_rows=player_rows,
        team_roster=dict(ctx.team_roster or {}),
        team_excluded=dict(ctx.team_excluded or {}),
    )


def _event_log_context_from_long_events(
    long_events: List[LongEvent],
    *,
    jersey_to_players: Dict[str, List[str]],
    focus_team: Optional[str],
    jerseys_by_team: Dict[str, set[int]],
    roster_name_by_team: Optional[Dict[str, Dict[str, str]]] = None,
) -> EventLogContext:
    event_counts_by_player: Dict[str, Dict[str, int]] = {}
    event_counts_by_type_team: Dict[Tuple[str, str], int] = {}
    event_instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    event_player_rows: List[Dict[str, Any]] = []

    MAX_TEAM_PLAYERS = 20
    team_roster: Dict[str, List[int]] = {}
    team_excluded: Dict[str, List[int]] = {}

    def _register_and_flag(team: str, jerseys: Iterable[int]) -> None:
        seen_local: set[int] = set()
        ordered: List[int] = []
        for j in jerseys:
            jj = int(j)
            if jj in seen_local:
                continue
            seen_local.add(jj)
            ordered.append(jj)
        roster = team_roster.setdefault(team, [])
        excluded = team_excluded.setdefault(team, [])
        for j in ordered:
            if j in roster:
                continue
            if len(roster) < MAX_TEAM_PLAYERS:
                roster.append(j)
            else:
                if j not in excluded:
                    excluded.append(j)

    # Seed rosters from the observed jerseys for better inference/debugging.
    for team, nums in (jerseys_by_team or {}).items():
        if not team or not nums:
            continue
        _register_and_flag(team, sorted(nums))

    for ev in long_events:
        team = ev.team
        etype = ev.event_type
        key = (etype, team)
        event_counts_by_type_team[key] = event_counts_by_type_team.get(key, 0) + 1
        event_instances.setdefault(key, []).append(
            {"period": ev.period, "video_s": ev.video_s, "game_s": ev.game_s}
        )

        if not ev.jerseys:
            continue

        _register_and_flag(team, ev.jerseys)
        for jersey in ev.jerseys:
            jersey_norm = _normalize_jersey_number(jersey)
            player_keys: List[str] = []
            if focus_team is not None and team == focus_team and jersey_norm:
                player_keys = jersey_to_players.get(jersey_norm, [])
            if not player_keys and jersey_norm and roster_name_by_team:
                nm = (roster_name_by_team.get(team) or {}).get(jersey_norm)
                if nm:
                    player_keys = [f"{sanitize_name(jersey_norm)}_{sanitize_name(nm)}"]
            if not player_keys:
                player_keys = [f"{team}_{int(jersey)}"]
            for pk in player_keys:
                d = event_counts_by_player.setdefault(pk, {})
                d[etype] = d.get(etype, 0) + 1
                event_player_rows.append(
                    {
                        "event_type": etype,
                        "team": team,
                        "player": pk,
                        "jersey": int(jersey),
                        "period": ev.period,
                        "video_s": ev.video_s,
                        "game_s": ev.game_s,
                    }
                )

    return EventLogContext(
        event_counts_by_player=event_counts_by_player,
        event_counts_by_type_team=event_counts_by_type_team,
        event_instances=event_instances,
        event_player_rows=event_player_rows,
        team_roster=team_roster,
        team_excluded=team_excluded,
    )


def _merge_event_log_contexts(
    a: Optional[EventLogContext],
    b: Optional[EventLogContext],
) -> Optional[EventLogContext]:
    if a is None:
        return b
    if b is None:
        return a

    merged_counts_by_player: Dict[str, Dict[str, int]] = {}
    for src in (a.event_counts_by_player or {}, b.event_counts_by_player or {}):
        for pk, kinds in src.items():
            dest = merged_counts_by_player.setdefault(pk, {})
            for kind, cnt in (kinds or {}).items():
                try:
                    inc = int(cnt)
                except Exception:
                    inc = 0
                dest[kind] = dest.get(kind, 0) + inc

    merged_counts_by_type_team: Dict[Tuple[str, str], int] = {}
    for src in (a.event_counts_by_type_team or {}, b.event_counts_by_type_team or {}):
        for k, cnt in src.items():
            try:
                inc = int(cnt)
            except Exception:
                inc = 0
            merged_counts_by_type_team[k] = merged_counts_by_type_team.get(k, 0) + inc

    merged_instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for src in (a.event_instances or {}, b.event_instances or {}):
        for k, rows in src.items():
            merged_instances.setdefault(k, []).extend(list(rows or []))

    merged_player_rows: List[Dict[str, Any]] = []
    merged_player_rows.extend(list(a.event_player_rows or []))
    merged_player_rows.extend(list(b.event_player_rows or []))

    def _merge_rosters(r1: Dict[str, List[int]], r2: Dict[str, List[int]]) -> Dict[str, List[int]]:
        out: Dict[str, List[int]] = {}
        for src in (r1 or {}, r2 or {}):
            for team, nums in src.items():
                if not team:
                    continue
                cur = out.setdefault(team, [])
                for n in nums or []:
                    try:
                        nn = int(n)
                    except Exception:
                        continue
                    if nn not in cur:
                        cur.append(nn)
        return out

    merged_roster = _merge_rosters(a.team_roster or {}, b.team_roster or {})
    merged_excluded = _merge_rosters(a.team_excluded or {}, b.team_excluded or {})

    return EventLogContext(
        event_counts_by_player=merged_counts_by_player,
        event_counts_by_type_team=merged_counts_by_type_team,
        event_instances=merged_instances,
        event_player_rows=merged_player_rows,
        team_roster=merged_roster,
        team_excluded=merged_excluded,
    )


def _detect_event_log_headers(
    df: pd.DataFrame,
) -> Tuple[Optional[Tuple[int, int]], Optional[Tuple[int, int]]]:
    def _find_cell(value: str) -> Optional[Tuple[int, int]]:
        needle = value.strip().lower()
        for rr in range(df.shape[0]):
            for cc in range(df.shape[1]):
                try:
                    v = df.iat[rr, cc]
                except Exception:
                    continue
                if pd.isna(v):
                    continue
                if str(v).strip().lower() == needle:
                    return (rr, cc)
        return None

    blue_hdr = _find_cell("Shifts (Blue)") or _find_cell("Shifts (blue)")
    white_hdr = _find_cell("Shifts (white)")
    return blue_hdr, white_hdr


def _parse_event_log_layout(df: pd.DataFrame) -> Tuple[
    bool,
    Dict[str, List[Tuple[str, str]]],
    Dict[str, List[Tuple[int, str, str]]],
    Dict[int, List[Tuple[int, int, int, int]]],
    Optional[EventLogContext],
]:
    blue_hdr, white_hdr = _detect_event_log_headers(df)
    if not (blue_hdr or white_hdr):
        return False, {}, {}, {}, None

    # Accumulators
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]] = {}
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]] = {}
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]] = {}

    # Roster cap handling
    MAX_TEAM_PLAYERS = 20
    team_roster: Dict[str, List[int]] = {}
    team_excluded: Dict[str, List[int]] = {}

    def _register_and_flag(team: str, jerseys: List[int]) -> List[int]:
        if not team:
            return []
        seen_local = set()
        ordered: List[int] = []
        for j in jerseys:
            if j in seen_local:
                continue
            seen_local.add(j)
            ordered.append(j)
        roster = team_roster.setdefault(team, [])
        excluded = team_excluded.setdefault(team, [])
        for j in ordered:
            if j in roster:
                continue
            if len(roster) < MAX_TEAM_PLAYERS:
                roster.append(j)
            else:
                if j not in excluded:
                    excluded.append(j)
        return ordered

    def _parse_event_time(cell: Any) -> Optional[int]:
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            return None
        s = str(cell).strip()
        parts = s.split(":")
        try:
            if len(parts) == 3:
                h, m, _s = parts
                return int(h) * 60 + int(m)
            if len(parts) == 2:
                m, sec = parts
                return int(m) * 60 + int(sec)
            if len(parts) == 1:
                return int(float(parts[0]))
        except Exception:
            return None
        return None

    def _period_num_from_label(lbl: Any) -> Optional[int]:
        return parse_period_token(lbl)

    def _parse_event_block(header_rc: Tuple[int, int], team_prefix: str) -> None:
        base_r, base_c = header_rc
        # Locate columns for video/game time near header
        video_col = None
        game_col = None
        period_col = None
        for r in range(base_r + 1, min(df.shape[0], base_r + 4)):
            for c in range(max(0, base_c - 4), min(df.shape[1], base_c + 6)):
                val = df.iat[r, c]
                if pd.notna(val) and isinstance(val, str):
                    s = val.strip().lower()
                    if s == "video time":
                        video_col = c
                    elif s == "game time":
                        game_col = c
                    elif "period" in s:
                        period_col = c
            if video_col is not None and game_col is not None:
                break
        if video_col is None:
            video_col = max(0, base_c - 2)
        if game_col is None:
            game_col = max(0, base_c - 1)
        if period_col is None:
            period_col = base_c + 2

        players_start = base_c
        players_width = 12

        # Build events
        current_period_label: Optional[str] = None
        events: List[Dict[str, Any]] = []
        for r in range(base_r + 1, df.shape[0]):
            vcell = df.iat[r, video_col] if video_col < df.shape[1] else None
            if isinstance(vcell, str) and vcell.strip().lower() == "video time":
                plbl = df.iat[r, period_col] if period_col < df.shape[1] else None
                current_period_label = str(plbl).strip() if pd.notna(plbl) else current_period_label
                continue
            gcell = df.iat[r, game_col] if game_col < df.shape[1] else None
            vsec = _parse_event_time(vcell)
            gsec = _parse_event_time(gcell)
            players: List[int] = []
            for k in range(players_width):
                c = players_start + k
                if c >= df.shape[1]:
                    break
                val = df.iat[r, c]
                if pd.isna(val):
                    continue
                if isinstance(val, (int, float)):
                    n = int(val)
                    if 1 <= n <= 98:
                        players.append(n)
                    continue
                if hasattr(val, "hour") and hasattr(val, "minute"):
                    continue
                s = str(val).strip()
                if not s:
                    continue
                if s.upper() in {"PP", "SH"}:
                    continue
                if re.match(r"^\d{1,2}:\d{2}(?::\d{2})?$", s):
                    continue
                for m in re.finditer(r"#?(\d{1,2})(?!\d)", s):
                    try:
                        n = int(m.group(1))
                    except Exception:
                        continue
                    if 1 <= n <= 98:
                        players.append(n)
            if players:
                players = sorted(set(players))
            if vsec is None and gsec is None and not players:
                continue
            players = _register_and_flag(team_prefix, players)
            events.append(
                {
                    "period": _period_num_from_label(current_period_label),
                    "v": vsec,
                    "g": gsec,
                    "players": players,
                }
            )

        # Walk events to generate per-player shifts
        open_shift: Dict[int, Dict[str, Any]] = {}
        last_period: Optional[int] = None
        for ev in events:
            cur_p = ev.get("period")
            # Close open shifts at period change (scoreboard end -> 0:00)
            if last_period is not None and cur_p is not None and cur_p != last_period:
                for pid, sh in list(open_shift.items()):
                    sv, sg = sh.get("sv"), sh.get("sg")
                    evv, egg = ev.get("v"), 0
                    key = f"{team_prefix}_{pid}"
                    if sv is not None and evv is not None:
                        video_pairs_by_player.setdefault(key, []).append(
                            (seconds_to_hhmmss(int(sv)), seconds_to_hhmmss(int(evv)))
                        )
                    if sg is not None:
                        sb_pairs_by_player.setdefault(key, []).append(
                            (
                                last_period,
                                seconds_to_mmss_or_hhmmss(int(sg)),
                                seconds_to_mmss_or_hhmmss(int(egg)),
                            )
                        )
                        if sv is not None and evv is not None:
                            conv_segments_by_period.setdefault(last_period, []).append(
                                (int(sg), int(egg), int(sv), int(evv))
                            )
                    del open_shift[pid]
            if cur_p is not None:
                last_period = cur_p

            on_ice = set(ev.get("players") or [])
            # Close shifts for players no longer on ice
            for pid, sh in list(open_shift.items()):
                if pid not in on_ice:
                    sv, sg = sh.get("sv"), sh.get("sg")
                    evv, egg = ev.get("v"), ev.get("g")
                    key = f"{team_prefix}_{pid}"
                    if sv is not None and evv is not None:
                        video_pairs_by_player.setdefault(key, []).append(
                            (seconds_to_hhmmss(int(sv)), seconds_to_hhmmss(int(evv)))
                        )
                    if sg is not None and cur_p is not None:
                        end_g = egg if egg is not None else 0
                        sb_pairs_by_player.setdefault(key, []).append(
                            (
                                cur_p,
                                seconds_to_mmss_or_hhmmss(int(sg)),
                                seconds_to_mmss_or_hhmmss(int(end_g)),
                            )
                        )
                        if sv is not None and evv is not None:
                            conv_segments_by_period.setdefault(cur_p, []).append(
                                (int(sg), int(end_g), int(sv), int(evv))
                            )
                    del open_shift[pid]

            # Open shifts for players now on
            for pid in on_ice:
                if pid not in open_shift:
                    open_shift[pid] = {
                        "sv": ev.get("v"),
                        "sg": ev.get("g"),
                        "period": ev.get("period"),
                    }

        # Close any remaining open shifts at last event, scoreboard -> 0:00
        if events and open_shift:
            last_ev = events[-1]
            for pid, sh in list(open_shift.items()):
                key = f"{team_prefix}_{pid}"
                sv, sg = sh.get("sv"), sh.get("sg")
                evv, egg = last_ev.get("v"), 0
                per = sh.get("period") or last_ev.get("period")
                if sv is not None and evv is not None:
                    video_pairs_by_player.setdefault(key, []).append(
                        (seconds_to_hhmmss(int(sv)), seconds_to_hhmmss(int(evv)))
                    )
                if sg is not None and per is not None:
                    sb_pairs_by_player.setdefault(key, []).append(
                        (
                            per,
                            seconds_to_mmss_or_hhmmss(int(sg)),
                            seconds_to_mmss_or_hhmmss(int(egg)),
                        )
                    )
                    if sv is not None and evv is not None:
                        conv_segments_by_period.setdefault(int(per), []).append(
                            (int(sg), int(egg), int(sv), int(evv))
                        )

    if blue_hdr:
        _parse_event_block(blue_hdr, "Blue")
    if white_hdr:
        _parse_event_block(white_hdr, "White")

    # ---- Parse left-side event columns ----
    event_counts_by_player: Dict[str, Dict[str, int]] = {}
    event_counts_by_type_team: Dict[Tuple[str, str], int] = {}
    event_instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    event_player_rows: List[Dict[str, Any]] = []

    # Find left header row and relevant columns up to the Blue/White shifts header column
    limit_col = blue_hdr[1] if blue_hdr else df.shape[1]
    left_header_row: Optional[int] = None
    for r in range(min(8, df.shape[0])):
        for c in range(min(limit_col, df.shape[1])):
            v = df.iat[r, c]
            if pd.notna(v) and isinstance(v, str) and v.strip().lower() == "video time":
                left_header_row = r
                break
        if left_header_row is not None:
            break
    if left_header_row is not None:
        # Map labels -> columns from the header row
        label_to_col: Dict[str, int] = {}
        for c in range(min(limit_col, df.shape[1])):
            v = df.iat[left_header_row, c]
            if pd.notna(v) and isinstance(v, str) and v.strip():
                label_to_col[v.strip().lower()] = c

        vt_col = label_to_col.get("video time")
        gt_col = label_to_col.get("scoreboard")
        shots_col = label_to_col.get("shots")
        goals_col = label_to_col.get("goals")
        assists_col = label_to_col.get("assist")
        sog_col = None
        # tolerate wording variations and capitalization
        entries_col = None
        exits_col = None
        for k, c in label_to_col.items():
            kl = k.lower()
            if "controlled" in kl and "blue" in kl and "entr" in kl:
                entries_col = c
            if "controlled" in kl and "exit" in kl:
                exits_col = c
            if "shots on goal" in kl or "shot on goal" in kl or kl == "sog":
                sog_col = c

        # Guess team column
        team_col: Optional[int] = None
        best_count = 0
        for c in range(min(limit_col, df.shape[1])):
            cnt = 0
            for r in range(left_header_row + 1, min(df.shape[0], left_header_row + 80)):
                v = df.iat[r, c]
                if isinstance(v, str) and v.strip() in ("Blue", "White"):
                    cnt += 1
            if cnt > best_count:
                best_count = cnt
                team_col = c

        def _parse_team_from_text(s: Optional[str]) -> Optional[str]:
            if not s:
                return None
            t = s.lower()
            if "blue" in t:
                return "Blue"
            if "white" in t:
                return "White"
            return None

        def _extract_nums(s: Optional[str]) -> List[int]:
            if not s:
                return []
            s = s.strip()
            if re.match(r"^\d{1,2}:\d{2}(?::\d{2})?$", s):
                return []
            nums: List[int] = []
            for m in re.finditer(r"#?(\d{1,2})(?!\d)", s):
                try:
                    n = int(m.group(1))
                except Exception:
                    continue
                if 1 <= n <= 98:
                    nums.append(n)
            seen = set()
            out: List[int] = []
            for n in nums:
                if n in seen:
                    continue
                seen.add(n)
                out.append(n)
            return out

        def _record_event(
            kind: str,
            team: Optional[str],
            jersey_list: List[int],
            period_label: Optional[str],
            vsec: Optional[int],
            gsec: Optional[int],
        ) -> None:
            if not team:
                return
            event_counts_by_type_team[(kind, team)] = (
                event_counts_by_type_team.get((kind, team), 0) + 1
            )
            filtered = _register_and_flag(team, jersey_list)
            period_num = None
            if period_label is not None:
                m = re.search(r"(\d+)", str(period_label))
                if m:
                    period_num = int(m.group(1))
            for j in filtered:
                pk = f"{team}_{int(j)}"
                d = event_counts_by_player.setdefault(pk, {})
                d[kind] = d.get(kind, 0) + 1
                event_player_rows.append(
                    {
                        "event_type": kind,
                        "team": team,
                        "player": pk,
                        "jersey": int(j),
                        "period": period_num,
                        "video_s": vsec,
                        "game_s": gsec,
                    }
                )
            event_instances.setdefault((kind, team), []).append(
                {"period": period_num, "video_s": vsec, "game_s": gsec}
            )

        # Walk data rows to collect events
        current_period: Optional[str] = None
        for r in range(left_header_row + 1, df.shape[0]):
            row_vals = [df.iat[r, c] for c in range(min(limit_col, df.shape[1]))]
            for v in row_vals:
                if isinstance(v, str) and "period" in v.lower():
                    current_period = v.strip()
                    break

            team_val = None
            if team_col is not None:
                tv = df.iat[r, team_col]
                if isinstance(tv, str) and tv.strip() in ("Blue", "White"):
                    team_val = tv.strip()
            row_vsec = _parse_event_time(df.iat[r, vt_col]) if vt_col is not None else None
            row_gsec = _parse_event_time(df.iat[r, gt_col]) if gt_col is not None else None

            if shots_col is not None:
                sv = df.iat[r, shots_col]
                if isinstance(sv, str) and sv.strip():
                    t = team_val or _parse_team_from_text(sv)
                    jerseys = _extract_nums(sv)
                    _record_event("Shot", t, jerseys, current_period, row_vsec, row_gsec)
            if sog_col is not None:
                sogv = df.iat[r, sog_col]
                if isinstance(sogv, str) and sogv.strip():
                    t = team_val or _parse_team_from_text(sogv)
                    jerseys = _extract_nums(sogv)
                    # SOG is also a shot attempt.
                    _record_event("Shot", t, jerseys, current_period, row_vsec, row_gsec)
                    _record_event("SOG", t, jerseys, current_period, row_vsec, row_gsec)
            if goals_col is not None:
                gv = df.iat[r, goals_col]
                if isinstance(gv, str) and gv.strip():
                    t = team_val or _parse_team_from_text(gv)
                    jerseys = _extract_nums(gv)
                    _record_event("Goal", t, jerseys, current_period, row_vsec, row_gsec)
                    # Goals also count as expected goals (xG).
                    _record_event("ExpectedGoal", t, jerseys, current_period, row_vsec, row_gsec)
            if assists_col is not None:
                av = df.iat[r, assists_col]
                if isinstance(av, str) and av.strip():
                    t = team_val or _parse_team_from_text(av)
                    jerseys = _extract_nums(av)
                    _record_event("Assist", t, jerseys, current_period, row_vsec, row_gsec)
            if entries_col is not None:
                ev = df.iat[r, entries_col]
                if isinstance(ev, str) and ev.strip():
                    t = team_val or _parse_team_from_text(ev)
                    jerseys = _extract_nums(ev)
                    _record_event("ControlledEntry", t, jerseys, current_period, row_vsec, row_gsec)
            if exits_col is not None:
                xv = df.iat[r, exits_col]
                if isinstance(xv, str) and xv.strip():
                    t = team_val or _parse_team_from_text(xv)
                    jerseys = _extract_nums(xv)
                    _record_event("ControlledExit", t, jerseys, current_period, row_vsec, row_gsec)
            label = df.iat[r, 0]
            if isinstance(label, str) and label.strip().lower() == "expected goal":
                text_cell = None
                if shots_col is not None:
                    text_cell = df.iat[r, shots_col]
                if not (isinstance(text_cell, str) and text_cell.strip()) and goals_col is not None:
                    text_cell = df.iat[r, goals_col]
                t = None
                jerseys: List[int] = []
                if isinstance(text_cell, str) and text_cell.strip():
                    t = team_val or _parse_team_from_text(text_cell)
                    jerseys = _extract_nums(text_cell)
                _record_event("ExpectedGoal", t, jerseys, current_period, row_vsec, row_gsec)

    event_log_context = EventLogContext(
        event_counts_by_player=event_counts_by_player,
        event_counts_by_type_team=event_counts_by_type_team,
        event_instances=event_instances,
        event_player_rows=event_player_rows,
        team_roster=team_roster,
        team_excluded=team_excluded,
    )

    return (
        True,
        video_pairs_by_player,
        sb_pairs_by_player,
        conv_segments_by_period,
        event_log_context,
    )


def _parse_per_player_layout(df: pd.DataFrame, keep_goalies: bool, skip_validation: bool) -> Tuple[
    Dict[str, List[Tuple[str, str]]],
    Dict[str, List[Tuple[int, str, str]]],
    Dict[int, List[Tuple[int, int, int, int]]],
    int,
]:
    blocks = find_period_blocks(df)
    if not blocks:
        raise ValueError("No 'Period N' sections found in column A.")

    video_pairs_by_player: Dict[str, List[Tuple[str, str]]] = {}
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]] = {}
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]] = {}

    MAX_SHIFT_SECONDS = 30 * 60  # 30 minutes

    def _report_validation(
        kind: str, period: int, player_key: str, a: str, b: str, reason: str
    ) -> None:
        print(
            f"[validation] {kind} | Player={player_key} | Period={period} | start='{a}' end='{b}' -> {reason}",
            file=sys.stderr,
        )

    validation_errors = 0

    # Parse each period block
    for period_num, blk_start, blk_end in blocks:
        header_row_idx = find_header_row(df, blk_start, blk_end)
        if header_row_idx is None:
            raise ValueError(f"Could not locate header row for Period {period_num}")
        data_start = header_row_idx + 1

        header = df.iloc[header_row_idx]
        groups = forward_fill_header_labels(header)

        start_sb_cols = _resolve_header_columns(
            groups,
            LABEL_START_SB,
            "Shift start (Scoreboard time)",
            "Shift Start (Scoreboard time)",
        )
        end_sb_cols = _resolve_header_columns(groups, LABEL_END_SB, "Shift end (Scoreboard time)")
        start_v_cols = _resolve_header_columns(groups, LABEL_START_V)
        end_v_cols = _resolve_header_columns(groups, LABEL_END_V)
        for lab, cols in [
            (LABEL_START_SB, start_sb_cols),
            (LABEL_END_SB, end_sb_cols),
            (LABEL_START_V, start_v_cols),
            (LABEL_END_V, end_v_cols),
        ]:
            if not cols:
                raise ValueError(f"Missing header columns for '{lab}' in Period {period_num}")

        # Iterate rows (players) in the block
        for r in range(data_start, blk_end):
            jersey = str(df.iloc[r, 0]).strip()
            name = str(df.iloc[r, 1]).strip()

            if is_period_label(df.iloc[r, 0]) or (not jersey and not name):
                break
            jersey_lower = jersey.lower()
            # Skip header-like rows that may appear again (e.g., overtime sections)
            if jersey_lower in {"jersey no", "jersey number"}:
                continue
            if not jersey or jersey_lower == "nan":
                continue
            # Skip goalies like "(G) 37"
            if not keep_goalies and "(" in jersey and ")" in jersey:
                continue

            player_key = f"{sanitize_name(jersey)}_{sanitize_name(name)}"

            video_pairs = extract_pairs_from_row(df.iloc[r], start_v_cols, end_v_cols)
            sb_pairs = extract_pairs_from_row(df.iloc[r], start_sb_cols, end_sb_cols)
            if sb_pairs:
                sb_pairs = [(a, _normalize_sb_end_time(b)) for a, b in sb_pairs]

            if not skip_validation:
                for va, vb in video_pairs:
                    try:
                        vsa = parse_flex_time_to_seconds(va)
                        vsb = parse_flex_time_to_seconds(vb)
                    except Exception as e:
                        _report_validation(
                            "VIDEO", period_num, player_key, va, vb, f"unparseable time: {e}"
                        )
                        validation_errors += 1
                        continue
                    if vsa >= vsb:
                        _report_validation(
                            "VIDEO",
                            period_num,
                            player_key,
                            va,
                            vb,
                            "start must be before end (strictly increasing)",
                        )
                        validation_errors += 1
                    dur = vsb - vsa if vsb >= vsa else 0
                    if dur > MAX_SHIFT_SECONDS:
                        _report_validation(
                            "VIDEO",
                            period_num,
                            player_key,
                            va,
                            vb,
                            f"duration {seconds_to_mmss_or_hhmmss(dur)} exceeds limit 30:00",
                        )
                        validation_errors += 1

                for sa, sb in sb_pairs:
                    try:
                        ssa = parse_flex_time_to_seconds(sa)
                        ssb = parse_flex_time_to_seconds(sb)
                    except Exception as e:
                        _report_validation(
                            "SCOREBOARD", period_num, player_key, sa, sb, f"unparseable time: {e}"
                        )
                        validation_errors += 1
                        continue
                    if ssa == ssb:
                        _report_validation(
                            "SCOREBOARD",
                            period_num,
                            player_key,
                            sa,
                            sb,
                            "start equals end (zero-length shift)",
                        )
                        validation_errors += 1
                    dur = abs(ssb - ssa)
                    if dur > MAX_SHIFT_SECONDS:
                        _report_validation(
                            "SCOREBOARD",
                            period_num,
                            player_key,
                            sa,
                            sb,
                            f"duration {seconds_to_mmss_or_hhmmss(dur)} exceeds limit 30:00",
                        )
                        validation_errors += 1

            if video_pairs:
                video_pairs_by_player.setdefault(player_key, []).extend(video_pairs)
            if sb_pairs:
                sb_pairs_by_player.setdefault(player_key, []).extend(
                    (period_num, a, b) for a, b in sb_pairs
                )

            nseg = min(len(video_pairs), len(sb_pairs))
            for idx in range(nseg):
                sva, svb = video_pairs[idx]
                sba, sbb = sb_pairs[idx]
                try:
                    v1 = parse_flex_time_to_seconds(sva)
                    v2 = parse_flex_time_to_seconds(svb)
                    s1 = parse_flex_time_to_seconds(sba)
                    s2 = parse_flex_time_to_seconds(sbb)
                except Exception:
                    continue
                conv_segments_by_period.setdefault(period_num, []).append((s1, s2, v1, v2))

    return video_pairs_by_player, sb_pairs_by_player, conv_segments_by_period, validation_errors


def _write_video_times_and_scripts(
    outdir: Path,
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]],
    create_scripts: bool,
) -> None:
    if not create_scripts:
        return
    for player_key, v_pairs in video_pairs_by_player.items():
        norm_pairs = []
        for a, b in v_pairs:
            try:
                sa = parse_flex_time_to_seconds(a)
                sb = parse_flex_time_to_seconds(b)
            except Exception:
                continue
            norm_pairs.append((seconds_to_hhmmss(sa), seconds_to_hhmmss(sb)))
        p = outdir / f"{player_key}_video_times.txt"
        p.write_text(
            "\n".join(f"{a} {b}" for a, b in norm_pairs) + ("\n" if norm_pairs else ""),
            encoding="utf-8",
        )

        script_path = outdir / f"clip_{player_key}.sh"
        player_label = player_key.replace("_", " ")
        script_body = """#!/usr/bin/env bash
set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
THIS_DIR=\"$(cd \"$(dirname \"${{BASH_SOURCE[0]}}\")\" && pwd)\"
TS_FILE=\"$THIS_DIR/{player_key}_video_times.txt\"
# Parse optional flags
QUICK=0
HQ=0
shift 2 || true
for ARG in \"$@\"; do
  if [ \"$ARG\" = \"--quick\" ] || [ \"$ARG\" = \"-q\" ]; then
    QUICK=1
  elif [ \"$ARG\" = \"--hq\" ]; then
    HQ=1
  fi
done

EXTRA_FLAGS=()
if [ \"$QUICK\" -gt 0 ]; then
  EXTRA_FLAGS+=(\"--quick\" \"1\")
fi
if [ \"$HQ\" -gt 0 ]; then
  export VIDEO_CLIPPER_HQ=1
fi

python -m hmlib.cli.video_clipper -j {nr_jobs} --input \"$INPUT\" --timestamps \"$TS_FILE\" --temp-dir \"$THIS_DIR/temp_clips/{player_key}\" \"{player_label} vs $OPP\" \"${{EXTRA_FLAGS[@]}}\"
""".format(
            nr_jobs=4, player_key=player_key, player_label=player_label
        )
        script_path.write_text(script_body, encoding="utf-8")
        try:
            import os

            os.chmod(script_path, 0o755)
        except Exception:
            pass


def _write_scoreboard_times(
    outdir: Path,
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    *,
    create_scripts: bool,
) -> None:
    if not create_scripts:
        return
    for player_key, sb_list in sb_pairs_by_player.items():
        p = outdir / f"{player_key}_scoreboard_times.txt"
        lines = [f"{period} {a} {b}" for (period, a, b) in sb_list]
        p.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def _write_global_summary_csv(
    stats_dir: Path, sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]]
) -> None:
    summary_rows = []
    for player_key, sb_list in sb_pairs_by_player.items():
        parts = _parse_player_key(player_key)
        all_pairs = [(a, b) for (_, a, b) in sb_list]
        shift_summary = summarize_shift_lengths_sec(all_pairs)
        row = {
            "jersey": parts.jersey or "",
            "player": parts.name,
            "num_shifts": int(shift_summary["num_shifts"]),
            "toi_total_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_total"])
                if ":" in shift_summary["toi_total"]
                else 0
            ),
            "toi_avg_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_avg"])
                if ":" in shift_summary["toi_avg"]
                else 0
            ),
            "toi_median_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_median"])
                if ":" in shift_summary["toi_median"]
                else 0
            ),
            "toi_longest_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_longest"])
                if ":" in shift_summary["toi_longest"]
                else 0
            ),
            "toi_shortest_sec": (
                parse_flex_time_to_seconds(shift_summary["toi_shortest"])
                if ":" in shift_summary["toi_shortest"]
                else 0
            ),
        }
        summary_rows.append(row)
    if summary_rows:
        df = pd.DataFrame(summary_rows).sort_values(by=["player", "jersey"])
        df.to_csv(stats_dir / "summary_stats.csv", index=False)
        _write_styled_xlsx_table(
            stats_dir / "summary_stats.xlsx",
            df,
            sheet_name="summary_stats",
            title="Shift Summary",
        )


def _write_shift_rows_csv(
    stats_dir: Path,
    *,
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    video_pairs_by_player: Optional[Dict[str, List[Tuple[str, str]]]] = None,
    source: str = "shift_spreadsheet",
) -> None:
    """
    Write per-shift intervals for webapp ingestion (stats/shift_rows.csv).

    These are later converted to on/off-ice markers for the timeline and used to derive TOI/Shifts
    at runtime (instead of importing TOI aggregates into PlayerStat).
    """
    import csv  # local import
    import hashlib  # local import
    import io  # local import

    if not sb_pairs_by_player:
        return

    video_pairs_by_player = dict(video_pairs_by_player or {})

    rows: List[Dict[str, Any]] = []
    for player_key, sb_list in (sb_pairs_by_player or {}).items():
        if not sb_list:
            continue
        parts = _parse_player_key(player_key)
        jersey = _normalize_jersey_number(parts.jersey)
        name = str(parts.name or "").replace("_", " ").strip()
        v_list = list(video_pairs_by_player.get(player_key) or [])
        for idx, (period, sb_start, sb_end) in enumerate(sb_list):
            try:
                per_i = int(period)
            except Exception:
                continue
            if per_i <= 0:
                continue
            try:
                start_gs = parse_flex_time_to_seconds(str(sb_start))
                end_gs = parse_flex_time_to_seconds(str(sb_end))
            except Exception:
                continue

            start_vs = None
            end_vs = None
            if idx < len(v_list):
                va, vb = v_list[idx]
                try:
                    start_vs = parse_flex_time_to_seconds(str(va))
                    end_vs = parse_flex_time_to_seconds(str(vb))
                except Exception:
                    start_vs = None
                    end_vs = None

            import_key_raw = "|".join(
                [
                    str(jersey or ""),
                    str(name or ""),
                    str(per_i),
                    str(start_gs),
                    str(end_gs),
                    str(start_vs if start_vs is not None else ""),
                    str(end_vs if end_vs is not None else ""),
                ]
            )
            import_key = hashlib.sha1(import_key_raw.encode("utf-8")).hexdigest()[:40]

            rows.append(
                {
                    "Jersey #": str(jersey or ""),
                    "Player": str(name or ""),
                    "Period": int(per_i),
                    "Game Seconds": int(start_gs),
                    "Game Seconds End": int(end_gs),
                    "Video Seconds": int(start_vs) if start_vs is not None else "",
                    "Video Seconds End": int(end_vs) if end_vs is not None else "",
                    "Source": str(source or ""),
                    "Import Key": str(import_key),
                }
            )

    if not rows:
        return

    # Keep file deterministic for clean diffs / stable imports.
    rows.sort(
        key=lambda r: (
            int(str(r.get("Period") or 0) or 0),
            str(r.get("Jersey #") or ""),
            str(r.get("Player") or ""),
            int(str(r.get("Game Seconds") or 0) or 0),
            int(str(r.get("Game Seconds End") or 0) or 0),
        )
    )

    out_path = stats_dir / "shift_rows.csv"
    buf = io.StringIO()
    fieldnames = [
        "Jersey #",
        "Player",
        "Period",
        "Game Seconds",
        "Game Seconds End",
        "Video Seconds",
        "Video Seconds End",
        "Source",
        "Import Key",
    ]
    w = csv.DictWriter(buf, fieldnames=fieldnames)
    w.writeheader()
    for r in rows:
        w.writerow({k: r.get(k, "") for k in fieldnames})
    out_path.write_text(buf.getvalue(), encoding="utf-8")


def _write_pair_on_ice_csv(
    stats_dir: Path, rows: List[Dict[str, Any]], *, include_toi: bool
) -> None:
    if not rows:
        return

    def _blank_if_zero(x: Any) -> Any:
        try:
            v = int(x or 0)
        except Exception:
            v = 0
        return "" if v == 0 else v

    out_rows: List[Dict[str, Any]] = []
    for r in rows:
        p_parts = _parse_player_key(r.get("player", ""))
        t_parts = _parse_player_key(r.get("teammate", ""))
        row_out: Dict[str, Any] = {
            "Player Jersey #": p_parts.jersey or "",
            "Player": p_parts.name,
            "Teammate Jersey #": t_parts.jersey or "",
            "Teammate": t_parts.name,
            "Games with Data": int(r.get("shift_games", 0) or 0),
            "Overlap %": float(r.get("overlap_pct", 0.0) or 0.0),
            "GF Together": int(r.get("gf_together", 0) or 0),
            "GA Together": int(r.get("ga_together", 0) or 0),
            "Player Goals (On Ice Together)": _blank_if_zero(
                r.get("player_goals_on_ice_together", 0)
            ),
            "Player Assists (On Ice Together)": _blank_if_zero(
                r.get("player_assists_on_ice_together", 0)
            ),
            "Goals Collaborated": _blank_if_zero(r.get("goals_collab_with_teammate", 0)),
            "Assists Collaborated": _blank_if_zero(r.get("assists_collab_with_teammate", 0)),
            "+/- Together": int(r.get("plus_minus_together", 0) or 0),
            "Player Total +/-": int(r.get("player_total_plus_minus", 0) or 0),
            "Teammate Total +/-": int(r.get("teammate_total_plus_minus", 0) or 0),
        }
        # Only publish absolute TOI/overlap time when explicitly enabled.
        if include_toi:
            row_out["Player TOI"] = _seconds_to_compact_hms(
                int(r.get("player_toi_seconds", 0) or 0)
            )
            row_out["Overlap"] = _seconds_to_compact_hms(int(r.get("overlap_seconds", 0) or 0))
        out_rows.append(row_out)
    df = pd.DataFrame(out_rows)
    try:
        df.sort_values(by=["Player", "Overlap %"], ascending=[True, False], inplace=True)
    except Exception:
        pass
    df.to_csv(stats_dir / "pair_on_ice.csv", index=False)
    _write_styled_xlsx_table(
        stats_dir / "pair_on_ice.xlsx",
        df,
        sheet_name="pair_on_ice",
        title="Pair On-Ice",
        number_formats={"Overlap %": "0.0"},
        text_columns=["Player", "Teammate"],
        align_right_columns=["Player TOI", "Overlap"],
        merge_columns=["Player Jersey #", "Player"],
    )


def _write_all_events_summary(
    stats_dir: Path,
    *,
    sb_pairs_by_player: Dict[str, List[Tuple[int, str, str]]],
    sb_pairs_by_player_by_side: Optional[Dict[str, Dict[str, List[Tuple[int, str, str]]]]] = None,
    goals: List[GoalEvent],
    goals_by_period: Dict[int, List[GoalEvent]],
    event_log_context: Optional["EventLogContext"],
    focus_team: Optional[str],
    team_side: Optional[str],
    t2s_game_id: Optional[int] = None,
    t2s_events: Optional[List[Dict[str, Any]]] = None,
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    spreadsheet_event_mapping_summary: Optional[SpreadsheetEventMappingSummary] = None,
) -> None:
    """
    Write a single per-game table of all known events, including:
      - long-sheet / event-log events (player-attributed and team-only)
      - goal/assist events from the goals list (when available)
      - on-ice player lists for each event (home/away when available, based on shift times)

    This is intended to be sufficient to recompute most per-game and per-player event-based stats.
    """

    def _side_label(side: str) -> str:
        sl = str(side or "").strip().lower()
        if sl == "home":
            return "Home"
        if sl == "away":
            return "Away"
        return ""

    def _team_side_from_color(team_color: Any) -> str:
        team_str = str(team_color) if team_color is not None else ""
        our_side_l = str(team_side or "").strip().lower()
        if our_side_l not in {"home", "away"}:
            return ""
        if focus_team not in {"Blue", "White"}:
            return ""
        if team_str not in {"Blue", "White"}:
            return ""
        if team_str == focus_team:
            return _side_label(our_side_l)
        opp = "away" if our_side_l == "home" else "home"
        return _side_label(opp)

    def _for_against_for_event(event_type: Any) -> str:
        et = str(event_type or "").strip()
        etl = et.casefold()
        # "For/Against" is relative to the owning team (Team Side), not a synonym for Home/Away.
        if etl in {"giveaway", "createdturnover", "created turnover"}:
            return "Against"
        if etl in {"penalty"}:
            return "Against"
        if etl in {"penalty expired"}:
            return "For"
        if etl in {
            "shot",
            "sog",
            "expectedgoal",
            "xg",
            "goal",
            "assist",
            "controlledentry",
            "controlled exit",
            "controlledexit",
            "takeaway",
            "turnoverforced",
            "turnovers (forced)",
        }:
            return "For"
        return ""

    # Precompute merged shift intervals and shift-start times for on-ice membership (Home/Away).
    intervals_by_side_player_period: Dict[str, Dict[str, Dict[int, List[Tuple[int, int]]]]] = {}
    start_times_by_side_player_period: Dict[str, Dict[str, Dict[int, set[int]]]] = {}

    if sb_pairs_by_player_by_side is None:
        # Back-compat: treat sb_pairs_by_player as the "our team" shift table.
        our_side_label = _side_label(str(team_side or ""))
        opp_side_label = (
            "Away" if our_side_label == "Home" else ("Home" if our_side_label == "Away" else "")
        )
        sb_pairs_by_player_by_side = {
            our_side_label: dict(sb_pairs_by_player or {}),
            opp_side_label: {},
        }

    for side_label, side_map in (sb_pairs_by_player_by_side or {}).items():
        if side_label not in {"Home", "Away"}:
            continue
        for player, sb_list in (side_map or {}).items():
            per_period: Dict[int, List[Tuple[int, int]]] = {}
            start_times: Dict[int, set[int]] = {}
            for period, a, b in sb_list or []:
                lo, hi = compute_interval_seconds(a, b)
                per_period.setdefault(int(period), []).append((lo, hi))
                try:
                    start_times.setdefault(int(period), set()).add(parse_flex_time_to_seconds(a))
                except Exception:
                    pass
            merged: Dict[int, List[Tuple[int, int]]] = {
                p: _merge_intervals(iv) for p, iv in per_period.items()
            }
            intervals_by_side_player_period.setdefault(side_label, {})[player] = merged
            start_times_by_side_player_period.setdefault(side_label, {})[player] = start_times

    def _on_ice_players_side(side_label: str, period: int, game_s: int) -> List[str]:
        out: List[str] = []
        for pk, per_map in (intervals_by_side_player_period.get(side_label) or {}).items():
            for lo, hi in per_map.get(period, []):
                if interval_contains(game_s, lo, hi):
                    out.append(pk)
                    break
        out.sort(key=_player_sort_key)
        return out

    def _on_ice_players_pm_side(side_label: str, period: int, game_s: int) -> List[str]:
        # Plus/minus skips events exactly at shift start for each player.
        base = _on_ice_players_side(side_label, period, game_s)
        out: List[str] = []
        for pk in base:
            if game_s in (start_times_by_side_player_period.get(side_label) or {}).get(pk, {}).get(
                period, set()
            ):
                continue
            out.append(pk)
        return out

    def _map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    # Group player-attributed long-sheet rows by event identity.
    player_rows_by_event: Dict[
        Tuple[str, str, int, Optional[int], Optional[int]], Dict[str, Any]
    ] = {}
    if event_log_context is not None:
        for r in event_log_context.event_player_rows or []:
            try:
                et = str(r.get("event_type") or "")
                tm = str(r.get("team") or "")
                per = int(r.get("period") or 0)
                vs = r.get("video_s")
                gs = r.get("game_s")
                key = (
                    et,
                    tm,
                    per,
                    int(vs) if isinstance(vs, (int, float)) else None,
                    int(gs) if isinstance(gs, (int, float)) else None,
                )
                dest = player_rows_by_event.setdefault(key, {"players": set(), "jerseys": set()})
                pk = str(r.get("player") or "").strip()
                if pk:
                    dest["players"].add(pk)
                j = r.get("jersey")
                if j is not None:
                    try:
                        dest["jerseys"].add(int(j))
                    except Exception:
                        pass
            except Exception:
                continue

    rows: List[Dict[str, Any]] = []
    event_id = 0

    # Long-sheet / event-log events.
    if event_log_context is not None:
        for (etype, team), inst_list in sorted((event_log_context.event_instances or {}).items()):
            if goals and str(etype) == "Goal":
                # Prefer canonical Goal/Assist rows from the goals list when available.
                # Exception: for TimeToScore-linked games, we keep long-sheet goal rows in this CSV so the webapp
                # can enrich TimeToScore goal events with long-sheet video times (server ignores long-only goals).
                if t2s_game_id is None:
                    continue
            for inst in inst_list or []:
                period = int(inst.get("period") or 0)
                video_s = inst.get("video_s")
                game_s = inst.get("game_s")
                vs_i = int(video_s) if isinstance(video_s, (int, float)) else None
                gs_i = int(game_s) if isinstance(game_s, (int, float)) else None
                key = (str(etype), str(team), period, vs_i, gs_i)
                pj = player_rows_by_event.get(key, {"players": set(), "jerseys": set()})
                attrib_players = sorted(
                    [str(x) for x in pj.get("players", set()) if x], key=_player_sort_key
                )
                attrib_jerseys = sorted(
                    [int(x) for x in pj.get("jerseys", set()) if isinstance(x, int)]
                )
                home_on_ice = (
                    _on_ice_players_pm_side("Home", period, gs_i)
                    if (gs_i is not None and period > 0)
                    else []
                )
                away_on_ice = (
                    _on_ice_players_pm_side("Away", period, gs_i)
                    if (gs_i is not None and period > 0)
                    else []
                )
                team_side_txt = _team_side_from_color(team)
                if team_side_txt == "Home":
                    on_ice_team = home_on_ice
                elif team_side_txt == "Away":
                    on_ice_team = away_on_ice
                else:
                    on_ice_team = []

                event_id += 1
                rows.append(
                    {
                        "Event ID": event_id,
                        "Source": "long",
                        "Event Type": _display_event_type(str(etype)),
                        "Team Raw": str(team),
                        "Team Side": team_side_txt,
                        "For/Against": _for_against_for_event(str(etype)),
                        "Team Rel": team_side_txt,
                        "Period": period if period > 0 else "",
                        "Game Time": seconds_to_mmss_or_hhmmss(gs_i) if gs_i is not None else "",
                        "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                        "Game Seconds": gs_i if gs_i is not None else "",
                        "Video Seconds": vs_i if vs_i is not None else "",
                        "Details": "",
                        "Attributed Players": ",".join(
                            _format_player_name_with_jersey(x) for x in attrib_players
                        ),
                        "Attributed Jerseys": ",".join(str(j) for j in attrib_jerseys),
                        "On-Ice Players": ",".join(
                            _format_player_name_with_jersey(x) for x in on_ice_team
                        ),
                        "On-Ice Players (Home)": ",".join(
                            _format_player_name_with_jersey(x) for x in home_on_ice
                        ),
                        "On-Ice Players (Away)": ",".join(
                            _format_player_name_with_jersey(x) for x in away_on_ice
                        ),
                    }
                )

    # Extra TimeToScore-derived events (penalties/goalie changes), even when long sheets exist.
    for ev in t2s_events or []:
        try:
            etype = str(ev.get("event_type") or "").strip() or "Event"
            period = int(ev.get("period") or 0)
        except Exception:
            continue
        gs_raw = ev.get("game_s")
        gs_i = int(gs_raw) if isinstance(gs_raw, (int, float)) else None
        vs_i = _map_sb_to_video(period, gs_i) if (gs_i is not None and period > 0) else None
        home_on_ice = (
            _on_ice_players_pm_side("Home", period, gs_i)
            if (gs_i is not None and period > 0)
            else []
        )
        away_on_ice = (
            _on_ice_players_pm_side("Away", period, gs_i)
            if (gs_i is not None and period > 0)
            else []
        )

        team_raw = str(ev.get("team_raw") or "")
        side_txt = _side_label(str(ev.get("team_side") or ""))
        for_against_txt = str(ev.get("for_against") or "")
        details = str(ev.get("details") or "")
        jerseys_val = ev.get("attributed_jerseys") or []
        jerseys_list: List[str] = []
        if isinstance(jerseys_val, (list, tuple)):
            for j in jerseys_val:
                norm = _normalize_jersey_number(j)
                if norm:
                    jerseys_list.append(norm)
        else:
            norm = _normalize_jersey_number(jerseys_val)
            if norm:
                jerseys_list.append(norm)

        event_id += 1
        rows.append(
            {
                "Event ID": event_id,
                "Source": str(ev.get("source") or "t2s"),
                "Event Type": _display_event_type(etype),
                "Team Raw": team_raw,
                "Team Side": side_txt,
                "For/Against": (
                    for_against_txt if for_against_txt else _for_against_for_event(etype)
                ),
                "Team Rel": side_txt,
                "Period": period if period > 0 else "",
                "Game Time": seconds_to_mmss_or_hhmmss(gs_i) if gs_i is not None else "",
                "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                "Game Seconds": gs_i if gs_i is not None else "",
                "Video Seconds": vs_i if vs_i is not None else "",
                "Details": details,
                "Attributed Players": "",
                "Attributed Jerseys": ",".join(jerseys_list),
                "On-Ice Players": ",".join(
                    _format_player_name_with_jersey(x)
                    for x in (
                        home_on_ice
                        if side_txt == "Home"
                        else (away_on_ice if side_txt == "Away" else [])
                    )
                ),
                "On-Ice Players (Home)": ",".join(
                    _format_player_name_with_jersey(x) for x in home_on_ice
                ),
                "On-Ice Players (Away)": ",".join(
                    _format_player_name_with_jersey(x) for x in away_on_ice
                ),
            }
        )

    # Goal/assist events from the goals list.
    for ev in sorted(
        goals or [],
        key=lambda e: (int(getattr(e, "period", 0) or 0), int(getattr(e, "t_sec", 0) or 0)),
    ):
        period = int(getattr(ev, "period", 0) or 0)
        gs_i = int(getattr(ev, "t_sec", 0) or 0)
        vs_i = getattr(ev, "video_t_sec", None)
        if vs_i is None:
            vs_i = _map_sb_to_video(period, gs_i)
        kind = str(getattr(ev, "kind", "") or "").strip().upper()
        our_side_l = str(team_side or "").strip().lower()
        if our_side_l in {"home", "away"}:
            opp_side_l = "away" if our_side_l == "home" else "home"
            goal_side_txt = (
                _side_label(our_side_l)
                if kind == "GF"
                else (_side_label(opp_side_l) if kind == "GA" else "")
            )
        else:
            goal_side_txt = ""
        for_against_txt = "For" if goal_side_txt else ""
        home_on_ice = _on_ice_players_pm_side("Home", period, gs_i) if period > 0 else []
        away_on_ice = _on_ice_players_pm_side("Away", period, gs_i) if period > 0 else []
        on_ice_team = (
            home_on_ice
            if goal_side_txt == "Home"
            else (away_on_ice if goal_side_txt == "Away" else [])
        )

        # Goal scorer row (if known).
        scorer = getattr(ev, "scorer", None)
        if scorer:
            event_id += 1
            rows.append(
                {
                    "Event ID": event_id,
                    "Source": "goals",
                    "Event Type": "Goal",
                    "Team Raw": "",
                    "Team Side": goal_side_txt,
                    "For/Against": for_against_txt,
                    "Team Rel": goal_side_txt,
                    "Period": period if period > 0 else "",
                    "Game Time": seconds_to_mmss_or_hhmmss(gs_i),
                    "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                    "Game Seconds": gs_i,
                    "Video Seconds": vs_i if vs_i is not None else "",
                    "Details": "",
                    "Attributed Players": "",
                    "Attributed Jerseys": str(scorer),
                    "On-Ice Players": ",".join(
                        _format_player_name_with_jersey(x) for x in on_ice_team
                    ),
                    "On-Ice Players (Home)": ",".join(
                        _format_player_name_with_jersey(x) for x in home_on_ice
                    ),
                    "On-Ice Players (Away)": ",".join(
                        _format_player_name_with_jersey(x) for x in away_on_ice
                    ),
                }
            )
        # Assist rows (if any).
        for ast in getattr(ev, "assists", []) or []:
            if not ast:
                continue
            event_id += 1
            rows.append(
                {
                    "Event ID": event_id,
                    "Source": "goals",
                    "Event Type": "Assist",
                    "Team Raw": "",
                    "Team Side": goal_side_txt,
                    "For/Against": for_against_txt,
                    "Team Rel": goal_side_txt,
                    "Period": period if period > 0 else "",
                    "Game Time": seconds_to_mmss_or_hhmmss(gs_i),
                    "Video Time": _seconds_to_compact_hms(vs_i) if vs_i is not None else "",
                    "Game Seconds": gs_i,
                    "Video Seconds": vs_i if vs_i is not None else "",
                    "Details": "",
                    "Attributed Players": "",
                    "Attributed Jerseys": str(ast),
                    "On-Ice Players": ",".join(
                        _format_player_name_with_jersey(x) for x in on_ice_team
                    ),
                    "On-Ice Players (Home)": ",".join(
                        _format_player_name_with_jersey(x) for x in home_on_ice
                    ),
                    "On-Ice Players (Away)": ",".join(
                        _format_player_name_with_jersey(x) for x in away_on_ice
                    ),
                }
            )

    # ---------------------------------------------------------------------
    # Video-time propagation for events without video time
    #
    # Goal: if an event row has no video time, but another event (or a shift boundary)
    # at the same game time has a video time, assign it.
    #
    # Special-case: penalties often occur at the end of a shift. If a penalty has no
    # video time OR a more precise nearby video anchor exists, prefer the earliest
    # available video time within +/- 3 seconds of the penalty game time.
    # ---------------------------------------------------------------------

    def _parse_int_or_none(v: Any) -> Optional[int]:
        try:
            if v is None:
                return None
            if isinstance(v, str) and not v.strip():
                return None
            return int(float(v))
        except Exception:
            return None

    def _is_missing_video(row: Dict[str, Any]) -> bool:
        vs = row.get("Video Seconds")
        if vs is None:
            return True
        if isinstance(vs, str) and not vs.strip():
            return True
        try:
            return not isinstance(int(float(vs)), int)
        except Exception:
            return True

    # Build a per-period anchor map: (period, game_s) -> earliest video_s.
    anchors_by_period: Dict[int, Dict[int, int]] = {}

    def _add_anchor(period: int, game_s: int, video_s: int) -> None:
        if period <= 0:
            return
        if game_s is None or video_s is None:
            return
        try:
            p = int(period)
            gs = int(game_s)
            vs = int(video_s)
        except Exception:
            return
        cur = anchors_by_period.setdefault(p, {})
        prev = cur.get(gs)
        if prev is None or vs < int(prev):
            cur[gs] = vs

    # 1) Existing event rows that already have video time.
    for r in rows:
        per = _parse_int_or_none(r.get("Period")) or 0
        gs = _parse_int_or_none(r.get("Game Seconds"))
        vs = _parse_int_or_none(r.get("Video Seconds"))
        if per > 0 and gs is not None and vs is not None:
            _add_anchor(per, gs, vs)

    # 2) Shift boundaries from conversion segments (primary/long shift tables).
    # Each segment (s1,s2,v1,v2) corresponds to one shift; the endpoints are "shift begin/end".
    for per, segs in (conv_segments_by_period or {}).items():
        try:
            p = int(per)
        except Exception:
            continue
        for s1, s2, v1, v2 in segs or []:
            try:
                _add_anchor(p, int(s1), int(v1))
                _add_anchor(p, int(s2), int(v2))
            except Exception:
                continue

    # Materialize anchors list per period for +/- window scans.
    anchors_list_by_period: Dict[int, List[Tuple[int, int]]] = {}
    for per, m in anchors_by_period.items():
        items = [(int(gs), int(vs)) for gs, vs in m.items()]
        items.sort(key=lambda t: t[0])
        anchors_list_by_period[int(per)] = items

    def _best_video_for_game_time(period: int, game_s: int) -> Optional[int]:
        rec = anchors_by_period.get(int(period), {}).get(int(game_s))
        if rec is None:
            return None
        return int(rec)

    def _best_video_within_window(
        period: int, game_s: int, window_s: int
    ) -> Optional[Tuple[int, int]]:
        items = anchors_list_by_period.get(int(period)) or []
        if not items:
            return None
        lo = int(game_s) - int(window_s)
        hi = int(game_s) + int(window_s)
        best_delta: Optional[int] = None
        best_vs: Optional[int] = None
        for gs, vs in items:
            if gs < lo:
                continue
            if gs > hi:
                break
            d = abs(int(gs) - int(game_s))
            if best_delta is None or d < best_delta:
                best_delta = d
                best_vs = int(vs)
            elif d == best_delta and best_vs is not None and int(vs) < best_vs:
                best_vs = int(vs)
        if best_delta is None or best_vs is None:
            return None
        return int(best_vs), int(best_delta)

    for r in rows:
        per = _parse_int_or_none(r.get("Period")) or 0
        gs = _parse_int_or_none(r.get("Game Seconds"))
        if per <= 0 or gs is None:
            continue
        et = str(r.get("Event Type") or "").strip().casefold()
        if et == "penalty":
            candidate = _best_video_within_window(per, gs, 3)
            if candidate is not None:
                vs_i, _delta = candidate
                r["Video Seconds"] = int(vs_i)
                r["Video Time"] = _seconds_to_compact_hms(int(vs_i))
            continue

        if _is_missing_video(r):
            candidate = _best_video_for_game_time(per, gs)
            if candidate is not None:
                vs_i = candidate
                r["Video Seconds"] = int(vs_i)
                r["Video Time"] = _seconds_to_compact_hms(int(vs_i))

    # Stable sort by time, then id.
    def _sort_key(r: Dict[str, Any]) -> Tuple[int, int, int]:
        try:
            per = int(r.get("Period") or 0)
        except Exception:
            per = 0
        try:
            gs = int(r.get("Game Seconds") or 0)
        except Exception:
            gs = 0
        try:
            eid = int(r.get("Event ID") or 0)
        except Exception:
            eid = 0
        return (per, gs, eid)

    rows.sort(key=_sort_key)

    cols = [
        "Event Type",
        "Event ID",
        "Source",
        "Team Raw",
        "Team Side",
        "For/Against",
        "Team Rel",
        "Period",
        "Game Time",
        "Video Time",
        "Game Seconds",
        "Video Seconds",
        "Details",
        "Attributed Players",
        "Attributed Jerseys",
        "On-Ice Players",
        "On-Ice Players (Home)",
        "On-Ice Players (Away)",
    ]
    df = pd.DataFrame([{c: r.get(c, "") for c in cols} for r in rows], columns=cols)
    df.to_csv(stats_dir / "all_events_summary.csv", index=False)
    text_cols = [
        "Event Type",
        "Team Raw",
        "Team Side",
        "For/Against",
        "Team Rel",
        "Details",
        "Attributed Players",
        "Attributed Jerseys",
        "On-Ice Players",
        "On-Ice Players (Home)",
        "On-Ice Players (Away)",
    ]
    xlsx_path = stats_dir / "all_events_summary.xlsx"

    # If we have spreadsheet mapping info, include a mapping table as the first sheet for inspection.
    if spreadsheet_event_mapping_summary and (
        spreadsheet_event_mapping_summary.raw_row_counts or {}
    ):
        mapping_rows: List[Dict[str, Any]] = []
        keys = sorted(
            (spreadsheet_event_mapping_summary.raw_row_counts or {}).keys(),
            key=lambda k: (
                -int((spreadsheet_event_mapping_summary.raw_row_counts or {}).get(k, 0) or 0),
                k[0],
                k[1],
            ),
        )
        for k in keys:
            raw_label, raw_marker = k
            nrows = int((spreadsheet_event_mapping_summary.raw_row_counts or {}).get(k, 0) or 0)
            mapped = (spreadsheet_event_mapping_summary.mapped_counts or {}).get(k, {}) or {}
            unmapped = int(
                (spreadsheet_event_mapping_summary.unmapped_row_counts or {}).get(k, 0) or 0
            )
            mapped_types = sorted(mapped.keys())
            mapping_rows.append(
                {
                    "Spreadsheet Label": raw_label,
                    "Marker": raw_marker,
                    "Rows": nrows,
                    "Unmapped Rows": unmapped,
                    "Mapped Event Types": ", ".join(mapped_types),
                    "Mapped Display Types": ", ".join(
                        _display_event_type(et) for et in mapped_types
                    ),
                    "Mapped Outputs": "; ".join(
                        f"{et}={int(cnt or 0)}"
                        for et, cnt in sorted(mapped.items(), key=lambda x: (-int(x[1] or 0), x[0]))
                    ),
                }
            )
        mapping_df = pd.DataFrame(mapping_rows)
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            # Mapping sheet (first).
            mapping_df_excel = mapping_df.copy()
            mapping_df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in mapping_df_excel.columns
            ]
            mapping_df_excel.to_excel(writer, sheet_name="event_mapping", index=False, startrow=1)
            _apply_excel_table_style(
                writer, "event_mapping", title="Event Mapping", df=mapping_df_excel
            )
            _autosize_columns(writer, "event_mapping", mapping_df_excel)

            # Main events table sheet.
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="all_events", index=False, startrow=1)
            _apply_excel_table_style(writer, "all_events", title="All Events", df=df_excel)
            try:
                ws = writer.sheets.get("all_events")
                if ws is not None:
                    header_row = 2
                    data_start_row = 3
                    nrows = int(getattr(df_excel, "shape", (0, 0))[0] or 0)
                    if nrows > 0:
                        header_cells = list(ws[header_row])
                        col_idx_by_name: Dict[str, int] = {}
                        for idx, cell in enumerate(header_cells, start=1):
                            if cell.value is None:
                                continue
                            col_idx_by_name[str(cell.value).replace("\n", " ").strip()] = idx
                        for col_name in text_cols or []:
                            col_idx = col_idx_by_name.get(str(col_name).strip())
                            if not col_idx:
                                continue
                            for rr in range(data_start_row, data_start_row + nrows):
                                ws.cell(row=rr, column=col_idx).number_format = "@"
            except Exception as e:  # noqa: BLE001
                print(
                    f"[warning] Failed to apply text formatting to 'all_events' sheet: {e}",
                    file=sys.stderr,
                )
            _autosize_columns(writer, "all_events", df_excel)
        return

    _write_styled_xlsx_table(
        xlsx_path,
        df,
        sheet_name="all_events",
        title="All Events",
        text_columns=text_cols,
    )


def _compute_player_stats(
    player_key: str,
    sb_list: List[Tuple[int, str, str]],
    video_pairs_by_player: Dict[str, List[Tuple[str, str]]],
    goals_by_period: Dict[int, List[GoalEvent]],
) -> Tuple[Dict[str, str], Dict[str, int], Dict[str, int], Dict[int, str]]:
    sb_by_period: Dict[int, List[Tuple[str, str]]] = {}
    for period, a, b in sb_list:
        sb_by_period.setdefault(period, []).append((a, b))

    all_pairs = [(a, b) for (_, a, b) in sb_list]
    shift_summary = summarize_shift_lengths_sec(all_pairs)
    per_period_toi_map = per_period_toi(sb_by_period)

    plus_minus = 0
    counted_gf: List[str] = []
    counted_ga: List[str] = []
    counted_gf_by_period: Dict[int, int] = {}
    counted_ga_by_period: Dict[int, int] = {}
    for period, pairs in sb_by_period.items():
        if period not in goals_by_period:
            continue
        for ev in goals_by_period[period]:
            matched = False
            for a, b in pairs:
                a_sec = parse_flex_time_to_seconds(a)
                b_sec = parse_flex_time_to_seconds(b)
                lo, hi = (a_sec, b_sec) if a_sec <= b_sec else (b_sec, a_sec)
                if not (lo <= ev.t_sec <= hi):
                    continue
                if ev.kind == "GA" and ev.t_sec == a_sec:
                    continue
                elif ev.kind == "GF" and ev.t_sec == a_sec:
                    continue
                matched = True
                break
            if matched:
                if ev.kind == "GF":
                    plus_minus += 1
                    counted_gf.append(f"P{period}:{ev.t_str}")
                    counted_gf_by_period[period] = counted_gf_by_period.get(period, 0) + 1
                else:
                    plus_minus -= 1
                    counted_ga.append(f"P{period}:{ev.t_str}")
                    counted_ga_by_period[period] = counted_ga_by_period.get(period, 0) + 1

    row_map: Dict[str, str] = {
        "player": player_key,
        "shifts": shift_summary["num_shifts"],
        "plus_minus": str(plus_minus),
        "sb_toi_total": shift_summary["toi_total"],
        "sb_avg": shift_summary["toi_avg"],
        "sb_median": shift_summary["toi_median"],
        "sb_longest": shift_summary["toi_longest"],
        "sb_shortest": shift_summary["toi_shortest"],
    }
    row_map["gf_counted"] = str(len(counted_gf))
    row_map["ga_counted"] = str(len(counted_ga))

    v_pairs = video_pairs_by_player.get(player_key, [])
    if v_pairs:
        v_sum = 0
        for a, b in v_pairs:
            lo, hi = compute_interval_seconds(a, b)
            v_sum += hi - lo
        row_map["video_toi_total"] = _format_duration(v_sum)
    else:
        row_map["video_toi_total"] = ""

    # per-period values
    per_counts: Dict[str, int] = {}
    per_counts_gf: Dict[str, int] = {}
    for period, toi in per_period_toi_map.items():
        row_map[f"P{period}_toi"] = toi
    for period, pairs in sb_by_period.items():
        per_counts[f"P{period}_shifts"] = len(pairs)
    for period, cnt in counted_gf_by_period.items():
        per_counts_gf[f"P{period}_GF"] = cnt
    for period, cnt in counted_ga_by_period.items():
        per_counts_gf[f"P{period}_GA"] = per_counts_gf.get(
            f"P{period}_GA", 0
        )  # placeholder to ensure keys
    # Return row_map and per-period counts; plus per_period_toi_map for columns
    return row_map, per_counts, {**{k: 0 for k in []}}, per_period_toi_map


def _build_stats_dataframe(
    stats_table_rows: List[Dict[str, str]],
    all_periods_seen: List[int],
    sort_for_cumulative: bool = False,
    *,
    include_shifts_in_stats: bool,
    include_per_game_columns: bool = True,
    include_gp_column: bool = True,
) -> Tuple[pd.DataFrame, List[str]]:
    def _has_any_value(key: str) -> bool:
        for r in stats_table_rows:
            if not r:
                continue
            v = r.get(key, "")
            if v is None:
                continue
            if str(v).strip() != "":
                return True
        return False

    periods = sorted(all_periods_seen)
    # Keep key summary metrics grouped early for easier reading in consolidated sheets.
    summary_cols = ["player", "gp", "goals", "assists", "points", "ppg"]
    summary_cols += [
        "plus_minus",
        "plus_minus_per_game",
        "gf_counted",
        "gf_per_game",
        "ga_counted",
        "ga_per_game",
    ]

    summary_cols += [
        "shots",
        "shots_per_game",
        "sog",
        "expected_goals",
        "sog_per_game",
        "expected_goals_per_game",
        "expected_goals_per_sog",
        "turnovers_forced",
        "turnovers_forced_per_game",
        "created_turnovers",
        "created_turnovers_per_game",
        "giveaways",
        "giveaways_per_game",
        "takeaways",
        "takeaways_per_game",
        "completed_passes",
        "completed_passes_per_game",
        "controlled_entry_for",
        "controlled_entry_for_per_game",
        "controlled_entry_against",
        "controlled_entry_against_per_game",
        "controlled_exit_for",
        "controlled_exit_for_per_game",
        "controlled_exit_against",
        "controlled_exit_against_per_game",
        "gt_goals",
        "gw_goals",
        "ot_goals",
        "ot_assists",
    ]
    if include_shifts_in_stats:
        summary_cols += [
            "shifts",
            "shifts_per_game",
        ]

    sb_cols = (
        ["sb_toi_total", "sb_toi_per_game", "sb_avg", "sb_median", "sb_longest", "sb_shortest"]
        if include_shifts_in_stats
        else []
    )
    if not include_per_game_columns:
        summary_cols = [c for c in summary_cols if c != "ppg" and "_per_game" not in c]
        sb_cols = [c for c in sb_cols if "_per_game" not in c]
    if not include_gp_column:
        summary_cols = [c for c in summary_cols if c != "gp"]
    video_cols = ["video_toi_total"] if include_shifts_in_stats else []
    period_toi_cols = (
        [f"P{p}_toi" for p in periods if _has_any_value(f"P{p}_toi")]
        if include_shifts_in_stats
        else []
    )
    period_shift_cols = (
        [f"P{p}_shifts" for p in periods if _has_any_value(f"P{p}_shifts")]
        if include_shifts_in_stats
        else []
    )
    # Only include per-period goal columns when they have at least one entry.
    period_gf_cols = [f"P{p}_GF" for p in periods if _has_any_value(f"P{p}_GF")]
    period_ga_cols = [f"P{p}_GA" for p in periods if _has_any_value(f"P{p}_GA")]
    # Place video TOI as the last column in the table so that
    # scoreboard-based stats and per-period splits appear first.
    cols = (
        summary_cols
        + sb_cols
        + period_toi_cols
        + period_shift_cols
        + period_gf_cols
        + period_ga_cols
        + video_cols
    )

    rows_sorted: List[Dict[str, str]] = list(stats_table_rows)
    if sort_for_cumulative:
        # Stable multi-key sort:
        #  - final primary key: points (descending)
        #  - tie-breakers (in order of increasing precedence due to stable sort):
        #      player name, then assists, then goals.
        def _intval(row: Dict[str, str], key: str) -> int:
            try:
                return int(str(row.get(key, 0) or 0))
            except Exception:
                return 0

        rows_sorted.sort(key=lambda r: _player_sort_key(r.get("player", "")))
        rows_sorted.sort(key=lambda r: _intval(r, "assists"))
        rows_sorted.sort(key=lambda r: _intval(r, "goals"))
        rows_sorted.sort(key=lambda r: _intval(r, "points"), reverse=True)
    else:
        # Per-game sheets: simple alphabetical order by player.
        rows_sorted.sort(key=lambda r: _player_sort_key(r.get("player", "")))

    rows_for_print: List[List[str]] = [[r.get(c, "") for c in cols] for r in rows_sorted]
    df = pd.DataFrame(rows_for_print, columns=cols)
    return df, cols


def _display_col_name(key: str) -> str:
    """
    Human-friendly column names: remove internal prefixes/suffixes,
    replace underscores with spaces, and capitalize words.
    """
    # Explicit overrides for common fields
    overrides = {
        "jersey": "Jersey #",
        "player": "Player",
        "gp": "GP",
        "goals": "Goals",
        "assists": "Assists",
        "points": "Points",
        "ppg": "PPG",
        "shots": "Shots",
        "shots_per_game": "Shots per Game",
        "sog": "SOG",
        "sog_per_game": "SOG per Game",
        "expected_goals": "xG",
        "expected_goals_per_game": "xG per Game",
        "expected_goals_per_sog": "xG per SOG",
        "turnovers_forced": "Turnovers (forced)",
        "turnovers_forced_per_game": "Turnovers (forced) per Game",
        "created_turnovers": "Created Turnovers",
        "created_turnovers_per_game": "Created Turnovers per Game",
        "giveaways": "Giveaways",
        "giveaways_per_game": "Giveaways per Game",
        "takeaways": "Takeaways",
        "takeaways_per_game": "Takeaways per Game",
        "completed_passes": "Completed Passes",
        "completed_passes_per_game": "Completed Passes per Game",
        "controlled_entry_for": "Controlled Entry For (On-Ice)",
        "controlled_entry_for_per_game": "Controlled Entry For (On-Ice) per Game",
        "controlled_entry_against": "Controlled Entry Against (On-Ice)",
        "controlled_entry_against_per_game": "Controlled Entry Against (On-Ice) per Game",
        "controlled_exit_for": "Controlled Exit For (On-Ice)",
        "controlled_exit_for_per_game": "Controlled Exit For (On-Ice) per Game",
        "controlled_exit_against": "Controlled Exit Against (On-Ice)",
        "controlled_exit_against_per_game": "Controlled Exit Against (On-Ice) per Game",
        "gt_goals": "GT Goals",
        "gw_goals": "GW Goals",
        "ot_goals": "OT Goals",
        "ot_assists": "OT Assists",
        "shifts": "Shifts",
        "shifts_per_game": "Shifts per Game",
        "plus_minus": "Goal +/-",
        "plus_minus_per_game": "Goal +/- per Game",
        "gf_counted": "GF Counted",
        "gf_per_game": "GF per Game",
        "ga_counted": "GA Counted",
        "ga_per_game": "GA per Game",
        "sb_toi_total": "TOI Total",
        "sb_toi_per_game": "TOI per Game",
        "sb_avg": "Average Shift",
        "sb_median": "Median Shift",
        "sb_longest": "Longest Shift",
        "sb_shortest": "Shortest Shift",
        "video_toi_total": "TOI Total (Video)",
    }
    if key in overrides:
        return overrides[key]

    # Period-specific columns
    m = re.fullmatch(r"P(\d+)_toi", key)
    if m:
        return f"Period {m.group(1)} TOI"
    m = re.fullmatch(r"P(\d+)_shifts", key)
    if m:
        return f"Period {m.group(1)} Shifts"
    m = re.fullmatch(r"P(\d+)_GF", key)
    if m:
        return f"Period {m.group(1)} GF"
    m = re.fullmatch(r"P(\d+)_GA", key)
    if m:
        return f"Period {m.group(1)} GA"

    # Generic fallback: split on underscores and capitalize words,
    # preserving common hockey/stat acronyms.
    parts = key.split("_")
    out_parts = []
    for part in parts:
        up = part.upper()
        if up in {"TOI", "GF", "GA", "GT", "GW"}:
            out_parts.append(up)
        else:
            out_parts.append(part.capitalize())
    return " ".join(out_parts)


def _display_player_name(raw: str) -> str:
    """
    Human-friendly player label from an internal key like '59_Ryan_S_Donahue'.
    Format: two-character jersey (right-aligned) + space + name with spaces.
    Example: '59_Ryan_S_Donahue' -> '59 Ryan S Donahue'
             '8_Adam_Ro'        -> ' 8 Adam Ro'
    """
    if not raw:
        return ""
    text = str(raw)
    parts = text.split("_", 1)
    if len(parts) == 2:
        jersey_part, name_part = parts
        # Extract numeric jersey if present; otherwise use the raw jersey_part.
        m = re.search(r"(\\d+)", jersey_part)
        num = m.group(1) if m else jersey_part
        jersey_fmt = f"{num:>2}"
        name = name_part.replace("_", " ").strip()
        return f"{jersey_fmt} {name}"
    # Fallback: just replace underscores with spaces
    return text.replace("_", " ")


def _display_event_type(event_type: str) -> str:
    et = str(event_type or "").strip()
    if et == "ExpectedGoal":
        return "xG"
    if et == "TurnoverForced":
        return "Turnovers (forced)"
    if et == "CreatedTurnover":
        return "Created Turnovers"
    if et == "CompletedPass":
        return "Completed Pass"
    if et == "ControlledEntry":
        return "Controlled Entry"
    if et == "ControlledExit":
        return "Controlled Exit"
    if et in {"Rush", "OddManRush"}:
        return "Odd-Man Rushes"
    return et


def _blink_event_label(event_type: str) -> str:
    """
    Label used for the flashing (blink) overlay in event clip scripts.

    Keep this short, uppercase, and singular (even when the displayed stat is plural).
    """
    et = str(event_type or "").strip()
    if not et:
        return ""

    # Prefer explicit mappings for cases where the displayed stat is plural or includes annotations.
    mapping: Dict[str, str] = {
        "ExpectedGoal": "xG",
        "SOG": "SOG",
        "Goal": "GOAL",
        "Assist": "ASSIST",
        "TurnoverForced": "FORCED TURNOVER",
        "CreatedTurnover": "CREATED TURNOVER",
        "Giveaway": "GIVEAWAY",
        "Takeaway": "TAKEAWAY",
        "CompletedPass": "COMPLETED PASS",
    }
    if et in mapping:
        return mapping[et]

    label = re.sub(r"[()]", "", _display_event_type(et)).strip()
    if label and label != "xG":
        label = label.upper()
    return label


def _write_player_stats_text_and_csv(
    stats_dir: Path,
    stats_table_rows: List[Dict[str, str]],
    all_periods_seen: List[int],
    *,
    include_shifts_in_stats: bool,
) -> None:
    df, cols = _build_stats_dataframe(
        stats_table_rows,
        all_periods_seen,
        sort_for_cumulative=False,
        include_shifts_in_stats=include_shifts_in_stats,
        include_per_game_columns=False,
        include_gp_column=False,
    )
    # Pretty-print player identity for display tables: separate jersey + name.
    if "player" in df.columns and "jersey" not in df.columns:
        jerseys = df["player"].apply(lambda x: _parse_player_key(x).jersey or "")
        df.insert(0, "jersey", jerseys)
        df["player"] = df["player"].apply(_format_player_name_only)
        cols = ["jersey"] + cols
    rows_for_print = df.values.tolist()

    # Human-friendly display column names
    disp_cols = [_display_col_name(c) for c in cols]

    widths = [len(c) for c in disp_cols]
    for row in rows_for_print:
        for i, cell in enumerate(row):
            if len(str(cell)) > widths[i]:
                widths[i] = len(str(cell))

    def fmt_row(values: List[str]) -> str:
        return "  ".join(str(v).ljust(widths[i]) for i, v in enumerate(values))

    lines = [fmt_row(disp_cols)]
    lines.append(fmt_row(["-" * w for w in widths]))
    for row in rows_for_print:
        lines.append(fmt_row(row))
    (stats_dir / "player_stats.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")

    import csv  # local import

    # DataFrame with display headers for CSV/XLSX output
    df_display = df.copy()
    df_display.columns = disp_cols

    csv_rows = [dict(zip(disp_cols, row)) for row in rows_for_print]
    try:
        df_display.to_csv(stats_dir / "player_stats.csv", index=False)
    except Exception:
        with (stats_dir / "player_stats.csv").open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=disp_cols)
            w.writeheader()
            for r in csv_rows:
                w.writerow(r)
    try:
        with pd.ExcelWriter(stats_dir / "player_stats.xlsx", engine="openpyxl") as writer:
            df_excel = df_display.copy()
            df_excel.columns = [_wrap_header_after_words(c, words_per_line=2) for c in disp_cols]
            df_excel.to_excel(writer, sheet_name="player_stats", index=False, startrow=1)
            _apply_excel_table_style(writer, "player_stats", title="Player Stats", df=df_excel)
            _autosize_columns(writer, "player_stats", df_excel)
    except Exception:
        pass


def _write_game_stats_files_t2s_only(
    stats_dir: Path,
    *,
    label: str,
    t2s_id: int,
    goals: List[GoalEvent],
    periods: Optional[List[int]] = None,
) -> None:
    """
    Write per-game stats for a TimeToScore-only game (no spreadsheets).

    This mirrors `_write_game_stats_files(...)` but doesn't require an input XLSX path.
    """
    gf = sum(1 for g in goals if getattr(g, "kind", None) == "GF")
    ga = sum(1 for g in goals if getattr(g, "kind", None) == "GA")
    goal_diff = gf - ga

    row: Dict[str, Any] = {
        "T2S ID": str(int(t2s_id)),
        "Score (For-Against)": f"{gf}-{ga}",
        "Goals For": gf,
        "Goals Against": ga,
        "Goal Diff": goal_diff,
    }

    # Per-period goal totals (from goal list).
    period_set: set[int] = set()
    for x in periods or []:
        if isinstance(x, int):
            period_set.add(int(x))
    for g in goals or []:
        try:
            period_set.add(int(getattr(g, "period", 0) or 0))
        except Exception:
            continue
    for p in sorted({pp for pp in period_set if isinstance(pp, int) and pp > 0}):
        row[f"Period {p} Goals For"] = sum(1 for g in goals if g.kind == "GF" and g.period == p)
        row[f"Period {p} Goals Against"] = sum(1 for g in goals if g.kind == "GA" and g.period == p)

    # No event-log/long-sheet stats available for T2S-only games; keep keys stable with blanks.
    row.update(
        {
            "Shots For": "",
            "Shots Against": "",
            "SOG For": "",
            "SOG Against": "",
            "xG For": "",
            "xG Against": "",
            "xG per SOG (For)": "",
            "xG per SOG (Against)": "",
            "Turnovers (forced) For": "",
            "Turnovers (forced) Against": "",
            "Created Turnovers For": "",
            "Created Turnovers Against": "",
            "Giveaways For": "",
            "Giveaways Against": "",
            "Takeaways For": "",
            "Takeaways Against": "",
            "Controlled Entry For": "",
            "Controlled Entry Against": "",
            "Controlled Exit For": "",
            "Controlled Exit Against": "",
            "Odd-Man Rushes For": "",
            "Odd-Man Rushes Against": "",
        }
    )

    # Transpose: stats are rows, and the game label is the column header.
    df = pd.DataFrame({"Stat": list(row.keys()), label: list(row.values())})
    df.to_csv(stats_dir / "game_stats.csv", index=False)

    try:
        with pd.ExcelWriter(stats_dir / "game_stats.xlsx", engine="openpyxl") as writer:
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="game_stats", index=False, startrow=1)
            _apply_excel_table_style(writer, "game_stats", title="Game Stats", df=df_excel)
            _autosize_columns(writer, "game_stats", df_excel)
    except Exception:
        pass


def _write_game_stats_files(
    stats_dir: Path,
    *,
    xls_path: Path,
    periods: List[int],
    goals: List[GoalEvent],
    event_log_context: Optional[EventLogContext],
    focus_team: Optional[str],
) -> None:
    """
    Write per-game stats as a compact 2-column table (Stat x Value) as CSV + XLSX.

    This intentionally excludes any shift/TOI information.
    """
    label = _base_label_from_path(xls_path)
    t2s_id = _infer_t2s_from_filename(xls_path)

    gf = sum(1 for g in goals if getattr(g, "kind", None) == "GF")
    ga = sum(1 for g in goals if getattr(g, "kind", None) == "GA")
    goal_diff = gf - ga

    row: Dict[str, Any] = {
        "T2S ID": str(t2s_id) if t2s_id is not None else "",
        "Score (For-Against)": f"{gf}-{ga}",
        "Goals For": gf,
        "Goals Against": ga,
        "Goal Diff": goal_diff,
    }

    # Per-period goal totals (from goal list). Include OT if present even when the shift
    # sheet doesn't have an OT section.
    period_set: set[int] = set()
    for x in periods or []:
        if isinstance(x, int):
            period_set.add(int(x))
    for g in goals or []:
        try:
            period_set.add(int(getattr(g, "period", 0) or 0))
        except Exception:
            continue
    for p in sorted({pp for pp in period_set if isinstance(pp, int) and pp > 0}):
        row[f"Period {p} Goals For"] = sum(1 for g in goals if g.kind == "GF" and g.period == p)
        row[f"Period {p} Goals Against"] = sum(1 for g in goals if g.kind == "GA" and g.period == p)

    # Event-based team stats (from event log context; usually from '*-long*' sheets)
    counts = (event_log_context.event_counts_by_type_team if event_log_context else None) or {}
    has_event_counts = bool(counts)

    def _for_against(event_type: str) -> Tuple[Any, Any]:
        if not has_event_counts or focus_team not in {"Blue", "White"}:
            return "", ""
        opp = "White" if focus_team == "Blue" else "Blue"
        k_for = (event_type, focus_team)
        k_against = (event_type, opp)
        # If this stat type wasn't collected for this game, leave blank (not 0).
        if k_for not in counts and k_against not in counts:
            return "", ""
        return int(counts.get(k_for, 0) or 0), int(counts.get(k_against, 0) or 0)

    shots_for, shots_against = _for_against("Shot")
    sog_for, sog_against = _for_against("SOG")
    # Fallback: when sheets only log generic Shots (no explicit SOG), treat Shots as SOG.
    if sog_for == "" and isinstance(shots_for, (int, float)):
        sog_for = shots_for
    if sog_against == "" and isinstance(shots_against, (int, float)):
        sog_against = shots_against
    xg_for, xg_against = _for_against("ExpectedGoal")
    turnovers_forced_for, turnovers_forced_against = _for_against("TurnoverForced")
    created_turnovers_for, created_turnovers_against = _for_against("CreatedTurnover")
    giveaways_for, giveaways_against = _for_against("Giveaway")
    takeaways_for, takeaways_against = _for_against("Takeaway")
    completed_pass_for, completed_pass_against = _for_against("CompletedPass")
    ce_for, ce_against = _for_against("ControlledEntry")
    cx_for, cx_against = _for_against("ControlledExit")
    rush_for, rush_against = _for_against("Rush")

    # Keep columns stable; leave blank when mapping isn't available.
    row.update(
        {
            "Shots For": shots_for,
            "Shots Against": shots_against,
            "SOG For": sog_for,
            "SOG Against": sog_against,
            "xG For": xg_for,
            "xG Against": xg_against,
            "xG per SOG (For)": (
                f"{(int(xg_for) / int(sog_for)):.2f}"
                if isinstance(xg_for, (int, float))
                and isinstance(sog_for, (int, float))
                and int(sog_for) > 0
                else ""
            ),
            "xG per SOG (Against)": (
                f"{(int(xg_against) / int(sog_against)):.2f}"
                if isinstance(xg_against, (int, float))
                and isinstance(sog_against, (int, float))
                and int(sog_against) > 0
                else ""
            ),
            "Turnovers (forced) For": turnovers_forced_for,
            "Turnovers (forced) Against": turnovers_forced_against,
            "Created Turnovers For": created_turnovers_for,
            "Created Turnovers Against": created_turnovers_against,
            "Giveaways For": giveaways_for,
            "Giveaways Against": giveaways_against,
            "Takeaways For": takeaways_for,
            "Takeaways Against": takeaways_against,
            "Completed Pass For": completed_pass_for,
            "Completed Pass Against": completed_pass_against,
            "Controlled Entry For": ce_for,
            "Controlled Entry Against": ce_against,
            "Controlled Exit For": cx_for,
            "Controlled Exit Against": cx_against,
            "Odd-Man Rushes For": rush_for,
            "Odd-Man Rushes Against": rush_against,
        }
    )

    # Transpose: stats are rows, and the game label is the column header.
    df = pd.DataFrame({"Stat": list(row.keys()), label: list(row.values())})
    df.to_csv(stats_dir / "game_stats.csv", index=False)

    try:
        with pd.ExcelWriter(stats_dir / "game_stats.xlsx", engine="openpyxl") as writer:
            df_excel = df.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="game_stats", index=False, startrow=1)
            _apply_excel_table_style(writer, "game_stats", title="Game Stats", df=df_excel)
            _autosize_columns(writer, "game_stats", df_excel)
    except Exception:
        pass


def _load_game_stats_csv_as_series(path: Path) -> Tuple[str, pd.Series]:
    """
    Load stats/game_stats.csv (2-column "Stat", "<game_label>") into a Series indexed by Stat.
    Returns (value_column_label, series).
    """
    df = pd.read_csv(path)
    if "Stat" not in df.columns:
        raise ValueError("missing Stat column")
    value_cols = [c for c in df.columns if c != "Stat"]
    if len(value_cols) != 1:
        raise ValueError("expected exactly one value column")
    value_col = str(value_cols[0])
    series = df.set_index("Stat")[value_col]
    # Preserve row order from the CSV.
    return value_col, series


def _write_game_stats_consolidated_files(
    base_outdir: Path,
    ordered_results: List[Dict[str, Any]],
) -> bool:
    """
    Join per-game stats/game_stats.csv tables into a single consolidated CSV/XLSX:
      - rows: Stat
      - columns: each game's label (value column header from game_stats.csv)
    """
    series_list: List[pd.Series] = []
    for r in ordered_results:
        outdir_val = r.get("outdir")
        if not outdir_val:
            continue
        outdir = Path(outdir_val)
        stats_csv = outdir / "stats" / "game_stats.csv"
        if not stats_csv.exists():
            continue
        try:
            value_col, series = _load_game_stats_csv_as_series(stats_csv)
        except Exception as e:  # noqa: BLE001
            print(f"[game_stats_consolidated] Failed to read {stats_csv}: {e}", file=sys.stderr)
            continue
        # Prefer the label from the CSV (matches per-game output); fall back to the group's label.
        label = str(value_col or r.get("label") or outdir.name)
        series_list.append(series.rename(label))

    if not series_list:
        return False

    df = pd.concat(series_list, axis=1, sort=False)
    df.index.name = "Stat"
    df_out = df.reset_index()

    csv_path = base_outdir / "game_stats_consolidated.csv"
    xlsx_path = base_outdir / "game_stats_consolidated.xlsx"
    try:
        df_out.to_csv(csv_path, index=False)
    except Exception as e:  # noqa: BLE001
        print(f"[game_stats_consolidated] Failed to write {csv_path}: {e}", file=sys.stderr)
        return False

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df_excel = df_out.copy()
            df_excel.columns = [
                _wrap_header_after_words(str(c), words_per_line=2) for c in df_excel.columns
            ]
            df_excel.to_excel(writer, sheet_name="game_stats", index=False, startrow=1)
            _apply_excel_table_style(
                writer, "game_stats", title="Game Stats (All Games)", df=df_excel
            )
            _autosize_columns(writer, "game_stats", df_excel)
    except Exception as e:  # noqa: BLE001
        print(f"[game_stats_consolidated] Failed to write {xlsx_path}: {e}", file=sys.stderr)
        return False

    return True


def _write_pair_on_ice_consolidated_files(
    base_outdir: Path,
    ordered_results: List[Dict[str, Any]],
    *,
    include_toi: bool,
) -> bool:
    """
    Join per-game pair-on-ice rows into a single consolidated CSV/XLSX.

    This avoids reading the per-game CSVs so we can:
      - keep full-precision overlap math internally
      - optionally omit absolute TOI publishing when `--shifts` is not set
    """
    agg: Dict[Tuple[str, str], Dict[str, int]] = {}
    # Compute per-player total +/- from the per-game player stats rows so that:
    #  - totals match the consolidated player stats sheets
    #  - games with shift data for only one skater still contribute to totals
    total_pm_by_player: Dict[str, int] = {}
    for r in ordered_results:
        for prow in r.get("stats") or []:
            try:
                player = str(prow.get("player") or "").strip()
                pm_raw = prow.get("plus_minus", "")
                if not player or str(pm_raw).strip() == "":
                    continue
                total_pm_by_player[player] = int(total_pm_by_player.get(player, 0) or 0) + int(
                    str(pm_raw)
                )
            except Exception:
                continue

        for raw in r.get("pair_on_ice") or []:
            try:
                player = str(raw.get("player") or "").strip()
                teammate = str(raw.get("teammate") or "").strip()
                if not player or not teammate:
                    continue

                key = (player, teammate)
                dest = agg.setdefault(
                    key,
                    {
                        "shift_games": 0,
                        "player_toi_seconds": 0,
                        "overlap_seconds": 0,
                        "gf_together": 0,
                        "ga_together": 0,
                        "player_goals_on_ice_together": 0,
                        "player_assists_on_ice_together": 0,
                        "goals_collab_with_teammate": 0,
                        "assists_collab_with_teammate": 0,
                    },
                )
                dest["shift_games"] += int(raw.get("shift_games", 1) or 1)
                dest["player_toi_seconds"] += int(raw.get("player_toi_seconds", 0) or 0)
                dest["overlap_seconds"] += int(raw.get("overlap_seconds", 0) or 0)
                dest["gf_together"] += int(raw.get("gf_together", 0) or 0)
                dest["ga_together"] += int(raw.get("ga_together", 0) or 0)
                dest["player_goals_on_ice_together"] += int(
                    raw.get("player_goals_on_ice_together", 0) or 0
                )
                dest["player_assists_on_ice_together"] += int(
                    raw.get("player_assists_on_ice_together", 0) or 0
                )
                dest["goals_collab_with_teammate"] += int(
                    raw.get("goals_collab_with_teammate", 0) or 0
                )
                dest["assists_collab_with_teammate"] += int(
                    raw.get("assists_collab_with_teammate", 0) or 0
                )
            except Exception:
                continue

    rows: List[Dict[str, Any]] = []
    if not agg:
        # Still write an empty consolidated file set so users always see it
        # alongside other consolidated outputs.
        empty_cols = [
            "Player Jersey #",
            "Player",
            "Teammate Jersey #",
            "Teammate",
            "Games with Data",
            "Overlap %",
            "GF Together",
            "GA Together",
            "Player Goals (On Ice Together)",
            "Player Assists (On Ice Together)",
            "Goals Collaborated",
            "Assists Collaborated",
            "+/- Together",
            "Player Total +/-",
            "Teammate Total +/-",
        ]
        if include_toi:
            try:
                idx = empty_cols.index("Overlap %")
            except Exception:
                idx = len(empty_cols)
            empty_cols.insert(idx, "Player TOI")
            empty_cols.insert(idx + 1, "Overlap")
        df_out = pd.DataFrame([], columns=empty_cols)
        csv_out = base_outdir / "pair_on_ice_consolidated.csv"
        xlsx_out = base_outdir / "pair_on_ice_consolidated.xlsx"
        try:
            df_out.to_csv(csv_out, index=False)
        except Exception:
            return False
        _write_styled_xlsx_table(
            xlsx_out,
            df_out,
            sheet_name="pair_on_ice",
            title="Pair On-Ice (All Games)",
            number_formats={"Overlap %": "0.0"},
            text_columns=["Player", "Teammate"],
        )
        return True

    def _blank_if_zero(x: Any) -> Any:
        try:
            v = int(x or 0)
        except Exception:
            v = 0
        return "" if v == 0 else v

    for (player, teammate), d in agg.items():
        p_parts = _parse_player_key(player)
        t_parts = _parse_player_key(teammate)
        toi = int(d.get("player_toi_seconds", 0) or 0)
        overlap = int(d.get("overlap_seconds", 0) or 0)
        gf = int(d.get("gf_together", 0) or 0)
        ga = int(d.get("ga_together", 0) or 0)
        pct = 100.0 * overlap / toi if toi > 0 else 0.0
        row_out: Dict[str, Any] = {
            "Player Jersey #": p_parts.jersey or "",
            "Player": p_parts.name,
            "Teammate Jersey #": t_parts.jersey or "",
            "Teammate": t_parts.name,
            "Games with Data": int(d.get("shift_games", 0) or 0),
            "Overlap %": pct,
            "GF Together": gf,
            "GA Together": ga,
            "Player Goals (On Ice Together)": _blank_if_zero(
                d.get("player_goals_on_ice_together", 0)
            ),
            "Player Assists (On Ice Together)": _blank_if_zero(
                d.get("player_assists_on_ice_together", 0)
            ),
            "Goals Collaborated": _blank_if_zero(d.get("goals_collab_with_teammate", 0)),
            "Assists Collaborated": _blank_if_zero(d.get("assists_collab_with_teammate", 0)),
            "+/- Together": gf - ga,
            "Player Total +/-": int(total_pm_by_player.get(player, 0) or 0),
            "Teammate Total +/-": int(total_pm_by_player.get(teammate, 0) or 0),
        }
        if include_toi:
            row_out["Player TOI"] = _seconds_to_compact_hms(toi)
            row_out["Overlap"] = _seconds_to_compact_hms(overlap)
        rows.append(row_out)

    df_out = pd.DataFrame(rows)
    try:
        df_out.sort_values(by=["Player", "Overlap %"], ascending=[True, False], inplace=True)
    except Exception:
        pass

    csv_out = base_outdir / "pair_on_ice_consolidated.csv"
    xlsx_out = base_outdir / "pair_on_ice_consolidated.xlsx"
    try:
        df_out.to_csv(csv_out, index=False)
    except Exception:
        return False

    _write_styled_xlsx_table(
        xlsx_out,
        df_out,
        sheet_name="pair_on_ice",
        title="Pair On-Ice (All Games)",
        number_formats={"Overlap %": "0.0"},
        text_columns=["Player", "Teammate"],
        align_right_columns=["Player TOI", "Overlap"],
    )
    return True


def _aggregate_stats_rows(
    stats_sets: List[Tuple[List[Dict[str, str]], List[int]]],
) -> Tuple[List[Dict[str, str]], List[int], Dict[str, int]]:
    def _game_has_any_value(rows: List[Dict[str, str]], key: str) -> bool:
        for r in rows:
            if not r:
                continue
            v = r.get(key, "")
            if v is None:
                continue
            if str(v).strip() != "":
                return True
        return False

    # Count how many games have player-attributed stats available for each
    # long-sheet-derived stat key. This is used as the denominator for per-game
    # averages so that games without a `*-long*` sheet (or without a newer stat
    # type) don't dilute the per-game numbers.
    long_stat_keys = [
        "shots",
        "sog",
        "expected_goals",
        "turnovers_forced",
        "created_turnovers",
        "giveaways",
        "takeaways",
        "completed_passes",
        "controlled_entry_for",
        "controlled_entry_against",
        "controlled_exit_for",
        "controlled_exit_against",
    ]
    games_with_long_stat: Dict[str, int] = {k: 0 for k in long_stat_keys}
    for rows, _periods in stats_sets:
        for k in long_stat_keys:
            if _game_has_any_value(rows, k):
                games_with_long_stat[k] += 1

    per_game_denoms: Dict[str, int] = {f"{k}_per_game": v for k, v in games_with_long_stat.items()}

    agg: Dict[str, Dict[str, Any]] = {}
    all_periods: set[int] = set()

    def _ensure(player: str) -> Dict[str, Any]:
        if player not in agg:
            agg[player] = {
                "player": player,
                "goals": 0,
                "assists": 0,
                "ot_goals": 0,
                "ot_assists": 0,
                "shots": 0,
                "sog": 0,
                "expected_goals": 0,
                "turnovers_forced": 0,
                "created_turnovers": 0,
                "giveaways": 0,
                "takeaways": 0,
                "completed_passes": 0,
                "controlled_entry_for": 0,
                "controlled_entry_against": 0,
                "controlled_exit_for": 0,
                "controlled_exit_against": 0,
                "gp": 0,
                "shifts": 0,
                "plus_minus": 0,
                "gf_counted": 0,
                "ga_counted": 0,
                "sb_toi_total_sec": 0,
                "video_toi_total_sec": 0,
                "sb_longest_sec": 0,
                "sb_shortest_sec": None,
            }
        return agg[player]

    for rows, periods in stats_sets:
        for p in periods:
            all_periods.add(p)
        for row in rows:
            player = row.get("player", "")
            if not player:
                continue
            dest = _ensure(player)
            dest["goals"] += int(str(row.get("goals", 0) or 0))
            dest["assists"] += int(str(row.get("assists", 0) or 0))
            dest["ot_goals"] += int(str(row.get("ot_goals", 0) or 0))
            dest["ot_assists"] += int(str(row.get("ot_assists", 0) or 0))
            dest["shots"] += int(str(row.get("shots", 0) or 0))
            dest["sog"] += int(str(row.get("sog", 0) or 0))
            dest["expected_goals"] += int(str(row.get("expected_goals", 0) or 0))
            dest["turnovers_forced"] += int(str(row.get("turnovers_forced", 0) or 0))
            dest["created_turnovers"] += int(str(row.get("created_turnovers", 0) or 0))
            dest["giveaways"] += int(str(row.get("giveaways", 0) or 0))
            dest["takeaways"] += int(str(row.get("takeaways", 0) or 0))
            dest["completed_passes"] += int(str(row.get("completed_passes", 0) or 0))
            dest["controlled_entry_for"] += int(str(row.get("controlled_entry_for", 0) or 0))
            dest["controlled_entry_against"] += int(
                str(row.get("controlled_entry_against", 0) or 0)
            )
            dest["controlled_exit_for"] += int(str(row.get("controlled_exit_for", 0) or 0))
            dest["controlled_exit_against"] += int(str(row.get("controlled_exit_against", 0) or 0))
            # Each per-game stats row corresponds to one game played (GP),
            # including cases where the player only appears on the T2S roster.
            dest["gp"] += 1
            shifts_raw = str(row.get("shifts", "") or "").strip()
            shifts_i = int(str(shifts_raw or 0))
            dest["shifts"] += shifts_i
            dest["plus_minus"] += int(str(row.get("plus_minus", 0) or 0))
            dest["gf_counted"] += int(str(row.get("gf_counted", 0) or 0))
            dest["ga_counted"] += int(str(row.get("ga_counted", 0) or 0))

            # Per-shift rates are intentionally not computed (too small/noisy).

            dest["sb_toi_total_sec"] += _duration_to_seconds(row.get("sb_toi_total", ""))
            dest["video_toi_total_sec"] += _duration_to_seconds(row.get("video_toi_total", ""))
            longest = _duration_to_seconds(row.get("sb_longest", ""))
            if longest > dest["sb_longest_sec"]:
                dest["sb_longest_sec"] = longest
            shortest = _duration_to_seconds(row.get("sb_shortest", ""))
            if shortest > 0:
                if dest["sb_shortest_sec"] is None or shortest < dest["sb_shortest_sec"]:
                    dest["sb_shortest_sec"] = shortest
            # per-period counts and toi
            for key, val in row.items():
                if not isinstance(key, str):
                    continue
                if re.fullmatch(r"P\d+_shifts", key):
                    dest[key] = dest.get(key, 0) + int(str(val or 0))
                elif re.fullmatch(r"P\d+_GF", key):
                    dest[key] = dest.get(key, 0) + int(str(val or 0))
                elif re.fullmatch(r"P\d+_GA", key):
                    dest[key] = dest.get(key, 0) + int(str(val or 0))
                elif re.fullmatch(r"P\d+_toi", key):
                    dest[key] = dest.get(key, 0) + _duration_to_seconds(val or "")

    aggregated_rows: List[Dict[str, str]] = []
    for player, data in sorted(agg.items(), key=lambda x: x[0]):
        gp = data.get("gp", 0) or 0
        shifts = data["shifts"] or 0
        total_sec = data["sb_toi_total_sec"]
        avg_sec = int(total_sec / shifts) if shifts else 0
        total_goals = data["goals"]
        total_assists = data["assists"]
        total_ot_goals = data.get("ot_goals", 0) or 0
        total_ot_assists = data.get("ot_assists", 0) or 0
        total_points = total_goals + total_assists
        total_shots = data.get("shots", 0) or 0
        total_sog = data.get("sog", 0) or 0
        total_expected_goals = data.get("expected_goals", 0) or 0
        total_turnovers_forced = data.get("turnovers_forced", 0) or 0
        total_created_turnovers = data.get("created_turnovers", 0) or 0
        total_giveaways = data.get("giveaways", 0) or 0
        total_takeaways = data.get("takeaways", 0) or 0
        total_completed_passes = data.get("completed_passes", 0) or 0
        total_ce_for = data.get("controlled_entry_for", 0) or 0
        total_ce_against = data.get("controlled_entry_against", 0) or 0
        total_cx_for = data.get("controlled_exit_for", 0) or 0
        total_cx_against = data.get("controlled_exit_against", 0) or 0
        shots_games = per_game_denoms.get("shots_per_game", 0) or 0
        sog_games = per_game_denoms.get("sog_per_game", 0) or 0
        xg_games = per_game_denoms.get("expected_goals_per_game", 0) or 0
        turnovers_forced_games = per_game_denoms.get("turnovers_forced_per_game", 0) or 0
        created_turnovers_games = per_game_denoms.get("created_turnovers_per_game", 0) or 0
        giveaway_games = per_game_denoms.get("giveaways_per_game", 0) or 0
        takeaway_games = per_game_denoms.get("takeaways_per_game", 0) or 0
        completed_pass_games = per_game_denoms.get("completed_passes_per_game", 0) or 0
        ce_for_games = per_game_denoms.get("controlled_entry_for_per_game", 0) or 0
        ce_against_games = per_game_denoms.get("controlled_entry_against_per_game", 0) or 0
        cx_for_games = per_game_denoms.get("controlled_exit_for_per_game", 0) or 0
        cx_against_games = per_game_denoms.get("controlled_exit_against_per_game", 0) or 0
        row: Dict[str, str] = {
            "player": player,
            "gp": str(gp),
            "goals": str(total_goals),
            "assists": str(total_assists),
            "ot_goals": str(total_ot_goals),
            "ot_assists": str(total_ot_assists),
            "points": str(total_points),
            "ppg": f"{(total_points / gp):.1f}" if gp > 0 else "0.0",
            "shots": str(total_shots) if shots_games > 0 else "",
            "shots_per_game": f"{(total_shots / shots_games):.1f}" if shots_games > 0 else "",
            "sog": str(total_sog) if sog_games > 0 else "",
            "sog_per_game": f"{(total_sog / sog_games):.1f}" if sog_games > 0 else "",
            "expected_goals": str(total_expected_goals) if xg_games > 0 else "",
            "expected_goals_per_game": (
                f"{(total_expected_goals / xg_games):.1f}" if xg_games > 0 else ""
            ),
            "expected_goals_per_sog": (
                f"{(total_expected_goals / total_sog):.2f}"
                if xg_games > 0 and sog_games > 0 and total_sog > 0
                else ""
            ),
            "turnovers_forced": str(total_turnovers_forced) if turnovers_forced_games > 0 else "",
            "turnovers_forced_per_game": (
                f"{(total_turnovers_forced / turnovers_forced_games):.1f}"
                if turnovers_forced_games > 0
                else ""
            ),
            "created_turnovers": (
                str(total_created_turnovers) if created_turnovers_games > 0 else ""
            ),
            "created_turnovers_per_game": (
                f"{(total_created_turnovers / created_turnovers_games):.1f}"
                if created_turnovers_games > 0
                else ""
            ),
            "giveaways": str(total_giveaways) if giveaway_games > 0 else "",
            "giveaways_per_game": (
                f"{(total_giveaways / giveaway_games):.1f}" if giveaway_games > 0 else ""
            ),
            "takeaways": str(total_takeaways) if takeaway_games > 0 else "",
            "takeaways_per_game": (
                f"{(total_takeaways / takeaway_games):.1f}" if takeaway_games > 0 else ""
            ),
            "completed_passes": str(total_completed_passes) if completed_pass_games > 0 else "",
            "completed_passes_per_game": (
                f"{(total_completed_passes / completed_pass_games):.1f}"
                if completed_pass_games > 0
                else ""
            ),
            "controlled_entry_for": str(total_ce_for) if ce_for_games > 0 else "",
            "controlled_entry_for_per_game": (
                f"{(total_ce_for / ce_for_games):.1f}" if ce_for_games > 0 else ""
            ),
            "controlled_entry_against": str(total_ce_against) if ce_against_games > 0 else "",
            "controlled_entry_against_per_game": (
                f"{(total_ce_against / ce_against_games):.1f}" if ce_against_games > 0 else ""
            ),
            "controlled_exit_for": str(total_cx_for) if cx_for_games > 0 else "",
            "controlled_exit_for_per_game": (
                f"{(total_cx_for / cx_for_games):.1f}" if cx_for_games > 0 else ""
            ),
            "controlled_exit_against": str(total_cx_against) if cx_against_games > 0 else "",
            "controlled_exit_against_per_game": (
                f"{(total_cx_against / cx_against_games):.1f}" if cx_against_games > 0 else ""
            ),
            "shifts": str(shifts),
            "shifts_per_game": f"{(shifts / gp):.1f}" if gp > 0 else "",
            "plus_minus": str(data["plus_minus"]),
            "plus_minus_per_game": (f"{(data['plus_minus'] / gp):.1f}" if gp > 0 else ""),
            "gf_counted": str(data["gf_counted"]),
            "gf_per_game": f"{(data['gf_counted'] / gp):.1f}" if gp > 0 else "",
            "ga_counted": str(data["ga_counted"]),
            "ga_per_game": f"{(data['ga_counted'] / gp):.1f}" if gp > 0 else "",
            "sb_toi_total": _format_duration(total_sec),
            "sb_toi_per_game": (
                _format_duration(int(total_sec / gp)) if gp > 0 and total_sec > 0 else ""
            ),
            "sb_avg": _format_duration(avg_sec) if shifts else "",
            "sb_median": "",
            "sb_longest": _format_duration(data["sb_longest_sec"]),
            "sb_shortest": (
                _format_duration(data["sb_shortest_sec"] or 0) if data["sb_shortest_sec"] else ""
            ),
            "video_toi_total": _format_duration(data["video_toi_total_sec"]),
        }

        for p in sorted(all_periods):
            toi_key = f"P{p}_toi"
            shift_key = f"P{p}_shifts"
            gf_key = f"P{p}_GF"
            ga_key = f"P{p}_GA"
            if toi_key in data:
                row[toi_key] = _format_duration(int(data[toi_key]))
            if shift_key in data:
                row[shift_key] = str(data[shift_key])
            if gf_key in data:
                row[gf_key] = str(data[gf_key])
            if ga_key in data:
                row[ga_key] = str(data[ga_key])
        aggregated_rows.append(row)

    return aggregated_rows, sorted(all_periods), per_game_denoms


def _augment_aggregate_with_goal_details(
    aggregated_rows: List[Dict[str, str]],
    per_player_events: Dict[str, Dict[str, List[Tuple[str, GoalEvent]]]],
) -> None:
    """
    Enrich aggregated rows with game-tying / game-winning goal counts
    derived from per-player scoring events across all games.
    """
    if not aggregated_rows:
        return

    totals: Dict[str, Dict[str, int]] = {}
    for player, events in per_player_events.items():
        gt = 0
        gw = 0
        goals = events.get("goals", [])
        for _label, ev in goals:
            if getattr(ev, "is_game_tying", False):
                gt += 1
            if getattr(ev, "is_game_winning", False):
                gw += 1
        totals[player] = {"gt_goals": gt, "gw_goals": gw}

    for row in aggregated_rows:
        player = row.get("player", "")
        t = totals.get(player, {})
        row["gt_goals"] = str(t.get("gt_goals", 0))
        row["gw_goals"] = str(t.get("gw_goals", 0))


def _count_goal_role_flags(goal_events: List[GoalEvent]) -> Tuple[int, int]:
    """
    Returns (gt_goals, gw_goals) for the given list of goal events.
    """
    gt = 0
    gw = 0
    for ev in goal_events or []:
        if getattr(ev, "is_game_tying", False):
            gt += 1
        if getattr(ev, "is_game_winning", False):
            gw += 1
    return gt, gw


def _write_consolidated_workbook(
    out_path: Path,
    sheets: List[Tuple[str, pd.DataFrame]],
    *,
    per_game_denoms: Optional[Dict[str, int]] = None,
) -> None:
    def _disp_col(key: str, *, is_cumulative: bool) -> str:
        base = _display_col_name(key)
        if is_cumulative and per_game_denoms and key in per_game_denoms:
            return f"{base} ({per_game_denoms[key]})"
        return base

    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for name, df in sheets:
                safe_name = re.sub(r"[:\\\\/?*\\[\\]]", "_", name or "Sheet")[:31]
                df_display = df.copy()
                # Pretty player identity if present: separate jersey + name.
                if "player" in df_display.columns:
                    if "jersey" not in df_display.columns:
                        jerseys = df_display["player"].apply(
                            lambda x: _parse_player_key(x).jersey or ""
                        )
                        df_display.insert(0, "jersey", jerseys)
                    df_display["player"] = df_display["player"].apply(_format_player_name_only)
                is_cumulative = str(name or "").strip().lower() == "cumulative"
                disp_cols = [
                    _wrap_header_after_words(
                        _disp_col(c, is_cumulative=is_cumulative), words_per_line=2
                    )
                    for c in df_display.columns
                ]
                df_display.columns = disp_cols
                df_display.to_excel(writer, sheet_name=safe_name, index=False, startrow=1)
                _apply_excel_table_style(
                    writer, safe_name, title=(name or safe_name), df=df_display
                )
                _autosize_columns(writer, safe_name, df_display)
    except Exception:
        pass


def _write_cumulative_player_detail_files(
    base_outdir: Path,
    aggregated_rows: List[Dict[str, str]],
    per_player_events: Dict[str, Dict[str, List[Tuple[str, GoalEvent]]]],
    per_game_stats_by_label: Dict[str, List[Dict[str, str]]],
    per_game_pair_on_ice_by_label: Dict[str, List[Dict[str, Any]]],
    *,
    include_shifts_in_stats: bool,
) -> None:
    """
    Write one cumulative per-player stats file summarizing all games,
    including lists of goals, assists, and goals-against with game/period/time
    and game-tying / game-winning annotations for goals.
    """
    if not aggregated_rows:
        return

    outdir = base_outdir / "cumulative_per_player"
    outdir.mkdir(parents=True, exist_ok=True)

    rows_by_player: Dict[str, Dict[str, str]] = {
        r.get("player", ""): r for r in aggregated_rows if r.get("player")
    }

    # Determine which game each player's longest / shortest shift came from (optional).
    longest_by_player: Dict[str, Tuple[int, str, str]] = {}
    shortest_by_player: Dict[str, Tuple[int, str, str]] = {}
    if include_shifts_in_stats:
        for game_label, rows in per_game_stats_by_label.items():
            for row in rows:
                player = row.get("player", "")
                if not player:
                    continue
                sb_long = _duration_to_seconds(row.get("sb_longest", ""))
                if sb_long > 0:
                    cur = longest_by_player.get(player)
                    if cur is None or sb_long > cur[0]:
                        longest_by_player[player] = (sb_long, row.get("sb_longest", ""), game_label)
                sb_short = _duration_to_seconds(row.get("sb_shortest", ""))
                if sb_short > 0:
                    cur_s = shortest_by_player.get(player)
                    if cur_s is None or sb_short < cur_s[0]:
                        shortest_by_player[player] = (
                            sb_short,
                            row.get("sb_shortest", ""),
                            game_label,
                        )

    pair_overlap: Dict[Tuple[str, str], Dict[str, int]] = {}
    if per_game_pair_on_ice_by_label:
        for _game_label, rows in per_game_pair_on_ice_by_label.items():
            for raw in rows or []:
                try:
                    player = str(raw.get("player") or "").strip()
                    teammate = str(raw.get("teammate") or "").strip()
                    if not player or not teammate:
                        continue
                    key = (player, teammate)
                    dest = pair_overlap.setdefault(
                        key,
                        {
                            "overlap_seconds": 0,
                            "player_toi_seconds": 0,
                            "gf_together": 0,
                            "ga_together": 0,
                            "shift_games": 0,
                        },
                    )
                    dest["overlap_seconds"] += int(raw.get("overlap_seconds", 0) or 0)
                    dest["player_toi_seconds"] += int(raw.get("player_toi_seconds", 0) or 0)
                    dest["gf_together"] += int(raw.get("gf_together", 0) or 0)
                    dest["ga_together"] += int(raw.get("ga_together", 0) or 0)
                    dest["shift_games"] += int(raw.get("shift_games", 1) or 1)
                except Exception:
                    continue

    def _fmt_tags(ev: GoalEvent) -> str:
        tags: List[str] = []
        if getattr(ev, "is_game_tying", False):
            tags.append("GT")
        if getattr(ev, "is_game_winning", False):
            tags.append("GW")
        return f" [{' '.join(tags)}]" if tags else ""

    for player, row in sorted(rows_by_player.items()):
        events = per_player_events.get(player, {})
        goals_list = events.get("goals", [])
        assists_list = events.get("assists", [])
        gf_on_ice_list = events.get("gf_on_ice", [])
        ga_list = events.get("ga_on_ice", [])

        lines: List[str] = []
        lines.append(f"Player: {_display_player_name(player)}")
        lines.append("")
        lines.append("Overall stats (all games):")
        gp_str = row.get("gp", "0")
        points_str = row.get("points", "0")
        ppg_str = row.get("ppg", "0.0")
        lines.append(f"  Games Played (GP): {gp_str}")
        lines.append(f"  Points (G+A): {points_str} (PPG: {ppg_str})")
        lines.append(
            f"  Goals: {row.get('goals', '0')} "
            f"(GT: {row.get('gt_goals', '0')}, GW: {row.get('gw_goals', '0')}, OT: {row.get('ot_goals', '0')})"
        )
        lines.append(f"  Assists: {row.get('assists', '0')} (OT: {row.get('ot_assists', '0')})")
        lines.append(f"  Goal +/-: {row.get('plus_minus', '0')}")
        if include_shifts_in_stats and row.get("shifts_per_game"):
            lines.append(f"  Shifts per game: {row.get('shifts_per_game')}")
        if row.get("plus_minus_per_game"):
            lines.append(f"  Goal +/- per game: {row.get('plus_minus_per_game')}")
        lines.append(
            f"  Goals For while on ice: {row.get('gf_counted', '0')}, "
            f"Goals Against while on ice: {row.get('ga_counted', '0')}"
        )
        if row.get("gf_per_game") or row.get("ga_per_game"):
            lines.append(
                f"  GF per game: {row.get('gf_per_game', '') or '0.0'}, "
                f"GA per game: {row.get('ga_per_game', '') or '0.0'}"
            )
        if include_shifts_in_stats:
            if row.get("sb_toi_total"):
                lines.append(f"  TOI total (scoreboard): {row.get('sb_toi_total')}")
                if row.get("sb_toi_per_game"):
                    lines.append(f"  TOI per game (scoreboard): {row.get('sb_toi_per_game')}")
            if row.get("video_toi_total"):
                lines.append(f"  TOI total (video): {row.get('video_toi_total')}")
            # Longest/shortest shift games
            long_info = longest_by_player.get(player)
            if long_info is not None:
                _, dur, game_label = long_info
                lines.append(f"  Longest shift (scoreboard): {dur} ({game_label})")
            short_info = shortest_by_player.get(player)
            if short_info is not None:
                _, dur_s, game_s = short_info
                lines.append(f"  Shortest shift (scoreboard): {dur_s} ({game_s})")

            # Pair on-ice overlap across all games that had shift data.
            teammates = sorted({t for (p, t) in pair_overlap.keys() if p == player and t})
            if teammates:
                rows_disp: List[Tuple[str, float, int, int, int, int, int]] = []
                for teammate in teammates:
                    d = pair_overlap.get(
                        (player, teammate),
                        {
                            "overlap_seconds": 0,
                            "player_toi_seconds": 0,
                            "gf_together": 0,
                            "ga_together": 0,
                            "shift_games": 0,
                        },
                    )
                    overlap_s = int(d.get("overlap_seconds", 0) or 0)
                    denom_toi = int(d.get("player_toi_seconds", 0) or 0)
                    gf = int(d.get("gf_together", 0) or 0)
                    ga = int(d.get("ga_together", 0) or 0)
                    shift_games = int(d.get("shift_games", 0) or 0)
                    pct = 100.0 * overlap_s / denom_toi if denom_toi > 0 else 0.0
                    rows_disp.append(
                        (
                            _display_player_name(teammate),
                            float(pct),
                            int(shift_games),
                            int(gf) - int(ga),
                            int(gf),
                            int(ga),
                            int(overlap_s),
                        )
                    )
                rows_disp.sort(key=lambda x: x[1], reverse=True)
                name_w = max([len(r[0]) for r in rows_disp] + [len("Teammate")])
                lines.append("")
                lines.append("On-ice with teammates (by TOI%, shift games only):")
                if include_shifts_in_stats:
                    lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA  Overlap")
                    for teammate_disp, pct, shift_games, pm, gf, ga, overlap_s in rows_disp:
                        lines.append(
                            f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}  {_format_duration(overlap_s):>7}"
                        )
                else:
                    # Without --shifts, do not publish any absolute time values.
                    lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA")
                    for teammate_disp, pct, shift_games, pm, gf, ga, _overlap_s in rows_disp:
                        lines.append(
                            f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}"
                        )

        # Goals
        lines.append("")
        lines.append("Goals detail:")
        if not goals_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(goals_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)):
                lines.append(f"  {game_label}: Period {ev.period}, {ev.t_str}{_fmt_tags(ev)}")

        # Assists
        lines.append("")
        lines.append("Assists detail:")
        if not assists_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(
                assists_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)
            ):
                lines.append(f"  {game_label}: Period {ev.period}, {ev.t_str}{_fmt_tags(ev)}")

        # Goals for / against while on ice
        lines.append("")
        lines.append("Goals for while on ice:")
        if not gf_on_ice_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(
                gf_on_ice_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)
            ):
                lines.append(f"  {game_label}: Period {ev.period}, {ev.t_str}{_fmt_tags(ev)}")

        lines.append("")
        lines.append("Goals against while on ice:")
        if not ga_list:
            lines.append("  (none)")
        else:
            for game_label, ev in sorted(ga_list, key=lambda x: (x[0], x[1].period, x[1].t_sec)):
                lines.append(f"  {game_label}: Period {ev.period}, {ev.t_str}")

        (outdir / f"{player}_cumulative_stats.txt").write_text(
            "\n".join(lines) + "\n", encoding="utf-8"
        )


def _write_season_highlight_scripts(
    base_outdir: Path,
    results: List[Dict[str, Any]],
    *,
    create_scripts: bool,
) -> None:
    """
    For multi-game runs, write no-arg per-player season highlight scripts that:
      - clip per-game highlights from `events_Highlights_<player>_video_times.txt`
      - use each game's `tracking_output-with-audio*.mp4` automatically
      - join per-game highlight clips in game order
    """
    if not create_scripts:
        return
    if not results:
        return

    season_dir = base_outdir / "season_highlights"
    season_dir.mkdir(parents=True, exist_ok=True)

    prefix = "events_Highlights_"
    suffix = "_video_times.txt"

    def _has_timestamps(path: Path) -> bool:
        try:
            with path.open("r", encoding="utf-8") as f:
                for line in f:
                    s = line.strip()
                    if not s or s.startswith("#"):
                        continue
                    return True
        except Exception:
            return False
        return False

    players: set[str] = set()
    for r in results:
        outdir = r.get("outdir")
        if not outdir:
            continue
        try:
            for p in Path(outdir).glob(f"{prefix}*{suffix}"):
                name = p.name
                if not name.startswith(prefix) or not name.endswith(suffix):
                    continue
                pk = name[len(prefix) : -len(suffix)]
                if pk:
                    players.add(pk)
        except Exception:
            continue

    if not players:
        return

    written_scripts: List[str] = []
    for player_key in sorted(players):
        # Determine which games this player has highlight events for, in game order.
        game_entries: List[Tuple[str, Path, Path, str]] = []
        missing_videos: List[Tuple[str, str]] = []
        for r in results:
            sheet_path = r.get("sheet_path")
            if sheet_path is None:
                # T2S-only games don't have a reliable scoreboard->video mapping.
                continue
            game_label = str(r.get("label") or "")
            outdir = Path(r.get("outdir"))
            ts_file = outdir / f"{prefix}{player_key}{suffix}"
            if not ts_file.exists() or not _has_timestamps(ts_file):
                continue

            video_path = r.get("video_path")
            video = (
                Path(video_path)
                if video_path is not None
                else _find_tracking_output_video_for_sheet_path(Path(sheet_path))
            )
            if video is None or not video.exists():
                missing_videos.append(
                    (game_label, str(video) if video is not None else "<missing>")
                )
                continue

            game_entries.append((game_label, video, ts_file, sanitize_name(game_label)))

        if missing_videos:
            miss_str = ", ".join(f"{g} -> {vp}" for g, vp in missing_videos)
            print(
                f"[season-highlights] WARNING: Missing video(s) for {_display_player_name(player_key)}: {miss_str}",
                file=sys.stderr,
            )

        if not game_entries:
            continue

        script_path = season_dir / f"clip_season_highlights_{player_key}.sh"

        script_lines: List[str] = [
            "#!/usr/bin/env bash",
            "set -euo pipefail",
            "",
            "# Optional flags:",
            "#   --quick / -q   lower quality, faster",
            "#   --hq           lossless intermediates (requires NVENC)",
            "",
            "QUICK=0",
            "HQ=0",
            'for ARG in "$@"; do',
            '  if [ "$ARG" = "--quick" ] || [ "$ARG" = "-q" ]; then',
            "    QUICK=1",
            '  elif [ "$ARG" = "--hq" ]; then',
            "    HQ=1",
            "  fi",
            "done",
            "",
            "EXTRA_FLAGS=()",
            'if [ "$QUICK" -gt 0 ]; then',
            '  EXTRA_FLAGS+=("--quick" "1")',
            "fi",
            'if [ "$HQ" -gt 0 ]; then',
            "  export VIDEO_CLIPPER_HQ=1",
            "fi",
            "",
            'THIS_DIR="$(cd "$(dirname "${BASH_SOURCE[0]}")" && pwd)"',
            f'OUT_DIR="$THIS_DIR/{player_key}"',
            'mkdir -p "$OUT_DIR"',
            "",
            "GAME_CLIPS=()",
            "",
        ]

        for game_label, video, ts_file, game_safe in game_entries:
            try:
                video_abs = str(Path(video).resolve())
            except Exception:
                video_abs = str(video)
            try:
                ts_abs = str(Path(ts_file).resolve())
            except Exception:
                ts_abs = str(ts_file)
            label_safe = sanitize_name(f"{player_key}__{game_label}__highlights")
            temp_dir = f"$THIS_DIR/temp_clips/{player_key}/{game_safe}"
            script_lines.extend(
                [
                    f'echo "[{_display_player_name(player_key)}] {game_label}"',
                    f'VIDEO="{video_abs}"',
                    f'TS_FILE="{ts_abs}"',
                    f'TEMP_DIR="{temp_dir}"',
                    'mkdir -p "$TEMP_DIR"',
                    "(",
                    '  cd "$OUT_DIR"',
                    '  python -m hmlib.cli.video_clipper -j 4 --input "$VIDEO" --timestamps "$TS_FILE" --temp-dir "$TEMP_DIR" '
                    f'"{label_safe}" "${{EXTRA_FLAGS[@]}}"',
                    ")",
                    f'GAME_CLIPS+=("$OUT_DIR/clips-{label_safe}.mp4")',
                    "",
                ]
            )

        list_file = "$OUT_DIR/season_clips.txt"
        season_label_safe = sanitize_name(f"{player_key}__season_highlights")

        script_lines.extend(
            [
                f'LIST_FILE="{list_file}"',
                ': > "$LIST_FILE"',
                'for f in "${GAME_CLIPS[@]}"; do',
                '  echo "file \'$f\'" >> "$LIST_FILE"',
                "done",
                "",
                f'OUT_FILE="$OUT_DIR/clips-{season_label_safe}.mp4"',
                'ffmpeg -f concat -safe 0 -i "$LIST_FILE" -c copy "$OUT_FILE"',
                "",
                'echo "Wrote: $OUT_FILE"',
                "",
            ]
        )

        script_path.write_text("\n".join(script_lines) + "\n", encoding="utf-8")
        try:
            os.chmod(script_path, 0o755)
        except Exception:
            pass

        written_scripts.append(script_path.name)

    if written_scripts:
        runner = season_dir / "clip_season_highlights_all.sh"
        runner_body = """#!/usr/bin/env bash
set -euo pipefail
THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"
for s in \"$THIS_DIR\"/clip_season_highlights_*.sh; do
  [ -x \"$s\" ] || continue
  if [ \"$s\" = \"$THIS_DIR/clip_season_highlights_all.sh\" ]; then
    continue
  fi
  echo \"Running $s...\"
  \"$s\" \"$@\"
done
"""
        runner.write_text(runner_body, encoding="utf-8")
        try:
            os.chmod(runner, 0o755)
        except Exception:
            pass


def _infer_side_from_rosters(
    t2s_id: int,
    jersey_numbers: set[str],
    hockey_db_dir: Path,
    *,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
    debug: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        import_err = _t2s_api_import_error
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": False,
                    "t2s_api_import_error": import_err,
                    "jerseys_in_sheet_count": len(jersey_numbers or set()),
                    "jerseys_in_sheet": sorted(list(jersey_numbers or set()), key=lambda x: int(x)),
                    "failure": "t2s_api_not_available",
                }
            )
        details = f": {import_err}" if import_err else ""
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: TimeToScore API not available "
            f"(failed to import hmlib.time2score.api){details}.",
            file=sys.stderr,
        )
        return None
    try:
        with _working_directory(hockey_db_dir):
            if allow_remote and not allow_full_sync:
                _log_t2s_scrape(int(t2s_id), "game details (roster/side inference)")
            try:
                info = t2s_api.get_game_details(
                    int(t2s_id),
                    sync_if_missing=(bool(allow_full_sync) if allow_remote else False),
                    fetch_stats_if_missing=bool(allow_remote),
                )
            except TypeError:
                info = t2s_api.get_game_details(int(t2s_id))
    except Exception as e:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "jerseys_in_sheet_count": len(jersey_numbers or set()),
                    "jerseys_in_sheet": sorted(list(jersey_numbers or set()), key=lambda x: int(x)),
                    "failure": "t2s_fetch_failed",
                    "exception": str(e),
                }
            )
        print(f"[t2s] Failed to load game {t2s_id} for side inference: {e}", file=sys.stderr)
        return None
    stats = (info or {}).get("stats") or {}
    home_players = stats.get("homePlayers") or []
    away_players = stats.get("awayPlayers") or []

    def _nums(rows: Any) -> set[str]:
        out: set[str] = set()
        for r in rows or []:
            num = _normalize_jersey_number((r or {}).get("number"))
            if num:
                out.add(num)
        return out

    home_set = _nums(home_players)
    away_set = _nums(away_players)

    if not jersey_numbers:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "home_roster_count": len(home_set),
                    "away_roster_count": len(away_set),
                    "failure": "no_jersey_numbers_in_sheet",
                }
            )
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: no jersey numbers found in sheet.",
            file=sys.stderr,
        )
        return None
    if not home_set and not away_set:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "jerseys_in_sheet_count": len(jersey_numbers),
                    "jerseys_in_sheet": sorted(list(jersey_numbers), key=lambda x: int(x)),
                    "home_roster_count": 0,
                    "away_roster_count": 0,
                    "failure": "no_roster_numbers_in_t2s",
                }
            )
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: no roster numbers in TimeToScore stats.",
            file=sys.stderr,
        )
        return None

    home_overlap = len(jersey_numbers & home_set)
    away_overlap = len(jersey_numbers & away_set)

    if home_overlap == away_overlap:
        if debug is not None:
            debug.clear()
            debug.update(
                {
                    "t2s_id": int(t2s_id),
                    "t2s_api_available": True,
                    "jerseys_in_sheet_count": len(jersey_numbers),
                    "jerseys_in_sheet": sorted(list(jersey_numbers), key=lambda x: int(x)),
                    "home_roster_count": len(home_set),
                    "away_roster_count": len(away_set),
                    "home_overlap": home_overlap,
                    "away_overlap": away_overlap,
                    "home_overlap_jerseys": sorted(
                        list(jersey_numbers & home_set), key=lambda x: int(x)
                    ),
                    "away_overlap_jerseys": sorted(
                        list(jersey_numbers & away_set), key=lambda x: int(x)
                    ),
                    "failure": "overlap_tie",
                }
            )
        print(
            f"[t2s] Cannot infer side for game {t2s_id}: overlap tie (home={home_overlap}, away={away_overlap}).",
            file=sys.stderr,
        )
        return None
    side = "home" if home_overlap > away_overlap else "away"
    if debug is not None:
        debug.clear()
        debug.update(
            {
                "t2s_id": int(t2s_id),
                "t2s_api_available": True,
                "jerseys_in_sheet_count": len(jersey_numbers),
                "jerseys_in_sheet": sorted(list(jersey_numbers), key=lambda x: int(x)),
                "home_roster_count": len(home_set),
                "away_roster_count": len(away_set),
                "home_overlap": home_overlap,
                "away_overlap": away_overlap,
                "chosen_side": side,
            }
        )
    return side


def _get_t2s_team_roster(
    t2s_id: int,
    side: str,
    hockey_db_dir: Path,
    *,
    keep_goalies: bool = True,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
) -> Dict[str, str]:
    """
    Return a mapping of normalized jersey number -> player name for the given
    TimeToScore game id and team side ('home' or 'away').

    This is used to credit Games Played (GP) for players who appear on the
    official game roster even if they have no recorded shifts in the sheet.
    """
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        details = f": {_t2s_api_import_error}" if _t2s_api_import_error else ""
        raise RuntimeError(
            f"TimeToScore API not available (failed to import hmlib.time2score.api){details}"
        )
    try:
        with _working_directory(hockey_db_dir):
            if allow_remote and not allow_full_sync:
                _log_t2s_scrape(int(t2s_id), "game details (team roster)")
            try:
                info = t2s_api.get_game_details(
                    int(t2s_id),
                    sync_if_missing=(bool(allow_full_sync) if allow_remote else False),
                    fetch_stats_if_missing=bool(allow_remote),
                )
            except TypeError:
                info = t2s_api.get_game_details(int(t2s_id))
    except Exception as e:
        raise RuntimeError(f"TimeToScore API failed to load game {t2s_id} for roster: {e}") from e

    stats = (info or {}).get("stats")
    if not isinstance(stats, dict) or not stats:
        suffix = (
            " (cache-only; run import_time2score or omit --t2s-cache-only)"
            if not allow_remote
            else ""
        )
        raise RuntimeError(
            f"TimeToScore returned no usable stats for game {t2s_id}; cannot load roster.{suffix}"
        )
    players_key = "homePlayers" if side == "home" else "awayPlayers"
    rows = stats.get(players_key) or []

    roster: Dict[str, str] = {}

    for r in rows:
        pos = str((r or {}).get("position") or "").strip()
        if not keep_goalies and pos and pos.upper().startswith("G"):
            continue
        try:
            raw_num = str((r or {}).get("number")).strip()
        except Exception:
            continue
        if not raw_num:
            continue
        num_norm = _normalize_jersey_number(raw_num)
        if not num_norm:
            continue
        name = str((r or {}).get("name") or "").strip()
        if not name:
            continue
        # One name per jersey; later entries with the same jersey overwrite.
        roster[num_norm] = name

    return roster


def _get_t2s_game_rosters(
    t2s_id: int,
    hockey_db_dir: Path,
    *,
    keep_goalies: bool = True,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
) -> Dict[str, Dict[str, str]]:
    """
    Return both rosters (home/away) as normalized jersey number -> player name mappings.

    This is used to resolve placeholder spreadsheet player labels like 'Blue_12' / 'White_7'
    to real player names when event logs include opponent jerseys.
    """
    t2s_api = _get_t2s_api()
    if t2s_api is None:
        details = f": {_t2s_api_import_error}" if _t2s_api_import_error else ""
        raise RuntimeError(
            f"TimeToScore API not available (failed to import hmlib.time2score.api){details}"
        )
    try:
        with _working_directory(hockey_db_dir):
            if allow_remote and not allow_full_sync:
                _log_t2s_scrape(int(t2s_id), "game details (home/away rosters)")
            try:
                info = t2s_api.get_game_details(
                    int(t2s_id),
                    sync_if_missing=(bool(allow_full_sync) if allow_remote else False),
                    fetch_stats_if_missing=bool(allow_remote),
                )
            except TypeError:
                info = t2s_api.get_game_details(int(t2s_id))
    except Exception as e:
        raise RuntimeError(f"TimeToScore API failed to load game {t2s_id} for rosters: {e}") from e

    stats = (info or {}).get("stats")
    if not isinstance(stats, dict) or not stats:
        suffix = (
            " (cache-only; run import_time2score or omit --t2s-cache-only)"
            if not allow_remote
            else ""
        )
        raise RuntimeError(
            f"TimeToScore returned no usable stats for game {t2s_id}; cannot load rosters.{suffix}"
        )

    out: Dict[str, Dict[str, str]] = {"home": {}, "away": {}}
    for side, key in (("home", "homePlayers"), ("away", "awayPlayers")):
        rows = stats.get(key) or []
        roster: Dict[str, str] = {}
        for r in rows:
            pos = str((r or {}).get("position") or "").strip()
            if not keep_goalies and pos and pos.upper().startswith("G"):
                continue
            raw_num = (r or {}).get("number")
            num_norm = _normalize_jersey_number(raw_num)
            if not num_norm:
                continue
            name = str((r or {}).get("name") or "").strip()
            if not name:
                continue
            roster[num_norm] = name
        out[side] = roster
    return out


def _write_event_summaries_and_clips(
    outdir: Path,
    stats_dir: Path,
    event_log_context: EventLogContext,
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    create_scripts: bool,
    *,
    focus_team: Optional[str] = None,
) -> None:
    raw_evt_by_team = dict(event_log_context.event_counts_by_type_team or {})
    raw_instances = event_log_context.event_instances or {}
    invert_event_type: set[str] = set()
    if focus_team in {"Blue", "White"}:
        invert_event_type.add("Giveaway")
        turnover_teams = {
            str(team) for (etype, team) in raw_evt_by_team.keys() if str(etype) == "TurnoverForced"
        }
        if not turnover_teams:
            turnover_teams = {
                str(team)
                for (etype, team) in raw_instances.keys()
                if str(etype) == "TurnoverForced"
            }
        turnover_teams = {t for t in turnover_teams if t in {"Blue", "White"}}
        if len(turnover_teams) > 1:
            invert_event_type.add("TurnoverForced")

    # If the sheet only collected generic Shots (no explicit SOG), mirror those counts to SOG so
    # downstream summaries aren't empty.
    if not any(str(et) == "SOG" for et, _ in raw_evt_by_team.keys()):
        for (et, tm), cnt in list(raw_evt_by_team.items()):
            if str(et) == "Shot":
                raw_evt_by_team[("SOG", tm)] = cnt

    def _no_parens_label(s: str) -> str:
        return re.sub(r"[()]", "", str(s or "")).strip()

    def _team_label(team: Any, *, event_type: Optional[str] = None) -> str:
        team_str = str(team) if team is not None else ""
        if focus_team in {"Blue", "White"} and team_str in {"Blue", "White"}:
            # Some event types are recorded for the team that loses possession, but
            # "For/Against" is intended to be relative to the focus team.
            et = str(event_type or "")
            invert = et in invert_event_type
            if invert:
                return "Against" if team_str == focus_team else "For"
            return "For" if team_str == focus_team else "Against"
        return team_str or "Unknown"

    evt_by_team: Dict[Tuple[str, str], int] = {}
    for (et, tm), cnt in sorted(raw_evt_by_team.items()):
        label = _team_label(tm, event_type=str(et))
        try:
            inc = int(cnt)
        except Exception:
            inc = 0
        evt_by_team[(et, label)] = evt_by_team.get((et, label), 0) + inc

    rows_evt = [
        {"event_type": _display_event_type(et), "team": tm, "count": cnt}
        for (et, tm), cnt in sorted(evt_by_team.items())
    ]
    if rows_evt:
        df_evt = pd.DataFrame(rows_evt)
        df_evt.to_csv(stats_dir / "event_summary.csv", index=False)
        _write_styled_xlsx_table(
            stats_dir / "event_summary.xlsx",
            df_evt,
            sheet_name="event_summary",
            title="Event Summary",
        )

    player_event_rows = event_log_context.event_player_rows or []
    if player_event_rows:

        def _fmt_v(x):
            return (
                seconds_to_hhmmss(int(x))
                if isinstance(x, int)
                else (seconds_to_hhmmss(int(x)) if isinstance(x, float) else "")
            )

        def _fmt_g(x):
            return (
                seconds_to_mmss_or_hhmmss(int(x))
                if isinstance(x, int)
                else (seconds_to_mmss_or_hhmmss(int(x)) if isinstance(x, float) else "")
            )

        rows = []
        for r in player_event_rows:
            pk_raw = r.get("player")
            pk_parts = _parse_player_key(pk_raw)
            jersey_val = r.get("jersey")
            if (jersey_val is None or str(jersey_val).strip() == "") and pk_parts.jersey:
                jersey_val = pk_parts.jersey
            rows.append(
                {
                    "event_type": _display_event_type(str(r.get("event_type") or "")),
                    "jersey": jersey_val,
                    "player": pk_parts.name,
                    "period": r.get("period"),
                    "video_time": _fmt_v(r.get("video_s")),
                    "game_time": _fmt_g(r.get("game_s")),
                    "team": _team_label(r.get("team"), event_type=str(r.get("event_type") or "")),
                }
            )
        df_players = pd.DataFrame(rows)
        df_players.to_csv(stats_dir / "event_players.csv", index=False)
        _write_styled_xlsx_table(
            stats_dir / "event_players.xlsx",
            df_players,
            sheet_name="event_players",
            title="Event Players",
        )

    # When --no-scripts is set, we still keep event summary CSVs but skip all
    # clip-related timestamp files and helper scripts.
    if not create_scripts:
        return

    instances: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for (etype, team), lst in raw_instances.items():
        label = _team_label(team, event_type=str(etype))
        instances.setdefault((etype, label), []).extend(list(lst or []))

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    max_sb_by_period: Dict[int, int] = {}
    for period, segs in conv_segments_by_period.items():
        mx = 0
        for s1, s2, _, _ in segs:
            mx = max(mx, s1, s2)
        if mx > 0:
            max_sb_by_period[period] = mx
    for _, lst in instances.items():
        for it in lst:
            p = it.get("period")
            gs = it.get("game_s")
            if isinstance(p, int) and isinstance(gs, (int, float)):
                v = int(gs)
                max_sb_by_period[p] = max(max_sb_by_period.get(p, 0), v)

    def merge_windows(win: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        if not win:
            return []
        win = sorted(win)
        out = [list(win[0])]
        for a, b in win[1:]:
            la, lb = out[-1]
            if a <= lb + 10:
                out[-1][1] = max(lb, b)
            else:
                out.append([a, b])
        return [(a, b) for a, b in out]

    def merge_windows_with_centers(
        win: List[Tuple[int, int, int]],
    ) -> List[Tuple[int, int, List[int]]]:
        """
        Merge overlapping/nearby (<=10s gap) video windows, preserving the set of
        event-center timestamps inside each merged window.
        """
        if not win:
            return []
        win_sorted = sorted(win, key=lambda x: (x[0], x[1], x[2]))
        out: List[Tuple[int, int, List[int]]] = [
            (win_sorted[0][0], win_sorted[0][1], [win_sorted[0][2]])
        ]
        for a, b, c in win_sorted[1:]:
            la, lb, centers = out[-1]
            if a <= lb + 10:
                out[-1] = (la, max(lb, b), centers + [c])
            else:
                out.append((a, b, [c]))

        merged: List[Tuple[int, int, List[int]]] = []
        for a, b, centers in out:
            seen: set[int] = set()
            centers_sorted: List[int] = []
            for x in sorted(int(v) for v in centers):
                if x in seen:
                    continue
                seen.add(x)
                centers_sorted.append(x)
            merged.append((int(a), int(b), centers_sorted))
        return merged

    clip_scripts: List[str] = []
    for (etype, team), lst in sorted(instances.items()):
        # Skip team-level assist clip scripts; assists are handled in per-player highlights.
        if str(etype) == "Assist":
            continue

        v_windows: List[Tuple[int, int, int]] = []
        sb_windows_by_period: Dict[int, List[Tuple[int, int]]] = {}
        pre_s, post_s = _clip_pre_post_s_for_event_type(str(etype))
        etype_disp = _display_event_type(str(etype))
        etype_safe = sanitize_name(etype_disp) or "Event"
        team_safe = sanitize_name(team) or "Team"
        for it in lst:
            p = it.get("period")
            v = it.get("video_s")
            g = it.get("game_s")
            vsec = None
            if isinstance(v, (int, float)):
                vsec = int(v)
            elif isinstance(g, (int, float)) and isinstance(p, int):
                vsec = map_sb_to_video(int(p), int(g))
            if vsec is not None:
                start = max(0, vsec - int(pre_s))
                end = vsec + int(post_s)
                v_windows.append((start, end, int(vsec)))
            if isinstance(g, (int, float)) and isinstance(p, int):
                gsec = int(g)
                sb_max = max_sb_by_period.get(int(p), None)
                sb_start = gsec + int(pre_s)
                if sb_max is not None:
                    sb_start = min(sb_max, sb_start)
                sb_end = max(0, gsec - int(post_s))
                lo, hi = (sb_end, sb_start) if sb_end <= sb_start else (sb_start, sb_end)
                sb_windows_by_period.setdefault(int(p), []).append((lo, hi))

        v_windows_merged = merge_windows_with_centers(v_windows)
        if v_windows_merged:
            vfile = outdir / f"events_{etype_safe}_{team_safe}_video_times.txt"
            v_lines = []
            for a, b, centers in v_windows_merged:
                tokens = [seconds_to_hhmmss(a), seconds_to_hhmmss(b)]
                tokens.extend(seconds_to_hhmmss(int(c)) for c in (centers or []))
                v_lines.append(" ".join(tokens))
            vfile.write_text("\n".join(v_lines) + "\n", encoding="utf-8")
            if create_scripts:
                script = outdir / f"clip_events_{etype_safe}_{team_safe}.sh"
                label = f"{_no_parens_label(etype_disp)} {_no_parens_label(team)}"
                blink_label = _blink_event_label(str(etype))
                body = f"""#!/usr/bin/env bash
        set -euo pipefail
        if [ $# -lt 2 ]; then
          echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
    THIS_DIR=\"$(cd \"$(dirname \"${{BASH_SOURCE[0]}}\")\" && pwd)\"
    TS_FILE=\"$THIS_DIR/{vfile.name}\"
    shift 2 || true
    python -m hmlib.cli.video_clipper -j 4 --blink-event-text --blink-event-label \"{blink_label}\" --input \"$INPUT\" --timestamps \"$TS_FILE\" --temp-dir \"$THIS_DIR/temp_clips/{etype_safe}_{team_safe}\" \"{label} vs $OPP\" \"$@\"
    """
                script.write_text(body, encoding="utf-8")
                try:
                    import os as _os

                    _os.chmod(script, 0o755)
                except Exception:
                    pass
                clip_scripts.append(script.name)

        if sb_windows_by_period:
            sfile = outdir / f"events_{etype_safe}_{team_safe}_scoreboard_times.txt"
            s_lines = []
            for p, wins in sorted(sb_windows_by_period.items()):
                wins = merge_windows(wins)
                for lo, hi in wins:
                    s_lines.append(
                        f"{p} {seconds_to_mmss_or_hhmmss(hi)} {seconds_to_mmss_or_hhmmss(lo)}"
                    )
            if s_lines:
                sfile.write_text("\n".join(s_lines) + "\n", encoding="utf-8")

    if clip_scripts and create_scripts:
        all_script = outdir / "clip_events_all.sh"
        all_body = """#!/usr/bin/env bash
set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
shift 2 || true
THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"
for s in {scripts}; do
  echo \"Running $s...\"
  \"$THIS_DIR/$s\" \"$INPUT\" \"$OPP\" \"$@\"
done
""".replace(
            "{scripts}", " ".join(sorted(clip_scripts))
        )
        all_script.write_text(all_body, encoding="utf-8")
        try:
            import os as _os

            _os.chmod(all_script, 0o755)
        except Exception:
            pass

    team_excluded = event_log_context.team_excluded or {}
    if any(v for v in team_excluded.values()):
        import sys as _sys

        for team, excl in team_excluded.items():
            if not excl:
                continue
            excl_str = ", ".join(str(x) for x in excl[:20])
            print(
                f"[warning] Team {team} exceeded 20 unique jerseys; additional jerseys seen: {excl_str} (data kept)",
                file=_sys.stderr,
            )


def _write_player_event_highlights(
    outdir: Path,
    event_log_context: EventLogContext,
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    player_keys: Iterable[str],
    create_scripts: bool,
    *,
    highlight_types: Tuple[str, ...] = ("Goal", "SOG"),
) -> None:
    """
    Generate per-player highlight timestamp files + helper scripts for selected event types.

    Intended for '-long' sheets where events include video time. Falls back to
    mapping scoreboard->video using conv_segments when needed.
    """
    if not create_scripts:
        return
    if not event_log_context or not (event_log_context.event_player_rows or []):
        return

    player_set = set(player_keys or [])
    if not player_set:
        return

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    def merge_windows(win: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        if not win:
            return []
        win_sorted = sorted(win)
        out: List[Tuple[int, int]] = [win_sorted[0]]
        for a, b in win_sorted[1:]:
            la, lb = out[-1]
            if a <= lb:
                out[-1] = (la, max(lb, b))
            else:
                out.append((a, b))
        return out

    def merge_windows_with_centers(
        win: List[Tuple[int, int, int]],
    ) -> List[Tuple[int, int, List[int]]]:
        if not win:
            return []
        win_sorted = sorted(win, key=lambda x: (x[0], x[1], x[2]))
        out: List[Tuple[int, int, List[int]]] = [
            (win_sorted[0][0], win_sorted[0][1], [win_sorted[0][2]])
        ]
        for a, b, c in win_sorted[1:]:
            la, lb, centers = out[-1]
            if a <= lb:
                out[-1] = (la, max(lb, b), centers + [c])
            else:
                out.append((a, b, [c]))

        merged: List[Tuple[int, int, List[int]]] = []
        for a, b, centers in out:
            seen: set[int] = set()
            centers_sorted: List[int] = []
            for x in sorted(int(v) for v in centers):
                if x in seen:
                    continue
                seen.add(x)
                centers_sorted.append(x)
            merged.append((int(a), int(b), centers_sorted))
        return merged

    by_player_type: Dict[Tuple[str, str], List[Tuple[int, Optional[int], Optional[int]]]] = {}
    for row in event_log_context.event_player_rows or []:
        et = row.get("event_type")
        pk = row.get("player")
        if not isinstance(et, str) or et not in highlight_types:
            continue
        if not isinstance(pk, str) or pk not in player_set:
            continue
        p = row.get("period")
        if not isinstance(p, int):
            continue
        v = row.get("video_s")
        g = row.get("game_s")
        vsec = int(v) if isinstance(v, (int, float)) else None
        gsec = int(g) if isinstance(g, (int, float)) else None
        by_player_type.setdefault((pk, et), []).append((p, vsec, gsec))

    for (pk, etype), rows in sorted(by_player_type.items(), key=lambda x: (x[0][0], x[0][1])):
        v_windows: List[Tuple[int, int, int]] = []
        pre_s, post_s = _clip_pre_post_s_for_event_type(str(etype))
        for period, vsec, gsec in rows:
            vv = vsec
            if vv is None and gsec is not None:
                vv = map_sb_to_video(period, gsec)
            if vv is None:
                continue
            v_windows.append((max(0, int(vv) - int(pre_s)), int(vv) + int(post_s), int(vv)))

        v_windows_merged = merge_windows_with_centers(v_windows)
        if not v_windows_merged:
            continue

        vfile = outdir / f"events_{etype}_{pk}_video_times.txt"
        v_lines = []
        for a, b, centers in v_windows_merged:
            tokens = [seconds_to_hhmmss(a), seconds_to_hhmmss(b)]
            tokens.extend(seconds_to_hhmmss(int(c)) for c in (centers or []))
            v_lines.append(" ".join(tokens))
        vfile.write_text("\n".join(v_lines) + "\n", encoding="utf-8")

        if not create_scripts:
            continue

        label = f"{_display_event_type(etype)} - {_display_player_name(pk)}"
        blink_label = _blink_event_label(str(etype))
        script_name = f"clip_{etype.lower()}_{pk}.sh"
        script = outdir / script_name
        body = f"""#!/usr/bin/env bash
    set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
THIS_DIR=\"$(cd \"$(dirname \"${{BASH_SOURCE[0]}}\")\" && pwd)\"
TS_FILE=\"$THIS_DIR/{vfile.name}\"
shift 2 || true
python -m hmlib.cli.video_clipper -j 4 --blink-event-text --blink-event-label \"{blink_label}\" --input \"$INPUT\" --timestamps \"$TS_FILE\" --temp-dir \"$THIS_DIR/temp_clips/{etype}_{pk}\" \"{label} vs $OPP\" \"$@\"
"""
        script.write_text(body, encoding="utf-8")
        try:
            import os as _os

            _os.chmod(script, 0o755)
        except Exception:
            pass


def _write_player_combined_highlights(
    outdir: Path,
    *,
    event_log_context: Optional[EventLogContext],
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]],
    player_keys: Iterable[str],
    create_scripts: bool,
    highlight_types: Tuple[str, ...] = ("Goal", "Assist", "ExpectedGoal", "Takeaway"),
) -> None:
    """
    Generate a per-player "Highlights" timestamp file that mixes multiple event types
    in chronological order (within the game).

    This is used for season (multi-game) highlight reels so a player's events are not
    grouped by type (e.g., all goals then all assists).
    """
    if not create_scripts:
        return
    player_set = set(player_keys or [])
    if not player_set:
        return

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    def merge_windows(win: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
        if not win:
            return []
        win_sorted = sorted(win)
        out: List[Tuple[int, int]] = [win_sorted[0]]
        for a, b in win_sorted[1:]:
            la, lb = out[-1]
            if a <= lb:
                out[-1] = (la, max(lb, b))
            else:
                out.append((a, b))
        return out

    highlight_types_set = set(highlight_types or ())

    # First, seed goal timestamps (period, game_s) from the scoring source so we can
    # dedupe xG rows that are just goals (goals count as xG internally).
    goal_keys_by_player: Dict[str, set[Tuple[int, int]]] = {}
    for pk, info in (per_player_goal_events or {}).items():
        if pk not in player_set:
            continue
        for ev in (info or {}).get("goals", []) or []:
            try:
                goal_keys_by_player.setdefault(pk, set()).add((int(ev.period), int(ev.t_sec)))
            except Exception:
                continue

    # Collect raw event centers (video seconds) per player.
    # We dedupe by (event_type, period, game_s) when possible.
    events_by_player: Dict[str, List[Tuple[int, str, int, Optional[int]]]] = {}
    seen_by_player: Dict[str, set[Tuple[str, int, int]]] = {}

    def _register_event(
        *,
        player: str,
        event_type: str,
        period: int,
        video_s: Optional[int],
        game_s: Optional[int],
    ) -> None:
        if player not in player_set:
            return
        et = str(event_type or "")
        if et not in highlight_types_set:
            return
        try:
            p = int(period)
        except Exception:
            return

        # Dedupe ExpectedGoal rows that correspond to goals for the same player/time.
        if et == "ExpectedGoal":
            if game_s is not None and (p, int(game_s)) in goal_keys_by_player.get(player, set()):
                return

        vsec = (
            int(video_s)
            if isinstance(video_s, int)
            else (int(video_s) if isinstance(video_s, float) else None)
        )
        gsec = (
            int(game_s)
            if isinstance(game_s, int)
            else (int(game_s) if isinstance(game_s, float) else None)
        )

        if vsec is None and gsec is not None:
            vsec = map_sb_to_video(p, int(gsec))
        if vsec is None:
            return

        # Use game time for dedupe when we have it; otherwise fall back to video time.
        key_time = gsec if gsec is not None else int(vsec)
        seen = seen_by_player.setdefault(player, set())
        if (et, p, int(key_time)) in seen:
            return
        seen.add((et, p, int(key_time)))
        events_by_player.setdefault(player, []).append((int(vsec), et, p, gsec))

        if et == "Goal" and gsec is not None:
            goal_keys_by_player.setdefault(player, set()).add((p, int(gsec)))

    # Prefer long-sheet / event-log times when available.
    if event_log_context is not None:
        for row in event_log_context.event_player_rows or []:
            try:
                pk = row.get("player")
                etype = row.get("event_type")
                period = row.get("period")
                if (
                    not isinstance(pk, str)
                    or not isinstance(etype, str)
                    or not isinstance(period, int)
                ):
                    continue
                _register_event(
                    player=pk,
                    event_type=etype,
                    period=period,
                    video_s=row.get("video_s"),
                    game_s=row.get("game_s"),
                )
            except Exception:
                continue

    # Fallback for goals/assists when a game has no long sheet.
    for pk, info in (per_player_goal_events or {}).items():
        if pk not in player_set:
            continue
        for ev in (info or {}).get("goals", []) or []:
            _register_event(
                player=pk,
                event_type="Goal",
                period=int(getattr(ev, "period", 0) or 0),
                video_s=None,
                game_s=int(getattr(ev, "t_sec", 0) or 0),
            )
        for ev in (info or {}).get("assists", []) or []:
            _register_event(
                player=pk,
                event_type="Assist",
                period=int(getattr(ev, "period", 0) or 0),
                video_s=None,
                game_s=int(getattr(ev, "t_sec", 0) or 0),
            )

    for pk, evs in sorted(events_by_player.items(), key=lambda x: x[0]):
        if not evs:
            continue
        # Chronological within this game.
        evs_sorted = sorted(evs, key=lambda x: x[0])
        windows: List[Tuple[int, int]] = []
        for vsec, etype, _period, _gsec in evs_sorted:
            pre_s, post_s = _clip_pre_post_s_for_event_type(str(etype))
            windows.append((max(0, int(vsec) - int(pre_s)), int(vsec) + int(post_s)))

        windows = merge_windows(windows)
        if not windows:
            continue

        vfile = outdir / f"events_Highlights_{pk}_video_times.txt"
        v_lines = [f"{seconds_to_hhmmss(a)} {seconds_to_hhmmss(b)}" for a, b in windows]
        vfile.write_text("\n".join(v_lines) + "\n", encoding="utf-8")


def _write_goal_window_files(
    outdir: Path,
    goals: List[GoalEvent],
    conv_segments_by_period: Dict[int, List[Tuple[int, int, int, int]]],
) -> None:
    if not goals:
        return

    def map_sb_to_video(period: int, t_sb: int) -> Optional[int]:
        segs = conv_segments_by_period.get(period)
        if not segs:
            return None
        for s1, s2, v1, v2 in segs:
            lo, hi = (s1, s2) if s1 <= s2 else (s2, s1)
            if lo <= t_sb <= hi and s1 != s2:
                return int(round(v1 + (t_sb - s1) * (v2 - v1) / (s2 - s1)))
        return None

    max_sb_by_period: Dict[int, int] = {}
    for period, segs in conv_segments_by_period.items():
        mx = 0
        for s1, s2, _, _ in segs:
            mx = max(mx, s1, s2)
        max_sb_by_period[period] = mx

    gf_lines: List[str] = []
    ga_lines: List[str] = []
    for ev in goals:
        sb_max = max_sb_by_period.get(ev.period, None)
        start_sb = ev.t_sec + GOAL_CLIP_PRE_S
        end_sb = ev.t_sec - GOAL_CLIP_POST_S
        if sb_max is not None:
            start_sb = max(0, min(sb_max, start_sb))
            end_sb = max(0, min(sb_max, end_sb))
        else:
            start_sb = max(0, start_sb)
            end_sb = max(0, end_sb)

        v_center = map_sb_to_video(ev.period, ev.t_sec)
        if v_center is not None:
            v_start = max(0, v_center - GOAL_CLIP_PRE_S)
            v_end = v_center + GOAL_CLIP_POST_S
            start_str = seconds_to_hhmmss(v_start)
            end_str = seconds_to_hhmmss(v_end)
        else:
            v_start = map_sb_to_video(ev.period, start_sb)
            v_end = map_sb_to_video(ev.period, end_sb)
            if v_start is not None and v_end is not None:
                lo, hi = (
                    (int(v_start), int(v_end)) if v_start <= v_end else (int(v_end), int(v_start))
                )
                start_str = seconds_to_hhmmss(max(0, lo))
                end_str = seconds_to_hhmmss(max(0, hi))
            else:
                start_str = seconds_to_hhmmss(max(0, start_sb))
                end_str = seconds_to_hhmmss(max(0, end_sb))

        line = f"{start_str} {end_str}"
        if ev.kind == "GF":
            gf_lines.append(line)
        else:
            ga_lines.append(line)

    (outdir / "goals_for.txt").write_text(
        "\n".join(gf_lines) + ("\n" if gf_lines else ""), encoding="utf-8"
    )
    (outdir / "goals_against.txt").write_text(
        "\n".join(ga_lines) + ("\n" if ga_lines else ""), encoding="utf-8"
    )


def _write_clip_all_runner(outdir: Path, create_scripts: bool) -> None:
    if not create_scripts:
        return
    clip_all_path = outdir / "clip_all.sh"
    clip_all_body = """#!/usr/bin/env bash
set -euo pipefail
if [ $# -lt 2 ]; then
  echo "Usage: $0 <input_video> <opposing_team> [--quick|-q] [--hq]"
  exit 1
fi
INPUT=\"$1\"
OPP=\"$2\"
shift 2 || true
THIS_DIR=\"$(cd \"$(dirname \"${BASH_SOURCE[0]}\")\" && pwd)\"
for s in \"$THIS_DIR\"/clip_*.sh; do
  [ -x \"$s\" ] || continue
  base=\"$(basename \"$s\")\"
  if [ \"$base\" = \"clip_all.sh\" ] || [ \"$base\" = \"clip_events_all.sh\" ]; then
    continue
  fi
  echo \"Running $s...\"
  \"$s\" \"$INPUT\" \"$OPP\" \"$@\"
done
"""
    clip_all_path.write_text(clip_all_body, encoding="utf-8")
    try:
        import os as _os

        _os.chmod(clip_all_path, 0o755)
    except Exception:
        pass


def process_sheet(
    xls_path: Path,
    sheet_name: Optional[str],
    outdir: Path,
    keep_goalies: bool,
    goals: List[GoalEvent],
    roster_map: Optional[Dict[str, str]] = None,
    t2s_rosters_by_side: Optional[Dict[str, Dict[str, str]]] = None,
    t2s_side: Optional[str] = None,
    t2s_game_id: Optional[int] = None,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
    long_xls_paths: Optional[List[Path]] = None,
    focus_team_override: Optional[str] = None,
    include_shifts_in_stats: bool = False,
    write_events_summary: Optional[bool] = None,
    skip_validation: bool = False,
    create_scripts: bool = True,
    write_opponent_stats_from_long_shifts: bool = True,
    verbose: bool = False,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
    List[Dict[str, Any]],
]:
    target_sheet = 0 if sheet_name is None else sheet_name
    df = pd.read_excel(xls_path, sheet_name=target_sheet, header=None)
    outdir.mkdir(parents=True, exist_ok=True)

    # Try event-log layout first
    (
        used_event_log,
        video_pairs_by_player,
        sb_pairs_by_player,
        conv_segments_by_period,
        event_log_context,
    ) = _parse_event_log_layout(df)

    validation_errors = 0
    if not used_event_log:
        (
            video_pairs_by_player,
            sb_pairs_by_player,
            conv_segments_by_period,
            validation_errors,
        ) = _parse_per_player_layout(df, keep_goalies=keep_goalies, skip_validation=skip_validation)

    # Output subdir depends on format
    format_dir = "event_log" if used_event_log else "per_player"
    side = str(t2s_side or "").strip().lower()
    team_subdir = "Away" if side == "away" else "Home"

    # Top-level Home/Away split, then format (per_player/event_log).
    (outdir / "Home").mkdir(parents=True, exist_ok=True)
    (outdir / "Away").mkdir(parents=True, exist_ok=True)
    outdir = outdir / team_subdir / format_dir
    outdir.mkdir(parents=True, exist_ok=True)
    stats_dir = outdir / "stats"
    stats_dir.mkdir(parents=True, exist_ok=True)

    # Per-player time files and clip scripts
    if include_shifts_in_stats:
        _write_video_times_and_scripts(outdir, video_pairs_by_player, create_scripts=create_scripts)
        _write_scoreboard_times(outdir, sb_pairs_by_player, create_scripts=create_scripts)

    if write_events_summary is None:
        write_events_summary = include_shifts_in_stats

    stats_table_rows: List[Dict[str, str]] = []
    all_periods_seen: set[int] = set()

    roster_tables: List[Tuple[Optional[str], Dict[str, str]]] = []
    roster_tables.extend(_extract_roster_tables_from_df(df))

    jersey_to_players: Dict[str, List[str]] = {}
    for pk in sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
        if norm:
            jersey_to_players.setdefault(norm, []).append(pk)

    # Optionally add TimeToScore roster-only players (no shifts in this sheet)
    roster_only_players: List[str] = []
    if roster_map:
        # Jersey numbers already present in this sheet (normalized).
        seen_normals: set[str] = set()
        for pk in sb_pairs_by_player.keys():
            norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
            if norm:
                seen_normals.add(norm)

        for jersey_norm, name in roster_map.items():
            if not jersey_norm:
                continue
            if jersey_norm in seen_normals:
                # Already have at least one player with this jersey from shifts.
                continue
            jersey_label = jersey_norm
            player_key = f"{sanitize_name(jersey_label)}_{sanitize_name(name)}"
            if player_key in sb_pairs_by_player:
                continue
            # No recorded shifts; keep an explicit empty list.
            sb_pairs_by_player[player_key] = []
            roster_only_players.append(player_key)
            # Map normalized jersey to this player for scoring lookups.
            jersey_to_players.setdefault(jersey_norm, []).append(player_key)

    # Optional '-long' event sheet: merge event context and optionally infer goals.
    focus_team: Optional[str] = focus_team_override
    merged_event_context: Optional[EventLogContext] = event_log_context
    inferred_long_goals: List[GoalEvent] = []

    long_events_all: List[LongEvent] = []
    long_goal_rows_all: List[Dict[str, Any]] = []
    jerseys_by_team_all: Dict[str, set[int]] = {}
    spreadsheet_event_mapping_summary = SpreadsheetEventMappingSummary.empty()
    long_shift_tables_by_team: Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]] = {}
    long_sheet_paths_used: List[Path] = []
    shift_cmp_summary: Dict[str, Any] = {}

    if long_xls_paths:
        for long_path in long_xls_paths:
            if long_path is None:
                continue
            lp = Path(long_path).expanduser()
            if not lp.exists():
                continue
            long_sheet_paths_used.append(lp)
            try:
                long_df = pd.read_excel(lp, sheet_name=0, header=None)
            except Exception as e:  # noqa: BLE001
                print(f"[long] Failed to read {lp}: {e}", file=sys.stderr)
                continue

            # Parse embedded shift tables (both teams) from the long sheet.
            try:
                parsed_shift_tables = _parse_long_shift_tables(long_df)
            except Exception as e:  # noqa: BLE001
                parsed_shift_tables = {}
                print(f"[long] Failed to parse long shift tables from {lp}: {e}", file=sys.stderr)

            for team_name, info in (parsed_shift_tables or {}).items():
                dest = long_shift_tables_by_team.setdefault(
                    str(team_name),
                    {"sb_pairs_by_player": {}, "video_pairs_by_player": {}},
                )
                for k in ("sb_pairs_by_player", "video_pairs_by_player"):
                    src_map = (info or {}).get(k) or {}
                    dst_map = dest.setdefault(k, {})
                    for pk, lst in (src_map or {}).items():
                        dst_map.setdefault(pk, []).extend(list(lst or []))

            (
                long_events,
                long_goal_rows,
                jerseys_by_team,
                long_mapping_summary,
            ) = _parse_long_left_event_table_with_mapping(long_df)
            if not long_events and not long_goal_rows:
                # Still keep long shift tables / rosters (useful for validation).
                continue

            roster_tables.extend(_extract_roster_tables_from_df(long_df))
            long_events_all.extend(long_events or [])
            for team, nums in (jerseys_by_team or {}).items():
                jerseys_by_team_all.setdefault(team, set()).update(nums or set())
            long_goal_rows_all.extend(long_goal_rows or [])
            spreadsheet_event_mapping_summary.merge(long_mapping_summary)

            if focus_team is None:
                our_jerseys: set[str] = set(jersey_to_players.keys())
                if roster_map:
                    our_jerseys |= set(roster_map.keys())
                if t2s_rosters_by_side and t2s_side in {"home", "away"}:
                    our_jerseys |= set((t2s_rosters_by_side.get(str(t2s_side)) or {}).keys())
                focus_team = _infer_focus_team_from_long_sheet(our_jerseys, jerseys_by_team_all)

    # If we have a primary shift sheet and a long sheet, compare shift boundaries (diagnostic).
    if long_shift_tables_by_team:
        try:
            for team_name, info in sorted(long_shift_tables_by_team.items(), key=lambda x: x[0]):
                sb_any = (info or {}).get("sb_pairs_by_player") or {}
                n_players = len(sb_any)
                n_shifts = sum(len(v or []) for v in sb_any.values())
                print(
                    f"[long-shifts]{' [' + str(xls_path.name) + ']' if xls_path.name else ''} "
                    f"parsed long shift tables: team={team_name!r} players={n_players} shifts={n_shifts}",
                    file=sys.stderr,
                )
        except Exception:
            pass
        shift_cmp_summary = _compare_primary_shifts_to_long_shifts(
            primary_sb_pairs_by_player=sb_pairs_by_player,
            long_shift_tables_by_team=long_shift_tables_by_team,
            threshold_seconds=5,
            warn_label=str(xls_path.name),
            long_sheet_paths=long_sheet_paths_used,
        )
        if shift_cmp_summary:
            _write_shift_discrepancy_xlsx(stats_dir, shift_cmp_summary)

    # If we still don't know our color, try inferring from the primary sheet's event-log rosters.
    if (
        focus_team is None
        and merged_event_context is not None
        and (merged_event_context.team_roster or {})
    ):
        our_jerseys2: set[str] = set()
        if roster_map:
            our_jerseys2 |= set(roster_map.keys())
        if t2s_rosters_by_side and t2s_side in {"home", "away"}:
            our_jerseys2 |= set((t2s_rosters_by_side.get(str(t2s_side)) or {}).keys())
        focus_team = _infer_focus_team_from_color_rosters(
            our_jerseys2, merged_event_context.team_roster
        )

    # Build roster name maps by team color for resolving placeholder keys (Blue_#/White_#).
    roster_name_by_team: Dict[str, Dict[str, str]] = {"Blue": {}, "White": {}}

    def _assign_unknown_roster(roster: Dict[str, str]) -> Optional[str]:
        if not roster:
            return None
        roster_nums = set(roster.keys())
        # Prefer long-sheet observed rosters when available.
        if jerseys_by_team_all:
            blue = {str(int(x)) for x in (jerseys_by_team_all.get("Blue") or set())}
            white = {str(int(x)) for x in (jerseys_by_team_all.get("White") or set())}
            blue_ov = len(roster_nums & blue)
            white_ov = len(roster_nums & white)
            if blue_ov != white_ov and (blue_ov > 0 or white_ov > 0):
                return "Blue" if blue_ov > white_ov else "White"
        if merged_event_context is not None and (merged_event_context.team_roster or {}):
            blue = {str(int(x)) for x in (merged_event_context.team_roster.get("Blue") or [])}
            white = {str(int(x)) for x in (merged_event_context.team_roster.get("White") or [])}
            blue_ov = len(roster_nums & blue)
            white_ov = len(roster_nums & white)
            if blue_ov != white_ov and (blue_ov > 0 or white_ov > 0):
                return "Blue" if blue_ov > white_ov else "White"
        return None

    for team_guess, roster in roster_tables:
        team = team_guess if team_guess in {"Blue", "White"} else _assign_unknown_roster(roster)
        if team in {"Blue", "White"}:
            roster_name_by_team[team].update(roster or {})

    # Use player names from embedded long shift tables (both teams) to resolve opponent jerseys when
    # roster tables / TimeToScore rosters are not available.
    if long_shift_tables_by_team:
        long_team_rosters: Dict[str, Dict[str, str]] = {}
        for team_name, info in (long_shift_tables_by_team or {}).items():
            sb_any = (info or {}).get("sb_pairs_by_player") or {}
            roster: Dict[str, str] = {}
            for pk in sb_any.keys():
                parts = _parse_player_key(pk)
                jersey_norm = _normalize_jersey_number(parts.jersey)
                name = str(parts.name or "").strip()
                if not jersey_norm or not name:
                    continue
                # Keep a stable display name; it will be sanitized again when building player keys.
                roster.setdefault(jersey_norm, name.replace("_", " "))
            if roster:
                long_team_rosters[str(team_name)] = roster

        matched_long_team = str((shift_cmp_summary or {}).get("matched_team") or "").strip()
        if (
            focus_team in {"Blue", "White"}
            and matched_long_team
            and matched_long_team in long_team_rosters
        ):
            roster_name_by_team[focus_team].update(long_team_rosters.get(matched_long_team) or {})
            opp_team = "White" if focus_team == "Blue" else "Blue"
            other_names = [t for t in long_team_rosters.keys() if t != matched_long_team]
            if len(other_names) == 1:
                roster_name_by_team[opp_team].update(long_team_rosters.get(other_names[0]) or {})
        else:
            for _team_name, roster in long_team_rosters.items():
                team = _assign_unknown_roster(roster)
                if team in {"Blue", "White"}:
                    roster_name_by_team[team].update(roster or {})

    # Prefer TimeToScore rosters when available (resolve opponent names too).
    if t2s_rosters_by_side and t2s_side in {"home", "away"} and focus_team in {"Blue", "White"}:
        our_side = str(t2s_side)
        opp_side = "away" if our_side == "home" else "home"
        opp_team = "White" if focus_team == "Blue" else "Blue"
        roster_name_by_team[focus_team].update(t2s_rosters_by_side.get(our_side) or {})
        roster_name_by_team[opp_team].update(t2s_rosters_by_side.get(opp_side) or {})
    elif roster_map and focus_team in {"Blue", "White"}:
        roster_name_by_team[focus_team].update(roster_map or {})

    # Resolve placeholder player keys (e.g., 'Blue_12') to named keys when possible.
    key_map: Dict[str, str] = {}
    for pk in list(sb_pairs_by_player.keys()):
        nk = _resolve_team_color_player_key(pk, roster_name_by_team)
        if nk != pk:
            key_map[pk] = nk
    for pk in list(video_pairs_by_player.keys()):
        nk = _resolve_team_color_player_key(pk, roster_name_by_team)
        if nk != pk:
            key_map[pk] = nk
    if merged_event_context is not None:
        for pk in list((merged_event_context.event_counts_by_player or {}).keys()):
            nk = _resolve_team_color_player_key(pk, roster_name_by_team)
            if nk != pk:
                key_map[pk] = nk

    if key_map:
        sb_pairs_by_player = _rename_dict_keys_merge_lists(sb_pairs_by_player, key_map)  # type: ignore[arg-type]
        video_pairs_by_player = _rename_dict_keys_merge_lists(video_pairs_by_player, key_map)  # type: ignore[arg-type]
        merged_event_context = _rename_event_log_context_players(merged_event_context, key_map)

        # Rebuild jersey->player mapping after key renames.
        jersey_to_players = {}
        for pk in sb_pairs_by_player.keys():
            norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
            if norm:
                jersey_to_players.setdefault(norm, []).append(pk)

    # Convert parsed long-sheet events into an EventLogContext (with roster name resolution).
    if long_events_all:
        long_ctx = _event_log_context_from_long_events(
            long_events_all,
            jersey_to_players=jersey_to_players,
            focus_team=focus_team,
            jerseys_by_team=jerseys_by_team_all,
            roster_name_by_team=roster_name_by_team,
        )
        merged_event_context = _merge_event_log_contexts(merged_event_context, long_ctx)

        if focus_team is None and (long_goal_rows_all or jerseys_by_team_all):
            print(
                "[long] Could not infer whether your team is Blue or White; "
                "use --dark/--light to enable team-relative stats and goal inference.",
                file=sys.stderr,
            )

        if (not goals) and focus_team is not None and long_goal_rows_all:
            for row in long_goal_rows_all:
                team = row.get("team")
                period = row.get("period")
                gsec = row.get("game_s")
                scorer = row.get("scorer")
                assists = row.get("assists") or []
                if team not in {"Blue", "White"}:
                    continue
                if not isinstance(period, int):
                    continue
                if not isinstance(gsec, (int, float)):
                    continue
                kind = "GF" if team == focus_team else "GA"
                t_str = seconds_to_mmss_or_hhmmss(int(gsec))
                scorer_str = str(int(scorer)) if scorer is not None else None
                assists_str: List[str] = []
                for a in assists:
                    if a is None:
                        continue
                    try:
                        assists_str.append(str(int(a)))
                    except Exception:
                        continue
                inferred_long_goals.append(
                    GoalEvent(kind, int(period), t_str, scorer=scorer_str, assists=assists_str)
                )
            inferred_long_goals.sort(key=lambda e: (e.period, e.t_sec))

    # Use inferred long-sheet goals only as a fallback when no goals were provided.
    if not goals and inferred_long_goals:
        goals = inferred_long_goals

    # Stats and plus/minus
    goals_by_period: Dict[int, List[GoalEvent]] = {}
    for ev in goals:
        goals_by_period.setdefault(ev.period, []).append(ev)

    # Annotate game-tying / game-winning roles for goals in this game.
    _annotate_goal_roles(goals)

    pair_on_ice_rows: List[Dict[str, Any]] = []
    pair_on_ice_by_player: Dict[str, List[Dict[str, Any]]] = {}
    if sb_pairs_by_player:
        try:
            pair_on_ice_rows = _compute_pair_on_ice_rows(sb_pairs_by_player, goals_by_period)
            for r in pair_on_ice_rows:
                pk = str(r.get("player") or "")
                if pk:
                    pair_on_ice_by_player.setdefault(pk, []).append(r)
            for pk, lst in pair_on_ice_by_player.items():
                lst.sort(key=lambda x: float(x.get("overlap_pct", 0.0) or 0.0), reverse=True)
            _write_pair_on_ice_csv(stats_dir, pair_on_ice_rows, include_toi=include_shifts_in_stats)
        except Exception:
            pair_on_ice_rows = []
            pair_on_ice_by_player = {}

    # Per-player event details for this game (for multi-game summaries).
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]] = {
        pk: {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []}
        for pk in sb_pairs_by_player.keys()
    }

    goal_assist_counts: Dict[str, Dict[str, int]] = {
        pk: {"goals": 0, "assists": 0} for pk in sb_pairs_by_player.keys()
    }

    def _match_player_keys(num_token: Any) -> List[str]:
        matches: List[str] = []
        candidates: set[str] = set()
        if num_token is not None:
            try:
                txt = str(num_token).strip()
                if txt:
                    candidates.add(txt)
            except Exception as e:  # noqa: BLE001
                print(
                    f"[warning] Failed to parse jersey token {num_token!r}: {e}",
                    file=sys.stderr,
                )
        norm = _normalize_jersey_number(num_token)
        if norm:
            candidates.add(norm)
        for cand in candidates:
            matches.extend(jersey_to_players.get(cand, []))
        return list(dict.fromkeys(matches))  # dedupe while preserving order

    for ev in goals:
        if ev.kind != "GF":
            continue
        if ev.scorer:
            for pk in _match_player_keys(ev.scorer):
                goal_assist_counts[pk]["goals"] += 1
                per_player_goal_events[pk]["goals"].append(ev)
        for ast in ev.assists:
            for pk in _match_player_keys(ast):
                goal_assist_counts[pk]["assists"] += 1
                per_player_goal_events[pk]["assists"].append(ev)

    event_log_context = merged_event_context

    # Determine which long-sheet/player-attributed event types exist in this game.
    # This is used to avoid treating missing stats as 0 when a game lacks a `*-long*` sheet
    # or when a newer stat type wasn't collected for older games.
    player_event_types_present: set[str] = set()
    if event_log_context is not None:
        for counts in (event_log_context.event_counts_by_player or {}).values():
            for et in (counts or {}).keys():
                if et:
                    player_event_types_present.add(str(et))

    has_player_shots = "Shot" in player_event_types_present
    has_player_sog = "SOG" in player_event_types_present
    has_player_expected_goals = "ExpectedGoal" in player_event_types_present
    has_player_turnovers_forced = "TurnoverForced" in player_event_types_present
    has_player_created_turnovers = "CreatedTurnover" in player_event_types_present
    has_player_giveaways = "Giveaway" in player_event_types_present
    has_player_takeaways = "Takeaway" in player_event_types_present

    # Team-level event availability for on-ice for/against metrics.
    has_controlled_entry_events = False
    has_controlled_exit_events = False
    if focus_team in {"Blue", "White"} and event_log_context is not None:
        for etype, _team in (event_log_context.event_instances or {}).keys():
            if etype == "ControlledEntry":
                has_controlled_entry_events = True
            elif etype == "ControlledExit":
                has_controlled_exit_events = True

    if write_events_summary:
        t2s_events: List[Dict[str, Any]] = []
        if t2s_game_id is not None:
            t2s_events = t2s_events_from_scoresheet(
                int(t2s_game_id),
                our_side=t2s_side,
                allow_remote=allow_remote,
                allow_full_sync=allow_full_sync,
            )

        # Build Home/Away shift maps (for on-ice players in all_events_summary).
        side_l = str(t2s_side or "").strip().lower()
        our_side_label = "Home" if side_l == "home" else ("Away" if side_l == "away" else "")
        opp_side_label = (
            "Away" if our_side_label == "Home" else ("Home" if our_side_label == "Away" else "")
        )
        our_shifts = dict(sb_pairs_by_player or {})
        matched_long_team = str((shift_cmp_summary or {}).get("matched_team") or "").strip()
        if (
            (not any((v or []) for v in our_shifts.values()))
            and matched_long_team
            and long_shift_tables_by_team
        ):
            our_shifts = dict(
                (long_shift_tables_by_team.get(matched_long_team) or {}).get("sb_pairs_by_player")
                or {}
            )
        opp_shifts: Dict[str, List[Tuple[int, str, str]]] = {}
        if matched_long_team and long_shift_tables_by_team:
            other_teams = [
                t for t in long_shift_tables_by_team.keys() if str(t) != matched_long_team
            ]
            if len(other_teams) == 1:
                opp_shifts = dict(
                    (long_shift_tables_by_team.get(other_teams[0]) or {}).get("sb_pairs_by_player")
                    or {}
                )
        sb_by_side: Dict[str, Dict[str, List[Tuple[int, str, str]]]] = {}
        if our_side_label in {"Home", "Away"}:
            sb_by_side[our_side_label] = our_shifts
        if opp_side_label in {"Home", "Away"}:
            sb_by_side[opp_side_label] = opp_shifts

        # Merge long-sheet shift conversion segments too (enables mapping scoreboard->video and shift-boundary anchors).
        conv_segments_full: Dict[int, List[Tuple[int, int, int, int]]] = {
            int(k): list(v or []) for k, v in (conv_segments_by_period or {}).items()
        }
        if long_shift_tables_by_team:
            for _tname, info in (long_shift_tables_by_team or {}).items():
                sb_map = (info or {}).get("sb_pairs_by_player") or {}
                v_map = (info or {}).get("video_pairs_by_player") or {}
                for pk, sb_list in (sb_map or {}).items():
                    v_list = v_map.get(pk) or []
                    nseg = min(len(sb_list or []), len(v_list or []))
                    for idx in range(nseg):
                        try:
                            per, sba, sbb = sb_list[idx]
                            sva, svb = v_list[idx]
                            p_i = int(per)
                            s1 = parse_flex_time_to_seconds(str(sba))
                            s2 = parse_flex_time_to_seconds(str(sbb))
                            v1 = parse_flex_time_to_seconds(str(sva))
                            v2 = parse_flex_time_to_seconds(str(svb))
                            conv_segments_full.setdefault(p_i, []).append((s1, s2, v1, v2))
                        except Exception:
                            continue
        _write_all_events_summary(
            stats_dir,
            sb_pairs_by_player=sb_pairs_by_player,
            sb_pairs_by_player_by_side=sb_by_side if sb_by_side else None,
            goals=goals,
            goals_by_period=goals_by_period,
            event_log_context=event_log_context,
            focus_team=focus_team,
            team_side=t2s_side,
            t2s_game_id=t2s_game_id,
            t2s_events=t2s_events,
            conv_segments_by_period=conv_segments_full,
            spreadsheet_event_mapping_summary=spreadsheet_event_mapping_summary,
        )

    # Pre-group team-level events by period for on-ice for/against counts.
    on_ice_event_types = {"ControlledEntry", "ControlledExit"}
    team_events_by_period: Dict[int, List[Tuple[str, str, int]]] = {}
    if focus_team is not None and event_log_context is not None:
        for (etype, team), inst_list in (event_log_context.event_instances or {}).items():
            if etype not in on_ice_event_types:
                continue
            for it in inst_list or []:
                p = it.get("period")
                gs = it.get("game_s")
                if not isinstance(p, int) or not isinstance(gs, (int, float)):
                    continue
                team_events_by_period.setdefault(int(p), []).append((etype, str(team), int(gs)))

    for player_key, sb_list in sb_pairs_by_player.items():
        sb_by_period: Dict[int, List[Tuple[str, str]]] = {}
        for period, a, b in sb_list:
            sb_by_period.setdefault(period, []).append((a, b))
        for period in sb_by_period.keys():
            all_periods_seen.add(period)
        all_pairs = [(a, b) for (_, a, b) in sb_list]
        if include_shifts_in_stats:
            shift_summary = summarize_shift_lengths_sec(all_pairs)
            per_period_toi_map = per_period_toi(sb_by_period)
        else:
            shift_summary = {}
            per_period_toi_map = {}

        plus_minus = 0
        counted_gf: List[str] = []
        counted_ga: List[str] = []
        counted_gf_by_period: Dict[int, int] = {}
        counted_ga_by_period: Dict[int, int] = {}
        for period, pairs in sb_by_period.items():
            if period not in goals_by_period:
                continue
            for ev in goals_by_period[period]:
                matched = False
                for a, b in pairs:
                    a_sec = parse_flex_time_to_seconds(a)
                    b_sec = parse_flex_time_to_seconds(b)
                    lo, hi = (a_sec, b_sec) if a_sec <= b_sec else (b_sec, a_sec)
                    if not (lo <= ev.t_sec <= hi):
                        continue
                    if ev.kind == "GA" and ev.t_sec == a_sec:
                        continue
                    elif ev.kind == "GF" and ev.t_sec == a_sec:
                        continue
                    matched = True
                    break
                if matched:
                    if ev.kind == "GF":
                        plus_minus += 1
                        counted_gf.append(f"P{period}:{ev.t_str}")
                        counted_gf_by_period[period] = counted_gf_by_period.get(period, 0) + 1
                        per_player_goal_events[player_key]["gf_on_ice"].append(ev)
                    else:
                        plus_minus -= 1
                        counted_ga.append(f"P{period}:{ev.t_str}")
                        counted_ga_by_period[period] = counted_ga_by_period.get(period, 0) + 1
                        per_player_goal_events[player_key]["ga_on_ice"].append(ev)

        scoring_counts = goal_assist_counts.get(player_key, {"goals": 0, "assists": 0})
        goals_cnt = int(scoring_counts.get("goals", 0) or 0)
        assists_cnt = int(scoring_counts.get("assists", 0) or 0)
        points_val = goals_cnt + assists_cnt
        # Overtime scoring (OT period is period 4; OT2 -> 5, etc.).
        ot_goals_cnt = 0
        ot_assists_cnt = 0
        try:
            ev_map = per_player_goal_events.get(player_key, {}) or {}
            ot_goals_cnt = sum(
                1 for ev in (ev_map.get("goals") or []) if int(getattr(ev, "period", 0) or 0) >= 4
            )
            ot_assists_cnt = sum(
                1 for ev in (ev_map.get("assists") or []) if int(getattr(ev, "period", 0) or 0) >= 4
            )
        except Exception:
            ot_goals_cnt = 0
            ot_assists_cnt = 0

        # Game-tying / game-winning goals for this game.
        gt_goals_cnt = 0
        gw_goals_cnt = 0
        try:
            ev_map = per_player_goal_events.get(player_key, {}) or {}
            gt_goals_cnt, gw_goals_cnt = _count_goal_role_flags(list(ev_map.get("goals") or []))
        except Exception:
            gt_goals_cnt = 0
            gw_goals_cnt = 0

        # On-ice for/against metrics for team-level events (e.g., controlled exits)
        on_ice: Dict[str, int] = {
            "controlled_entry_for": 0,
            "controlled_entry_against": 0,
            "controlled_exit_for": 0,
            "controlled_exit_against": 0,
        }
        if focus_team is not None and team_events_by_period and sb_by_period:
            for period, pairs in sb_by_period.items():
                period_events = team_events_by_period.get(period, [])
                if not period_events or not pairs:
                    continue
                intervals = [compute_interval_seconds(a, b) for a, b in pairs]
                for etype, team, gsec in period_events:
                    in_any = False
                    for lo, hi in intervals:
                        if lo <= gsec <= hi:
                            in_any = True
                            break
                    if not in_any:
                        continue
                    is_for = team == focus_team
                    if etype == "ControlledEntry":
                        key = "controlled_entry_for" if is_for else "controlled_entry_against"
                        on_ice[key] += 1
                    elif etype == "ControlledExit":
                        key = "controlled_exit_for" if is_for else "controlled_exit_against"
                        on_ice[key] += 1

        stats_lines = []
        stats_lines.append(f"Player: {_display_player_name(player_key)}")
        stats_lines.append(f"Goals: {goals_cnt}")
        stats_lines.append(f"Assists: {assists_cnt}")
        stats_lines.append(f"OT Goals: {ot_goals_cnt}")
        stats_lines.append(f"OT Assists: {ot_assists_cnt}")
        stats_lines.append(f"Points (G+A): {points_val}")
        if include_shifts_in_stats:
            stats_lines.append(f"Shifts (scoreboard): {shift_summary.get('num_shifts', '0')}")
            stats_lines.append(f"TOI total (scoreboard): {shift_summary.get('toi_total', '0:00')}")
            stats_lines.append(f"Avg shift: {shift_summary.get('toi_avg', '0:00')}")
            stats_lines.append(f"Median shift: {shift_summary.get('toi_median', '0:00')}")
            stats_lines.append(f"Longest shift: {shift_summary.get('toi_longest', '0:00')}")
            stats_lines.append(f"Shortest shift: {shift_summary.get('toi_shortest', '0:00')}")
            if per_period_toi_map:
                stats_lines.append("Per-period TOI (scoreboard):")
                for period in sorted(per_period_toi_map.keys()):
                    stats_lines.append(f"  Period {period}: {per_period_toi_map[period]}")
        stats_lines.append(f"Goal +/-: {plus_minus}")
        if counted_gf:
            stats_lines.append("  GF counted at: " + ", ".join(sorted(counted_gf)))
        if counted_ga:
            stats_lines.append("  GA counted at: " + ", ".join(sorted(counted_ga)))

        if player_key in pair_on_ice_by_player:
            stats_lines.append("")
            stats_lines.append("On-ice with teammates (by TOI%):")
            disp_rows: List[Tuple[str, float, int, int, int, int, int]] = []
            for r in pair_on_ice_by_player.get(player_key, []):
                teammate = str(r.get("teammate") or "")
                if not teammate:
                    continue
                disp_rows.append(
                    (
                        _display_player_name(teammate),
                        float(r.get("overlap_pct", 0.0) or 0.0),
                        int(r.get("shift_games", 0) or 0),
                        int(r.get("plus_minus_together", 0) or 0),
                        int(r.get("gf_together", 0) or 0),
                        int(r.get("ga_together", 0) or 0),
                        int(r.get("overlap_seconds", 0) or 0),
                    )
                )
            name_w = max([len(r[0]) for r in disp_rows] + [len("Teammate")])
            if include_shifts_in_stats:
                stats_lines.append(
                    f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA  Overlap"
                )
                for teammate_disp, pct, shift_games, pm, gf, ga, overlap_s in disp_rows:
                    stats_lines.append(
                        f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}  {_format_duration(overlap_s):>7}"
                    )
            else:
                # Without --shifts, do not publish any absolute time values.
                stats_lines.append(f"  {'Teammate'.ljust(name_w)}  TOI%  GWD  +/-   GF   GA")
                for teammate_disp, pct, shift_games, pm, gf, ga, _overlap_s in disp_rows:
                    stats_lines.append(
                        f"  {teammate_disp.ljust(name_w)}  {pct:5.1f}  {shift_games:>3}  {_fmt_plus_minus(pm):>3}  {gf:>3}  {ga:>3}"
                    )

        if event_log_context is not None:
            per_player_events = event_log_context.event_counts_by_player
            ev_counts = per_player_events.get(player_key, {})
            if ev_counts:
                stats_lines.append("")
                stats_lines.append("Event Counts:")
                order = [
                    "Shot",
                    "SOG",
                    "Goal",
                    "Assist",
                    "ExpectedGoal",
                    "TurnoverForced",
                    "CreatedTurnover",
                    "Giveaway",
                    "Takeaway",
                    "CompletedPass",
                    "ControlledEntry",
                    "ControlledExit",
                ]
                sog_cnt = int(ev_counts.get("SOG", 0) or 0)
                xg_cnt = int(ev_counts.get("ExpectedGoal", 0) or 0)
                for kind in order:
                    if kind == "ExpectedGoal":
                        # If the player has at least one SOG, always show xG even when it's 0
                        # (so parents can see the full SOG/xG line without missing fields).
                        if sog_cnt > 0 or xg_cnt > 0:
                            stats_lines.append(f"  {_display_event_type(kind)}: {xg_cnt}")
                        continue
                    if kind in ev_counts and ev_counts[kind] > 0:
                        stats_lines.append(f"  {_display_event_type(kind)}: {ev_counts[kind]}")
                for kind, cnt in sorted(ev_counts.items()):
                    if kind in order:
                        continue
                    stats_lines.append(f"  {_display_event_type(kind)}: {cnt}")

        if focus_team is not None and any(v > 0 for v in on_ice.values()):
            stats_lines.append("")
            stats_lines.append("On-ice team events (for/against):")
            stats_lines.append(
                f"  ControlledEntry: {on_ice['controlled_entry_for']} for, {on_ice['controlled_entry_against']} against"
            )
            stats_lines.append(
                f"  ControlledExit: {on_ice['controlled_exit_for']} for, {on_ice['controlled_exit_against']} against"
            )

        if include_shifts_in_stats:
            for period, pairs in sorted(sb_by_period.items()):
                stats_lines.append(f"Shifts in Period {period}: {len(pairs)}")

        (stats_dir / f"{player_key}_stats.txt").write_text(
            "\n".join(stats_lines) + "\n", encoding="utf-8"
        )

        row_map: Dict[str, str] = {
            "player": player_key,
            "goals": str(goals_cnt),
            "assists": str(assists_cnt),
            "gt_goals": str(gt_goals_cnt),
            "gw_goals": str(gw_goals_cnt),
            "ot_goals": str(ot_goals_cnt),
            "ot_assists": str(ot_assists_cnt),
            "points": str(points_val),
            "gp": "1",
            "plus_minus": str(plus_minus),
        }
        if include_shifts_in_stats:
            shifts_cnt_row = 0
            try:
                shifts_cnt_row = int(str(shift_summary.get("num_shifts", "0") or 0))
            except Exception:
                shifts_cnt_row = 0
            row_map["shifts"] = str(shifts_cnt_row)
            row_map["sb_toi_total"] = str(shift_summary.get("toi_total", "0:00"))
            row_map["sb_avg"] = str(shift_summary.get("toi_avg", "0:00"))
            row_map["sb_median"] = str(shift_summary.get("toi_median", "0:00"))
            row_map["sb_longest"] = str(shift_summary.get("toi_longest", "0:00"))
            row_map["sb_shortest"] = str(shift_summary.get("toi_shortest", "0:00"))
        # Event counts (from event logs / long sheets), per game.
        if event_log_context is not None:
            ev_counts = (event_log_context.event_counts_by_player or {}).get(player_key, {})
        else:
            ev_counts = {}
        has_completed_passes_local = "CompletedPass" in player_event_types_present
        shots_cnt = int(ev_counts.get("Shot", 0) or 0)
        sog_cnt = int(ev_counts.get("SOG", 0) or 0)
        if not has_player_sog and has_player_shots:
            sog_cnt = shots_cnt
            if sog_cnt > 0 and "SOG" not in ev_counts:
                ev_counts = dict(ev_counts)
                ev_counts["SOG"] = sog_cnt
        expected_goals_cnt = int(ev_counts.get("ExpectedGoal", 0) or 0)
        turnovers_forced_cnt = int(ev_counts.get("TurnoverForced", 0) or 0)
        created_turnovers_cnt = int(ev_counts.get("CreatedTurnover", 0) or 0)
        giveaways_cnt = int(ev_counts.get("Giveaway", 0) or 0)
        takeaways_cnt = int(ev_counts.get("Takeaway", 0) or 0)
        completed_passes_cnt = int(ev_counts.get("CompletedPass", 0) or 0)
        if has_player_shots:
            row_map["shots"] = str(shots_cnt)
        else:
            row_map["shots"] = ""

        row_map["sog"] = (
            str(sog_cnt) if (has_player_sog or (has_player_shots and sog_cnt >= 0)) else ""
        )

        if has_player_expected_goals:
            row_map["expected_goals"] = str(expected_goals_cnt)
        else:
            row_map["expected_goals"] = ""

        if has_player_expected_goals and (has_player_sog or has_player_shots):
            row_map["expected_goals_per_sog"] = (
                f"{(expected_goals_cnt / sog_cnt):.2f}" if sog_cnt > 0 else ""
            )
        else:
            row_map["expected_goals_per_sog"] = ""

        if has_player_turnovers_forced:
            row_map["turnovers_forced"] = str(turnovers_forced_cnt)
        else:
            row_map["turnovers_forced"] = ""

        if has_player_created_turnovers:
            row_map["created_turnovers"] = str(created_turnovers_cnt)
        else:
            row_map["created_turnovers"] = ""

        if has_player_giveaways:
            row_map["giveaways"] = str(giveaways_cnt)
        else:
            row_map["giveaways"] = ""

        if has_player_takeaways:
            row_map["takeaways"] = str(takeaways_cnt)
        else:
            row_map["takeaways"] = ""

        if has_completed_passes_local:
            row_map["completed_passes"] = str(completed_passes_cnt)
        else:
            row_map["completed_passes"] = ""

        if has_controlled_entry_events:
            row_map["controlled_entry_for"] = str(on_ice["controlled_entry_for"])
            row_map["controlled_entry_against"] = str(on_ice["controlled_entry_against"])
        else:
            row_map["controlled_entry_for"] = ""
            row_map["controlled_entry_against"] = ""

        if has_controlled_exit_events:
            row_map["controlled_exit_for"] = str(on_ice["controlled_exit_for"])
            row_map["controlled_exit_against"] = str(on_ice["controlled_exit_against"])
        else:
            row_map["controlled_exit_for"] = ""
            row_map["controlled_exit_against"] = ""

        row_map["gf_counted"] = str(len(counted_gf))
        row_map["ga_counted"] = str(len(counted_ga))
        if include_shifts_in_stats:
            v_pairs = video_pairs_by_player.get(player_key, [])
            if v_pairs:
                v_sum = 0
                for a, b in v_pairs:
                    lo, hi = compute_interval_seconds(a, b)
                    v_sum += hi - lo
                row_map["video_toi_total"] = _format_duration(v_sum)
            else:
                row_map["video_toi_total"] = ""
            for period, toi in per_period_toi_map.items():
                row_map[f"P{period}_toi"] = toi
                all_periods_seen.add(period)
            for period, pairs in sb_by_period.items():
                row_map[f"P{period}_shifts"] = str(len(pairs))
        for period, cnt in counted_gf_by_period.items():
            row_map[f"P{period}_GF"] = str(cnt)
            all_periods_seen.add(period)
        for period, cnt in counted_ga_by_period.items():
            row_map[f"P{period}_GA"] = str(cnt)
            all_periods_seen.add(period)

        stats_table_rows.append(row_map)

    # Always write shift_rows.csv when writing event summaries (webapp upload/auditing), even if
    # shift/TOI columns are not being included in the parent-facing stats outputs.
    if write_events_summary:
        try:
            _write_shift_rows_csv(
                stats_dir,
                sb_pairs_by_player=sb_pairs_by_player,
                video_pairs_by_player=video_pairs_by_player,
                source="shift_spreadsheet",
            )
        except Exception as exc:  # noqa: BLE001
            print(f"Warning: failed to write shift_rows.csv: {exc!r}", file=sys.stderr)
    if include_shifts_in_stats:
        _write_global_summary_csv(stats_dir, sb_pairs_by_player)

    period_list = sorted(all_periods_seen)
    # Consolidated player stats
    if stats_table_rows:
        _write_player_stats_text_and_csv(
            stats_dir,
            stats_table_rows,
            period_list,
            include_shifts_in_stats=include_shifts_in_stats,
        )

    # Per-game team stats (no TOI)
    _write_game_stats_files(
        stats_dir,
        xls_path=xls_path,
        periods=period_list,
        goals=goals,
        event_log_context=event_log_context,
        focus_team=focus_team,
    )

    # If a long sheet provides embedded shift tables for both teams, also write the opponent's
    # per-player stats under the opposite Home/Away subtree (primarily for webapp import).
    if write_opponent_stats_from_long_shifts and long_shift_tables_by_team and shift_cmp_summary:
        try:
            opp_outdir = _write_opponent_team_stats_from_long_shifts(
                game_out_root=outdir.parent.parent,
                format_dir=format_dir,
                our_side=str(t2s_side or ""),
                long_shift_tables_by_team=long_shift_tables_by_team,
                shift_cmp_summary=shift_cmp_summary,
                goals=goals,
                event_log_context=event_log_context,
                focus_team=focus_team,
                include_shifts_in_stats=include_shifts_in_stats,
                write_shift_rows_csv=bool(write_events_summary),
                xls_path=xls_path,
                t2s_rosters_by_side=t2s_rosters_by_side,
                create_scripts=create_scripts,
                skip_if_exists=False,
            )
            # Mirror the combined event summary into the opponent subtree (it contains both Home/Away on-ice lists).
            if opp_outdir is not None:
                opp_stats = Path(opp_outdir) / "stats"
                for fn in ("all_events_summary.csv", "all_events_summary.xlsx"):
                    src = stats_dir / fn
                    dst = opp_stats / fn
                    try:
                        if src.exists() and src.is_file():
                            dst.write_bytes(src.read_bytes())
                    except Exception:
                        pass
        except Exception:
            pass

    # Goals windows
    _write_goal_window_files(outdir, goals, conv_segments_by_period)

    # Event summaries
    if event_log_context is not None:
        _write_event_summaries_and_clips(
            outdir,
            stats_dir,
            event_log_context,
            conv_segments_by_period,
            create_scripts=create_scripts,
            focus_team=focus_team,
        )
        _write_player_event_highlights(
            outdir,
            event_log_context,
            conv_segments_by_period,
            sb_pairs_by_player.keys(),
            create_scripts=create_scripts,
        )
        _write_player_combined_highlights(
            outdir,
            event_log_context=event_log_context,
            conv_segments_by_period=conv_segments_by_period,
            per_player_goal_events=per_player_goal_events,
            player_keys=sb_pairs_by_player.keys(),
            create_scripts=create_scripts,
        )
    else:
        # Still write combined per-player highlights for goals/assists when available.
        _write_player_combined_highlights(
            outdir,
            event_log_context=None,
            conv_segments_by_period=conv_segments_by_period,
            per_player_goal_events=per_player_goal_events,
            player_keys=sb_pairs_by_player.keys(),
            create_scripts=create_scripts,
        )

    # Aggregate clip runner (optional scripts)
    _write_clip_all_runner(outdir, create_scripts=create_scripts)

    # Validation summary
    if (not used_event_log) and (not skip_validation) and validation_errors > 0:
        print(
            f"[validation] Completed with {validation_errors} issue(s). See messages above.",
            file=sys.stderr,
        )

    # At end, print a Rich summary table for primary-vs-long shift discrepancies (when available).
    if shift_cmp_summary:
        _print_shift_discrepancy_rich_summary(shift_cmp_summary)

    if verbose:
        _print_spreadsheet_event_mapping_summary(
            label=str(xls_path.name), summary=spreadsheet_event_mapping_summary
        )
        if spreadsheet_event_mapping_summary.unmapped_row_counts:
            unmapped_pretty = ", ".join(
                f"(label={lbl!r} marker={mkr!r} rows={int(cnt or 0)})"
                for (lbl, mkr), cnt in sorted(
                    (spreadsheet_event_mapping_summary.unmapped_row_counts or {}).items(),
                    key=lambda kv: (-int(kv[1] or 0), kv[0][0], kv[0][1]),
                )
            )
            raise ValueError(f"Unmapped spreadsheet event rows: {unmapped_pretty}")

    return (
        outdir,
        stats_table_rows,
        sorted(all_periods_seen),
        per_player_goal_events,
        pair_on_ice_rows,
    )


def process_goals_only_xlsx(
    *,
    goals_xlsx: Path,
    outdir: Path,
    label: str,
    team_side: Optional[str],
    our_team_name: Optional[str],
    write_events_summary: bool,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
    List[Dict[str, Any]],
]:
    """
    Process a goals-only game where no shift sheet exists and goals are provided via goals.xlsx.

    Output is minimal: `stats/all_events_summary.csv` (and .xlsx) so the webapp can upsert
    goal/assist events for an external game.
    """
    outdir.mkdir(parents=True, exist_ok=True)

    # Match process_sheet() directory structure so the webapp uploader finds the stats:
    #   <outdir>/<Home|Away>/per_player/stats/*
    side = str(team_side or "").strip().lower()
    team_subdir = "Away" if side == "away" else "Home"
    format_dir = "per_player"
    (outdir / "Home").mkdir(parents=True, exist_ok=True)
    (outdir / "Away").mkdir(parents=True, exist_ok=True)
    outdir = outdir / team_subdir / format_dir
    outdir.mkdir(parents=True, exist_ok=True)
    stats_dir = outdir / "stats"
    stats_dir.mkdir(parents=True, exist_ok=True)

    goals = _goals_from_goals_xlsx(goals_xlsx, our_team_name=our_team_name)
    goals_by_period: Dict[int, List[GoalEvent]] = {}
    for ev in goals or []:
        goals_by_period.setdefault(int(ev.period), []).append(ev)

    periods = sorted({int(ev.period) for ev in goals or [] if int(ev.period) > 0})

    if write_events_summary:
        _write_all_events_summary(
            stats_dir,
            sb_pairs_by_player={},
            goals=goals,
            goals_by_period=goals_by_period,
            event_log_context=None,
            focus_team=None,
            team_side=team_side,
            t2s_game_id=None,
            t2s_events=None,
            conv_segments_by_period={},
            spreadsheet_event_mapping_summary=None,
        )

    # Write game_stats.csv so the webapp can set the final score for external games.
    try:
        fake_sheet_path = Path(f"{str(label or 'game').strip() or 'game'}.xlsx")
        _write_game_stats_files(
            stats_dir,
            xls_path=fake_sheet_path,
            periods=periods,
            goals=goals,
            event_log_context=None,
            focus_team=None,
        )
    except Exception as e:  # noqa: BLE001
        print(
            f"[warn] goals-only: failed to write game_stats.csv for {label!r}: {type(e).__name__}: {e}",
            file=sys.stderr,
        )

    stats_table_rows: List[Dict[str, str]] = []
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]] = {}
    pair_on_ice_rows: List[Dict[str, Any]] = []
    return outdir, stats_table_rows, periods, per_player_goal_events, pair_on_ice_rows


def process_long_only_sheets(
    *,
    long_xls_paths: List[Path],
    outdir: Path,
    goals: List[GoalEvent],
    roster_map: Optional[Dict[str, str]] = None,
    t2s_rosters_by_side: Optional[Dict[str, Dict[str, str]]] = None,
    t2s_side: Optional[str] = None,
    t2s_game_id: Optional[int] = None,
    focus_team_override: Optional[str] = None,
    include_shifts_in_stats: bool = False,
    write_events_summary: Optional[bool] = None,
    create_scripts: bool = True,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
    verbose: bool = False,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
    List[Dict[str, Any]],
]:
    """
    Process a game when only '*-long*' spreadsheets are available.

    Uses the embedded long-sheet shift table for our team as the primary shift source, and (when
    available) writes opponent player stats from the other embedded shift table.
    """
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "Home").mkdir(parents=True, exist_ok=True)
    (outdir / "Away").mkdir(parents=True, exist_ok=True)

    side = str(t2s_side or "").strip().lower()
    if side not in {"home", "away"}:
        raise ValueError("process_long_only_sheets requires t2s_side='home' or 'away'")
    format_dir = "per_player"

    if write_events_summary is None:
        write_events_summary = include_shifts_in_stats

    # Parse all long sheets: embedded shift tables + left event table.
    long_shift_tables_by_team: Dict[str, Dict[str, Dict[str, List[Tuple[Any, ...]]]]] = {}
    long_events_all: List[LongEvent] = []
    long_goal_rows_all: List[Dict[str, Any]] = []
    jerseys_by_team_all: Dict[str, set[int]] = {}
    spreadsheet_event_mapping_summary = SpreadsheetEventMappingSummary.empty()
    long_sheet_paths_used: List[Path] = []
    for lp in long_xls_paths or []:
        if lp is None:
            continue
        p = Path(lp).expanduser()
        if not p.exists():
            continue
        long_sheet_paths_used.append(p)
        long_df = pd.read_excel(p, sheet_name=0, header=None)

        try:
            parsed_shift_tables = _parse_long_shift_tables(long_df)
        except Exception as e:  # noqa: BLE001
            parsed_shift_tables = {}
            print(f"[long] Failed to parse long shift tables from {p}: {e}", file=sys.stderr)

        for team_name, info in (parsed_shift_tables or {}).items():
            dest = long_shift_tables_by_team.setdefault(
                str(team_name),
                {"sb_pairs_by_player": {}, "video_pairs_by_player": {}},
            )
            for k in ("sb_pairs_by_player", "video_pairs_by_player"):
                src_map = (info or {}).get(k) or {}
                dst_map = dest.setdefault(k, {})
                for pk, lst in (src_map or {}).items():
                    dst_map.setdefault(pk, []).extend(list(lst or []))

        try:
            (
                long_events,
                long_goal_rows,
                jerseys_by_team,
                long_mapping_summary,
            ) = _parse_long_left_event_table_with_mapping(long_df)
        except Exception as e:  # noqa: BLE001
            long_events, long_goal_rows, jerseys_by_team, long_mapping_summary = (
                [],
                [],
                {},
                SpreadsheetEventMappingSummary.empty(),
            )
            print(f"[long] Failed to parse long left event table from {p}: {e}", file=sys.stderr)

        long_events_all.extend(list(long_events or []))
        long_goal_rows_all.extend(list(long_goal_rows or []))
        for team, nums in (jerseys_by_team or {}).items():
            jerseys_by_team_all.setdefault(str(team), set()).update(set(nums or set()))
        spreadsheet_event_mapping_summary.merge(long_mapping_summary)

    if not long_shift_tables_by_team:
        raise ValueError("No embedded long-sheet shift tables were found; cannot compute shifts.")

    # Normalize jerseys seen in each long shift table (for matching).
    jerseys_by_long_team: Dict[str, set[str]] = {}
    for team_name, info in long_shift_tables_by_team.items():
        sb_any = (info or {}).get("sb_pairs_by_player") or {}
        jerseys: set[str] = set()
        for pk in sb_any.keys():
            parts = _parse_player_key(pk)
            norm = _normalize_jersey_number(parts.jersey)
            if norm:
                jerseys.add(norm)
        # Some malformed / partial long sheets can produce placeholder team names
        # with no parsed players. Ignore those for "our team" selection.
        if jerseys:
            jerseys_by_long_team[str(team_name)] = jerseys

    # Determine our Blue/White focus team from long events (when possible).
    focus_team: Optional[str] = focus_team_override
    if focus_team not in {"Blue", "White"}:
        our_jerseys: set[str] = set()
        if roster_map:
            our_jerseys |= set(roster_map.keys())
        if t2s_rosters_by_side and side in {"home", "away"}:
            our_jerseys |= set((t2s_rosters_by_side.get(side) or {}).keys())
        focus_team = _infer_focus_team_from_long_sheet(our_jerseys, jerseys_by_team_all)

    # Map each embedded shift table team name -> Blue/White (best-effort) by overlap with event rosters.
    blue_set = {str(int(x)) for x in (jerseys_by_team_all.get("Blue") or set())}
    white_set = {str(int(x)) for x in (jerseys_by_team_all.get("White") or set())}
    long_team_color: Dict[str, Optional[str]] = {}
    if blue_set or white_set:
        for team_name, nums in jerseys_by_long_team.items():
            ob = len(nums & blue_set) if nums else 0
            ow = len(nums & white_set) if nums else 0
            if ob > ow and ob > 0:
                long_team_color[team_name] = "Blue"
            elif ow > ob and ow > 0:
                long_team_color[team_name] = "White"
            else:
                long_team_color[team_name] = None

    # Pick which embedded shift table is ours.
    our_team_name: Optional[str] = None
    if focus_team in {"Blue", "White"} and long_team_color:
        best = -1
        for team_name, color in long_team_color.items():
            if color != focus_team:
                continue
            ov = len(
                jerseys_by_long_team.get(team_name, set())
                & (blue_set if focus_team == "Blue" else white_set)
            )
            if ov > best:
                best = ov
                our_team_name = team_name
    if our_team_name is None and roster_map:
        our_roster = set(roster_map.keys())
        if our_roster:
            best2 = -1
            for team_name, nums in jerseys_by_long_team.items():
                ov = len(nums & our_roster)
                if ov > best2:
                    best2 = ov
                    our_team_name = team_name
    if our_team_name is None:
        if len(jerseys_by_long_team) == 1:
            our_team_name = next(iter(jerseys_by_long_team.keys()))
        else:
            teams = ", ".join(sorted(jerseys_by_long_team.keys()))
            raise ValueError(
                "Cannot infer which long shift table is your team. "
                f"Found teams: {teams}. Provide a TimeToScore id (for roster matching) or include the primary shift sheet."
            )

    our_team_name = str(our_team_name)

    # Build roster_name_by_team for resolving opponent player names in long events.
    roster_name_by_team: Dict[str, Dict[str, str]] = {"Blue": {}, "White": {}}
    if long_team_color:
        for team_name, info in long_shift_tables_by_team.items():
            color = long_team_color.get(team_name)
            if color not in {"Blue", "White"}:
                continue
            sb_any = (info or {}).get("sb_pairs_by_player") or {}
            for pk in sb_any.keys():
                parts = _parse_player_key(pk)
                jersey_norm = _normalize_jersey_number(parts.jersey)
                name = str(parts.name or "").replace("_", " ").strip()
                if jersey_norm and name:
                    roster_name_by_team[color].setdefault(jersey_norm, name)

    if t2s_rosters_by_side and side in {"home", "away"} and focus_team in {"Blue", "White"}:
        our_side = side
        opp_side = "away" if our_side == "home" else "home"
        opp_team = "White" if focus_team == "Blue" else "Blue"
        roster_name_by_team[focus_team].update(t2s_rosters_by_side.get(our_side) or {})
        roster_name_by_team[opp_team].update(t2s_rosters_by_side.get(opp_side) or {})
    elif roster_map and focus_team in {"Blue", "White"}:
        roster_name_by_team[focus_team].update(roster_map or {})

    # Build jersey_to_players mapping for our team from the long shift table.
    our_sb_pairs_by_player = (long_shift_tables_by_team.get(our_team_name) or {}).get(
        "sb_pairs_by_player"
    ) or {}
    jersey_to_players: Dict[str, List[str]] = {}
    for pk in our_sb_pairs_by_player.keys():
        norm = _normalize_jersey_number(_parse_player_key(pk).jersey)
        if norm:
            jersey_to_players.setdefault(norm, []).append(pk)

    merged_event_context: Optional[EventLogContext] = None
    if long_events_all:
        merged_event_context = _event_log_context_from_long_events(
            long_events_all,
            jersey_to_players=jersey_to_players,
            focus_team=focus_team,
            jerseys_by_team=jerseys_by_team_all,
            roster_name_by_team=roster_name_by_team,
        )

    # Write our team stats from the selected long shift table.
    primary_long_path = (
        long_sheet_paths_used[0]
        if long_sheet_paths_used
        else (long_xls_paths[0] if long_xls_paths else Path("game-long.xlsx"))
    )
    our_outdir, stats_rows, periods, per_player_events, pair_on_ice_rows = (
        _write_team_stats_from_long_shift_team(
            game_out_root=outdir,
            format_dir=format_dir,
            team_side=side,
            team_name=our_team_name,
            long_shift_tables_by_team=long_shift_tables_by_team,
            goals=goals,
            event_log_context=merged_event_context,
            focus_team=focus_team,
            include_shifts_in_stats=include_shifts_in_stats,
            write_shift_rows_csv=bool(write_events_summary),
            xls_path=Path(primary_long_path),
            t2s_rosters_by_side=t2s_rosters_by_side,
            create_scripts=create_scripts,
            skip_if_exists=False,
        )
    )

    # Write all_events_summary (for webapp import / auditing).
    if write_events_summary:
        t2s_events: List[Dict[str, Any]] = []
        if t2s_game_id is not None:
            t2s_events = t2s_events_from_scoresheet(
                int(t2s_game_id),
                our_side=side,
                allow_remote=allow_remote,
                allow_full_sync=allow_full_sync,
            )
        goals_by_period: Dict[int, List[GoalEvent]] = {}
        for ev in goals or []:
            goals_by_period.setdefault(int(ev.period), []).append(ev)

        our_side_label = "Home" if str(side or "").strip().lower() == "home" else "Away"
        opp_side_label = "Away" if our_side_label == "Home" else "Home"
        opp_team_names = [
            t for t in (long_shift_tables_by_team or {}).keys() if str(t) != str(our_team_name)
        ]
        opp_shifts: Dict[str, List[Tuple[int, str, str]]] = {}
        if len(opp_team_names) == 1:
            opp_shifts = dict(
                (long_shift_tables_by_team.get(opp_team_names[0]) or {}).get("sb_pairs_by_player")
                or {}
            )
        sb_by_side = {our_side_label: dict(our_sb_pairs_by_player), opp_side_label: opp_shifts}

        # Build conversion segments from long shift tables for mapping scoreboard->video and anchors.
        conv_segments_full: Dict[int, List[Tuple[int, int, int, int]]] = {}
        for _tname, info in (long_shift_tables_by_team or {}).items():
            sb_map = (info or {}).get("sb_pairs_by_player") or {}
            v_map = (info or {}).get("video_pairs_by_player") or {}
            for pk, sb_list in (sb_map or {}).items():
                v_list = v_map.get(pk) or []
                nseg = min(len(sb_list or []), len(v_list or []))
                for idx in range(nseg):
                    try:
                        per, sba, sbb = sb_list[idx]
                        sva, svb = v_list[idx]
                        p_i = int(per)
                        s1 = parse_flex_time_to_seconds(str(sba))
                        s2 = parse_flex_time_to_seconds(str(sbb))
                        v1 = parse_flex_time_to_seconds(str(sva))
                        v2 = parse_flex_time_to_seconds(str(svb))
                        conv_segments_full.setdefault(p_i, []).append((s1, s2, v1, v2))
                    except Exception:
                        continue
        _write_all_events_summary(
            our_outdir / "stats",
            sb_pairs_by_player=dict(our_sb_pairs_by_player),
            sb_pairs_by_player_by_side=sb_by_side,
            goals=list(goals or []),
            goals_by_period=goals_by_period,
            event_log_context=merged_event_context,
            focus_team=focus_team,
            team_side=side,
            t2s_game_id=t2s_game_id,
            t2s_events=t2s_events,
            conv_segments_by_period=conv_segments_full,
            spreadsheet_event_mapping_summary=spreadsheet_event_mapping_summary,
        )

    if verbose:
        _print_spreadsheet_event_mapping_summary(
            label=str(outdir.name), summary=spreadsheet_event_mapping_summary
        )
        if spreadsheet_event_mapping_summary.unmapped_row_counts:
            unmapped_pretty = ", ".join(
                f"(label={lbl!r} marker={mkr!r} rows={int(cnt or 0)})"
                for (lbl, mkr), cnt in sorted(
                    (spreadsheet_event_mapping_summary.unmapped_row_counts or {}).items(),
                    key=lambda kv: (-int(kv[1] or 0), kv[0][0], kv[0][1]),
                )
            )
            raise ValueError(f"Unmapped spreadsheet event rows: {unmapped_pretty}")

    # Write opponent player stats (when another embedded shift table exists).
    try:
        opp_outdir = _write_opponent_team_stats_from_long_shifts(
            game_out_root=outdir,
            format_dir=format_dir,
            our_side=side,
            long_shift_tables_by_team=long_shift_tables_by_team,
            shift_cmp_summary={
                "matched_team": our_team_name,
                "long_sheets": [str(p) for p in long_sheet_paths_used],
            },
            goals=goals,
            event_log_context=merged_event_context,
            focus_team=focus_team,
            include_shifts_in_stats=include_shifts_in_stats,
            write_shift_rows_csv=bool(write_events_summary),
            xls_path=Path(primary_long_path),
            t2s_rosters_by_side=t2s_rosters_by_side,
            create_scripts=create_scripts,
            skip_if_exists=False,
        )
        if opp_outdir is not None:
            opp_stats = Path(opp_outdir) / "stats"
            src_stats = our_outdir / "stats"
            for fn in ("all_events_summary.csv", "all_events_summary.xlsx"):
                src = src_stats / fn
                dst = opp_stats / fn
                try:
                    if src.exists() and src.is_file():
                        dst.write_bytes(src.read_bytes())
                except Exception:
                    pass
    except Exception:
        pass

    # Return our team's outputs (for season aggregation).
    return our_outdir, stats_rows, periods, per_player_events, pair_on_ice_rows


def process_t2s_only_game(
    *,
    t2s_id: int,
    side: str,
    outdir: Path,
    label: str,
    hockey_db_dir: Path,
    include_shifts_in_stats: bool,
    keep_goalies: bool = False,
    allow_remote: bool = True,
    allow_full_sync: bool = True,
) -> Tuple[
    Path,
    List[Dict[str, str]],
    List[int],
    Dict[str, Dict[str, List[GoalEvent]]],
]:
    """
    Process a game using only TimeToScore data (no shift spreadsheets).

    Writes the same `per_player/stats/*` outputs as `process_sheet`, but leaves
    shift/TOI and on-ice (+/-) fields blank since they cannot be derived without
    shift timing data.
    """
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "Home").mkdir(parents=True, exist_ok=True)
    (outdir / "Away").mkdir(parents=True, exist_ok=True)
    side_l = str(side or "").strip().lower()
    team_subdir = "Away" if side_l == "away" else "Home"
    outdir = outdir / team_subdir / "per_player"
    outdir.mkdir(parents=True, exist_ok=True)
    stats_dir = outdir / "stats"
    stats_dir.mkdir(parents=True, exist_ok=True)

    # Load roster + scoring from TimeToScore.
    with _working_directory(hockey_db_dir):
        goals = goals_from_t2s(
            int(t2s_id),
            side=side,
            allow_remote=allow_remote,
            allow_full_sync=allow_full_sync,
        )
    roster_map = _get_t2s_team_roster(
        int(t2s_id),
        side,
        hockey_db_dir,
        keep_goalies=bool(keep_goalies),
        allow_remote=allow_remote,
        allow_full_sync=allow_full_sync,
    )

    # Annotate game-tying / game-winning roles.
    _annotate_goal_roles(goals)

    periods_seen = sorted(
        {
            int(getattr(ev, "period", 0) or 0)
            for ev in goals
            if isinstance(getattr(ev, "period", None), int)
            and int(getattr(ev, "period", 0) or 0) > 0
        }
    )

    # Build internal player keys from T2S roster (normalized jersey -> name).
    player_keys: List[str] = []
    jersey_to_player: Dict[str, str] = {}
    for jersey_norm, name in sorted((roster_map or {}).items(), key=lambda x: x[0]):
        if not jersey_norm or not name:
            continue
        player_key = f"{sanitize_name(jersey_norm)}_{sanitize_name(name)}"
        player_keys.append(player_key)
        jersey_to_player[jersey_norm] = player_key

    def _sort_player_key(pk: str) -> Tuple[int, str]:
        try:
            norm = _normalize_jersey_number(_parse_player_key(pk).jersey) or ""
            return int(norm), pk
        except Exception:
            return 9999, pk

    player_keys.sort(key=_sort_player_key)

    # Per-player event details for this game (for multi-game summaries).
    per_player_goal_events: Dict[str, Dict[str, List[GoalEvent]]] = {
        pk: {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []} for pk in player_keys
    }
    goal_assist_counts: Dict[str, Dict[str, int]] = {
        pk: {"goals": 0, "assists": 0} for pk in player_keys
    }

    def _match_player_keys(num_token: Any) -> List[str]:
        norm = _normalize_jersey_number(num_token)
        if not norm:
            return []
        pk = jersey_to_player.get(norm)
        return [pk] if pk else []

    for ev in goals:
        if ev.kind != "GF":
            continue
        if ev.scorer:
            for pk in _match_player_keys(ev.scorer):
                goal_assist_counts[pk]["goals"] += 1
                per_player_goal_events[pk]["goals"].append(ev)
        for ast in ev.assists:
            for pk in _match_player_keys(ast):
                goal_assist_counts[pk]["assists"] += 1
                per_player_goal_events[pk]["assists"].append(ev)

    stats_table_rows: List[Dict[str, str]] = []
    for player_key in player_keys:
        scoring_counts = goal_assist_counts.get(player_key, {"goals": 0, "assists": 0})
        goals_cnt = int(scoring_counts.get("goals", 0) or 0)
        assists_cnt = int(scoring_counts.get("assists", 0) or 0)
        points_val = goals_cnt + assists_cnt
        ppg_val = float(points_val)

        ev_map = per_player_goal_events.get(player_key, {}) or {}
        ot_goals_cnt = sum(
            1 for ev in (ev_map.get("goals") or []) if int(getattr(ev, "period", 0) or 0) >= 4
        )
        ot_assists_cnt = sum(
            1 for ev in (ev_map.get("assists") or []) if int(getattr(ev, "period", 0) or 0) >= 4
        )
        gt_goals_cnt, gw_goals_cnt = _count_goal_role_flags(list(ev_map.get("goals") or []))

        # Individual per-player stats file (best-effort, without shift-derived fields).
        try:
            lines: List[str] = []
            lines.append(f"Player: {_display_player_name(player_key)}")
            lines.append("")
            lines.append("TimeToScore-only game (no shift spreadsheet)")
            lines.append("")
            lines.append(f"Goals: {goals_cnt}")
            lines.append(f"Assists: {assists_cnt}")
            lines.append(f"OT Goals: {ot_goals_cnt}")
            lines.append(f"OT Assists: {ot_assists_cnt}")
            lines.append(f"Points (G+A): {points_val}")
            goals_list = ev_map.get("goals") or []
            assists_list = ev_map.get("assists") or []
            if goals_list:
                lines.append("")
                lines.append("Goals:")
                for ev in sorted(goals_list, key=lambda e: (e.period, e.t_sec)):
                    tags: List[str] = []
                    if getattr(ev, "is_game_tying", False):
                        tags.append("GT")
                    if getattr(ev, "is_game_winning", False):
                        tags.append("GW")
                    tag_str = f" [{' '.join(tags)}]" if tags else ""
                    lines.append(f"  Period {ev.period}, {ev.t_str}{tag_str}")
            if assists_list:
                lines.append("")
                lines.append("Assists:")
                for ev in sorted(assists_list, key=lambda e: (e.period, e.t_sec)):
                    tags2: List[str] = []
                    if getattr(ev, "is_game_tying", False):
                        tags2.append("GT")
                    if getattr(ev, "is_game_winning", False):
                        tags2.append("GW")
                    tag_str2 = f" [{' '.join(tags2)}]" if tags2 else ""
                    lines.append(f"  Period {ev.period}, {ev.t_str}{tag_str2}")
            (stats_dir / f"{player_key}_stats.txt").write_text(
                "\n".join(lines) + "\n", encoding="utf-8"
            )
        except Exception:
            pass

        row_map: Dict[str, str] = {
            "player": player_key,
            "goals": str(goals_cnt),
            "assists": str(assists_cnt),
            "gt_goals": str(gt_goals_cnt),
            "gw_goals": str(gw_goals_cnt),
            "ot_goals": str(ot_goals_cnt),
            "ot_assists": str(ot_assists_cnt),
            "points": str(points_val),
            "gp": "1",
            "ppg": f"{ppg_val:.1f}",
            # Long-sheet-derived stats are not available via T2S; leave blank.
            "shots": "",
            "shots_per_game": "",
            "sog": "",
            "sog_per_game": "",
            "expected_goals": "",
            "expected_goals_per_game": "",
            "expected_goals_per_sog": "",
            "giveaways": "",
            "giveaways_per_game": "",
            "takeaways": "",
            "takeaways_per_game": "",
            "controlled_entry_for": "",
            "controlled_entry_for_per_game": "",
            "controlled_entry_against": "",
            "controlled_entry_against_per_game": "",
            "controlled_exit_for": "",
            "controlled_exit_for_per_game": "",
            "controlled_exit_against": "",
            "controlled_exit_against_per_game": "",
            # Shift-derived metrics are not available; leave blank.
            "plus_minus": "",
            "plus_minus_per_game": "",
            "gf_counted": "",
            "gf_per_game": "",
            "ga_counted": "",
            "ga_per_game": "",
        }
        if include_shifts_in_stats:
            row_map.update(
                {
                    "shifts": "",
                    "shifts_per_game": "",
                    "sb_toi_total": "",
                    "sb_toi_per_game": "",
                    "sb_avg": "",
                    "sb_median": "",
                    "sb_longest": "",
                    "sb_shortest": "",
                    "video_toi_total": "",
                }
            )
        stats_table_rows.append(row_map)

    if stats_table_rows:
        _write_player_stats_text_and_csv(
            stats_dir,
            stats_table_rows,
            periods_seen,
            include_shifts_in_stats=include_shifts_in_stats,
        )

    _write_game_stats_files_t2s_only(
        stats_dir,
        label=label,
        t2s_id=int(t2s_id),
        goals=goals,
        periods=periods_seen,
    )

    return outdir, stats_table_rows, periods_seen, per_player_goal_events


# ----------------------------- CLI -----------------------------


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Extract per-player shifts & stats from an Excel sheet like 'dh-tv-12-1.xls'."
    )
    p.add_argument(
        "--input",
        "-i",
        dest="inputs",
        action="append",
        type=str,
        default=[],
        help="Path to input .xls/.xlsx file, or a directory containing one primary sheet plus optional '*-long*' companion sheets. "
        "Can repeat for multiple games. "
        "Append ':HOME' or ':AWAY' to override side for that file.",
    )
    p.add_argument(
        "--file-list",
        type=Path,
        default=None,
        help=(
            "Path to a text or YAML file containing one .xls/.xlsx path or directory per line (comments/# allowed). "
            "Directories are expanded to the primary sheet plus optional '*-long*' companion sheets. "
            "You can append ':HOME' or ':AWAY' per line. "
            "For YAML (.yaml/.yml), prefer mapping entries with subkeys (e.g., path/side/date/time/home_team/away_team/game_video); "
            "legacy '|key=value' metadata is still supported but deprecated. "
            "To force a game to be treated as non-TimeToScore even if its filename ends with a T2S id: "
            "in YAML add 'no_t2s: 1' (or under metadata/meta); in text file-lists add '|no_t2s=1'. "
            "Lines may also be 't2s=<game_id>[:HOME|AWAY][:game_label]' to process a TimeToScore-only game with no spreadsheets."
        ),
    )
    p.add_argument(
        "--sheet", "-s", type=str, default=None, help="Worksheet name (default: first sheet)."
    )
    p.add_argument(
        "--ignore-primary",
        action="store_true",
        help=(
            "Prefer '*-long*' companion spreadsheets when both primary and '*-long*' are present. "
            "Falls back to primary-only if no '*-long*' sheet exists for a game."
        ),
    )
    p.add_argument(
        "--ignore-long",
        action="store_true",
        help=(
            "Prefer primary shift spreadsheets when both primary and '*-long*' are present. "
            "Falls back to '*-long*' only if no primary sheet exists for a game."
        ),
    )
    p.add_argument(
        "--verbose",
        action="store_true",
        help=(
            "Verbose output. Prints a per-game summary of spreadsheet event labels and how they were mapped "
            "to output event types (and fails the run if any event rows were not mapped)."
        ),
    )
    p.add_argument(
        "--outdir", "-o", type=Path, default=Path("player_focus"), help="Output directory."
    )
    p.add_argument(
        "--dump-events",
        action="store_true",
        help=(
            "After processing each game, print the parsed per-game events (from stats/all_events_summary.csv) "
            "to stdout. Implies writing all_events_summary even without --shifts/--upload-webapp."
        ),
    )
    p.add_argument(
        "--keep-goalies",
        action="store_true",
        help="By default, rows like '(G) 37' are skipped. Use this flag to include them.",
    )
    # Goals
    p.add_argument(
        "--goal",
        "-g",
        action="append",
        default=[],
        help="Goal event token. Example: 'GF:2/13:45' or 'GA:1/05:12'. Can repeat.",
    )
    p.add_argument(
        "--goals-file",
        type=Path,
        default=None,
        help="Path to a text file with one goal per line (GF:period/time or GA:period/time). '#' lines ignored.",
    )
    # TimeToScore integration
    p.add_argument(
        "--t2s",
        type=str,
        default=None,
        help=(
            "TimeToScore spec: '<game_id>[:HOME|AWAY][:game_label]'. "
            "If set and no --goal/--goals-file provided, fetch goals from T2S. "
            "If provided without --input/--file-list, runs a TimeToScore-only game (no spreadsheets)."
        ),
    )
    p.add_argument(
        "--no-time2score",
        "--no-t2s",
        action="store_true",
        help=(
            "Disable all TimeToScore usage: ignore `t2s=...` file-list lines, do not infer T2S ids from filenames, "
            "and do not fetch goals/start times from TimeToScore."
        ),
    )
    p.add_argument(
        "--t2s-cache-only",
        action="store_true",
        help=(
            "Allow TimeToScore usage, but only from the local cache/DB (no sync/scrape). "
            "If a game is missing locally, TimeToScore lookups will fail."
        ),
    )
    p.add_argument(
        "--t2s-scrape-only",
        action="store_true",
        help=(
            "Allow TimeToScore usage, but avoid full season syncs; only scrape the specific game "
            "when it is missing locally."
        ),
    )
    p.add_argument(
        "--hockey-db-dir",
        type=Path,
        default=Path.home() / ".cache" / "hockeymom",
        help="Directory for the hockey_league.db used when fetching TimeToScore goals (default: ~/.cache/hockeymom).",
    )
    side_group = p.add_mutually_exclusive_group()
    side_group.add_argument(
        "--home",
        action="store_true",
        help="Your team is the home team (with --t2s).",
    )
    side_group.add_argument(
        "--away",
        action="store_true",
        help="Your team is the away team (with --t2s).",
    )
    team_color_group = p.add_mutually_exclusive_group()
    team_color_group.add_argument(
        "--light",
        action="store_true",
        help="For '*-long*' sheets: treat the White team as your team when mapping events to players.",
    )
    team_color_group.add_argument(
        "--dark",
        action="store_true",
        help="For '*-long*' sheets: treat the Blue team as your team when mapping events to players.",
    )
    p.add_argument(
        "--shifts",
        action="store_true",
        help=(
            "Include shift/TOI metrics in parent-facing stats outputs "
            "(stats/*.txt, stats/player_stats.* and consolidated workbook). "
            "Also enables writing per-player shift clip scripts and `*_video_times.txt` / `*_scoreboard_times.txt` files. "
            "By default these are omitted."
        ),
    )
    p.add_argument(
        "--no-scripts",
        action="store_true",
        help="Do not generate clip helper scripts or any '*_times.txt' timestamp files used for clipping.",
    )
    p.add_argument(
        "--skip-validation",
        action="store_true",
        help="Skip validation checks on start/end ordering and excessive durations.",
    )
    p.add_argument(
        "--upload-webapp",
        action="store_true",
        help="Upload per-game CSV outputs (all_events_summary and optional shift_rows) to the HockeyMOM webapp via REST.",
    )
    p.add_argument(
        "--corrections-yaml",
        type=Path,
        default=None,
        help=(
            "Apply event corrections to the webapp via REST from a YAML file (requires --upload-webapp). "
            "The YAML may be a list of correction objects, or a mapping containing `corrections:`."
        ),
    )
    p.add_argument(
        "--webapp-url",
        type=str,
        default="http://127.0.0.1:8008",
        help="Webapp base URL for --upload-webapp (default: http://127.0.0.1:8008).",
    )
    p.add_argument(
        "--webapp-token",
        type=str,
        default=None,
        help="Optional import token for webapp REST upload (sent as Authorization: Bearer ... and X-HM-Import-Token).",
    )
    p.add_argument(
        "--import-token",
        type=str,
        default=None,
        help="Alias for --webapp-token.",
    )
    p.add_argument(
        "--webapp-replace",
        action="store_true",
        help="For --upload-webapp: overwrite existing stats/events for the game (server-side).",
    )
    p.add_argument(
        "--webapp-owner-email",
        type=str,
        default=None,
        help="Owner email for creating/mapping external games when uploading to the webapp.",
    )
    p.add_argument(
        "--webapp-league-name",
        type=str,
        default=None,
        help="League name for mapping/creating games when uploading to the webapp (required for external games).",
    )
    p.add_argument(
        "--webapp-division-name",
        type=str,
        default=None,
        help="Division name to use for external games when uploading to the webapp (default: External).",
    )
    p.add_argument(
        "--webapp-create-missing-players",
        action="store_true",
        help=(
            "When uploading to the webapp, allow creating missing player records from the uploaded CSVs. "
            "Default is off; prefer providing rosters via TimeToScore/shift spreadsheets."
        ),
    )
    return p


def _dump_game_events_from_outdir(*, label: str, outdir: Path) -> None:
    stats_dir = outdir / "stats"
    csv_path = stats_dir / "all_events_summary.csv"
    if not csv_path.exists():
        print(
            f"[dump-events] {label}: no all_events_summary.csv found at {csv_path}",
            file=sys.stderr,
        )
        return

    try:
        df = pd.read_csv(csv_path)
    except Exception as e:  # noqa: BLE001
        print(
            f"[dump-events] {label}: failed to read {csv_path}: {type(e).__name__}: {e}",
            file=sys.stderr,
        )
        return

    try:
        team_side = ""
        try:
            team_side = str(outdir.parent.name)
        except Exception:
            team_side = ""
        hdr = f"[dump-events] {label}"
        if team_side in {"Home", "Away"}:
            hdr += f" ({team_side})"
        print(hdr)

        if df.empty:
            print("  (no events)")
            return

        # Keep output readable by focusing on common, stable columns.
        preferred_cols = [
            "Event ID",
            "Event Type",
            "Team Side",
            "Period",
            "Game Time",
            "Video Time",
            "Attributed Players",
            "Attributed Jerseys",
            "Details",
            "Source",
        ]
        cols = [c for c in preferred_cols if c in df.columns]
        if "Event ID" in df.columns:
            try:
                df = df.sort_values(by=["Event ID"], kind="stable")
            except Exception:
                pass

        for _, r in df[cols].iterrows():
            d = r.to_dict()
            event_id = str(d.get("Event ID") or "").strip()
            etype = str(d.get("Event Type") or "").strip()
            side = str(d.get("Team Side") or "").strip()
            per = str(d.get("Period") or "").strip()
            game_t = str(d.get("Game Time") or "").strip()
            video_t = str(d.get("Video Time") or "").strip()
            ap = str(d.get("Attributed Players") or "").strip()
            aj = str(d.get("Attributed Jerseys") or "").strip()
            details = str(d.get("Details") or "").strip()
            src = str(d.get("Source") or "").strip()

            attrib = aj or ap
            extra = []
            if attrib:
                extra.append(f"attrib={attrib}")
            if details:
                extra.append(f"details={details}")
            if src:
                extra.append(f"src={src}")
            extra_txt = f"  ({', '.join(extra)})" if extra else ""
            print(
                f"  {event_id:>3} {etype:<10} {side:<4} P{per} {game_t:<6} v{video_t:<6}{extra_txt}"
            )
    finally:
        sys.stdout.flush()


def _apply_event_corrections_to_webapp(
    *,
    webapp_url: str,
    webapp_token: Optional[str],
    corrections_yaml: Path,
) -> Dict[str, Any]:
    try:
        import requests  # type: ignore
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"requests is required for --corrections-yaml: {e}") from e

    data_raw = corrections_yaml.read_text(encoding="utf-8", errors="ignore").lstrip("\ufeff")
    data = yaml.safe_load(data_raw) if data_raw.strip() else None
    if data is None:
        return {}
    create_missing_players = False
    if isinstance(data, dict):
        create_missing_players = bool(data.get("create_missing_players", False))
        corrections = data.get("corrections")
    else:
        corrections = data
    if not isinstance(corrections, list) or not corrections:
        raise ValueError("corrections YAML must contain a non-empty list (or `corrections:` list)")

    headers: Dict[str, str] = {}
    if webapp_token:
        tok = str(webapp_token).strip()
        if tok:
            headers["Authorization"] = f"Bearer {tok}"
            headers["X-HM-Import-Token"] = tok

    base = str(webapp_url or "").rstrip("/")
    req_payload: Dict[str, Any] = {"corrections": corrections}
    if create_missing_players:
        req_payload["create_missing_players"] = True
    r = requests.post(
        f"{base}/api/internal/apply_event_corrections",
        json=req_payload,
        headers=headers,
        timeout=60,
    )
    if r.status_code != 200:
        raise RuntimeError(f"apply_event_corrections failed: {r.status_code}: {r.text}")
    try:
        payload = r.json()
    except Exception:  # noqa: BLE001
        payload = None
    if not isinstance(payload, dict) or not payload.get("ok"):
        raise RuntimeError(f"apply_event_corrections failed: {payload!r}")
    stats = payload.get("stats")
    return stats if isinstance(stats, dict) else {}


def _apply_event_corrections_payload_to_webapp(
    *,
    webapp_url: str,
    webapp_token: Optional[str],
    corrections: list[dict[str, Any]],
    create_missing_players: bool = False,
) -> Dict[str, Any]:
    try:
        import requests  # type: ignore
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"requests is required to apply event corrections: {e}") from e

    if not corrections:
        return {}

    headers: Dict[str, str] = {}
    if webapp_token:
        tok = str(webapp_token).strip()
        if tok:
            headers["Authorization"] = f"Bearer {tok}"
            headers["X-HM-Import-Token"] = tok

    base = str(webapp_url or "").rstrip("/")
    req_payload: Dict[str, Any] = {"corrections": list(corrections)}
    if create_missing_players:
        req_payload["create_missing_players"] = True
    r = requests.post(
        f"{base}/api/internal/apply_event_corrections",
        json=req_payload,
        headers=headers,
        timeout=60,
    )
    if r.status_code != 200:
        raise RuntimeError(f"apply_event_corrections failed: {r.status_code}: {r.text}")
    try:
        payload = r.json()
    except Exception:  # noqa: BLE001
        payload = None
    if not isinstance(payload, dict) or not payload.get("ok"):
        raise RuntimeError(f"apply_event_corrections failed: {payload!r}")
    stats = payload.get("stats")
    return stats if isinstance(stats, dict) else {}


def _upload_shift_package_to_webapp(
    *,
    webapp_url: str,
    webapp_token: Optional[str],
    t2s_game_id: Optional[int],
    external_game_key: Optional[str],
    label: str,
    stats_dir: Path,
    replace: bool,
    owner_email: Optional[str] = None,
    league_name: Optional[str] = None,
    division_name: Optional[str] = None,
    sort_order: Optional[int] = None,
    team_side: Optional[str] = None,
    home_team_name: Optional[str] = None,
    away_team_name: Optional[str] = None,
    starts_at: Optional[str] = None,
    home_logo_b64: Optional[str] = None,
    home_logo_content_type: Optional[str] = None,
    away_logo_b64: Optional[str] = None,
    away_logo_content_type: Optional[str] = None,
    home_logo_url: Optional[str] = None,
    away_logo_url: Optional[str] = None,
    game_video_url: Optional[str] = None,
    stats_note: Optional[str] = None,
    roster_home: Optional[List[Dict[str, Any]]] = None,
    roster_away: Optional[List[Dict[str, Any]]] = None,
    create_missing_players: bool = False,
    include_events: bool = True,
    source_label_suffix: Optional[str] = None,
) -> None:
    try:
        import requests  # type: ignore
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"requests is required for --upload-webapp: {e}") from e

    def _read_text(p: Path) -> str:
        try:
            if p.exists():
                return p.read_text(encoding="utf-8")
        except Exception:
            pass
        return ""

    events_csv = _read_text(stats_dir / "all_events_summary.csv") if include_events else ""
    shift_rows_csv = _read_text(stats_dir / "shift_rows.csv")

    payload: Dict[str, Any] = {
        "events_csv": events_csv,
        "source_label": (
            f"parse_stats_inputs:{label}{str(source_label_suffix) if source_label_suffix else ''}"
        ),
        "replace": bool(replace),
    }
    if t2s_game_id is not None:
        payload["timetoscore_game_id"] = int(t2s_game_id)
    if external_game_key:
        payload["external_game_key"] = str(external_game_key)
    if owner_email:
        payload["owner_email"] = str(owner_email)
    if league_name:
        payload["league_name"] = str(league_name)
    if division_name:
        payload["division_name"] = str(division_name)
    if sort_order is not None:
        payload["sort_order"] = int(sort_order)
    if team_side in {"home", "away"}:
        payload["team_side"] = str(team_side)
    if home_team_name:
        payload["home_team_name"] = str(home_team_name)
    if away_team_name:
        payload["away_team_name"] = str(away_team_name)
    if starts_at:
        payload["starts_at"] = str(starts_at)
    if home_logo_b64:
        payload["home_logo_b64"] = str(home_logo_b64)
    if home_logo_content_type:
        payload["home_logo_content_type"] = str(home_logo_content_type)
    if away_logo_b64:
        payload["away_logo_b64"] = str(away_logo_b64)
    if away_logo_content_type:
        payload["away_logo_content_type"] = str(away_logo_content_type)
    if home_logo_url:
        payload["home_logo_url"] = str(home_logo_url)
    if away_logo_url:
        payload["away_logo_url"] = str(away_logo_url)
    if roster_home:
        payload["roster_home"] = roster_home
    if roster_away:
        payload["roster_away"] = roster_away
    payload["create_missing_players"] = bool(create_missing_players)
    if game_video_url:
        payload["game_video_url"] = str(game_video_url)
    if stats_note:
        payload["stats_note"] = str(stats_note)
    if shift_rows_csv and shift_rows_csv.strip():
        payload["shift_rows_csv"] = str(shift_rows_csv)
        payload["replace_shift_rows"] = True
    headers: Dict[str, str] = {}
    if webapp_token:
        tok = str(webapp_token).strip()
        if tok:
            headers["Authorization"] = f"Bearer {tok}"
            headers["X-HM-Import-Token"] = tok

    base = str(webapp_url or "").rstrip("/")
    try:
        r = requests.post(
            f"{base}/api/import/hockey/shift_package",
            json=payload,
            headers=headers,
            timeout=180,
        )
    except requests.RequestException as e:  # type: ignore[attr-defined]
        raise RuntimeError(f"shift_package request failed: {e}") from e

    if r.status_code != 200:
        detail = str(r.text or "").strip()
        msg = f"shift_package failed: {r.status_code}"
        if detail:
            msg = f"{msg}: {detail}"
        raise RuntimeError(msg)
    try:
        out = r.json()
    except Exception as e:  # noqa: BLE001
        detail = str(r.text or "").strip()
        msg = f"shift_package returned non-JSON: {e}"
        if detail:
            msg = f"{msg}: {detail}"
        raise RuntimeError(msg) from e
    if not out.get("ok"):
        detail = str(out.get("error") or out.get("message") or "").strip()
        msg = f"shift_package failed: {out!r}"
        if detail:
            msg = f"shift_package failed: {detail}"
        raise RuntimeError(msg)
    unmatched = out.get("unmatched") or []
    if unmatched:
        if t2s_game_id is not None:
            print(
                f"[webapp] Uploaded shift package for t2s={t2s_game_id} ({label}) with unmatched: {unmatched}"
            )
        else:
            print(
                f"[webapp] Uploaded shift package for external={external_game_key} ({label}) with unmatched: {unmatched}"
            )
    else:
        if t2s_game_id is not None:
            print(f"[webapp] Uploaded shift package for t2s={t2s_game_id} ({label})")
        else:
            print(f"[webapp] Uploaded shift package for external={external_game_key} ({label})")


def main() -> None:
    args = build_arg_parser().parse_args()
    hockey_db_dir = args.hockey_db_dir.expanduser()
    use_t2s = not bool(getattr(args, "no_time2score", False))
    t2s_cache_only = bool(getattr(args, "t2s_cache_only", False))
    t2s_scrape_only = bool(getattr(args, "t2s_scrape_only", False))
    ignore_primary = bool(args.ignore_primary)
    ignore_long = bool(args.ignore_long)
    if ignore_primary and ignore_long:
        print("Error: --ignore-primary cannot be combined with --ignore-long.", file=sys.stderr)
        sys.exit(2)
    if t2s_cache_only and t2s_scrape_only:
        print("Error: --t2s-cache-only cannot be combined with --t2s-scrape-only.", file=sys.stderr)
        sys.exit(2)
    if not use_t2s and (t2s_cache_only or t2s_scrape_only):
        print(
            "Error: TimeToScore is disabled; drop --t2s-cache-only/--t2s-scrape-only.",
            file=sys.stderr,
        )
        sys.exit(2)
    t2s_allow_remote = not t2s_cache_only
    t2s_allow_full_sync = bool(t2s_allow_remote and not t2s_scrape_only)
    dump_events = bool(args.dump_events)
    create_scripts = not args.no_scripts
    include_shifts_in_stats = bool(getattr(args, "shifts", False))
    if dump_events and not include_shifts_in_stats:
        # Dumping events is usually a debug workflow; avoid writing clip scripts/timestamp files unless
        # the user explicitly asked for shift outputs.
        create_scripts = False
    write_events_summary = (
        include_shifts_in_stats or bool(getattr(args, "upload_webapp", False)) or dump_events
    )
    if getattr(args, "corrections_yaml", None) and not getattr(args, "upload_webapp", False):
        print("Error: --corrections-yaml requires --upload-webapp.", file=sys.stderr)
        sys.exit(2)
    focus_team_override: Optional[str] = None
    if getattr(args, "light", False):
        focus_team_override = "White"
    elif getattr(args, "dark", False):
        focus_team_override = "Blue"

    t2s_arg_id: Optional[int] = None
    t2s_arg_side: Optional[str] = None
    t2s_arg_label: Optional[str] = None
    if args.t2s is not None:
        if not use_t2s:
            print("Error: --no-time2score cannot be combined with --t2s.", file=sys.stderr)
            sys.exit(2)
        parsed = _parse_t2s_spec(args.t2s)
        if parsed is None:
            print(
                f"Error: invalid --t2s spec '{args.t2s}' (expected '<id>[:HOME|AWAY][:label]').",
                file=sys.stderr,
            )
            sys.exit(2)
        t2s_arg_id, t2s_arg_side, t2s_arg_label = parsed

    input_entries: List[InputEntry] = []
    if args.file_list:
        try:
            file_list_path = args.file_list.expanduser()
            base_dir = _stats_base_dir_from_env() or file_list_path.resolve().parent
            if file_list_path.suffix.lower() in {".yaml", ".yml"}:
                input_entries.extend(
                    _load_input_entries_from_yaml_file_list(
                        file_list_path, base_dir=base_dir, use_t2s=use_t2s
                    )
                )
            else:
                with file_list_path.open("r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        # Strip UTF-8 BOM if present (common when files are created on Windows).
                        line = line.lstrip("\ufeff")
                        if not line or line.startswith("#"):
                            continue
                        parts = [p.strip() for p in str(line).split("|") if p.strip()]
                        token = parts[0] if parts else ""
                        meta: dict[str, str] = {}
                        for seg in parts[1:]:
                            if "=" not in seg:
                                continue
                            k, v = seg.split("=", 1)
                            kk = str(k or "").strip().lower()
                            vv = str(v or "").strip()
                            if kk:
                                meta[kk] = vv

                        t2s_only = _parse_t2s_only_token(token)
                        if t2s_only is not None:
                            if not use_t2s:
                                print(
                                    f"[no-time2score] Skipping file-list entry: {token}",
                                    file=sys.stderr,
                                )
                                continue
                            t2s_id, side, label = t2s_only
                            input_entries.append(
                                InputEntry(
                                    path=None,
                                    side=side,
                                    t2s_id=t2s_id,
                                    label=label,
                                    meta=meta,
                                )
                            )
                            continue
                        p, side, inline_meta = _parse_input_token_with_meta(
                            token, base_dir=base_dir
                        )
                        merged_meta = dict(meta or {})
                        for k, v in (inline_meta or {}).items():
                            kk = str(k or "").strip().lower()
                            vv = str(v or "").strip()
                            if not kk or not vv:
                                continue
                            if kk in merged_meta and str(merged_meta[kk]) != vv:
                                raise ValueError(
                                    f"conflicting metadata for key '{kk}': {merged_meta[kk]!r} vs {vv!r}"
                                )
                            merged_meta[kk] = vv
                        input_entries.append(InputEntry(path=p, side=side, meta=merged_meta))
        except Exception as e:
            print(f"Error reading --file-list: {e}", file=sys.stderr)
            sys.exit(2)
    for tok in args.inputs or []:
        parts = [p.strip() for p in str(tok).split("|") if p.strip()]
        token = parts[0] if parts else ""
        meta: dict[str, str] = {}
        for seg in parts[1:]:
            if "=" not in seg:
                continue
            k, v = seg.split("=", 1)
            kk = str(k or "").strip().lower()
            vv = str(v or "").strip()
            if kk:
                meta[kk] = vv

        t2s_only = _parse_t2s_only_token(token)
        if t2s_only is not None:
            if not use_t2s:
                print(f"[no-time2score] Skipping input entry: {token}", file=sys.stderr)
                continue
            t2s_id, side, label = t2s_only
            input_entries.append(
                InputEntry(path=None, side=side, t2s_id=t2s_id, label=label, meta=meta)
            )
            continue
        p, side, inline_meta = _parse_input_token_with_meta(token)
        merged_meta = dict(meta or {})
        for k, v in (inline_meta or {}).items():
            kk = str(k or "").strip().lower()
            vv = str(v or "").strip()
            if not kk or not vv:
                continue
            if kk in merged_meta and str(merged_meta[kk]) != vv:
                raise ValueError(
                    f"conflicting metadata for key '{kk}': {merged_meta[kk]!r} vs {vv!r}"
                )
            merged_meta[kk] = vv
        input_entries.append(InputEntry(path=p, side=side, meta=merged_meta))

    if not input_entries:
        # Corrections-only mode (no games processed).
        if getattr(args, "corrections_yaml", None):
            stats = _apply_event_corrections_to_webapp(
                webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                or "http://127.0.0.1:8008",
                webapp_token=(
                    getattr(args, "webapp_token", None)
                    or getattr(args, "import_token", None)
                    or None
                ),
                corrections_yaml=Path(getattr(args, "corrections_yaml")).expanduser(),
            )
            print(f"[webapp] Applied event corrections: {stats}")
            return
        # Allow a TimeToScore-only run by specifying just `--t2s`.
        if t2s_arg_id is not None:
            input_entries.append(
                InputEntry(
                    path=None, side=t2s_arg_side, t2s_id=int(t2s_arg_id), label=t2s_arg_label
                )
            )
        else:
            print(
                "Error: at least one --input/--file-list entry or --t2s is required.",
                file=sys.stderr,
            )
            sys.exit(2)

    # Support passing a directory to --input for single-game runs: discover the
    # primary shift sheet plus optional '*-long*' companion sheet(s) from that
    # directory (or `<dir>/stats`).
    expanded_entries: List[InputEntry] = []
    for entry in input_entries:
        if entry.path is None:
            expanded_entries.append(entry)
            continue
        pp = Path(entry.path).expanduser()
        if pp.is_dir():
            try:
                discovered = _expand_dir_input_to_game_sheets(
                    pp, ignore_primary=ignore_primary, ignore_long=ignore_long
                )
            except Exception as e:  # noqa: BLE001
                print(f"Error expanding directory input {pp}: {e}", file=sys.stderr)
                sys.exit(2)
            for fp in discovered:
                expanded_entries.append(
                    InputEntry(path=fp, side=entry.side, meta=dict(entry.meta or {}))
                )
        else:
            try:
                if pp.is_file() and _is_spreadsheet_input_path(pp):
                    if ignore_primary and (not _is_long_sheet_path(pp)):
                        base_label = _base_label_from_path(pp)
                        found_long: list[Path] = []
                        for cand in _discover_spreadsheet_inputs_in_dir(pp.parent):
                            if not _is_long_sheet_path(cand):
                                continue
                            if _base_label_from_path(cand) != base_label:
                                continue
                            found_long.append(cand)
                        if found_long:
                            for cand in found_long:
                                expanded_entries.append(
                                    InputEntry(
                                        path=cand, side=entry.side, meta=dict(entry.meta or {})
                                    )
                                )
                            continue
                    if ignore_long and _is_long_sheet_path(pp):
                        base_label = _base_label_from_path(pp)
                        found_primary: list[Path] = []
                        for cand in _discover_spreadsheet_inputs_in_dir(pp.parent):
                            if _is_long_sheet_path(cand):
                                continue
                            if _base_label_from_path(cand) != base_label:
                                continue
                            found_primary.append(cand)
                        if len(found_primary) == 1:
                            expanded_entries.append(
                                InputEntry(
                                    path=found_primary[0],
                                    side=entry.side,
                                    meta=dict(entry.meta or {}),
                                )
                            )
                            continue
                        if len(found_primary) > 1:
                            print(
                                f"Error: --ignore-long is set but expected exactly 1 primary sheet for {pp} "
                                f"(found {len(found_primary)}).",
                                file=sys.stderr,
                            )
                            sys.exit(2)
                    expanded_entries.append(
                        InputEntry(path=pp, side=entry.side, meta=dict(entry.meta or {}))
                    )
                    # If a primary sheet is provided directly, auto-discover any companion '*-long*' sheet(s)
                    # with the same base label in the same directory.
                    if (not ignore_primary) and (not ignore_long) and (not _is_long_sheet_path(pp)):
                        base_label = _base_label_from_path(pp)
                        for cand in _discover_spreadsheet_inputs_in_dir(pp.parent):
                            if not _is_long_sheet_path(cand):
                                continue
                            if _base_label_from_path(cand) != base_label:
                                continue
                            expanded_entries.append(
                                InputEntry(path=cand, side=entry.side, meta=dict(entry.meta or {}))
                            )
                else:
                    expanded_entries.append(
                        InputEntry(path=pp, side=entry.side, meta=dict(entry.meta or {}))
                    )
            except Exception as e:  # noqa: BLE001
                print(f"Error processing input path {pp}: {e}", file=sys.stderr)
    input_entries = expanded_entries

    base_outdir = args.outdir.expanduser()
    # Group '-long' companion sheets with their non-long counterpart so a game is processed once.
    groups_by_label: Dict[str, Dict[str, Any]] = {}
    display_label_to_key: Dict[str, str] = {}
    for order_idx, entry in enumerate(input_entries):
        side = entry.side
        if entry.t2s_id is not None and entry.path is None:
            group_key = f"t2s-{int(entry.t2s_id)}"
            display_label = str(entry.label or group_key)
            existing_key = display_label_to_key.get(display_label)
            if existing_key is not None and existing_key != group_key:
                print(
                    f"Error: duplicate game label '{display_label}' (conflicts between {existing_key} and {group_key}).",
                    file=sys.stderr,
                )
                sys.exit(2)
            display_label_to_key.setdefault(display_label, group_key)
            g = groups_by_label.setdefault(
                group_key,
                {
                    "label": display_label,
                    "primary": None,
                    "long_paths": [],
                    "side": None,
                    "order": order_idx,
                    "t2s_id_only": int(entry.t2s_id),
                    "meta": {},
                    "event_corrections": None,
                },
            )
        else:
            if entry.path is None:
                continue
            p = Path(entry.path)
            if p.stem.lower() == "goals":
                label = str(entry.label or _label_from_goals_xlsx_path(p))
            else:
                label = str(entry.label or _base_label_from_path(p))
            existing_key = display_label_to_key.get(label)
            if existing_key is not None and existing_key != label:
                print(
                    f"Error: duplicate game label '{label}' (conflicts between {existing_key} and {label}).",
                    file=sys.stderr,
                )
                sys.exit(2)
            display_label_to_key.setdefault(label, label)
            g = groups_by_label.setdefault(
                label,
                {
                    "label": label,
                    "primary": None,
                    "primaries": [],
                    "long_paths": [],
                    "side": None,
                    "order": order_idx,
                    "meta": {},
                    "event_corrections": None,
                },
            )
        if side:
            if g.get("side") is None:
                g["side"] = side
            elif g.get("side") != side:
                # Allow multiple per-sheet side overrides for the same game label (e.g., one HOME sheet and one AWAY sheet).
                # Downstream processing uses per-primary sides; this group-level side is only a default.
                g["side"] = None

        meta = getattr(entry, "meta", None) or {}
        if meta:
            gm = g.get("meta") or {}
            for k, v in meta.items():
                kk = str(k or "").strip().lower()
                vv = str(v or "").strip()
                if not kk or not vv:
                    continue
                if kk in gm and str(gm[kk]) != vv:
                    print(
                        f"Error: conflicting metadata for '{g.get('label')}' key '{kk}': {gm[kk]!r} vs {vv!r}",
                        file=sys.stderr,
                    )
                    sys.exit(2)
                gm[kk] = vv
            g["meta"] = gm

        event_corrections = getattr(entry, "event_corrections", None)
        if event_corrections:
            existing = g.get("event_corrections")
            if existing is None:
                g["event_corrections"] = event_corrections
            elif existing != event_corrections:
                print(
                    f"Error: conflicting event_corrections for '{g.get('label')}'.",
                    file=sys.stderr,
                )
                sys.exit(2)

        if entry.path is None:
            continue
        if _is_long_sheet_path(p):
            if not ignore_long:
                g["long_paths"].append(p)
        else:
            if g.get("primary") is None:
                g["primary"] = p
            primaries = g.get("primaries") or []
            primaries.append({"path": p, "side": side, "order": order_idx})
            g["primaries"] = primaries

    groups = sorted(groups_by_label.values(), key=lambda x: int(x.get("order", 0)))
    multiple_inputs = len(groups) > 1
    results: List[Dict[str, Any]] = []
    goal_discrepancy_rows: List[Dict[str, Any]] = []
    upload_ok = 0
    upload_failed = 0
    upload_skipped_external_missing_meta = 0

    def _meta_truthy(meta: dict[str, str], *keys: str) -> bool:
        for k in keys:
            kk = str(k or "").strip().lower()
            if not kk:
                continue
            if kk not in meta:
                continue
            raw = str(meta.get(kk) or "").strip().lower()
            if raw in {"", "0", "false", "no", "n", "off"}:
                return False
            return raw in {"1", "true", "yes", "y", "on"} or True
        return False

    def _meta_t2s_id(meta: dict[str, str]) -> Optional[int]:
        for k in (
            "t2s",
            "t2s_id",
            "timetoscore_game_id",
            "timetoscore",
            "time2score",
        ):
            raw = meta.get(str(k).strip().lower())
            if raw is None:
                continue
            try:
                return int(str(raw).strip())
            except Exception:
                continue
        return None

    def _no_t2s_for_game(meta: dict[str, str]) -> bool:
        return _meta_truthy(
            meta,
            "no_t2s",
            "no-t2s",
            "no_time2score",
            "no-time2score",
            "disable_t2s",
            "disable-t2s",
        )

    def _meta_str(meta: dict[str, str], *keys: str) -> Optional[str]:
        for k in keys:
            kk = str(k or "").strip().lower()
            if not kk:
                continue
            v = meta.get(kk)
            if v is None:
                continue
            vv = str(v).strip()
            if vv:
                return vv
        return None

    def _match_team_name(a: Optional[str], b: Optional[str]) -> bool:
        """
        Best-effort match for team names that may be shortened (e.g., 'Utah' vs
        'Utah ...') or contain varying whitespace/punctuation.
        """
        aa = _normalize_header_label(str(a or ""))
        bb = _normalize_header_label(str(b or ""))
        if not aa or not bb:
            return False
        return aa == bb or aa in bb or bb in aa

    def _infer_side_from_meta_and_long_sheet(
        in_path: Path, *, meta: dict[str, str], sheet: Any
    ) -> Tuple[Optional[str], Dict[str, Any]]:
        """
        If the caller did not specify HOME/AWAY (and no TimeToScore id is in use),
        try to infer which side this spreadsheet represents by matching the long
        sheet's embedded team names against `home_team` / `away_team` metadata.
        """
        debug: Dict[str, Any] = {}
        home_team = _meta_str(meta, "home_team", "home_team_name", "home")
        away_team = _meta_str(meta, "away_team", "away_team_name", "away")
        debug["meta_home_team"] = home_team or ""
        debug["meta_away_team"] = away_team or ""
        if not home_team or not away_team:
            debug["failure"] = "missing home_team/away_team metadata"
            return None, debug

        if not _is_long_sheet_path(in_path):
            debug["failure"] = "not a '*-long*' sheet"
            return None, debug

        try:
            target_sheet = 0 if sheet is None else sheet
            long_df = pd.read_excel(in_path, sheet_name=target_sheet, header=None)
        except Exception as e:  # noqa: BLE001
            debug["failure"] = f"failed to read spreadsheet: {type(e).__name__}: {e}"
            return None, debug

        try:
            parsed = _parse_long_shift_tables(long_df) or {}
        except Exception as e:  # noqa: BLE001
            debug["failure"] = f"failed to parse long shift tables: {type(e).__name__}: {e}"
            return None, debug

        teams: List[Tuple[str, int]] = []
        for team_name, info in parsed.items():
            sb_pairs = (info or {}).get("sb_pairs_by_player") or {}
            teams.append((str(team_name), len(sb_pairs)))
        teams_sorted = sorted(teams, key=lambda x: (-int(x[1] or 0), x[0]))
        debug["long_teams"] = [{"team": t, "players": int(n)} for t, n in teams_sorted]
        if not teams_sorted:
            debug["failure"] = "no teams found in long shift tables"
            return None, debug

        best_team = teams_sorted[0][0]
        debug["best_team"] = best_team
        if _match_team_name(best_team, away_team):
            return "away", debug
        if _match_team_name(best_team, home_team):
            return "home", debug

        matches: List[str] = []
        for t, _n in teams_sorted:
            if _match_team_name(t, away_team):
                matches.append("away")
            if _match_team_name(t, home_team):
                matches.append("home")
        matches = sorted(set(matches))
        debug["matches"] = matches
        if len(matches) == 1:
            return matches[0], debug

        debug["failure"] = "no unique match to home_team/away_team"
        return None, debug

    # Webapp upload arg validation (fail fast; don't silently skip uploads).
    default_webapp_url = "http://127.0.0.1:8008"
    if str(getattr(args, "webapp_url", "") or "").strip() != default_webapp_url and not getattr(
        args, "upload_webapp", False
    ):
        print("Error: --webapp-url requires --upload-webapp.", file=sys.stderr)
        sys.exit(2)

    if getattr(args, "upload_webapp", False):
        owner_email_req = str(getattr(args, "webapp_owner_email", "") or "").strip()
        league_name_req = str(getattr(args, "webapp_league_name", "") or "").strip()
        if not owner_email_req:
            print("Error: --upload-webapp requires --webapp-owner-email.", file=sys.stderr)
            sys.exit(2)
        if not league_name_req:
            print("Error: --upload-webapp requires --webapp-league-name.", file=sys.stderr)
            sys.exit(2)

        missing_external_meta: list[str] = []
        missing_external_date: list[str] = []
        for gg in groups:
            t2s_only_id = gg.get("t2s_id_only")
            if t2s_only_id is not None:
                continue
            primary = gg.get("primary")
            meta = dict(gg.get("meta") or {})
            t2s_id_inferred = None
            if use_t2s and not _no_t2s_for_game(meta):
                t2s_id_inferred = _meta_t2s_id(meta)
                if t2s_id_inferred is None:
                    if primary:
                        t2s_id_inferred = _infer_t2s_from_filename(Path(primary))
                    else:
                        long_paths = list(gg.get("long_paths") or [])
                        for lp in long_paths:
                            t2s_id_inferred = _infer_t2s_from_filename(Path(lp))
                            if t2s_id_inferred is not None:
                                break
            if t2s_id_inferred is not None:
                continue

            def _m(*keys: str) -> Optional[str]:
                for k in keys:
                    v = meta.get(str(k).strip().lower())
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return None

            if not (
                _m("home_team", "home_team_name", "home")
                and _m("away_team", "away_team_name", "away")
            ):
                missing_external_meta.append(str(gg.get("label") or gg.get("primary") or "UNKNOWN"))
            if _starts_at_from_meta(meta, warn_label=str(gg.get("label") or "")) is None:
                missing_external_date.append(str(gg.get("label") or gg.get("primary") or "UNKNOWN"))
        if missing_external_meta:
            print(
                "Error: --upload-webapp external games require per-game metadata for team names.\n"
                "Add `home_team` and `away_team` metadata for these games in --file-list "
                "(YAML: mapping keys or under `metadata:`; text: `|home_team=...|away_team=...`):\n"
                "  - " + "\n  - ".join(missing_external_meta),
                file=sys.stderr,
            )
            sys.exit(2)
        if missing_external_date:
            print(
                "Error: --upload-webapp requires a resolvable game date for external games.\n"
                "Add `date` (and optional `time`) metadata for these games in --file-list "
                "(YAML: mapping keys or under `metadata:`; text: `|date=...|time=...`):\n"
                "  - " + "\n  - ".join(missing_external_date),
                file=sys.stderr,
            )
            sys.exit(2)

    if t2s_arg_id is not None and len(groups) > 1:
        print(
            "Error: --t2s can only be used with a single game. "
            "For multi-game runs, use filename suffixes (e.g. 'game-51602.xlsx') "
            "or `t2s=<id>[:HOME|AWAY][:label]` lines in --file-list.",
            file=sys.stderr,
        )
        sys.exit(2)

    def _resolve_goals_for_file(
        in_path: Path,
        t2s_id: Optional[int],
        side: Optional[str],
        *,
        meta: Optional[dict[str, str]] = None,
        allow_remote: bool,
        allow_full_sync: bool,
    ) -> List[GoalEvent]:
        g = load_goals(args.goal, args.goals_file)
        if g:
            return g
        # goals.xlsx always takes precedence when present.
        goals_xlsx = in_path.parent / "goals.xlsx"
        if goals_xlsx.exists() and goals_xlsx.is_file():
            meta_for_file = dict(meta or {})
            our_team_name: Optional[str] = None
            side_l = str(side or "").strip().lower()
            if side_l == "home":
                our_team_name = _meta_str(meta_for_file, "home_team", "home_team_name", "home")
            elif side_l == "away":
                our_team_name = _meta_str(meta_for_file, "away_team", "away_team_name", "away")

            our_jerseys: Optional[set[str]] = None

            # For long-sheet-only runs, the generic roster-table extractor can
            # return Blue/White blocks that include both teams, which flips GF/GA
            # mapping. Prefer the embedded long-shift roster for our matched team.
            if our_team_name and _is_long_sheet_path(in_path):
                try:
                    target_sheet = 0 if args.sheet is None else args.sheet
                    df_long = pd.read_excel(in_path, sheet_name=target_sheet, header=None)
                    parsed = _parse_long_shift_tables(df_long) or {}
                    best = None
                    best_n = -1
                    for team_name, info in parsed.items():
                        if not _match_team_name(str(team_name), our_team_name):
                            continue
                        sb_pairs = (info or {}).get("sb_pairs_by_player") or {}
                        n = len(sb_pairs)
                        if n > best_n:
                            best_n = n
                            best = (str(team_name), sb_pairs)
                    if best is not None:
                        _team_name, sb_pairs = best
                        nums: set[str] = set()
                        for pk in sb_pairs.keys():
                            parts = _parse_player_key(pk)
                            norm = _normalize_jersey_number(parts.jersey)
                            if norm:
                                nums.add(norm)
                        if nums:
                            our_jerseys = nums
                except Exception:
                    our_jerseys = None

            # Fall back to extracting a roster from the sheet (primary sheet case).
            if our_jerseys is None:
                try:
                    target_sheet = 0 if args.sheet is None else args.sheet
                    df_primary = pd.read_excel(in_path, sheet_name=target_sheet, header=None)
                    roster_tables = _extract_roster_tables_from_df(df_primary)
                    if roster_tables:
                        # If we know the team name, avoid taking a union across
                        # multiple rosters (which can invert GF/GA).
                        if our_team_name:
                            best_roster = None
                            best_size = -1
                            for team_name, roster in roster_tables:
                                if not _match_team_name(str(team_name), our_team_name):
                                    continue
                                n = len((roster or {}).keys())
                                if n > best_size:
                                    best_size = n
                                    best_roster = roster
                            if best_roster is not None:
                                our_jerseys = {
                                    str(x).strip()
                                    for x in (best_roster or {}).keys()
                                    if str(x).strip()
                                }
                        elif len(roster_tables) == 1:
                            _team, roster = roster_tables[0]
                            our_jerseys = {
                                str(x).strip() for x in (roster or {}).keys() if str(x).strip()
                            }
                except Exception:
                    our_jerseys = None
            try:
                gx = _goals_from_goals_xlsx(
                    goals_xlsx, our_jerseys=our_jerseys, our_team_name=our_team_name
                )
            except Exception as e:  # noqa: BLE001
                print(f"[goals.xlsx] Failed to parse {goals_xlsx}: {e}", file=sys.stderr)
                gx = []
            if gx:
                for gg in reversed(sorted([str(x) for x in gx])):
                    print(f"[goals.xlsx:{in_path.name}] {gg}")
                return gx

        if t2s_id is None:
            return g
        # With a t2s id, require a side and use TimeToScore data.
        if side is None:
            print(
                "Error: T2S game id "
                f"{t2s_id} provided but side could not be determined while processing "
                f"'{_base_label_from_path(in_path)}' ({in_path}). "
                "Provide --home/--away (single game) or ':HOME' / ':AWAY' in --file-list.",
                file=sys.stderr,
            )
            sys.exit(2)
        try:
            with _working_directory(hockey_db_dir):
                g = goals_from_t2s(
                    int(t2s_id),
                    side=side,
                    allow_remote=allow_remote,
                    allow_full_sync=allow_full_sync,
                )
        except Exception as e:  # noqa: BLE001
            print(
                f"Error: failed to fetch goals from TimeToScore for game {t2s_id} while processing "
                f"'{_base_label_from_path(in_path)}' ({in_path}).",
                file=sys.stderr,
            )
            if args.file_list:
                print(f"  File list: {args.file_list}", file=sys.stderr)
            print(f"  Side: {side}", file=sys.stderr)
            print(f"  Cause: {e}", file=sys.stderr)
            sys.exit(2)

        for gg in reversed(sorted([str(x) for x in g])):
            print(f"[t2s:{t2s_id}] {gg}")
        return g

    for idx, g in enumerate(groups):
        meta_for_group = dict(g.get("meta") or {})
        no_t2s_for_group = _no_t2s_for_game(meta_for_group)

        t2s_only_id = g.get("t2s_id_only")
        if t2s_only_id is not None:
            if no_t2s_for_group:
                print(
                    f"Error: game '{g.get('label')}' is a TimeToScore-only entry (t2s={t2s_only_id}) "
                    "but metadata requested no TimeToScore (no_t2s=1). Remove the t2s=... line.",
                    file=sys.stderr,
                )
                sys.exit(2)
            if not use_t2s:
                print(
                    f"[no-time2score] Skipping TimeToScore-only game: t2s={t2s_only_id}",
                    file=sys.stderr,
                )
                continue
            label = str(g.get("label") or f"t2s-{int(t2s_only_id)}")
            outdir = base_outdir if not multiple_inputs else base_outdir / label
            side_override: Optional[str] = g.get("side") or (
                "home" if args.home else ("away" if args.away else None)
            )
            if side_override is None:
                print(
                    f"Error: T2S-only game {t2s_only_id} requires a side (provide --home/--away or ':HOME'/'AWAY' on the t2s=... line).",
                    file=sys.stderr,
                )
                sys.exit(2)
            try:
                final_outdir, stats_rows, periods, per_player_events = process_t2s_only_game(
                    t2s_id=int(t2s_only_id),
                    side=str(side_override),
                    outdir=outdir,
                    label=label,
                    hockey_db_dir=hockey_db_dir,
                    include_shifts_in_stats=include_shifts_in_stats,
                    keep_goalies=bool(args.keep_goalies),
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                )
            except Exception as e:  # noqa: BLE001
                print(
                    f"Error: failed to process TimeToScore-only game {t2s_only_id} ('{label}').",
                    file=sys.stderr,
                )
                if args.file_list:
                    print(f"  File list: {args.file_list}", file=sys.stderr)
                print(f"  Side: {side_override}", file=sys.stderr)
                print(f"  Cause: {e}", file=sys.stderr)
                sys.exit(2)
            results.append(
                {
                    "label": label,
                    "t2s_id": int(t2s_only_id),
                    "order": idx,
                    "outdir": final_outdir,
                    "stats": stats_rows,
                    "periods": periods,
                    "events": per_player_events,
                    "pair_on_ice": [],
                    "sheet_path": None,
                    "video_path": None,
                    "side": str(side_override or "") or None,
                    "meta": dict(g.get("meta") or {}),
                }
            )
            try:
                print(f"✅ Done. Wrote per-player files to: {final_outdir.resolve()}")
            except Exception:
                print("✅ Done.")

            if dump_events:
                _dump_game_events_from_outdir(label=label, outdir=final_outdir)
            continue

        primary_specs_raw = list(g.get("primaries") or [])
        primary_specs_sorted = sorted(primary_specs_raw, key=lambda x: int(x.get("order", 0)))
        primary_specs: list[dict[str, Any]] = []
        seen_primary_paths: set[str] = set()
        for s in primary_specs_sorted:
            try:
                p0 = Path(s.get("path"))
            except Exception:
                continue
            k0 = str(p0)
            if not k0 or k0 in seen_primary_paths:
                continue
            seen_primary_paths.add(k0)
            primary_specs.append({"path": p0, "side": s.get("side")})

        multi_primary_processed = False
        if (
            len(primary_specs) == 1
            and isinstance(primary_specs[0].get("path"), Path)
            and str(primary_specs[0]["path"].stem).lower() == "goals"
        ):
            multi_primary_processed = True
            goals_xlsx = Path(primary_specs[0]["path"])
            label = str(g.get("label") or _label_from_goals_xlsx_path(goals_xlsx))
            outdir = base_outdir if not multiple_inputs else base_outdir / label
            meta = dict(g.get("meta") or {})

            def _meta(*keys: str) -> Optional[str]:
                for k in keys:
                    v = meta.get(str(k).strip().lower())
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return None

            side_to_use: Optional[str] = g.get("side") or (
                "home" if args.home else ("away" if args.away else None)
            )
            if side_to_use not in {"home", "away"}:
                side_to_use = "home"

            our_team_name = None
            if side_to_use == "home":
                our_team_name = _meta("home_team", "home_team_name", "home")
            elif side_to_use == "away":
                our_team_name = _meta("away_team", "away_team_name", "away")

            (
                final_outdir,
                stats_rows,
                periods,
                per_player_events,
                pair_on_ice_rows,
            ) = process_goals_only_xlsx(
                goals_xlsx=goals_xlsx,
                outdir=outdir,
                label=label,
                team_side=side_to_use,
                our_team_name=our_team_name,
                write_events_summary=write_events_summary,
            )
            in_path = goals_xlsx
            primary_path = goals_xlsx
            long_paths: List[Path] = []
            t2s_id = None

        if len(primary_specs) > 1:
            # Multi-primary game: process each primary sheet (HOME/AWAY) into its own subtree, sharing any long sheets.
            multi_primary_processed = True
            primary_path = Path(primary_specs[0]["path"])
            in_path = Path(primary_path)

            label = str(g.get("label") or _base_label_from_path(in_path))
            outdir = base_outdir if not multiple_inputs else base_outdir / label

            long_paths_all: list[Path] = []
            _seen_lp: set[str] = set()
            primary_path_set: set[str] = {str(Path(s["path"])) for s in primary_specs}
            for lp in g.get("long_paths") or []:
                try:
                    p = Path(lp)
                except Exception:
                    continue
                kp = str(p)
                if not kp or kp in _seen_lp or kp in primary_path_set:
                    continue
                _seen_lp.add(kp)
                long_paths_all.append(p)

            # Prefer an explicit --t2s value; otherwise infer from filename (if enabled).
            t2s_id = None
            if use_t2s and (not no_t2s_for_group):
                t2s_id = t2s_arg_id if t2s_arg_id is not None else _meta_t2s_id(meta_for_group)
                if t2s_id is None:
                    t2s_id = _infer_t2s_from_filename(in_path)

            manual_goals = load_goals(args.goal, args.goals_file)
            if manual_goals:
                print(
                    "Error: manual --goal/--goals-file is not supported when multiple primary sheets are provided for one game.",
                    file=sys.stderr,
                )
                print(f"  Game: {label}", file=sys.stderr)
                sys.exit(2)

            rep: Optional[dict[str, Any]] = None
            rep_side: Optional[str] = None
            rep_in_path: Optional[Path] = None

            for spec in primary_specs:
                sheet_path = Path(spec["path"])
                sheet_side = str(spec.get("side") or "").strip().lower() or None
                if sheet_side not in {"home", "away"}:
                    print(
                        f"Error: multi-sheet game '{label}' requires a per-sheet HOME/AWAY side for '{sheet_path}'.",
                        file=sys.stderr,
                    )
                    print(
                        "Fix: set `side: HOME|AWAY` for each entry under `sheets:` in the YAML file-list.",
                        file=sys.stderr,
                    )
                    sys.exit(2)

                goals = _resolve_goals_for_file(
                    sheet_path,
                    t2s_id,
                    sheet_side,
                    meta=meta_for_group,
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                )

                roster_map: Optional[Dict[str, str]] = None
                t2s_rosters_by_side: Optional[Dict[str, Dict[str, str]]] = None
                if t2s_id is not None:
                    try:
                        t2s_rosters_by_side = _get_t2s_game_rosters(
                            int(t2s_id),
                            hockey_db_dir,
                            keep_goalies=bool(args.keep_goalies),
                            allow_remote=t2s_allow_remote,
                            allow_full_sync=t2s_allow_full_sync,
                        )
                        roster_map = dict(t2s_rosters_by_side.get(str(sheet_side), {}) or {})
                    except Exception as e:  # noqa: BLE001
                        print(
                            f"Error: failed to fetch roster from TimeToScore for game {t2s_id} while processing '{label}'.",
                            file=sys.stderr,
                        )
                        print(f"  Input sheet: {sheet_path}", file=sys.stderr)
                        if args.file_list:
                            print(f"  File list: {args.file_list}", file=sys.stderr)
                        print(f"  Side: {sheet_side}", file=sys.stderr)
                        print(f"  Cause: {e}", file=sys.stderr)
                        sys.exit(2)

                long_paths_for_sheet = [p for p in long_paths_all if Path(p) != sheet_path]
                (
                    final_outdir,
                    stats_rows,
                    periods,
                    per_player_events,
                    pair_on_ice_rows,
                ) = process_sheet(
                    xls_path=sheet_path,
                    sheet_name=args.sheet,
                    outdir=outdir,
                    keep_goalies=args.keep_goalies,
                    goals=goals,
                    roster_map=roster_map,
                    t2s_rosters_by_side=t2s_rosters_by_side,
                    t2s_side=sheet_side,
                    t2s_game_id=t2s_id,
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                    long_xls_paths=long_paths_for_sheet,
                    focus_team_override=focus_team_override,
                    include_shifts_in_stats=include_shifts_in_stats,
                    write_events_summary=write_events_summary,
                    skip_validation=args.skip_validation,
                    create_scripts=create_scripts,
                    write_opponent_stats_from_long_shifts=False,
                    verbose=bool(args.verbose),
                )

                if rep is None or sheet_side == "home":
                    rep = {
                        "final_outdir": final_outdir,
                        "stats_rows": stats_rows,
                        "periods": periods,
                        "per_player_events": per_player_events,
                        "pair_on_ice_rows": pair_on_ice_rows,
                    }
                    rep_side = sheet_side
                    rep_in_path = sheet_path

            if rep is None or rep_in_path is None:
                continue
            # Representative outputs (used for summary + webapp upload). Prefer Home when present.
            final_outdir = Path(rep["final_outdir"])
            stats_rows = list(rep["stats_rows"] or [])
            periods = list(rep["periods"] or [])
            per_player_events = dict(rep["per_player_events"] or {})
            pair_on_ice_rows = list(rep["pair_on_ice_rows"] or [])
            in_path = Path(rep_in_path)
            side_to_use = rep_side
            primary_path = Path(primary_path)
            long_paths = list(long_paths_all)

        if not multi_primary_processed:
            primary_path = g.get("primary")
            in_path = primary_path
            if in_path is None:
                long_paths = g.get("long_paths") or []
                if long_paths:
                    in_path = long_paths[0]
                else:
                    continue
            in_path = Path(in_path)
            path_side = (
                str(primary_specs[0].get("side") or "").strip().lower() or None
                if primary_specs
                else g.get("side")
            )
            long_paths: List[Path] = [
                Path(p) for p in (g.get("long_paths") or []) if Path(p) != in_path
            ]

            # Prefer an explicit --t2s value; otherwise, infer a T2S id from the
            # filename only when the trailing numeric suffix is large enough
            # (>= 10000). Smaller suffixes (e.g., 'chicago-1') remain part of the
            # game name and do not trigger T2S usage.
            t2s_id = None
            if use_t2s and (not no_t2s_for_group):
                t2s_id = t2s_arg_id if t2s_arg_id is not None else _meta_t2s_id(meta_for_group)
                if t2s_id is None:
                    t2s_id = _infer_t2s_from_filename(in_path)
            label = str(g.get("label") or _base_label_from_path(in_path))
            outdir = base_outdir if not multiple_inputs else base_outdir / label
            manual_goals = load_goals(args.goal, args.goals_file)

            jersey_numbers = set()
            if t2s_id is not None and not manual_goals:
                try:
                    jersey_numbers = _collect_sheet_jerseys(in_path, args.sheet, args.keep_goalies)
                except Exception as e:
                    print(
                        f"Error parsing sheet for jersey numbers ({in_path}): {e}",
                        file=sys.stderr,
                    )

            side_override: Optional[str] = (
                path_side
                or t2s_arg_side
                or ("home" if args.home else ("away" if args.away else None))
            )
            inferred_side: Optional[str] = None
            side_infer_debug: Dict[str, Any] = {}
            if manual_goals:
                side_to_use = side_override
            else:
                if side_override:
                    side_to_use = side_override
                elif t2s_id is not None:
                    inferred_side = _infer_side_from_rosters(
                        int(t2s_id),
                        jersey_numbers,
                        hockey_db_dir,
                        allow_remote=t2s_allow_remote,
                        allow_full_sync=t2s_allow_full_sync,
                        debug=side_infer_debug,
                    )
                    side_to_use = inferred_side
                else:
                    side_to_use = None

            # If TimeToScore is not in use, attempt to infer HOME/AWAY from file-list
            # metadata + the embedded long-sheet shift table team names.
            if (
                (not manual_goals)
                and side_to_use is None
                and t2s_id is None
                and args.file_list
                and meta_for_group
            ):
                inferred_side2, debug2 = _infer_side_from_meta_and_long_sheet(
                    Path(in_path), meta=meta_for_group, sheet=args.sheet
                )
                if inferred_side2 in {"home", "away"}:
                    side_to_use = inferred_side2
                    side_infer_debug = debug2 or {}

            # If a TimeToScore id is in use and we don't have a side override, we
            # must determine whether our team is Home or Away to map GF/GA.
            if t2s_id is not None and (not manual_goals) and side_to_use is None:
                t2s_source = (
                    f"--t2s {args.t2s}"
                    if t2s_arg_id is not None
                    else f"inferred from filename suffix: {in_path.name}"
                )
                cli_side = "--home" if args.home else ("--away" if args.away else "<none>")

                def _safe_int(token: Any) -> int:
                    try:
                        return int(str(token))
                    except Exception:
                        return 0

                jerseys_sorted = (
                    sorted(list(jersey_numbers or set()), key=_safe_int) if jersey_numbers else []
                )
                sample = ", ".join(jerseys_sorted[:20])
                extra = (
                    f", ... (+{len(jerseys_sorted) - 20} more)" if len(jerseys_sorted) > 20 else ""
                )

                print(
                    f"Error: cannot determine HOME/AWAY side for TimeToScore game {t2s_id} while processing '{label}'.",
                    file=sys.stderr,
                )
                print(f"  Input sheet: {in_path}", file=sys.stderr)
                if args.file_list:
                    print(f"  File list: {args.file_list}", file=sys.stderr)
                print(f"  T2S id source: {t2s_source}", file=sys.stderr)
                print("  Side overrides checked:", file=sys.stderr)
                print(
                    f"    - from --file-list ':HOME' / ':AWAY': {path_side or '<none>'}",
                    file=sys.stderr,
                )
                print(
                    f"    - from --t2s spec side: {t2s_arg_side or '<none>'}",
                    file=sys.stderr,
                )
                print(f"    - from CLI flags: {cli_side}", file=sys.stderr)
                print(
                    f"  Jerseys found in sheet: {len(jerseys_sorted)}"
                    + (f" ({sample}{extra})" if jerseys_sorted else ""),
                    file=sys.stderr,
                )
                if side_infer_debug:
                    failure = side_infer_debug.get("failure")
                    if failure:
                        print(f"  Side inference result: {failure}", file=sys.stderr)
                    if side_infer_debug.get("t2s_api_available") is False:
                        import_err = side_infer_debug.get("t2s_api_import_error")
                        details = f": {import_err}" if import_err else ""
                        print(
                            f"  TimeToScore API: not available (failed to import hmlib.time2score.api){details}.",
                            file=sys.stderr,
                        )
                    exc = side_infer_debug.get("exception")
                    if exc:
                        print(f"  TimeToScore fetch error: {exc}", file=sys.stderr)
                    hr = side_infer_debug.get("home_roster_count")
                    ar = side_infer_debug.get("away_roster_count")
                    if isinstance(hr, int) or isinstance(ar, int):
                        print(
                            f"  TimeToScore roster sizes: home={hr or 0}, away={ar or 0}",
                            file=sys.stderr,
                        )
                    ho = side_infer_debug.get("home_overlap")
                    ao = side_infer_debug.get("away_overlap")
                    if isinstance(ho, int) or isinstance(ao, int):
                        print(
                            f"  Overlap with sheet jerseys: home={ho or 0}, away={ao or 0}",
                            file=sys.stderr,
                        )
                    hoj = side_infer_debug.get("home_overlap_jerseys")
                    aoj = side_infer_debug.get("away_overlap_jerseys")
                    if isinstance(hoj, list) or isinstance(aoj, list):
                        try:
                            h_txt = ", ".join(str(x) for x in (hoj or [])[:20])
                            a_txt = ", ".join(str(x) for x in (aoj or [])[:20])
                        except Exception:
                            h_txt = ""
                            a_txt = ""
                        if h_txt or a_txt:
                            print(
                                f"  Overlap jersey numbers: home=[{h_txt}], away=[{a_txt}]",
                                file=sys.stderr,
                            )
                print(
                    "Fix: add ':HOME' or ':AWAY' to this game in --file-list (or run a single game with --home/--away).",
                    file=sys.stderr,
                )
                sys.exit(2)

            # For output organization (Home/Away subdirs) and webapp upload, we always
            # require a known HOME/AWAY side for spreadsheet-backed games.
            if side_to_use not in {"home", "away"}:
                if t2s_id is None:
                    print(
                        f"Error: cannot determine HOME/AWAY side for '{label}' while processing '{in_path.name}'.",
                        file=sys.stderr,
                    )
                    print(
                        "  Reason: no ':HOME' / ':AWAY' provided and no TimeToScore id inferred.",
                        file=sys.stderr,
                    )
                    if side_infer_debug:
                        meta_home = side_infer_debug.get("meta_home_team")
                        meta_away = side_infer_debug.get("meta_away_team")
                        if meta_home or meta_away:
                            print(
                                f"  Metadata: home_team={meta_home!r}, away_team={meta_away!r}",
                                file=sys.stderr,
                            )
                        long_teams = side_infer_debug.get("long_teams")
                        if isinstance(long_teams, list) and long_teams:
                            try:
                                txt = ", ".join(
                                    f"{str(x.get('team') or '')}({int(x.get('players') or 0)})"
                                    for x in long_teams[:8]
                                )
                            except Exception:
                                txt = ""
                            if txt:
                                print(f"  Long-sheet teams (players): {txt}", file=sys.stderr)
                    if args.file_list:
                        print(f"  File list: {args.file_list}", file=sys.stderr)
                    print(
                        "Fix: set `side: HOME|AWAY` for this game in the YAML file-list (or add ':HOME' / ':AWAY' for text file-lists; or run a single game with --home/--away).",
                        file=sys.stderr,
                    )
                else:
                    print(
                        f"Error: cannot determine HOME/AWAY side for TimeToScore game {t2s_id} while processing '{label}'.",
                        file=sys.stderr,
                    )
                    print(
                        "Fix: add ':HOME' or ':AWAY' to this game in --file-list (or run a single game with --home/--away).",
                        file=sys.stderr,
                    )
                sys.exit(2)

            goals = _resolve_goals_for_file(
                in_path,
                t2s_id,
                side_to_use,
                meta=meta_for_group,
                allow_remote=t2s_allow_remote,
                allow_full_sync=t2s_allow_full_sync,
            )

            # Build a TimeToScore roster map (normalized jersey -> name) for GP
            # accounting, so that players listed on the official roster are
            # credited with a game played even if they have no recorded shifts.
            roster_map: Optional[Dict[str, str]] = None
            t2s_rosters_by_side: Optional[Dict[str, Dict[str, str]]] = None
            if t2s_id is not None and side_to_use is not None:
                try:
                    t2s_rosters_by_side = _get_t2s_game_rosters(
                        int(t2s_id),
                        hockey_db_dir,
                        keep_goalies=bool(args.keep_goalies),
                        allow_remote=t2s_allow_remote,
                        allow_full_sync=t2s_allow_full_sync,
                    )
                    roster_map = dict(t2s_rosters_by_side.get(str(side_to_use), {}) or {})
                except Exception as e:  # noqa: BLE001
                    print(
                        f"Error: failed to fetch roster from TimeToScore for game {t2s_id} while processing '{label}'.",
                        file=sys.stderr,
                    )
                    print(f"  Input sheet: {in_path}", file=sys.stderr)
                    if args.file_list:
                        print(f"  File list: {args.file_list}", file=sys.stderr)
                    print(f"  Side: {side_to_use}", file=sys.stderr)
                    print(f"  Cause: {e}", file=sys.stderr)
                    sys.exit(2)

            # Compare TimeToScore's official scoring vs long-sheet recorded goals/assists (auditing).
            long_paths_for_compare: List[Path] = []
            if primary_path is None:
                long_paths_for_compare.append(in_path)
            long_paths_for_compare.extend(list(long_paths or []))
            long_paths_for_compare = [Path(p) for p in long_paths_for_compare if p is not None]
            # De-dupe while preserving order.
            _seen_lp: set[str] = set()
            _tmp_lp: List[Path] = []
            for _p in long_paths_for_compare:
                k = str(Path(_p))
                if k in _seen_lp:
                    continue
                _seen_lp.add(k)
                _tmp_lp.append(Path(_p))
            long_paths_for_compare = _tmp_lp

            if t2s_id is not None and long_paths_for_compare:
                try:
                    with _working_directory(hockey_db_dir):
                        t2s_goals_for_compare = goals_from_t2s(
                            int(t2s_id),
                            side=str(side_to_use),
                            allow_remote=t2s_allow_remote,
                            allow_full_sync=t2s_allow_full_sync,
                        )
                except Exception:
                    t2s_goals_for_compare = []

                if t2s_goals_for_compare:
                    # Best-effort fetch of team names from the TimeToScore scoresheet HTML.
                    home_team_name: Optional[str] = None
                    away_team_name: Optional[str] = None
                    if t2s_allow_remote:
                        try:
                            html = _fetch_t2s_scoresheet_html(int(t2s_id))
                            visitor, home = _t2s_team_names_from_scoresheet_html(html)
                            away_team_name = str(visitor or "").strip() or None
                            home_team_name = str(home or "").strip() or None
                        except Exception:
                            home_team_name = None
                            away_team_name = None

                    long_goal_rows_all: List[Dict[str, Any]] = []
                    jerseys_by_team_all: Dict[str, set[int]] = {}
                    for lp in long_paths_for_compare:
                        try:
                            long_df = pd.read_excel(
                                Path(lp).expanduser(), sheet_name=0, header=None
                            )
                        except Exception:
                            continue
                        try:
                            _long_events, long_goal_rows, jerseys_by_team, _mapping = (
                                _parse_long_left_event_table_with_mapping(long_df)
                            )
                        except Exception:
                            continue
                        long_goal_rows_all.extend(list(long_goal_rows or []))
                        for team, nums in (jerseys_by_team or {}).items():
                            jerseys_by_team_all.setdefault(str(team), set()).update(
                                set(nums or set())
                            )

                    if long_goal_rows_all:
                        our_jerseys: set[str] = set((roster_map or {}).keys())
                        focus_team_for_compare = _infer_focus_team_from_long_sheet(
                            our_jerseys, jerseys_by_team_all
                        )
                        goal_discrepancy_rows.extend(
                            _compare_t2s_vs_long_goals(
                                label=str(label or ""),
                                t2s_id=int(t2s_id),
                                side=str(side_to_use),
                                t2s_goals=t2s_goals_for_compare,
                                long_goal_rows=long_goal_rows_all,
                                focus_team=focus_team_for_compare,
                                home_team_name=home_team_name,
                                away_team_name=away_team_name,
                            )
                        )

            if primary_path is None and long_paths_for_compare:
                final_outdir, stats_rows, periods, per_player_events, pair_on_ice_rows = (
                    process_long_only_sheets(
                        long_xls_paths=list(long_paths_for_compare),
                        outdir=outdir,
                        goals=goals,
                        roster_map=roster_map,
                        t2s_rosters_by_side=t2s_rosters_by_side,
                        t2s_side=side_to_use,
                        t2s_game_id=t2s_id,
                        focus_team_override=focus_team_override,
                        include_shifts_in_stats=include_shifts_in_stats,
                        write_events_summary=write_events_summary,
                        create_scripts=create_scripts,
                        allow_remote=t2s_allow_remote,
                        verbose=bool(args.verbose),
                    )
                )
            else:
                (
                    final_outdir,
                    stats_rows,
                    periods,
                    per_player_events,
                    pair_on_ice_rows,
                ) = process_sheet(
                    xls_path=in_path,
                    sheet_name=args.sheet,
                    outdir=outdir,
                    keep_goalies=args.keep_goalies,
                    goals=goals,
                    roster_map=roster_map,
                    t2s_rosters_by_side=t2s_rosters_by_side,
                    t2s_side=side_to_use,
                    t2s_game_id=t2s_id,
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                    long_xls_paths=long_paths,
                    focus_team_override=focus_team_override,
                    include_shifts_in_stats=include_shifts_in_stats,
                    write_events_summary=write_events_summary,
                    skip_validation=args.skip_validation,
                    create_scripts=create_scripts,
                    verbose=bool(args.verbose),
                )
        video_path = _find_tracking_output_video_for_sheet_path(in_path) if create_scripts else None
        results.append(
            {
                "label": label,
                "t2s_id": t2s_id,
                "order": idx,
                "outdir": final_outdir,
                "stats": stats_rows,
                "periods": periods,
                "events": per_player_events,
                "pair_on_ice": pair_on_ice_rows,
                "sheet_path": in_path,
                "primary_path": str(primary_path) if primary_path is not None else None,
                "long_paths": [str(p) for p in (long_paths or [])],
                "video_path": video_path,
                "side": str(side_to_use or "") or None,
                "meta": dict(g.get("meta") or {}),
            }
        )
        try:
            print(f"✅ Done. Wrote per-player files to: {final_outdir.resolve()}")
        except Exception:
            print("✅ Done.")
        if dump_events:
            _dump_game_events_from_outdir(label=label, outdir=final_outdir)

        if getattr(args, "upload_webapp", False):
            meta = dict(g.get("meta") or {})

            def _meta(*keys: str) -> Optional[str]:
                for k in keys:
                    v = meta.get(str(k).strip().lower())
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return None

            owner_email = _meta("owner_email") or (
                str(getattr(args, "webapp_owner_email", "") or "").strip() or None
            )
            league_name = _meta("league", "league_name") or (
                str(getattr(args, "webapp_league_name", "") or "").strip() or None
            )
            # For T2S-linked games, division is already known from the TimeToScore import,
            # so do not default to "External" (which can cause incorrect mappings).
            if t2s_id is not None:
                division_name = _meta("division", "division_name") or None
            else:
                division_name = _meta("division", "division_name") or (
                    str(getattr(args, "webapp_division_name", "") or "").strip() or None
                )
            sort_order = int((idx + 1) * 100)
            team_side = str(side_to_use or "").strip().lower() or None
            # If this is a non-TimeToScore game and side isn't specified, treat the "for" team as home for now.
            # TODO: infer home/away robustly from spreadsheet metadata when available.
            if team_side not in {"home", "away"}:
                team_side = "home"

            external_home = _meta("home_team", "home_team_name", "home")
            external_away = _meta("away_team", "away_team_name", "away")
            file_list_base_dir: Optional[Path] = None
            try:
                file_list_base_dir = _stats_base_dir_from_env() or (
                    args.file_list.expanduser().resolve().parent if args.file_list else None
                )
            except Exception:
                file_list_base_dir = None
            logo_fields = _load_logo_fields_from_meta(
                meta, base_dir=file_list_base_dir, warn_label=str(label or "")
            )
            starts_at = _starts_at_from_meta(meta, warn_label=str(label or ""))
            if starts_at is None and t2s_id is not None:
                starts_at = _starts_at_from_t2s_game_id(
                    int(t2s_id),
                    hockey_db_dir=hockey_db_dir,
                    warn_label=str(label or ""),
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                )
            if starts_at is None:
                print(
                    "Error: --upload-webapp requires a resolvable game date.\n"
                    "Provide `date` (and optional `time`) metadata in --file-list (YAML mapping preferred; "
                    "text file-lists can use `|date=...|time=...`), or ensure TimeToScore provides start_time.\n"
                    f"  Game: {label} (t2s={t2s_id})",
                    file=sys.stderr,
                )
                sys.exit(2)

            upload_home = external_home
            upload_away = external_away
            upload_home_logo_url = None
            upload_away_logo_url = None
            roster_home_payload: Optional[List[Dict[str, Any]]] = None
            roster_away_payload: Optional[List[Dict[str, Any]]] = None
            if t2s_id is not None and (not upload_home or not upload_away):
                t2s_home, t2s_away = _t2s_team_names_for_game(
                    int(t2s_id),
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                    hockey_db_dir=hockey_db_dir,
                )
                if not upload_home:
                    upload_home = t2s_home
                if not upload_away:
                    upload_away = t2s_away
            if t2s_id is not None:
                try:
                    rosters = _get_t2s_game_rosters(
                        int(t2s_id),
                        hockey_db_dir,
                        allow_remote=t2s_allow_remote,
                        allow_full_sync=t2s_allow_full_sync,
                    )

                    def _roster_list(side: str) -> List[Dict[str, Any]]:
                        out: List[Dict[str, Any]] = []
                        for jersey, name in (rosters.get(side) or {}).items():
                            jj = str(jersey or "").strip()
                            nn = str(name or "").strip()
                            if not jj or not nn:
                                continue
                            out.append({"jersey_number": jj, "name": nn})
                        out.sort(
                            key=lambda r: int(
                                re.sub(r"[^0-9]+", "", str(r.get("jersey_number") or "0")) or "0"
                            )
                        )
                        return out

                    roster_home_payload = _roster_list("home")
                    roster_away_payload = _roster_list("away")
                except Exception:
                    roster_home_payload = None
                    roster_away_payload = None

            # Optional roster seed from goals.xlsx (external games / non-T2S goals-only games).
            # This supports goals.xlsx variants where a "Roster" table appears under each team's goal table.
            if roster_home_payload is None or roster_away_payload is None:
                try:
                    candidates: List[Path] = []
                    if isinstance(primary_path, Path):
                        candidates.append(primary_path)
                        candidates.append(primary_path.parent / "goals.xlsx")
                    if isinstance(in_path, Path):
                        candidates.append(in_path)
                        candidates.append(in_path.parent / "goals.xlsx")
                    for lp in long_paths or []:
                        try:
                            candidates.append(Path(lp).parent / "goals.xlsx")
                        except Exception:
                            pass

                    goals_xlsx_path: Optional[Path] = None
                    for c in candidates:
                        try:
                            if c.name.lower() == "goals.xlsx" and c.exists() and c.is_file():
                                goals_xlsx_path = c
                                break
                        except Exception:
                            continue

                    if goals_xlsx_path is not None:
                        tables = _rosters_from_goals_xlsx(goals_xlsx_path)

                        def _norm_team(s: Optional[str]) -> str:
                            return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().casefold())

                        def _match_score(label: Optional[str], target: Optional[str]) -> int:
                            a = _norm_team(label)
                            b = _norm_team(target)
                            if not a or not b:
                                return 0
                            if a == b:
                                return 3
                            if a in b or b in a:
                                return 2
                            return 0

                        home_name = str(upload_home or "").strip() or None
                        away_name = str(upload_away or "").strip() or None

                        unassigned: List[Tuple[Optional[str], List[Dict[str, Any]]]] = []
                        for team_label, roster in tables:
                            if not roster:
                                continue
                            sh = _match_score(team_label, home_name)
                            sa = _match_score(team_label, away_name)
                            if sh > sa and sh > 0 and roster_home_payload is None:
                                roster_home_payload = roster
                            elif sa > sh and sa > 0 and roster_away_payload is None:
                                roster_away_payload = roster
                            else:
                                unassigned.append((team_label, roster))

                        # Fallback: if labels are missing/ambiguous, assign by order.
                        if (
                            roster_home_payload is None
                            and roster_away_payload is None
                            and len(unassigned) == 2
                        ):
                            roster_home_payload = unassigned[0][1]
                            roster_away_payload = unassigned[1][1]
                        elif len(unassigned) == 1:
                            if roster_home_payload is None and roster_away_payload is None:
                                roster_home_payload = unassigned[0][1]
                            elif roster_home_payload is None:
                                roster_home_payload = unassigned[0][1]
                            elif roster_away_payload is None:
                                roster_away_payload = unassigned[0][1]
                except Exception:
                    # Best-effort: roster extraction must not block uploads.
                    pass
            if t2s_id is not None and not (
                logo_fields.get("home_logo_b64") or logo_fields.get("away_logo_b64")
            ):
                home_logo_url, away_logo_url = _t2s_team_logo_urls_for_game(
                    int(t2s_id),
                    allow_remote=t2s_allow_remote,
                    allow_full_sync=t2s_allow_full_sync,
                    hockey_db_dir=hockey_db_dir,
                )
                upload_home_logo_url = home_logo_url or None
                upload_away_logo_url = away_logo_url or None

            print(f"[webapp] Uploading {idx + 1}/{len(groups)}: {label}")

            try:
                # If both Home/Away outputs exist (e.g., when long-sheet shift tables were used to
                # generate both teams), upload both. To avoid overwriting the per-game raw CSV blobs
                # and game/event stats, we only upload the non-primary side's shift rows (no events).
                def _team_stats_dirs(final_out: Path) -> Dict[str, Path]:
                    # final_out is "<game>/<Home|Away>/<format_dir>"
                    root = final_out.parent.parent
                    format_dir = final_out.name
                    return {
                        "home": root / "Home" / format_dir / "stats",
                        "away": root / "Away" / format_dir / "stats",
                    }

                def _has_uploadable_stats(sd: Path) -> bool:
                    try:
                        for fn in ("all_events_summary.csv", "shift_rows.csv"):
                            p = sd / fn
                            if p.exists() and p.is_file() and p.stat().st_size > 0:
                                return True
                    except Exception:
                        pass
                    return False

                if t2s_id is not None:
                    dirs = _team_stats_dirs(final_outdir)
                    primary_side = str(team_side or "").strip().lower()
                    if primary_side not in {"home", "away"}:
                        primary_side = "home"
                    primary_stats = dirs.get(primary_side, final_outdir / "stats")
                    if _has_uploadable_stats(primary_stats):
                        _upload_shift_package_to_webapp(
                            webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                            or "http://127.0.0.1:8008",
                            webapp_token=(
                                getattr(args, "webapp_token", None)
                                or getattr(args, "import_token", None)
                                or None
                            ),
                            t2s_game_id=int(t2s_id),
                            external_game_key=str(label or ""),
                            label=str(label or ""),
                            stats_dir=primary_stats,
                            replace=bool(getattr(args, "webapp_replace", False)),
                            owner_email=owner_email,
                            league_name=league_name,
                            division_name=division_name,
                            sort_order=sort_order,
                            team_side=primary_side,
                            starts_at=starts_at,
                            home_team_name=upload_home,
                            away_team_name=upload_away,
                            home_logo_b64=logo_fields.get("home_logo_b64"),
                            home_logo_content_type=logo_fields.get("home_logo_content_type"),
                            away_logo_b64=logo_fields.get("away_logo_b64"),
                            away_logo_content_type=logo_fields.get("away_logo_content_type"),
                            home_logo_url=upload_home_logo_url,
                            away_logo_url=upload_away_logo_url,
                            roster_home=roster_home_payload,
                            roster_away=roster_away_payload,
                            game_video_url=_meta("game_video", "game_video_url", "video_url"),
                            stats_note=_meta("stats_note", "schedule_note"),
                            create_missing_players=False,
                            source_label_suffix=f":{primary_side}",
                        )
                        upload_ok += 1
                        embedded = g.get("event_corrections")
                        if embedded:
                            corr_objs: list[dict[str, Any]] = []
                            if isinstance(embedded, list):
                                corr_objs = [{"patch": embedded}]
                            elif isinstance(embedded, dict):
                                corr_objs = [dict(embedded)]
                            else:
                                raise ValueError(
                                    f"invalid event_corrections for {label!r} (must be list or mapping)"
                                )
                            for cobj in corr_objs:
                                if not any(
                                    cobj.get(k)
                                    for k in (
                                        "game_id",
                                        "timetoscore_game_id",
                                        "tts_game_id",
                                        "external_game_key",
                                        "external_key",
                                        "label",
                                    )
                                ):
                                    cobj["timetoscore_game_id"] = int(t2s_id)
                            stats = _apply_event_corrections_payload_to_webapp(
                                webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                                or "http://127.0.0.1:8008",
                                webapp_token=(
                                    getattr(args, "webapp_token", None)
                                    or getattr(args, "import_token", None)
                                    or None
                                ),
                                corrections=corr_objs,
                                create_missing_players=False,
                            )
                            print(
                                f"[webapp] Applied embedded event corrections for {label}: {stats}"
                            )

                    other_side = "away" if primary_side == "home" else "home"
                    other_stats = dirs.get(other_side)
                    if other_stats is not None and _has_uploadable_stats(other_stats):
                        # Avoid overwriting per-game events: upload only the other side's shift rows (no events).
                        _upload_shift_package_to_webapp(
                            webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                            or "http://127.0.0.1:8008",
                            webapp_token=(
                                getattr(args, "webapp_token", None)
                                or getattr(args, "import_token", None)
                                or None
                            ),
                            t2s_game_id=int(t2s_id),
                            external_game_key=str(label or ""),
                            label=str(label or ""),
                            stats_dir=other_stats,
                            replace=False,
                            owner_email=owner_email,
                            league_name=league_name,
                            division_name=division_name,
                            sort_order=sort_order,
                            team_side=other_side,
                            starts_at=starts_at,
                            home_team_name=upload_home,
                            away_team_name=upload_away,
                            home_logo_b64=logo_fields.get("home_logo_b64"),
                            home_logo_content_type=logo_fields.get("home_logo_content_type"),
                            away_logo_b64=logo_fields.get("away_logo_b64"),
                            away_logo_content_type=logo_fields.get("away_logo_content_type"),
                            home_logo_url=upload_home_logo_url,
                            away_logo_url=upload_away_logo_url,
                            roster_home=roster_home_payload,
                            roster_away=roster_away_payload,
                            game_video_url=_meta("game_video", "game_video_url", "video_url"),
                            stats_note=_meta("stats_note", "schedule_note"),
                            create_missing_players=False,
                            include_events=False,
                            source_label_suffix=f":{other_side}",
                        )
                        upload_ok += 1
                else:
                    if not (owner_email and league_name and external_home and external_away):
                        upload_skipped_external_missing_meta += 1
                        missing = []
                        if not owner_email:
                            missing.append("owner_email")
                        if not league_name:
                            missing.append("league")
                        if not external_home:
                            missing.append("home_team")
                        if not external_away:
                            missing.append("away_team")
                        try:
                            print(
                                f"[webapp] Skipping external game '{label}': missing {', '.join(missing)} "
                                "(add metadata in --file-list: YAML mapping keys, or legacy `|key=value` in text file-lists).",
                                file=sys.stderr,
                            )
                        except Exception:
                            pass
                    else:
                        dirs = _team_stats_dirs(final_outdir)
                        primary_side = str(team_side or "").strip().lower()
                        if primary_side not in {"home", "away"}:
                            primary_side = "home"
                        primary_stats = dirs.get(primary_side, final_outdir / "stats")
                        _upload_shift_package_to_webapp(
                            webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                            or "http://127.0.0.1:8008",
                            webapp_token=(
                                getattr(args, "webapp_token", None)
                                or getattr(args, "import_token", None)
                                or None
                            ),
                            t2s_game_id=None,
                            external_game_key=str(label or ""),
                            label=str(label or ""),
                            stats_dir=primary_stats,
                            replace=bool(getattr(args, "webapp_replace", False)),
                            owner_email=owner_email,
                            league_name=league_name,
                            division_name=division_name or "External",
                            sort_order=sort_order,
                            team_side=primary_side,
                            home_team_name=external_home,
                            away_team_name=external_away,
                            starts_at=starts_at,
                            home_logo_b64=logo_fields.get("home_logo_b64"),
                            home_logo_content_type=logo_fields.get("home_logo_content_type"),
                            away_logo_b64=logo_fields.get("away_logo_b64"),
                            away_logo_content_type=logo_fields.get("away_logo_content_type"),
                            home_logo_url=upload_home_logo_url,
                            away_logo_url=upload_away_logo_url,
                            roster_home=roster_home_payload,
                            roster_away=roster_away_payload,
                            game_video_url=_meta("game_video", "game_video_url", "video_url"),
                            stats_note=_meta("stats_note", "schedule_note"),
                            create_missing_players=bool(
                                getattr(args, "webapp_create_missing_players", False)
                            ),
                            source_label_suffix=f":{primary_side}",
                        )
                        upload_ok += 1
                        embedded = g.get("event_corrections")
                        if embedded:
                            corr_objs: list[dict[str, Any]] = []
                            if isinstance(embedded, list):
                                corr_objs = [{"patch": embedded}]
                            elif isinstance(embedded, dict):
                                corr_objs = [dict(embedded)]
                            else:
                                raise ValueError(
                                    f"invalid event_corrections for {label!r} (must be list or mapping)"
                                )
                            for cobj in corr_objs:
                                if not any(
                                    cobj.get(k)
                                    for k in (
                                        "game_id",
                                        "timetoscore_game_id",
                                        "tts_game_id",
                                        "external_game_key",
                                        "external_key",
                                        "label",
                                    )
                                ):
                                    cobj["external_game_key"] = str(label or "")
                                    if owner_email:
                                        cobj["owner_email"] = str(owner_email)
                            stats = _apply_event_corrections_payload_to_webapp(
                                webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                                or "http://127.0.0.1:8008",
                                webapp_token=(
                                    getattr(args, "webapp_token", None)
                                    or getattr(args, "import_token", None)
                                    or None
                                ),
                                corrections=corr_objs,
                                create_missing_players=bool(
                                    getattr(args, "webapp_create_missing_players", False)
                                ),
                            )
                            print(
                                f"[webapp] Applied embedded event corrections for {label}: {stats}"
                            )

                        other_side = "away" if primary_side == "home" else "home"
                        other_stats = dirs.get(other_side)
                        if other_stats is not None and _has_uploadable_stats(other_stats):
                            _upload_shift_package_to_webapp(
                                webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
                                or "http://127.0.0.1:8008",
                                webapp_token=(
                                    getattr(args, "webapp_token", None)
                                    or getattr(args, "import_token", None)
                                    or None
                                ),
                                t2s_game_id=None,
                                external_game_key=str(label or ""),
                                label=str(label or ""),
                                stats_dir=other_stats,
                                replace=False,
                                owner_email=owner_email,
                                league_name=league_name,
                                division_name=division_name or "External",
                                sort_order=sort_order,
                                team_side=other_side,
                                home_team_name=external_home,
                                away_team_name=external_away,
                                starts_at=starts_at,
                                home_logo_b64=logo_fields.get("home_logo_b64"),
                                home_logo_content_type=logo_fields.get("home_logo_content_type"),
                                away_logo_b64=logo_fields.get("away_logo_b64"),
                                away_logo_content_type=logo_fields.get("away_logo_content_type"),
                                home_logo_url=upload_home_logo_url,
                                away_logo_url=upload_away_logo_url,
                                roster_home=roster_home_payload,
                                roster_away=roster_away_payload,
                                game_video_url=_meta("game_video", "game_video_url", "video_url"),
                                stats_note=_meta("stats_note", "schedule_note"),
                                create_missing_players=bool(
                                    getattr(args, "webapp_create_missing_players", False)
                                ),
                                include_events=False,
                                source_label_suffix=f":{other_side}",
                            )
                            upload_ok += 1
            except Exception as e:  # noqa: BLE001
                upload_failed += 1
                if t2s_id is not None:
                    print(
                        f"[webapp] Upload failed for t2s={t2s_id} ({label}): {e}", file=sys.stderr
                    )
                else:
                    print(
                        f"[webapp] Upload failed for external={label} ({label}): {e}",
                        file=sys.stderr,
                    )

    if getattr(args, "upload_webapp", False):
        print(
            f"[webapp] Upload summary: ok={upload_ok} failed={upload_failed} skipped_external_missing_meta={upload_skipped_external_missing_meta}"
        )
        if upload_ok == 0 and upload_skipped_external_missing_meta > 0:
            print(
                "[webapp] External games were skipped because metadata was missing. Add file-list metadata like "
                "'|owner_email=you@example.com|league=CAHA|home_team=...|away_team=...' for non-TimeToScore games.",
                file=sys.stderr,
            )

    if getattr(args, "corrections_yaml", None):
        stats = _apply_event_corrections_to_webapp(
            webapp_url=str(getattr(args, "webapp_url", "") or "").strip()
            or "http://127.0.0.1:8008",
            webapp_token=(
                getattr(args, "webapp_token", None) or getattr(args, "import_token", None) or None
            ),
            corrections_yaml=Path(getattr(args, "corrections_yaml")).expanduser(),
        )
        print(f"[webapp] Applied event corrections: {stats}")

    _print_game_inputs_rich_summary(results)

    if goal_discrepancy_rows:
        _print_goal_discrepancy_rich_table(goal_discrepancy_rows)

    if multiple_inputs:
        agg_rows, agg_periods, per_game_denoms = _aggregate_stats_rows(
            [(r["stats"], r["periods"]) for r in results]
        )

        # Build per-player event lists across all games (with game labels).
        per_player_events: Dict[str, Dict[str, List[Tuple[str, GoalEvent]]]] = {}
        per_game_stats_by_label: Dict[str, List[Dict[str, str]]] = {}
        for r in results:
            game_label = r.get("label", "")
            ev_map: Dict[str, Dict[str, List[GoalEvent]]] = r.get("events", {}) or {}
            per_game_stats_by_label[game_label] = r.get("stats", []) or []
            for player_key, info in ev_map.items():
                dest = per_player_events.setdefault(
                    player_key, {"goals": [], "assists": [], "gf_on_ice": [], "ga_on_ice": []}
                )
                for ev in info.get("goals", []):
                    dest["goals"].append((game_label, ev))
                for ev in info.get("assists", []):
                    dest["assists"].append((game_label, ev))
                for ev in info.get("gf_on_ice", []):
                    dest["gf_on_ice"].append((game_label, ev))
                for ev in info.get("ga_on_ice", []):
                    dest["ga_on_ice"].append((game_label, ev))

        # Add GT/GW goal counts into aggregated rows.
        _augment_aggregate_with_goal_details(agg_rows, per_player_events)

        # Cumulative sheet: points-based ordering.
        agg_df, _ = _build_stats_dataframe(
            agg_rows,
            agg_periods,
            sort_for_cumulative=True,
            include_shifts_in_stats=include_shifts_in_stats,
            include_per_game_columns=True,
            include_gp_column=True,
        )
        sheets: List[Tuple[str, pd.DataFrame]] = [("Cumulative", agg_df)]
        has_t2s = any(r.get("t2s_id") is not None for r in results)
        ordered_results = (
            sorted(
                results,
                key=lambda r: (
                    0 if r.get("t2s_id") is not None else 1,
                    r.get("t2s_id") if r.get("t2s_id") is not None else float("inf"),
                    r.get("order", 0),
                ),
            )
            if has_t2s
            else results
        )

        # Workbook tab order: keep Cumulative first, then list games in reverse
        # `--file-list` order so the most recent games (often listed last) appear first.
        per_game_sheet_results = list(reversed(results)) if args.file_list else ordered_results
        for r in per_game_sheet_results:
            # Per-game sheets: keep simple alphabetical ordering by player.
            df, _ = _build_stats_dataframe(
                r["stats"],
                r["periods"],
                sort_for_cumulative=False,
                include_shifts_in_stats=include_shifts_in_stats,
                include_per_game_columns=False,
                include_gp_column=False,
            )
            sheets.append((r["label"], df))
        consolidated_path = base_outdir / "player_stats_consolidated.xlsx"
        _write_consolidated_workbook(consolidated_path, sheets, per_game_denoms=per_game_denoms)
        try:
            print(f"📊 Consolidated workbook: {consolidated_path.resolve()}")
        except Exception:
            print("📊 Consolidated workbook written.")

        # Also write a one-sheet CSV summary for the cumulative sheet.
        try:
            df_csv = agg_df.copy()
            if "player" in df_csv.columns:
                if "jersey" not in df_csv.columns:
                    jerseys = df_csv["player"].apply(lambda x: _parse_player_key(x).jersey or "")
                    df_csv.insert(0, "jersey", jerseys)
                df_csv["player"] = df_csv["player"].apply(_format_player_name_only)
            df_csv.columns = [
                (
                    f"{_display_col_name(c)} ({per_game_denoms[c]})"
                    if c in per_game_denoms
                    else _display_col_name(c)
                )
                for c in df_csv.columns
            ]
            player_csv_path = base_outdir / "player_stats_consolidated.csv"
            df_csv.to_csv(player_csv_path, index=False)
        except Exception:
            pass

        # Team-level (game) stats consolidated across all games.
        if _write_game_stats_consolidated_files(base_outdir, results):
            try:
                print(
                    f"📊 Game stats consolidated: {(base_outdir / 'game_stats_consolidated.xlsx').resolve()}"
                )
            except Exception:
                print("📊 Game stats consolidated workbook written.")

        # Pair on-ice stats consolidated across all games (only includes games that had shift data).
        if _write_pair_on_ice_consolidated_files(
            base_outdir, results, include_toi=include_shifts_in_stats
        ):
            try:
                print(
                    f"📊 Pair on-ice consolidated: {(base_outdir / 'pair_on_ice_consolidated.xlsx').resolve()}"
                )
            except Exception:
                print("📊 Pair on-ice consolidated workbook written.")

        # Per-player cumulative detail files across all games.
        _write_cumulative_player_detail_files(
            base_outdir,
            agg_rows,
            per_player_events,
            per_game_stats_by_label,
            {
                str(r.get("label") or ""): (r.get("pair_on_ice") or [])
                for r in results
                if r.get("label")
            },
            include_shifts_in_stats=include_shifts_in_stats,
        )

        # Season (multi-game) highlight reel scripts (skipped with --no-scripts).
        _write_season_highlight_scripts(base_outdir, results, create_scripts=create_scripts)


if __name__ == "__main__":
    main()
